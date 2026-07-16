
"""
AQSD Professional
Module: Trend Intelligence Engine
Version: 1.0

Builds trend intelligence for NSE F&O candidates.

Features
--------
- EMA 20 / 50 / 200 alignment
- EMA slope analysis
- ADX trend-strength analysis
- ATR-normalized momentum
- Price position versus moving averages
- Bullish / Bearish / Neutral trend classification
- Trend Score from 0 to 100
- Excel output: Trend Intelligence

Run
---
python trend_intelligence_engine.py
python trend_intelligence_engine.py --period 2y --limit 100
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"
FNO_FILE = BASE / "Data" / "FnO_Stocks.xlsx"


# ============================================================
# SETTINGS
# ============================================================

DEFAULT_PERIOD = "2y"
DEFAULT_LIMIT = 50

EMA_FAST = 20
EMA_MEDIUM = 50
EMA_SLOW = 200
ADX_PERIOD = 14
ATR_PERIOD = 14
SLOPE_LOOKBACK = 10


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def normalize_symbol(value) -> str:
    symbol = str(value or "").strip().upper()

    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"

    return symbol


def load_symbols(limit: int) -> list[str]:
    symbols: list[str] = []

    if DASHBOARD.exists():
        try:
            wb = load_workbook(
                DASHBOARD,
                read_only=True,
                data_only=True,
            )

            for sheet_name in ("CALL Candidates", "PUT Candidates"):
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                headers = {
                    str(cell.value).strip(): cell.column
                    for cell in ws[1]
                    if cell.value is not None
                }

                symbol_col = headers.get("Symbol")

                if not symbol_col:
                    continue

                for row in range(2, ws.max_row + 1):
                    symbol = normalize_symbol(
                        ws.cell(row, symbol_col).value
                    )

                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

                    if len(symbols) >= limit:
                        wb.close()
                        return symbols

            wb.close()

        except Exception:
            pass

    if FNO_FILE.exists() and len(symbols) < limit:
        try:
            df = pd.read_excel(FNO_FILE)

            symbol_col = next(
                (
                    column
                    for column in (
                        "Yahoo Symbol",
                        "Symbol",
                        "SYMBOL",
                        "Ticker",
                    )
                    if column in df.columns
                ),
                None,
            )

            if symbol_col:
                for value in df[symbol_col].dropna():
                    symbol = normalize_symbol(value)

                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

                    if len(symbols) >= limit:
                        return symbols

        except Exception:
            pass

    fallback = [
        "RELIANCE.NS",
        "HDFCBANK.NS",
        "ICICIBANK.NS",
        "SBIN.NS",
        "INFY.NS",
        "TCS.NS",
        "LT.NS",
        "SUNPHARMA.NS",
        "BIOCON.NS",
        "TATAMOTORS.NS",
    ]

    for symbol in fallback:
        if symbol not in symbols:
            symbols.append(symbol)

        if len(symbols) >= limit:
            break

    return symbols


def download_ohlc(symbol: str, period: str) -> pd.DataFrame:
    df = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
    )

    if df.empty:
        raise RuntimeError("No price data")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    required = ["High", "Low", "Close"]

    if any(column not in df.columns for column in required):
        raise RuntimeError("Required OHLC columns missing")

    return df.dropna(subset=required).copy()


def calculate_atr(df: pd.DataFrame, period: int) -> pd.Series:
    previous_close = df["Close"].shift(1)

    true_range = pd.concat(
        [
            df["High"] - df["Low"],
            (df["High"] - previous_close).abs(),
            (df["Low"] - previous_close).abs(),
        ],
        axis=1,
    ).max(axis=1)

    return true_range.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()


def calculate_adx(df: pd.DataFrame, period: int) -> pd.Series:
    up_move = df["High"].diff()
    down_move = -df["Low"].diff()

    plus_dm = up_move.where(
        (up_move > down_move) & (up_move > 0),
        0.0,
    )

    minus_dm = down_move.where(
        (down_move > up_move) & (down_move > 0),
        0.0,
    )

    atr = calculate_atr(df, period)

    plus_di = 100 * (
        plus_dm.ewm(
            alpha=1 / period,
            adjust=False,
            min_periods=period,
        ).mean()
        / atr
    )

    minus_di = 100 * (
        minus_dm.ewm(
            alpha=1 / period,
            adjust=False,
            min_periods=period,
        ).mean()
        / atr
    )

    denominator = plus_di + minus_di

    dx = 100 * (
        (plus_di - minus_di).abs()
        / denominator.replace(0, pd.NA)
    )

    return dx.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()


def percent_change(current: float, previous: float) -> float:
    if previous == 0:
        return 0.0

    return (current - previous) / previous * 100


def analyse_symbol(symbol: str, df: pd.DataFrame) -> dict:
    df[f"EMA{EMA_FAST}"] = df["Close"].ewm(
        span=EMA_FAST,
        adjust=False,
    ).mean()

    df[f"EMA{EMA_MEDIUM}"] = df["Close"].ewm(
        span=EMA_MEDIUM,
        adjust=False,
    ).mean()

    df[f"EMA{EMA_SLOW}"] = df["Close"].ewm(
        span=EMA_SLOW,
        adjust=False,
    ).mean()

    df["ATR"] = calculate_atr(df, ATR_PERIOD)
    df["ADX"] = calculate_adx(df, ADX_PERIOD)

    df = df.dropna()

    if len(df) <= SLOPE_LOOKBACK:
        raise RuntimeError("Insufficient history")

    close = float(df["Close"].iloc[-1])
    ema20 = float(df[f"EMA{EMA_FAST}"].iloc[-1])
    ema50 = float(df[f"EMA{EMA_MEDIUM}"].iloc[-1])
    ema200 = float(df[f"EMA{EMA_SLOW}"].iloc[-1])

    ema20_old = float(
        df[f"EMA{EMA_FAST}"].iloc[-1 - SLOPE_LOOKBACK]
    )
    ema50_old = float(
        df[f"EMA{EMA_MEDIUM}"].iloc[-1 - SLOPE_LOOKBACK]
    )

    ema20_slope = percent_change(ema20, ema20_old)
    ema50_slope = percent_change(ema50, ema50_old)

    adx = float(df["ADX"].iloc[-1])
    atr = float(df["ATR"].iloc[-1])

    momentum_20 = percent_change(
        close,
        float(df["Close"].iloc[-21]),
    ) if len(df) >= 21 else 0.0

    atr_momentum = (
        (close - float(df["Close"].iloc[-6])) / atr
        if atr > 0 and len(df) >= 6
        else 0.0
    )

    bullish_alignment = ema20 > ema50 > ema200
    bearish_alignment = ema20 < ema50 < ema200

    above_count = sum(
        [
            close > ema20,
            close > ema50,
            close > ema200,
        ]
    )

    below_count = sum(
        [
            close < ema20,
            close < ema50,
            close < ema200,
        ]
    )

    score = 50
    reasons: list[str] = []

    if bullish_alignment:
        score += 20
        reasons.append("Bullish EMA alignment")
    elif bearish_alignment:
        score -= 20
        reasons.append("Bearish EMA alignment")

    if close > ema20:
        score += 5
        reasons.append("Above EMA20")
    else:
        score -= 5
        reasons.append("Below EMA20")

    if close > ema50:
        score += 5
        reasons.append("Above EMA50")
    else:
        score -= 5
        reasons.append("Below EMA50")

    if close > ema200:
        score += 10
        reasons.append("Above EMA200")
    else:
        score -= 10
        reasons.append("Below EMA200")

    if ema20_slope > 0.5:
        score += 8
        reasons.append("EMA20 rising")
    elif ema20_slope < -0.5:
        score -= 8
        reasons.append("EMA20 falling")

    if ema50_slope > 0.3:
        score += 7
        reasons.append("EMA50 rising")
    elif ema50_slope < -0.3:
        score -= 7
        reasons.append("EMA50 falling")

    if momentum_20 > 5:
        score += 8
        reasons.append("Strong 20-day momentum")
    elif momentum_20 < -5:
        score -= 8
        reasons.append("Weak 20-day momentum")

    if atr_momentum > 1:
        score += 5
        reasons.append("Positive ATR momentum")
    elif atr_momentum < -1:
        score -= 5
        reasons.append("Negative ATR momentum")

    score = max(0, min(100, score))

    if score >= 75:
        trend = "STRONG BULLISH"
    elif score >= 60:
        trend = "BULLISH"
    elif score <= 25:
        trend = "STRONG BEARISH"
    elif score <= 40:
        trend = "BEARISH"
    else:
        trend = "NEUTRAL"

    if adx >= 30:
        strength = "STRONG TREND"
    elif adx >= 20:
        strength = "MODERATE TREND"
    else:
        strength = "WEAK / RANGE"

    if bullish_alignment and above_count == 3:
        regime = "MARKUP"
    elif bearish_alignment and below_count == 3:
        regime = "MARKDOWN"
    elif adx < 20:
        regime = "RANGE"
    else:
        regime = "TRANSITION"

    return {
        "Symbol": symbol,
        "Close": round(close, 2),
        "EMA20": round(ema20, 2),
        "EMA50": round(ema50, 2),
        "EMA200": round(ema200, 2),
        "EMA20 Slope %": round(ema20_slope, 2),
        "EMA50 Slope %": round(ema50_slope, 2),
        "ADX": round(adx, 2),
        "20-Day Momentum %": round(momentum_20, 2),
        "ATR Momentum": round(atr_momentum, 2),
        "Trend Classification": trend,
        "Trend Strength": strength,
        "Trend Regime": regime,
        "Trend Score": score,
        "Reason": " | ".join(reasons),
    }


def write_results(results: list[dict]) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Trend Intelligence" in wb.sheetnames:
        del wb["Trend Intelligence"]

    ws = wb.create_sheet("Trend Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:O2")
    ws["A1"] = "AQSD PROFESSIONAL - TREND INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Stocks Analysed"
    ws["B4"] = len(results)
    ws["A4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill(
        "solid",
        fgColor=BLUE,
    )

    headers = [
        "Symbol",
        "Close",
        "EMA20",
        "EMA50",
        "EMA200",
        "EMA20 Slope %",
        "EMA50 Slope %",
        "ADX",
        "20-Day Momentum %",
        "ATR Momentum",
        "Trend Classification",
        "Trend Strength",
        "Trend Regime",
        "Trend Score",
        "Reason",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(6, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    ranked = sorted(
        results,
        key=lambda item: item["Trend Score"],
        reverse=True,
    )

    for row_no, result in enumerate(ranked, start=7):
        for col, heading in enumerate(headers, start=1):
            cell = ws.cell(
                row_no,
                col,
                result.get(heading),
            )
            cell.border = Border(bottom=THIN)

        for col in (2, 3, 4, 5):
            ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

        for col in (6, 7, 9):
            ws.cell(
                row_no,
                col,
            ).number_format = '0.00"%"'

        trend = result["Trend Classification"]
        trend_fill = (
            GREEN
            if "BULLISH" in trend
            else RED
            if "BEARISH" in trend
            else YELLOW
        )

        ws.cell(
            row_no,
            11,
        ).fill = PatternFill(
            "solid",
            fgColor=trend_fill,
        )

        ws.cell(
            row_no,
            11,
        ).font = Font(bold=True)

        score = result["Trend Score"]
        score_fill = (
            GREEN
            if score >= 70
            else RED
            if score <= 30
            else YELLOW
        )

        ws.cell(
            row_no,
            14,
        ).fill = PatternFill(
            "solid",
            fgColor=score_fill,
        )

        ws.cell(
            row_no,
            14,
        ).font = Font(bold=True)

    widths = {
        "A": 18,
        "B": 12,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 14,
        "G": 14,
        "H": 10,
        "I": 16,
        "J": 14,
        "K": 20,
        "L": 18,
        "M": 16,
        "N": 12,
        "O": 55,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build AQSD trend intelligence."
    )

    parser.add_argument(
        "--period",
        default=DEFAULT_PERIOD,
        help="Yahoo Finance period, e.g. 1y, 2y, 5y.",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_LIMIT,
        help="Maximum number of symbols to analyse.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    symbols = load_symbols(args.limit)

    print("\nAQSD TREND INTELLIGENCE")
    print("=" * 72)
    print(f"Symbols: {len(symbols)}")
    print(f"Period: {args.period}")

    results: list[dict] = []

    for index, symbol in enumerate(symbols, start=1):
        print(f"[{index}/{len(symbols)}] {symbol}")

        try:
            df = download_ohlc(
                symbol,
                args.period,
            )

            results.append(
                analyse_symbol(
                    symbol,
                    df,
                )
            )

        except Exception as error:
            print(f"  Skipped: {error}")

    write_results(results)

    print("=" * 72)
    print(f"Stocks completed: {len(results)}")

    if results:
        strongest = max(
            results,
            key=lambda item: item["Trend Score"],
        )

        weakest = min(
            results,
            key=lambda item: item["Trend Score"],
        )

        print(
            f"Strongest trend: {strongest['Symbol']} "
            f"({strongest['Trend Score']})"
        )

        print(
            f"Weakest trend: {weakest['Symbol']} "
            f"({weakest['Trend Score']})"
        )

    print(DASHBOARD)


if __name__ == "__main__":
    main()
