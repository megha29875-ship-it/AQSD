
"""
AQSD Professional
Module: Portfolio Price & P/L Updater
Version: 2.0

Updates CMP, P/L, P/L %, status and portfolio summary
inside Dashboard.xlsx using Yahoo Finance prices.
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
BLUE = "D9EAF7"


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[7]
        if cell.value is not None
    }


def download_prices(symbols: list[str]) -> dict[str, float]:
    if not symbols:
        return {}

    data = yf.download(
        tickers=symbols,
        period="5d",
        interval="1d",
        group_by="ticker",
        auto_adjust=True,
        progress=False,
        threads=True,
    )

    prices: dict[str, float] = {}

    if len(symbols) == 1:
        symbol = symbols[0]

        if not data.empty and "Close" in data.columns:
            close = data["Close"].dropna()

            if not close.empty:
                prices[symbol] = float(close.iloc[-1])

        return prices

    for symbol in symbols:
        try:
            close = data[symbol]["Close"].dropna()

            if not close.empty:
                prices[symbol] = float(close.iloc[-1])

        except (KeyError, TypeError):
            continue

    return prices


def calculate_status(
    side: str,
    cmp_price: float,
    stop_loss: float | None,
    target: float | None,
) -> str:
    side = side.upper().strip()

    if stop_loss is None or target is None:
        return "OPEN"

    if side in {"BUY", "CALL"}:
        if cmp_price <= stop_loss:
            return "STOPPED"
        if cmp_price >= target:
            return "TARGET"

    if side in {"SELL", "PUT"}:
        if cmp_price >= stop_loss:
            return "STOPPED"
        if cmp_price <= target:
            return "TARGET"

    return "OPEN"


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(f"Dashboard not found:\n{DASHBOARD}")

    wb = load_workbook(DASHBOARD)

    if "Portfolio" not in wb.sheetnames:
        raise RuntimeError(
            "Portfolio sheet not found. Run portfolio_manager.py first."
        )

    ws = wb["Portfolio"]
    headers = header_map(ws)

    required = [
        "Symbol",
        "Side",
        "Qty",
        "Entry",
        "CMP",
        "Stop Loss",
        "Capital Used",
        "Risk Amount",
        "P/L",
        "P/L %",
        "Target",
        "Status",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            "Missing Portfolio columns: " + ", ".join(missing)
        )

    active_rows: list[int] = []
    symbols: list[str] = []

    for row in range(8, ws.max_row + 1):
        symbol = str(
            ws.cell(row, headers["Symbol"]).value or ""
        ).strip().upper()

        qty = ws.cell(row, headers["Qty"]).value
        entry = ws.cell(row, headers["Entry"]).value

        if symbol and qty and entry:
            if not symbol.endswith(".NS"):
                symbol += ".NS"
                ws.cell(row, headers["Symbol"]).value = symbol

            active_rows.append(row)
            symbols.append(symbol)

    prices = download_prices(sorted(set(symbols)))

    total_invested = 0.0
    total_pl = 0.0
    open_positions = 0
    updated = 0

    for row in active_rows:
        symbol = str(ws.cell(row, headers["Symbol"]).value).strip()
        side = str(
            ws.cell(row, headers["Side"]).value or "BUY"
        ).upper().strip()

        qty = float(ws.cell(row, headers["Qty"]).value)
        entry = float(ws.cell(row, headers["Entry"]).value)

        stop_value = ws.cell(row, headers["Stop Loss"]).value
        target_value = ws.cell(row, headers["Target"]).value

        stop_loss = float(stop_value) if stop_value not in (None, "") else None
        target = float(target_value) if target_value not in (None, "") else None

        cmp_price = prices.get(symbol)

        if cmp_price is None:
            ws.cell(row, headers["Status"]).value = "PRICE ERROR"
            ws.cell(row, headers["Status"]).fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )
            continue

        capital_used = qty * entry
        risk_amount = (
            qty * abs(entry - stop_loss)
            if stop_loss is not None
            else 0.0
        )

        if side in {"BUY", "CALL"}:
            profit_loss = (cmp_price - entry) * qty
        else:
            profit_loss = (entry - cmp_price) * qty

        pl_percent = (
            profit_loss / capital_used * 100
            if capital_used
            else 0.0
        )

        status = calculate_status(
            side,
            cmp_price,
            stop_loss,
            target,
        )

        ws.cell(row, headers["CMP"]).value = round(cmp_price, 2)
        ws.cell(row, headers["Capital Used"]).value = round(capital_used, 2)
        ws.cell(row, headers["Risk Amount"]).value = round(risk_amount, 2)
        ws.cell(row, headers["P/L"]).value = round(profit_loss, 2)
        ws.cell(row, headers["P/L %"]).value = round(pl_percent, 2)
        ws.cell(row, headers["Status"]).value = status

        pl_fill = GREEN if profit_loss >= 0 else RED

        ws.cell(row, headers["P/L"]).fill = PatternFill(
            "solid",
            fgColor=pl_fill,
        )
        ws.cell(row, headers["P/L %"]).fill = PatternFill(
            "solid",
            fgColor=pl_fill,
        )

        if status == "TARGET":
            status_fill = GREEN
        elif status == "STOPPED":
            status_fill = RED
        else:
            status_fill = YELLOW

        ws.cell(row, headers["Status"]).fill = PatternFill(
            "solid",
            fgColor=status_fill,
        )
        ws.cell(row, headers["Status"]).font = Font(bold=True)

        total_invested += capital_used
        total_pl += profit_loss

        if status == "OPEN":
            open_positions += 1

        updated += 1

    trading_capital = float(ws["B4"].value or 0)
    available_cash = trading_capital - total_invested
    portfolio_return = (
        total_pl / total_invested * 100
        if total_invested
        else 0.0
    )

    ws["B5"] = round(total_invested, 2)
    ws["B6"] = round(available_cash, 2)
    ws["B7"] = open_positions
    ws["B8"] = round(total_pl, 2)
    ws["B9"] = round(portfolio_return, 2)

    ws["D4"] = "Last Price Update"
    ws["E4"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["D4"].font = Font(bold=True)
    ws["D4"].fill = PatternFill("solid", fgColor=BLUE)

    ws["B5"].number_format = '₹#,##0.00'
    ws["B6"].number_format = '₹#,##0.00'
    ws["B8"].number_format = '₹#,##0.00'
    ws["B9"].number_format = '0.00"%"'

    wb.save(DASHBOARD)

    print("Portfolio updated successfully.")
    print(f"Positions updated: {updated}")
    print(f"Capital invested: ₹{total_invested:,.2f}")
    print(f"Total P/L: ₹{total_pl:,.2f}")
    print(f"Open positions: {open_positions}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
