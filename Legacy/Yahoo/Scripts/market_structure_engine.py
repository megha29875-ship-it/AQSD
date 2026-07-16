
"""
AQSD Professional
Market Structure Intelligence Engine v1.0
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"
FNO_FILE = BASE / "Data" / "FnO_Stocks.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


def normalize_symbol(value) -> str:
    symbol = str(value or "").strip().upper()
    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"
    return symbol


def load_symbols(limit: int) -> list[str]:
    symbols: list[str] = []

    if DASHBOARD.exists():
        try:
            wb = load_workbook(DASHBOARD, read_only=True, data_only=True)
            for sheet_name in ("CALL Candidates", "PUT Candidates"):
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                headers = {
                    str(cell.value).strip(): cell.column
                    for cell in ws[1]
                    if cell.value is not None
                }
                symbol_col = headers.get("Symbol")
                if not symbol_col:
                    continue

                for row in range(2, ws.max_row + 1):
                    symbol = normalize_symbol(ws.cell(row, symbol_col).value)
                    if symbol and symbol not in symbols:
                        symbols.append(symbol)
                    if len(symbols) >= limit:
                        wb.close()
                        return symbols
            wb.close()
        except Exception:
            pass

    if FNO_FILE.exists() and len(symbols) < limit:
        try:
            df = pd.read_excel(FNO_FILE)
            symbol_col = next(
                (
                    col
                    for col in ("Yahoo Symbol", "Symbol", "SYMBOL", "Ticker")
                    if col in df.columns
                ),
                None,
            )
            if symbol_col:
                for value in df[symbol_col].dropna():
                    symbol = normalize_symbol(value)
                    if symbol and symbol not in symbols:
                        symbols.append(symbol)
                    if len(symbols) >= limit:
                        return symbols
        except Exception:
            pass

    fallback = [
        "RELIANCE.NS", "HDFCBANK.NS", "ICICIBANK.NS", "SBIN.NS",
        "INFY.NS", "TCS.NS", "LT.NS", "SUNPHARMA.NS",
        "BIOCON.NS", "TATAMOTORS.NS",
    ]

    for symbol in fallback:
        if symbol not in symbols:
            symbols.append(symbol)
        if len(symbols) >= limit:
            break

    return symbols


def download_ohlc(symbol: str, period: str) -> pd.DataFrame:
    df = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
    )

    if df.empty:
        raise RuntimeError("No price data")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    required = ["High", "Low", "Close"]
    if any(column not in df.columns for column in required):
        raise RuntimeError("OHLC columns missing")

    return df.dropna(subset=required).copy()


def detect_swings(df: pd.DataFrame, window: int):
    swing_highs = []
    swing_lows = []

    for i in range(window, len(df) - window):
        current_high = float(df["High"].iloc[i])
        current_low = float(df["Low"].iloc[i])

        left_high = float(df["High"].iloc[i-window:i].max())
        right_high = float(df["High"].iloc[i+1:i+window+1].max())
        left_low = float(df["Low"].iloc[i-window:i].min())
        right_low = float(df["Low"].iloc[i+1:i+window+1].min())

        if current_high > left_high and current_high >= right_high:
            swing_highs.append((df.index[i], current_high))

        if current_low < left_low and current_low <= right_low:
            swing_lows.append((df.index[i], current_low))

    return swing_highs, swing_lows


def pct_distance(current: float, reference):
    if reference in (None, 0):
        return None
    return round((current - reference) / reference * 100, 2)


def analyse_symbol(symbol: str, df: pd.DataFrame, window: int) -> dict:
    swing_highs, swing_lows = detect_swings(df, window)

    close = float(df["Close"].iloc[-1])
    previous_close = float(df["Close"].iloc[-2]) if len(df) >= 2 else close

    latest_high = swing_highs[-1][1] if swing_highs else None
    previous_high = swing_highs[-2][1] if len(swing_highs) >= 2 else None
    latest_low = swing_lows[-1][1] if swing_lows else None
    previous_low = swing_lows[-2][1] if len(swing_lows) >= 2 else None

    high_structure = "UNCONFIRMED"
    low_structure = "UNCONFIRMED"

    if latest_high is not None and previous_high is not None:
        high_structure = "HIGHER HIGH" if latest_high > previous_high else "LOWER HIGH"

    if latest_low is not None and previous_low is not None:
        low_structure = "HIGHER LOW" if latest_low > previous_low else "LOWER LOW"

    event = "NONE"

    bullish_break = (
        latest_high is not None
        and close > latest_high
        and previous_close <= latest_high
    )
    bearish_break = (
        latest_low is not None
        and close < latest_low
        and previous_close >= latest_low
    )

    if bullish_break:
        event = "BULLISH CHOCH" if low_structure == "LOWER LOW" else "BULLISH BOS"
    elif bearish_break:
        event = "BEARISH CHOCH" if high_structure == "HIGHER HIGH" else "BEARISH BOS"

    if high_structure == "HIGHER HIGH" and low_structure == "HIGHER LOW":
        phase = "MARKUP"
    elif high_structure == "LOWER HIGH" and low_structure == "LOWER LOW":
        phase = "MARKDOWN"
    elif high_structure == "LOWER HIGH" and low_structure == "HIGHER LOW":
        phase = "COMPRESSION"
    elif high_structure == "HIGHER HIGH" and low_structure == "LOWER LOW":
        phase = "EXPANSION / VOLATILE"
    else:
        phase = "TRANSITION"

    score = 50
    reasons = []

    if high_structure == "HIGHER HIGH":
        score += 15
        reasons.append("Higher High")
    elif high_structure == "LOWER HIGH":
        score -= 15
        reasons.append("Lower High")

    if low_structure == "HIGHER LOW":
        score += 15
        reasons.append("Higher Low")
    elif low_structure == "LOWER LOW":
        score -= 15
        reasons.append("Lower Low")

    if event == "BULLISH BOS":
        score += 15
        reasons.append("Bullish BOS")
    elif event == "BEARISH BOS":
        score -= 15
        reasons.append("Bearish BOS")
    elif event == "BULLISH CHOCH":
        score += 10
        reasons.append("Bullish CHOCH")
    elif event == "BEARISH CHOCH":
        score -= 10
        reasons.append("Bearish CHOCH")

    score = max(0, min(100, score))

    return {
        "Symbol": symbol,
        "Close": round(close, 2),
        "Latest Swing High": round(latest_high, 2) if latest_high is not None else None,
        "Latest Swing Low": round(latest_low, 2) if latest_low is not None else None,
        "Previous Swing High": round(previous_high, 2) if previous_high is not None else None,
        "Previous Swing Low": round(previous_low, 2) if previous_low is not None else None,
        "High Structure": high_structure,
        "Low Structure": low_structure,
        "Structure Event": event,
        "Trend Phase": phase,
        "Structure Score": score,
        "Distance to Swing High %": pct_distance(close, latest_high),
        "Distance to Swing Low %": pct_distance(close, latest_low),
        "Reason": " | ".join(reasons) if reasons else "Structure forming",
    }


def write_results(results: list[dict]) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Market Structure" in wb.sheetnames:
        del wb["Market Structure"]

    ws = wb.create_sheet("Market Structure", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:N2")
    ws["A1"] = "AQSD PROFESSIONAL - MARKET STRUCTURE INTELLIGENCE"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A4"] = "Stocks Analysed"
    ws["B4"] = len(results)
    ws["A4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE)

    headers = [
        "Symbol", "Close", "Latest Swing High", "Latest Swing Low",
        "Previous Swing High", "Previous Swing Low", "High Structure",
        "Low Structure", "Structure Event", "Trend Phase",
        "Structure Score", "Distance to Swing High %",
        "Distance to Swing Low %", "Reason",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(6, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)

    ranked = sorted(results, key=lambda item: item["Structure Score"], reverse=True)

    for row_no, result in enumerate(ranked, start=7):
        for col, heading in enumerate(headers, start=1):
            cell = ws.cell(row_no, col, result.get(heading))
            cell.border = Border(bottom=THIN)

        for col in range(2, 7):
            ws.cell(row_no, col).number_format = '₹#,##0.00'

        for col in (12, 13):
            ws.cell(row_no, col).number_format = '0.00"%"'

        score = result["Structure Score"]
        score_fill = GREEN if score >= 70 else RED if score <= 30 else YELLOW
        ws.cell(row_no, 11).fill = PatternFill("solid", fgColor=score_fill)
        ws.cell(row_no, 11).font = Font(bold=True)

        event = result["Structure Event"]
        event_fill = GREEN if "BULLISH" in event else RED if "BEARISH" in event else GREY
        ws.cell(row_no, 9).fill = PatternFill("solid", fgColor=event_fill)

    widths = {
        "A": 18, "B": 12, "C": 16, "D": 16, "E": 18, "F": 18,
        "G": 16, "H": 16, "I": 18, "J": 20, "K": 15,
        "L": 18, "M": 18, "N": 42,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError("Close Dashboard.xlsx in Excel and run again.") from error


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build AQSD market-structure intelligence."
    )
    parser.add_argument("--period", default="1y")
    parser.add_argument("--limit", type=int, default=30)
    parser.add_argument("--swing-window", type=int, default=3)
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    symbols = load_symbols(args.limit)

    print("\nAQSD MARKET STRUCTURE INTELLIGENCE")
    print("=" * 72)
    print(f"Symbols: {len(symbols)}")
    print(f"Period: {args.period}")
    print(f"Swing window: {args.swing_window}")

    results = []

    for index, symbol in enumerate(symbols, start=1):
        print(f"[{index}/{len(symbols)}] {symbol}")

        try:
            df = download_ohlc(symbol, args.period)
            results.append(analyse_symbol(symbol, df, args.swing_window))
        except Exception as error:
            print(f"  Skipped: {error}")

    write_results(results)

    print("=" * 72)
    print(f"Stocks completed: {len(results)}")

    if results:
        best = max(results, key=lambda item: item["Structure Score"])
        worst = min(results, key=lambda item: item["Structure Score"])

        print(
            f"Strongest structure: {best['Symbol']} "
            f"({best['Structure Score']})"
        )
        print(
            f"Weakest structure: {worst['Symbol']} "
            f"({worst['Structure Score']})"
        )

    print(DASHBOARD)


if __name__ == "__main__":
    main()
