
"""
AQSD Professional
Module: Strategy Optimizer
Version: 1.0

Tests multiple EMA, RSI, ATR-stop and reward:risk combinations
for one NSE symbol and ranks the results.

Outputs:
- Optimizer Summary
- Optimizer Trades

Example:
    python strategy_optimizer.py --symbol RELIANCE.NS --period 5y
"""

from __future__ import annotations

import argparse
from itertools import product
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

INITIAL_CAPITAL = 200000
RISK_PERCENT = 1.0
MAX_HOLD_DAYS = 10

EMA_FAST_VALUES = [10, 20]
EMA_SLOW_VALUES = [50, 100]
RSI_TRIGGER_VALUES = [50, 55, 60]
ATR_MULTIPLIERS = [1.5, 2.0, 2.5]
TARGET_RR_VALUES = [1.5, 2.0, 3.0]


def calculate_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    avg_gain = gain.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    avg_loss = loss.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def calculate_atr(df: pd.DataFrame, period: int = 14) -> pd.Series:
    previous_close = df["Close"].shift(1)

    true_range = pd.concat(
        [
            df["High"] - df["Low"],
            (df["High"] - previous_close).abs(),
            (df["Low"] - previous_close).abs(),
        ],
        axis=1,
    ).max(axis=1)

    return true_range.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()


def download_data(symbol: str, period: str) -> pd.DataFrame:
    df = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
    )

    if df.empty:
        raise RuntimeError(f"No data downloaded for {symbol}")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df = df.dropna(subset=["Open", "High", "Low", "Close"])
    df["RSI14"] = calculate_rsi(df["Close"], 14)
    df["ATR14"] = calculate_atr(df, 14)

    return df.dropna()


def quantity_for_trade(
    capital: float,
    entry: float,
    stop_loss: float,
) -> int:
    risk_amount = capital * (RISK_PERCENT / 100)
    risk_per_share = abs(entry - stop_loss)

    if risk_per_share <= 0:
        return 0

    qty_by_risk = int(risk_amount // risk_per_share)
    qty_by_capital = int(capital // entry)

    return max(0, min(qty_by_risk, qty_by_capital))


def run_configuration(
    symbol: str,
    df: pd.DataFrame,
    ema_fast: int,
    ema_slow: int,
    rsi_trigger: int,
    atr_multiplier: float,
    target_rr: float,
) -> tuple[dict, list[dict]]:
    working = df.copy()

    fast_col = f"EMA{ema_fast}"
    slow_col = f"EMA{ema_slow}"

    working[fast_col] = working["Close"].ewm(
        span=ema_fast,
        adjust=False,
    ).mean()

    working[slow_col] = working["Close"].ewm(
        span=ema_slow,
        adjust=False,
    ).mean()

    working = working.dropna()

    bullish = (
        (working[fast_col] > working[slow_col])
        & (working["RSI14"] > rsi_trigger)
        & (working["Close"] > working[fast_col])
    )

    signals = bullish & (~bullish.shift(1).fillna(False))

    capital = INITIAL_CAPITAL
    trades: list[dict] = []
    i = 0

    while i < len(working) - 1:
        if not bool(signals.iloc[i]):
            i += 1
            continue

        entry_index = i + 1
        entry_row = working.iloc[entry_index]

        entry = float(entry_row["Open"])
        atr_value = float(working.iloc[i]["ATR14"])

        stop_loss = entry - atr_multiplier * atr_value
        target = entry + target_rr * (entry - stop_loss)

        quantity = quantity_for_trade(
            capital,
            entry,
            stop_loss,
        )

        if quantity <= 0:
            i += 1
            continue

        final_index = min(
            entry_index + MAX_HOLD_DAYS,
            len(working) - 1,
        )

        exit_index = final_index
        exit_price = float(working.iloc[final_index]["Close"])
        result = "TIME EXIT"

        for j in range(entry_index, final_index + 1):
            row = working.iloc[j]

            if float(row["Low"]) <= stop_loss:
                exit_index = j
                exit_price = stop_loss
                result = "LOSS"
                break

            if float(row["High"]) >= target:
                exit_index = j
                exit_price = target
                result = "WIN"
                break

        pnl = (exit_price - entry) * quantity
        capital += pnl

        trades.append(
            {
                "Symbol": symbol,
                "Entry Date": str(working.index[entry_index].date()),
                "Exit Date": str(working.index[exit_index].date()),
                "Entry": round(entry, 2),
                "Stop Loss": round(stop_loss, 2),
                "Target": round(target, 2),
                "Exit": round(exit_price, 2),
                "Quantity": quantity,
                "P/L": round(pnl, 2),
                "Result": result,
                "EMA Fast": ema_fast,
                "EMA Slow": ema_slow,
                "RSI Trigger": rsi_trigger,
                "ATR Multiplier": atr_multiplier,
                "Target RR": target_rr,
            }
        )

        i = exit_index + 1

    wins = [trade for trade in trades if trade["P/L"] > 0]
    losses = [trade for trade in trades if trade["P/L"] < 0]

    gross_profit = sum(trade["P/L"] for trade in wins)
    gross_loss = abs(sum(trade["P/L"] for trade in losses))
    net_profit = sum(trade["P/L"] for trade in trades)

    win_rate = (
        len(wins) / len(trades) * 100
        if trades
        else 0.0
    )

    profit_factor = (
        gross_profit / gross_loss
        if gross_loss
        else gross_profit
    )

    equity = INITIAL_CAPITAL
    peak = INITIAL_CAPITAL
    max_drawdown = 0.0

    for trade in trades:
        equity += trade["P/L"]
        peak = max(peak, equity)
        max_drawdown = min(
            max_drawdown,
            equity - peak,
        )

    score = (
        net_profit
        + profit_factor * 1000
        - abs(max_drawdown) * 0.5
    )

    summary = {
        "EMA Fast": ema_fast,
        "EMA Slow": ema_slow,
        "RSI Trigger": rsi_trigger,
        "ATR Multiplier": atr_multiplier,
        "Target RR": target_rr,
        "Trades": len(trades),
        "Wins": len(wins),
        "Losses": len(losses),
        "Win Rate %": round(win_rate, 2),
        "Net Profit": round(net_profit, 2),
        "Return %": round(
            net_profit / INITIAL_CAPITAL * 100,
            2,
        ),
        "Profit Factor": round(profit_factor, 2),
        "Max Drawdown": round(abs(max_drawdown), 2),
        "Final Capital": round(capital, 2),
        "Optimizer Score": round(score, 2),
    }

    return summary, trades


def write_results(
    symbol: str,
    summaries: list[dict],
    all_trades: list[dict],
) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_name in (
        "Optimizer Summary",
        "Optimizer Trades",
    ):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    summary_ws = wb.create_sheet("Optimizer Summary")
    trades_ws = wb.create_sheet("Optimizer Trades")

    navy = "17365D"
    white = "FFFFFF"
    green = "C6EFCE"

    summary_ws.merge_cells("A1:O2")
    summary_ws["A1"] = f"AQSD STRATEGY OPTIMIZER - {symbol}"
    summary_ws["A1"].font = Font(
        size=18,
        bold=True,
        color=white,
    )
    summary_ws["A1"].fill = PatternFill(
        "solid",
        fgColor=navy,
    )
    summary_ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    headers = [
        "Rank",
        "EMA Fast",
        "EMA Slow",
        "RSI Trigger",
        "ATR Multiplier",
        "Target RR",
        "Trades",
        "Wins",
        "Losses",
        "Win Rate %",
        "Net Profit",
        "Return %",
        "Profit Factor",
        "Max Drawdown",
        "Final Capital",
        "Optimizer Score",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = summary_ws.cell(4, col, heading)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)

    ranked = sorted(
        summaries,
        key=lambda item: item["Optimizer Score"],
        reverse=True,
    )

    for row_no, item in enumerate(ranked, start=5):
        values = [
            row_no - 4,
            item["EMA Fast"],
            item["EMA Slow"],
            item["RSI Trigger"],
            item["ATR Multiplier"],
            item["Target RR"],
            item["Trades"],
            item["Wins"],
            item["Losses"],
            item["Win Rate %"],
            item["Net Profit"],
            item["Return %"],
            item["Profit Factor"],
            item["Max Drawdown"],
            item["Final Capital"],
            item["Optimizer Score"],
        ]

        for col, value in enumerate(values, start=1):
            summary_ws.cell(row_no, col, value)

        for col in (11, 14, 15, 16):
            summary_ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

        for col in (10, 12):
            summary_ws.cell(
                row_no,
                col,
            ).number_format = '0.00"%"'

        if row_no == 5:
            for col in range(1, len(headers) + 1):
                summary_ws.cell(
                    row_no,
                    col,
                ).fill = PatternFill(
                    "solid",
                    fgColor=green,
                )

    if ranked:
        last_row = 4 + min(len(ranked), 15)

        chart = BarChart()
        chart.title = "Top Strategy Net Profit"
        chart.y_axis.title = "Net Profit"
        chart.x_axis.title = "Strategy Rank"

        chart.add_data(
            Reference(
                summary_ws,
                min_col=11,
                min_row=4,
                max_row=last_row,
            ),
            titles_from_data=True,
        )

        chart.set_categories(
            Reference(
                summary_ws,
                min_col=1,
                min_row=5,
                max_row=last_row,
            )
        )

        chart.height = 8
        chart.width = 14
        summary_ws.add_chart(chart, "R4")

    trade_headers = [
        "Trade No.",
        "Symbol",
        "Entry Date",
        "Exit Date",
        "Entry",
        "Stop Loss",
        "Target",
        "Exit",
        "Quantity",
        "P/L",
        "Result",
        "EMA Fast",
        "EMA Slow",
        "RSI Trigger",
        "ATR Multiplier",
        "Target RR",
    ]

    for col, heading in enumerate(trade_headers, start=1):
        cell = trades_ws.cell(1, col, heading)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)

    for row_no, trade in enumerate(all_trades, start=2):
        values = [
            row_no - 1,
            trade["Symbol"],
            trade["Entry Date"],
            trade["Exit Date"],
            trade["Entry"],
            trade["Stop Loss"],
            trade["Target"],
            trade["Exit"],
            trade["Quantity"],
            trade["P/L"],
            trade["Result"],
            trade["EMA Fast"],
            trade["EMA Slow"],
            trade["RSI Trigger"],
            trade["ATR Multiplier"],
            trade["Target RR"],
        ]

        for col, value in enumerate(values, start=1):
            trades_ws.cell(row_no, col, value)

        for col in (5, 6, 7, 8, 10):
            trades_ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

    summary_ws.freeze_panes = "A5"
    trades_ws.freeze_panes = "A2"

    wb.save(DASHBOARD)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Optimize AQSD swing-strategy parameters."
    )

    parser.add_argument(
        "--symbol",
        default="RELIANCE.NS",
        help="Yahoo Finance symbol.",
    )

    parser.add_argument(
        "--period",
        default="5y",
        help="Yahoo Finance period, e.g. 3y or 5y.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    symbol = args.symbol.upper().strip()

    if not symbol.endswith(".NS"):
        symbol += ".NS"

    df = download_data(symbol, args.period)

    combinations = list(
        product(
            EMA_FAST_VALUES,
            EMA_SLOW_VALUES,
            RSI_TRIGGER_VALUES,
            ATR_MULTIPLIERS,
            TARGET_RR_VALUES,
        )
    )

    summaries: list[dict] = []
    all_trades: list[dict] = []

    print("\nAQSD STRATEGY OPTIMIZER")
    print("=" * 68)
    print(f"Symbol: {symbol}")
    print(f"Period: {args.period}")
    print(f"Configurations: {len(combinations)}")

    for index, params in enumerate(combinations, start=1):
        (
            ema_fast,
            ema_slow,
            rsi_trigger,
            atr_multiplier,
            target_rr,
        ) = params

        print(
            f"[{index}/{len(combinations)}] "
            f"EMA {ema_fast}/{ema_slow}, "
            f"RSI {rsi_trigger}, "
            f"ATR {atr_multiplier}, "
            f"RR {target_rr}"
        )

        summary, trades = run_configuration(
            symbol,
            df,
            ema_fast,
            ema_slow,
            rsi_trigger,
            atr_multiplier,
            target_rr,
        )

        summaries.append(summary)
        all_trades.extend(trades)

    write_results(
        symbol,
        summaries,
        all_trades,
    )

    best = max(
        summaries,
        key=lambda item: item["Optimizer Score"],
    )

    print("=" * 68)
    print("Best configuration")
    print(
        f"EMA {best['EMA Fast']}/{best['EMA Slow']} | "
        f"RSI {best['RSI Trigger']} | "
        f"ATR {best['ATR Multiplier']} | "
        f"RR {best['Target RR']}"
    )
    print(f"Net profit: ₹{best['Net Profit']:,.2f}")
    print(f"Win rate: {best['Win Rate %']:.2f}%")
    print(f"Profit factor: {best['Profit Factor']:.2f}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
