
"""
AQSD Professional
Module: Market Pulse
Version: 1.0
"""

from pathlib import Path
from datetime import datetime

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BASE = Path(__file__).resolve().parent.parent
OUTPUT_FILE = BASE / "Output" / "Dashboard.xlsx"
OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

MARKETS = {
    "NIFTY": "^NSEI",
    "BANKNIFTY": "^NSEBANK",
    "INDIA VIX": "^INDIAVIX",
}


def calculate_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    avg_gain = gain.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    avg_loss = loss.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))

    rsi.loc[(avg_loss == 0) & (avg_gain > 0)] = 100
    rsi.loc[(avg_gain == 0) & (avg_loss > 0)] = 0

    return rsi


def get_market_data(symbol: str) -> pd.DataFrame | None:
    df = yf.download(
        symbol,
        period="6mo",
        interval="1d",
        progress=False,
        auto_adjust=True,
    )

    if df.empty:
        return None

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df = df.dropna(subset=["Close"])

    if len(df) < 50:
        return None

    df["EMA20"] = df["Close"].ewm(span=20, adjust=False).mean()
    df["EMA50"] = df["Close"].ewm(span=50, adjust=False).mean()
    df["RSI14"] = calculate_rsi(df["Close"], period=14)
    df["Daily %"] = df["Close"].pct_change() * 100
    df["Weekly %"] = df["Close"].pct_change(5) * 100

    return df


def classify_index(
    close: float,
    ema20: float,
    ema50: float,
    rsi: float,
) -> tuple[str, int]:
    score = 0

    if close > ema20:
        score += 20

    if ema20 > ema50:
        score += 20

    if rsi >= 55:
        score += 20
    elif rsi >= 50:
        score += 10

    if close > ema20 > ema50 and rsi >= 55:
        trend = "Bullish"
    elif close < ema20 < ema50 and rsi < 45:
        trend = "Bearish"
    else:
        trend = "Neutral"

    return trend, score


def classify_vix(
    close: float,
    ema20: float,
    ema50: float,
) -> tuple[str, int]:
    if close < ema20 and ema20 < ema50:
        return "Falling", 20

    if close > ema20 and ema20 > ema50:
        return "Rising", 0

    return "Neutral", 10


def final_market_view(score: int) -> tuple[str, str]:
    if score >= 80:
        return "BULLISH", "CALL BUYING"

    if score <= 35:
        return "BEARISH", "PUT BUYING"

    return "NEUTRAL", "NO TRADE / WAIT"


def save_market_pulse(
    results: list[list],
    bias: str,
    confidence: int,
    strategy: str,
) -> None:
    if OUTPUT_FILE.exists():
        wb = load_workbook(OUTPUT_FILE)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Market Pulse" in wb.sheetnames:
        del wb["Market Pulse"]

    ws = wb.create_sheet("Market Pulse", 0)

    navy = "1F4E78"
    blue = "D9EAF7"
    green = "C6EFCE"
    yellow = "FFF2CC"
    red = "FFC7CE"
    white = "FFFFFF"
    grey = "D9E1F2"

    thin = Side(style="thin", color="D9D9D9")

    ws["A1"] = "AQSD PROFESSIONAL - MARKET PULSE"
    ws["A1"].font = Font(size=18, bold=True, color=white)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Last Updated"
    ws["B2"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    headers = [
        "Index",
        "Close",
        "EMA20",
        "EMA50",
        "RSI14",
        "Daily %",
        "Weekly %",
        "Trend",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=heading)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=grey)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    for row_index, row_data in enumerate(results, start=5):
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = Border(bottom=thin)

        trend = str(ws.cell(row=row_index, column=8).value)

        if trend in {"Bullish", "Falling"}:
            fill_color = green
        elif trend in {"Bearish", "Rising"}:
            fill_color = red
        else:
            fill_color = yellow

        ws.cell(row=row_index, column=8).fill = PatternFill(
            fill_type="solid",
            fgColor=fill_color,
        )
        ws.cell(row=row_index, column=8).font = Font(bold=True)

    summary_row = 10

    ws[f"A{summary_row}"] = "Market Bias"
    ws[f"B{summary_row}"] = bias
    ws[f"A{summary_row + 1}"] = "Confidence"
    ws[f"B{summary_row + 1}"] = f"{confidence}%"
    ws[f"A{summary_row + 2}"] = "Strategy"
    ws[f"B{summary_row + 2}"] = strategy

    for row in range(summary_row, summary_row + 3):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(fill_type="solid", fgColor=blue)

    if bias == "BULLISH":
        bias_fill = green
    elif bias == "BEARISH":
        bias_fill = red
    else:
        bias_fill = yellow

    ws[f"B{summary_row}"].fill = PatternFill(
        fill_type="solid",
        fgColor=bias_fill,
    )
    ws[f"B{summary_row}"].font = Font(bold=True)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:H7"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 16,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 16,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=5, max_row=7, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = "0.00"

    wb.save(OUTPUT_FILE)

    print("\nMarket Pulse saved to:")
    print(OUTPUT_FILE)


def main() -> None:
    print("\nAQSD MARKET PULSE\n")

    results = []
    market_score = 0

    for name, symbol in MARKETS.items():
        print(f"Downloading {name}...")

        df = get_market_data(symbol)

        if df is None:
            print(f"Failed to download {name}\n")
            continue

        last = df.iloc[-1]

        close = float(last["Close"])
        ema20 = float(last["EMA20"])
        ema50 = float(last["EMA50"])
        rsi = float(last["RSI14"])
        daily_change = float(last["Daily %"])
        weekly_change = float(last["Weekly %"])

        if name == "INDIA VIX":
            trend, score = classify_vix(close, ema20, ema50)
        else:
            trend, score = classify_index(close, ema20, ema50, rsi)

        market_score += score

        results.append(
            [
                name,
                round(close, 2),
                round(ema20, 2),
                round(ema50, 2),
                round(rsi, 2),
                round(daily_change, 2),
                round(weekly_change, 2),
                trend,
            ]
        )

        print("-" * 35)
        print(name)
        print(f"Close    : {close:.2f}")
        print(f"EMA20    : {ema20:.2f}")
        print(f"EMA50    : {ema50:.2f}")
        print(f"RSI14    : {rsi:.2f}")
        print(f"Daily %  : {daily_change:.2f}")
        print(f"Weekly % : {weekly_change:.2f}")
        print(f"Trend    : {trend}\n")

    if len(results) < 3:
        print("Market Pulse could not download all required markets.")
        raise SystemExit(1)

    confidence = min(market_score, 100)
    bias, strategy = final_market_view(confidence)

    print("=" * 45)
    print("AQSD MARKET PULSE")
    print("=" * 45)
    print(f"Market Bias : {bias}")
    print(f"Confidence  : {confidence}%")
    print(f"Strategy    : {strategy}")
    print("=" * 45)

    save_market_pulse(
        results,
        bias,
        confidence,
        strategy,
    )


if __name__ == "__main__":
    main()
