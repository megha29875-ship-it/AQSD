
"""
AQSD Market Intelligence
Module: Global Markets & Commodities Engine
Version: 1.0

Downloads major global indices, volatility, currencies and commodities,
stores daily snapshots in aqsd_core.db, and creates a "Global Intelligence"
sheet in Dashboard.xlsx.

Commands
--------
python aqsd_global_intelligence.py --update
python aqsd_global_intelligence.py --status
python aqsd_global_intelligence.py --report
python aqsd_global_intelligence.py --update --period 3mo

Notes
-----
- Yahoo Finance is used as the initial data source.
- This module is designed so the data source can be replaced later.
- Market data may be delayed and should be treated as decision support,
  not as an execution feed.
"""

from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# ASSET UNIVERSE
# ============================================================

GLOBAL_ASSETS = {
    "^DJI": {
        "name": "Dow Jones",
        "asset_type": "Index",
        "group": "US Equity",
    },
    "^IXIC": {
        "name": "Nasdaq Composite",
        "asset_type": "Index",
        "group": "US Equity",
    },
    "^GSPC": {
        "name": "S&P 500",
        "asset_type": "Index",
        "group": "US Equity",
    },
    "^VIX": {
        "name": "CBOE VIX",
        "asset_type": "Volatility",
        "group": "Risk",
    },
    "^FTSE": {
        "name": "FTSE 100",
        "asset_type": "Index",
        "group": "Europe",
    },
    "^GDAXI": {
        "name": "DAX",
        "asset_type": "Index",
        "group": "Europe",
    },
    "^N225": {
        "name": "Nikkei 225",
        "asset_type": "Index",
        "group": "Asia",
    },
    "^HSI": {
        "name": "Hang Seng",
        "asset_type": "Index",
        "group": "Asia",
    },
    "DX-Y.NYB": {
        "name": "US Dollar Index",
        "asset_type": "Currency Index",
        "group": "Macro",
    },
    "INR=X": {
        "name": "USD/INR",
        "asset_type": "Currency",
        "group": "Macro",
    },
    "^TNX": {
        "name": "US 10Y Yield",
        "asset_type": "Yield",
        "group": "Macro",
    },
}

COMMODITIES = {
    "BZ=F": {
        "name": "Brent Crude",
        "affected_sectors": "Energy|Aviation|Paints|Tyres|Chemicals",
    },
    "CL=F": {
        "name": "WTI Crude",
        "affected_sectors": "Energy|Aviation|Paints|Tyres|Chemicals",
    },
    "NG=F": {
        "name": "Natural Gas",
        "affected_sectors": "Gas|Fertiliser|Power|Chemicals",
    },
    "GC=F": {
        "name": "Gold",
        "affected_sectors": "Jewellery|NBFC|Safe Haven",
    },
    "SI=F": {
        "name": "Silver",
        "affected_sectors": "Metals|Electronics|Solar",
    },
    "HG=F": {
        "name": "Copper",
        "affected_sectors": "Metals|Capital Goods|Power|Infrastructure",
    },
}

DEFAULT_PERIOD = "3mo"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DOWNLOAD HELPERS
# ============================================================

def flatten_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if isinstance(frame.columns, pd.MultiIndex):
        frame.columns = frame.columns.get_level_values(0)

    if "Close" not in frame.columns:
        raise RuntimeError("Close column missing")

    output = frame[["Close"]].dropna().copy()
    output.index = pd.to_datetime(output.index)

    return output


def download_asset(symbol: str, period: str) -> pd.Series:
    frame = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
        threads=False,
    )

    if frame.empty:
        raise RuntimeError("No data downloaded")

    clean = flatten_frame(frame)
    return clean["Close"]


def calculate_return(series: pd.Series, periods: int) -> float | None:
    if len(series) <= periods:
        return None

    latest = float(series.iloc[-1])
    previous = float(series.iloc[-1 - periods])

    if previous == 0:
        return None

    return round((latest / previous - 1) * 100, 2)


def latest_snapshot(series: pd.Series) -> dict[str, Any]:
    if series.empty:
        raise RuntimeError("Empty series")

    latest_value = float(series.iloc[-1])
    previous_value = (
        float(series.iloc[-2])
        if len(series) >= 2
        else latest_value
    )

    day_change = (
        0.0
        if previous_value == 0
        else (latest_value / previous_value - 1) * 100
    )

    return {
        "snapshot_date": pd.to_datetime(series.index[-1]).date().isoformat(),
        "close_value": round(latest_value, 4),
        "day_change_percent": round(day_change, 2),
        "five_day_change_percent": calculate_return(series, 5),
    }


# ============================================================
# SCORING
# ============================================================

def risk_signal(
    symbol: str,
    day_change: float | None,
    five_day_change: float | None,
) -> str:
    day_change = float(day_change or 0)
    five_day_change = float(five_day_change or 0)

    if symbol == "^VIX":
        if day_change >= 5 or five_day_change >= 10:
            return "RISK OFF"
        if day_change <= -5 or five_day_change <= -10:
            return "RISK ON"
        return "NEUTRAL"

    if symbol in {"DX-Y.NYB", "^TNX"}:
        if day_change >= 1 or five_day_change >= 2:
            return "INDIA NEGATIVE"
        if day_change <= -1 or five_day_change <= -2:
            return "INDIA POSITIVE"
        return "NEUTRAL"

    if symbol == "INR=X":
        if day_change >= 0.5 or five_day_change >= 1:
            return "RUPEE WEAK"
        if day_change <= -0.5 or five_day_change <= -1:
            return "RUPEE STRONG"
        return "NEUTRAL"

    score = day_change * 0.4 + five_day_change * 0.6

    if score >= 1.5:
        return "RISK ON"
    if score <= -1.5:
        return "RISK OFF"
    return "NEUTRAL"


def commodity_signal(
    symbol: str,
    day_change: float | None,
    five_day_change: float | None,
) -> str:
    day_change = float(day_change or 0)
    five_day_change = float(five_day_change or 0)

    momentum = day_change * 0.4 + five_day_change * 0.6

    if symbol in {"BZ=F", "CL=F", "NG=F"}:
        if momentum >= 2:
            return "INPUT COST PRESSURE"
        if momentum <= -2:
            return "INPUT COST RELIEF"
        return "NEUTRAL"

    if momentum >= 2:
        return "STRONG UP"
    if momentum <= -2:
        return "STRONG DOWN"
    return "NEUTRAL"


def global_risk_score(rows: list[dict]) -> float:
    """
    0 = strong risk-off, 100 = strong risk-on.
    """

    score = 50.0
    lookup = {row["symbol"]: row for row in rows}

    for symbol in ("^DJI", "^IXIC", "^GSPC", "^FTSE", "^GDAXI", "^N225", "^HSI"):
        row = lookup.get(symbol)
        if not row:
            continue

        five_day = float(row["five_day_change_percent"] or 0)
        score += max(-4, min(4, five_day)) * 1.5

    vix = lookup.get("^VIX")
    if vix:
        vix_five = float(vix["five_day_change_percent"] or 0)
        score -= max(-10, min(10, vix_five)) * 1.2

    dollar = lookup.get("DX-Y.NYB")
    if dollar:
        score -= float(dollar["five_day_change_percent"] or 0) * 2

    yield_row = lookup.get("^TNX")
    if yield_row:
        score -= float(yield_row["five_day_change_percent"] or 0) * 1.5

    return round(max(0, min(100, score)), 2)


# ============================================================
# DATABASE STORAGE
# ============================================================

def store_global_rows(rows: list[dict]) -> None:
    with connect() as connection:
        for row in rows:
            connection.execute(
                """
                INSERT INTO global_markets(
                    snapshot_date,
                    symbol,
                    market_name,
                    asset_type,
                    close_value,
                    day_change_percent,
                    five_day_change_percent,
                    risk_signal,
                    source,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(snapshot_date, symbol)
                DO UPDATE SET
                    market_name = excluded.market_name,
                    asset_type = excluded.asset_type,
                    close_value = excluded.close_value,
                    day_change_percent = excluded.day_change_percent,
                    five_day_change_percent = excluded.five_day_change_percent,
                    risk_signal = excluded.risk_signal,
                    source = excluded.source,
                    created_at = excluded.created_at
                """,
                (
                    row["snapshot_date"],
                    row["symbol"],
                    row["market_name"],
                    row["asset_type"],
                    row["close_value"],
                    row["day_change_percent"],
                    row["five_day_change_percent"],
                    row["risk_signal"],
                    "Yahoo Finance",
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )

        connection.commit()


def store_commodity_rows(rows: list[dict]) -> None:
    with connect() as connection:
        for row in rows:
            connection.execute(
                """
                INSERT INTO commodities(
                    snapshot_date,
                    symbol,
                    commodity_name,
                    close_value,
                    day_change_percent,
                    five_day_change_percent,
                    affected_sectors,
                    source,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(snapshot_date, symbol)
                DO UPDATE SET
                    commodity_name = excluded.commodity_name,
                    close_value = excluded.close_value,
                    day_change_percent = excluded.day_change_percent,
                    five_day_change_percent = excluded.five_day_change_percent,
                    affected_sectors = excluded.affected_sectors,
                    source = excluded.source,
                    created_at = excluded.created_at
                """,
                (
                    row["snapshot_date"],
                    row["symbol"],
                    row["commodity_name"],
                    row["close_value"],
                    row["day_change_percent"],
                    row["five_day_change_percent"],
                    row["affected_sectors"],
                    "Yahoo Finance",
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )

        connection.commit()


# ============================================================
# UPDATE ENGINE
# ============================================================

def update_global_intelligence(period: str) -> tuple[list[dict], list[dict]]:
    setup_database()

    run_id = start_run(
        "aqsd_global_intelligence",
        f"Updating global markets and commodities for {period}",
    )

    global_rows: list[dict] = []
    commodity_rows: list[dict] = []
    failures = 0

    try:
        print("\nAQSD GLOBAL MARKETS & COMMODITIES")
        print("=" * 78)

        for symbol, metadata in GLOBAL_ASSETS.items():
            print(f"{symbol:<12} {metadata['name']:<24}", end="")

            try:
                series = download_asset(symbol, period)
                snapshot = latest_snapshot(series)

                row = {
                    **snapshot,
                    "symbol": symbol,
                    "market_name": metadata["name"],
                    "asset_type": metadata["asset_type"],
                    "group": metadata["group"],
                    "risk_signal": risk_signal(
                        symbol,
                        snapshot["day_change_percent"],
                        snapshot["five_day_change_percent"],
                    ),
                }

                global_rows.append(row)
                print(" OK")

            except Exception as error:
                failures += 1
                print(f" FAILED: {error}")

        for symbol, metadata in COMMODITIES.items():
            print(f"{symbol:<12} {metadata['name']:<24}", end="")

            try:
                series = download_asset(symbol, period)
                snapshot = latest_snapshot(series)

                row = {
                    **snapshot,
                    "symbol": symbol,
                    "commodity_name": metadata["name"],
                    "affected_sectors": metadata["affected_sectors"],
                    "signal": commodity_signal(
                        symbol,
                        snapshot["day_change_percent"],
                        snapshot["five_day_change_percent"],
                    ),
                }

                commodity_rows.append(row)
                print(" OK")

            except Exception as error:
                failures += 1
                print(f" FAILED: {error}")

        store_global_rows(global_rows)
        store_commodity_rows(commodity_rows)

        finish_run(
            run_id,
            status="SUCCESS" if failures == 0 else "PARTIAL",
            records_processed=len(global_rows) + len(commodity_rows),
            errors_count=failures,
            message=(
                f"Global={len(global_rows)}; "
                f"Commodities={len(commodity_rows)}"
            ),
        )

        return global_rows, commodity_rows

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=len(global_rows) + len(commodity_rows),
            errors_count=failures + 1,
            message=str(error),
        )
        raise


# ============================================================
# EXCEL REPORT
# ============================================================

def write_dashboard(
    global_rows: list[dict],
    commodity_rows: list[dict],
) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Global Intelligence" in wb.sheetnames:
        del wb["Global Intelligence"]

    ws = wb.create_sheet("Global Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:J2")
    ws["A1"] = "AQSD PROFESSIONAL - GLOBAL & COMMODITY INTELLIGENCE"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    risk_score = global_risk_score(global_rows)

    ws["A4"] = "Global Risk Score"
    ws["B4"] = risk_score
    ws["D4"] = "Market Regime"
    ws["E4"] = (
        "RISK ON"
        if risk_score >= 60
        else "RISK OFF"
        if risk_score <= 40
        else "NEUTRAL"
    )
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill("solid", fgColor=BLUE)

    ws["B4"].fill = PatternFill(
        "solid",
        fgColor=(
            GREEN if risk_score >= 60
            else RED if risk_score <= 40
            else YELLOW
        ),
    )
    ws["B4"].font = Font(bold=True)

    global_headers = [
        "Symbol",
        "Market",
        "Group",
        "Asset Type",
        "Close",
        "Day Change %",
        "5-Day Change %",
        "Risk Signal",
        "Snapshot Date",
    ]

    for col, heading in enumerate(global_headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    for row_no, row in enumerate(global_rows, start=8):
        values = [
            row["symbol"],
            row["market_name"],
            row["group"],
            row["asset_type"],
            row["close_value"],
            row["day_change_percent"],
            row["five_day_change_percent"],
            row["risk_signal"],
            row["snapshot_date"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        ws.cell(row_no, 6).number_format = '0.00"%"'
        ws.cell(row_no, 7).number_format = '0.00"%"'

        signal = row["risk_signal"]
        ws.cell(row_no, 8).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if signal in {"RISK ON", "INDIA POSITIVE", "RUPEE STRONG"}
                else RED
                if signal in {"RISK OFF", "INDIA NEGATIVE", "RUPEE WEAK"}
                else GREY
            ),
        )

    commodity_start = 10 + len(global_rows)

    ws.cell(
        commodity_start,
        1,
        "COMMODITY INTELLIGENCE",
    ).font = Font(size=14, bold=True, color=WHITE)
    ws.cell(
        commodity_start,
        1,
    ).fill = PatternFill("solid", fgColor=NAVY)

    commodity_headers = [
        "Symbol",
        "Commodity",
        "Close",
        "Day Change %",
        "5-Day Change %",
        "Signal",
        "Affected Sectors",
        "Snapshot Date",
    ]

    for col, heading in enumerate(commodity_headers, start=1):
        cell = ws.cell(commodity_start + 2, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    for row_no, row in enumerate(
        commodity_rows,
        start=commodity_start + 3,
    ):
        values = [
            row["symbol"],
            row["commodity_name"],
            row["close_value"],
            row["day_change_percent"],
            row["five_day_change_percent"],
            row["signal"],
            row["affected_sectors"],
            row["snapshot_date"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        ws.cell(row_no, 4).number_format = '0.00"%"'
        ws.cell(row_no, 5).number_format = '0.00"%"'

        signal = row["signal"]
        ws.cell(row_no, 6).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if signal in {"INPUT COST RELIEF", "STRONG UP"}
                else RED
                if signal in {"INPUT COST PRESSURE", "STRONG DOWN"}
                else GREY
            ),
        )

    widths = {
        "A": 14,
        "B": 24,
        "C": 18,
        "D": 18,
        "E": 14,
        "F": 16,
        "G": 18,
        "H": 24,
        "I": 16,
        "J": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


# ============================================================
# STATUS
# ============================================================

def load_latest_global() -> pd.DataFrame:
    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM global_markets
            WHERE snapshot_date = (
                SELECT MAX(snapshot_date)
                FROM global_markets
            )
            ORDER BY market_name
            """,
            connection,
        )


def load_latest_commodities() -> pd.DataFrame:
    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM commodities
            WHERE snapshot_date = (
                SELECT MAX(snapshot_date)
                FROM commodities
            )
            ORDER BY commodity_name
            """,
            connection,
        )


def show_status() -> None:
    global_frame = load_latest_global()
    commodity_frame = load_latest_commodities()

    print("\nAQSD GLOBAL INTELLIGENCE STATUS")
    print("=" * 78)
    print(f"Latest global records:    {len(global_frame)}")
    print(f"Latest commodity records: {len(commodity_frame)}")

    if not global_frame.empty:
        print(
            f"Latest global date:       "
            f"{global_frame['snapshot_date'].max()}"
        )

    if not commodity_frame.empty:
        print(
            f"Latest commodity date:    "
            f"{commodity_frame['snapshot_date'].max()}"
        )

    print("=" * 78)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Global Markets and Commodities Engine."
    )

    parser.add_argument(
        "--update",
        action="store_true",
        help="Download and store latest global intelligence.",
    )

    parser.add_argument(
        "--period",
        default=DEFAULT_PERIOD,
        help="Yahoo Finance history period, e.g. 1mo or 3mo.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show latest database status.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild the Excel sheet from latest stored data.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_database()

    if args.update:
        global_rows, commodity_rows = update_global_intelligence(
            args.period
        )

        write_dashboard(
            global_rows,
            commodity_rows,
        )

        print("=" * 78)
        print(f"Global records stored:    {len(global_rows)}")
        print(f"Commodity records stored: {len(commodity_rows)}")
        print(f"Global Risk Score:        {global_risk_score(global_rows)}")
        print(DASHBOARD)
        return

    if args.report:
        global_frame = load_latest_global()
        commodity_frame = load_latest_commodities()

        global_rows = [
            {
                "snapshot_date": row["snapshot_date"],
                "symbol": row["symbol"],
                "market_name": row["market_name"],
                "asset_type": row["asset_type"],
                "group": GLOBAL_ASSETS.get(
                    row["symbol"],
                    {},
                ).get("group", ""),
                "close_value": row["close_value"],
                "day_change_percent": row["day_change_percent"],
                "five_day_change_percent": row["five_day_change_percent"],
                "risk_signal": row["risk_signal"],
            }
            for _, row in global_frame.iterrows()
        ]

        commodity_rows = [
            {
                "snapshot_date": row["snapshot_date"],
                "symbol": row["symbol"],
                "commodity_name": row["commodity_name"],
                "close_value": row["close_value"],
                "day_change_percent": row["day_change_percent"],
                "five_day_change_percent": row["five_day_change_percent"],
                "affected_sectors": row["affected_sectors"],
                "signal": commodity_signal(
                    row["symbol"],
                    row["day_change_percent"],
                    row["five_day_change_percent"],
                ),
            }
            for _, row in commodity_frame.iterrows()
        ]

        write_dashboard(global_rows, commodity_rows)
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
