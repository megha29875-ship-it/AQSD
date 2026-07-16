
"""
AQSD Professional
Module: Portfolio Live Assistant
Version: 1.0

Updates portfolio prices, calculates MTM, moves trailing stops,
marks TARGET / STOPPED trades, and automatically records completed
trades in the Trade Journal.

Latest Yahoo Finance prices may be delayed and are not exchange real-time.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ============================================================
# SETTINGS
# ============================================================

TRAILING_STOP_ENABLED = True
TRAIL_TRIGGER_PERCENT = 2.0
TRAIL_DISTANCE_PERCENT = 1.5

PORTFOLIO_HEADER_ROW = 12
PORTFOLIO_FIRST_ROW = 13
JOURNAL_HEADER_ROW = 10
JOURNAL_FIRST_ROW = 11


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
BLUE = "D9EAF7"
GREY = "E7E6E6"


# ============================================================
# HELPERS
# ============================================================

def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def latest_price(symbol: str) -> float | None:
    data = yf.download(
        symbol,
        period="5d",
        interval="1d",
        progress=False,
        auto_adjust=True,
    )

    if data.empty:
        return None

    if hasattr(data.columns, "nlevels") and data.columns.nlevels > 1:
        data.columns = data.columns.get_level_values(0)

    close = data["Close"].dropna()

    if close.empty:
        return None

    return float(close.iloc[-1])


def calculate_pl(
    side: str,
    qty: float,
    entry: float,
    cmp_price: float,
) -> float:
    side = side.upper().strip()

    if side in {"BUY", "CALL"}:
        return (cmp_price - entry) * qty

    return (entry - cmp_price) * qty


def update_trailing_stop(
    side: str,
    entry: float,
    cmp_price: float,
    current_stop: float | None,
) -> float | None:
    if not TRAILING_STOP_ENABLED:
        return current_stop

    side = side.upper().strip()

    if side in {"BUY", "CALL"}:
        gain_percent = ((cmp_price - entry) / entry) * 100

        if gain_percent < TRAIL_TRIGGER_PERCENT:
            return current_stop

        proposed_stop = cmp_price * (
            1 - TRAIL_DISTANCE_PERCENT / 100
        )

        if current_stop is None:
            return proposed_stop

        return max(current_stop, proposed_stop)

    gain_percent = ((entry - cmp_price) / entry) * 100

    if gain_percent < TRAIL_TRIGGER_PERCENT:
        return current_stop

    proposed_stop = cmp_price * (
        1 + TRAIL_DISTANCE_PERCENT / 100
    )

    if current_stop is None:
        return proposed_stop

    return min(current_stop, proposed_stop)


def calculate_status(
    side: str,
    cmp_price: float,
    stop_loss: float | None,
    target: float | None,
) -> str:
    side = side.upper().strip()

    if side in {"BUY", "CALL"}:
        if stop_loss is not None and cmp_price <= stop_loss:
            return "STOPPED"

        if target is not None and cmp_price >= target:
            return "TARGET"

    else:
        if stop_loss is not None and cmp_price >= stop_loss:
            return "STOPPED"

        if target is not None and cmp_price <= target:
            return "TARGET"

    return "OPEN"


def journal_existing_keys(ws) -> set[tuple]:
    keys: set[tuple] = set()

    for row in range(JOURNAL_FIRST_ROW, ws.max_row + 1):
        trade_id = ws.cell(row, 1).value
        symbol = ws.cell(row, 2).value
        entry_date = ws.cell(row, 4).value

        if trade_id and symbol:
            keys.add(
                (
                    str(trade_id),
                    str(symbol),
                    str(entry_date),
                )
            )

    return keys


def append_to_journal(
    journal_ws,
    trade: dict,
    existing_keys: set[tuple],
) -> bool:
    key = (
        str(trade["Trade ID"]),
        str(trade["Symbol"]),
        str(trade["Entry Date"]),
    )

    if key in existing_keys:
        return False

    result = "WIN" if trade["P/L"] > 0 else "LOSS"
    row = max(journal_ws.max_row + 1, JOURNAL_FIRST_ROW)

    values = [
        trade["Trade ID"],
        trade["Symbol"],
        trade["Side"],
        trade["Entry Date"],
        datetime.now(),
        trade["Qty"],
        trade["Entry"],
        trade["CMP"],
        trade["Stop Loss"],
        trade["Target"],
        trade["P/L"],
        trade["P/L %"],
        result,
        "Auto-recorded by Portfolio Live Assistant",
    ]

    for col, value in enumerate(values, start=1):
        journal_ws.cell(row, col, value)

    for col in [7, 8, 9, 10, 11]:
        journal_ws.cell(row, col).number_format = '₹#,##0.00'

    journal_ws.cell(row, 12).number_format = '0.00"%"'
    journal_ws.cell(row, 4).number_format = "dd-mm-yyyy"
    journal_ws.cell(row, 5).number_format = "dd-mm-yyyy"

    journal_ws.cell(row, 13).fill = PatternFill(
        "solid",
        fgColor=GREEN if result == "WIN" else RED,
    )
    journal_ws.cell(row, 13).font = Font(bold=True)

    existing_keys.add(key)
    return True


def update_journal_summary(ws) -> None:
    total = 0
    wins = 0
    losses = 0
    total_pl = 0.0

    for row in range(JOURNAL_FIRST_ROW, ws.max_row + 1):
        symbol = ws.cell(row, 2).value

        if not symbol:
            continue

        total += 1
        pl = float(ws.cell(row, 11).value or 0)
        total_pl += pl

        if pl > 0:
            wins += 1
        elif pl < 0:
            losses += 1

    win_rate = (wins / total * 100) if total else 0.0
    average_pl = (total_pl / total) if total else 0.0

    ws["B4"] = total
    ws["B5"] = wins
    ws["B6"] = losses
    ws["B7"] = round(win_rate, 2)
    ws["B8"] = round(total_pl, 2)
    ws["B9"] = round(average_pl, 2)

    ws["B7"].number_format = '0.00"%"'
    ws["B8"].number_format = '₹#,##0.00'
    ws["B9"].number_format = '₹#,##0.00'


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    if "Portfolio" not in wb.sheetnames:
        raise RuntimeError(
            "Portfolio sheet not found. "
            "Run portfolio_manager.py first."
        )

    if "Trade Journal" not in wb.sheetnames:
        raise RuntimeError(
            "Trade Journal sheet not found. "
            "Run trade_journal.py first."
        )

    portfolio_ws = wb["Portfolio"]
    journal_ws = wb["Trade Journal"]

    p_headers = header_map(
        portfolio_ws,
        PORTFOLIO_HEADER_ROW,
    )

    required = [
        "Trade ID",
        "Symbol",
        "Side",
        "Entry Date",
        "Qty",
        "Entry",
        "CMP",
        "Stop Loss",
        "Capital Used",
        "Risk Amount",
        "P/L",
        "P/L %",
        "Target",
        "Status",
    ]

    missing = [
        name
        for name in required
        if name not in p_headers
    ]

    if missing:
        raise RuntimeError(
            "Missing Portfolio columns: "
            + ", ".join(missing)
        )

    existing_keys = journal_existing_keys(journal_ws)

    total_invested = 0.0
    total_pl = 0.0
    open_positions = 0
    updated = 0
    journal_added = 0

    for row in range(
        PORTFOLIO_FIRST_ROW,
        portfolio_ws.max_row + 1,
    ):
        symbol = str(
            portfolio_ws.cell(
                row,
                p_headers["Symbol"],
            ).value
            or ""
        ).strip().upper()

        if not symbol:
            continue

        if not symbol.endswith(".NS"):
            symbol += ".NS"
            portfolio_ws.cell(
                row,
                p_headers["Symbol"],
            ).value = symbol

        qty_value = portfolio_ws.cell(
            row,
            p_headers["Qty"],
        ).value

        entry_value = portfolio_ws.cell(
            row,
            p_headers["Entry"],
        ).value

        if not qty_value or not entry_value:
            continue

        qty = float(qty_value)
        entry = float(entry_value)

        side = str(
            portfolio_ws.cell(
                row,
                p_headers["Side"],
            ).value
            or "CALL"
        ).upper().strip()

        stop_value = portfolio_ws.cell(
            row,
            p_headers["Stop Loss"],
        ).value

        target_value = portfolio_ws.cell(
            row,
            p_headers["Target"],
        ).value

        stop_loss = (
            float(stop_value)
            if stop_value not in (None, "")
            else None
        )

        target = (
            float(target_value)
            if target_value not in (None, "")
            else None
        )

        cmp_price = latest_price(symbol)

        if cmp_price is None:
            portfolio_ws.cell(
                row,
                p_headers["Status"],
            ).value = "PRICE ERROR"

            portfolio_ws.cell(
                row,
                p_headers["Status"],
            ).fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )
            continue

        new_stop = update_trailing_stop(
            side,
            entry,
            cmp_price,
            stop_loss,
        )

        if new_stop is not None:
            stop_loss = round(new_stop, 2)

            portfolio_ws.cell(
                row,
                p_headers["Stop Loss"],
            ).value = stop_loss

        capital_used = qty * entry

        risk_amount = (
            qty * abs(entry - stop_loss)
            if stop_loss is not None
            else 0.0
        )

        profit_loss = calculate_pl(
            side,
            qty,
            entry,
            cmp_price,
        )

        pl_percent = (
            profit_loss / capital_used * 100
            if capital_used
            else 0.0
        )

        status = calculate_status(
            side,
            cmp_price,
            stop_loss,
            target,
        )

        portfolio_ws.cell(
            row,
            p_headers["CMP"],
        ).value = round(cmp_price, 2)

        portfolio_ws.cell(
            row,
            p_headers["Capital Used"],
        ).value = round(capital_used, 2)

        portfolio_ws.cell(
            row,
            p_headers["Risk Amount"],
        ).value = round(risk_amount, 2)

        portfolio_ws.cell(
            row,
            p_headers["P/L"],
        ).value = round(profit_loss, 2)

        portfolio_ws.cell(
            row,
            p_headers["P/L %"],
        ).value = round(pl_percent, 2)

        portfolio_ws.cell(
            row,
            p_headers["Status"],
        ).value = status

        pl_fill = GREEN if profit_loss >= 0 else RED

        portfolio_ws.cell(
            row,
            p_headers["P/L"],
        ).fill = PatternFill(
            "solid",
            fgColor=pl_fill,
        )

        portfolio_ws.cell(
            row,
            p_headers["P/L %"],
        ).fill = PatternFill(
            "solid",
            fgColor=pl_fill,
        )

        if status == "TARGET":
            status_fill = GREEN
        elif status == "STOPPED":
            status_fill = RED
        else:
            status_fill = YELLOW

        portfolio_ws.cell(
            row,
            p_headers["Status"],
        ).fill = PatternFill(
            "solid",
            fgColor=status_fill,
        )

        portfolio_ws.cell(
            row,
            p_headers["Status"],
        ).font = Font(bold=True)

        if status == "OPEN":
            total_invested += capital_used
            total_pl += profit_loss
            open_positions += 1

        if status in {"TARGET", "STOPPED"}:
            trade = {
                "Trade ID": portfolio_ws.cell(
                    row,
                    p_headers["Trade ID"],
                ).value,
                "Symbol": symbol,
                "Side": side,
                "Entry Date": portfolio_ws.cell(
                    row,
                    p_headers["Entry Date"],
                ).value,
                "Qty": qty,
                "Entry": entry,
                "CMP": cmp_price,
                "Stop Loss": stop_loss,
                "Target": target,
                "P/L": profit_loss,
                "P/L %": pl_percent,
            }

            if append_to_journal(
                journal_ws,
                trade,
                existing_keys,
            ):
                journal_added += 1

        updated += 1

    trading_capital = float(
        portfolio_ws["B4"].value or 0
    )

    available_cash = (
        trading_capital - total_invested
    )

    portfolio_return = (
        total_pl / total_invested * 100
        if total_invested
        else 0.0
    )

    portfolio_ws["B5"] = round(total_invested, 2)
    portfolio_ws["B6"] = round(available_cash, 2)
    portfolio_ws["B7"] = open_positions
    portfolio_ws["B8"] = round(total_pl, 2)
    portfolio_ws["B9"] = round(portfolio_return, 2)

    portfolio_ws["D4"] = "Last Live Update"
    portfolio_ws["E4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )
    portfolio_ws["D4"].font = Font(bold=True)
    portfolio_ws["D4"].fill = PatternFill(
        "solid",
        fgColor=BLUE,
    )

    update_journal_summary(journal_ws)

    wb.save(DASHBOARD)

    print("AQSD Portfolio Live Assistant completed.")
    print(f"Positions updated: {updated}")
    print(f"Open positions: {open_positions}")
    print(f"Journal entries added: {journal_added}")
    print(f"Open-position MTM: ₹{total_pl:,.2f}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
