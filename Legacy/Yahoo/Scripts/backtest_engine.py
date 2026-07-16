
"""
AQSD Professional
Module: Backtesting Engine
Version: 1.0

Backtests a simple bullish swing strategy:
- EMA20 above EMA50
- RSI14 above 55
- Entry on next day's open
- ATR-based stop loss
- Fixed reward:risk target
- Maximum holding period

Outputs:
- Backtest Trades
- Backtest Summary
- Equity Curve chart

Examples
--------
python backtest_engine.py --symbol RELIANCE.NS
python backtest_engine.py --symbol BIOCON.NS --period 5y
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

DEFAULT_SYMBOL = "RELIANCE.NS"
DEFAULT_PERIOD = "3y"
ATR_PERIOD = 14
ATR_MULTIPLIER = 2.0
TARGET_RR = 2.0
MAX_HOLD_DAYS = 10
INITIAL_CAPITAL = 200000
RISK_PERCENT = 1.0


@dataclass
class Trade:
    symbol: str
    entry_date: str
    exit_date: str
    entry: float
    stop_loss: float
    target: float
    exit_price: float
    quantity: int
    pnl: float
    pnl_percent: float
    result: str
    holding_days: int


def rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    avg_gain = gain.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    avg_loss = loss.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def atr(df: pd.DataFrame, period: int = 14) -> pd.Series:
    previous_close = df["Close"].shift(1)

    tr = pd.concat(
        [
            df["High"] - df["Low"],
            (df["High"] - previous_close).abs(),
            (df["Low"] - previous_close).abs(),
        ],
        axis=1,
    ).max(axis=1)

    return tr.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()


def download_data(symbol: str, period: str) -> pd.DataFrame:
    df = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
    )

    if df.empty:
        raise RuntimeError(f"No data downloaded for {symbol}")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df = df.dropna(subset=["Open", "High", "Low", "Close"])

    df["EMA20"] = df["Close"].ewm(span=20, adjust=False).mean()
    df["EMA50"] = df["Close"].ewm(span=50, adjust=False).mean()
    df["RSI14"] = rsi(df["Close"], 14)
    df["ATR14"] = atr(df, ATR_PERIOD)

    return df.dropna()


def build_signals(df: pd.DataFrame) -> pd.Series:
    bullish = (
        (df["EMA20"] > df["EMA50"])
        & (df["RSI14"] > 55)
        & (df["Close"] > df["EMA20"])
    )

    fresh_signal = bullish & (~bullish.shift(1).fillna(False))
    return fresh_signal


def position_size(
    capital: float,
    risk_percent: float,
    entry: float,
    stop_loss: float,
) -> int:
    risk_amount = capital * (risk_percent / 100)
    risk_per_share = abs(entry - stop_loss)

    if risk_per_share <= 0:
        return 0

    qty_by_risk = int(risk_amount // risk_per_share)
    qty_by_capital = int(capital // entry)

    return max(0, min(qty_by_risk, qty_by_capital))


def run_backtest(
    symbol: str,
    df: pd.DataFrame,
) -> list[Trade]:
    signals = build_signals(df)
    trades: list[Trade] = []

    capital = INITIAL_CAPITAL
    i = 0

    while i < len(df) - 1:
        if not bool(signals.iloc[i]):
            i += 1
            continue

        entry_index = i + 1
        entry_row = df.iloc[entry_index]

        entry = float(entry_row["Open"])
        atr_value = float(df.iloc[i]["ATR14"])

        stop_loss = entry - ATR_MULTIPLIER * atr_value
        risk_per_share = entry - stop_loss
        target = entry + TARGET_RR * risk_per_share

        quantity = position_size(
            capital,
            RISK_PERCENT,
            entry,
            stop_loss,
        )

        if quantity <= 0:
            i += 1
            continue

        exit_price = float(entry_row["Close"])
        exit_index = entry_index
        result = "TIME EXIT"

        final_index = min(
            entry_index + MAX_HOLD_DAYS,
            len(df) - 1,
        )

        for j in range(entry_index, final_index + 1):
            row = df.iloc[j]
            low = float(row["Low"])
            high = float(row["High"])
            close = float(row["Close"])

            if low <= stop_loss:
                exit_price = stop_loss
                exit_index = j
                result = "LOSS"
                break

            if high >= target:
                exit_price = target
                exit_index = j
                result = "WIN"
                break

            exit_price = close
            exit_index = j

        pnl = (exit_price - entry) * quantity
        pnl_percent = (
            pnl / (entry * quantity) * 100
            if entry * quantity
            else 0.0
        )

        capital += pnl

        trade = Trade(
            symbol=symbol,
            entry_date=str(df.index[entry_index].date()),
            exit_date=str(df.index[exit_index].date()),
            entry=round(entry, 2),
            stop_loss=round(stop_loss, 2),
            target=round(target, 2),
            exit_price=round(exit_price, 2),
            quantity=quantity,
            pnl=round(pnl, 2),
            pnl_percent=round(pnl_percent, 2),
            result=result,
            holding_days=max(exit_index - entry_index + 1, 1),
        )

        trades.append(trade)
        i = exit_index + 1

    return trades


def summary_metrics(trades: list[Trade]) -> dict:
    total = len(trades)
    wins = [t for t in trades if t.pnl > 0]
    losses = [t for t in trades if t.pnl < 0]

    gross_profit = sum(t.pnl for t in wins)
    gross_loss = abs(sum(t.pnl for t in losses))
    net_profit = sum(t.pnl for t in trades)

    win_rate = len(wins) / total * 100 if total else 0.0
    average_win = gross_profit / len(wins) if wins else 0.0
    average_loss = gross_loss / len(losses) if losses else 0.0

    profit_factor = (
        gross_profit / gross_loss
        if gross_loss
        else gross_profit
    )

    expectancy = net_profit / total if total else 0.0

    equity = INITIAL_CAPITAL
    peak = INITIAL_CAPITAL
    max_drawdown = 0.0

    for trade in trades:
        equity += trade.pnl
        peak = max(peak, equity)
        drawdown = equity - peak
        max_drawdown = min(max_drawdown, drawdown)

    return {
        "Initial Capital": INITIAL_CAPITAL,
        "Final Capital": INITIAL_CAPITAL + net_profit,
        "Net Profit": net_profit,
        "Return %": (
            net_profit / INITIAL_CAPITAL * 100
            if INITIAL_CAPITAL
            else 0.0
        ),
        "Total Trades": total,
        "Winning Trades": len(wins),
        "Losing Trades": len(losses),
        "Win Rate %": win_rate,
        "Average Win": average_win,
        "Average Loss": average_loss,
        "Profit Factor": profit_factor,
        "Expectancy / Trade": expectancy,
        "Maximum Drawdown": abs(max_drawdown),
    }


def write_results(
    symbol: str,
    trades: list[Trade],
    metrics: dict,
) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_name in ["Backtest Trades", "Backtest Summary"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    summary_ws = wb.create_sheet("Backtest Summary")
    trades_ws = wb.create_sheet("Backtest Trades")

    navy = "17365D"
    white = "FFFFFF"
    blue = "D9EAF7"

    summary_ws.merge_cells("A1:D2")
    summary_ws["A1"] = f"AQSD BACKTEST SUMMARY - {symbol}"
    summary_ws["A1"].font = Font(
        size=18,
        bold=True,
        color=white,
    )
    summary_ws["A1"].fill = PatternFill(
        "solid",
        fgColor=navy,
    )
    summary_ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    for row_no, (label, value) in enumerate(
        metrics.items(),
        start=4,
    ):
        summary_ws[f"A{row_no}"] = label
        summary_ws[f"B{row_no}"] = value
        summary_ws[f"A{row_no}"].font = Font(bold=True)
        summary_ws[f"A{row_no}"].fill = PatternFill(
            "solid",
            fgColor=blue,
        )

        if label in {
            "Initial Capital",
            "Final Capital",
            "Net Profit",
            "Average Win",
            "Average Loss",
            "Expectancy / Trade",
            "Maximum Drawdown",
        }:
            summary_ws[f"B{row_no}"].number_format = '₹#,##0.00'
        elif label in {"Return %", "Win Rate %"}:
            summary_ws[f"B{row_no}"].number_format = '0.00"%"'
        elif label == "Profit Factor":
            summary_ws[f"B{row_no}"].number_format = "0.00"

    headers = [
        "Trade No.",
        "Symbol",
        "Entry Date",
        "Exit Date",
        "Entry",
        "Stop Loss",
        "Target",
        "Exit Price",
        "Quantity",
        "P/L",
        "P/L %",
        "Result",
        "Holding Days",
        "Equity",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = trades_ws.cell(1, col, heading)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)

    equity = INITIAL_CAPITAL

    for row_no, trade in enumerate(trades, start=2):
        equity += trade.pnl

        values = [
            row_no - 1,
            trade.symbol,
            trade.entry_date,
            trade.exit_date,
            trade.entry,
            trade.stop_loss,
            trade.target,
            trade.exit_price,
            trade.quantity,
            trade.pnl,
            trade.pnl_percent,
            trade.result,
            trade.holding_days,
            round(equity, 2),
        ]

        for col, value in enumerate(values, start=1):
            trades_ws.cell(row_no, col, value)

        for col in [5, 6, 7, 8, 10, 14]:
            trades_ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

        trades_ws.cell(
            row_no,
            11,
        ).number_format = '0.00"%"'

    if trades:
        chart = LineChart()
        chart.title = "Backtest Equity Curve"
        chart.y_axis.title = "Equity"
        chart.x_axis.title = "Trade Number"

        chart.add_data(
            Reference(
                trades_ws,
                min_col=14,
                min_row=1,
                max_row=len(trades) + 1,
            ),
            titles_from_data=True,
        )

        chart.set_categories(
            Reference(
                trades_ws,
                min_col=1,
                min_row=2,
                max_row=len(trades) + 1,
            )
        )

        chart.height = 8
        chart.width = 14
        summary_ws.add_chart(chart, "D4")

    trades_ws.freeze_panes = "A2"
    trades_ws.auto_filter.ref = trades_ws.dimensions

    wb.save(DASHBOARD)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backtest the AQSD EMA/RSI swing strategy."
    )

    parser.add_argument(
        "--symbol",
        default=DEFAULT_SYMBOL,
        help="Yahoo Finance symbol, e.g. RELIANCE.NS",
    )

    parser.add_argument(
        "--period",
        default=DEFAULT_PERIOD,
        help="Yahoo Finance history period, e.g. 1y, 3y, 5y",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    symbol = args.symbol.upper().strip()

    if not symbol.endswith(".NS"):
        symbol += ".NS"

    print("\nAQSD BACKTESTING ENGINE")
    print("=" * 60)
    print(f"Symbol: {symbol}")
    print(f"Period: {args.period}")

    df = download_data(symbol, args.period)
    trades = run_backtest(symbol, df)
    metrics = summary_metrics(trades)

    write_results(
        symbol,
        trades,
        metrics,
    )

    print(f"Trades completed: {len(trades)}")
    print(f"Net profit: ₹{metrics['Net Profit']:,.2f}")
    print(f"Win rate: {metrics['Win Rate %']:.2f}%")
    print(f"Profit factor: {metrics['Profit Factor']:.2f}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
