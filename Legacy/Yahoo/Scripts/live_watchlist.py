
"""
AQSD Professional
Module: Live Watchlist
Version: 1.0

Creates a Live Watchlist sheet using the top CALL and PUT candidates.

Features
--------
- Top 10 CALL candidates
- Top 10 PUT candidates
- Latest available price from Yahoo Finance
- Daily percentage change
- Entry, stop loss and target
- Distance to entry, stop and target
- Live status
- Colour-coded watchlist
- Timestamped refresh

Note:
Yahoo Finance prices may be delayed and are not exchange real-time.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# SETTINGS
# ============================================================

TOP_CALLS = 10
TOP_PUTS = 10


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
DARK_GREEN = "006100"
RED = "FFC7CE"
DARK_RED = "9C0006"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def header_map(ws, row_number: int = 1) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def normalise_symbol(symbol: str) -> str:
    symbol = symbol.strip().upper()

    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"

    return symbol


def get_candidates(
    wb,
    sheet_name: str,
    limit: int,
) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    headers = header_map(ws, 1)

    required = [
        "Symbol",
        "Trade Score",
        "Trade Confidence",
        "Trade Grade",
        "Recommendation",
        "Entry",
        "Stop Loss",
        "Target 1",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            f"Missing columns in {sheet_name}: "
            + ", ".join(missing)
        )

    rows: list[dict] = []

    for row in range(2, min(ws.max_row, limit + 1) + 1):
        symbol = normalise_symbol(
            str(ws.cell(row, headers["Symbol"]).value or "")
        )

        if not symbol:
            continue

        rows.append(
            {
                "Symbol": symbol,
                "Score": ws.cell(
                    row,
                    headers["Trade Score"],
                ).value,
                "Confidence": ws.cell(
                    row,
                    headers["Trade Confidence"],
                ).value,
                "Grade": ws.cell(
                    row,
                    headers["Trade Grade"],
                ).value,
                "Recommendation": ws.cell(
                    row,
                    headers["Recommendation"],
                ).value,
                "Entry": ws.cell(
                    row,
                    headers["Entry"],
                ).value,
                "Stop Loss": ws.cell(
                    row,
                    headers["Stop Loss"],
                ).value,
                "Target": ws.cell(
                    row,
                    headers["Target 1"],
                ).value,
            }
        )

    return rows


def download_market_data(symbols: list[str]) -> dict[str, dict]:
    if not symbols:
        return {}

    data = yf.download(
        tickers=sorted(set(symbols)),
        period="5d",
        interval="1d",
        group_by="ticker",
        auto_adjust=True,
        progress=False,
        threads=True,
    )

    result: dict[str, dict] = {}

    if len(symbols) == 1:
        symbol = symbols[0]

        if not data.empty and "Close" in data.columns:
            close = data["Close"].dropna()

            if not close.empty:
                latest = float(close.iloc[-1])
                previous = (
                    float(close.iloc[-2])
                    if len(close) >= 2
                    else latest
                )

                result[symbol] = {
                    "CMP": latest,
                    "Previous": previous,
                }

        return result

    for symbol in sorted(set(symbols)):
        try:
            close = data[symbol]["Close"].dropna()

            if close.empty:
                continue

            latest = float(close.iloc[-1])
            previous = (
                float(close.iloc[-2])
                if len(close) >= 2
                else latest
            )

            result[symbol] = {
                "CMP": latest,
                "Previous": previous,
            }

        except (KeyError, TypeError):
            continue

    return result


def safe_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def percentage_change(
    current: float,
    previous: float,
) -> float:
    if previous == 0:
        return 0.0

    return (current - previous) / previous * 100


def distance_percent(
    current: float,
    reference: float | None,
) -> float | None:
    if reference in (None, 0):
        return None

    return (current - reference) / reference * 100


def trade_status(
    side: str,
    cmp_price: float,
    entry: float | None,
    stop_loss: float | None,
    target: float | None,
) -> str:
    side = side.upper().strip()

    if entry is None:
        return "NO ENTRY"

    if side == "CALL":
        if stop_loss is not None and cmp_price <= stop_loss:
            return "STOP ZONE"

        if target is not None and cmp_price >= target:
            return "TARGET HIT"

        if cmp_price >= entry:
            return "ACTIVE"

        distance = (entry - cmp_price) / entry * 100

        if distance <= 1:
            return "NEAR ENTRY"

        return "WAIT"

    if stop_loss is not None and cmp_price >= stop_loss:
        return "STOP ZONE"

    if target is not None and cmp_price <= target:
        return "TARGET HIT"

    if cmp_price <= entry:
        return "ACTIVE"

    distance = (cmp_price - entry) / entry * 100

    if distance <= 1:
        return "NEAR ENTRY"

    return "WAIT"


def add_section(
    ws,
    start_row: int,
    title: str,
    side: str,
    candidates: list[dict],
    prices: dict[str, dict],
) -> int:
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=15,
    )

    title_cell = ws.cell(start_row, 1, title)
    title_cell.font = Font(
        size=14,
        bold=True,
        color=WHITE,
    )
    title_cell.fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    title_cell.alignment = Alignment(
        horizontal="center",
    )

    header_row = start_row + 1

    headers = [
        "Rank",
        "Symbol",
        "Score",
        "Confidence",
        "Grade",
        "Recommendation",
        "CMP",
        "Day %",
        "Entry",
        "Stop Loss",
        "Target",
        "Entry Dist %",
        "Stop Dist %",
        "Target Dist %",
        "Live Status",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    first_data_row = header_row + 1

    for index, candidate in enumerate(candidates, start=1):
        row = first_data_row + index - 1
        symbol = candidate["Symbol"]
        market = prices.get(symbol)

        cmp_price = None
        day_change = None

        if market:
            cmp_price = float(market["CMP"])
            day_change = percentage_change(
                cmp_price,
                float(market["Previous"]),
            )

        entry = safe_float(candidate["Entry"])
        stop_loss = safe_float(candidate["Stop Loss"])
        target = safe_float(candidate["Target"])

        status = (
            trade_status(
                side,
                cmp_price,
                entry,
                stop_loss,
                target,
            )
            if cmp_price is not None
            else "PRICE ERROR"
        )

        entry_distance = (
            distance_percent(cmp_price, entry)
            if cmp_price is not None
            else None
        )

        stop_distance = (
            distance_percent(cmp_price, stop_loss)
            if cmp_price is not None
            else None
        )

        target_distance = (
            distance_percent(cmp_price, target)
            if cmp_price is not None
            else None
        )

        values = [
            index,
            symbol,
            candidate["Score"],
            candidate["Confidence"],
            candidate["Grade"],
            candidate["Recommendation"],
            round(cmp_price, 2)
            if cmp_price is not None
            else "",
            round(day_change, 2)
            if day_change is not None
            else "",
            entry,
            stop_loss,
            target,
            round(entry_distance, 2)
            if entry_distance is not None
            else "",
            round(stop_distance, 2)
            if stop_distance is not None
            else "",
            round(target_distance, 2)
            if target_distance is not None
            else "",
            status,
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = Border(bottom=THIN)

        ws.cell(row, 1).fill = PatternFill(
            "solid",
            fgColor=YELLOW,
        )

        if day_change is not None:
            day_fill = GREEN if day_change >= 0 else RED
            day_font = DARK_GREEN if day_change >= 0 else DARK_RED

            ws.cell(row, 8).fill = PatternFill(
                "solid",
                fgColor=day_fill,
            )
            ws.cell(row, 8).font = Font(
                bold=True,
                color=day_font,
            )

        if status == "ACTIVE":
            status_fill = GREEN
            status_font = DARK_GREEN
        elif status == "TARGET HIT":
            status_fill = GREEN
            status_font = DARK_GREEN
        elif status == "STOP ZONE":
            status_fill = RED
            status_font = DARK_RED
        elif status == "NEAR ENTRY":
            status_fill = YELLOW
            status_font = "7F6000"
        elif status == "WAIT":
            status_fill = GREY
            status_font = "666666"
        else:
            status_fill = RED
            status_font = DARK_RED

        ws.cell(row, 15).fill = PatternFill(
            "solid",
            fgColor=status_fill,
        )
        ws.cell(row, 15).font = Font(
            bold=True,
            color=status_font,
        )

        for col in [7, 9, 10, 11]:
            ws.cell(row, col).number_format = '₹#,##0.00'

        for col in [8, 12, 13, 14]:
            ws.cell(row, col).number_format = '0.00"%"'

    return first_data_row + len(candidates)


def create_watchlist_sheet(
    wb,
    calls: list[dict],
    puts: list[dict],
    prices: dict[str, dict],
) -> None:
    if "Live Watchlist" in wb.sheetnames:
        del wb["Live Watchlist"]

    ws = wb.create_sheet("Live Watchlist", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    ws.merge_cells("A1:O2")
    ws["A1"] = "AQSD PROFESSIONAL - LIVE WATCHLIST"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Last Updated"
    ws["B4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )
    ws["A4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill(
        "solid",
        fgColor=BLUE,
    )

    next_row = add_section(
        ws,
        6,
        "TOP CALL WATCHLIST",
        "CALL",
        calls,
        prices,
    )

    add_section(
        ws,
        next_row + 2,
        "TOP PUT WATCHLIST",
        "PUT",
        puts,
        prices,
    )

    widths = {
        "A": 8,
        "B": 18,
        "C": 10,
        "D": 12,
        "E": 9,
        "F": 18,
        "G": 12,
        "H": 10,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    calls = get_candidates(
        wb,
        "CALL Candidates",
        TOP_CALLS,
    )

    puts = get_candidates(
        wb,
        "PUT Candidates",
        TOP_PUTS,
    )

    symbols = [
        item["Symbol"]
        for item in calls + puts
    ]

    prices = download_market_data(symbols)

    create_watchlist_sheet(
        wb,
        calls,
        puts,
        prices,
    )

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Live Watchlist created successfully.")
    print(f"CALL candidates: {len(calls)}")
    print(f"PUT candidates: {len(puts)}")
    print(f"Prices downloaded: {len(prices)}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
