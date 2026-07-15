
"""
AQSD Derivatives Intelligence
Module: Options OI & PCR Intelligence Engine
Version: 1.0

Purpose
-------
Builds options-chain intelligence from imported EOD or live snapshots.

The module is provider-agnostic. It can later accept data from NSE,
Quantsapp, TrueData, broker APIs, Opstra, Sensibull or any other source,
provided the data is converted into the expected CSV structure.

Core features
-------------
- Call OI and Put OI
- Change in Call OI and Put OI
- OI PCR
- Volume PCR
- Modified PCR
- Max Pain
- Highest Call OI wall
- Highest Put OI wall
- Call writing / Put writing detection
- Short covering / Long unwinding detection
- Support / Resistance from options OI
- Options Intelligence Score from 0 to 100
- Excel report: Options Intelligence
- SQLite storage in aqsd_core.db

Minimum CSV columns
-------------------
trade_date
nse_symbol
expiry_date
strike_price
option_type
open_interest
volume
last_price
underlying_price

Recommended additional columns
------------------------------
previous_open_interest
previous_last_price
implied_volatility
bid_price
ask_price

Commands
--------
python aqsd_options_intelligence.py --setup
python aqsd_options_intelligence.py --import-csv C:\\Users\\megha\\AQSD\\Data\\options_data.csv
python aqsd_options_intelligence.py --analyse
python aqsd_options_intelligence.py --report
python aqsd_options_intelligence.py --status
"""

from __future__ import annotations

import argparse
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS options_chain (
    option_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    nse_symbol TEXT NOT NULL,
    expiry_date TEXT NOT NULL,
    strike_price REAL NOT NULL,
    option_type TEXT NOT NULL,
    open_interest REAL NOT NULL,
    previous_open_interest REAL,
    oi_change REAL,
    oi_change_percent REAL,
    volume REAL,
    last_price REAL,
    previous_last_price REAL,
    price_change_percent REAL,
    implied_volatility REAL,
    bid_price REAL,
    ask_price REAL,
    underlying_price REAL NOT NULL,
    activity_type TEXT,
    source TEXT,
    event_hash TEXT NOT NULL UNIQUE,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_options_symbol_date
ON options_chain(nse_symbol, trade_date);

CREATE INDEX IF NOT EXISTS idx_options_symbol_expiry
ON options_chain(nse_symbol, expiry_date);

CREATE TABLE IF NOT EXISTS options_intelligence (
    intelligence_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    nse_symbol TEXT NOT NULL,
    expiry_date TEXT NOT NULL,
    underlying_price REAL,
    call_oi REAL,
    put_oi REAL,
    call_oi_change REAL,
    put_oi_change REAL,
    call_volume REAL,
    put_volume REAL,
    oi_pcr REAL,
    volume_pcr REAL,
    modified_pcr REAL,
    max_pain REAL,
    call_wall REAL,
    put_wall REAL,
    options_support REAL,
    options_resistance REAL,
    directional_bias TEXT,
    options_score REAL,
    explanation TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(trade_date, nse_symbol, expiry_date)
);

CREATE INDEX IF NOT EXISTS idx_options_intelligence_symbol_date
ON options_intelligence(nse_symbol, trade_date);
"""


def setup_options_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# HELPERS
# ============================================================

def clean_text(value: Any) -> str:
    text = str(value or "").strip()

    if text.lower() in {"nan", "none", "null"}:
        return ""

    return text


def normalize_symbol(value: Any) -> str:
    symbol = clean_text(value).upper()

    if symbol.endswith(".NS"):
        symbol = symbol[:-3]

    return symbol.replace(" ", "")


def normalize_option_type(value: Any) -> str:
    text = clean_text(value).upper()

    if text in {"CE", "CALL", "C"}:
        return "CE"

    if text in {"PE", "PUT", "P"}:
        return "PE"

    raise ValueError(f"Unknown option type: {value}")


def safe_float(
    value: Any,
    default: float | None = None,
) -> float | None:
    try:
        if value is None or value == "":
            return default

        return float(value)

    except (TypeError, ValueError):
        return default


def parse_date(value: Any) -> str:
    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        raise ValueError(f"Invalid date: {value}")

    return parsed.date().isoformat()


def pct_change(
    current: float | None,
    previous: float | None,
) -> float | None:
    if current is None or previous in (None, 0):
        return None

    return round((current / previous - 1) * 100, 2)


def build_hash(record: dict) -> str:
    raw = "|".join(
        [
            record["trade_date"],
            record["nse_symbol"],
            record["expiry_date"],
            str(record["strike_price"]),
            record["option_type"],
        ]
    )

    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ============================================================
# OPTION ACTIVITY CLASSIFICATION
# ============================================================

def classify_option_activity(
    price_change_percent: float | None,
    oi_change_percent: float | None,
    option_type: str,
) -> str:
    if price_change_percent is None or oi_change_percent is None:
        return "UNCLASSIFIED"

    if price_change_percent > 0 and oi_change_percent > 0:
        return (
            "CALL BUYING"
            if option_type == "CE"
            else "PUT BUYING"
        )

    if price_change_percent < 0 and oi_change_percent > 0:
        return (
            "CALL WRITING"
            if option_type == "CE"
            else "PUT WRITING"
        )

    if price_change_percent > 0 and oi_change_percent < 0:
        return "SHORT COVERING"

    if price_change_percent < 0 and oi_change_percent < 0:
        return "LONG UNWINDING"

    return "NEUTRAL"


# ============================================================
# CSV IMPORT
# ============================================================

ALIASES = {
    "date": "trade_date",
    "symbol": "nse_symbol",
    "ticker": "nse_symbol",
    "expiry": "expiry_date",
    "strike": "strike_price",
    "type": "option_type",
    "option": "option_type",
    "oi": "open_interest",
    "prev_oi": "previous_open_interest",
    "previous_oi": "previous_open_interest",
    "ltp": "last_price",
    "prev_ltp": "previous_last_price",
    "iv": "implied_volatility",
    "underlying": "underlying_price",
    "spot_price": "underlying_price",
}


def normalize_input_frame(frame: pd.DataFrame) -> pd.DataFrame:
    renamed = {}

    for column in frame.columns:
        key = (
            str(column)
            .strip()
            .lower()
            .replace(" ", "_")
            .replace("/", "_")
        )

        renamed[column] = ALIASES.get(key, key)

    return frame.rename(columns=renamed)


def import_csv(path: Path, source: str) -> tuple[int, int]:
    setup_options_schema()

    if not path.exists():
        raise FileNotFoundError(path)

    frame = normalize_input_frame(
        pd.read_csv(path)
    )

    required = {
        "trade_date",
        "nse_symbol",
        "expiry_date",
        "strike_price",
        "option_type",
        "open_interest",
        "volume",
        "last_price",
        "underlying_price",
    }

    if not required.issubset(frame.columns):
        missing = sorted(required - set(frame.columns))

        raise RuntimeError(
            "Missing required columns: " + ", ".join(missing)
        )

    inserted = 0
    duplicates = 0

    run_id = start_run(
        "aqsd_options_intelligence",
        f"Importing {path.name}",
    )

    try:
        with connect() as connection:
            for _, row in frame.iterrows():
                record = {
                    "trade_date": parse_date(row.get("trade_date")),
                    "nse_symbol": normalize_symbol(
                        row.get("nse_symbol")
                    ),
                    "expiry_date": parse_date(
                        row.get("expiry_date")
                    ),
                    "strike_price": safe_float(
                        row.get("strike_price")
                    ),
                    "option_type": normalize_option_type(
                        row.get("option_type")
                    ),
                    "open_interest": safe_float(
                        row.get("open_interest")
                    ),
                    "previous_open_interest": safe_float(
                        row.get("previous_open_interest")
                    ),
                    "volume": safe_float(
                        row.get("volume"),
                        0,
                    ),
                    "last_price": safe_float(
                        row.get("last_price")
                    ),
                    "previous_last_price": safe_float(
                        row.get("previous_last_price")
                    ),
                    "implied_volatility": safe_float(
                        row.get("implied_volatility")
                    ),
                    "bid_price": safe_float(
                        row.get("bid_price")
                    ),
                    "ask_price": safe_float(
                        row.get("ask_price")
                    ),
                    "underlying_price": safe_float(
                        row.get("underlying_price")
                    ),
                }

                if (
                    not record["nse_symbol"]
                    or record["strike_price"] is None
                    or record["open_interest"] is None
                    or record["last_price"] is None
                    or record["underlying_price"] is None
                ):
                    continue

                oi_change = (
                    record["open_interest"]
                    - record["previous_open_interest"]
                    if record["previous_open_interest"] is not None
                    else None
                )

                oi_change_percent = pct_change(
                    record["open_interest"],
                    record["previous_open_interest"],
                )

                price_change_percent = pct_change(
                    record["last_price"],
                    record["previous_last_price"],
                )

                activity_type = classify_option_activity(
                    price_change_percent,
                    oi_change_percent,
                    record["option_type"],
                )

                hashed = build_hash(record)

                try:
                    connection.execute(
                        """
                        INSERT INTO options_chain(
                            trade_date,
                            nse_symbol,
                            expiry_date,
                            strike_price,
                            option_type,
                            open_interest,
                            previous_open_interest,
                            oi_change,
                            oi_change_percent,
                            volume,
                            last_price,
                            previous_last_price,
                            price_change_percent,
                            implied_volatility,
                            bid_price,
                            ask_price,
                            underlying_price,
                            activity_type,
                            source,
                            event_hash,
                            created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            record["trade_date"],
                            record["nse_symbol"],
                            record["expiry_date"],
                            record["strike_price"],
                            record["option_type"],
                            record["open_interest"],
                            record["previous_open_interest"],
                            oi_change,
                            oi_change_percent,
                            record["volume"],
                            record["last_price"],
                            record["previous_last_price"],
                            price_change_percent,
                            record["implied_volatility"],
                            record["bid_price"],
                            record["ask_price"],
                            record["underlying_price"],
                            activity_type,
                            source,
                            hashed,
                            datetime.now().isoformat(
                                timespec="seconds"
                            ),
                        ),
                    )

                    inserted += 1

                except Exception as error:
                    if "UNIQUE constraint failed" in str(error):
                        duplicates += 1
                    else:
                        raise

            connection.commit()

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=inserted,
            errors_count=duplicates,
            message=(
                f"Inserted={inserted}; "
                f"duplicates={duplicates}"
            ),
        )

        return inserted, duplicates

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=inserted,
            errors_count=duplicates + 1,
            message=str(error),
        )
        raise


# ============================================================
# INTELLIGENCE CALCULATIONS
# ============================================================

def latest_chain() -> pd.DataFrame:
    setup_options_schema()

    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM options_chain
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM options_chain
            )
            ORDER BY nse_symbol, expiry_date, strike_price, option_type
            """,
            connection,
        )


def max_pain(chain: pd.DataFrame) -> float | None:
    if chain.empty:
        return None

    strikes = sorted(
        chain["strike_price"].dropna().unique()
    )

    if not strikes:
        return None

    pain_values = []

    calls = chain[
        chain["option_type"] == "CE"
    ][["strike_price", "open_interest"]]

    puts = chain[
        chain["option_type"] == "PE"
    ][["strike_price", "open_interest"]]

    for settlement in strikes:
        call_pain = (
            (
                (settlement - calls["strike_price"]).clip(lower=0)
                * calls["open_interest"]
            ).sum()
        )

        put_pain = (
            (
                (puts["strike_price"] - settlement).clip(lower=0)
                * puts["open_interest"]
            ).sum()
        )

        pain_values.append(
            (settlement, call_pain + put_pain)
        )

    return float(
        min(
            pain_values,
            key=lambda item: item[1],
        )[0]
    )


def calculate_options_score(
    oi_pcr: float,
    volume_pcr: float,
    modified_pcr: float,
    underlying: float,
    put_wall: float | None,
    call_wall: float | None,
) -> tuple[float, str, str]:
    score = 50.0
    reasons = []

    if oi_pcr >= 1.2:
        score += 15
        reasons.append("Strong Put OI dominance")
    elif oi_pcr >= 1.0:
        score += 8
        reasons.append("Put OI moderately dominant")
    elif oi_pcr <= 0.7:
        score -= 15
        reasons.append("Call OI strongly dominant")
    elif oi_pcr <= 0.9:
        score -= 8
        reasons.append("Call OI moderately dominant")

    if volume_pcr >= 1.2:
        score += 8
        reasons.append("Put volume dominance")
    elif volume_pcr <= 0.8:
        score -= 8
        reasons.append("Call volume dominance")

    if modified_pcr >= 1.2:
        score += 10
        reasons.append("Positive modified PCR")
    elif modified_pcr <= 0.8:
        score -= 10
        reasons.append("Negative modified PCR")

    if put_wall is not None and underlying >= put_wall:
        score += 5
        reasons.append("Price above Put OI support")

    if call_wall is not None and underlying <= call_wall:
        score += 2
        reasons.append("Price below Call OI resistance")

    score = round(max(0, min(100, score)), 2)

    if score >= 75:
        bias = "STRONG BULLISH"
    elif score >= 60:
        bias = "BULLISH"
    elif score <= 25:
        bias = "STRONG BEARISH"
    elif score <= 40:
        bias = "BEARISH"
    else:
        bias = "NEUTRAL"

    return score, bias, " | ".join(reasons)


def analyse_options() -> pd.DataFrame:
    chain = latest_chain()

    if chain.empty:
        return pd.DataFrame()

    rows = []

    for (symbol, expiry), group in chain.groupby(
        ["nse_symbol", "expiry_date"]
    ):
        calls = group[group["option_type"] == "CE"]
        puts = group[group["option_type"] == "PE"]

        call_oi = float(calls["open_interest"].sum())
        put_oi = float(puts["open_interest"].sum())

        call_oi_change = float(
            calls["oi_change"].fillna(0).sum()
        )
        put_oi_change = float(
            puts["oi_change"].fillna(0).sum()
        )

        call_volume = float(
            calls["volume"].fillna(0).sum()
        )
        put_volume = float(
            puts["volume"].fillna(0).sum()
        )

        oi_pcr = (
            put_oi / call_oi
            if call_oi
            else 0.0
        )

        volume_pcr = (
            put_volume / call_volume
            if call_volume
            else 0.0
        )

        modified_pcr = (
            (put_oi + max(0, put_oi_change))
            / (call_oi + max(0, call_oi_change))
            if (call_oi + max(0, call_oi_change))
            else 0.0
        )

        underlying = float(
            group["underlying_price"].dropna().iloc[-1]
        )

        call_wall = (
            float(
                calls.loc[
                    calls["open_interest"].idxmax(),
                    "strike_price",
                ]
            )
            if not calls.empty
            else None
        )

        put_wall = (
            float(
                puts.loc[
                    puts["open_interest"].idxmax(),
                    "strike_price",
                ]
            )
            if not puts.empty
            else None
        )

        pain = max_pain(group)

        score, bias, explanation = calculate_options_score(
            oi_pcr,
            volume_pcr,
            modified_pcr,
            underlying,
            put_wall,
            call_wall,
        )

        rows.append(
            {
                "trade_date": group["trade_date"].max(),
                "nse_symbol": symbol,
                "expiry_date": expiry,
                "underlying_price": underlying,
                "call_oi": call_oi,
                "put_oi": put_oi,
                "call_oi_change": call_oi_change,
                "put_oi_change": put_oi_change,
                "call_volume": call_volume,
                "put_volume": put_volume,
                "oi_pcr": round(oi_pcr, 3),
                "volume_pcr": round(volume_pcr, 3),
                "modified_pcr": round(modified_pcr, 3),
                "max_pain": pain,
                "call_wall": call_wall,
                "put_wall": put_wall,
                "options_support": put_wall,
                "options_resistance": call_wall,
                "directional_bias": bias,
                "options_score": score,
                "explanation": explanation,
            }
        )

    result = pd.DataFrame(rows).sort_values(
        "options_score",
        ascending=False,
    )

    with connect() as connection:
        for _, row in result.iterrows():
            connection.execute(
                """
                INSERT INTO options_intelligence(
                    trade_date,
                    nse_symbol,
                    expiry_date,
                    underlying_price,
                    call_oi,
                    put_oi,
                    call_oi_change,
                    put_oi_change,
                    call_volume,
                    put_volume,
                    oi_pcr,
                    volume_pcr,
                    modified_pcr,
                    max_pain,
                    call_wall,
                    put_wall,
                    options_support,
                    options_resistance,
                    directional_bias,
                    options_score,
                    explanation,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(trade_date, nse_symbol, expiry_date)
                DO UPDATE SET
                    underlying_price = excluded.underlying_price,
                    call_oi = excluded.call_oi,
                    put_oi = excluded.put_oi,
                    call_oi_change = excluded.call_oi_change,
                    put_oi_change = excluded.put_oi_change,
                    call_volume = excluded.call_volume,
                    put_volume = excluded.put_volume,
                    oi_pcr = excluded.oi_pcr,
                    volume_pcr = excluded.volume_pcr,
                    modified_pcr = excluded.modified_pcr,
                    max_pain = excluded.max_pain,
                    call_wall = excluded.call_wall,
                    put_wall = excluded.put_wall,
                    options_support = excluded.options_support,
                    options_resistance = excluded.options_resistance,
                    directional_bias = excluded.directional_bias,
                    options_score = excluded.options_score,
                    explanation = excluded.explanation,
                    created_at = excluded.created_at
                """,
                (
                    row["trade_date"],
                    row["nse_symbol"],
                    row["expiry_date"],
                    row["underlying_price"],
                    row["call_oi"],
                    row["put_oi"],
                    row["call_oi_change"],
                    row["put_oi_change"],
                    row["call_volume"],
                    row["put_volume"],
                    row["oi_pcr"],
                    row["volume_pcr"],
                    row["modified_pcr"],
                    row["max_pain"],
                    row["call_wall"],
                    row["put_wall"],
                    row["options_support"],
                    row["options_resistance"],
                    row["directional_bias"],
                    row["options_score"],
                    row["explanation"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()

    return result


# ============================================================
# EXCEL REPORT
# ============================================================

def write_report() -> None:
    summary = analyse_options()
    chain = latest_chain()

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Options Intelligence" in wb.sheetnames:
        del wb["Options Intelligence"]

    ws = wb.create_sheet("Options Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:V2")
    ws["A1"] = "AQSD PROFESSIONAL - OPTIONS OI & PCR INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Symbols Analysed"
    ws["B4"] = len(summary)
    ws["D4"] = "Option Rows"
    ws["E4"] = len(chain)
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Rank",
        "Symbol",
        "Expiry",
        "Underlying",
        "Call OI",
        "Put OI",
        "Call OI Change",
        "Put OI Change",
        "OI PCR",
        "Volume PCR",
        "Modified PCR",
        "Max Pain",
        "Call Wall",
        "Put Wall",
        "Support",
        "Resistance",
        "Options Bias",
        "Options Score",
        "Explanation",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        summary.iterrows(),
        start=8,
    ):
        values = [
            row_no - 7,
            row["nse_symbol"],
            row["expiry_date"],
            row["underlying_price"],
            row["call_oi"],
            row["put_oi"],
            row["call_oi_change"],
            row["put_oi_change"],
            row["oi_pcr"],
            row["volume_pcr"],
            row["modified_pcr"],
            row["max_pain"],
            row["call_wall"],
            row["put_wall"],
            row["options_support"],
            row["options_resistance"],
            row["directional_bias"],
            row["options_score"],
            row["explanation"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        score = float(row["options_score"])

        ws.cell(row_no, 18).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )

        bias = str(row["directional_bias"])

        ws.cell(row_no, 17).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BULLISH" in bias
                else RED
                if "BEARISH" in bias
                else GREY
            ),
        )

    widths = {
        "A": 8,
        "B": 16,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 16,
        "H": 16,
        "I": 12,
        "J": 12,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 14,
        "P": 14,
        "Q": 18,
        "R": 14,
        "S": 60,
        "T": 14,
        "U": 14,
        "V": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


# ============================================================
# STATUS
# ============================================================

def show_status() -> None:
    setup_options_schema()

    with connect() as connection:
        chain_row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT nse_symbol) AS symbols,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM options_chain
            """
        ).fetchone()

        intel_row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT nse_symbol) AS symbols,
                MAX(trade_date) AS latest_date
            FROM options_intelligence
            """
        ).fetchone()

    print("\nAQSD OPTIONS INTELLIGENCE STATUS")
    print("=" * 72)
    print(f"Stored option rows:       {chain_row['total'] or 0}")
    print(f"Option symbols covered:   {chain_row['symbols'] or 0}")
    print(f"First chain date:         {chain_row['first_date'] or 'No data'}")
    print(f"Latest chain date:        {chain_row['latest_date'] or 'No data'}")
    print("-" * 72)
    print(f"Intelligence records:     {intel_row['total'] or 0}")
    print(f"Intelligence symbols:     {intel_row['symbols'] or 0}")
    print(f"Latest intelligence date: {intel_row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Options OI and PCR Intelligence Engine."
    )

    parser.add_argument(
        "--setup",
        action="store_true",
        help="Create options intelligence tables.",
    )

    parser.add_argument(
        "--import-csv",
        type=Path,
        help="Import structured options-chain CSV data.",
    )

    parser.add_argument(
        "--source",
        default="Manual CSV",
        help="Data-source label.",
    )

    parser.add_argument(
        "--analyse",
        action="store_true",
        help="Calculate latest options intelligence.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Create the Options Intelligence Excel report.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show options intelligence status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_options_schema()

    if args.setup:
        print("AQSD Options Intelligence schema is ready.")
        return

    if args.import_csv:
        inserted, duplicates = import_csv(
            args.import_csv,
            args.source,
        )

        print("\nAQSD OPTIONS IMPORT")
        print("=" * 72)
        print(f"Inserted:   {inserted}")
        print(f"Duplicates: {duplicates}")
        return

    if args.analyse:
        frame = analyse_options()

        if frame.empty:
            print("No options intelligence available.")
        else:
            print(frame.to_string(index=False))
        return

    if args.report:
        write_report()
        print(f"Options Intelligence report created:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
