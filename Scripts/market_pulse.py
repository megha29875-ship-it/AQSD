from openpyxl import load_workbook
from openpyxl.styles import Font

"""
=========================================================
AQSD Professional
Module : Market Pulse
Version: 1.0
=========================================================
"""

from pathlib import Path
import pandas as pd
import yfinance as yf

# -----------------------------
# Paths
# -----------------------------

BASE = Path(__file__).resolve().parent.parent

OUTPUT_FILE = BASE / "Output" / "Dashboard.xlsx"

# -----------------------------
# Market Symbols
# -----------------------------

MARKETS = {
    "NIFTY": "^NSEI",
    "BANKNIFTY": "^NSEBANK",
    "INDIA VIX": "^INDIAVIX"
}

# -----------------------------
# RSI Function
# -----------------------------

def calculate_rsi(close, period=14):

    delta = close.diff()

    gain = delta.clip(lower=0)

    loss = -delta.clip(upper=0)

    avg_gain = gain.ewm(alpha=1/period, adjust=False).mean()

    avg_loss = loss.ewm(alpha=1/period, adjust=False).mean()

    rs = avg_gain / avg_loss

    return 100 - (100/(1+rs))

# -----------------------------
# Download Market Data
# -----------------------------

def get_market(symbol):

    df = yf.download(
        symbol,
        period="6mo",
        interval="1d",
        progress=False,
        auto_adjust=True
    )

    if df.empty:
        return None

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df["EMA20"] = df["Close"].ewm(span=20, adjust=False).mean()

    df["EMA50"] = df["Close"].ewm(span=50, adjust=False).mean()

    df["RSI"] = calculate_rsi(df["Close"])

    return df


def save_market_pulse(results, bias, confidence, strategy):

    if not OUTPUT_FILE.exists():
        print("Dashboard.xlsx not found.")
        return

    wb = load_workbook(OUTPUT_FILE)

    if "Market Pulse" in wb.sheetnames:
        del wb["Market Pulse"]

    ws = wb.create_sheet("Market Pulse", 0)

    ws["A1"] = "AQSD MARKET PULSE"
    ws["A1"].font = Font(size=16, bold=True)

    ws.append([])

    ws.append(["Index", "Close", "EMA20", "EMA50", "RSI"])

    for row in results:
        ws.append(row)

    ws.append([])

    ws.append(["Market Bias", bias])
    ws.append(["Confidence", f"{confidence}%"])
    ws.append(["Strategy", strategy])

    wb.save(OUTPUT_FILE)

    print("\nMarket Pulse saved to Dashboard.xlsx")


# =======================================================
# MARKET PULSE
# =======================================================

print("\nAQSD MARKET PULSE\n")

results = []

market_score = 0

for name, symbol in MARKETS.items():

    print(f"Downloading {name}...")

    df = get_market(symbol)

    if df is None:
        continue

    last = df.iloc[-1]

    close = last["Close"]
    ema20 = last["EMA20"]
    ema50 = last["EMA50"]
    rsi = last["RSI"]

    if close > ema20:
        market_score += 15

    if ema20 > ema50:
        market_score += 15

    if rsi > 55:
        market_score += 10

    print("--------------------------------")
    print(name)
    print(f"Close : {close:.2f}")
    print(f"EMA20 : {ema20:.2f}")
    print(f"EMA50 : {ema50:.2f}")
    print(f"RSI   : {rsi:.2f}")
    print()

# INDIA VIX Bonus
# Lower VIX is positive

vix = get_market("^INDIAVIX")

if vix is not None:

    last = vix.iloc[-1]

    if last["Close"] < last["EMA20"]:
        market_score += 20

confidence = min(market_score,100)

if confidence >= 75:
    bias = "🟢 BULLISH"
    strategy = "CALL BUYING"

elif confidence >= 50:
    bias = "🟡 NEUTRAL"
    strategy = "WAIT"

else:
    bias = "🔴 BEARISH"
    strategy = "PUT BUYING"

print("="*45)
print("AQSD MARKET PULSE")
print("="*45)
print("Market Bias :", bias)
print("Confidence  :", confidence,"%")
print("Strategy    :", strategy)
print("="*45)

save_market_pulse(
    results,
    bias,
    confidence,
    strategy
)