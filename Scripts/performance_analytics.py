
"""
AQSD Professional
Module: Performance Analytics
Version: 1.0
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_closed_trades(wb) -> list[dict]:
    if "Trade Journal" not in wb.sheetnames:
        return []

    ws = wb["Trade Journal"]
    headers = header_map(ws, 10)
    required = ["Symbol", "Entry Date", "Exit Date", "P/L", "P/L %", "Result"]

    if any(name not in headers for name in required):
        return []

    trades = []

    for row in range(11, ws.max_row + 1):
        symbol = ws.cell(row, headers["Symbol"]).value

        if not symbol:
            continue

        entry_date = ws.cell(row, headers["Entry Date"]).value
        exit_date = ws.cell(row, headers["Exit Date"]).value
        holding_days = 0

        if isinstance(entry_date, datetime) and isinstance(exit_date, datetime):
            holding_days = max((exit_date - entry_date).days, 0)

        trades.append(
            {
                "Symbol": str(symbol),
                "Entry Date": entry_date,
                "Exit Date": exit_date,
                "P/L": safe_float(ws.cell(row, headers["P/L"]).value),
                "P/L %": safe_float(ws.cell(row, headers["P/L %"]).value),
                "Result": str(ws.cell(row, headers["Result"]).value or "").upper(),
                "Holding Days": holding_days,
            }
        )

    return trades


def read_open_positions(wb) -> list[dict]:
    if "Portfolio" not in wb.sheetnames:
        return []

    ws = wb["Portfolio"]
    headers = header_map(ws, 12)

    if any(name not in headers for name in ["Symbol", "P/L", "Status"]):
        return []

    positions = []

    for row in range(13, ws.max_row + 1):
        symbol = ws.cell(row, headers["Symbol"]).value

        if not symbol:
            continue

        status = str(ws.cell(row, headers["Status"]).value or "").upper().strip()

        if status == "OPEN":
            positions.append(
                {
                    "Symbol": str(symbol),
                    "P/L": safe_float(ws.cell(row, headers["P/L"]).value),
                }
            )

    return positions


def max_drawdown(values: list[float]) -> float:
    peak = 0.0
    worst = 0.0

    for value in values:
        peak = max(peak, value)
        worst = min(worst, value - peak)

    return abs(worst)


def calculate_metrics(closed_trades: list[dict], open_positions: list[dict]) -> dict:
    wins = [trade for trade in closed_trades if trade["P/L"] > 0]
    losses = [trade for trade in closed_trades if trade["P/L"] < 0]

    gross_profit = sum(trade["P/L"] for trade in wins)
    gross_loss = abs(sum(trade["P/L"] for trade in losses))
    closed_pl = sum(trade["P/L"] for trade in closed_trades)
    open_mtm = sum(position["P/L"] for position in open_positions)

    closed_count = len(closed_trades)
    average_win = gross_profit / len(wins) if wins else 0.0
    average_loss = gross_loss / len(losses) if losses else 0.0
    win_rate = len(wins) / closed_count * 100 if closed_count else 0.0
    loss_rate = len(losses) / closed_count if closed_count else 0.0
    profit_factor = gross_profit / gross_loss if gross_loss else (gross_profit if gross_profit else 0.0)
    expectancy = (win_rate / 100) * average_win - loss_rate * average_loss

    running = 0.0
    equity = []

    for trade in closed_trades:
        running += trade["P/L"]
        equity.append(running)

    return {
        "Total Trades": closed_count + len(open_positions),
        "Open Trades": len(open_positions),
        "Closed Trades": closed_count,
        "Winning Trades": len(wins),
        "Losing Trades": len(losses),
        "Win Rate": win_rate,
        "Closed P/L": closed_pl,
        "Open MTM": open_mtm,
        "Combined P/L": closed_pl + open_mtm,
        "Average Win": average_win,
        "Average Loss": average_loss,
        "Profit Factor": profit_factor,
        "Expectancy": expectancy,
        "Largest Winner": max((t["P/L"] for t in closed_trades), default=0.0),
        "Largest Loser": min((t["P/L"] for t in closed_trades), default=0.0),
        "Average Holding Days": (
            sum(t["Holding Days"] for t in closed_trades) / closed_count
            if closed_count else 0.0
        ),
        "Maximum Drawdown": max_drawdown(equity),
    }


def create_analytics_sheet(wb, closed_trades: list[dict], open_positions: list[dict]) -> None:
    if "Analytics" in wb.sheetnames:
        del wb["Analytics"]

    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H2")
    ws["A1"] = "AQSD PROFESSIONAL - PERFORMANCE ANALYTICS"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    metrics = calculate_metrics(closed_trades, open_positions)

    for row_no, (label, value) in enumerate(metrics.items(), start=4):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)
        ws[f"A{row_no}"].border = Border(bottom=THIN)
        ws[f"B{row_no}"].border = Border(bottom=THIN)

        if label == "Win Rate":
            ws[f"B{row_no}"].number_format = '0.00"%"'
        elif label in {
            "Closed P/L", "Open MTM", "Combined P/L", "Average Win",
            "Average Loss", "Expectancy", "Largest Winner",
            "Largest Loser", "Maximum Drawdown",
        }:
            ws[f"B{row_no}"].number_format = '₹#,##0.00'
        elif label in {"Profit Factor", "Average Holding Days"}:
            ws[f"B{row_no}"].number_format = "0.00"

    ws["D3"] = "Trade No."
    ws["E3"] = "Equity P/L"
    ws["G3"] = "Month"
    ws["H3"] = "Monthly P/L"
    ws["J3"] = "Result"
    ws["K3"] = "Count"

    for ref in ("D3", "E3", "G3", "H3", "J3", "K3"):
        ws[ref].font = Font(bold=True, color=WHITE)
        ws[ref].fill = PatternFill("solid", fgColor=NAVY)

    running = 0.0

    for index, trade in enumerate(closed_trades, start=1):
        running += trade["P/L"]
        ws.cell(index + 3, 4, index)
        ws.cell(index + 3, 5, round(running, 2))
        ws.cell(index + 3, 5).number_format = '₹#,##0.00'

    monthly = defaultdict(float)

    for trade in closed_trades:
        exit_date = trade["Exit Date"]
        key = exit_date.strftime("%Y-%m") if isinstance(exit_date, datetime) else "Unknown"
        monthly[key] += trade["P/L"]

    for index, key in enumerate(sorted(monthly), start=4):
        ws.cell(index, 7, key)
        ws.cell(index, 8, round(monthly[key], 2))
        ws.cell(index, 8).number_format = '₹#,##0.00'

    wins = sum(1 for trade in closed_trades if trade["P/L"] > 0)
    losses = sum(1 for trade in closed_trades if trade["P/L"] < 0)
    ws["J4"] = "Wins"
    ws["K4"] = wins
    ws["J5"] = "Losses"
    ws["K5"] = losses

    if closed_trades:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.y_axis.title = "Cumulative P/L"
        chart.x_axis.title = "Trade Number"
        chart.add_data(
            Reference(ws, min_col=5, min_row=3, max_row=3 + len(closed_trades)),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(ws, min_col=4, min_row=4, max_row=3 + len(closed_trades))
        )
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, "D22")

    if monthly:
        chart = BarChart()
        chart.title = "Monthly P/L"
        chart.add_data(
            Reference(ws, min_col=8, min_row=3, max_row=3 + len(monthly)),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(ws, min_col=7, min_row=4, max_row=3 + len(monthly))
        )
        chart.height = 7
        chart.width = 12
        ws.add_chart(chart, "J22")

    pie = PieChart()
    pie.title = "Win / Loss Distribution"
    pie.add_data(Reference(ws, min_col=11, min_row=3, max_row=5), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=10, min_row=4, max_row=5))
    pie.height = 7
    pie.width = 9
    ws.add_chart(pie, "J7")

    widths = {
        "A": 24, "B": 16, "C": 4, "D": 12, "E": 16,
        "F": 4, "G": 12, "H": 16, "I": 4, "J": 12, "K": 10,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    combined_pl_row = 12
    combined_pl = metrics["Combined P/L"]
    ws[f"B{combined_pl_row}"].fill = PatternFill(
        "solid",
        fgColor=GREEN if combined_pl > 0 else RED if combined_pl < 0 else YELLOW,
    )


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(f"Dashboard not found:\n{DASHBOARD}")

    wb = load_workbook(DASHBOARD)
    closed_trades = read_closed_trades(wb)
    open_positions = read_open_positions(wb)

    create_analytics_sheet(wb, closed_trades, open_positions)
    wb.save(DASHBOARD)

    print("Performance Analytics created successfully.")
    print(f"Closed trades analysed: {len(closed_trades)}")
    print(f"Open positions analysed: {len(open_positions)}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
