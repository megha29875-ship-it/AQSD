
"""
AQSD Professional
Module: System Health Check
Version: 1.0

Checks the AQSD installation and creates a Health Check sheet
inside Dashboard.xlsx.

Checks include:
- Required folders
- Required Python modules
- Required scripts
- Configuration file
- Dashboard file
- Important Excel sheets
- Git repository
- Backup folder
"""

from __future__ import annotations

import importlib.util
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

SCRIPTS_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPTS_DIR.parent

DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"
CONFIG_FILE = BASE_DIR / "Config" / "AQSD_Config.json"
BACKUP_DIR = BASE_DIR / "Backups"
GIT_DIR = BASE_DIR / ".git"


# ============================================================
# SETTINGS
# ============================================================

REQUIRED_FOLDERS = [
    BASE_DIR / "Scripts",
    BASE_DIR / "Data",
    BASE_DIR / "Output",
    BASE_DIR / "Config",
]

REQUIRED_MODULES = [
    "pandas",
    "yfinance",
    "openpyxl",
]

REQUIRED_SCRIPTS = [
    "AQSD.py",
    "scanner.py",
    "market_pulse.py",
    "format_dashboard.py",
    "risk_manager.py",
    "risk_dashboard.py",
    "portfolio_manager.py",
    "portfolio_live_assistant.py",
    "trade_journal.py",
    "performance_analytics.py",
    "live_watchlist.py",
    "portfolio_heatmap.py",
    "auto_backup.py",
    "smart_alerts.py",
    "daily_trading_report.py",
    "aqsd_launcher.py",
]

IMPORTANT_SHEETS = [
    "HOME",
    "Market Pulse",
    "Option Buying",
    "CALL Candidates",
    "PUT Candidates",
    "Portfolio",
    "Trade Journal",
    "Analytics",
    "Live Watchlist",
    "Alerts",
    "Daily Report",
]


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# CHECKS
# ============================================================

def add_result(
    results: list[dict],
    category: str,
    item: str,
    ok: bool,
    details: str,
) -> None:
    results.append(
        {
            "Category": category,
            "Item": item,
            "Status": "PASS" if ok else "FAIL",
            "Details": details,
        }
    )


def check_folders(results: list[dict]) -> None:
    for folder in REQUIRED_FOLDERS:
        add_result(
            results,
            "Folder",
            folder.name,
            folder.exists(),
            str(folder),
        )


def check_modules(results: list[dict]) -> None:
    for module_name in REQUIRED_MODULES:
        available = importlib.util.find_spec(module_name) is not None

        add_result(
            results,
            "Python Module",
            module_name,
            available,
            "Installed" if available else "Missing",
        )


def check_scripts(results: list[dict]) -> None:
    for script_name in REQUIRED_SCRIPTS:
        script_path = SCRIPTS_DIR / script_name

        add_result(
            results,
            "Script",
            script_name,
            script_path.exists(),
            str(script_path),
        )


def check_project_files(results: list[dict]) -> None:
    add_result(
        results,
        "Project",
        "Configuration File",
        CONFIG_FILE.exists(),
        str(CONFIG_FILE),
    )

    add_result(
        results,
        "Project",
        "Dashboard Workbook",
        DASHBOARD.exists(),
        str(DASHBOARD),
    )

    add_result(
        results,
        "Project",
        "Git Repository",
        GIT_DIR.exists(),
        str(GIT_DIR),
    )

    add_result(
        results,
        "Project",
        "Backup Folder",
        BACKUP_DIR.exists(),
        str(BACKUP_DIR),
    )


def check_dashboard_sheets(results: list[dict]) -> None:
    if not DASHBOARD.exists():
        for sheet_name in IMPORTANT_SHEETS:
            add_result(
                results,
                "Dashboard Sheet",
                sheet_name,
                False,
                "Dashboard.xlsx not found",
            )
        return

    try:
        wb = load_workbook(
            DASHBOARD,
            read_only=True,
            data_only=False,
        )

        sheetnames = set(wb.sheetnames)
        wb.close()

        for sheet_name in IMPORTANT_SHEETS:
            add_result(
                results,
                "Dashboard Sheet",
                sheet_name,
                sheet_name in sheetnames,
                "Available" if sheet_name in sheetnames else "Missing",
            )

    except Exception as error:
        add_result(
            results,
            "Dashboard",
            "Workbook Read Test",
            False,
            str(error),
        )


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_health_sheet(results: list[dict]) -> None:
    if not DASHBOARD.exists():
        return

    wb = load_workbook(DASHBOARD)

    if "Health Check" in wb.sheetnames:
        del wb["Health Check"]

    ws = wb.create_sheet("Health Check", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:D2")
    ws["A1"] = "AQSD PROFESSIONAL - SYSTEM HEALTH CHECK"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    passed = sum(1 for item in results if item["Status"] == "PASS")
    failed = len(results) - passed
    score = round(passed / len(results) * 100, 2) if results else 0.0

    summary = [
        ("Last Checked", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("Checks Passed", passed),
        ("Checks Failed", failed),
        ("Health Score", score),
    ]

    for row_no, (label, value) in enumerate(summary, start=4):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)

    ws["B7"].number_format = '0.00"%"'
    ws["B7"].fill = PatternFill(
        "solid",
        fgColor=GREEN if score >= 90 else YELLOW if score >= 75 else RED,
    )
    ws["B7"].font = Font(bold=True)

    headers = ["Category", "Item", "Status", "Details"]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(9, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    for row_no, item in enumerate(results, start=10):
        values = [
            item["Category"],
            item["Item"],
            item["Status"],
            item["Details"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            cell.border = Border(bottom=THIN)

        ws.cell(row_no, 3).fill = PatternFill(
            "solid",
            fgColor=GREEN if item["Status"] == "PASS" else RED,
        )
        ws.cell(row_no, 3).font = Font(bold=True)

    widths = {
        "A": 22,
        "B": 28,
        "C": 12,
        "D": 70,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    results: list[dict] = []

    check_folders(results)
    check_modules(results)
    check_scripts(results)
    check_project_files(results)
    check_dashboard_sheets(results)

    passed = sum(1 for item in results if item["Status"] == "PASS")
    failed = len(results) - passed
    score = round(passed / len(results) * 100, 2) if results else 0.0

    print("\nAQSD SYSTEM HEALTH CHECK")
    print("=" * 72)

    for item in results:
        print(
            f"{item['Status']:<5} | "
            f"{item['Category']:<18} | "
            f"{item['Item']}"
        )

    print("=" * 72)
    print(f"Checks passed: {passed}")
    print(f"Checks failed: {failed}")
    print(f"Health score: {score:.2f}%")

    if DASHBOARD.exists():
        try:
            write_health_sheet(results)
            print("Health Check sheet added to Dashboard.xlsx")
        except PermissionError:
            print(
                "Dashboard.xlsx is open. "
                "Close Excel to write the Health Check sheet."
            )

    print("=" * 72)


if __name__ == "__main__":
    main()
