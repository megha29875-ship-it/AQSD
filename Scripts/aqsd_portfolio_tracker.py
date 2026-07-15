
"""
AQSD Professional
Module: Portfolio Tracker
Version: 1.0

Purpose
-------
Tracks portfolio transactions, open positions, realized and unrealized P&L.

Commands
--------
python aqsd_portfolio_tracker.py --setup
python aqsd_portfolio_tracker.py --add
python aqsd_portfolio_tracker.py --import-csv PATH
python aqsd_portfolio_tracker.py --update
python aqsd_portfolio_tracker.py --status
python aqsd_portfolio_tracker.py --report
python aqsd_portfolio_tracker.py --positions
python aqsd_portfolio_tracker.py --transactions
"""

from __future__ import annotations

import argparse
import csv
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database


BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data"
OUTPUT_DIR = BASE_DIR / "Output"

DB_PATH = DATA_DIR / "aqsd_core.db"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"

POSITIONS_CSV = OUTPUT_DIR / "AQSD_Portfolio_Positions.csv"
TRANSACTIONS_CSV = OUTPUT_DIR / "AQSD_Portfolio_Transactions.csv"
SUMMARY_CSV = OUTPUT_DIR / "AQSD_Portfolio_Summary.csv"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


SCHEMA = """
CREATE TABLE IF NOT EXISTS portfolio_transactions (
    transaction_id INTEGER PRIMARY KEY AUTOINCREMENT,
    transaction_date TEXT NOT NULL,
    nse_symbol TEXT NOT NULL,
    transaction_type TEXT NOT NULL,
    quantity REAL NOT NULL,
    price REAL NOT NULL,
    charges REAL NOT NULL DEFAULT 0,
    notes TEXT,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_portfolio_transactions_symbol
ON portfolio_transactions(nse_symbol, transaction_date);

CREATE TABLE IF NOT EXISTS portfolio_positions (
    nse_symbol TEXT PRIMARY KEY,
    quantity REAL NOT NULL,
    average_price REAL NOT NULL,
    invested_value REAL NOT NULL,
    latest_price REAL,
    market_value REAL,
    unrealized_pnl REAL,
    unrealized_pnl_percent REAL,
    realized_pnl REAL NOT NULL DEFAULT 0,
    sector TEXT,
    last_updated TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS portfolio_summary (
    summary_date TEXT PRIMARY KEY,
    invested_value REAL,
    market_value REAL,
    unrealized_pnl REAL,
    realized_pnl REAL,
    total_pnl REAL,
    total_return_percent REAL,
    open_positions INTEGER,
    cash_balance REAL,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS portfolio_settings (
    setting_key TEXT PRIMARY KEY,
    setting_value TEXT NOT NULL
);
"""


def setup_tracker() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.execute(
            """
            INSERT OR IGNORE INTO portfolio_settings(
                setting_key,
                setting_value
            )
            VALUES ('initial_cash', '0')
            """
        )
        connection.commit()

    print("AQSD Portfolio Tracker schema is ready.")


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_symbol(symbol: str) -> str:
    text = symbol.strip().upper()
    if text.endswith(".NS"):
        text = text[:-3]
    return text


def add_transaction_interactive() -> None:
    setup_tracker()

    transaction_date = input(
        "Transaction date YYYY-MM-DD: "
    ).strip()

    symbol = normalize_symbol(
        input("NSE symbol: ").strip()
    )

    transaction_type = input(
        "Type BUY or SELL: "
    ).strip().upper()

    if transaction_type not in {"BUY", "SELL"}:
        raise ValueError("Type must be BUY or SELL.")

    quantity = float(
        input("Quantity: ").strip()
    )

    price = float(
        input("Price: ").strip()
    )

    charges_text = input(
        "Charges [0]: "
    ).strip()

    charges = (
        float(charges_text)
        if charges_text
        else 0.0
    )

    notes = input(
        "Notes [optional]: "
    ).strip()

    insert_transaction(
        transaction_date=transaction_date,
        symbol=symbol,
        transaction_type=transaction_type,
        quantity=quantity,
        price=price,
        charges=charges,
        notes=notes,
    )

    rebuild_positions()
    print("Transaction added successfully.")


def insert_transaction(
    transaction_date: str,
    symbol: str,
    transaction_type: str,
    quantity: float,
    price: float,
    charges: float = 0.0,
    notes: str = "",
) -> None:
    if quantity <= 0:
        raise ValueError("Quantity must be greater than zero.")

    if price <= 0:
        raise ValueError("Price must be greater than zero.")

    with connect() as connection:
        connection.execute(
            """
            INSERT INTO portfolio_transactions(
                transaction_date,
                nse_symbol,
                transaction_type,
                quantity,
                price,
                charges,
                notes,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                transaction_date,
                normalize_symbol(symbol),
                transaction_type.upper(),
                quantity,
                price,
                charges,
                notes,
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )
        connection.commit()


def import_csv(path_text: str) -> None:
    setup_tracker()

    path = Path(path_text)

    if not path.exists():
        raise FileNotFoundError(path)

    frame = pd.read_csv(path)

    required = {
        "transaction_date",
        "nse_symbol",
        "transaction_type",
        "quantity",
        "price",
    }

    missing = required - set(frame.columns)

    if missing:
        raise ValueError(
            f"Missing CSV columns: {sorted(missing)}"
        )

    inserted = 0

    for _, row in frame.iterrows():
        insert_transaction(
            transaction_date=str(
                row["transaction_date"]
            ),
            symbol=str(
                row["nse_symbol"]
            ),
            transaction_type=str(
                row["transaction_type"]
            ),
            quantity=float(
                row["quantity"]
            ),
            price=float(
                row["price"]
            ),
            charges=safe_float(
                row.get("charges"),
                0.0,
            ),
            notes=str(
                row.get("notes") or ""
            ),
        )
        inserted += 1

    rebuild_positions()

    print(f"Transactions imported: {inserted}")


def latest_prices() -> dict[str, float]:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                s.nse_symbol,
                p.close
            FROM daily_prices p
            JOIN symbols s
                ON s.symbol_id = p.symbol_id
            JOIN (
                SELECT
                    symbol_id,
                    MAX(trade_date) AS max_date
                FROM daily_prices
                GROUP BY symbol_id
            ) latest
                ON latest.symbol_id = p.symbol_id
               AND latest.max_date = p.trade_date
            """,
            connection,
        )

    return {
        normalize_symbol(str(row["nse_symbol"])): float(
            row["close"]
        )
        for _, row in frame.iterrows()
    }


def sector_lookup() -> dict[str, str]:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                nse_symbol,
                COALESCE(NULLIF(TRIM(sector), ''), 'Unmapped') AS sector
            FROM symbols
            """,
            connection,
        )

    return {
        normalize_symbol(str(row["nse_symbol"])): str(
            row["sector"]
        )
        for _, row in frame.iterrows()
    }


def rebuild_positions() -> None:
    setup_tracker()

    with connect() as connection:
        transactions = pd.read_sql_query(
            """
            SELECT *
            FROM portfolio_transactions
            ORDER BY transaction_date, transaction_id
            """,
            connection,
        )

    prices = latest_prices()
    sectors = sector_lookup()

    position_map: dict[str, dict[str, float]] = {}

    for _, row in transactions.iterrows():
        symbol = normalize_symbol(
            str(row["nse_symbol"])
        )

        transaction_type = str(
            row["transaction_type"]
        ).upper()

        quantity = float(
            row["quantity"]
        )

        price = float(
            row["price"]
        )

        charges = safe_float(
            row.get("charges"),
            0.0,
        )

        state = position_map.setdefault(
            symbol,
            {
                "quantity": 0.0,
                "average_price": 0.0,
                "realized_pnl": 0.0,
            },
        )

        if transaction_type == "BUY":
            old_qty = state["quantity"]
            old_cost = (
                old_qty * state["average_price"]
            )

            new_cost = (
                quantity * price + charges
            )

            new_qty = old_qty + quantity

            state["quantity"] = new_qty
            state["average_price"] = (
                (old_cost + new_cost) / new_qty
                if new_qty > 0
                else 0.0
            )

        elif transaction_type == "SELL":
            sell_qty = min(
                quantity,
                state["quantity"],
            )

            realized = (
                sell_qty
                * (
                    price
                    - state["average_price"]
                )
                - charges
            )

            state["realized_pnl"] += realized
            state["quantity"] -= sell_qty

            if state["quantity"] <= 0:
                state["quantity"] = 0.0
                state["average_price"] = 0.0

    rows = []

    for symbol, state in position_map.items():
        quantity = state["quantity"]
        average_price = state["average_price"]
        invested_value = (
            quantity * average_price
        )

        latest_price = prices.get(symbol)
        market_value = (
            quantity * latest_price
            if latest_price is not None
            else None
        )

        unrealized_pnl = (
            market_value - invested_value
            if market_value is not None
            else None
        )

        unrealized_percent = (
            unrealized_pnl
            / invested_value
            * 100
            if (
                unrealized_pnl is not None
                and invested_value > 0
            )
            else None
        )

        rows.append(
            {
                "nse_symbol": symbol,
                "quantity": round(
                    quantity,
                    4,
                ),
                "average_price": round(
                    average_price,
                    2,
                ),
                "invested_value": round(
                    invested_value,
                    2,
                ),
                "latest_price": (
                    round(
                        latest_price,
                        2,
                    )
                    if latest_price is not None
                    else None
                ),
                "market_value": (
                    round(
                        market_value,
                        2,
                    )
                    if market_value is not None
                    else None
                ),
                "unrealized_pnl": (
                    round(
                        unrealized_pnl,
                        2,
                    )
                    if unrealized_pnl is not None
                    else None
                ),
                "unrealized_pnl_percent": (
                    round(
                        unrealized_percent,
                        2,
                    )
                    if unrealized_percent is not None
                    else None
                ),
                "realized_pnl": round(
                    state["realized_pnl"],
                    2,
                ),
                "sector": sectors.get(
                    symbol,
                    "Unmapped",
                ),
                "last_updated": datetime.now().isoformat(
                    timespec="seconds"
                ),
            }
        )

    with connect() as connection:
        connection.execute(
            "DELETE FROM portfolio_positions"
        )

        for row in rows:
            connection.execute(
                """
                INSERT INTO portfolio_positions(
                    nse_symbol,
                    quantity,
                    average_price,
                    invested_value,
                    latest_price,
                    market_value,
                    unrealized_pnl,
                    unrealized_pnl_percent,
                    realized_pnl,
                    sector,
                    last_updated
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["nse_symbol"],
                    row["quantity"],
                    row["average_price"],
                    row["invested_value"],
                    row["latest_price"],
                    row["market_value"],
                    row["unrealized_pnl"],
                    row["unrealized_pnl_percent"],
                    row["realized_pnl"],
                    row["sector"],
                    row["last_updated"],
                ),
            )

        connection.commit()

    save_summary()


def get_cash_balance(
    transactions: pd.DataFrame,
) -> float:
    with connect() as connection:
        row = connection.execute(
            """
            SELECT setting_value
            FROM portfolio_settings
            WHERE setting_key='initial_cash'
            """
        ).fetchone()

    initial_cash = safe_float(
        row["setting_value"] if row else 0,
        0.0,
    )

    cash = initial_cash

    for _, tx in transactions.iterrows():
        gross = (
            float(tx["quantity"])
            * float(tx["price"])
        )

        charges = safe_float(
            tx.get("charges"),
            0.0,
        )

        if str(
            tx["transaction_type"]
        ).upper() == "BUY":
            cash -= gross + charges
        else:
            cash += gross - charges

    return round(cash, 2)


def save_summary() -> None:
    with connect() as connection:
        positions = pd.read_sql_query(
            """
            SELECT *
            FROM portfolio_positions
            """,
            connection,
        )

        transactions = pd.read_sql_query(
            """
            SELECT *
            FROM portfolio_transactions
            """,
            connection,
        )

    invested_value = safe_float(
        positions["invested_value"].sum()
        if not positions.empty
        else 0.0,
        0.0,
    )

    market_value = safe_float(
        positions["market_value"].fillna(0).sum()
        if not positions.empty
        else 0.0,
        0.0,
    )

    unrealized_pnl = safe_float(
        positions["unrealized_pnl"].fillna(0).sum()
        if not positions.empty
        else 0.0,
        0.0,
    )

    realized_pnl = safe_float(
        positions["realized_pnl"].fillna(0).sum()
        if not positions.empty
        else 0.0,
        0.0,
    )

    total_pnl = unrealized_pnl + realized_pnl

    total_return_percent = (
        total_pnl / invested_value * 100
        if invested_value > 0
        else 0.0
    )

    open_positions = int(
        (
            positions["quantity"] > 0
        ).sum()
        if not positions.empty
        else 0
    )

    cash_balance = get_cash_balance(
        transactions
    )

    summary_date = datetime.now().date().isoformat()

    with connect() as connection:
        connection.execute(
            """
            INSERT INTO portfolio_summary(
                summary_date,
                invested_value,
                market_value,
                unrealized_pnl,
                realized_pnl,
                total_pnl,
                total_return_percent,
                open_positions,
                cash_balance,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(summary_date)
            DO UPDATE SET
                invested_value=excluded.invested_value,
                market_value=excluded.market_value,
                unrealized_pnl=excluded.unrealized_pnl,
                realized_pnl=excluded.realized_pnl,
                total_pnl=excluded.total_pnl,
                total_return_percent=
                    excluded.total_return_percent,
                open_positions=excluded.open_positions,
                cash_balance=excluded.cash_balance,
                created_at=excluded.created_at
            """,
            (
                summary_date,
                round(
                    invested_value,
                    2,
                ),
                round(
                    market_value,
                    2,
                ),
                round(
                    unrealized_pnl,
                    2,
                ),
                round(
                    realized_pnl,
                    2,
                ),
                round(
                    total_pnl,
                    2,
                ),
                round(
                    total_return_percent,
                    2,
                ),
                open_positions,
                cash_balance,
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )
        connection.commit()


def latest_frames() -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
]:
    setup_tracker()

    with connect() as connection:
        positions = pd.read_sql_query(
            """
            SELECT *
            FROM portfolio_positions
            ORDER BY market_value DESC
            """,
            connection,
        )

        transactions = pd.read_sql_query(
            """
            SELECT *
            FROM portfolio_transactions
            ORDER BY transaction_date DESC, transaction_id DESC
            """,
            connection,
        )

        summary = pd.read_sql_query(
            """
            SELECT *
            FROM portfolio_summary
            ORDER BY summary_date DESC
            LIMIT 1
            """,
            connection,
        )

    return positions, transactions, summary


def write_reports() -> None:
    positions, transactions, summary = latest_frames()

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    positions.to_csv(
        POSITIONS_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    transactions.to_csv(
        TRANSACTIONS_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    summary.to_csv(
        SUMMARY_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    if DASHBOARD.exists():
        workbook = load_workbook(
            DASHBOARD
        )
    else:
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    sheet_name = "Portfolio Tracker"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(
        sheet_name,
        1,
    )

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"

    ws.merge_cells("A1:L2")
    ws["A1"] = "AQSD PROFESSIONAL - PORTFOLIO TRACKER"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    latest = (
        summary.iloc[0]
        if not summary.empty
        else {}
    )

    summary_values = [
        ("Invested Value", latest.get("invested_value")),
        ("Market Value", latest.get("market_value")),
        ("Unrealized P&L", latest.get("unrealized_pnl")),
        ("Realized P&L", latest.get("realized_pnl")),
        ("Total P&L", latest.get("total_pnl")),
        ("Return %", latest.get("total_return_percent")),
        ("Open Positions", latest.get("open_positions")),
        ("Cash Balance", latest.get("cash_balance")),
    ]

    for index, (label, value) in enumerate(
        summary_values,
        start=1,
    ):
        col = ((index - 1) % 4) * 3 + 1
        row = 4 if index <= 4 else 6

        ws.cell(
            row,
            col,
            label,
        )
        ws.cell(
            row,
            col + 1,
            value,
        )

        ws.cell(
            row,
            col,
        ).font = Font(
            bold=True
        )
        ws.cell(
            row,
            col,
        ).fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Symbol",
        "Sector",
        "Quantity",
        "Average Price",
        "Latest Price",
        "Invested Value",
        "Market Value",
        "Unrealized P&L",
        "Unrealized P&L %",
        "Realized P&L",
        "Last Updated",
    ]

    for col, heading in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(
            8,
            col,
            heading,
        )
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        positions.iterrows(),
        start=9,
    ):
        values = [
            row.get("nse_symbol"),
            row.get("sector"),
            row.get("quantity"),
            row.get("average_price"),
            row.get("latest_price"),
            row.get("invested_value"),
            row.get("market_value"),
            row.get("unrealized_pnl"),
            row.get("unrealized_pnl_percent"),
            row.get("realized_pnl"),
            row.get("last_updated"),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(
                bottom=THIN
            )

        pnl = safe_float(
            row.get("unrealized_pnl"),
            0.0,
        )

        ws.cell(
            row_no,
            8,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if pnl > 0
                else RED
                if pnl < 0
                else YELLOW
            ),
        )

    widths = {
        "A": 16,
        "B": 20,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 16,
        "G": 16,
        "H": 16,
        "I": 18,
        "J": 16,
        "K": 22,
        "L": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    workbook.save(
        DASHBOARD
    )


def show_positions() -> None:
    positions, _, _ = latest_frames()

    if positions.empty:
        print("No portfolio positions found.")
        return

    print(
        positions.to_string(
            index=False
        )
    )


def show_transactions() -> None:
    _, transactions, _ = latest_frames()

    if transactions.empty:
        print("No portfolio transactions found.")
        return

    print(
        transactions.to_string(
            index=False
        )
    )


def show_status() -> None:
    setup_tracker()

    with connect() as connection:
        transaction_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM portfolio_transactions
            """
        ).fetchone()[0]

        position_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM portfolio_positions
            WHERE quantity > 0
            """
        ).fetchone()[0]

        summary = connection.execute(
            """
            SELECT *
            FROM portfolio_summary
            ORDER BY summary_date DESC
            LIMIT 1
            """
        ).fetchone()

    print("\nAQSD PORTFOLIO TRACKER STATUS")
    print("=" * 72)
    print(f"Transactions:   {transaction_count}")
    print(f"Open positions: {position_count}")

    if summary:
        print(
            f"Market value:   "
            f"{summary['market_value']}"
        )
        print(
            f"Total P&L:      "
            f"{summary['total_pnl']}"
        )
        print(
            f"Return %:       "
            f"{summary['total_return_percent']}"
        )
        print(
            f"Cash balance:   "
            f"{summary['cash_balance']}"
        )

    print("=" * 72)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Portfolio Tracker."
    )

    parser.add_argument(
        "--setup",
        action="store_true",
    )

    parser.add_argument(
        "--add",
        action="store_true",
    )

    parser.add_argument(
        "--import-csv",
        metavar="PATH",
    )

    parser.add_argument(
        "--update",
        action="store_true",
    )

    parser.add_argument(
        "--report",
        action="store_true",
    )

    parser.add_argument(
        "--status",
        action="store_true",
    )

    parser.add_argument(
        "--positions",
        action="store_true",
    )

    parser.add_argument(
        "--transactions",
        action="store_true",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if args.setup:
        setup_tracker()
        return

    if args.add:
        add_transaction_interactive()
        return

    if args.import_csv:
        import_csv(
            args.import_csv
        )
        return

    if args.update:
        rebuild_positions()
        write_reports()
        print("Portfolio updated successfully.")
        print(f"Dashboard: {DASHBOARD}")
        return

    if args.report:
        write_reports()
        print(f"Positions CSV:    {POSITIONS_CSV}")
        print(f"Transactions CSV: {TRANSACTIONS_CSV}")
        print(f"Summary CSV:      {SUMMARY_CSV}")
        print(f"Dashboard:        {DASHBOARD}")
        return

    if args.positions:
        show_positions()
        return

    if args.transactions:
        show_transactions()
        return

    show_status()


if __name__ == "__main__":
    main()
