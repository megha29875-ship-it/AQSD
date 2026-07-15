
"""
AQSD Professional
Module: Trade Decision & Watchlist Engine
Version: 1.0

Purpose
-------
Converts Unified Master Intelligence into an actionable daily watchlist.

Inputs
------
- unified_master_intelligence
- price_structure_intelligence
- daily_prices
- sector_rotation_intelligence

Outputs
-------
- Action: BUY / BUY ON DIP / WATCH / AVOID / EXIT
- Entry zone
- Stop loss
- Target 1 / Target 2
- Reward-to-risk ratio
- Position-quality score
- Priority rank
- Explainable trade rationale
- SQLite history
- Excel sheet: AQSD Decision Engine

Commands
--------
python aqsd_decision_engine.py --run
python aqsd_decision_engine.py --status
python aqsd_decision_engine.py --report
python aqsd_decision_engine.py --symbol RELIANCE
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# SETTINGS
# ============================================================

MINIMUM_COMPLETENESS = 45.0
MINIMUM_CONFIDENCE = 40.0
ATR_STOP_MULTIPLIER = 1.5
TARGET_ONE_R_MULTIPLE = 1.5
TARGET_TWO_R_MULTIPLE = 2.5


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS aqsd_trade_decisions (
    decision_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    symbol_id INTEGER NOT NULL,
    nse_symbol TEXT NOT NULL,
    sector TEXT,
    master_score REAL,
    confidence_percent REAL,
    completeness_percent REAL,
    risk_level TEXT,
    directional_bias TEXT,
    action TEXT,
    entry_quality TEXT,
    last_price REAL,
    entry_low REAL,
    entry_high REAL,
    stop_loss REAL,
    target_1 REAL,
    target_2 REAL,
    reward_risk_1 REAL,
    reward_risk_2 REAL,
    priority_score REAL,
    priority_rank INTEGER,
    rationale TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(trade_date, symbol_id),
    FOREIGN KEY(symbol_id)
        REFERENCES symbols(symbol_id)
        ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_trade_decisions_date
ON aqsd_trade_decisions(trade_date);

CREATE INDEX IF NOT EXISTS idx_trade_decisions_rank
ON aqsd_trade_decisions(trade_date, priority_rank);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# HELPERS
# ============================================================

def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None

        return float(value)

    except (TypeError, ValueError):
        return None


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def round_price(value: float | None) -> float | None:
    return round(value, 2) if value is not None else None


# ============================================================
# LOADERS
# ============================================================

def latest_master() -> pd.DataFrame:
    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM unified_master_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM unified_master_intelligence
            )
            ORDER BY master_score DESC
            """,
            connection,
        )


def latest_structure() -> pd.DataFrame:
    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT
                symbol_id,
                trade_date,
                close_price,
                atr_14,
                support_level,
                resistance_level,
                weekly_pivot,
                monthly_pivot,
                cpr_low,
                cpr_high,
                bos_signal,
                choch_signal,
                market_structure,
                trend_strength,
                structure_score
            FROM price_structure_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
            """,
            connection,
        )


def latest_sector_rotation() -> pd.DataFrame:
    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT
                sector,
                sector_rotation_score,
                rotation_state,
                trend_state
            FROM sector_rotation_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            """,
            connection,
        )


# ============================================================
# DECISION LOGIC
# ============================================================

def determine_action(
    master_score: float,
    confidence: float,
    completeness: float,
    bias: str,
    risk: str,
) -> str:
    if completeness < MINIMUM_COMPLETENESS:
        return "INSUFFICIENT DATA"

    if confidence < MINIMUM_CONFIDENCE:
        return "WATCH"

    if master_score >= 85 and "BULLISH" in bias and risk != "HIGH":
        return "STRONG BUY"

    if master_score >= 72 and "BULLISH" in bias:
        return "BUY"

    if master_score >= 60 and "BULLISH" in bias:
        return "BUY ON DIP"

    if master_score <= 25 and "BEARISH" in bias:
        return "EXIT / AVOID"

    if master_score <= 40 and "BEARISH" in bias:
        return "AVOID"

    return "WATCH"


def build_trade_levels(
    last_price: float,
    atr: float | None,
    support: float | None,
    resistance: float | None,
    cpr_low: float | None,
    cpr_high: float | None,
    action: str,
) -> dict:
    atr = atr if atr and atr > 0 else last_price * 0.02

    if action in {"STRONG BUY", "BUY"}:
        entry_low = max(
            support or (last_price - atr * 0.5),
            cpr_low or (last_price - atr * 0.5),
        )
        entry_high = last_price
        stop_loss = min(
            support or (last_price - atr),
            last_price - ATR_STOP_MULTIPLIER * atr,
        )

    elif action == "BUY ON DIP":
        entry_low = max(
            support or (last_price - atr),
            cpr_low or (last_price - atr),
        )
        entry_high = min(
            last_price,
            cpr_high or last_price,
        )
        stop_loss = min(
            support or (last_price - atr),
            entry_low - ATR_STOP_MULTIPLIER * atr,
        )

    else:
        return {
            "entry_low": None,
            "entry_high": None,
            "stop_loss": None,
            "target_1": None,
            "target_2": None,
            "reward_risk_1": None,
            "reward_risk_2": None,
        }

    entry_mid = (entry_low + entry_high) / 2
    risk_per_share = max(0.01, entry_mid - stop_loss)

    target_1 = entry_mid + risk_per_share * TARGET_ONE_R_MULTIPLE
    target_2 = entry_mid + risk_per_share * TARGET_TWO_R_MULTIPLE

    if resistance and resistance > entry_mid:
        target_1 = max(target_1, resistance)

    rr1 = (target_1 - entry_mid) / risk_per_share
    rr2 = (target_2 - entry_mid) / risk_per_share

    return {
        "entry_low": round_price(entry_low),
        "entry_high": round_price(entry_high),
        "stop_loss": round_price(stop_loss),
        "target_1": round_price(target_1),
        "target_2": round_price(target_2),
        "reward_risk_1": round(rr1, 2),
        "reward_risk_2": round(rr2, 2),
    }


def priority_score(
    master_score: float,
    confidence: float,
    completeness: float,
    sector_score: float | None,
    rr1: float | None,
    risk_level: str,
) -> float:
    score = (
        master_score * 0.45
        + confidence * 0.25
        + completeness * 0.10
        + (sector_score if sector_score is not None else 50) * 0.10
        + min((rr1 or 0) * 20, 100) * 0.10
    )

    if risk_level == "HIGH":
        score -= 10
    elif risk_level == "LOW":
        score += 5

    return round(clamp(score, 0, 100), 2)


def build_decisions() -> list[dict]:
    master = latest_master()
    structure = latest_structure()
    sectors = latest_sector_rotation()

    if master.empty:
        raise RuntimeError(
            "No Unified Master Intelligence found. "
            "Run aqsd_unified_master_intelligence.py --run first."
        )

    merged = master.merge(
        structure,
        on="symbol_id",
        how="left",
        suffixes=("", "_structure"),
    )

    if not sectors.empty:
        sectors = sectors.copy()
        sectors["sector_key"] = sectors["sector"].astype(str).str.upper()

    decisions = []

    for _, row in merged.iterrows():
        master_score = float(row["master_score"])
        confidence = float(row["confidence_percent"])
        completeness = float(row["data_completeness_percent"])
        bias = str(row["directional_bias"] or "")
        risk = str(row["risk_level"] or "HIGH")
        action = determine_action(
            master_score,
            confidence,
            completeness,
            bias,
            risk,
        )

        last_price = safe_float(row.get("close_price"))
        atr = safe_float(row.get("atr_14"))
        support = safe_float(row.get("support_level"))
        resistance = safe_float(row.get("resistance_level"))
        cpr_low = safe_float(row.get("cpr_low"))
        cpr_high = safe_float(row.get("cpr_high"))

        levels = (
            build_trade_levels(
                last_price,
                atr,
                support,
                resistance,
                cpr_low,
                cpr_high,
                action,
            )
            if last_price is not None
            else {
                "entry_low": None,
                "entry_high": None,
                "stop_loss": None,
                "target_1": None,
                "target_2": None,
                "reward_risk_1": None,
                "reward_risk_2": None,
            }
        )

        sector = str(row["sector"] or "Unmapped")
        sector_score = safe_float(row.get("sector_rotation_score"))
        rotation_state = ""
        sector_trend = ""

        if not sectors.empty:
            selected = sectors[
                sectors["sector_key"] == sector.upper()
            ]

            if not selected.empty:
                sector_score = safe_float(
                    selected["sector_rotation_score"].iloc[0]
                )
                rotation_state = str(
                    selected["rotation_state"].iloc[0] or ""
                )
                sector_trend = str(
                    selected["trend_state"].iloc[0] or ""
                )

        priority = priority_score(
            master_score,
            confidence,
            completeness,
            sector_score,
            levels["reward_risk_1"],
            risk,
        )

        rationale_parts = [
            f"Master {master_score:.1f}",
            f"Confidence {confidence:.1f}%",
            f"Completeness {completeness:.1f}%",
            f"Bias {bias}",
            f"Risk {risk}",
        ]

        if sector_score is not None:
            rationale_parts.append(
                f"Sector {sector_score:.1f}"
            )

        if rotation_state:
            rationale_parts.append(
                f"Rotation {rotation_state}"
            )

        if sector_trend:
            rationale_parts.append(
                f"Sector trend {sector_trend}"
            )

        market_structure = str(
            row.get("market_structure") or ""
        )
        bos = str(row.get("bos_signal") or "")
        choch = str(row.get("choch_signal") or "")
        trend_strength = str(
            row.get("trend_strength") or ""
        )

        if market_structure:
            rationale_parts.append(market_structure)

        if bos and bos != "NONE":
            rationale_parts.append(bos)

        if choch and choch != "NONE":
            rationale_parts.append(choch)

        if trend_strength:
            rationale_parts.append(
                f"Trend {trend_strength}"
            )

        decisions.append(
            {
                "trade_date": str(row["trade_date"]),
                "symbol_id": int(row["symbol_id"]),
                "nse_symbol": str(row["nse_symbol"]),
                "sector": sector,
                "master_score": master_score,
                "confidence_percent": confidence,
                "completeness_percent": completeness,
                "risk_level": risk,
                "directional_bias": bias,
                "action": action,
                "entry_quality": str(
                    row["entry_quality"] or ""
                ),
                "last_price": round_price(last_price),
                **levels,
                "priority_score": priority,
                "priority_rank": 0,
                "rationale": " | ".join(rationale_parts),
            }
        )

    action_priority = {
        "STRONG BUY": 0,
        "BUY": 1,
        "BUY ON DIP": 2,
        "WATCH": 3,
        "AVOID": 4,
        "EXIT / AVOID": 5,
        "INSUFFICIENT DATA": 6,
    }

    decisions = sorted(
        decisions,
        key=lambda item: (
            action_priority.get(item["action"], 99),
            -item["priority_score"],
            -item["master_score"],
        ),
    )

    for rank, item in enumerate(decisions, start=1):
        item["priority_rank"] = rank

    return decisions


def save_decisions(decisions: list[dict]) -> None:
    with connect() as connection:
        for item in decisions:
            connection.execute(
                """
                INSERT INTO aqsd_trade_decisions(
                    trade_date,
                    symbol_id,
                    nse_symbol,
                    sector,
                    master_score,
                    confidence_percent,
                    completeness_percent,
                    risk_level,
                    directional_bias,
                    action,
                    entry_quality,
                    last_price,
                    entry_low,
                    entry_high,
                    stop_loss,
                    target_1,
                    target_2,
                    reward_risk_1,
                    reward_risk_2,
                    priority_score,
                    priority_rank,
                    rationale,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(trade_date, symbol_id)
                DO UPDATE SET
                    sector = excluded.sector,
                    master_score = excluded.master_score,
                    confidence_percent = excluded.confidence_percent,
                    completeness_percent = excluded.completeness_percent,
                    risk_level = excluded.risk_level,
                    directional_bias = excluded.directional_bias,
                    action = excluded.action,
                    entry_quality = excluded.entry_quality,
                    last_price = excluded.last_price,
                    entry_low = excluded.entry_low,
                    entry_high = excluded.entry_high,
                    stop_loss = excluded.stop_loss,
                    target_1 = excluded.target_1,
                    target_2 = excluded.target_2,
                    reward_risk_1 = excluded.reward_risk_1,
                    reward_risk_2 = excluded.reward_risk_2,
                    priority_score = excluded.priority_score,
                    priority_rank = excluded.priority_rank,
                    rationale = excluded.rationale,
                    created_at = excluded.created_at
                """,
                (
                    item["trade_date"],
                    item["symbol_id"],
                    item["nse_symbol"],
                    item["sector"],
                    item["master_score"],
                    item["confidence_percent"],
                    item["completeness_percent"],
                    item["risk_level"],
                    item["directional_bias"],
                    item["action"],
                    item["entry_quality"],
                    item["last_price"],
                    item["entry_low"],
                    item["entry_high"],
                    item["stop_loss"],
                    item["target_1"],
                    item["target_2"],
                    item["reward_risk_1"],
                    item["reward_risk_2"],
                    item["priority_score"],
                    item["priority_rank"],
                    item["rationale"],
                    datetime.now().isoformat(timespec="seconds"),
                ),
            )

        connection.commit()


def run_engine() -> list[dict]:
    setup_schema()

    run_id = start_run(
        "aqsd_decision_engine",
        "Building actionable trade decisions",
    )

    try:
        decisions = build_decisions()
        save_decisions(decisions)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=len(decisions),
            errors_count=0,
            message=f"Decisions created={len(decisions)}",
        )

        return decisions

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


# ============================================================
# REPORTING
# ============================================================

def latest_decisions() -> pd.DataFrame:
    setup_schema()

    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM aqsd_trade_decisions
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
            ORDER BY priority_rank
            """,
            connection,
        )


def write_report(
    decisions: list[dict] | None = None,
) -> None:
    if decisions is None:
        decisions = latest_decisions().to_dict("records")

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "AQSD Decision Engine" in wb.sheetnames:
        del wb["AQSD Decision Engine"]

    ws = wb.create_sheet("AQSD Decision Engine", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:V2")
    ws["A1"] = "AQSD PROFESSIONAL - TRADE DECISION & WATCHLIST ENGINE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    actionable = sum(
        item["action"] in {
            "STRONG BUY",
            "BUY",
            "BUY ON DIP",
        }
        for item in decisions
    )

    ws["A4"] = "Stocks Ranked"
    ws["B4"] = len(decisions)
    ws["D4"] = "Actionable Longs"
    ws["E4"] = actionable
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill("solid", fgColor=BLUE)

    headers = [
        "Priority",
        "Symbol",
        "Sector",
        "Action",
        "Master Score",
        "Confidence %",
        "Completeness %",
        "Risk",
        "Entry Quality",
        "Last Price",
        "Entry Low",
        "Entry High",
        "Stop Loss",
        "Target 1",
        "Target 2",
        "R:R 1",
        "R:R 2",
        "Priority Score",
        "Directional Bias",
        "Rationale",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, item in enumerate(decisions, start=8):
        values = [
            item["priority_rank"],
            item["nse_symbol"],
            item["sector"],
            item["action"],
            item["master_score"],
            item["confidence_percent"],
            item["completeness_percent"],
            item["risk_level"],
            item["entry_quality"],
            item["last_price"],
            item["entry_low"],
            item["entry_high"],
            item["stop_loss"],
            item["target_1"],
            item["target_2"],
            item["reward_risk_1"],
            item["reward_risk_2"],
            item["priority_score"],
            item["directional_bias"],
            item["rationale"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        action = item["action"]

        ws.cell(row_no, 4).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if action in {
                    "STRONG BUY",
                    "BUY",
                    "BUY ON DIP",
                }
                else RED
                if action in {
                    "AVOID",
                    "EXIT / AVOID",
                }
                else YELLOW
            ),
        )
        ws.cell(row_no, 4).font = Font(bold=True)

        risk = item["risk_level"]

        ws.cell(row_no, 8).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if risk == "LOW"
                else RED
                if risk == "HIGH"
                else YELLOW
            ),
        )

        score = float(item["master_score"])

        ws.cell(row_no, 5).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )

    widths = {
        "A": 10,
        "B": 16,
        "C": 20,
        "D": 18,
        "E": 14,
        "F": 14,
        "G": 16,
        "H": 12,
        "I": 16,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 10,
        "Q": 10,
        "R": 14,
        "S": 18,
        "T": 90,
        "U": 14,
        "V": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


def show_symbol(symbol: str) -> None:
    text = symbol.strip().upper()

    if text.endswith(".NS"):
        text = text[:-3]

    frame = latest_decisions()

    if frame.empty:
        print("No trade decisions found.")
        return

    selected = frame[
        frame["nse_symbol"].str.upper() == text
    ]

    if selected.empty:
        print(f"Symbol not found: {text}")
        return

    row = selected.iloc[0]

    print("\nAQSD TRADE DECISION")
    print("=" * 72)

    for field in [
        "priority_rank",
        "nse_symbol",
        "sector",
        "action",
        "master_score",
        "confidence_percent",
        "completeness_percent",
        "risk_level",
        "entry_quality",
        "last_price",
        "entry_low",
        "entry_high",
        "stop_loss",
        "target_1",
        "target_2",
        "reward_risk_1",
        "reward_risk_2",
        "priority_score",
        "directional_bias",
        "rationale",
    ]:
        print(f"{field:<26}{row[field]}")

    print("=" * 72)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT symbol_id) AS symbols,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM aqsd_trade_decisions
            """
        ).fetchone()

    print("\nAQSD DECISION ENGINE STATUS")
    print("=" * 72)
    print(f"Stored decisions:  {row['total'] or 0}")
    print(f"Symbols covered:   {row['symbols'] or 0}")
    print(f"First date:        {row['first_date'] or 'No data'}")
    print(f"Latest date:       {row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Trade Decision and Watchlist Engine."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Build daily trade decisions.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild Excel report from stored data.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show decision-engine status.",
    )

    parser.add_argument(
        "--symbol",
        metavar="SYMBOL",
        help="Show one symbol's trade decision.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.symbol:
        show_symbol(args.symbol)
        return

    if args.run:
        decisions = run_engine()
        write_report(decisions)

        print("\nAQSD TRADE DECISION ENGINE")
        print("=" * 72)
        print(f"Stocks ranked: {len(decisions)}")

        actionable = [
            item
            for item in decisions
            if item["action"] in {
                "STRONG BUY",
                "BUY",
                "BUY ON DIP",
            }
        ]

        print(f"Actionable longs: {len(actionable)}")

        if decisions:
            top = decisions[0]
            print(
                f"Top priority: "
                f"{top['nse_symbol']} | "
                f"{top['action']} | "
                f"{top['priority_score']}"
            )

        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report()
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
