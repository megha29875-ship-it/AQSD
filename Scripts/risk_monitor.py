
"""
AQSD Professional
Module: Risk Monitor
Version: 1.0

Creates a Risk Monitor sheet highlighting:
- Position risk
- Portfolio risk
- Oversized positions
- Highest risk trades
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
FILE = BASE / "Output" / "Dashboard.xlsx"

wb = load_workbook(FILE)

if "Portfolio" not in wb.sheetnames:
    raise RuntimeError("Portfolio sheet not found.")

pws = wb["Portfolio"]

if "Risk Monitor" in wb.sheetnames:
    del wb["Risk Monitor"]

ws = wb.create_sheet("Risk Monitor")

ws["A1"] = "AQSD PROFESSIONAL - RISK MONITOR"
ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="17365D")

headers = {}
for c in pws[12]:
    if c.value:
        headers[str(c.value)] = c.column

required = ["Symbol","Capital Used","Risk Amount","Status"]
for r in required:
    if r not in headers:
        raise RuntimeError(f"Missing column: {r}")

ws.append([])
ws.append(["Symbol","Capital Used","Risk Amount","Risk %","Status","Alert"])

capital = 0.0
rows = []

for row in range(13, pws.max_row+1):
    if str(pws.cell(row, headers["Status"]).value).upper() != "OPEN":
        continue
    sym = pws.cell(row, headers["Symbol"]).value
    cap = float(pws.cell(row, headers["Capital Used"]).value or 0)
    risk = float(pws.cell(row, headers["Risk Amount"]).value or 0)
    capital += cap
    rows.append((sym, cap, risk))

total_risk = sum(r[2] for r in rows)

excel_row = 4
for sym, cap, risk in sorted(rows, key=lambda x:x[2], reverse=True):
    pct = (risk/cap*100) if cap else 0
    alert = "HIGH" if pct > 1 else "NORMAL"
    ws.cell(excel_row,1).value = sym
    ws.cell(excel_row,2).value = cap
    ws.cell(excel_row,3).value = risk
    ws.cell(excel_row,4).value = pct
    ws.cell(excel_row,5).value = "OPEN"
    ws.cell(excel_row,6).value = alert
    ws.cell(excel_row,2).number_format='₹#,##0.00'
    ws.cell(excel_row,3).number_format='₹#,##0.00'
    ws.cell(excel_row,4).number_format='0.00"%"'
    if alert=="HIGH":
        ws.cell(excel_row,6).fill=PatternFill("solid", fgColor="FFC7CE")
    else:
        ws.cell(excel_row,6).fill=PatternFill("solid", fgColor="C6EFCE")
    excel_row += 1

ws["H2"]="Portfolio Capital"
ws["I2"]=capital
ws["H3"]="Total Portfolio Risk"
ws["I3"]=total_risk
ws["H4"]="Portfolio Risk %"
ws["I4"]=(total_risk/capital*100) if capital else 0
ws["I2"].number_format='₹#,##0.00'
ws["I3"].number_format='₹#,##0.00'
ws["I4"].number_format='0.00"%"'

wb.save(FILE)
print("Risk Monitor created.")
print(FILE)
