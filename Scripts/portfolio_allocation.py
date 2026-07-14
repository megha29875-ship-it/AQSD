
"""
AQSD Professional
Module: Portfolio Allocation Dashboard
Version: 1.0

Creates a Portfolio Allocation sheet using current open positions.

Features
--------
- Capital allocation by position
- Risk allocation by position
- Sector allocation using SECTOR_MAP
- Largest position
- Largest sector
- Diversification score
- Concentration warnings
- Capital allocation pie chart
- Risk allocation bar chart
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# SETTINGS
# ============================================================

MAX_POSITION_PERCENT = 25.0
MAX_SECTOR_PERCENT = 40.0

SECTOR_MAP = {
    "RELIANCE.NS": "Energy",
    "ONGC.NS": "Energy",
    "IOC.NS": "Energy",
    "BPCL.NS": "Energy",
    "BIOCON.NS": "Pharma",
    "DIVISLAB.NS": "Pharma",
    "SUNPHARMA.NS": "Pharma",
    "CIPLA.NS": "Pharma",
    "DRREDDY.NS": "Pharma",
    "ICICIBANK.NS": "Banking",
    "HDFCBANK.NS": "Banking",
    "SBIN.NS": "Banking",
    "AXISBANK.NS": "Banking",
    "KOTAKBANK.NS": "Banking",
    "TCS.NS": "IT",
    "INFY.NS": "IT",
    "HCLTECH.NS": "IT",
    "WIPRO.NS": "IT",
    "LTIM.NS": "IT",
    "LT.NS": "Infrastructure",
    "SIEMENS.NS": "Capital Goods",
    "ABB.NS": "Capital Goods",
    "HAL.NS": "Defence",
    "BEL.NS": "Defence",
    "TATAMOTORS.NS": "Auto",
    "M&M.NS": "Auto",
    "MARUTI.NS": "Auto",
    "BAJAJ-AUTO.NS": "Auto",
    "HINDUNILVR.NS": "FMCG",
    "ITC.NS": "FMCG",
    "NESTLEIND.NS": "FMCG",
    "TATASTEEL.NS": "Metals",
    "JSWSTEEL.NS": "Metals",
    "HINDALCO.NS": "Metals",
}


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_open_positions(wb) -> list[dict]:
    if "Portfolio" not in wb.sheetnames:
        raise RuntimeError(
            "Portfolio sheet not found. "
            "Run portfolio_manager.py first."
        )

    ws = wb["Portfolio"]
    headers = header_map(ws, 12)

    required = [
        "Symbol",
        "Capital Used",
        "Risk Amount",
        "P/L",
        "Status",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            "Missing Portfolio columns: "
            + ", ".join(missing)
        )

    positions: list[dict] = []

    for row in range(13, ws.max_row + 1):
        symbol = str(
            ws.cell(row, headers["Symbol"]).value or ""
        ).strip().upper()

        if not symbol:
            continue

        status = str(
            ws.cell(row, headers["Status"]).value or ""
        ).strip().upper()

        if status != "OPEN":
            continue

        positions.append(
            {
                "Symbol": symbol,
                "Sector": SECTOR_MAP.get(symbol, "Others"),
                "Capital Used": safe_float(
                    ws.cell(
                        row,
                        headers["Capital Used"],
                    ).value
                ),
                "Risk Amount": safe_float(
                    ws.cell(
                        row,
                        headers["Risk Amount"],
                    ).value
                ),
                "P/L": safe_float(
                    ws.cell(row, headers["P/L"]).value
                ),
            }
        )

    return positions


def diversification_score(
    position_count: int,
    largest_position_percent: float,
    largest_sector_percent: float,
) -> int:
    score = 100

    if position_count < 3:
        score -= 30
    elif position_count < 5:
        score -= 15

    if largest_position_percent > 50:
        score -= 35
    elif largest_position_percent > 35:
        score -= 20
    elif largest_position_percent > 25:
        score -= 10

    if largest_sector_percent > 60:
        score -= 25
    elif largest_sector_percent > 45:
        score -= 15
    elif largest_sector_percent > 40:
        score -= 8

    return max(0, min(100, score))


def concentration_status(
    largest_position_percent: float,
    largest_sector_percent: float,
) -> str:
    if (
        largest_position_percent > MAX_POSITION_PERCENT
        or largest_sector_percent > MAX_SECTOR_PERCENT
    ):
        return "HIGH CONCENTRATION"

    return "ACCEPTABLE"


# ============================================================
# SHEET CREATION
# ============================================================

def create_allocation_sheet(
    wb,
    positions: list[dict],
) -> None:
    if "Portfolio Allocation" in wb.sheetnames:
        del wb["Portfolio Allocation"]

    ws = wb.create_sheet("Portfolio Allocation")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A15"

    ws.merge_cells("A1:H2")
    ws["A1"] = "AQSD PROFESSIONAL - PORTFOLIO ALLOCATION"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    total_capital = sum(
        position["Capital Used"]
        for position in positions
    )

    total_risk = sum(
        position["Risk Amount"]
        for position in positions
    )

    sector_totals = defaultdict(float)

    for position in positions:
        sector_totals[position["Sector"]] += position["Capital Used"]

    largest_position = max(
        positions,
        key=lambda item: item["Capital Used"],
        default=None,
    )

    largest_sector = max(
        sector_totals.items(),
        key=lambda item: item[1],
        default=("", 0.0),
    )

    largest_position_percent = (
        largest_position["Capital Used"] / total_capital * 100
        if largest_position and total_capital
        else 0.0
    )

    largest_sector_percent = (
        largest_sector[1] / total_capital * 100
        if total_capital
        else 0.0
    )

    diversification = diversification_score(
        len(positions),
        largest_position_percent,
        largest_sector_percent,
    )

    concentration = concentration_status(
        largest_position_percent,
        largest_sector_percent,
    )

    summary = [
        ("Last Updated", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("Open Positions", len(positions)),
        ("Total Capital Used", total_capital),
        ("Total Risk", total_risk),
        (
            "Largest Position",
            largest_position["Symbol"]
            if largest_position
            else "",
        ),
        ("Largest Position %", largest_position_percent),
        ("Largest Sector", largest_sector[0]),
        ("Largest Sector %", largest_sector_percent),
        ("Diversification Score", diversification),
        ("Concentration Status", concentration),
    ]

    for row_no, (label, value) in enumerate(summary, start=4):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value

        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )
        ws[f"A{row_no}"].border = Border(bottom=THIN)
        ws[f"B{row_no}"].border = Border(bottom=THIN)

    ws["B6"].number_format = '₹#,##0.00'
    ws["B7"].number_format = '₹#,##0.00'
    ws["B9"].number_format = '0.00"%"'
    ws["B11"].number_format = '0.00"%"'
    ws["B12"].number_format = "0"

    ws["B13"].fill = PatternFill(
        "solid",
        fgColor=(
            GREEN
            if concentration == "ACCEPTABLE"
            else RED
        ),
    )
    ws["B13"].font = Font(bold=True)

    # --------------------------------------------------------
    # Position allocation table
    # --------------------------------------------------------
    position_headers = [
        "Rank",
        "Symbol",
        "Sector",
        "Capital Used",
        "Capital %",
        "Risk Amount",
        "Risk %",
        "P/L",
    ]

    header_row = 15

    for col, heading in enumerate(position_headers, start=1):
        cell = ws.cell(header_row, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    sorted_positions = sorted(
        positions,
        key=lambda item: item["Capital Used"],
        reverse=True,
    )

    for index, item in enumerate(sorted_positions, start=1):
        row = header_row + index

        capital_percent = (
            item["Capital Used"] / total_capital * 100
            if total_capital
            else 0.0
        )

        risk_percent = (
            item["Risk Amount"] / total_risk * 100
            if total_risk
            else 0.0
        )

        values = [
            index,
            item["Symbol"],
            item["Sector"],
            item["Capital Used"],
            capital_percent,
            item["Risk Amount"],
            risk_percent,
            item["P/L"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = Border(bottom=THIN)

        ws.cell(row, 4).number_format = '₹#,##0.00'
        ws.cell(row, 5).number_format = '0.00"%"'
        ws.cell(row, 6).number_format = '₹#,##0.00'
        ws.cell(row, 7).number_format = '0.00"%"'
        ws.cell(row, 8).number_format = '₹#,##0.00'

        ws.cell(row, 5).fill = PatternFill(
            "solid",
            fgColor=(
                RED
                if capital_percent > MAX_POSITION_PERCENT
                else GREEN
            ),
        )

        ws.cell(row, 8).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if item["P/L"] >= 0
                else RED
            ),
        )

    # --------------------------------------------------------
    # Sector allocation table
    # --------------------------------------------------------
    ws["J15"] = "Sector"
    ws["K15"] = "Capital"
    ws["L15"] = "Allocation %"

    for ref in ("J15", "K15", "L15"):
        ws[ref].font = Font(bold=True, color=WHITE)
        ws[ref].fill = PatternFill("solid", fgColor=NAVY)

    for index, (sector, capital) in enumerate(
        sorted(
            sector_totals.items(),
            key=lambda item: item[1],
            reverse=True,
        ),
        start=16,
    ):
        allocation_percent = (
            capital / total_capital * 100
            if total_capital
            else 0.0
        )

        ws.cell(index, 10, sector)
        ws.cell(index, 11, capital)
        ws.cell(index, 12, allocation_percent)

        ws.cell(index, 11).number_format = '₹#,##0.00'
        ws.cell(index, 12).number_format = '0.00"%"'

        ws.cell(index, 12).fill = PatternFill(
            "solid",
            fgColor=(
                RED
                if allocation_percent > MAX_SECTOR_PERCENT
                else GREEN
            ),
        )

    # --------------------------------------------------------
    # Charts
    # --------------------------------------------------------
    if sorted_positions:
        last_position_row = header_row + len(sorted_positions)

        pie = PieChart()
        pie.title = "Capital Allocation by Position"

        pie.add_data(
            Reference(
                ws,
                min_col=4,
                min_row=15,
                max_row=last_position_row,
            ),
            titles_from_data=True,
        )

        pie.set_categories(
            Reference(
                ws,
                min_col=2,
                min_row=16,
                max_row=last_position_row,
            )
        )

        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, "J3")

        risk_chart = BarChart()
        risk_chart.title = "Risk Allocation by Position"
        risk_chart.y_axis.title = "Risk Amount"
        risk_chart.x_axis.title = "Symbol"

        risk_chart.add_data(
            Reference(
                ws,
                min_col=6,
                min_row=15,
                max_row=last_position_row,
            ),
            titles_from_data=True,
        )

        risk_chart.set_categories(
            Reference(
                ws,
                min_col=2,
                min_row=16,
                max_row=last_position_row,
            )
        )

        risk_chart.height = 8
        risk_chart.width = 12
        ws.add_chart(risk_chart, "J25")

    widths = {
        "A": 24,
        "B": 18,
        "C": 18,
        "D": 15,
        "E": 12,
        "F": 14,
        "G": 12,
        "H": 14,
        "J": 18,
        "K": 15,
        "L": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)
    positions = read_open_positions(wb)

    create_allocation_sheet(
        wb,
        positions,
    )

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Portfolio Allocation created successfully.")
    print(f"Open positions analysed: {len(positions)}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
