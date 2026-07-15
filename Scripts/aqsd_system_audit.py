"""
AQSD Professional
Module: System Audit
Version: 1.0

Purpose
-------
Performs a complete health audit of the AQSD installation.

Commands
--------
python aqsd_system_audit.py --run
python aqsd_system_audit.py --quick
python aqsd_system_audit.py --json
python aqsd_system_audit.py --status
"""

from __future__ import annotations

import argparse
import json
import shutil
import sqlite3
import subprocess
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPTS_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPTS_DIR.parent
DATA_DIR = BASE_DIR / "Data"
OUTPUT_DIR = BASE_DIR / "Output"
LOG_DIR = BASE_DIR / "Logs"
DATABASE = DATA_DIR / "aqsd_core.db"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"
ORCHESTRATOR_LOG = LOG_DIR / "aqsd_orchestrator_runs.jsonl"
TEXT_REPORT = OUTPUT_DIR / "AQSD_System_Audit.txt"
JSON_REPORT = OUTPUT_DIR / "AQSD_System_Audit.json"

REQUIRED_FOLDERS = [BASE_DIR, SCRIPTS_DIR, DATA_DIR, OUTPUT_DIR]

REQUIRED_SCRIPTS = [
    "aqsd_database.py",
    "aqsd_symbol_master.py",
    "aqsd_price_cache.py",
    "aqsd_incremental_updater.py",
    "aqsd_symbol_resolver.py",
    "aqsd_datahub.py",
    "aqsd_database_optimizer.py",
    "aqsd_intelligence_recorder.py",
    "aqsd_global_intelligence.py",
    "aqsd_news_intelligence.py",
    "aqsd_futures_intelligence.py",
    "aqsd_options_intelligence.py",
    "aqsd_macro_intelligence.py",
    "aqsd_price_structure.py",
    "aqsd_sector_rotation.py",
    "aqsd_relative_strength.py",
    "aqsd_market_breadth.py",
    "aqsd_unified_master_intelligence.py",
    "aqsd_decision_engine.py",
    "aqsd_portfolio_allocator.py",
    "aqsd_copilot.py",
    "aqsd_daily_orchestrator.py",
]

REQUIRED_TABLES = [
    "symbols",
    "daily_prices",
    "intelligence_scores",
    "system_runs",
    "settings",
]

OPTIONAL_TABLES = [
    "symbol_aliases",
    "symbol_validation_log",
    "global_markets",
    "commodities",
    "news_events",
    "macro_policy_events",
    "futures_oi",
    "options_chain",
    "options_intelligence",
    "price_structure_intelligence",
    "sector_rotation_intelligence",
    "relative_strength",
    "market_breadth_intelligence",
    "unified_master_intelligence",
    "aqsd_trade_decisions",
    "portfolio_allocation",
]

EXPECTED_SHEETS = [
    "Price Structure Intelligence",
    "Sector Rotation Intelligence",
    "Market Breadth Intelligence",
    "Global Intelligence",
    "Unified Master Intelligence",
    "AQSD Decision Engine",
]

BACKUP_BATCH_CANDIDATES = [
    BASE_DIR / "AQSD_BACKUP.bat",
    BASE_DIR / "backup_aqsd.bat",
    BASE_DIR / "robocopy_backup.bat",
    BASE_DIR / "Backup_AQSD.bat",
]

GIT_BATCH_CANDIDATES = [
    BASE_DIR / "AQSD_GIT.bat",
    BASE_DIR / "git_aqsd.bat",
    BASE_DIR / "git_update.bat",
    BASE_DIR / "Git_AQSD.bat",
]


@dataclass
class AuditItem:
    section: str
    check: str
    status: str
    message: str
    value: Any = None


def add_result(results, section, check, status, message, value=None):
    results.append(AuditItem(section, check, status, message, value))


def connect():
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    return con


def table_exists(con, table_name):
    return con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone() is not None


def latest_value(con, table, column):
    try:
        row = con.execute(f"SELECT MAX({column}) AS value FROM {table}").fetchone()
        return row["value"] if row else None
    except Exception:
        return None


def days_old(value):
    if not value:
        return None
    for parser in (
        lambda x: datetime.fromisoformat(x).date(),
        lambda x: datetime.strptime(x, "%Y-%m-%d").date(),
    ):
        try:
            parsed = parser(value)
            return (date.today() - parsed).days
        except Exception:
            pass
    return None


def human_size(size_bytes):
    size = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size < 1024:
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size:.2f} PB"


def audit_folders(results):
    for folder in REQUIRED_FOLDERS:
        exists = folder.exists()
        add_result(results, "Folders", str(folder), "PASS" if exists else "FAIL",
                   "Folder exists." if exists else "Required folder is missing.")


def audit_scripts(results):
    for script in REQUIRED_SCRIPTS:
        path = SCRIPTS_DIR / script
        exists = path.exists()
        add_result(results, "Scripts", script, "PASS" if exists else "FAIL",
                   "Script found." if exists else f"Missing script: {path}")


def audit_database(results):
    if not DATABASE.exists():
        add_result(results, "Database", "Database file", "FAIL", f"Database not found: {DATABASE}")
        return

    add_result(results, "Database", "Database file", "PASS", "Database file exists.", human_size(DATABASE.stat().st_size))

    try:
        with connect() as con:
            integrity = con.execute("PRAGMA integrity_check").fetchone()[0]
            add_result(results, "Database", "SQLite integrity", "PASS" if integrity == "ok" else "FAIL",
                       f"Integrity result: {integrity}", integrity)

            for table in REQUIRED_TABLES:
                exists = table_exists(con, table)
                add_result(results, "Database Tables", table, "PASS" if exists else "FAIL",
                           "Required table exists." if exists else "Required table is missing.")

            for table in OPTIONAL_TABLES:
                exists = table_exists(con, table)
                add_result(results, "Optional Tables", table, "PASS" if exists else "WARN",
                           "Optional table exists." if exists else "Optional table not created yet.")
    except Exception as error:
        add_result(results, "Database", "Database connection", "FAIL", str(error))


def audit_symbols_and_prices(results, stale_days):
    if not DATABASE.exists():
        return
    try:
        with connect() as con:
            if not table_exists(con, "symbols"):
                return
            row = con.execute(
                """
                SELECT COUNT(*) total,
                       SUM(CASE WHEN active=1 THEN 1 ELSE 0 END) active,
                       SUM(CASE WHEN fno_eligible=1 THEN 1 ELSE 0 END) fno
                FROM symbols
                """
            ).fetchone()
            total, active, fno = int(row["total"] or 0), int(row["active"] or 0), int(row["fno"] or 0)
            add_result(results, "Symbols", "Symbol Master", "PASS" if active > 0 else "FAIL",
                       f"Total={total}, Active={active}, F&O={fno}",
                       {"total": total, "active": active, "fno": fno})

            if not table_exists(con, "daily_prices"):
                return
            row = con.execute(
                """
                SELECT COUNT(*) rows_count,
                       COUNT(DISTINCT symbol_id) symbols,
                       MIN(trade_date) first_date,
                       MAX(trade_date) latest_date
                FROM daily_prices
                """
            ).fetchone()
            latest = row["latest_date"]
            age = days_old(latest)
            add_result(results, "Prices", "Price Cache",
                       "PASS" if age is not None and age <= stale_days else "WARN",
                       f"Rows={row['rows_count'] or 0}, Symbols={row['symbols'] or 0}, First={row['first_date'] or '-'}, Latest={latest or '-'}, Age={age if age is not None else '-'} days",
                       {"rows": int(row['rows_count'] or 0), "symbols": int(row['symbols'] or 0), "first_date": row['first_date'], "latest_date": latest, "age_days": age})

            cutoff = (date.today() - timedelta(days=stale_days)).isoformat()
            stale_rows = con.execute(
                """
                SELECT s.nse_symbol, MAX(p.trade_date) latest_date
                FROM symbols s
                LEFT JOIN daily_prices p ON p.symbol_id=s.symbol_id
                WHERE s.active=1
                GROUP BY s.symbol_id, s.nse_symbol
                HAVING latest_date IS NULL OR latest_date < ?
                ORDER BY latest_date, s.nse_symbol
                """,
                (cutoff,),
            ).fetchall()
            stale = [{"symbol": r["nse_symbol"], "latest_date": r["latest_date"]} for r in stale_rows]
            add_result(results, "Prices", "Stale / Missing Symbols", "PASS" if not stale else "WARN",
                       "No stale symbols." if not stale else f"{len(stale)} stale or missing symbols found.", stale[:50])
    except Exception as error:
        add_result(results, "Prices", "Price audit", "FAIL", str(error))


def audit_intelligence(results):
    checks = [
        ("price_structure_intelligence", "trade_date", "Price Structure"),
        ("sector_rotation_intelligence", "trade_date", "Sector Rotation"),
        ("relative_strength", "trade_date", "Relative Strength"),
        ("market_breadth_intelligence", "trade_date", "Market Breadth"),
        ("global_markets", "snapshot_date", "Global Markets"),
        ("unified_master_intelligence", "trade_date", "Unified Master"),
        ("aqsd_trade_decisions", "trade_date", "Decision Engine"),
        ("portfolio_allocation", "trade_date", "Portfolio Allocation"),
    ]
    if not DATABASE.exists():
        return
    with connect() as con:
        for table, date_col, label in checks:
            if not table_exists(con, table):
                add_result(results, "Intelligence", label, "WARN", f"Table {table} does not exist.")
                continue
            count = con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            latest = latest_value(con, table, date_col)
            age = days_old(latest)
            status = "PASS"
            if count == 0 or (age is not None and age > 7):
                status = "WARN"
            add_result(results, "Intelligence", label, status,
                       f"Rows={count}, Latest={latest or '-'}, Age={age if age is not None else '-'} days",
                       {"rows": int(count), "latest": latest, "age_days": age})


def audit_dashboard(results):
    if not DASHBOARD.exists():
        add_result(results, "Dashboard", "Dashboard.xlsx", "FAIL", f"Dashboard not found: {DASHBOARD}")
        return
    add_result(results, "Dashboard", "Dashboard.xlsx", "PASS", "Dashboard workbook exists.", human_size(DASHBOARD.stat().st_size))
    try:
        wb = load_workbook(DASHBOARD, read_only=True, data_only=False)
        try:
            for sheet in EXPECTED_SHEETS:
                exists = sheet in wb.sheetnames
                add_result(results, "Dashboard Sheets", sheet, "PASS" if exists else "WARN",
                           "Sheet exists." if exists else "Expected sheet is missing.")
            add_result(results, "Dashboard", "Workbook sheet count", "PASS",
                       f"Workbook contains {len(wb.sheetnames)} sheets.", wb.sheetnames)
        finally:
            wb.close()
    except Exception as error:
        add_result(results, "Dashboard", "Workbook open test", "FAIL", str(error))


def audit_orchestrator_log(results):
    if not ORCHESTRATOR_LOG.exists():
        add_result(results, "Orchestrator", "Run log", "WARN", "Orchestrator log does not exist yet.")
        return
    try:
        records = []
        for line in ORCHESTRATOR_LOG.read_text(encoding="utf-8").splitlines()[-200:]:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except Exception:
                pass
        failed = [r for r in records if r.get("status") == "FAILED"]
        add_result(results, "Orchestrator", "Run log", "PASS" if records else "WARN",
                   f"Parsed {len(records)} recent records. Failures={len(failed)}.",
                   {"records": len(records), "failures": len(failed), "latest": records[-1] if records else None})
        if failed:
            add_result(results, "Orchestrator", "Recent failures", "WARN",
                       f"{len(failed)} failed steps found in recent log entries.", failed[-20:])
    except Exception as error:
        add_result(results, "Orchestrator", "Run log parse", "FAIL", str(error))


def audit_git(results):
    git = shutil.which("git")
    if not git:
        add_result(results, "Git", "Git executable", "WARN", "Git is not available in PATH.")
        return
    add_result(results, "Git", "Git executable", "PASS", git)
    if not (BASE_DIR / ".git").exists():
        add_result(results, "Git", "Git repository", "WARN", "AQSD folder is not detected as a Git repository.")
        return
    try:
        process = subprocess.run([git, "-C", str(BASE_DIR), "status", "--porcelain"], capture_output=True, text=True, timeout=30, check=False)
        changed = [line for line in process.stdout.splitlines() if line.strip()]
        add_result(results, "Git", "Repository status", "PASS" if not changed else "WARN",
                   "Working tree is clean." if not changed else f"{len(changed)} uncommitted changes found.", changed[:100])
    except Exception as error:
        add_result(results, "Git", "Repository status", "FAIL", str(error))


def audit_batch_files(results):
    backup = [p for p in BACKUP_BATCH_CANDIDATES if p.exists()]
    git_files = [p for p in GIT_BATCH_CANDIDATES if p.exists()]
    add_result(results, "Workflow", "Robocopy backup batch", "PASS" if backup else "WARN",
               f"Found: {backup[0]}" if backup else "No recognised Robocopy backup batch file found.", [str(p) for p in backup])
    add_result(results, "Workflow", "Git batch", "PASS" if git_files else "WARN",
               f"Found: {git_files[0]}" if git_files else "No recognised Git batch file found.", [str(p) for p in git_files])


def audit_disk(results):
    usage = shutil.disk_usage(BASE_DIR)
    free_pct = usage.free / usage.total * 100 if usage.total else 0
    add_result(results, "System", "Disk space", "PASS" if free_pct >= 10 else "WARN",
               f"Free={human_size(usage.free)} ({free_pct:.1f}%), Total={human_size(usage.total)}",
               {"free_bytes": usage.free, "total_bytes": usage.total, "free_percent": round(free_pct, 2)})


def overall_status(results):
    statuses = {r.status for r in results}
    if "FAIL" in statuses:
        return "FAIL"
    if "WARN" in statuses:
        return "WARN"
    return "PASS"


def build_text_report(results):
    lines = [
        "AQSD PROFESSIONAL - SYSTEM AUDIT",
        "=" * 100,
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Base folder: {BASE_DIR}",
        f"Overall status: {overall_status(results)}",
        "=" * 100,
    ]
    current = None
    for item in results:
        if item.section != current:
            current = item.section
            lines.extend(["", current.upper(), "-" * 100])
        lines.append(f"[{item.status:<4}] {item.check:<38} {item.message}")
        if item.value not in (None, "", [], {}):
            value_text = json.dumps(item.value, ensure_ascii=False, default=str)
            if len(value_text) <= 300:
                lines.append(f"       Value: {value_text}")
    summary = {s: sum(r.status == s for r in results) for s in ("PASS", "WARN", "FAIL")}
    lines.extend([
        "", "=" * 100, "SUMMARY", "-" * 100,
        f"PASS: {summary['PASS']}", f"WARN: {summary['WARN']}", f"FAIL: {summary['FAIL']}",
        f"OVERALL: {overall_status(results)}", "=" * 100,
    ])
    return "\n".join(lines)


def save_reports(results):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TEXT_REPORT.write_text(build_text_report(results), encoding="utf-8")
    JSON_REPORT.write_text(json.dumps({
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "base_dir": str(BASE_DIR),
        "overall_status": overall_status(results),
        "results": [asdict(item) for item in results],
    }, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def run_audit(quick, stale_days):
    results = []
    audit_folders(results)
    audit_scripts(results)
    audit_database(results)
    audit_symbols_and_prices(results, stale_days)
    audit_intelligence(results)
    if not quick:
        audit_dashboard(results)
        audit_orchestrator_log(results)
        audit_git(results)
        audit_batch_files(results)
        audit_disk(results)
    save_reports(results)
    return results


def parse_arguments():
    parser = argparse.ArgumentParser(description="Run a complete AQSD system health audit.")
    parser.add_argument("--run", action="store_true", help="Run the full system audit.")
    parser.add_argument("--quick", action="store_true", help="Run a faster core audit.")
    parser.add_argument("--json", action="store_true", help="Print only JSON output.")
    parser.add_argument("--status", action="store_true", help="Run the full audit and show summary.")
    parser.add_argument("--stale-days", type=int, default=5,
                        help="Calendar days after which price data is considered stale. Default: 5.")
    return parser.parse_args()


def main():
    args = parse_arguments()
    results = run_audit(quick=args.quick, stale_days=max(1, args.stale_days))
    if args.json:
        print(json.dumps({
            "overall_status": overall_status(results),
            "results": [asdict(item) for item in results],
        }, ensure_ascii=False, indent=2, default=str))
        return
    print(build_text_report(results))
    print()
    print(f"Text report: {TEXT_REPORT}")
    print(f"JSON report: {JSON_REPORT}")
    if overall_status(results) == "FAIL":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
