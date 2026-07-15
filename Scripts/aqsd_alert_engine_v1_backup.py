
"""
AQSD Professional
Module: Alert Intelligence Engine
Version: 1.0

Purpose
-------
Creates actionable AQSD alerts from the latest:

- Unified Master Intelligence
- Trade Decision Engine
- Market Regime Intelligence
- Sector Rotation Intelligence
- Price Structure Intelligence

Alert types
-----------
- STRONG BUY
- BUY
- BUY ON DIP
- BREAKOUT
- BULLISH CHOCH
- BEARISH CHOCH
- EXIT / AVOID
- HIGH RISK
- LOW DATA COMPLETENESS
- SECTOR INFLOW
- SECTOR OUTFLOW
- MARKET REGIME CHANGE

Outputs
-------
- SQLite alert history
- CSV alert report
- Excel sheet: AQSD Alerts
- Console summary

Commands
--------
python aqsd_alert_engine.py --run
python aqsd_alert_engine.py --status
python aqsd_alert_engine.py --report
python aqsd_alert_engine.py --unread
python aqsd_alert_engine.py --mark-read
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "Output"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"
ALERT_CSV = OUTPUT_DIR / "AQSD_Alerts.csv"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS aqsd_alerts (
    alert_id INTEGER PRIMARY KEY AUTOINCREMENT,
    alert_date TEXT NOT NULL,
    alert_time TEXT NOT NULL,
    alert_type TEXT NOT NULL,
    severity TEXT NOT NULL,
    nse_symbol TEXT,
    sector TEXT,
    title TEXT NOT NULL,
    message TEXT NOT NULL,
    score REAL,
    status TEXT NOT NULL DEFAULT 'UNREAD',
    source_module TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(
        alert_date,
        alert_type,
        nse_symbol,
        title
    )
);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_date
ON aqsd_alerts(alert_date);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_status
ON aqsd_alerts(status);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_symbol
ON aqsd_alerts(nse_symbol);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# HELPERS
# ============================================================

def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table'
          AND name=?
        """,
        (table_name,),
    ).fetchone()

    return row is not None


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def severity_rank(value: str) -> int:
    mapping = {
        "CRITICAL": 1,
        "HIGH": 2,
        "MEDIUM": 3,
        "LOW": 4,
        "INFO": 5,
    }
    return mapping.get(value.upper(), 99)


def add_alert(
    alerts: list[dict],
    *,
    alert_type: str,
    severity: str,
    title: str,
    message: str,
    nse_symbol: str = "",
    sector: str = "",
    score: float | None = None,
    source_module: str = "",
) -> None:
    now = datetime.now()

    alerts.append(
        {
            "alert_date": now.date().isoformat(),
            "alert_time": now.strftime("%H:%M:%S"),
            "alert_type": alert_type,
            "severity": severity,
            "nse_symbol": nse_symbol,
            "sector": sector,
            "title": title,
            "message": message,
            "score": score,
            "status": "UNREAD",
            "source_module": source_module,
            "created_at": now.isoformat(timespec="seconds"),
        }
    )


# ============================================================
# LOADERS
# ============================================================

def load_latest_decisions() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "aqsd_trade_decisions"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM aqsd_trade_decisions
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
            ORDER BY priority_rank
            """,
            connection,
        )


def load_latest_structure() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "price_structure_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM price_structure_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
            """,
            connection,
        )


def load_latest_sectors() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "sector_rotation_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM sector_rotation_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            ORDER BY sector_rotation_score DESC
            """,
            connection,
        )


def load_latest_regime() -> dict | None:
    with connect() as connection:
        if not table_exists(connection, "market_regime_intelligence"):
            return None

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else None


def load_previous_regime() -> dict | None:
    with connect() as connection:
        if not table_exists(connection, "market_regime_intelligence"):
            return None

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1 OFFSET 1
            """
        ).fetchone()

    return dict(row) if row else None


# ============================================================
# ALERT GENERATION
# ============================================================

def build_decision_alerts(
    alerts: list[dict],
    decisions: pd.DataFrame,
) -> None:
    if decisions.empty:
        return

    for _, row in decisions.iterrows():
        action = str(row.get("action") or "")
        symbol = str(row.get("nse_symbol") or "")
        sector = str(row.get("sector") or "")
        score = safe_float(row.get("master_score"))
        confidence = safe_float(row.get("confidence_percent"))
        completeness = safe_float(row.get("completeness_percent"))
        risk = str(row.get("risk_level") or "")
        rank = row.get("priority_rank")

        if action == "STRONG BUY":
            add_alert(
                alerts,
                alert_type="STRONG BUY",
                severity="HIGH",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: Strong Buy",
                message=(
                    f"Priority rank {rank}. Master score {score}. "
                    f"Confidence {confidence}%. "
                    f"Entry zone {row.get('entry_low')} to {row.get('entry_high')}. "
                    f"Stop {row.get('stop_loss')}. "
                    f"Targets {row.get('target_1')} / {row.get('target_2')}."
                ),
                score=score,
                source_module="aqsd_decision_engine",
            )

        elif action == "BUY":
            add_alert(
                alerts,
                alert_type="BUY",
                severity="MEDIUM",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: Buy Candidate",
                message=(
                    f"Master score {score}, confidence {confidence}%, "
                    f"risk {risk}. Entry zone "
                    f"{row.get('entry_low')} to {row.get('entry_high')}."
                ),
                score=score,
                source_module="aqsd_decision_engine",
            )

        elif action == "BUY ON DIP":
            add_alert(
                alerts,
                alert_type="BUY ON DIP",
                severity="MEDIUM",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: Buy on Dip",
                message=(
                    f"Wait for the preferred entry zone "
                    f"{row.get('entry_low')} to {row.get('entry_high')}. "
                    f"Stop {row.get('stop_loss')}."
                ),
                score=score,
                source_module="aqsd_decision_engine",
            )

        elif action in {"AVOID", "EXIT / AVOID"}:
            add_alert(
                alerts,
                alert_type="EXIT / AVOID",
                severity="HIGH",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: {action}",
                message=(
                    f"Master score {score}, confidence {confidence}%, "
                    f"directional bias {row.get('directional_bias')}."
                ),
                score=score,
                source_module="aqsd_decision_engine",
            )

        if risk == "HIGH":
            add_alert(
                alerts,
                alert_type="HIGH RISK",
                severity="HIGH",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: High Risk",
                message=(
                    f"AQSD has classified this setup as high risk. "
                    f"Master score {score}; confidence {confidence}%."
                ),
                score=score,
                source_module="aqsd_decision_engine",
            )

        if completeness is not None and completeness < 50:
            add_alert(
                alerts,
                alert_type="LOW DATA COMPLETENESS",
                severity="LOW",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: Limited Data",
                message=(
                    f"Only {completeness}% of configured intelligence "
                    f"inputs are available."
                ),
                score=completeness,
                source_module="aqsd_unified_master_intelligence",
            )


def build_structure_alerts(
    alerts: list[dict],
    structure: pd.DataFrame,
) -> None:
    if structure.empty:
        return

    for _, row in structure.iterrows():
        symbol = str(row.get("nse_symbol") or "")
        score = safe_float(row.get("structure_score"))
        bos = str(row.get("bos_signal") or "")
        choch = str(row.get("choch_signal") or "")

        if bos == "BULLISH BOS":
            add_alert(
                alerts,
                alert_type="BREAKOUT",
                severity="MEDIUM",
                nse_symbol=symbol,
                title=f"{symbol}: Bullish Break of Structure",
                message=(
                    f"Close {row.get('close_price')} moved above "
                    f"the latest structural resistance. "
                    f"Structure score {score}."
                ),
                score=score,
                source_module="aqsd_price_structure",
            )

        elif bos == "BEARISH BOS":
            add_alert(
                alerts,
                alert_type="BREAKDOWN",
                severity="HIGH",
                nse_symbol=symbol,
                title=f"{symbol}: Bearish Break of Structure",
                message=(
                    f"Close {row.get('close_price')} moved below "
                    f"the latest structural support. "
                    f"Structure score {score}."
                ),
                score=score,
                source_module="aqsd_price_structure",
            )

        if choch == "BULLISH CHOCH":
            add_alert(
                alerts,
                alert_type="BULLISH CHOCH",
                severity="MEDIUM",
                nse_symbol=symbol,
                title=f"{symbol}: Bullish Change of Character",
                message="A possible bearish-to-bullish structural transition was detected.",
                score=score,
                source_module="aqsd_price_structure",
            )

        elif choch == "BEARISH CHOCH":
            add_alert(
                alerts,
                alert_type="BEARISH CHOCH",
                severity="HIGH",
                nse_symbol=symbol,
                title=f"{symbol}: Bearish Change of Character",
                message="A possible bullish-to-bearish structural transition was detected.",
                score=score,
                source_module="aqsd_price_structure",
            )


def build_sector_alerts(
    alerts: list[dict],
    sectors: pd.DataFrame,
) -> None:
    if sectors.empty:
        return

    for _, row in sectors.iterrows():
        sector = str(row.get("sector") or "")
        state = str(row.get("rotation_state") or "")
        score = safe_float(row.get("sector_rotation_score"))

        if state in {"STRONG INFLOW", "EARLY ROTATION"}:
            add_alert(
                alerts,
                alert_type="SECTOR INFLOW",
                severity="MEDIUM",
                sector=sector,
                title=f"{sector}: {state}",
                message=(
                    f"Sector rotation score {score}. "
                    f"Leader {row.get('leader_symbol')} "
                    f"with score {row.get('leader_score')}."
                ),
                score=score,
                source_module="aqsd_sector_rotation",
            )

        elif state in {"STRONG OUTFLOW", "WEAKENING LEADERSHIP"}:
            add_alert(
                alerts,
                alert_type="SECTOR OUTFLOW",
                severity="HIGH",
                sector=sector,
                title=f"{sector}: {state}",
                message=(
                    f"Sector rotation score {score}. "
                    f"Laggard {row.get('laggard_symbol')} "
                    f"with score {row.get('laggard_score')}."
                ),
                score=score,
                source_module="aqsd_sector_rotation",
            )


def build_regime_alerts(
    alerts: list[dict],
    current: dict | None,
    previous: dict | None,
) -> None:
    if not current:
        return

    current_regime = str(current.get("market_regime") or "")
    previous_regime = (
        str(previous.get("market_regime") or "")
        if previous
        else ""
    )

    if previous_regime and current_regime != previous_regime:
        add_alert(
            alerts,
            alert_type="MARKET REGIME CHANGE",
            severity="CRITICAL",
            title=f"Market Regime Changed to {current_regime}",
            message=(
                f"Previous regime: {previous_regime}. "
                f"Current strategy: {current.get('suggested_strategy')}. "
                f"Suggested exposure: "
                f"{current.get('capital_exposure_percent')}%."
            ),
            score=safe_float(current.get("regime_score")),
            source_module="aqsd_market_regime",
        )

    if current_regime in {"BEAR TREND", "RISK OFF", "HIGH VOLATILITY"}:
        add_alert(
            alerts,
            alert_type="MARKET RISK",
            severity="CRITICAL",
            title=f"Market Environment: {current_regime}",
            message=(
                f"Risk state {current.get('risk_state')}. "
                f"Volatility {current.get('volatility_state')}. "
                f"Suggested strategy: {current.get('suggested_strategy')}."
            ),
            score=safe_float(current.get("regime_score")),
            source_module="aqsd_market_regime",
        )


def build_alerts() -> list[dict]:
    alerts: list[dict] = []

    build_decision_alerts(
        alerts,
        load_latest_decisions(),
    )

    build_structure_alerts(
        alerts,
        load_latest_structure(),
    )

    build_sector_alerts(
        alerts,
        load_latest_sectors(),
    )

    build_regime_alerts(
        alerts,
        load_latest_regime(),
        load_previous_regime(),
    )

    alerts.sort(
        key=lambda item: (
            severity_rank(item["severity"]),
            -(item["score"] or 0),
            item["title"],
        )
    )

    return alerts


# ============================================================
# DATABASE WRITE
# ============================================================

def save_alerts(alerts: list[dict]) -> tuple[int, int]:
    inserted = 0
    duplicates = 0

    with connect() as connection:
        for alert in alerts:
            try:
                connection.execute(
                    """
                    INSERT INTO aqsd_alerts(
                        alert_date,
                        alert_time,
                        alert_type,
                        severity,
                        nse_symbol,
                        sector,
                        title,
                        message,
                        score,
                        status,
                        source_module,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        alert["alert_date"],
                        alert["alert_time"],
                        alert["alert_type"],
                        alert["severity"],
                        alert["nse_symbol"],
                        alert["sector"],
                        alert["title"],
                        alert["message"],
                        alert["score"],
                        alert["status"],
                        alert["source_module"],
                        alert["created_at"],
                    ),
                )
                inserted += 1

            except Exception as error:
                if "UNIQUE constraint failed" in str(error):
                    duplicates += 1
                else:
                    raise

        connection.commit()

    return inserted, duplicates


def run_engine() -> tuple[list[dict], int, int]:
    setup_schema()

    run_id = start_run(
        "aqsd_alert_engine",
        "Generating AQSD alerts",
    )

    try:
        alerts = build_alerts()
        inserted, duplicates = save_alerts(alerts)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=inserted,
            errors_count=duplicates,
            message=(
                f"Generated={len(alerts)}; "
                f"Inserted={inserted}; "
                f"Duplicates={duplicates}"
            ),
        )

        return alerts, inserted, duplicates

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


# ============================================================
# REPORTING
# ============================================================

def latest_alerts(
    unread_only: bool = False,
) -> pd.DataFrame:
    setup_schema()

    query = """
        SELECT *
        FROM aqsd_alerts
        WHERE alert_date=(
            SELECT MAX(alert_date)
            FROM aqsd_alerts
        )
    """

    if unread_only:
        query += " AND status='UNREAD'"

    query += """
        ORDER BY
            CASE severity
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'MEDIUM' THEN 3
                WHEN 'LOW' THEN 4
                ELSE 5
            END,
            score DESC,
            alert_id
    """

    with connect() as connection:
        return pd.read_sql_query(
            query,
            connection,
        )


def write_csv(frame: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    frame.to_csv(
        ALERT_CSV,
        index=False,
        encoding="utf-8-sig",
    )


def write_report(
    frame: pd.DataFrame | None = None,
) -> None:
    if frame is None:
        frame = latest_alerts()

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "AQSD Alerts" in wb.sheetnames:
        del wb["AQSD Alerts"]

    ws = wb.create_sheet("AQSD Alerts", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:L2")
    ws["A1"] = "AQSD PROFESSIONAL - ALERT INTELLIGENCE ENGINE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    unread = 0

    if not frame.empty:
        unread = int(
            (frame["status"] == "UNREAD").sum()
        )

    ws["A4"] = "Latest Alerts"
    ws["B4"] = len(frame)
    ws["D4"] = "Unread"
    ws["E4"] = unread
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Date",
        "Time",
        "Severity",
        "Alert Type",
        "Symbol",
        "Sector",
        "Title",
        "Message",
        "Score",
        "Status",
        "Source",
    ]

    for col, heading in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, (_, row) in enumerate(
        frame.iterrows(),
        start=8,
    ):
        values = [
            row.get("alert_date"),
            row.get("alert_time"),
            row.get("severity"),
            row.get("alert_type"),
            row.get("nse_symbol"),
            row.get("sector"),
            row.get("title"),
            row.get("message"),
            row.get("score"),
            row.get("status"),
            row.get("source_module"),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(
                bottom=THIN
            )

        severity = str(
            row.get("severity") or ""
        )

        ws.cell(row_no, 3).fill = PatternFill(
            "solid",
            fgColor=(
                RED
                if severity in {"CRITICAL", "HIGH"}
                else YELLOW
                if severity == "MEDIUM"
                else GREY
            ),
        )

    widths = {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 22,
        "E": 16,
        "F": 22,
        "G": 36,
        "H": 90,
        "I": 12,
        "J": 12,
        "K": 28,
        "L": 12,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


def mark_all_read() -> int:
    setup_schema()

    with connect() as connection:
        cursor = connection.execute(
            """
            UPDATE aqsd_alerts
            SET status='READ'
            WHERE status='UNREAD'
            """
        )
        connection.commit()

    return cursor.rowcount


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status='UNREAD' THEN 1 ELSE 0 END) AS unread,
                MIN(alert_date) AS first_date,
                MAX(alert_date) AS latest_date
            FROM aqsd_alerts
            """
        ).fetchone()

    print("\nAQSD ALERT ENGINE STATUS")
    print("=" * 72)
    print(f"Stored alerts: {row['total'] or 0}")
    print(f"Unread alerts: {row['unread'] or 0}")
    print(f"First date:    {row['first_date'] or 'No data'}")
    print(f"Latest date:   {row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Alert Intelligence Engine."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Generate the latest AQSD alerts.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild Excel and CSV reports.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show alert-engine status.",
    )

    parser.add_argument(
        "--unread",
        action="store_true",
        help="Show unread alerts only.",
    )

    parser.add_argument(
        "--mark-read",
        action="store_true",
        help="Mark all unread alerts as read.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.mark_read:
        count = mark_all_read()
        print(f"Alerts marked as read: {count}")
        return

    if args.unread:
        frame = latest_alerts(
            unread_only=True
        )

        if frame.empty:
            print("No unread alerts.")
        else:
            print(
                frame[
                    [
                        "severity",
                        "alert_type",
                        "nse_symbol",
                        "sector",
                        "title",
                        "message",
                    ]
                ].to_string(index=False)
            )
        return

    if args.run:
        alerts, inserted, duplicates = run_engine()

        frame = latest_alerts()
        write_csv(frame)
        write_report(frame)

        print("\nAQSD ALERT ENGINE")
        print("=" * 72)
        print(f"Alerts generated: {len(alerts)}")
        print(f"Inserted:         {inserted}")
        print(f"Duplicates:       {duplicates}")
        print(f"CSV:              {ALERT_CSV}")
        print(f"Dashboard:        {DASHBOARD}")
        return

    if args.report:
        frame = latest_alerts()
        write_csv(frame)
        write_report(frame)

        print(f"CSV rebuilt:\n{ALERT_CSV}")
        print(f"Dashboard rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()

