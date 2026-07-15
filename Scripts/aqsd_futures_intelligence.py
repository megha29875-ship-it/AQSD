
"""
AQSD Derivatives Intelligence
Module: Futures OI Intelligence Engine
Version: 1.0

Purpose
-------
Builds futures positioning intelligence from imported NSE/broker/vendor data.

This module does NOT assume a live derivatives feed. It accepts structured
CSV input and stores the results in AQSD's SQLite database.

Core classifications
--------------------
Price Up   + OI Up   = LONG BUILD-UP
Price Down + OI Up   = SHORT BUILD-UP
Price Up   + OI Down = SHORT COVERING
Price Down + OI Down = LONG UNWINDING

Features
--------
- CSV import
- Duplicate-safe storage
- Futures OI classification
- Positional and daily OI analysis
- Smart Money Score from 0 to 100
- Bullish / Bearish bias
- Excel report: Futures Intelligence
- Symbol-level derivatives score for AQSD Master Intelligence

Expected CSV columns
--------------------
trade_date
nse_symbol
contract_symbol
expiry_date
futures_price
previous_futures_price
open_interest
previous_open_interest
volume
spot_price
basis
rollover_percent
cost_of_carry

Only these columns are mandatory:
trade_date, nse_symbol, futures_price, open_interest

Commands
--------
python aqsd_futures_intelligence.py --setup
python aqsd_futures_intelligence.py --import-csv C:\\path\\futures_data.csv
python aqsd_futures_intelligence.py --analyse
python aqsd_futures_intelligence.py --report
python aqsd_futures_intelligence.py --status
"""

from __future__ import annotations

import argparse
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS futures_oi (
    futures_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    nse_symbol TEXT NOT NULL,
    contract_symbol TEXT,
    expiry_date TEXT,
    futures_price REAL NOT NULL,
    previous_futures_price REAL,
    price_change_percent REAL,
    open_interest REAL NOT NULL,
    previous_open_interest REAL,
    oi_change REAL,
    oi_change_percent REAL,
    volume REAL,
    spot_price REAL,
    basis REAL,
    rollover_percent REAL,
    cost_of_carry REAL,
    buildup_type TEXT,
    directional_bias TEXT,
    smart_money_score REAL,
    source TEXT,
    event_hash TEXT NOT NULL UNIQUE,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_futures_symbol_date
ON futures_oi(nse_symbol, trade_date);

CREATE INDEX IF NOT EXISTS idx_futures_buildup
ON futures_oi(buildup_type);
"""


def setup_futures_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# HELPERS
# ============================================================

def clean_text(value: Any) -> str:
    text = str(value or "").strip()

    if text.lower() in {"nan", "none", "null"}:
        return ""

    return text


def normalize_symbol(value: Any) -> str:
    symbol = clean_text(value).upper()

    if symbol.endswith(".NS"):
        symbol = symbol[:-3]

    return symbol.replace(" ", "")


def safe_float(
    value: Any,
    default: float | None = None,
) -> float | None:
    try:
        if value is None or value == "":
            return default

        return float(value)

    except (TypeError, ValueError):
        return default


def parse_date(value: Any) -> str:
    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        raise ValueError(f"Invalid date: {value}")

    return parsed.date().isoformat()


def percentage_change(
    current: float | None,
    previous: float | None,
) -> float | None:
    if current is None or previous in (None, 0):
        return None

    return round((current / previous - 1) * 100, 2)


def build_hash(record: dict) -> str:
    raw = "|".join(
        [
            record["trade_date"],
            record["nse_symbol"],
            record.get("contract_symbol", ""),
            record.get("expiry_date", ""),
        ]
    )

    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ============================================================
# INTELLIGENCE LOGIC
# ============================================================

def classify_buildup(
    price_change_percent: float | None,
    oi_change_percent: float | None,
) -> tuple[str, str]:
    if price_change_percent is None or oi_change_percent is None:
        return "UNCLASSIFIED", "NEUTRAL"

    if price_change_percent > 0 and oi_change_percent > 0:
        return "LONG BUILD-UP", "BULLISH"

    if price_change_percent < 0 and oi_change_percent > 0:
        return "SHORT BUILD-UP", "BEARISH"

    if price_change_percent > 0 and oi_change_percent < 0:
        return "SHORT COVERING", "BULLISH"

    if price_change_percent < 0 and oi_change_percent < 0:
        return "LONG UNWINDING", "BEARISH"

    return "NEUTRAL", "NEUTRAL"


def calculate_smart_money_score(record: dict) -> float:
    price_change = float(record.get("price_change_percent") or 0)
    oi_change = float(record.get("oi_change_percent") or 0)
    volume = float(record.get("volume") or 0)
    rollover = float(record.get("rollover_percent") or 0)
    basis = float(record.get("basis") or 0)
    cost_of_carry = float(record.get("cost_of_carry") or 0)
    buildup = record.get("buildup_type", "UNCLASSIFIED")

    score = 50.0

    if buildup == "LONG BUILD-UP":
        score += min(20, abs(price_change) * 4)
        score += min(20, abs(oi_change) * 1.5)

    elif buildup == "SHORT COVERING":
        score += min(15, abs(price_change) * 4)
        score += min(15, abs(oi_change) * 1.2)

    elif buildup == "SHORT BUILD-UP":
        score -= min(20, abs(price_change) * 4)
        score -= min(20, abs(oi_change) * 1.5)

    elif buildup == "LONG UNWINDING":
        score -= min(15, abs(price_change) * 4)
        score -= min(15, abs(oi_change) * 1.2)

    if volume > 0:
        score += min(5, max(0, volume / 1_000_000))

    if rollover > 70:
        score += 4 if buildup in {"LONG BUILD-UP", "SHORT COVERING"} else -4

    if basis > 0:
        score += 3
    elif basis < 0:
        score -= 3

    if cost_of_carry > 0:
        score += 2
    elif cost_of_carry < 0:
        score -= 2

    return round(max(0, min(100, score)), 2)


def derivatives_bias(score: float) -> str:
    if score >= 75:
        return "STRONG BULLISH"
    if score >= 60:
        return "BULLISH"
    if score <= 25:
        return "STRONG BEARISH"
    if score <= 40:
        return "BEARISH"
    return "NEUTRAL"


# ============================================================
# CSV IMPORT
# ============================================================

def normalize_input_frame(frame: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "date": "trade_date",
        "symbol": "nse_symbol",
        "ticker": "nse_symbol",
        "future_price": "futures_price",
        "fut_price": "futures_price",
        "prev_future_price": "previous_futures_price",
        "prev_futures_price": "previous_futures_price",
        "oi": "open_interest",
        "prev_oi": "previous_open_interest",
        "expiry": "expiry_date",
        "contract": "contract_symbol",
    }

    renamed = {}

    for column in frame.columns:
        key = str(column).strip().lower().replace(" ", "_")

        renamed[column] = aliases.get(key, key)

    return frame.rename(columns=renamed)


def import_csv(path: Path, source: str) -> tuple[int, int]:
    setup_futures_schema()

    if not path.exists():
        raise FileNotFoundError(path)

    frame = normalize_input_frame(
        pd.read_csv(path)
    )

    required = {
        "trade_date",
        "nse_symbol",
        "futures_price",
        "open_interest",
    }

    if not required.issubset(frame.columns):
        missing = sorted(required - set(frame.columns))
        raise RuntimeError(
            "Missing required columns: " + ", ".join(missing)
        )

    inserted = 0
    duplicates = 0

    run_id = start_run(
        "aqsd_futures_intelligence",
        f"Importing {path.name}",
    )

    try:
        with connect() as connection:
            for _, row in frame.iterrows():
                record = {
                    "trade_date": parse_date(row.get("trade_date")),
                    "nse_symbol": normalize_symbol(
                        row.get("nse_symbol")
                    ),
                    "contract_symbol": clean_text(
                        row.get("contract_symbol")
                    ),
                    "expiry_date": (
                        parse_date(row.get("expiry_date"))
                        if clean_text(row.get("expiry_date"))
                        else ""
                    ),
                    "futures_price": safe_float(
                        row.get("futures_price")
                    ),
                    "previous_futures_price": safe_float(
                        row.get("previous_futures_price")
                    ),
                    "open_interest": safe_float(
                        row.get("open_interest")
                    ),
                    "previous_open_interest": safe_float(
                        row.get("previous_open_interest")
                    ),
                    "volume": safe_float(row.get("volume"), 0),
                    "spot_price": safe_float(row.get("spot_price")),
                    "basis": safe_float(row.get("basis")),
                    "rollover_percent": safe_float(
                        row.get("rollover_percent")
                    ),
                    "cost_of_carry": safe_float(
                        row.get("cost_of_carry")
                    ),
                }

                if not record["nse_symbol"]:
                    continue

                if (
                    record["futures_price"] is None
                    or record["open_interest"] is None
                ):
                    continue

                price_change = percentage_change(
                    record["futures_price"],
                    record["previous_futures_price"],
                )

                oi_change = (
                    record["open_interest"]
                    - record["previous_open_interest"]
                    if record["previous_open_interest"] is not None
                    else None
                )

                oi_change_percent = percentage_change(
                    record["open_interest"],
                    record["previous_open_interest"],
                )

                buildup, bias = classify_buildup(
                    price_change,
                    oi_change_percent,
                )

                record["price_change_percent"] = price_change
                record["oi_change"] = oi_change
                record["oi_change_percent"] = oi_change_percent
                record["buildup_type"] = buildup
                record["directional_bias"] = bias
                record["smart_money_score"] = (
                    calculate_smart_money_score(record)
                )

                hashed = build_hash(record)

                try:
                    connection.execute(
                        """
                        INSERT INTO futures_oi(
                            trade_date,
                            nse_symbol,
                            contract_symbol,
                            expiry_date,
                            futures_price,
                            previous_futures_price,
                            price_change_percent,
                            open_interest,
                            previous_open_interest,
                            oi_change,
                            oi_change_percent,
                            volume,
                            spot_price,
                            basis,
                            rollover_percent,
                            cost_of_carry,
                            buildup_type,
                            directional_bias,
                            smart_money_score,
                            source,
                            event_hash,
                            created_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            record["trade_date"],
                            record["nse_symbol"],
                            record["contract_symbol"],
                            record["expiry_date"],
                            record["futures_price"],
                            record["previous_futures_price"],
                            record["price_change_percent"],
                            record["open_interest"],
                            record["previous_open_interest"],
                            record["oi_change"],
                            record["oi_change_percent"],
                            record["volume"],
                            record["spot_price"],
                            record["basis"],
                            record["rollover_percent"],
                            record["cost_of_carry"],
                            record["buildup_type"],
                            record["directional_bias"],
                            record["smart_money_score"],
                            source,
                            hashed,
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )

                    inserted += 1

                except Exception as error:
                    if "UNIQUE constraint failed" in str(error):
                        duplicates += 1
                    else:
                        raise

            connection.commit()

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=inserted,
            errors_count=duplicates,
            message=(
                f"Inserted={inserted}; duplicates={duplicates}"
            ),
        )

        return inserted, duplicates

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=inserted,
            errors_count=duplicates + 1,
            message=str(error),
        )
        raise


# ============================================================
# ANALYSIS
# ============================================================

def latest_futures_rows() -> pd.DataFrame:
    setup_futures_schema()

    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT *
            FROM futures_oi
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM futures_oi
            )
            ORDER BY smart_money_score DESC, nse_symbol
            """,
            connection,
        )

    return frame


def symbol_derivatives_scores() -> pd.DataFrame:
    frame = latest_futures_rows()

    if frame.empty:
        return frame

    grouped = (
        frame.groupby("nse_symbol")
        .agg(
            Futures_Contracts=("futures_id", "count"),
            Derivatives_Score=("smart_money_score", "mean"),
            Price_Change_Percent=("price_change_percent", "mean"),
            OI_Change_Percent=("oi_change_percent", "mean"),
            Latest_Trade_Date=("trade_date", "max"),
        )
        .reset_index()
    )

    grouped["Derivatives_Score"] = grouped[
        "Derivatives_Score"
    ].round(2)

    grouped["Derivatives_Bias"] = grouped[
        "Derivatives_Score"
    ].apply(derivatives_bias)

    return grouped.sort_values(
        "Derivatives_Score",
        ascending=False,
    )


# ============================================================
# EXCEL REPORT
# ============================================================

def write_report() -> None:
    frame = latest_futures_rows()
    scores = symbol_derivatives_scores()

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Futures Intelligence" in wb.sheetnames:
        del wb["Futures Intelligence"]

    ws = wb.create_sheet("Futures Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:U2")
    ws["A1"] = "AQSD PROFESSIONAL - FUTURES OI INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Contracts Analysed"
    ws["B4"] = len(frame)
    ws["D4"] = "Symbols Scored"
    ws["E4"] = len(scores)
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    summary_headers = [
        "Rank",
        "Symbol",
        "Contracts",
        "Derivatives Score",
        "Derivatives Bias",
        "Price Change %",
        "OI Change %",
        "Latest Date",
    ]

    for col, heading in enumerate(summary_headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)

    for row_no, (_, row) in enumerate(
        scores.iterrows(),
        start=8,
    ):
        values = [
            row_no - 7,
            row["nse_symbol"],
            row["Futures_Contracts"],
            row["Derivatives_Score"],
            row["Derivatives_Bias"],
            row["Price_Change_Percent"],
            row["OI_Change_Percent"],
            row["Latest_Trade_Date"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        score = float(row["Derivatives_Score"])

        ws.cell(row_no, 4).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )

    detail_start = max(12, 10 + len(scores))

    ws.cell(
        detail_start,
        1,
        "DETAILED FUTURES CONTRACTS",
    ).font = Font(
        size=14,
        bold=True,
        color=WHITE,
    )
    ws.cell(
        detail_start,
        1,
    ).fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    detail_headers = [
        "Trade Date",
        "Symbol",
        "Contract",
        "Expiry",
        "Futures Price",
        "Previous Price",
        "Price Change %",
        "Open Interest",
        "Previous OI",
        "OI Change",
        "OI Change %",
        "Volume",
        "Spot Price",
        "Basis",
        "Rollover %",
        "Cost of Carry",
        "Build-up Type",
        "Directional Bias",
        "Smart Money Score",
        "Source",
    ]

    for col, heading in enumerate(detail_headers, start=1):
        cell = ws.cell(detail_start + 2, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        frame.iterrows(),
        start=detail_start + 3,
    ):
        values = [
            row.get("trade_date"),
            row.get("nse_symbol"),
            row.get("contract_symbol"),
            row.get("expiry_date"),
            row.get("futures_price"),
            row.get("previous_futures_price"),
            row.get("price_change_percent"),
            row.get("open_interest"),
            row.get("previous_open_interest"),
            row.get("oi_change"),
            row.get("oi_change_percent"),
            row.get("volume"),
            row.get("spot_price"),
            row.get("basis"),
            row.get("rollover_percent"),
            row.get("cost_of_carry"),
            row.get("buildup_type"),
            row.get("directional_bias"),
            row.get("smart_money_score"),
            row.get("source"),
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        buildup = str(row.get("buildup_type") or "")

        ws.cell(row_no, 17).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if buildup in {"LONG BUILD-UP", "SHORT COVERING"}
                else RED
                if buildup in {"SHORT BUILD-UP", "LONG UNWINDING"}
                else GREY
            ),
        )

    widths = {
        "A": 14,
        "B": 16,
        "C": 22,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 15,
        "I": 15,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 12,
        "O": 13,
        "P": 14,
        "Q": 20,
        "R": 18,
        "S": 18,
        "T": 20,
        "U": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


# ============================================================
# STATUS
# ============================================================

def show_status() -> None:
    setup_futures_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT nse_symbol) AS symbols,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM futures_oi
            """
        ).fetchone()

    print("\nAQSD FUTURES INTELLIGENCE STATUS")
    print("=" * 72)
    print(f"Stored contracts:    {row['total'] or 0}")
    print(f"Symbols covered:     {row['symbols'] or 0}")
    print(f"First trade date:    {row['first_date'] or 'No data'}")
    print(f"Latest trade date:   {row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Futures OI Intelligence Engine."
    )

    parser.add_argument(
        "--setup",
        action="store_true",
        help="Create futures intelligence tables.",
    )

    parser.add_argument(
        "--import-csv",
        type=Path,
        help="Import structured futures/OI CSV data.",
    )

    parser.add_argument(
        "--source",
        default="Manual CSV",
        help="Data-source label for imported rows.",
    )

    parser.add_argument(
        "--analyse",
        action="store_true",
        help="Display latest symbol-level derivatives scores.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Create the Futures Intelligence Excel report.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show futures database status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_futures_schema()

    if args.setup:
        print("AQSD Futures Intelligence schema is ready.")
        return

    if args.import_csv:
        inserted, duplicates = import_csv(
            args.import_csv,
            args.source,
        )

        print("\nAQSD FUTURES IMPORT")
        print("=" * 72)
        print(f"Inserted:   {inserted}")
        print(f"Duplicates: {duplicates}")
        return

    if args.analyse:
        frame = symbol_derivatives_scores()

        if frame.empty:
            print("No futures/OI intelligence available.")
        else:
            print(frame.to_string(index=False))
        return

    if args.report:
        write_report()
        print(f"Futures Intelligence report created:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
