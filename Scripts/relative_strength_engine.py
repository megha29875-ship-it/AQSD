
"""
AQSD Professional
Module: Relative Strength Intelligence Engine
Version: 1.0

Compares NSE F&O stocks against NIFTY 50 and ranks true relative strength.

Features
--------
- 20-day relative strength
- 60-day relative strength
- 120-day relative strength
- Relative momentum acceleration
- Benchmark outperformance
- Relative Strength Score from 0 to 100
- Leader / Improving / Neutral / Weakening / Laggard classification
- Excel output: Relative Strength

Run
---
python relative_strength_engine.py
python relative_strength_engine.py --period 1y
python relative_strength_engine.py --limit 100

Notes
-----
--limit 0 means analyse the entire available F&O universe.
Yahoo Finance prices may be delayed.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"
FNO_FILE = BASE / "Data" / "FnO_Stocks.xlsx"


# ============================================================
# SETTINGS
# ============================================================

BENCHMARK = "^NSEI"
DEFAULT_PERIOD = "1y"
DEFAULT_LIMIT = 0

LOOKBACK_SHORT = 20
LOOKBACK_MEDIUM = 60
LOOKBACK_LONG = 120


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def normalize_symbol(value) -> str:
    symbol = str(value or "").strip().upper()

    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"

    return symbol


def load_symbols(limit: int) -> list[str]:
    """
    Prefer the full F&O file.
    Fall back to candidate sheets when the file is unavailable.
    """

    symbols: list[str] = []

    if FNO_FILE.exists():
        try:
            df = pd.read_excel(FNO_FILE)

            symbol_col = next(
                (
                    column
                    for column in (
                        "Yahoo Symbol",
                        "Symbol",
                        "SYMBOL",
                        "Ticker",
                    )
                    if column in df.columns
                ),
                None,
            )

            if symbol_col:
                for value in df[symbol_col].dropna():
                    symbol = normalize_symbol(value)

                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

                    if limit > 0 and len(symbols) >= limit:
                        return symbols

        except Exception:
            pass

    if DASHBOARD.exists() and (limit == 0 or len(symbols) < limit):
        try:
            wb = load_workbook(
                DASHBOARD,
                read_only=True,
                data_only=True,
            )

            for sheet_name in ("CALL Candidates", "PUT Candidates"):
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                headers = {
                    str(cell.value).strip(): cell.column
                    for cell in ws[1]
                    if cell.value is not None
                }

                symbol_col = headers.get("Symbol")

                if not symbol_col:
                    continue

                for row in range(2, ws.max_row + 1):
                    symbol = normalize_symbol(
                        ws.cell(row, symbol_col).value
                    )

                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

                    if limit > 0 and len(symbols) >= limit:
                        wb.close()
                        return symbols

            wb.close()

        except Exception:
            pass

    fallback = [
        "RELIANCE.NS",
        "HDFCBANK.NS",
        "ICICIBANK.NS",
        "SBIN.NS",
        "INFY.NS",
        "TCS.NS",
        "LT.NS",
        "SUNPHARMA.NS",
        "BIOCON.NS",
        "TATAMOTORS.NS",
    ]

    for symbol in fallback:
        if symbol not in symbols:
            symbols.append(symbol)

        if limit > 0 and len(symbols) >= limit:
            break

    return symbols


def total_return(series: pd.Series, lookback: int) -> float | None:
    clean = series.dropna()

    if len(clean) <= lookback:
        return None

    current = float(clean.iloc[-1])
    previous = float(clean.iloc[-1 - lookback])

    if previous == 0:
        return None

    return (current / previous - 1) * 100


def percentile_scores(values: pd.Series) -> pd.Series:
    """
    Convert a series into 0-100 percentile scores.
    """

    return values.rank(
        method="average",
        pct=True,
    ) * 100


def classify_relative_strength(
    score: float,
    rs20: float,
    rs60: float,
) -> str:
    if score >= 80 and rs20 > 0 and rs60 > 0:
        return "LEADER"

    if score >= 60 and rs20 > rs60:
        return "IMPROVING"

    if score <= 20 and rs20 < 0 and rs60 < 0:
        return "LAGGARD"

    if score <= 40 and rs20 < rs60:
        return "WEAKENING"

    return "NEUTRAL"


def build_reason(
    classification: str,
    rs20: float,
    rs60: float,
    rs120: float,
    acceleration: float,
) -> str:
    reasons = [classification.title()]

    if rs20 > 0:
        reasons.append("20D outperforming NIFTY")
    else:
        reasons.append("20D underperforming NIFTY")

    if rs60 > 0:
        reasons.append("60D relative trend positive")
    else:
        reasons.append("60D relative trend negative")

    if rs120 > 0:
        reasons.append("Long-term leadership")
    else:
        reasons.append("Long-term lagging")

    if acceleration > 0:
        reasons.append("Relative momentum improving")
    elif acceleration < 0:
        reasons.append("Relative momentum slowing")

    return " | ".join(reasons)


# ============================================================
# DATA DOWNLOAD
# ============================================================

def download_close_data(
    symbols: list[str],
    period: str,
) -> pd.DataFrame:
    tickers = [BENCHMARK, *symbols]

    data = yf.download(
        tickers=tickers,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
        group_by="column",
        threads=True,
    )

    if data.empty:
        raise RuntimeError("No market data downloaded.")

    if isinstance(data.columns, pd.MultiIndex):
        if "Close" not in data.columns.get_level_values(0):
            raise RuntimeError("Close data missing.")

        close = data["Close"].copy()
    else:
        close = data[["Close"]].copy()
        close.columns = tickers[:1]

    close = close.dropna(how="all")

    if BENCHMARK not in close.columns:
        raise RuntimeError(
            f"Benchmark data missing for {BENCHMARK}."
        )

    return close


# ============================================================
# ANALYSIS
# ============================================================

def analyse_relative_strength(
    symbols: list[str],
    close: pd.DataFrame,
) -> list[dict]:
    benchmark = close[BENCHMARK].dropna()

    benchmark_returns = {
        LOOKBACK_SHORT: total_return(
            benchmark,
            LOOKBACK_SHORT,
        ),
        LOOKBACK_MEDIUM: total_return(
            benchmark,
            LOOKBACK_MEDIUM,
        ),
        LOOKBACK_LONG: total_return(
            benchmark,
            LOOKBACK_LONG,
        ),
    }

    raw_rows: list[dict] = []

    for symbol in symbols:
        if symbol not in close.columns:
            continue

        stock = close[symbol].dropna()

        return20 = total_return(stock, LOOKBACK_SHORT)
        return60 = total_return(stock, LOOKBACK_MEDIUM)
        return120 = total_return(stock, LOOKBACK_LONG)

        if None in (
            return20,
            return60,
            return120,
            benchmark_returns[LOOKBACK_SHORT],
            benchmark_returns[LOOKBACK_MEDIUM],
            benchmark_returns[LOOKBACK_LONG],
        ):
            continue

        rs20 = return20 - benchmark_returns[LOOKBACK_SHORT]
        rs60 = return60 - benchmark_returns[LOOKBACK_MEDIUM]
        rs120 = return120 - benchmark_returns[LOOKBACK_LONG]

        acceleration = rs20 - rs60

        raw_rows.append(
            {
                "Symbol": symbol,
                "Close": round(float(stock.iloc[-1]), 2),
                "20D Return %": round(return20, 2),
                "60D Return %": round(return60, 2),
                "120D Return %": round(return120, 2),
                "NIFTY 20D %": round(
                    benchmark_returns[LOOKBACK_SHORT],
                    2,
                ),
                "NIFTY 60D %": round(
                    benchmark_returns[LOOKBACK_MEDIUM],
                    2,
                ),
                "NIFTY 120D %": round(
                    benchmark_returns[LOOKBACK_LONG],
                    2,
                ),
                "RS 20D %": round(rs20, 2),
                "RS 60D %": round(rs60, 2),
                "RS 120D %": round(rs120, 2),
                "RS Acceleration": round(acceleration, 2),
            }
        )

    if not raw_rows:
        return []

    frame = pd.DataFrame(raw_rows)

    frame["Score 20"] = percentile_scores(frame["RS 20D %"])
    frame["Score 60"] = percentile_scores(frame["RS 60D %"])
    frame["Score 120"] = percentile_scores(frame["RS 120D %"])
    frame["Score Acceleration"] = percentile_scores(
        frame["RS Acceleration"]
    )

    frame["Relative Strength Score"] = (
        frame["Score 20"] * 0.35
        + frame["Score 60"] * 0.30
        + frame["Score 120"] * 0.25
        + frame["Score Acceleration"] * 0.10
    ).round(0).astype(int)

    results: list[dict] = []

    for _, row in frame.iterrows():
        score = float(row["Relative Strength Score"])
        rs20 = float(row["RS 20D %"])
        rs60 = float(row["RS 60D %"])
        rs120 = float(row["RS 120D %"])
        acceleration = float(row["RS Acceleration"])

        classification = classify_relative_strength(
            score,
            rs20,
            rs60,
        )

        result = {
            "Symbol": row["Symbol"],
            "Close": row["Close"],
            "20D Return %": row["20D Return %"],
            "60D Return %": row["60D Return %"],
            "120D Return %": row["120D Return %"],
            "NIFTY 20D %": row["NIFTY 20D %"],
            "NIFTY 60D %": row["NIFTY 60D %"],
            "NIFTY 120D %": row["NIFTY 120D %"],
            "RS 20D %": row["RS 20D %"],
            "RS 60D %": row["RS 60D %"],
            "RS 120D %": row["RS 120D %"],
            "RS Acceleration": row["RS Acceleration"],
            "Relative Strength Score": int(score),
            "Classification": classification,
            "Reason": build_reason(
                classification,
                rs20,
                rs60,
                rs120,
                acceleration,
            ),
        }

        results.append(result)

    return sorted(
        results,
        key=lambda item: item["Relative Strength Score"],
        reverse=True,
    )


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_results(results: list[dict]) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Relative Strength" in wb.sheetnames:
        del wb["Relative Strength"]

    ws = wb.create_sheet("Relative Strength", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:O2")
    ws["A1"] = (
        "AQSD PROFESSIONAL - RELATIVE STRENGTH INTELLIGENCE"
    )
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Benchmark"
    ws["B4"] = "NIFTY 50"
    ws["D4"] = "Stocks Analysed"
    ws["E4"] = len(results)

    for ref in ("A4", "D4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Rank",
        "Symbol",
        "Close",
        "20D Return %",
        "60D Return %",
        "120D Return %",
        "NIFTY 20D %",
        "NIFTY 60D %",
        "NIFTY 120D %",
        "RS 20D %",
        "RS 60D %",
        "RS 120D %",
        "RS Acceleration",
        "Relative Strength Score",
        "Classification",
        "Reason",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(6, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, result in enumerate(results, start=7):
        values = [
            row_no - 6,
            result["Symbol"],
            result["Close"],
            result["20D Return %"],
            result["60D Return %"],
            result["120D Return %"],
            result["NIFTY 20D %"],
            result["NIFTY 60D %"],
            result["NIFTY 120D %"],
            result["RS 20D %"],
            result["RS 60D %"],
            result["RS 120D %"],
            result["RS Acceleration"],
            result["Relative Strength Score"],
            result["Classification"],
            result["Reason"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            cell.border = Border(bottom=THIN)

        ws.cell(
            row_no,
            3,
        ).number_format = '₹#,##0.00'

        for col in range(4, 14):
            ws.cell(
                row_no,
                col,
            ).number_format = '0.00"%"'

        score = result["Relative Strength Score"]

        score_fill = (
            GREEN
            if score >= 70
            else RED
            if score <= 30
            else YELLOW
        )

        ws.cell(
            row_no,
            14,
        ).fill = PatternFill(
            "solid",
            fgColor=score_fill,
        )

        ws.cell(
            row_no,
            14,
        ).font = Font(bold=True)

        classification = result["Classification"]

        class_fill = (
            GREEN
            if classification in {"LEADER", "IMPROVING"}
            else RED
            if classification in {"LAGGARD", "WEAKENING"}
            else GREY
        )

        ws.cell(
            row_no,
            15,
        ).fill = PatternFill(
            "solid",
            fgColor=class_fill,
        )

        ws.cell(
            row_no,
            15,
        ).font = Font(bold=True)

    widths = {
        "A": 8,
        "B": 18,
        "C": 12,
        "D": 13,
        "E": 13,
        "F": 14,
        "G": 13,
        "H": 13,
        "I": 14,
        "J": 12,
        "K": 12,
        "L": 13,
        "M": 15,
        "N": 18,
        "O": 15,
        "P": 60,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build AQSD relative-strength intelligence."
    )

    parser.add_argument(
        "--period",
        default=DEFAULT_PERIOD,
        help="Yahoo Finance period, normally 1y or 2y.",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_LIMIT,
        help=(
            "Maximum stocks to analyse. "
            "Use 0 for the entire available F&O universe."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    symbols = load_symbols(args.limit)

    print("\nAQSD RELATIVE STRENGTH INTELLIGENCE")
    print("=" * 72)
    print(f"Symbols requested: {len(symbols)}")
    print(f"Benchmark: {BENCHMARK}")
    print(f"Period: {args.period}")

    close = download_close_data(
        symbols,
        args.period,
    )

    results = analyse_relative_strength(
        symbols,
        close,
    )

    write_results(results)

    print("=" * 72)
    print(f"Stocks completed: {len(results)}")

    if results:
        leader = results[0]
        laggard = results[-1]

        print(
            f"Top leader: {leader['Symbol']} "
            f"({leader['Relative Strength Score']})"
        )

        print(
            f"Bottom laggard: {laggard['Symbol']} "
            f"({laggard['Relative Strength Score']})"
        )

    print(DASHBOARD)


if __name__ == "__main__":
    main()
