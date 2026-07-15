"""
AQSD Market Intelligence
Module: Macro & Policy Intelligence Engine v1.0

Stores and scores macroeconomic and policy events in aqsd_core.db.

Commands:
  python aqsd_macro_intelligence.py --setup
  python aqsd_macro_intelligence.py --import-csv C:\\Users\\megha\\AQSD\\Data\\macro_events.csv
  python aqsd_macro_intelligence.py --add-manual
  python aqsd_macro_intelligence.py --analyse
  python aqsd_macro_intelligence.py --report
  python aqsd_macro_intelligence.py --status
"""
from __future__ import annotations

import argparse
import hashlib
import math
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")

SCHEMA = """
CREATE TABLE IF NOT EXISTS macro_policy_events (
    macro_policy_id INTEGER PRIMARY KEY AUTOINCREMENT,
    event_date TEXT NOT NULL,
    country TEXT NOT NULL,
    indicator_name TEXT NOT NULL,
    event_type TEXT,
    actual_value REAL,
    expected_value REAL,
    previous_value REAL,
    unit TEXT,
    direction_rule TEXT,
    surprise_percent REAL,
    surprise_score REAL,
    materiality_score REAL,
    credibility_score REAL,
    recency_score REAL,
    macro_impact_score REAL,
    macro_bias TEXT,
    affected_sectors TEXT,
    policy_text TEXT,
    source TEXT NOT NULL,
    event_hash TEXT NOT NULL UNIQUE,
    created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_macro_policy_date
ON macro_policy_events(event_date);
"""

LOWER_IS_POSITIVE = {
    "CPI", "WPI", "INFLATION", "REPO RATE", "BOND YIELD",
    "FISCAL DEFICIT", "UNEMPLOYMENT"
}
HIGHER_IS_POSITIVE = {
    "GDP", "IIP", "PMI", "INDUSTRIAL PRODUCTION", "EXPORTS",
    "CREDIT GROWTH", "CAPEX", "PLI"
}
SECTOR_MAP = {
    "REPO RATE": ("Banks|NBFC|Auto|Realty", "Rate Sensitive Borrowers"),
    "CPI": ("Banks|NBFC|Auto|Realty", "High Input Cost Sectors"),
    "GDP": ("Banks|Capital Goods|Infrastructure|Auto", "Defensives"),
    "IIP": ("Capital Goods|Industrials|Metals|Power", "Defensives"),
    "PMI": ("Industrials|Capital Goods|Logistics|Banks", "Defensives"),
    "PLI": ("Manufacturing|Electronics|Pharma|Auto|Textiles|Solar", "Import Dependent Competitors"),
    "BUDGET": ("Policy Dependent", "Policy Dependent"),
}


def setup_schema() -> None:
    setup_database()
    with connect() as con:
        con.executescript(SCHEMA)
        con.commit()


def clean(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"nan", "none", "null"} else text


def safe_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_date(value: Any) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Invalid date: {value}")
    return parsed.date().isoformat()


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def event_hash(record: dict) -> str:
    raw = "|".join([
        record["event_date"], record["country"].lower(),
        record["indicator_name"].lower(), record["source"].lower()
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def infer_rule(name: str, explicit: str) -> str:
    explicit = explicit.upper().strip()
    if explicit in {"HIGHER_POSITIVE", "LOWER_POSITIVE", "TEXT_POLICY"}:
        return explicit
    upper = name.upper()
    if any(word in upper for word in LOWER_IS_POSITIVE):
        return "LOWER_POSITIVE"
    if any(word in upper for word in HIGHER_IS_POSITIVE):
        return "HIGHER_POSITIVE"
    if any(word in upper for word in {"BUDGET", "PLI", "POLICY", "DUTY"}):
        return "TEXT_POLICY"
    return "HIGHER_POSITIVE"


def surprise_percent(actual: float | None, expected: float | None) -> float | None:
    if actual is None or expected in (None, 0):
        return None
    return round((actual - expected) / abs(expected) * 100, 2)


def policy_text_score(text: str) -> float:
    positive = ["increase", "support", "incentive", "subsidy", "allocation", "capex", "approval", "duty cut"]
    negative = ["withdrawal", "penalty", "restriction", "ban", "tax increase", "duty increase"]
    score = sum(12 for word in positive if word in text.lower())
    score -= sum(12 for word in negative if word in text.lower())
    return clamp(score, -100, 100)


def recency_score(event_date: str) -> float:
    age = max(0, (date.today() - datetime.strptime(event_date, "%Y-%m-%d").date()).days)
    return round(100 * math.pow(0.5, age / 30), 2)


def bias(score: float) -> str:
    if score >= 60:
        return "STRONG POSITIVE"
    if score >= 20:
        return "POSITIVE"
    if score <= -60:
        return "STRONG NEGATIVE"
    if score <= -20:
        return "NEGATIVE"
    return "NEUTRAL"


def score_record(record: dict) -> dict:
    rule = infer_rule(record["indicator_name"], record.get("direction_rule", ""))
    raw = surprise_percent(record.get("actual_value"), record.get("expected_value"))
    surprise = 0.0
    if rule == "TEXT_POLICY":
        surprise = policy_text_score(record.get("policy_text", ""))
    elif raw is not None:
        surprise = clamp(raw * ( -4 if rule == "LOWER_POSITIVE" else 4), -100, 100)
    materiality = clamp(float(record.get("materiality_score") or 50), 0, 100)
    credibility = clamp(float(record.get("credibility_score") or 70), 0, 100)
    recency = recency_score(record["event_date"])
    magnitude = materiality * 0.45 + credibility * 0.30 + recency * 0.25
    impact = round(clamp((abs(surprise) / 100) * magnitude * (1 if surprise > 0 else -1 if surprise < 0 else 0), -100, 100), 2)
    sectors = clean(record.get("affected_sectors"))
    if not sectors:
        upper = record["indicator_name"].upper()
        sectors = "Broad Market"
        for key, mapping in SECTOR_MAP.items():
            if key in upper:
                sectors = mapping[0] if impact >= 0 else mapping[1]
                break
    return {
        "direction_rule": rule,
        "surprise_percent": raw,
        "surprise_score": round(surprise, 2),
        "recency_score": recency,
        "macro_impact_score": impact,
        "macro_bias": bias(impact),
        "affected_sectors": sectors,
    }


def insert_event(con, record: dict) -> bool:
    scores = score_record(record)
    try:
        con.execute("""
            INSERT INTO macro_policy_events(
                event_date,country,indicator_name,event_type,actual_value,
                expected_value,previous_value,unit,direction_rule,
                surprise_percent,surprise_score,materiality_score,
                credibility_score,recency_score,macro_impact_score,
                macro_bias,affected_sectors,policy_text,source,event_hash,created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            record["event_date"], record["country"], record["indicator_name"],
            record.get("event_type", ""), record.get("actual_value"),
            record.get("expected_value"), record.get("previous_value"),
            record.get("unit", ""), scores["direction_rule"],
            scores["surprise_percent"], scores["surprise_score"],
            record.get("materiality_score"), record.get("credibility_score"),
            scores["recency_score"], scores["macro_impact_score"],
            scores["macro_bias"], scores["affected_sectors"],
            record.get("policy_text", ""), record["source"], event_hash(record),
            datetime.now().isoformat(timespec="seconds")
        ))
        return True
    except Exception as exc:
        if "UNIQUE constraint failed" in str(exc):
            return False
        raise


def import_csv(path: Path) -> tuple[int, int]:
    if not path.exists():
        raise FileNotFoundError(path)
    frame = pd.read_csv(path)
    frame.columns = [str(c).strip().lower().replace(" ", "_") for c in frame.columns]
    required = {"event_date", "country", "indicator_name", "source"}
    if not required.issubset(frame.columns):
        raise RuntimeError("Missing required columns: " + ", ".join(sorted(required - set(frame.columns))))
    inserted = duplicates = 0
    run_id = start_run("aqsd_macro_intelligence", f"Importing {path.name}")
    try:
        with connect() as con:
            for _, row in frame.iterrows():
                record = {
                    "event_date": parse_date(row.get("event_date")),
                    "country": clean(row.get("country")),
                    "indicator_name": clean(row.get("indicator_name")),
                    "event_type": clean(row.get("event_type")),
                    "actual_value": safe_float(row.get("actual_value")),
                    "expected_value": safe_float(row.get("expected_value")),
                    "previous_value": safe_float(row.get("previous_value")),
                    "unit": clean(row.get("unit")),
                    "direction_rule": clean(row.get("direction_rule")),
                    "materiality_score": safe_float(row.get("materiality_score"), 50),
                    "credibility_score": safe_float(row.get("credibility_score"), 70),
                    "affected_sectors": clean(row.get("affected_sectors")),
                    "policy_text": clean(row.get("policy_text")),
                    "source": clean(row.get("source")),
                }
                if insert_event(con, record):
                    inserted += 1
                else:
                    duplicates += 1
            con.commit()
        finish_run(run_id, "SUCCESS", inserted, duplicates, f"Inserted={inserted}; duplicates={duplicates}")
        return inserted, duplicates
    except Exception as exc:
        finish_run(run_id, "FAILED", inserted, duplicates + 1, str(exc))
        raise


def add_manual() -> None:
    record = {
        "event_date": parse_date(input("Event date YYYY-MM-DD: ").strip()),
        "country": input("Country [India]: ").strip() or "India",
        "indicator_name": input("Indicator / policy name: ").strip(),
        "event_type": input("Event type: ").strip(),
        "actual_value": safe_float(input("Actual [optional]: ").strip()),
        "expected_value": safe_float(input("Expected [optional]: ").strip()),
        "previous_value": safe_float(input("Previous [optional]: ").strip()),
        "unit": input("Unit: ").strip(),
        "direction_rule": input("Rule HIGHER_POSITIVE/LOWER_POSITIVE/TEXT_POLICY: ").strip(),
        "materiality_score": safe_float(input("Materiality 0-100 [50]: ").strip(), 50),
        "credibility_score": safe_float(input("Credibility 0-100 [70]: ").strip(), 70),
        "affected_sectors": input("Affected sectors [optional]: ").strip(),
        "policy_text": input("Policy description [optional]: ").strip(),
        "source": input("Source: ").strip(),
    }
    with connect() as con:
        added = insert_event(con, record)
        con.commit()
    print("Event added." if added else "Duplicate event already exists.")


def recent_events(days: int = 180) -> pd.DataFrame:
    cutoff = (date.today() - timedelta(days=max(days, 1))).isoformat()
    with connect() as con:
        return pd.read_sql_query(
            "SELECT * FROM macro_policy_events WHERE event_date >= ? ORDER BY event_date DESC, macro_impact_score DESC",
            con, params=(cutoff,)
        )


def macro_summary(days: int = 90) -> dict:
    frame = recent_events(days)
    if frame.empty:
        return {"score": 0.0, "bias": "NO DATA", "events": 0}
    weights = frame["materiality_score"].fillna(50) * frame["credibility_score"].fillna(70)
    score = 0.0 if weights.sum() == 0 else float((frame["macro_impact_score"] * weights).sum() / weights.sum())
    score = round(clamp(score, -100, 100), 2)
    return {"score": score, "bias": bias(score), "events": len(frame)}


def sector_scores(days: int = 90) -> pd.DataFrame:
    frame = recent_events(days)
    records = []
    for _, row in frame.iterrows():
        for sector in clean(row.get("affected_sectors")).split("|"):
            if sector.strip():
                records.append({"Sector": sector.strip(), "Impact": float(row.get("macro_impact_score") or 0), "Event Date": row.get("event_date")})
    if not records:
        return pd.DataFrame()
    result = pd.DataFrame(records).groupby("Sector").agg(Event_Count=("Impact", "count"), Macro_Score=("Impact", "mean"), Latest_Event=("Event Date", "max")).reset_index()
    result["Macro_Score"] = result["Macro_Score"].round(2)
    result["Macro_Bias"] = result["Macro_Score"].apply(bias)
    return result.sort_values("Macro_Score", ascending=False)


def write_report(days: int = 180) -> None:
    events = recent_events(days)
    sectors = sector_scores(min(days, 90))
    summary = macro_summary(min(days, 90))
    wb = load_workbook(DASHBOARD) if DASHBOARD.exists() else Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]
    if "Macro Intelligence" in wb.sheetnames:
        del wb["Macro Intelligence"]
    ws = wb.create_sheet("Macro Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"
    ws.merge_cells("A1:S2")
    ws["A1"] = "AQSD PROFESSIONAL - MACRO & POLICY INTELLIGENCE"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A4"], ws["B4"] = "Macro Score", summary["score"]
    ws["D4"], ws["E4"] = "Macro Bias", summary["bias"]
    ws["G4"], ws["H4"] = "Events", summary["events"]
    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill("solid", fgColor=BLUE)
    headers = ["Rank", "Sector", "Event Count", "Macro Score", "Macro Bias", "Latest Event"]
    for col, heading in enumerate(headers, 1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    for row_no, (_, row) in enumerate(sectors.iterrows(), 8):
        vals = [row_no - 7, row["Sector"], row["Event_Count"], row["Macro_Score"], row["Macro_Bias"], row["Latest_Event"]]
        for col, value in enumerate(vals, 1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)
        score = float(row["Macro_Score"])
        ws.cell(row_no, 4).fill = PatternFill("solid", fgColor=GREEN if score >= 20 else RED if score <= -20 else YELLOW)
    start = max(12, 10 + len(sectors))
    detail_headers = ["Event Date", "Country", "Indicator", "Event Type", "Actual", "Expected", "Previous", "Unit", "Rule", "Surprise %", "Surprise Score", "Materiality", "Credibility", "Recency", "Macro Impact", "Bias", "Affected Sectors", "Policy Text", "Source"]
    for col, heading in enumerate(detail_headers, 1):
        cell = ws.cell(start, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_no, (_, row) in enumerate(events.iterrows(), start + 1):
        vals = [row.get("event_date"), row.get("country"), row.get("indicator_name"), row.get("event_type"), row.get("actual_value"), row.get("expected_value"), row.get("previous_value"), row.get("unit"), row.get("direction_rule"), row.get("surprise_percent"), row.get("surprise_score"), row.get("materiality_score"), row.get("credibility_score"), row.get("recency_score"), row.get("macro_impact_score"), row.get("macro_bias"), row.get("affected_sectors"), row.get("policy_text"), row.get("source")]
        for col, value in enumerate(vals, 1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)
    widths = {"A":14,"B":16,"C":28,"D":18,"E":12,"F":12,"G":12,"H":12,"I":18,"J":13,"K":14,"L":13,"M":13,"N":12,"O":14,"P":18,"Q":40,"R":60,"S":24}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    wb.save(DASHBOARD)


def show_status() -> None:
    setup_schema()
    with connect() as con:
        row = con.execute("SELECT COUNT(*) total, COUNT(DISTINCT indicator_name) indicators, MIN(event_date) first_date, MAX(event_date) latest_date FROM macro_policy_events").fetchone()
    summary = macro_summary(90)
    print("\nAQSD MACRO INTELLIGENCE STATUS")
    print("=" * 72)
    print(f"Stored events:      {row['total'] or 0}")
    print(f"Indicators covered: {row['indicators'] or 0}")
    print(f"First event date:   {row['first_date'] or 'No data'}")
    print(f"Latest event date:  {row['latest_date'] or 'No data'}")
    print(f"90-day score:       {summary['score']}")
    print(f"90-day bias:        {summary['bias']}")


def main() -> None:
    parser = argparse.ArgumentParser(description="AQSD Macro & Policy Intelligence Engine")
    parser.add_argument("--setup", action="store_true")
    parser.add_argument("--import-csv", type=Path)
    parser.add_argument("--add-manual", action="store_true")
    parser.add_argument("--analyse", action="store_true")
    parser.add_argument("--report", action="store_true")
    parser.add_argument("--days", type=int, default=180)
    parser.add_argument("--status", action="store_true")
    args = parser.parse_args()
    setup_schema()
    if args.setup:
        print("AQSD Macro Intelligence schema is ready.")
    elif args.import_csv:
        inserted, duplicates = import_csv(args.import_csv)
        print(f"Inserted: {inserted}\nDuplicates: {duplicates}")
    elif args.add_manual:
        add_manual()
    elif args.analyse:
        print(macro_summary(90))
        frame = sector_scores(90)
        print(frame.to_string(index=False) if not frame.empty else "No sector macro scores available.")
    elif args.report:
        write_report(args.days)
        print(f"Macro Intelligence report created:\n{DASHBOARD}")
    else:
        show_status()


if __name__ == "__main__":
    main()
