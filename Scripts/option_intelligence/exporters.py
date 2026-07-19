"""
AQSD
Option Intelligence Export Framework

Module: exporters.py
Version: 1.0
Author: AQSD

Description:
Provides standardized export functions for AQSD analytics engines.

Supported outputs:
- Summary CSV
- Table CSV
- History CSV
- Excel workbook
- JSON
- Run metadata
- EngineResult standardized object

The module can be used by:
- OI Engine
- PCR Engine
- Max Pain Engine
- Wall Engine
- Volatility Engine
- Probability Engine
- Option Intelligence Dashboard
"""

from __future__ import annotations

import json
import math
import re
from dataclasses import asdict, dataclass, field, is_dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

BASE_DIR = Path(__file__).resolve().parents[2]

try:
    from Scripts.option_intelligence import config as aqsd_config

    OUTPUT_DIR = Path(
        getattr(
            aqsd_config,
            "OUTPUT_DIR",
            BASE_DIR / "Output",
        )
    )

except (ImportError, AttributeError):
    OUTPUT_DIR = BASE_DIR / "Output"


DEFAULT_EXCEL_ENGINE = "openpyxl"

SUPPORTED_MODULES = {
    "OI": "OI",
    "OPEN_INTEREST": "OI",
    "PCR": "PCR",
    "MAX_PAIN": "MaxPain",
    "MAXPAIN": "MaxPain",
    "WALL": "Walls",
    "WALLS": "Walls",
    "VOLATILITY": "Volatility",
    "PROBABILITY": "Probability",
    "DASHBOARD": "Dashboard",
    "OPTIONS": "Options",
    "OPTION_INTELLIGENCE": "Options",
    "GENERAL": "General",
}

INVALID_FILENAME_CHARACTERS = re.compile(r'[<>:"/\\|?*\x00-\x1F]')


# ============================================================
# DATA MODELS
# ============================================================

@dataclass(slots=True)
class ExportMetadata:
    """
    Standard metadata attached to every AQSD export.
    """

    engine: str
    underlying: str
    timestamp: str = field(
        default_factory=lambda: datetime.now().astimezone().isoformat(
            timespec="seconds"
        )
    )
    expiry: str | None = None
    engine_version: str = "1.0"
    rows_processed: int | None = None
    status: str = "SUCCESS"
    source: str | None = None
    notes: str | None = None

    def to_dict(self) -> dict[str, Any]:
        """
        Convert metadata into a dictionary.
        """

        return {
            key: value
            for key, value in asdict(self).items()
            if value is not None
        }


@dataclass(slots=True)
class EngineResult:
    """
    Standard result object returned by AQSD analytics engines.

    Attributes:
        summary:
            Summary analytics in dataclass, dictionary, Series,
            DataFrame, or compatible object form.

        table:
            Main detailed analytics table.

        history:
            Optional history row or history DataFrame.

        metadata:
            Export metadata.

        extra_tables:
            Additional Excel or CSV tables.

        json_data:
            Optional JSON-specific data. If omitted, summary and
            table data are used automatically.
    """

    summary: Any
    table: pd.DataFrame | None
    metadata: ExportMetadata | Mapping[str, Any]

    history: Any | None = None

    extra_tables: dict[str, pd.DataFrame] = field(
        default_factory=dict
    )

    json_data: Any | None = None


@dataclass(slots=True)
class ExportPaths:
    """
    Paths created during an export operation.
    """

    module_directory: Path

    summary_csv: Path | None = None
    table_csv: Path | None = None
    history_csv: Path | None = None
    excel: Path | None = None
    json: Path | None = None

    extra_csv_files: dict[str, Path] = field(
        default_factory=dict
    )

    errors: list[str] = field(
        default_factory=list
    )

    def to_dict(self) -> dict[str, Any]:
        """
        Convert export paths to a dictionary.
        """

        result: dict[str, Any] = {
            "module_directory": self.module_directory,
        }

        if self.summary_csv is not None:
            result["summary_csv"] = self.summary_csv

        if self.table_csv is not None:
            result["table_csv"] = self.table_csv

        if self.history_csv is not None:
            result["history_csv"] = self.history_csv

        if self.excel is not None:
            result["excel"] = self.excel

        if self.json is not None:
            result["json"] = self.json

        if self.extra_csv_files:
            result["extra_csv_files"] = self.extra_csv_files

        if self.errors:
            result["errors"] = self.errors

        return result


# ============================================================
# GENERIC CONVERSION HELPERS
# ============================================================

def is_missing_value(value: Any) -> bool:
    """
    Check whether a value should be treated as missing.
    """

    if value is None:
        return True

    try:
        missing = pd.isna(value)

        if isinstance(missing, bool):
            return missing

    except (TypeError, ValueError):
        return False

    return False


def make_json_safe(value: Any) -> Any:
    """
    Convert Python, pandas, NumPy, date and Path values into
    JSON-compatible values.
    """

    if value is None:
        return None

    if isinstance(value, Path):
        return str(value)

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, pd.Timedelta):
        return str(value)

    if is_dataclass(value) and not isinstance(value, type):
        return make_json_safe(asdict(value))

    if isinstance(value, pd.DataFrame):
        records = value.to_dict(orient="records")
        return make_json_safe(records)

    if isinstance(value, pd.Series):
        return make_json_safe(value.to_dict())

    if isinstance(value, Mapping):
        return {
            str(key): make_json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, (list, tuple, set)):
        return [
            make_json_safe(item)
            for item in value
        ]

    item_method = getattr(value, "item", None)

    if callable(item_method):
        try:
            return make_json_safe(item_method())
        except (ValueError, TypeError):
            pass
    
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None

        return value

    if is_missing_value(value):
        return None

    return value
def object_to_dictionary(
    value: Any,
) -> dict[str, Any]:
    """
    Convert a supported summary or metadata object into a dictionary.
    """

    if value is None:
        return {}

    if isinstance(value, ExportMetadata):
        return value.to_dict()

    if is_dataclass(value) and not isinstance(value, type):
        dataclass_values = asdict(value)

        return {
            str(key): item
            for key, item in dataclass_values.items()
        }

    if isinstance(value, Mapping):
        return {
            str(key): item
            for key, item in value.items()
        }

    if isinstance(value, pd.Series):
        series_values = value.to_dict()

        return {
            str(key): item
            for key, item in series_values.items()
        }

    if isinstance(value, pd.DataFrame):
        if value.empty:
            return {}

        if {
            "metric",
            "value",
        }.issubset(value.columns):
            return {
                str(metric): item
                for metric, item in zip(
                    value["metric"],
                    value["value"],
                )
            }

        if len(value) == 1:
            row_values = value.iloc[0].to_dict()

            return {
                str(key): item
                for key, item in row_values.items()
            }

        raise ValueError(
            "A multi-row DataFrame cannot be converted into a "
            "single summary dictionary."
        )

    object_values = getattr(
        value,
        "__dict__",
        None,
    )

    if isinstance(object_values, Mapping):
        return {
            str(key): item
            for key, item in object_values.items()
            if not str(key).startswith("_")
        }

    raise TypeError(
        "Unsupported object type for dictionary conversion: "
        f"{type(value).__name__}"
    )

def object_to_dictionary(value: Any) -> dict[str, Any]:
    """
    Convert a supported summary or metadata object into a dictionary.
    """

    if value is None:
        return {}

    if isinstance(value, ExportMetadata):
        return value.to_dict()

    if is_dataclass(value) and not isinstance(value, type):
        return asdict(value)

    if isinstance(value, Mapping):
        return dict(value)

    if isinstance(value, pd.Series):
        return value.to_dict()

    if isinstance(value, pd.DataFrame):
        if value.empty:
            return {}

        if {
            "metric",
            "value",
        }.issubset(value.columns):
            return dict(
                zip(
                    value["metric"].astype(str),
                    value["value"],
                )
            )

        if len(value) == 1:
            return value.iloc[0].to_dict()

        raise ValueError(
            "A multi-row DataFrame cannot be converted into a "
            "single summary dictionary."
        )

    if hasattr(value, "__dict__"):
        return {
            key: item
            for key, item in vars(value).items()
            if not key.startswith("_")
        }

    raise TypeError(
        "Unsupported object type for dictionary conversion: "
        f"{type(value).__name__}"
    )


def summary_to_dataframe(summary: Any) -> pd.DataFrame:
    """
    Convert a summary object into a two-column DataFrame.

    Output columns:
        metric
        value
    """

    if summary is None:
        return pd.DataFrame(
            columns=["metric", "value"]
        )

    if isinstance(summary, pd.DataFrame):
        dataframe = summary.copy()

        if {
            "metric",
            "value",
        }.issubset(dataframe.columns):
            return dataframe[
                ["metric", "value"]
            ].copy()

        if len(dataframe) == 1:
            summary_dictionary = dataframe.iloc[0].to_dict()

            return pd.DataFrame(
                {
                    "metric": summary_dictionary.keys(),
                    "value": summary_dictionary.values(),
                }
            )

        return dataframe

    summary_dictionary = object_to_dictionary(summary)

    return pd.DataFrame(
        {
            "metric": summary_dictionary.keys(),
            "value": summary_dictionary.values(),
        }
    )


def history_to_dataframe(
    history: Any,
    metadata: Any | None = None,
) -> pd.DataFrame:
    """
    Convert history data into a row-oriented DataFrame.

    Summary-style history is converted into one row.
    """

    if history is None:
        return pd.DataFrame()

    if isinstance(history, pd.DataFrame):
        dataframe = history.copy()

    elif isinstance(history, pd.Series):
        dataframe = history.to_frame().T

    elif is_dataclass(history) and not isinstance(history, type):
        dataframe = pd.DataFrame(
            [asdict(history)]
        )

    elif isinstance(history, Mapping):
        dataframe = pd.DataFrame(
            [dict(history)]
        )

    elif isinstance(history, Sequence) and not isinstance(
        history,
        (str, bytes),
    ):
        dataframe = pd.DataFrame(history)

    else:
        dataframe = pd.DataFrame(
            [object_to_dictionary(history)]
        )

    if metadata is not None:
        metadata_dictionary = object_to_dictionary(metadata)

        for column_name, value in reversed(
            list(metadata_dictionary.items())
        ):
            if column_name not in dataframe.columns:
                dataframe.insert(
                    0,
                    column_name,
                    value,
                )

    return dataframe


def ensure_dataframe(
    value: Any,
    name: str,
) -> pd.DataFrame:
    """
    Convert a value into a DataFrame where possible.
    """

    if value is None:
        return pd.DataFrame()

    if isinstance(value, pd.DataFrame):
        return value.copy()

    if isinstance(value, pd.Series):
        return value.to_frame().T

    if is_dataclass(value) and not isinstance(value, type):
        return pd.DataFrame(
            [asdict(value)]
        )

    if isinstance(value, Mapping):
        try:
            return pd.DataFrame(value)
        except ValueError:
            return pd.DataFrame([dict(value)])

    if isinstance(value, Sequence) and not isinstance(
        value,
        (str, bytes),
    ):
        return pd.DataFrame(value)

    raise TypeError(
        f"{name} could not be converted to a DataFrame. "
        f"Received type: {type(value).__name__}"
    )


# ============================================================
# FILE AND FOLDER HELPERS
# ============================================================

def sanitize_filename(
    value: str,
    fallback: str = "AQSD",
) -> str:
    """
    Convert text into a safe Windows filename component.
    """

    text = str(value).strip()

    text = INVALID_FILENAME_CHARACTERS.sub(
        "_",
        text,
    )

    text = re.sub(
        r"\s+",
        "_",
        text,
    )

    text = re.sub(
        r"_+",
        "_",
        text,
    )

    text = text.strip(" ._")

    if not text:
        return fallback

    return text


def sanitize_sheet_name(
    value: str,
    fallback: str = "Sheet",
) -> str:
    """
    Create a valid Excel worksheet name.
    """

    text = str(value).strip()

    text = re.sub(
        r"[\[\]:*?/\\]",
        "_",
        text,
    )

    text = text.strip("'")

    if not text:
        text = fallback

    return text[:31]


def normalize_module_name(
    module_name: str,
) -> str:
    """
    Convert an engine name into the standard output folder name.
    """

    normalized = (
        str(module_name)
        .strip()
        .upper()
        .replace("-", "_")
        .replace(" ", "_")
    )

    return SUPPORTED_MODULES.get(
        normalized,
        sanitize_filename(
            module_name,
            fallback="General",
        ),
    )


def create_module_directory(
    module_name: str,
    output_root: Path | str = OUTPUT_DIR,
) -> Path:
    """
    Create and return the output directory for an AQSD module.
    """

    root = Path(output_root)

    module_folder = normalize_module_name(
        module_name
    )

    directory = root / module_folder

    directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    return directory


def build_base_filename(
    underlying: str,
    engine: str,
    suffix: str | None = None,
) -> str:
    """
    Build a consistent AQSD base filename.
    """

    safe_underlying = sanitize_filename(
        underlying.upper(),
        fallback="UNKNOWN",
    )

    safe_engine = sanitize_filename(
        engine,
        fallback="ENGINE",
    )

    parts = [
        safe_underlying,
        safe_engine,
    ]

    if suffix:
        parts.append(
            sanitize_filename(
                suffix,
                fallback="OUTPUT",
            )
        )

    return "_".join(parts)


# ============================================================
# CSV EXPORT FUNCTIONS
# ============================================================

def save_summary_csv(
    summary: Any,
    file_path: Path | str,
) -> Path:
    """
    Save summary analytics as CSV.
    """

    path = Path(file_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    dataframe = summary_to_dataframe(summary)

    dataframe.to_csv(
        path,
        index=False,
    )

    return path


def save_table_csv(
    table: Any,
    file_path: Path | str,
) -> Path:
    """
    Save a detailed analytics table as CSV.
    """

    path = Path(file_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    dataframe = ensure_dataframe(
        table,
        name="table",
    )

    dataframe.to_csv(
        path,
        index=False,
    )

    return path


def append_history_csv(
    history: Any,
    file_path: Path | str,
    metadata: Any | None = None,
    remove_exact_duplicates: bool = True,
) -> Path:
    """
    Append analytics history to an existing CSV.

    Exact duplicate rows can optionally be removed.
    """

    path = Path(file_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    new_history = history_to_dataframe(
        history=history,
        metadata=metadata,
    )

    if new_history.empty:
        return path

    if path.exists():
        try:
            existing_history = pd.read_csv(path)

        except (
            pd.errors.EmptyDataError,
            UnicodeDecodeError,
        ):
            existing_history = pd.DataFrame()

        combined_history = pd.concat(
            [
                existing_history,
                new_history,
            ],
            ignore_index=True,
            sort=False,
        )

    else:
        combined_history = new_history

    if remove_exact_duplicates:
        combined_history = combined_history.drop_duplicates(
            keep="last"
        )

    combined_history.to_csv(
        path,
        index=False,
    )

    return path


# ============================================================
# JSON EXPORT
# ============================================================

def save_json(
    data: Any,
    file_path: Path | str,
    indent: int = 4,
) -> Path:
    """
    Save JSON-compatible AQSD output.
    """

    path = Path(file_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    safe_data = make_json_safe(data)

    with path.open(
        mode="w",
        encoding="utf-8",
    ) as json_file:
        json.dump(
            safe_data,
            json_file,
            indent=indent,
            ensure_ascii=False,
            allow_nan=False,
        )

    return path


# ============================================================
# EXCEL FORMATTING
# ============================================================

def style_excel_header(
    worksheet: Any,
) -> None:
    """
    Apply standard formatting to the first row of an Excel sheet.
    """

    if worksheet.max_row < 1:
        return

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def auto_fit_excel_columns(
    worksheet: Any,
    minimum_width: int = 10,
    maximum_width: int = 50,
) -> None:
    """
    Automatically size Excel worksheet columns.
    """

    for column_cells in worksheet.columns:
        maximum_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            length = len(str(value))

            maximum_length = max(
                maximum_length,
                length,
            )

        adjusted_width = min(
            max(
                maximum_length + 2,
                minimum_width,
            ),
            maximum_width,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = adjusted_width


def apply_excel_number_formats(
    worksheet: Any,
) -> None:
    """
    Apply basic number formatting based on column names.
    """

    if worksheet.max_row < 2:
        return

    headers = {
        cell.column: str(cell.value).strip().lower()
        for cell in worksheet[1]
        if cell.value is not None
    }

    for column_index, header in headers.items():
        number_format = None

        if "percent" in header or header.endswith("_pct"):
            number_format = "0.00%"

        elif any(
            term in header
            for term in [
                "price",
                "strike",
                "ltp",
                "iv",
                "hv",
                "pcr",
                "score",
                "probability",
            ]
        ):
            number_format = "0.00"

        elif any(
            term in header
            for term in [
                "open_interest",
                "change_oi",
                "change_in_oi",
                "volume",
                "_oi",
            ]
        ):
            number_format = "#,##0"

        if number_format is None:
            continue

        for row_index in range(
            2,
            worksheet.max_row + 1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
            )

            if isinstance(
                cell.value,
                (int, float),
            ):
                cell.number_format = number_format


def freeze_excel_header(
    worksheet: Any,
) -> None:
    """
    Freeze the top row of a worksheet.
    """

    if worksheet.max_row > 1:
        worksheet.freeze_panes = "A2"


def format_excel_workbook(
    writer: pd.ExcelWriter,
) -> None:
    """
    Apply standard AQSD formatting to all workbook sheets.
    """

    workbook = writer.book

    for worksheet in workbook.worksheets:
        style_excel_header(worksheet)
        auto_fit_excel_columns(worksheet)
        apply_excel_number_formats(worksheet)
        freeze_excel_header(worksheet)

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


def save_excel(
    file_path: Path | str,
    sheets: Mapping[str, Any],
    engine: str = DEFAULT_EXCEL_ENGINE,
) -> Path:
    """
    Save multiple tables into a formatted Excel workbook.
    """

    path = Path(file_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    valid_sheets: dict[str, pd.DataFrame] = {}

    used_sheet_names: set[str] = set()

    for requested_name, value in sheets.items():
        if value is None:
            continue

        dataframe = ensure_dataframe(
            value,
            name=requested_name,
        )

        sheet_name = sanitize_sheet_name(
            requested_name
        )

        original_sheet_name = sheet_name
        counter = 2

        while sheet_name in used_sheet_names:
            suffix = f"_{counter}"

            sheet_name = (
                original_sheet_name[
                    : 31 - len(suffix)
                ]
                + suffix
            )

            counter += 1

        used_sheet_names.add(sheet_name)
        valid_sheets[sheet_name] = dataframe

    if not valid_sheets:
        valid_sheets["Summary"] = pd.DataFrame(
            {
                "message": [
                    "No exportable data was supplied."
                ]
            }
        )

    with pd.ExcelWriter(
        path,
        engine=engine,
    ) as writer:
        for sheet_name, dataframe in valid_sheets.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

        format_excel_workbook(writer)

    return path


# ============================================================
# RESULT VALIDATION
# ============================================================

def validate_engine_result(
    engine_result: EngineResult,
) -> None:
    """
    Validate an EngineResult before exporting.
    """

    if not isinstance(
        engine_result,
        EngineResult,
    ):
        raise TypeError(
            "engine_result must be an EngineResult object."
        )

    metadata = object_to_dictionary(
        engine_result.metadata
    )

    required_metadata = [
        "engine",
        "underlying",
    ]

    missing_fields = [
        field_name
        for field_name in required_metadata
        if not str(
            metadata.get(
                field_name,
                "",
            )
        ).strip()
    ]

    if missing_fields:
        raise ValueError(
            "Missing required metadata fields: "
            + ", ".join(missing_fields)
        )

    if engine_result.summary is None:
        raise ValueError(
            "EngineResult.summary cannot be None."
        )

    if engine_result.table is not None:
        ensure_dataframe(
            engine_result.table,
            name="EngineResult.table",
        )

    for table_name, table in engine_result.extra_tables.items():
        ensure_dataframe(
            table,
            name=f"extra_tables[{table_name}]",
        )


# ============================================================
# MAIN EXPORT ORCHESTRATOR
# ============================================================

def export_results(
    engine_result: EngineResult,
    output_root: Path | str = OUTPUT_DIR,
    base_filename: str | None = None,
    save_summary: bool = True,
    save_table: bool = True,
    save_history: bool = True,
    save_excel_file: bool = True,
    save_json_file: bool = True,
    save_extra_csv: bool = True,
    continue_on_error: bool = False,
) -> ExportPaths:
    """
    Export a standardized AQSD EngineResult.

    Returns:
        ExportPaths containing every generated file path.
    """

    validate_engine_result(engine_result)

    metadata_dictionary = object_to_dictionary(
        engine_result.metadata
    )

    engine_name = str(
        metadata_dictionary["engine"]
    )

    underlying = str(
        metadata_dictionary["underlying"]
    )

    module_directory = create_module_directory(
        module_name=engine_name,
        output_root=output_root,
    )

    if base_filename is None:
        base_filename = build_base_filename(
            underlying=underlying,
            engine=engine_name,
        )

    else:
        base_filename = sanitize_filename(
            base_filename,
            fallback="AQSD_OUTPUT",
        )

    export_paths = ExportPaths(
        module_directory=module_directory
    )

    summary_dataframe = summary_to_dataframe(
        engine_result.summary
    )

    metadata_dataframe = summary_to_dataframe(
        metadata_dictionary
    )

    table_dataframe = (
        ensure_dataframe(
            engine_result.table,
            name="table",
        )
        if engine_result.table is not None
        else pd.DataFrame()
    )

    def execute_export(
        description: str,
        export_function: Any,
    ) -> Any:
        try:
            return export_function()

        except Exception as error:
            message = (
                f"{description} failed: "
                f"{type(error).__name__}: {error}"
            )

            export_paths.errors.append(message)

            if not continue_on_error:
                raise

            return None

    if save_summary:
        summary_path = (
            module_directory
            / f"{base_filename}_Summary.csv"
        )

        export_paths.summary_csv = execute_export(
            "Summary CSV export",
            lambda: save_summary_csv(
                summary=summary_dataframe,
                file_path=summary_path,
            ),
        )

    if save_table and not table_dataframe.empty:
        table_path = (
            module_directory
            / f"{base_filename}_Table.csv"
        )

        export_paths.table_csv = execute_export(
            "Table CSV export",
            lambda: save_table_csv(
                table=table_dataframe,
                file_path=table_path,
            ),
        )

    history_source = engine_result.history

    if history_source is None:
        history_source = object_to_dictionary(
            engine_result.summary
        )

    if save_history:
        history_path = (
            module_directory
            / f"{base_filename}_History.csv"
        )

        export_paths.history_csv = execute_export(
            "History CSV export",
            lambda: append_history_csv(
                history=history_source,
                metadata=metadata_dictionary,
                file_path=history_path,
            ),
        )

    if save_extra_csv:
        for table_name, table_value in (
            engine_result.extra_tables.items()
        ):
            safe_table_name = sanitize_filename(
                table_name,
                fallback="Extra",
            )

            extra_path = (
                module_directory
                / (
                    f"{base_filename}_"
                    f"{safe_table_name}.csv"
                )
            )

            exported_path = execute_export(
                f"Extra CSV export: {table_name}",
                lambda value=table_value, path=extra_path: (
                    save_table_csv(
                        table=value,
                        file_path=path,
                    )
                ),
            )

            if exported_path is not None:
                export_paths.extra_csv_files[
                    table_name
                ] = exported_path

    if save_excel_file:
        excel_path = (
            module_directory
            / f"{base_filename}.xlsx"
        )

        excel_sheets: dict[str, Any] = {
            "Metadata": metadata_dataframe,
            "Summary": summary_dataframe,
        }

        if not table_dataframe.empty:
            excel_sheets["Table"] = table_dataframe

        if engine_result.history is not None:
            excel_sheets["Current History Row"] = (
                history_to_dataframe(
                    engine_result.history,
                    metadata=metadata_dictionary,
                )
            )

        for table_name, table_value in (
            engine_result.extra_tables.items()
        ):
            excel_sheets[table_name] = table_value

        export_paths.excel = execute_export(
            "Excel export",
            lambda: save_excel(
                file_path=excel_path,
                sheets=excel_sheets,
            ),
        )

    if save_json_file:
        json_path = (
            module_directory
            / f"{base_filename}.json"
        )

        if engine_result.json_data is not None:
            analytics_json = engine_result.json_data

        else:
            analytics_json = {
                "summary": object_to_dictionary(
                    engine_result.summary
                ),
                "table": (
                    table_dataframe.to_dict(
                        orient="records"
                    )
                    if not table_dataframe.empty
                    else []
                ),
                "extra_tables": {
                    table_name: ensure_dataframe(
                        table_value,
                        name=table_name,
                    ).to_dict(
                        orient="records"
                    )
                    for table_name, table_value
                    in engine_result.extra_tables.items()
                },
            }

        json_payload = {
            "metadata": metadata_dictionary,
            "analytics": analytics_json,
        }

        export_paths.json = execute_export(
            "JSON export",
            lambda: save_json(
                data=json_payload,
                file_path=json_path,
            ),
        )

    return export_paths


# ============================================================
# TERMINAL OUTPUT
# ============================================================

def print_export_report(
    export_paths: ExportPaths,
) -> None:
    """
    Print created export files in the terminal.
    """

    separator = "=" * 76

    print()
    print(separator)
    print("AQSD EXPORT REPORT")
    print(separator)

    print(
        f"Output Folder : "
        f"{export_paths.module_directory}"
    )

    path_dictionary = export_paths.to_dict()

    for name in [
        "summary_csv",
        "table_csv",
        "history_csv",
        "excel",
        "json",
    ]:
        path = path_dictionary.get(name)

        if path is not None:
            readable_name = (
                name.replace("_", " ").title()
            )

            print(
                f"{readable_name:<16}: {path}"
            )

    if export_paths.extra_csv_files:
        print()
        print("Additional CSV Files")
        print("-" * 76)

        for name, path in (
            export_paths.extra_csv_files.items()
        ):
            print(
                f"{name:<16}: {path}"
            )

    if export_paths.errors:
        print()
        print("Export Errors")
        print("-" * 76)

        for error in export_paths.errors:
            print(f"- {error}")

    else:
        print()
        print("Status          : SUCCESS")

    print(separator)
    print()


# ============================================================
# SAMPLE TEST DATA
# ============================================================

@dataclass(slots=True)
class SampleOISummary:
    """
    Sample summary used to test exporters.py independently.
    """

    total_call_oi: float
    total_put_oi: float
    oi_pcr: float
    positional_call_wall: float
    positional_put_wall: float
    market_bias: str


def create_sample_summary() -> SampleOISummary:
    """
    Create a sample OI summary.
    """

    return SampleOISummary(
        total_call_oi=1_725_000,
        total_put_oi=1_120_000,
        oi_pcr=0.649275,
        positional_call_wall=59_000,
        positional_put_wall=58_000,
        market_bias="BEARISH",
    )


def create_sample_table() -> pd.DataFrame:
    """
    Create a sample strike-wise OI table.
    """

    return pd.DataFrame(
        [
            {
                "strike": 57000,
                "call_oi": 125000,
                "put_oi": 185000,
                "call_change_oi": 15000,
                "put_change_oi": 24000,
                "strike_oi_pcr": 1.480,
            },
            {
                "strike": 57500,
                "call_oi": 210000,
                "put_oi": 260000,
                "call_change_oi": 38000,
                "put_change_oi": 42000,
                "strike_oi_pcr": 1.238,
            },
            {
                "strike": 58000,
                "call_oi": 395000,
                "put_oi": 340000,
                "call_change_oi": 72000,
                "put_change_oi": 51000,
                "strike_oi_pcr": 0.861,
            },
            {
                "strike": 58500,
                "call_oi": 470000,
                "put_oi": 190000,
                "call_change_oi": 93000,
                "put_change_oi": -18000,
                "strike_oi_pcr": 0.404,
            },
            {
                "strike": 59000,
                "call_oi": 525000,
                "put_oi": 145000,
                "call_change_oi": 65000,
                "put_change_oi": -26000,
                "strike_oi_pcr": 0.276,
            },
        ]
    )


def create_sample_extra_table() -> pd.DataFrame:
    """
    Create a sample top-wall table.
    """

    return pd.DataFrame(
        [
            {
                "wall_type": "POSITIONAL CALL",
                "strike": 59000,
                "value": 525000,
            },
            {
                "wall_type": "POSITIONAL PUT",
                "strike": 58000,
                "value": 340000,
            },
            {
                "wall_type": "FRESH CALL",
                "strike": 58500,
                "value": 93000,
            },
            {
                "wall_type": "FRESH PUT",
                "strike": 58000,
                "value": 51000,
            },
        ]
    )


# ============================================================
# INDEPENDENT MODULE TEST
# ============================================================

def main() -> None:
    """
    Test the exporter independently.
    """

    sample_summary = create_sample_summary()
    sample_table = create_sample_table()
    sample_walls = create_sample_extra_table()

    metadata = ExportMetadata(
        engine="OI",
        underlying="BANKNIFTY_SAMPLE",
        expiry="2026-07-30",
        engine_version="1.0",
        rows_processed=len(sample_table),
        status="SUCCESS",
        source="AQSD Sample Data",
        notes="Independent exporters.py module test.",
    )

    result = EngineResult(
        summary=sample_summary,
        table=sample_table,
        history=asdict(sample_summary),
        metadata=metadata,
        extra_tables={
            "Walls": sample_walls,
        },
    )

    export_paths = export_results(
        engine_result=result,
        base_filename="BANKNIFTY_SAMPLE_OI_Intelligence",
    )

    print_export_report(export_paths)


if __name__ == "__main__":
    main()