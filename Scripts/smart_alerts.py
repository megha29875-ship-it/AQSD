
"""
AQSD Professional
Module: Smart Alert Engine
Version: 1.0

Creates an Alerts sheet from:
- Live Watchlist
- Portfolio

Alerts include:
- Near Entry
- Target Hit
- Stop Zone
- Near Stop
- Near Target
- High Portfolio Risk
- High Position Concentration
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

PORTFOLIO_RISK_LIMIT_PERCENT = 5.0
POSITION_CAPITAL_LIMIT_PERCENT = 25.0

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
GREY = "E7E6E6"
THIN = Side(style="thin", color="D9D9D9")


def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def add_alert(
    alerts: list[dict],
    *,
    source: str,
    symbol: str,
    alert_type: str,
    priority: str,
    message: str,
) -> None:
    alerts.append(
        {
            "Source": source,
            "Symbol": symbol,
            "Alert": alert_type,
            "Priority": priority,
            "Message": message,
        }
    )


def read_watchlist_alerts(wb, alerts: list[dict]) -> None:
    if "Live Watchlist" not in wb.sheetnames:
        return

    ws = wb["Live Watchlist"]

    for row in range(1, ws.max_row + 1):
        symbol = str(ws.cell(row, 2).value or "").strip()
        status = str(ws.cell(row, 15).value or "").strip().upper()

        if not symbol or status not in {
            "ACTIVE",
            "NEAR ENTRY",
            "TARGET HIT",
            "STOP ZONE",
        }:
            continue

        if status == "TARGET HIT":
            priority = "HIGH"
            message = "Candidate has reached or crossed Target 1."
        elif status == "STOP ZONE":
            priority = "HIGH"
            message = "Candidate is at or beyond the stop-loss zone."
        elif status == "NEAR ENTRY":
            priority = "MEDIUM"
            message = "Candidate is within approximately 1% of entry."
        else:
            priority = "LOW"
            message = "Candidate is active beyond the planned entry."

        add_alert(
            alerts,
            source="Live Watchlist",
            symbol=symbol,
            alert_type=status,
            priority=priority,
            message=message,
        )


def read_portfolio_alerts(wb, alerts: list[dict]) -> None:
    if "Portfolio" not in wb.sheetnames:
        return

    ws = wb["Portfolio"]
    headers = header_map(ws, 12)

    required = [
        "Symbol",
        "Side",
        "CMP",
        "Stop Loss",
        "Target",
        "Capital Used",
        "Risk Amount",
        "Status",
    ]

    if any(name not in headers for name in required):
        return

    trading_capital = safe_float(ws["B4"].value)
    total_risk = 0.0

    for row in range(13, ws.max_row + 1):
        symbol = str(ws.cell(row, headers["Symbol"]).value or "").strip()

        if not symbol:
            continue

        status = str(
            ws.cell(row, headers["Status"]).value or ""
        ).upper().strip()

        if status != "OPEN":
            continue

        side = str(
            ws.cell(row, headers["Side"]).value or ""
        ).upper().strip()

        cmp_price = safe_float(ws.cell(row, headers["CMP"]).value)
        stop_loss = safe_float(ws.cell(row, headers["Stop Loss"]).value)
        target = safe_float(ws.cell(row, headers["Target"]).value)
        capital_used = safe_float(
            ws.cell(row, headers["Capital Used"]).value
        )
        risk_amount = safe_float(
            ws.cell(row, headers["Risk Amount"]).value
        )

        total_risk += risk_amount

        if side in {"BUY", "CALL"}:
            stop_distance = (
                (cmp_price - stop_loss) / cmp_price * 100
                if cmp_price
                else 0.0
            )
            target_distance = (
                (target - cmp_price) / cmp_price * 100
                if cmp_price
                else 0.0
            )
        else:
            stop_distance = (
                (stop_loss - cmp_price) / cmp_price * 100
                if cmp_price
                else 0.0
            )
            target_distance = (
                (cmp_price - target) / cmp_price * 100
                if cmp_price
                else 0.0
            )

        if stop_distance <= 1:
            add_alert(
                alerts,
                source="Portfolio",
                symbol=symbol,
                alert_type="NEAR STOP",
                priority="HIGH",
                message=f"Only {stop_distance:.2f}% away from stop loss.",
            )

        if target_distance <= 1:
            add_alert(
                alerts,
                source="Portfolio",
                symbol=symbol,
                alert_type="NEAR TARGET",
                priority="MEDIUM",
                message=f"Only {target_distance:.2f}% away from target.",
            )

        if trading_capital > 0:
            capital_percent = capital_used / trading_capital * 100

            if capital_percent > POSITION_CAPITAL_LIMIT_PERCENT:
                add_alert(
                    alerts,
                    source="Portfolio",
                    symbol=symbol,
                    alert_type="OVERSIZED POSITION",
                    priority="HIGH",
                    message=(
                        f"Position uses {capital_percent:.2f}% of capital; "
                        f"limit is {POSITION_CAPITAL_LIMIT_PERCENT:.2f}%."
                    ),
                )

    if trading_capital > 0:
        risk_percent = total_risk / trading_capital * 100

        if risk_percent > PORTFOLIO_RISK_LIMIT_PERCENT:
            add_alert(
                alerts,
                source="Portfolio",
                symbol="PORTFOLIO",
                alert_type="HIGH TOTAL RISK",
                priority="HIGH",
                message=(
                    f"Total portfolio risk is {risk_percent:.2f}% of capital; "
                    f"limit is {PORTFOLIO_RISK_LIMIT_PERCENT:.2f}%."
                ),
            )


def create_alerts_sheet(wb, alerts: list[dict]) -> None:
    if "Alerts" in wb.sheetnames:
        del wb["Alerts"]

    ws = wb.create_sheet("Alerts", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    ws.merge_cells("A1:F2")
    ws["A1"] = "AQSD PROFESSIONAL - SMART ALERTS"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Last Updated"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["A4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE)

    headers = [
        "Priority",
        "Source",
        "Symbol",
        "Alert",
        "Message",
        "Timestamp",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(6, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}

    alerts = sorted(
        alerts,
        key=lambda item: (
            priority_order.get(item["Priority"], 9),
            item["Symbol"],
        ),
    )

    if not alerts:
        ws["A8"] = "No active alerts."
        ws["A8"].fill = PatternFill("solid", fgColor=GREEN)
        ws["A8"].font = Font(bold=True)
    else:
        for row_no, alert in enumerate(alerts, start=7):
            values = [
                alert["Priority"],
                alert["Source"],
                alert["Symbol"],
                alert["Alert"],
                alert["Message"],
                datetime.now(),
            ]

            for col, value in enumerate(values, start=1):
                cell = ws.cell(row_no, col, value)
                cell.border = Border(bottom=THIN)

            priority = alert["Priority"]

            if priority == "HIGH":
                fill = RED
            elif priority == "MEDIUM":
                fill = YELLOW
            else:
                fill = GREY

            ws.cell(row_no, 1).fill = PatternFill(
                "solid",
                fgColor=fill,
            )
            ws.cell(row_no, 1).font = Font(bold=True)
            ws.cell(row_no, 6).number_format = "dd-mm-yyyy hh:mm"

    widths = {
        "A": 12,
        "B": 18,
        "C": 18,
        "D": 20,
        "E": 60,
        "F": 20,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)
    alerts: list[dict] = []

    read_watchlist_alerts(wb, alerts)
    read_portfolio_alerts(wb, alerts)
    create_alerts_sheet(wb, alerts)

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Smart Alert Engine completed.")
    print(f"Active alerts: {len(alerts)}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
