
"""
AQSD Professional
Module: Sector Exposure
Version: 1.0

Creates a Sector Exposure sheet by grouping open portfolio positions.
Edit the SECTOR_MAP dictionary to match your watchlist.
"""

from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference

BASE = Path(__file__).resolve().parent.parent
FILE = BASE / "Output" / "Dashboard.xlsx"

SECTOR_MAP = {
    "RELIANCE.NS":"Energy",
    "ONGC.NS":"Energy",
    "BIOCON.NS":"Pharma",
    "DIVISLAB.NS":"Pharma",
    "SUNPHARMA.NS":"Pharma",
    "CIPLA.NS":"Pharma",
    "ICICIBANK.NS":"Banking",
    "HDFCBANK.NS":"Banking",
    "SBIN.NS":"Banking",
    "LT.NS":"Infrastructure",
    "LTIM.NS":"IT",
    "TCS.NS":"IT",
    "INFY.NS":"IT",
}

wb = load_workbook(FILE)

if "Portfolio" not in wb.sheetnames:
    raise RuntimeError("Portfolio sheet not found.")

pws = wb["Portfolio"]

if "Sector Exposure" in wb.sheetnames:
    del wb["Sector Exposure"]

ws = wb.create_sheet("Sector Exposure")

ws["A1"]="AQSD PROFESSIONAL - SECTOR EXPOSURE"
ws["A1"].font=Font(size=18,bold=True,color="FFFFFF")
ws["A1"].fill=PatternFill("solid",fgColor="1F4E78")

headers={}
for c in pws[12]:
    if c.value:
        headers[str(c.value)] = c.column

totals=defaultdict(float)

for r in range(13,pws.max_row+1):
    sym=pws.cell(r,headers["Symbol"]).value
    if not sym:
        continue
    status=str(pws.cell(r,headers["Status"]).value).upper()
    if status!="OPEN":
        continue
    cap=float(pws.cell(r,headers["Capital Used"]).value or 0)
    sector=SECTOR_MAP.get(str(sym),"Others")
    totals[sector]+=cap

ws["A3"]="Sector"
ws["B3"]="Capital"

row=4
for sec,val in sorted(totals.items()):
    ws.cell(row,1).value=sec
    ws.cell(row,2).value=val
    row+=1

if row>4:
    chart=PieChart()
    labels=Reference(ws,min_col=1,min_row=4,max_row=row-1)
    data=Reference(ws,min_col=2,min_row=3,max_row=row-1)
    chart.add_data(data,titles_from_data=True)
    chart.set_categories(labels)
    chart.title="Portfolio Sector Allocation"
    chart.height=10
    chart.width=12
    ws.add_chart(chart,"D3")

wb.save(FILE)

print("Sector Exposure created.")
print(FILE)
