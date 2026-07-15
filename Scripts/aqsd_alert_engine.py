
"""
AQSD Professional
Module: Alert Intelligence Engine
Version: 2.1 - Migration Safe

Purpose
-------
Generates, stores, ranks and reports AQSD alerts.

This version safely upgrades the earlier Version 1 alert table.
If the existing aqsd_alerts table uses the old schema, it is renamed
to a timestamped backup table and its records are migrated.

Commands
--------
python aqsd_alert_engine.py --migrate
python aqsd_alert_engine.py --run
python aqsd_alert_engine.py --status
python aqsd_alert_engine.py --unread
python aqsd_alert_engine.py --mark-read
python aqsd_alert_engine.py --acknowledge ALERT_ID
python aqsd_alert_engine.py --snooze ALERT_ID --hours 4
python aqsd_alert_engine.py --report
"""

from __future__ import annotations

import argparse
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "Output"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"
ALERT_CSV = OUTPUT_DIR / "AQSD_Alerts.csv"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


V2_SCHEMA = """
CREATE TABLE IF NOT EXISTS aqsd_alerts (
    alert_id INTEGER PRIMARY KEY AUTOINCREMENT,
    alert_key TEXT NOT NULL UNIQUE,
    alert_date TEXT NOT NULL,
    alert_time TEXT NOT NULL,
    alert_type TEXT NOT NULL,
    severity TEXT NOT NULL,
    priority_score REAL NOT NULL DEFAULT 0,
    nse_symbol TEXT NOT NULL DEFAULT '',
    sector TEXT NOT NULL DEFAULT '',
    title TEXT NOT NULL,
    message TEXT NOT NULL,
    score REAL,
    status TEXT NOT NULL DEFAULT 'UNREAD',
    acknowledged_at TEXT,
    snoozed_until TEXT,
    expires_at TEXT,
    source_module TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_date
ON aqsd_alerts(alert_date);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_status
ON aqsd_alerts(status);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_symbol
ON aqsd_alerts(nse_symbol);

CREATE INDEX IF NOT EXISTS idx_aqsd_alerts_priority
ON aqsd_alerts(priority_score DESC);
"""


REQUIRED_V2_COLUMNS = {
    "alert_id",
    "alert_key",
    "alert_date",
    "alert_time",
    "alert_type",
    "severity",
    "priority_score",
    "nse_symbol",
    "sector",
    "title",
    "message",
    "score",
    "status",
    "acknowledged_at",
    "snoozed_until",
    "expires_at",
    "source_module",
    "created_at",
    "updated_at",
}


def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND name=?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def table_columns(connection, table_name: str) -> set[str]:
    if not table_exists(connection, table_name):
        return set()

    return {
        row["name"]
        for row in connection.execute(
            f"PRAGMA table_info({table_name})"
        ).fetchall()
    }


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def severity_weight(severity: str) -> float:
    return {
        "CRITICAL": 100,
        "HIGH": 80,
        "MEDIUM": 60,
        "LOW": 40,
        "INFO": 20,
    }.get(str(severity).upper(), 20)


def make_alert_key(
    alert_date: str,
    alert_type: str,
    symbol: str,
    sector: str,
    title: str,
) -> str:
    raw = "|".join(
        [
            str(alert_date or ""),
            str(alert_type or "").strip().upper(),
            str(symbol or "").strip().upper(),
            str(sector or "").strip().upper(),
            str(title or "").strip().upper(),
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def calculate_priority(
    severity: str,
    score: float | None,
    confidence: float | None = None,
    risk_level: str = "",
) -> float:
    priority = severity_weight(severity) * 0.60
    priority += clamp(abs((score or 50) - 50) * 2, 0, 100) * 0.25
    priority += clamp(confidence or 50, 0, 100) * 0.15

    if str(risk_level).upper() == "HIGH":
        priority += 10
    elif str(risk_level).upper() == "LOW":
        priority -= 5

    return round(clamp(priority, 0, 100), 2)


def create_v2_table(connection) -> None:
    connection.executescript(V2_SCHEMA)
    connection.commit()


def migrate_schema() -> tuple[bool, str]:
    """
    Returns:
        migrated, message
    """

    setup_database()

    with connect() as connection:
        if not table_exists(connection, "aqsd_alerts"):
            create_v2_table(connection)
            return True, "Created new Version 2.1 alert table."

        columns = table_columns(connection, "aqsd_alerts")

        if REQUIRED_V2_COLUMNS.issubset(columns):
            create_v2_table(connection)
            return False, "Alert table already uses Version 2 schema."

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        legacy_table = f"aqsd_alerts_legacy_{timestamp}"

        connection.execute(
            f"ALTER TABLE aqsd_alerts RENAME TO {legacy_table}"
        )
        connection.commit()

        create_v2_table(connection)

        legacy_rows = connection.execute(
            f"SELECT * FROM {legacy_table}"
        ).fetchall()

        migrated_count = 0

        for row in legacy_rows:
            data = dict(row)

            alert_date = str(
                data.get("alert_date")
                or datetime.now().date().isoformat()
            )
            alert_time = str(
                data.get("alert_time")
                or datetime.now().strftime("%H:%M:%S")
            )
            alert_type = str(data.get("alert_type") or "LEGACY ALERT")
            severity = str(data.get("severity") or "INFO")
            symbol = str(data.get("nse_symbol") or "")
            sector = str(data.get("sector") or "")
            title = str(data.get("title") or alert_type)
            message = str(data.get("message") or "")
            score = safe_float(data.get("score"))
            status = str(data.get("status") or "UNREAD")
            source = str(data.get("source_module") or "legacy_alert_engine")
            created_at = str(
                data.get("created_at")
                or datetime.now().isoformat(timespec="seconds")
            )
            updated_at = datetime.now().isoformat(timespec="seconds")

            expires_at = (
                datetime.now() + timedelta(hours=24)
            ).isoformat(timespec="seconds")

            key = make_alert_key(
                alert_date,
                alert_type,
                symbol,
                sector,
                title,
            )

            connection.execute(
                """
                INSERT OR IGNORE INTO aqsd_alerts(
                    alert_key,
                    alert_date,
                    alert_time,
                    alert_type,
                    severity,
                    priority_score,
                    nse_symbol,
                    sector,
                    title,
                    message,
                    score,
                    status,
                    acknowledged_at,
                    snoozed_until,
                    expires_at,
                    source_module,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    key,
                    alert_date,
                    alert_time,
                    alert_type,
                    severity,
                    calculate_priority(severity, score),
                    symbol,
                    sector,
                    title,
                    message,
                    score,
                    status,
                    None,
                    None,
                    expires_at,
                    source,
                    created_at,
                    updated_at,
                ),
            )
            migrated_count += 1

        connection.commit()

        return (
            True,
            f"Migrated {migrated_count} legacy alerts. "
            f"Backup table: {legacy_table}"
        )


def setup_schema() -> None:
    migrate_schema()


def add_alert(
    alerts: list[dict],
    *,
    alert_type: str,
    severity: str,
    title: str,
    message: str,
    nse_symbol: str = "",
    sector: str = "",
    score: float | None = None,
    confidence: float | None = None,
    risk_level: str = "",
    source_module: str = "",
    expiry_hours: int = 24,
) -> None:
    now = datetime.now()
    alert_date = now.date().isoformat()

    alerts.append(
        {
            "alert_key": make_alert_key(
                alert_date,
                alert_type,
                nse_symbol,
                sector,
                title,
            ),
            "alert_date": alert_date,
            "alert_time": now.strftime("%H:%M:%S"),
            "alert_type": alert_type,
            "severity": severity,
            "priority_score": calculate_priority(
                severity,
                score,
                confidence,
                risk_level,
            ),
            "nse_symbol": nse_symbol or "",
            "sector": sector or "",
            "title": title,
            "message": message,
            "score": score,
            "status": "UNREAD",
            "acknowledged_at": None,
            "snoozed_until": None,
            "expires_at": (
                now + timedelta(hours=expiry_hours)
            ).isoformat(timespec="seconds"),
            "source_module": source_module,
            "created_at": now.isoformat(timespec="seconds"),
            "updated_at": now.isoformat(timespec="seconds"),
        }
    )


def load_latest_decisions() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "aqsd_trade_decisions"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM aqsd_trade_decisions
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
            ORDER BY priority_rank
            """,
            connection,
        )


def load_latest_structure() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "price_structure_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM price_structure_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
            """,
            connection,
        )


def load_latest_sectors() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "sector_rotation_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM sector_rotation_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            ORDER BY sector_rotation_score DESC
            """,
            connection,
        )


def load_latest_regime() -> dict | None:
    with connect() as connection:
        if not table_exists(connection, "market_regime_intelligence"):
            return None

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else None


def load_previous_regime() -> dict | None:
    with connect() as connection:
        if not table_exists(connection, "market_regime_intelligence"):
            return None

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1 OFFSET 1
            """
        ).fetchone()

    return dict(row) if row else None


def build_alerts() -> list[dict]:
    alerts: list[dict] = []

    decisions = load_latest_decisions()

    for _, row in decisions.iterrows():
        action = str(row.get("action") or "")
        symbol = str(row.get("nse_symbol") or "")
        sector = str(row.get("sector") or "")
        score = safe_float(row.get("master_score"))
        confidence = safe_float(row.get("confidence_percent"))
        risk = str(row.get("risk_level") or "")

        if action in {"STRONG BUY", "BUY", "BUY ON DIP"}:
            severity = "HIGH" if action == "STRONG BUY" else "MEDIUM"

            add_alert(
                alerts,
                alert_type=action,
                severity=severity,
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: {action}",
                message=(
                    f"Rank {row.get('priority_rank')}; "
                    f"Master score {score}; confidence {confidence}%; "
                    f"entry {row.get('entry_low')} to {row.get('entry_high')}; "
                    f"stop {row.get('stop_loss')}; "
                    f"targets {row.get('target_1')} / {row.get('target_2')}."
                ),
                score=score,
                confidence=confidence,
                risk_level=risk,
                source_module="aqsd_decision_engine",
            )

        elif action in {"AVOID", "EXIT / AVOID"}:
            add_alert(
                alerts,
                alert_type="EXIT / AVOID",
                severity="HIGH",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: {action}",
                message=(
                    f"Master score {score}; confidence {confidence}%; "
                    f"bias {row.get('directional_bias')}."
                ),
                score=score,
                confidence=confidence,
                risk_level=risk,
                source_module="aqsd_decision_engine",
            )

        if risk.upper() == "HIGH":
            add_alert(
                alerts,
                alert_type="HIGH RISK",
                severity="HIGH",
                nse_symbol=symbol,
                sector=sector,
                title=f"{symbol}: High Risk",
                message=(
                    f"AQSD classified this setup as high risk. "
                    f"Score {score}; confidence {confidence}%."
                ),
                score=score,
                confidence=confidence,
                risk_level=risk,
                source_module="aqsd_decision_engine",
            )

    structure = load_latest_structure()

    for _, row in structure.iterrows():
        symbol = str(row.get("nse_symbol") or "")
        score = safe_float(row.get("structure_score"))
        bos = str(row.get("bos_signal") or "")
        choch = str(row.get("choch_signal") or "")

        if bos in {"BULLISH BOS", "BEARISH BOS"}:
            add_alert(
                alerts,
                alert_type="BREAKOUT" if bos == "BULLISH BOS" else "BREAKDOWN",
                severity="MEDIUM" if bos == "BULLISH BOS" else "HIGH",
                nse_symbol=symbol,
                title=f"{symbol}: {bos}",
                message=(
                    f"{bos} detected at close {row.get('close_price')}; "
                    f"structure score {score}."
                ),
                score=score,
                source_module="aqsd_price_structure",
            )

        if choch in {"BULLISH CHOCH", "BEARISH CHOCH"}:
            add_alert(
                alerts,
                alert_type=choch,
                severity="MEDIUM" if choch == "BULLISH CHOCH" else "HIGH",
                nse_symbol=symbol,
                title=f"{symbol}: {choch}",
                message="A structural change of character was detected.",
                score=score,
                source_module="aqsd_price_structure",
            )

    sectors = load_latest_sectors()

    for _, row in sectors.iterrows():
        sector = str(row.get("sector") or "")
        state = str(row.get("rotation_state") or "")
        score = safe_float(row.get("sector_rotation_score"))

        if state in {"STRONG INFLOW", "EARLY ROTATION"}:
            add_alert(
                alerts,
                alert_type="SECTOR INFLOW",
                severity="MEDIUM",
                sector=sector,
                title=f"{sector}: {state}",
                message=(
                    f"Rotation score {score}; "
                    f"leader {row.get('leader_symbol')}."
                ),
                score=score,
                source_module="aqsd_sector_rotation",
            )

        elif state in {"STRONG OUTFLOW", "WEAKENING LEADERSHIP"}:
            add_alert(
                alerts,
                alert_type="SECTOR OUTFLOW",
                severity="HIGH",
                sector=sector,
                title=f"{sector}: {state}",
                message=(
                    f"Rotation score {score}; "
                    f"laggard {row.get('laggard_symbol')}."
                ),
                score=score,
                source_module="aqsd_sector_rotation",
            )

    current = load_latest_regime()
    previous = load_previous_regime()

    if current:
        current_regime = str(current.get("market_regime") or "")
        previous_regime = (
            str(previous.get("market_regime") or "")
            if previous
            else ""
        )

        if previous_regime and current_regime != previous_regime:
            add_alert(
                alerts,
                alert_type="MARKET REGIME CHANGE",
                severity="CRITICAL",
                title=f"Market Regime Changed to {current_regime}",
                message=(
                    f"Previous regime: {previous_regime}. "
                    f"Strategy: {current.get('suggested_strategy')}. "
                    f"Exposure: {current.get('capital_exposure_percent')}%."
                ),
                score=safe_float(current.get("regime_score")),
                source_module="aqsd_market_regime",
                expiry_hours=48,
            )

    return sorted(
        alerts,
        key=lambda item: (
            -item["priority_score"],
            item["title"],
        ),
    )


def save_alerts(alerts: list[dict]) -> tuple[int, int]:
    inserted = 0
    updated = 0

    with connect() as connection:
        for alert in alerts:
            existing = connection.execute(
                """
                SELECT alert_id
                FROM aqsd_alerts
                WHERE alert_key=?
                """,
                (alert["alert_key"],),
            ).fetchone()

            if existing:
                connection.execute(
                    """
                    UPDATE aqsd_alerts
                    SET alert_time=?,
                        severity=?,
                        priority_score=?,
                        message=?,
                        score=?,
                        expires_at=?,
                        source_module=?,
                        updated_at=?
                    WHERE alert_key=?
                    """,
                    (
                        alert["alert_time"],
                        alert["severity"],
                        alert["priority_score"],
                        alert["message"],
                        alert["score"],
                        alert["expires_at"],
                        alert["source_module"],
                        alert["updated_at"],
                        alert["alert_key"],
                    ),
                )
                updated += 1
            else:
                connection.execute(
                    """
                    INSERT INTO aqsd_alerts(
                        alert_key,
                        alert_date,
                        alert_time,
                        alert_type,
                        severity,
                        priority_score,
                        nse_symbol,
                        sector,
                        title,
                        message,
                        score,
                        status,
                        acknowledged_at,
                        snoozed_until,
                        expires_at,
                        source_module,
                        created_at,
                        updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        alert["alert_key"],
                        alert["alert_date"],
                        alert["alert_time"],
                        alert["alert_type"],
                        alert["severity"],
                        alert["priority_score"],
                        alert["nse_symbol"],
                        alert["sector"],
                        alert["title"],
                        alert["message"],
                        alert["score"],
                        alert["status"],
                        alert["acknowledged_at"],
                        alert["snoozed_until"],
                        alert["expires_at"],
                        alert["source_module"],
                        alert["created_at"],
                        alert["updated_at"],
                    ),
                )
                inserted += 1

        connection.commit()

    return inserted, updated


def run_engine() -> tuple[list[dict], int, int]:
    setup_schema()

    run_id = start_run(
        "aqsd_alert_engine",
        "Generating AQSD alerts",
    )

    try:
        alerts = build_alerts()
        inserted, updated = save_alerts(alerts)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=inserted + updated,
            errors_count=0,
            message=(
                f"Generated={len(alerts)}; "
                f"Inserted={inserted}; Updated={updated}"
            ),
        )

        return alerts, inserted, updated

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


def latest_alerts(unread_only: bool = False) -> pd.DataFrame:
    setup_schema()

    query = """
        SELECT *
        FROM aqsd_alerts
        WHERE (
            expires_at IS NULL
            OR expires_at >= DATETIME('now')
        )
        AND (
            snoozed_until IS NULL
            OR snoozed_until <= DATETIME('now')
        )
    """

    if unread_only:
        query += " AND status='UNREAD'"

    query += """
        ORDER BY priority_score DESC, alert_id DESC
    """

    with connect() as connection:
        return pd.read_sql_query(query, connection)


def mark_all_read() -> int:
    now = datetime.now().isoformat(timespec="seconds")

    with connect() as connection:
        cursor = connection.execute(
            """
            UPDATE aqsd_alerts
            SET status='READ',
                updated_at=?
            WHERE status='UNREAD'
            """,
            (now,),
        )
        connection.commit()
        return cursor.rowcount


def acknowledge(alert_id: int) -> int:
    now = datetime.now().isoformat(timespec="seconds")

    with connect() as connection:
        cursor = connection.execute(
            """
            UPDATE aqsd_alerts
            SET status='ACKNOWLEDGED',
                acknowledged_at=?,
                updated_at=?
            WHERE alert_id=?
            """,
            (now, now, alert_id),
        )
        connection.commit()
        return cursor.rowcount


def snooze(alert_id: int, hours: int) -> int:
    now = datetime.now()
    until = (
        now + timedelta(hours=max(1, hours))
    ).isoformat(timespec="seconds")

    with connect() as connection:
        cursor = connection.execute(
            """
            UPDATE aqsd_alerts
            SET snoozed_until=?,
                updated_at=?
            WHERE alert_id=?
            """,
            (
                until,
                now.isoformat(timespec="seconds"),
                alert_id,
            ),
        )
        connection.commit()
        return cursor.rowcount


def write_reports(frame: pd.DataFrame | None = None) -> None:
    if frame is None:
        frame = latest_alerts()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    frame.to_csv(
        ALERT_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    if DASHBOARD.exists():
        workbook = load_workbook(DASHBOARD)
    else:
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    sheet_name = "AQSD Alerts"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(sheet_name, 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:N2")
    ws["A1"] = "AQSD PROFESSIONAL - ALERT INTELLIGENCE ENGINE V2.1"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    unread = (
        int((frame["status"] == "UNREAD").sum())
        if not frame.empty
        else 0
    )

    ws["A4"] = "Active Alerts"
    ws["B4"] = len(frame)
    ws["D4"] = "Unread"
    ws["E4"] = unread
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill("solid", fgColor=BLUE)

    headers = [
        "ID",
        "Date",
        "Time",
        "Severity",
        "Priority",
        "Alert Type",
        "Symbol",
        "Sector",
        "Title",
        "Message",
        "Score",
        "Status",
        "Expires",
        "Source",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, (_, row) in enumerate(frame.iterrows(), start=8):
        values = [
            row.get("alert_id"),
            row.get("alert_date"),
            row.get("alert_time"),
            row.get("severity"),
            row.get("priority_score"),
            row.get("alert_type"),
            row.get("nse_symbol"),
            row.get("sector"),
            row.get("title"),
            row.get("message"),
            row.get("score"),
            row.get("status"),
            row.get("expires_at"),
            row.get("source_module"),
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        severity = str(row.get("severity") or "")

        ws.cell(row_no, 4).fill = PatternFill(
            "solid",
            fgColor=(
                RED
                if severity in {"CRITICAL", "HIGH"}
                else YELLOW
                if severity == "MEDIUM"
                else GREY
            ),
        )

    widths = {
        "A": 8,
        "B": 14,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 22,
        "G": 16,
        "H": 22,
        "I": 38,
        "J": 90,
        "K": 12,
        "L": 16,
        "M": 21,
        "N": 28,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    workbook.save(DASHBOARD)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        columns = table_columns(connection, "aqsd_alerts")

        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status='UNREAD' THEN 1 ELSE 0 END) AS unread,
                SUM(CASE WHEN status='ACKNOWLEDGED' THEN 1 ELSE 0 END) AS acknowledged,
                MIN(alert_date) AS first_date,
                MAX(alert_date) AS latest_date
            FROM aqsd_alerts
            """
        ).fetchone()

    print("\nAQSD ALERT ENGINE V2.1 STATUS")
    print("=" * 72)
    print(f"Schema version:     {'V2.1' if REQUIRED_V2_COLUMNS.issubset(columns) else 'OLD'}")
    print(f"Stored alerts:      {row['total'] or 0}")
    print(f"Unread alerts:      {row['unread'] or 0}")
    print(f"Acknowledged:       {row['acknowledged'] or 0}")
    print(f"First date:         {row['first_date'] or 'No data'}")
    print(f"Latest date:        {row['latest_date'] or 'No data'}")
    print("=" * 72)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Alert Intelligence Engine V2.1."
    )

    parser.add_argument("--migrate", action="store_true")
    parser.add_argument("--run", action="store_true")
    parser.add_argument("--report", action="store_true")
    parser.add_argument("--status", action="store_true")
    parser.add_argument("--unread", action="store_true")
    parser.add_argument("--mark-read", action="store_true")
    parser.add_argument("--acknowledge", type=int, metavar="ALERT_ID")
    parser.add_argument("--snooze", type=int, metavar="ALERT_ID")
    parser.add_argument("--hours", type=int, default=4)

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if args.migrate:
        migrated, message = migrate_schema()
        print(message)
        return

    setup_schema()

    if args.mark_read:
        print(f"Alerts marked read: {mark_all_read()}")
        return

    if args.acknowledge is not None:
        print(
            f"Alerts acknowledged: "
            f"{acknowledge(args.acknowledge)}"
        )
        return

    if args.snooze is not None:
        print(
            f"Alerts snoozed: "
            f"{snooze(args.snooze, args.hours)}"
        )
        return

    if args.unread:
        frame = latest_alerts(unread_only=True)

        if frame.empty:
            print("No unread active alerts.")
        else:
            print(
                frame[
                    [
                        "alert_id",
                        "severity",
                        "priority_score",
                        "alert_type",
                        "nse_symbol",
                        "sector",
                        "title",
                    ]
                ].to_string(index=False)
            )
        return

    if args.run:
        alerts, inserted, updated = run_engine()
        frame = latest_alerts()
        write_reports(frame)

        print("\nAQSD ALERT ENGINE V2.1")
        print("=" * 72)
        print(f"Alerts generated: {len(alerts)}")
        print(f"Inserted:         {inserted}")
        print(f"Updated:          {updated}")
        print(f"CSV:              {ALERT_CSV}")
        print(f"Dashboard:        {DASHBOARD}")
        return

    if args.report:
        write_reports()
        print(f"CSV rebuilt:\n{ALERT_CSV}")
        print(f"Dashboard rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
