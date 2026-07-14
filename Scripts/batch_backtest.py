
"""
AQSD Professional
Module: Batch Backtester
Version: 1.0

Runs the AQSD backtest strategy across multiple NSE symbols and ranks
the results inside Dashboard.xlsx.

Input priority:
1. Symbols from CALL Candidates and PUT Candidates sheets
2. Symbols from FnO_Stocks.xlsx
3. Built-in fallback symbols

Outputs:
- Batch Backtest Summary
- Batch Backtest Trades
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"
FNO_FILE = BASE / "Data" / "FnO_Stocks.xlsx"


# ============================================================
# SETTINGS
# ============================================================

DEFAULT_PERIOD = "3y"
DEFAULT_LIMIT = 20

ATR_PERIOD = 14
ATR_MULTIPLIER = 2.0
TARGET_RR = 2.0
MAX_HOLD_DAYS = 10

INITIAL_CAPITAL = 200000
RISK_PERCENT = 1.0


# ============================================================
# INDICATORS
# ============================================================

def calculate_rsi(close: pd.Series, period: int = 14) -> pd.Series:
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    avg_gain = gain.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    avg_loss = loss.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()

    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def calculate_atr(df: pd.DataFrame, period: int = 14) -> pd.Series:
    previous_close = df["Close"].shift(1)

    true_range = pd.concat(
        [
            df["High"] - df["Low"],
            (df["High"] - previous_close).abs(),
            (df["Low"] - previous_close).abs(),
        ],
        axis=1,
    ).max(axis=1)

    return true_range.ewm(
        alpha=1 / period,
        adjust=False,
        min_periods=period,
    ).mean()


# ============================================================
# SYMBOL LOADING
# ============================================================

def normalize_symbol(symbol: str) -> str:
    symbol = str(symbol).strip().upper()

    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"

    return symbol


def symbols_from_dashboard(limit: int) -> list[str]:
    if not DASHBOARD.exists():
        return []

    try:
        wb = load_workbook(
            DASHBOARD,
            read_only=True,
            data_only=True,
        )
    except Exception:
        return []

    symbols: list[str] = []

    for sheet_name in ("CALL Candidates", "PUT Candidates"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        headers = {
            str(cell.value).strip(): cell.column
            for cell in ws[1]
            if cell.value is not None
        }

        symbol_col = headers.get("Symbol")

        if not symbol_col:
            continue

        for row in range(2, ws.max_row + 1):
            symbol = normalize_symbol(
                ws.cell(row, symbol_col).value or ""
            )

            if symbol and symbol not in symbols:
                symbols.append(symbol)

            if len(symbols) >= limit:
                wb.close()
                return symbols

    wb.close()
    return symbols


def symbols_from_fno_file(limit: int) -> list[str]:
    if not FNO_FILE.exists():
        return []

    try:
        df = pd.read_excel(FNO_FILE)
    except Exception:
        return []

    possible_columns = [
        "Yahoo Symbol",
        "Symbol",
        "SYMBOL",
        "Ticker",
    ]

    symbol_col = next(
        (column for column in possible_columns if column in df.columns),
        None,
    )

    if not symbol_col:
        return []

    symbols: list[str] = []

    for value in df[symbol_col].dropna():
        symbol = normalize_symbol(value)

        if symbol and symbol not in symbols:
            symbols.append(symbol)

        if len(symbols) >= limit:
            break

    return symbols


def load_symbols(limit: int) -> list[str]:
    symbols = symbols_from_dashboard(limit)

    if symbols:
        return symbols

    symbols = symbols_from_fno_file(limit)

    if symbols:
        return symbols

    fallback = [
        "RELIANCE.NS",
        "HDFCBANK.NS",
        "ICICIBANK.NS",
        "INFY.NS",
        "TCS.NS",
        "SBIN.NS",
        "LT.NS",
        "SUNPHARMA.NS",
        "TATAMOTORS.NS",
        "BEL.NS",
    ]

    return fallback[:limit]


# ============================================================
# BACKTEST LOGIC
# ============================================================

def download_data(symbol: str, period: str) -> pd.DataFrame:
    df = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
    )

    if df.empty:
        raise RuntimeError("No data")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    df = df.dropna(subset=["Open", "High", "Low", "Close"])

    df["EMA20"] = df["Close"].ewm(span=20, adjust=False).mean()
    df["EMA50"] = df["Close"].ewm(span=50, adjust=False).mean()
    df["RSI14"] = calculate_rsi(df["Close"], 14)
    df["ATR14"] = calculate_atr(df, ATR_PERIOD)

    return df.dropna()


def build_signals(df: pd.DataFrame) -> pd.Series:
    bullish = (
        (df["EMA20"] > df["EMA50"])
        & (df["RSI14"] > 55)
        & (df["Close"] > df["EMA20"])
    )

    return bullish & (~bullish.shift(1).fillna(False))


def calculate_quantity(
    capital: float,
    entry: float,
    stop_loss: float,
) -> int:
    risk_amount = capital * (RISK_PERCENT / 100)
    risk_per_share = abs(entry - stop_loss)

    if risk_per_share <= 0:
        return 0

    qty_by_risk = int(risk_amount // risk_per_share)
    qty_by_capital = int(capital // entry)

    return max(0, min(qty_by_risk, qty_by_capital))


def run_symbol_backtest(
    symbol: str,
    df: pd.DataFrame,
) -> tuple[dict, list[dict]]:
    signals = build_signals(df)

    capital = INITIAL_CAPITAL
    trades: list[dict] = []
    i = 0

    while i < len(df) - 1:
        if not bool(signals.iloc[i]):
            i += 1
            continue

        entry_index = i + 1
        entry_row = df.iloc[entry_index]

        entry = float(entry_row["Open"])
        atr_value = float(df.iloc[i]["ATR14"])

        stop_loss = entry - ATR_MULTIPLIER * atr_value
        target = entry + TARGET_RR * (entry - stop_loss)

        quantity = calculate_quantity(
            capital,
            entry,
            stop_loss,
        )

        if quantity <= 0:
            i += 1
            continue

        final_index = min(
            entry_index + MAX_HOLD_DAYS,
            len(df) - 1,
        )

        exit_index = final_index
        exit_price = float(df.iloc[final_index]["Close"])
        result = "TIME EXIT"

        for j in range(entry_index, final_index + 1):
            row = df.iloc[j]

            if float(row["Low"]) <= stop_loss:
                exit_index = j
                exit_price = stop_loss
                result = "LOSS"
                break

            if float(row["High"]) >= target:
                exit_index = j
                exit_price = target
                result = "WIN"
                break

        pnl = (exit_price - entry) * quantity
        capital += pnl

        trades.append(
            {
                "Symbol": symbol,
                "Entry Date": str(df.index[entry_index].date()),
                "Exit Date": str(df.index[exit_index].date()),
                "Entry": round(entry, 2),
                "Stop Loss": round(stop_loss, 2),
                "Target": round(target, 2),
                "Exit": round(exit_price, 2),
                "Quantity": quantity,
                "P/L": round(pnl, 2),
                "Result": result,
                "Holding Days": max(
                    exit_index - entry_index + 1,
                    1,
                ),
            }
        )

        i = exit_index + 1

    wins = [trade for trade in trades if trade["P/L"] > 0]
    losses = [trade for trade in trades if trade["P/L"] < 0]

    gross_profit = sum(trade["P/L"] for trade in wins)
    gross_loss = abs(sum(trade["P/L"] for trade in losses))
    net_profit = sum(trade["P/L"] for trade in trades)

    win_rate = (
        len(wins) / len(trades) * 100
        if trades
        else 0.0
    )

    profit_factor = (
        gross_profit / gross_loss
        if gross_loss
        else gross_profit
    )

    equity = INITIAL_CAPITAL
    peak = INITIAL_CAPITAL
    maximum_drawdown = 0.0

    for trade in trades:
        equity += trade["P/L"]
        peak = max(peak, equity)
        maximum_drawdown = min(
            maximum_drawdown,
            equity - peak,
        )

    summary = {
        "Symbol": symbol,
        "Trades": len(trades),
        "Wins": len(wins),
        "Losses": len(losses),
        "Win Rate %": round(win_rate, 2),
        "Net Profit": round(net_profit, 2),
        "Return %": round(
            net_profit / INITIAL_CAPITAL * 100,
            2,
        ),
        "Profit Factor": round(profit_factor, 2),
        "Max Drawdown": round(abs(maximum_drawdown), 2),
        "Final Capital": round(capital, 2),
    }

    return summary, trades


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_results(
    summaries: list[dict],
    trades: list[dict],
) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_name in (
        "Batch Backtest Summary",
        "Batch Backtest Trades",
    ):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    summary_ws = wb.create_sheet("Batch Backtest Summary")
    trades_ws = wb.create_sheet("Batch Backtest Trades")

    navy = "17365D"
    white = "FFFFFF"
    green = "C6EFCE"
    red = "FFC7CE"

    summary_ws.merge_cells("A1:J2")
    summary_ws["A1"] = "AQSD PROFESSIONAL - BATCH BACKTEST"
    summary_ws["A1"].font = Font(
        size=18,
        bold=True,
        color=white,
    )
    summary_ws["A1"].fill = PatternFill(
        "solid",
        fgColor=navy,
    )
    summary_ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    summary_headers = [
        "Rank",
        "Symbol",
        "Trades",
        "Wins",
        "Losses",
        "Win Rate %",
        "Net Profit",
        "Return %",
        "Profit Factor",
        "Max Drawdown",
        "Final Capital",
    ]

    for col, heading in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(4, col, heading)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)

    ranked = sorted(
        summaries,
        key=lambda item: (
            item["Net Profit"],
            item["Profit Factor"],
        ),
        reverse=True,
    )

    for row_no, item in enumerate(ranked, start=5):
        values = [
            row_no - 4,
            item["Symbol"],
            item["Trades"],
            item["Wins"],
            item["Losses"],
            item["Win Rate %"],
            item["Net Profit"],
            item["Return %"],
            item["Profit Factor"],
            item["Max Drawdown"],
            item["Final Capital"],
        ]

        for col, value in enumerate(values, start=1):
            summary_ws.cell(row_no, col, value)

        for col in (7, 10, 11):
            summary_ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

        for col in (6, 8):
            summary_ws.cell(
                row_no,
                col,
            ).number_format = '0.00"%"'

        summary_ws.cell(
            row_no,
            7,
        ).fill = PatternFill(
            "solid",
            fgColor=green if item["Net Profit"] >= 0 else red,
        )

    if ranked:
        last_row = 4 + len(ranked)

        chart = BarChart()
        chart.title = "Net Profit by Symbol"
        chart.y_axis.title = "Net Profit"
        chart.x_axis.title = "Symbol"

        chart.add_data(
            Reference(
                summary_ws,
                min_col=7,
                min_row=4,
                max_row=last_row,
            ),
            titles_from_data=True,
        )

        chart.set_categories(
            Reference(
                summary_ws,
                min_col=2,
                min_row=5,
                max_row=last_row,
            )
        )

        chart.height = 8
        chart.width = 14
        summary_ws.add_chart(chart, "M4")

    trade_headers = [
        "Trade No.",
        "Symbol",
        "Entry Date",
        "Exit Date",
        "Entry",
        "Stop Loss",
        "Target",
        "Exit",
        "Quantity",
        "P/L",
        "Result",
        "Holding Days",
    ]

    for col, heading in enumerate(trade_headers, start=1):
        cell = trades_ws.cell(1, col, heading)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)

    for row_no, trade in enumerate(trades, start=2):
        values = [
            row_no - 1,
            trade["Symbol"],
            trade["Entry Date"],
            trade["Exit Date"],
            trade["Entry"],
            trade["Stop Loss"],
            trade["Target"],
            trade["Exit"],
            trade["Quantity"],
            trade["P/L"],
            trade["Result"],
            trade["Holding Days"],
        ]

        for col, value in enumerate(values, start=1):
            trades_ws.cell(row_no, col, value)

        for col in (5, 6, 7, 8, 10):
            trades_ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

    summary_ws.freeze_panes = "A5"
    trades_ws.freeze_panes = "A2"
    summary_ws.auto_filter.ref = summary_ws.dimensions
    trades_ws.auto_filter.ref = trades_ws.dimensions

    wb.save(DASHBOARD)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Batch backtest AQSD candidates."
    )

    parser.add_argument(
        "--period",
        default=DEFAULT_PERIOD,
        help="Yahoo Finance period, e.g. 1y, 3y, 5y",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_LIMIT,
        help="Maximum number of symbols to test.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    symbols = load_symbols(args.limit)

    print("\nAQSD BATCH BACKTEST")
    print("=" * 68)
    print(f"Symbols: {len(symbols)}")
    print(f"Period: {args.period}")

    summaries: list[dict] = []
    all_trades: list[dict] = []

    for index, symbol in enumerate(symbols, start=1):
        print(f"[{index}/{len(symbols)}] {symbol}")

        try:
            df = download_data(symbol, args.period)
            summary, trades = run_symbol_backtest(
                symbol,
                df,
            )

            summaries.append(summary)
            all_trades.extend(trades)

        except Exception as error:
            print(f"  Skipped: {error}")

    write_results(
        summaries,
        all_trades,
    )

    print("=" * 68)
    print(f"Symbols completed: {len(summaries)}")
    print(f"Trades generated: {len(all_trades)}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
