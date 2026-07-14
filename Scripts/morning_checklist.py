
"""
AQSD Professional
Module: Morning Checklist
Version: 1.0

Creates a Morning Checklist sheet that summarizes the
most important items before the market opens.
"""

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).resolve().parent.parent
FILE = BASE / "Output" / "Dashboard.xlsx"

wb = load_workbook(FILE)

if "Morning Checklist" in wb.sheetnames:
    del wb["Morning Checklist"]

ws = wb.create_sheet("Morning Checklist", 0)
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F2")
c = ws["A1"]
c.value = "AQSD PROFESSIONAL - MORNING CHECKLIST"
c.font = Font(size=20, bold=True, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="17365D")
c.alignment = Alignment(horizontal="center")

ws["A4"] = "Generated"
ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

items = [
    ("☐", "Run AQSD Daily Scan"),
    ("☐", "Review Market Pulse"),
    ("☐", "Review Smart Alerts"),
    ("☐", "Check Top 10 CALL Candidates"),
    ("☐", "Check Top 10 PUT Candidates"),
    ("☐", "Review Portfolio Heat Map"),
    ("☐", "Verify Position Size & Risk"),
    ("☐", "Review Daily Trading Report"),
    ("☐", "Create Dashboard Backup"),
    ("☐", "Update Trade Journal After Market"),
]

ws["A6"] = "Status"
ws["B6"] = "Task"
ws["A6"].font = ws["B6"].font = Font(bold=True)
ws["A6"].fill = ws["B6"].fill = PatternFill("solid", fgColor="D9EAF7")

r = 7
for status, task in items:
    ws.cell(r,1).value = status
    ws.cell(r,2).value = task
    r += 1

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 55

wb.save(FILE)
print("Morning Checklist created.")
print(FILE)
