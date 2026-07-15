
"""
AQSD Professional
Module: Market Regime Intelligence Engine
Version: 1.0

Purpose
-------
Classifies the current market environment using AQSD's stored intelligence.

Regimes
-------
- BULL TREND
- BEAR TREND
- RANGE BOUND
- HIGH VOLATILITY
- LOW VOLATILITY
- ACCUMULATION
- DISTRIBUTION
- TRANSITION
- RISK ON
- RISK OFF

Inputs
------
- Market Breadth Intelligence
- Sector Rotation Intelligence
- Price Structure Intelligence
- Global Markets Intelligence
- VIX
- USD/INR
- US Dollar Index
- US 10Y Yield
- Commodities
- Unified Master Intelligence

Outputs
-------
- Market Regime
- Regime Score
- Bull Probability
- Bear Probability
- Volatility State
- Risk State
- Participation State
- Suggested Strategy
- Capital Exposure Guidance
- SQLite history
- Excel sheet: Market Regime Intelligence

Commands
--------
python aqsd_market_regime.py --run
python aqsd_market_regime.py --status
python aqsd_market_regime.py --report
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS market_regime_intelligence (
    regime_id INTEGER PRIMARY KEY AUTOINCREMENT,
    regime_date TEXT NOT NULL UNIQUE,
    market_regime TEXT NOT NULL,
    regime_score REAL,
    bull_probability REAL,
    bear_probability REAL,
    range_probability REAL,
    volatility_state TEXT,
    risk_state TEXT,
    participation_state TEXT,
    breadth_score REAL,
    sector_rotation_score REAL,
    structure_score REAL,
    global_risk_score REAL,
    vix_change_percent REAL,
    dollar_change_percent REAL,
    yield_change_percent REAL,
    bullish_stock_percent REAL,
    bearish_stock_percent REAL,
    suggested_strategy TEXT,
    capital_exposure_percent REAL,
    explanation TEXT,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_market_regime_date
ON market_regime_intelligence(regime_date);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# HELPERS
# ============================================================

def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table'
          AND name=?
        """,
        (table_name,),
    ).fetchone()

    return row is not None


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


# ============================================================
# LOADERS
# ============================================================

def load_market_breadth() -> dict:
    with connect() as connection:
        if not table_exists(connection, "market_breadth_intelligence"):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM market_breadth_intelligence
            WHERE scope_type='MARKET'
              AND trade_date=(
                  SELECT MAX(trade_date)
                  FROM market_breadth_intelligence
              )
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def load_sector_rotation() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "sector_rotation_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM sector_rotation_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            """,
            connection,
        )


def load_price_structure() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "price_structure_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                structure_score,
                directional_bias,
                market_structure,
                adx_14
            FROM price_structure_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
            """,
            connection,
        )


def load_global_markets() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "global_markets"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM global_markets
            WHERE snapshot_date=(
                SELECT MAX(snapshot_date)
                FROM global_markets
            )
            """,
            connection,
        )


def load_unified_master() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "unified_master_intelligence"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                master_score,
                directional_bias,
                recommendation
            FROM unified_master_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM unified_master_intelligence
            )
            """,
            connection,
        )


# ============================================================
# DERIVED COMPONENTS
# ============================================================

def global_component(frame: pd.DataFrame) -> dict:
    if frame.empty:
        return {
            "global_risk_score": 50.0,
            "vix_change_percent": None,
            "dollar_change_percent": None,
            "yield_change_percent": None,
        }

    lookup = {
        str(row["symbol"]): row
        for _, row in frame.iterrows()
    }

    score = 50.0

    for symbol in ("^DJI", "^IXIC", "^GSPC", "^FTSE", "^GDAXI", "^N225", "^HSI"):
        row = lookup.get(symbol)
        if row is None:
            continue

        five_day = safe_float(row.get("five_day_change_percent")) or 0
        score += clamp(five_day, -4, 4) * 1.5

    vix = lookup.get("^VIX")
    vix_change = None

    if vix is not None:
        vix_change = safe_float(
            vix.get("five_day_change_percent")
        )
        score -= clamp(vix_change or 0, -10, 10) * 1.2

    dollar = lookup.get("DX-Y.NYB")
    dollar_change = None

    if dollar is not None:
        dollar_change = safe_float(
            dollar.get("five_day_change_percent")
        )
        score -= (dollar_change or 0) * 2

    yield_row = lookup.get("^TNX")
    yield_change = None

    if yield_row is not None:
        yield_change = safe_float(
            yield_row.get("five_day_change_percent")
        )
        score -= (yield_change or 0) * 1.5

    return {
        "global_risk_score": round(
            clamp(score, 0, 100),
            2,
        ),
        "vix_change_percent": vix_change,
        "dollar_change_percent": dollar_change,
        "yield_change_percent": yield_change,
    }


def classify_volatility(
    vix_change: float | None,
    average_adx: float | None,
    breadth_score: float,
) -> str:
    if vix_change is not None and vix_change >= 10:
        return "HIGH VOLATILITY"

    if vix_change is not None and vix_change <= -10:
        return "LOW VOLATILITY"

    if average_adx is not None and average_adx >= 30:
        return "TRENDING VOLATILITY"

    if breadth_score < 35:
        return "ELEVATED VOLATILITY"

    return "NORMAL VOLATILITY"


def classify_participation(
    bullish_percent: float,
    bearish_percent: float,
    breadth_score: float,
) -> str:
    if bullish_percent >= 65 and breadth_score >= 65:
        return "BROAD PARTICIPATION"

    if bearish_percent >= 65 and breadth_score <= 35:
        return "BROAD WEAKNESS"

    if bullish_percent >= 50:
        return "SELECTIVE PARTICIPATION"

    if bearish_percent >= 50:
        return "SELECTIVE WEAKNESS"

    return "MIXED PARTICIPATION"


def classify_risk_state(
    global_risk_score: float,
    breadth_score: float,
    sector_score: float,
) -> str:
    combined = (
        global_risk_score * 0.40
        + breadth_score * 0.35
        + sector_score * 0.25
    )

    if combined >= 65:
        return "RISK ON"

    if combined <= 35:
        return "RISK OFF"

    return "NEUTRAL RISK"


def strategy_for_regime(regime: str) -> tuple[str, float]:
    mapping = {
        "BULL TREND": ("TREND FOLLOWING / BUY ON DIPS", 80.0),
        "BEAR TREND": ("CAPITAL PROTECTION / SHORT BIAS", 20.0),
        "RANGE BOUND": ("MEAN REVERSION / REDUCED SIZE", 45.0),
        "ACCUMULATION": ("EARLY POSITION BUILDING", 60.0),
        "DISTRIBUTION": ("REDUCE EXPOSURE / TIGHT STOPS", 30.0),
        "HIGH VOLATILITY": ("SMALL SIZE / WIDE STOPS / SELECTIVE", 25.0),
        "LOW VOLATILITY": ("BREAKOUT WATCH / NORMAL SIZE", 60.0),
        "TRANSITION": ("WAIT FOR CONFIRMATION", 35.0),
        "RISK ON": ("TREND FOLLOWING / SECTOR LEADERS", 75.0),
        "RISK OFF": ("DEFENSIVE / CASH HEAVY", 20.0),
    }

    return mapping.get(
        regime,
        ("SELECTIVE SWING TRADING", 50.0),
    )


# ============================================================
# REGIME ENGINE
# ============================================================

def calculate_regime() -> dict:
    breadth = load_market_breadth()
    sectors = load_sector_rotation()
    structure = load_price_structure()
    global_frame = load_global_markets()
    master = load_unified_master()

    breadth_score = safe_float(
        breadth.get("breadth_score")
    ) or 50.0

    sector_score = (
        float(
            sectors["sector_rotation_score"]
            .dropna()
            .mean()
        )
        if not sectors.empty
        else 50.0
    )

    structure_score = (
        float(
            structure["structure_score"]
            .dropna()
            .mean()
        )
        if not structure.empty
        else 50.0
    )

    average_adx = (
        float(
            structure["adx_14"]
            .dropna()
            .mean()
        )
        if (
            not structure.empty
            and structure["adx_14"].notna().any()
        )
        else None
    )

    bullish_stock_percent = (
        float(
            structure["directional_bias"]
            .fillna("")
            .str.contains("BULLISH")
            .mean()
            * 100
        )
        if not structure.empty
        else 0.0
    )

    bearish_stock_percent = (
        float(
            structure["directional_bias"]
            .fillna("")
            .str.contains("BEARISH")
            .mean()
            * 100
        )
        if not structure.empty
        else 0.0
    )

    master_score = (
        float(master["master_score"].dropna().mean())
        if not master.empty
        else 50.0
    )

    global_data = global_component(global_frame)
    global_risk_score = global_data["global_risk_score"]

    trend_score = (
        structure_score * 0.30
        + sector_score * 0.20
        + breadth_score * 0.20
        + master_score * 0.15
        + global_risk_score * 0.15
    )

    bull_probability = clamp(
        (
            trend_score * 0.70
            + bullish_stock_percent * 0.30
        ),
        0,
        100,
    )

    bear_probability = clamp(
        (
            (100 - trend_score) * 0.70
            + bearish_stock_percent * 0.30
        ),
        0,
        100,
    )

    range_probability = clamp(
        100 - abs(trend_score - 50) * 2,
        0,
        100,
    )

    volatility_state = classify_volatility(
        global_data["vix_change_percent"],
        average_adx,
        breadth_score,
    )

    participation_state = classify_participation(
        bullish_stock_percent,
        bearish_stock_percent,
        breadth_score,
    )

    risk_state = classify_risk_state(
        global_risk_score,
        breadth_score,
        sector_score,
    )

    if "HIGH VOLATILITY" in volatility_state:
        market_regime = "HIGH VOLATILITY"

    elif bull_probability >= 70 and risk_state == "RISK ON":
        market_regime = "BULL TREND"

    elif bear_probability >= 70 and risk_state == "RISK OFF":
        market_regime = "BEAR TREND"

    elif (
        bull_probability >= 60
        and breadth_score < 55
    ):
        market_regime = "ACCUMULATION"

    elif (
        bear_probability >= 60
        and breadth_score > 45
    ):
        market_regime = "DISTRIBUTION"

    elif range_probability >= 65:
        market_regime = "RANGE BOUND"

    elif risk_state == "RISK ON":
        market_regime = "RISK ON"

    elif risk_state == "RISK OFF":
        market_regime = "RISK OFF"

    else:
        market_regime = "TRANSITION"

    suggested_strategy, exposure = strategy_for_regime(
        market_regime
    )

    explanation = " | ".join(
        [
            f"Breadth {breadth_score:.1f}",
            f"Sector rotation {sector_score:.1f}",
            f"Structure {structure_score:.1f}",
            f"Master {master_score:.1f}",
            f"Global risk {global_risk_score:.1f}",
            f"Bullish stocks {bullish_stock_percent:.1f}%",
            f"Bearish stocks {bearish_stock_percent:.1f}%",
            f"Participation {participation_state}",
            f"Risk state {risk_state}",
            f"Volatility {volatility_state}",
        ]
    )

    regime_score = round(
        clamp(
            (
                bull_probability
                if market_regime in {
                    "BULL TREND",
                    "ACCUMULATION",
                    "RISK ON",
                }
                else 100 - bear_probability
                if market_regime in {
                    "BEAR TREND",
                    "DISTRIBUTION",
                    "RISK OFF",
                }
                else trend_score
            ),
            0,
            100,
        ),
        2,
    )

    dates = []

    if breadth.get("trade_date"):
        dates.append(str(breadth["trade_date"]))

    if not sectors.empty:
        dates.extend(
            sectors["trade_date"]
            .dropna()
            .astype(str)
            .tolist()
        )

    regime_date = (
        max(dates)
        if dates
        else datetime.now().date().isoformat()
    )

    return {
        "regime_date": regime_date,
        "market_regime": market_regime,
        "regime_score": regime_score,
        "bull_probability": round(bull_probability, 2),
        "bear_probability": round(bear_probability, 2),
        "range_probability": round(range_probability, 2),
        "volatility_state": volatility_state,
        "risk_state": risk_state,
        "participation_state": participation_state,
        "breadth_score": round(breadth_score, 2),
        "sector_rotation_score": round(sector_score, 2),
        "structure_score": round(structure_score, 2),
        "global_risk_score": round(global_risk_score, 2),
        "vix_change_percent": global_data["vix_change_percent"],
        "dollar_change_percent": global_data["dollar_change_percent"],
        "yield_change_percent": global_data["yield_change_percent"],
        "bullish_stock_percent": round(
            bullish_stock_percent,
            2,
        ),
        "bearish_stock_percent": round(
            bearish_stock_percent,
            2,
        ),
        "suggested_strategy": suggested_strategy,
        "capital_exposure_percent": exposure,
        "explanation": explanation,
    }


def save_result(result: dict) -> None:
    with connect() as connection:
        connection.execute(
            """
            INSERT INTO market_regime_intelligence(
                regime_date,
                market_regime,
                regime_score,
                bull_probability,
                bear_probability,
                range_probability,
                volatility_state,
                risk_state,
                participation_state,
                breadth_score,
                sector_rotation_score,
                structure_score,
                global_risk_score,
                vix_change_percent,
                dollar_change_percent,
                yield_change_percent,
                bullish_stock_percent,
                bearish_stock_percent,
                suggested_strategy,
                capital_exposure_percent,
                explanation,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(regime_date)
            DO UPDATE SET
                market_regime=excluded.market_regime,
                regime_score=excluded.regime_score,
                bull_probability=excluded.bull_probability,
                bear_probability=excluded.bear_probability,
                range_probability=excluded.range_probability,
                volatility_state=excluded.volatility_state,
                risk_state=excluded.risk_state,
                participation_state=excluded.participation_state,
                breadth_score=excluded.breadth_score,
                sector_rotation_score=excluded.sector_rotation_score,
                structure_score=excluded.structure_score,
                global_risk_score=excluded.global_risk_score,
                vix_change_percent=excluded.vix_change_percent,
                dollar_change_percent=excluded.dollar_change_percent,
                yield_change_percent=excluded.yield_change_percent,
                bullish_stock_percent=excluded.bullish_stock_percent,
                bearish_stock_percent=excluded.bearish_stock_percent,
                suggested_strategy=excluded.suggested_strategy,
                capital_exposure_percent=excluded.capital_exposure_percent,
                explanation=excluded.explanation,
                created_at=excluded.created_at
            """,
            (
                result["regime_date"],
                result["market_regime"],
                result["regime_score"],
                result["bull_probability"],
                result["bear_probability"],
                result["range_probability"],
                result["volatility_state"],
                result["risk_state"],
                result["participation_state"],
                result["breadth_score"],
                result["sector_rotation_score"],
                result["structure_score"],
                result["global_risk_score"],
                result["vix_change_percent"],
                result["dollar_change_percent"],
                result["yield_change_percent"],
                result["bullish_stock_percent"],
                result["bearish_stock_percent"],
                result["suggested_strategy"],
                result["capital_exposure_percent"],
                result["explanation"],
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )

        connection.commit()


def run_engine() -> dict:
    setup_schema()

    run_id = start_run(
        "aqsd_market_regime",
        "Calculating current market regime",
    )

    try:
        result = calculate_regime()
        save_result(result)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=1,
            errors_count=0,
            message=(
                f"Regime={result['market_regime']}; "
                f"score={result['regime_score']}"
            ),
        )

        return result

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


# ============================================================
# REPORTING
# ============================================================

def latest_result() -> dict | None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else None


def write_report(
    result: dict | None = None,
) -> None:
    if result is None:
        result = latest_result()

    if not result:
        raise RuntimeError(
            "No Market Regime result found. Run --run first."
        )

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Market Regime Intelligence" in wb.sheetnames:
        del wb["Market Regime Intelligence"]

    ws = wb.create_sheet(
        "Market Regime Intelligence",
        1,
    )
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H2")
    ws["A1"] = "AQSD PROFESSIONAL - MARKET REGIME INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    regime = str(result["market_regime"])

    ws["A4"] = "Current Regime"
    ws["B4"] = regime
    ws["D4"] = "Regime Score"
    ws["E4"] = result["regime_score"]
    ws["G4"] = "Date"
    ws["H4"] = result["regime_date"]

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    ws["B4"].font = Font(bold=True)

    ws["B4"].fill = PatternFill(
        "solid",
        fgColor=(
            GREEN
            if regime in {
                "BULL TREND",
                "ACCUMULATION",
                "RISK ON",
            }
            else RED
            if regime in {
                "BEAR TREND",
                "DISTRIBUTION",
                "RISK OFF",
                "HIGH VOLATILITY",
            }
            else YELLOW
        ),
    )

    rows = [
        ("Bull Probability", result["bull_probability"]),
        ("Bear Probability", result["bear_probability"]),
        ("Range Probability", result["range_probability"]),
        ("Volatility State", result["volatility_state"]),
        ("Risk State", result["risk_state"]),
        ("Participation State", result["participation_state"]),
        ("Breadth Score", result["breadth_score"]),
        ("Sector Rotation Score", result["sector_rotation_score"]),
        ("Structure Score", result["structure_score"]),
        ("Global Risk Score", result["global_risk_score"]),
        ("VIX 5D Change %", result["vix_change_percent"]),
        ("Dollar 5D Change %", result["dollar_change_percent"]),
        ("US 10Y 5D Change %", result["yield_change_percent"]),
        ("Bullish Stocks %", result["bullish_stock_percent"]),
        ("Bearish Stocks %", result["bearish_stock_percent"]),
        ("Suggested Strategy", result["suggested_strategy"]),
        ("Capital Exposure %", result["capital_exposure_percent"]),
        ("Explanation", result["explanation"]),
    ]

    start_row = 7

    for index, (label, value) in enumerate(
        rows,
        start=start_row,
    ):
        ws.cell(index, 1, label)
        ws.cell(index, 2, value)

        ws.cell(index, 1).font = Font(
            bold=True
        )
        ws.cell(index, 1).fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )
        ws.cell(index, 1).border = Border(
            bottom=THIN
        )
        ws.cell(index, 2).border = Border(
            bottom=THIN
        )

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    wb.save(DASHBOARD)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                MIN(regime_date) AS first_date,
                MAX(regime_date) AS latest_date
            FROM market_regime_intelligence
            """
        ).fetchone()

    print("\nAQSD MARKET REGIME STATUS")
    print("=" * 72)
    print(f"Stored records: {row['total'] or 0}")
    print(f"First date:     {row['first_date'] or 'No data'}")
    print(f"Latest date:    {row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Market Regime Intelligence Engine."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Calculate the latest market regime.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild Excel report from stored data.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show Market Regime status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.run:
        result = run_engine()
        write_report(result)

        print("\nAQSD MARKET REGIME")
        print("=" * 72)
        print(f"Regime:            {result['market_regime']}")
        print(f"Regime Score:      {result['regime_score']}")
        print(f"Bull Probability:  {result['bull_probability']}%")
        print(f"Bear Probability:  {result['bear_probability']}%")
        print(f"Range Probability: {result['range_probability']}%")
        print(f"Risk State:        {result['risk_state']}")
        print(f"Volatility:        {result['volatility_state']}")
        print(f"Strategy:          {result['suggested_strategy']}")
        print(
            f"Capital Exposure:  "
            f"{result['capital_exposure_percent']}%"
        )
        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report()
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
