
"""
AQSD Professional
Module: Portfolio Heat Map
Version: 1.0

Creates a visual Portfolio Heat Map sheet from the current Portfolio sheet.

Features
--------
- Open positions summary
- Winner / loser / near-stop classification
- P/L heat map
- Capital exposure
- Risk utilization
- Distance to stop and target
- Best and worst open position
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
DARK_GREEN = "006100"
RED = "FFC7CE"
DARK_RED = "9C0006"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
GREY = "E7E6E6"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def distance_percent(
    current: float,
    reference: float,
) -> float:
    if reference == 0:
        return 0.0

    return (current - reference) / reference * 100


def classify_position(
    side: str,
    cmp_price: float,
    entry: float,
    stop_loss: float,
    target: float,
    pl_percent: float,
) -> str:
    side = side.upper().strip()

    if side in {"BUY", "CALL"}:
        stop_gap = (
            (cmp_price - stop_loss) / cmp_price * 100
            if cmp_price
            else 0.0
        )

        target_gap = (
            (target - cmp_price) / cmp_price * 100
            if cmp_price
            else 0.0
        )
    else:
        stop_gap = (
            (stop_loss - cmp_price) / cmp_price * 100
            if cmp_price
            else 0.0
        )

        target_gap = (
            (cmp_price - target) / cmp_price * 100
            if cmp_price
            else 0.0
        )

    if stop_gap <= 1:
        return "NEAR STOP"

    if target_gap <= 1:
        return "NEAR TARGET"

    if pl_percent > 0.5:
        return "WINNER"

    if pl_percent < -0.5:
        return "LOSER"

    return "NEUTRAL"


def read_positions(wb) -> list[dict]:
    if "Portfolio" not in wb.sheetnames:
        raise RuntimeError(
            "Portfolio sheet not found. "
            "Run portfolio_manager.py first."
        )

    ws = wb["Portfolio"]
    headers = header_map(ws, 12)

    required = [
        "Trade ID",
        "Symbol",
        "Side",
        "Qty",
        "Entry",
        "CMP",
        "Stop Loss",
        "Capital Used",
        "Risk Amount",
        "P/L",
        "P/L %",
        "Target",
        "Status",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            "Missing Portfolio columns: "
            + ", ".join(missing)
        )

    positions: list[dict] = []

    for row in range(13, ws.max_row + 1):
        symbol = ws.cell(row, headers["Symbol"]).value

        if not symbol:
            continue

        status = str(
            ws.cell(row, headers["Status"]).value or ""
        ).upper().strip()

        if status != "OPEN":
            continue

        side = str(
            ws.cell(row, headers["Side"]).value or ""
        ).upper().strip()

        entry = safe_float(
            ws.cell(row, headers["Entry"]).value
        )
        cmp_price = safe_float(
            ws.cell(row, headers["CMP"]).value
        )
        stop_loss = safe_float(
            ws.cell(row, headers["Stop Loss"]).value
        )
        target = safe_float(
            ws.cell(row, headers["Target"]).value
        )
        pl_percent = safe_float(
            ws.cell(row, headers["P/L %"]).value
        )

        classification = classify_position(
            side,
            cmp_price,
            entry,
            stop_loss,
            target,
            pl_percent,
        )

        if side in {"BUY", "CALL"}:
            stop_distance = distance_percent(
                cmp_price,
                stop_loss,
            )
            target_distance = distance_percent(
                target,
                cmp_price,
            )
        else:
            stop_distance = distance_percent(
                stop_loss,
                cmp_price,
            )
            target_distance = distance_percent(
                cmp_price,
                target,
            )

        positions.append(
            {
                "Trade ID": ws.cell(
                    row,
                    headers["Trade ID"],
                ).value,
                "Symbol": str(symbol),
                "Side": side,
                "Qty": safe_float(
                    ws.cell(row, headers["Qty"]).value
                ),
                "Entry": entry,
                "CMP": cmp_price,
                "Stop Loss": stop_loss,
                "Target": target,
                "Capital Used": safe_float(
                    ws.cell(
                        row,
                        headers["Capital Used"],
                    ).value
                ),
                "Risk Amount": safe_float(
                    ws.cell(
                        row,
                        headers["Risk Amount"],
                    ).value
                ),
                "P/L": safe_float(
                    ws.cell(row, headers["P/L"]).value
                ),
                "P/L %": pl_percent,
                "Stop Distance %": stop_distance,
                "Target Distance %": target_distance,
                "Classification": classification,
            }
        )

    return positions


# ============================================================
# SHEET CREATION
# ============================================================

def create_heatmap_sheet(
    wb,
    positions: list[dict],
) -> None:
    if "Portfolio Heat Map" in wb.sheetnames:
        del wb["Portfolio Heat Map"]

    ws = wb.create_sheet("Portfolio Heat Map", 2)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    ws.merge_cells("A1:N2")
    ws["A1"] = "AQSD PROFESSIONAL - PORTFOLIO HEAT MAP"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    total_positions = len(positions)
    total_capital = sum(
        item["Capital Used"]
        for item in positions
    )
    total_risk = sum(
        item["Risk Amount"]
        for item in positions
    )
    total_pl = sum(
        item["P/L"]
        for item in positions
    )

    winners = sum(
        1
        for item in positions
        if item["Classification"] == "WINNER"
    )
    losers = sum(
        1
        for item in positions
        if item["Classification"] == "LOSER"
    )
    near_stop = sum(
        1
        for item in positions
        if item["Classification"] == "NEAR STOP"
    )

    best_trade = max(
        positions,
        key=lambda item: item["P/L"],
        default=None,
    )
    worst_trade = min(
        positions,
        key=lambda item: item["P/L"],
        default=None,
    )

    summary = [
        ("Last Updated", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("Open Positions", total_positions),
        ("Capital Used", total_capital),
        ("Total Risk", total_risk),
        ("Open MTM", total_pl),
        ("Winners", winners),
        ("Losers", losers),
        ("Near Stop", near_stop),
        (
            "Best Position",
            best_trade["Symbol"]
            if best_trade
            else "",
        ),
        (
            "Worst Position",
            worst_trade["Symbol"]
            if worst_trade
            else "",
        ),
    ]

    for row_no, (label, value) in enumerate(
        summary,
        start=4,
    ):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value

        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )
        ws[f"A{row_no}"].border = Border(bottom=THIN)
        ws[f"B{row_no}"].border = Border(bottom=THIN)

    for ref in ("B6", "B7", "B8"):
        ws[ref].number_format = '₹#,##0.00'

    ws["B9"].number_format = "0"
    ws["B10"].number_format = "0"
    ws["B11"].number_format = "0"

    header_row = 15

    headers = [
        "Rank",
        "Trade ID",
        "Symbol",
        "Side",
        "Qty",
        "Entry",
        "CMP",
        "P/L",
        "P/L %",
        "Capital Used",
        "Risk Amount",
        "Stop Dist %",
        "Target Dist %",
        "Heat Status",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    sorted_positions = sorted(
        positions,
        key=lambda item: item["P/L %"],
        reverse=True,
    )

    first_row = header_row + 1

    for index, item in enumerate(
        sorted_positions,
        start=1,
    ):
        row = first_row + index - 1

        values = [
            index,
            item["Trade ID"],
            item["Symbol"],
            item["Side"],
            item["Qty"],
            item["Entry"],
            item["CMP"],
            item["P/L"],
            item["P/L %"],
            item["Capital Used"],
            item["Risk Amount"],
            item["Stop Distance %"],
            item["Target Distance %"],
            item["Classification"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = Border(bottom=THIN)

        ws.cell(row, 1).fill = PatternFill(
            "solid",
            fgColor=YELLOW,
        )

        classification = item["Classification"]

        if classification == "WINNER":
            fill = GREEN
            font_color = DARK_GREEN
        elif classification == "LOSER":
            fill = RED
            font_color = DARK_RED
        elif classification == "NEAR STOP":
            fill = RED
            font_color = DARK_RED
        elif classification == "NEAR TARGET":
            fill = GREEN
            font_color = DARK_GREEN
        else:
            fill = GREY
            font_color = "666666"

        ws.cell(row, 14).fill = PatternFill(
            "solid",
            fgColor=fill,
        )
        ws.cell(row, 14).font = Font(
            bold=True,
            color=font_color,
        )

        for col in [6, 7, 8, 10, 11]:
            ws.cell(row, col).number_format = '₹#,##0.00'

        for col in [9, 12, 13]:
            ws.cell(row, col).number_format = '0.00"%"'

    if positions:
        last_row = first_row + len(positions) - 1

        ws.conditional_formatting.add(
            f"H{first_row}:H{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

        ws.conditional_formatting.add(
            f"I{first_row}:I{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="num",
                mid_value=0,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    widths = {
        "A": 8,
        "B": 10,
        "C": 18,
        "D": 10,
        "E": 10,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 10,
        "J": 15,
        "K": 14,
        "L": 12,
        "M": 12,
        "N": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)
    positions = read_positions(wb)

    create_heatmap_sheet(
        wb,
        positions,
    )

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Portfolio Heat Map created successfully.")
    print(f"Open positions analysed: {len(positions)}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
