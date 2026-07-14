
"""
AQSD Professional
Module: Drawdown Analyzer
Version: 1.0

Creates a Drawdown Analysis sheet from completed trades in Trade Journal.

Features
--------
- Cumulative equity curve
- Running peak
- Drawdown amount
- Drawdown percentage
- Maximum drawdown
- Maximum drawdown percentage
- Longest drawdown duration
- Recovery status
- Underwater curve chart
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def read_closed_trades(wb) -> list[dict]:
    if "Trade Journal" not in wb.sheetnames:
        raise RuntimeError("Trade Journal sheet not found.")

    ws = wb["Trade Journal"]
    headers = header_map(ws, 10)

    required = ["Symbol", "Exit Date", "P/L"]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            "Missing Trade Journal columns: " + ", ".join(missing)
        )

    trades: list[dict] = []

    for row in range(11, ws.max_row + 1):
        symbol = ws.cell(row, headers["Symbol"]).value

        if not symbol:
            continue

        trades.append(
            {
                "Symbol": str(symbol),
                "Exit Date": ws.cell(row, headers["Exit Date"]).value,
                "P/L": safe_float(ws.cell(row, headers["P/L"]).value),
            }
        )

    return trades


def build_drawdown_rows(trades: list[dict]) -> tuple[list[dict], dict]:
    rows: list[dict] = []

    equity = 0.0
    peak = 0.0

    max_drawdown = 0.0
    max_drawdown_percent = 0.0

    current_drawdown_days = 0
    longest_drawdown_days = 0
    recovery_status = "NO DRAWDOWN"

    for index, trade in enumerate(trades, start=1):
        equity += trade["P/L"]
        peak = max(peak, equity)

        drawdown = equity - peak
        drawdown_percent = (
            drawdown / peak * 100
            if peak > 0
            else 0.0
        )

        if drawdown < 0:
            current_drawdown_days += 1
            recovery_status = "IN DRAWDOWN"
        else:
            current_drawdown_days = 0
            recovery_status = "RECOVERED"

        longest_drawdown_days = max(
            longest_drawdown_days,
            current_drawdown_days,
        )

        max_drawdown = min(max_drawdown, drawdown)
        max_drawdown_percent = min(
            max_drawdown_percent,
            drawdown_percent,
        )

        rows.append(
            {
                "Trade No.": index,
                "Exit Date": trade["Exit Date"],
                "Symbol": trade["Symbol"],
                "Trade P/L": trade["P/L"],
                "Equity": equity,
                "Peak Equity": peak,
                "Drawdown": drawdown,
                "Drawdown %": drawdown_percent,
                "Drawdown Duration": current_drawdown_days,
                "Status": recovery_status,
            }
        )

    summary = {
        "Closed Trades": len(trades),
        "Ending Equity P/L": equity,
        "Peak Equity P/L": peak,
        "Maximum Drawdown": abs(max_drawdown),
        "Maximum Drawdown %": abs(max_drawdown_percent),
        "Longest Drawdown Duration": longest_drawdown_days,
        "Current Status": recovery_status if trades else "NO DATA",
    }

    return rows, summary


def create_drawdown_sheet(
    wb,
    rows: list[dict],
    summary: dict,
) -> None:
    if "Drawdown Analysis" in wb.sheetnames:
        del wb["Drawdown Analysis"]

    ws = wb.create_sheet("Drawdown Analysis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A14"

    ws.merge_cells("A1:J2")
    ws["A1"] = "AQSD PROFESSIONAL - DRAWDOWN ANALYSIS"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    for row_no, (label, value) in enumerate(
        summary.items(),
        start=4,
    ):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )
        ws[f"A{row_no}"].border = Border(bottom=THIN)
        ws[f"B{row_no}"].border = Border(bottom=THIN)

        if label in {
            "Ending Equity P/L",
            "Peak Equity P/L",
            "Maximum Drawdown",
        }:
            ws[f"B{row_no}"].number_format = '₹#,##0.00'
        elif label == "Maximum Drawdown %":
            ws[f"B{row_no}"].number_format = '0.00"%"'

    status_cell = ws["B10"]

    if summary["Current Status"] == "RECOVERED":
        status_cell.fill = PatternFill("solid", fgColor=GREEN)
    elif summary["Current Status"] == "IN DRAWDOWN":
        status_cell.fill = PatternFill("solid", fgColor=RED)
    else:
        status_cell.fill = PatternFill("solid", fgColor=YELLOW)

    headers = [
        "Trade No.",
        "Exit Date",
        "Symbol",
        "Trade P/L",
        "Equity",
        "Peak Equity",
        "Drawdown",
        "Drawdown %",
        "Drawdown Duration",
        "Status",
    ]

    header_row = 13

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, item in enumerate(rows, start=14):
        values = [
            item["Trade No."],
            item["Exit Date"],
            item["Symbol"],
            item["Trade P/L"],
            item["Equity"],
            item["Peak Equity"],
            item["Drawdown"],
            item["Drawdown %"],
            item["Drawdown Duration"],
            item["Status"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            cell.border = Border(bottom=THIN)

        ws.cell(row_no, 2).number_format = "dd-mm-yyyy"

        for col in [4, 5, 6, 7]:
            ws.cell(row_no, col).number_format = '₹#,##0.00'

        ws.cell(row_no, 8).number_format = '0.00"%"'

        if item["Drawdown"] < 0:
            ws.cell(row_no, 7).fill = PatternFill(
                "solid",
                fgColor=RED,
            )
            ws.cell(row_no, 8).fill = PatternFill(
                "solid",
                fgColor=RED,
            )
        else:
            ws.cell(row_no, 7).fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )
            ws.cell(row_no, 8).fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )

        ws.cell(row_no, 10).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if item["Status"] == "RECOVERED"
                else RED
            ),
        )
        ws.cell(row_no, 10).font = Font(bold=True)

    if rows:
        last_row = 13 + len(rows)

        chart = LineChart()
        chart.title = "Underwater Curve"
        chart.y_axis.title = "Drawdown %"
        chart.x_axis.title = "Trade Number"

        data = Reference(
            ws,
            min_col=8,
            min_row=13,
            max_row=last_row,
        )

        categories = Reference(
            ws,
            min_col=1,
            min_row=14,
            max_row=last_row,
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 15

        ws.add_chart(chart, "L4")

    widths = {
        "A": 12,
        "B": 14,
        "C": 18,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 12,
        "I": 18,
        "J": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    trades = read_closed_trades(wb)
    rows, summary = build_drawdown_rows(trades)

    create_drawdown_sheet(
        wb,
        rows,
        summary,
    )

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Drawdown Analysis created successfully.")
    print(f"Closed trades analysed: {len(trades)}")
    print(f"Maximum drawdown: ₹{summary['Maximum Drawdown']:,.2f}")
    print(f"Maximum drawdown %: {summary['Maximum Drawdown %']:.2f}%")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
