
"""
AQSD Professional
Module: Capital Protection Engine
Version: 1.0

Purpose
-------
Applies a portfolio-level risk overlay to AQSD trade decisions.

Inputs
------
- Market Regime Intelligence
- Dynamic Strategy Selector
- Trade Decision Engine
- Portfolio Allocation
- Market Breadth Intelligence

Outputs
-------
- Maximum portfolio exposure
- Maximum position size
- Maximum sector exposure
- Adjusted allocation percentages
- Risk flags
- Capital-protection status
- SQLite history
- CSV report
- Excel sheet: Capital Protection

Commands
--------
python aqsd_capital_protection.py --run
python aqsd_capital_protection.py --status
python aqsd_capital_protection.py --report
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "Output"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"
CSV_REPORT = OUTPUT_DIR / "AQSD_Capital_Protection.csv"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


SCHEMA = """
CREATE TABLE IF NOT EXISTS capital_protection_summary (
    summary_id INTEGER PRIMARY KEY AUTOINCREMENT,
    protection_date TEXT NOT NULL UNIQUE,
    market_regime TEXT,
    protection_status TEXT,
    maximum_portfolio_exposure REAL,
    maximum_position_size REAL,
    maximum_sector_exposure REAL,
    maximum_concurrent_positions INTEGER,
    cash_reserve_percent REAL,
    stop_loss_policy TEXT,
    portfolio_risk_score REAL,
    explanation TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS capital_protection_allocations (
    allocation_id INTEGER PRIMARY KEY AUTOINCREMENT,
    protection_date TEXT NOT NULL,
    nse_symbol TEXT NOT NULL,
    sector TEXT,
    action TEXT,
    original_allocation_percent REAL,
    adjusted_allocation_percent REAL,
    master_score REAL,
    confidence_percent REAL,
    risk_level TEXT,
    sector_exposure_percent REAL,
    risk_flag TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(protection_date, nse_symbol)
);

CREATE INDEX IF NOT EXISTS idx_capital_protection_date
ON capital_protection_allocations(protection_date);
"""


def setup_schema() -> None:
    setup_database()
    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND name=?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def safe_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value is None or value == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def load_regime() -> dict:
    with connect() as connection:
        if not table_exists(connection, "market_regime_intelligence"):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def load_strategy() -> dict:
    with connect() as connection:
        if not table_exists(connection, "aqsd_strategy_selector"):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM aqsd_strategy_selector
            ORDER BY selection_date DESC, selector_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def load_decisions() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "aqsd_trade_decisions"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM aqsd_trade_decisions
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
            ORDER BY priority_rank
            """,
            connection,
        )


def load_allocations() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "portfolio_allocation"):
            return pd.DataFrame()

        columns = {
            row["name"]
            for row in connection.execute(
                "PRAGMA table_info(portfolio_allocation)"
            ).fetchall()
        }

        symbol_column = (
            "symbol"
            if "symbol" in columns
            else "nse_symbol"
        )

        return pd.read_sql_query(
            f"""
            SELECT
                trade_date,
                {symbol_column} AS nse_symbol,
                action,
                priority_rank,
                allocation_percent,
                conviction
            FROM portfolio_allocation
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM portfolio_allocation
            )
            ORDER BY allocation_percent DESC
            """,
            connection,
        )


def protection_rules(
    regime: str,
    risk_state: str,
    volatility_state: str,
) -> dict:
    maximum_portfolio_exposure = 70.0
    maximum_position_size = 12.0
    maximum_sector_exposure = 25.0
    maximum_concurrent_positions = 7
    stop_loss_policy = "ATR OR STRUCTURE STOP"

    if regime in {"BEAR TREND", "RISK OFF"}:
        maximum_portfolio_exposure = 25.0
        maximum_position_size = 6.0
        maximum_sector_exposure = 12.0
        maximum_concurrent_positions = 3
        stop_loss_policy = "TIGHT STRUCTURE STOP"

    elif regime == "HIGH VOLATILITY":
        maximum_portfolio_exposure = 30.0
        maximum_position_size = 5.0
        maximum_sector_exposure = 12.0
        maximum_concurrent_positions = 3
        stop_loss_policy = "WIDE ATR STOP WITH SMALL POSITION"

    elif regime in {"DISTRIBUTION", "TRANSITION"}:
        maximum_portfolio_exposure = 40.0
        maximum_position_size = 8.0
        maximum_sector_exposure = 18.0
        maximum_concurrent_positions = 4
        stop_loss_policy = "TIGHT TRAILING STOP"

    elif regime in {"BULL TREND", "RISK ON"}:
        maximum_portfolio_exposure = 80.0
        maximum_position_size = 15.0
        maximum_sector_exposure = 30.0
        maximum_concurrent_positions = 8
        stop_loss_policy = "ATR TRAILING STOP"

    elif regime == "ACCUMULATION":
        maximum_portfolio_exposure = 60.0
        maximum_position_size = 10.0
        maximum_sector_exposure = 22.0
        maximum_concurrent_positions = 6
        stop_loss_policy = "SWING-LOW STOP"

    if risk_state == "RISK OFF":
        maximum_portfolio_exposure *= 0.80
        maximum_position_size *= 0.85

    if "HIGH" in volatility_state:
        maximum_portfolio_exposure *= 0.80
        maximum_position_size *= 0.80

    return {
        "maximum_portfolio_exposure": round(
            clamp(maximum_portfolio_exposure, 10, 90),
            2,
        ),
        "maximum_position_size": round(
            clamp(maximum_position_size, 3, 20),
            2,
        ),
        "maximum_sector_exposure": round(
            clamp(maximum_sector_exposure, 8, 35),
            2,
        ),
        "maximum_concurrent_positions": int(
            clamp(maximum_concurrent_positions, 1, 12)
        ),
        "stop_loss_policy": stop_loss_policy,
    }


def build_protection() -> tuple[dict, list[dict]]:
    regime = load_regime()
    strategy = load_strategy()
    decisions = load_decisions()
    allocations = load_allocations()

    if decisions.empty:
        raise RuntimeError(
            "No Trade Decision data found. "
            "Run aqsd_decision_engine.py --run first."
        )

    regime_name = str(
        regime.get("market_regime") or "TRANSITION"
    )
    risk_state = str(
        regime.get("risk_state") or "NEUTRAL RISK"
    )
    volatility_state = str(
        regime.get("volatility_state") or "NORMAL VOLATILITY"
    )

    rules = protection_rules(
        regime_name,
        risk_state,
        volatility_state,
    )

    strategy_exposure = safe_float(
        strategy.get("capital_exposure_percent"),
        rules["maximum_portfolio_exposure"],
    ) or rules["maximum_portfolio_exposure"]

    maximum_portfolio_exposure = min(
        rules["maximum_portfolio_exposure"],
        strategy_exposure,
    )

    actionable = decisions[
        decisions["action"].isin(
            ["STRONG BUY", "BUY", "BUY ON DIP"]
        )
    ].copy()

    if actionable.empty:
        summary = {
            "protection_date": str(
                regime.get("regime_date")
                or datetime.now().date().isoformat()
            ),
            "market_regime": regime_name,
            "protection_status": "CASH / NO ACTIONABLE LONGS",
            **rules,
            "maximum_portfolio_exposure": 0.0,
            "cash_reserve_percent": 100.0,
            "portfolio_risk_score": 100.0,
            "explanation": (
                "No actionable long trades were found in the latest "
                "Decision Engine output."
            ),
        }
        return summary, []

    if allocations.empty:
        actionable["allocation_percent"] = (
            100 / len(actionable)
        )
    else:
        actionable = actionable.merge(
            allocations[
                [
                    "nse_symbol",
                    "allocation_percent",
                ]
            ],
            on="nse_symbol",
            how="left",
        )

        actionable["allocation_percent"] = (
            actionable["allocation_percent"]
            .fillna(0)
        )

        if actionable["allocation_percent"].sum() == 0:
            actionable["allocation_percent"] = (
                100 / len(actionable)
            )

    raw_total = float(
        actionable["allocation_percent"].sum()
    )

    if raw_total <= 0:
        raw_total = 100.0

    actionable["scaled_allocation"] = (
        actionable["allocation_percent"]
        / raw_total
        * maximum_portfolio_exposure
    )

    actionable["scaled_allocation"] = (
        actionable["scaled_allocation"]
        .clip(
            upper=rules["maximum_position_size"]
        )
    )

    actionable = actionable.sort_values(
        [
            "priority_rank",
            "priority_score",
        ],
        ascending=[True, False],
    ).head(
        rules["maximum_concurrent_positions"]
    )

    sector_totals: dict[str, float] = {}
    adjusted_rows: list[dict] = []

    for _, row in actionable.iterrows():
        sector = str(row.get("sector") or "Unmapped")
        requested = float(
            row.get("scaled_allocation") or 0
        )

        current_sector = sector_totals.get(
            sector,
            0.0,
        )

        remaining_sector_capacity = max(
            0.0,
            rules["maximum_sector_exposure"]
            - current_sector,
        )

        adjusted = min(
            requested,
            remaining_sector_capacity,
        )

        risk_level = str(
            row.get("risk_level") or "HIGH"
        )

        if risk_level == "HIGH":
            adjusted *= 0.70
        elif risk_level == "MEDIUM":
            adjusted *= 0.90

        confidence = safe_float(
            row.get("confidence_percent"),
            0,
        ) or 0

        if confidence < 55:
            adjusted *= 0.80

        adjusted = round(
            max(0.0, adjusted),
            2,
        )

        sector_totals[sector] = (
            current_sector + adjusted
        )

        flags = []

        if requested > rules["maximum_position_size"]:
            flags.append("POSITION CAPPED")

        if remaining_sector_capacity < requested:
            flags.append("SECTOR CAP APPLIED")

        if risk_level == "HIGH":
            flags.append("HIGH RISK REDUCTION")

        if confidence < 55:
            flags.append("LOW CONFIDENCE REDUCTION")

        adjusted_rows.append(
            {
                "protection_date": str(
                    regime.get("regime_date")
                    or datetime.now().date().isoformat()
                ),
                "nse_symbol": str(
                    row.get("nse_symbol")
                ),
                "sector": sector,
                "action": str(
                    row.get("action")
                ),
                "original_allocation_percent": round(
                    float(
                        row.get("allocation_percent")
                        or 0
                    ),
                    2,
                ),
                "adjusted_allocation_percent": adjusted,
                "master_score": safe_float(
                    row.get("master_score")
                ),
                "confidence_percent": confidence,
                "risk_level": risk_level,
                "sector_exposure_percent": round(
                    sector_totals[sector],
                    2,
                ),
                "risk_flag": (
                    " | ".join(flags)
                    if flags
                    else "OK"
                ),
            }
        )

    total_adjusted = round(
        sum(
            row["adjusted_allocation_percent"]
            for row in adjusted_rows
        ),
        2,
    )

    cash_reserve = round(
        100 - total_adjusted,
        2,
    )

    high_risk_count = sum(
        row["risk_level"] == "HIGH"
        for row in adjusted_rows
    )

    low_confidence_count = sum(
        row["confidence_percent"] < 55
        for row in adjusted_rows
    )

    risk_score = (
        high_risk_count * 15
        + low_confidence_count * 10
        + max(
            0,
            total_adjusted
            - maximum_portfolio_exposure,
        )
        * 2
    )

    risk_score = round(
        clamp(risk_score, 0, 100),
        2,
    )

    if risk_score >= 70:
        protection_status = "CRITICAL PROTECTION"
    elif risk_score >= 40:
        protection_status = "DEFENSIVE"
    elif risk_score >= 20:
        protection_status = "CAUTIOUS"
    else:
        protection_status = "NORMAL"

    summary = {
        "protection_date": str(
            regime.get("regime_date")
            or datetime.now().date().isoformat()
        ),
        "market_regime": regime_name,
        "protection_status": protection_status,
        "maximum_portfolio_exposure": round(
            maximum_portfolio_exposure,
            2,
        ),
        "maximum_position_size": rules[
            "maximum_position_size"
        ],
        "maximum_sector_exposure": rules[
            "maximum_sector_exposure"
        ],
        "maximum_concurrent_positions": rules[
            "maximum_concurrent_positions"
        ],
        "cash_reserve_percent": cash_reserve,
        "stop_loss_policy": rules[
            "stop_loss_policy"
        ],
        "portfolio_risk_score": risk_score,
        "explanation": " | ".join(
            [
                f"Regime {regime_name}",
                f"Risk state {risk_state}",
                f"Volatility {volatility_state}",
                f"Allowed exposure {maximum_portfolio_exposure:.1f}%",
                f"Adjusted exposure {total_adjusted:.1f}%",
                f"Cash reserve {cash_reserve:.1f}%",
                f"Positions {len(adjusted_rows)}",
            ]
        ),
    }

    return summary, adjusted_rows


def save_results(
    summary: dict,
    allocations: list[dict],
) -> None:
    with connect() as connection:
        connection.execute(
            """
            INSERT INTO capital_protection_summary(
                protection_date,
                market_regime,
                protection_status,
                maximum_portfolio_exposure,
                maximum_position_size,
                maximum_sector_exposure,
                maximum_concurrent_positions,
                cash_reserve_percent,
                stop_loss_policy,
                portfolio_risk_score,
                explanation,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(protection_date)
            DO UPDATE SET
                market_regime=excluded.market_regime,
                protection_status=excluded.protection_status,
                maximum_portfolio_exposure=
                    excluded.maximum_portfolio_exposure,
                maximum_position_size=
                    excluded.maximum_position_size,
                maximum_sector_exposure=
                    excluded.maximum_sector_exposure,
                maximum_concurrent_positions=
                    excluded.maximum_concurrent_positions,
                cash_reserve_percent=
                    excluded.cash_reserve_percent,
                stop_loss_policy=
                    excluded.stop_loss_policy,
                portfolio_risk_score=
                    excluded.portfolio_risk_score,
                explanation=excluded.explanation,
                created_at=excluded.created_at
            """,
            (
                summary["protection_date"],
                summary["market_regime"],
                summary["protection_status"],
                summary[
                    "maximum_portfolio_exposure"
                ],
                summary["maximum_position_size"],
                summary["maximum_sector_exposure"],
                summary[
                    "maximum_concurrent_positions"
                ],
                summary["cash_reserve_percent"],
                summary["stop_loss_policy"],
                summary["portfolio_risk_score"],
                summary["explanation"],
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )

        for row in allocations:
            connection.execute(
                """
                INSERT INTO capital_protection_allocations(
                    protection_date,
                    nse_symbol,
                    sector,
                    action,
                    original_allocation_percent,
                    adjusted_allocation_percent,
                    master_score,
                    confidence_percent,
                    risk_level,
                    sector_exposure_percent,
                    risk_flag,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(protection_date, nse_symbol)
                DO UPDATE SET
                    sector=excluded.sector,
                    action=excluded.action,
                    original_allocation_percent=
                        excluded.original_allocation_percent,
                    adjusted_allocation_percent=
                        excluded.adjusted_allocation_percent,
                    master_score=excluded.master_score,
                    confidence_percent=
                        excluded.confidence_percent,
                    risk_level=excluded.risk_level,
                    sector_exposure_percent=
                        excluded.sector_exposure_percent,
                    risk_flag=excluded.risk_flag,
                    created_at=excluded.created_at
                """,
                (
                    row["protection_date"],
                    row["nse_symbol"],
                    row["sector"],
                    row["action"],
                    row[
                        "original_allocation_percent"
                    ],
                    row[
                        "adjusted_allocation_percent"
                    ],
                    row["master_score"],
                    row["confidence_percent"],
                    row["risk_level"],
                    row[
                        "sector_exposure_percent"
                    ],
                    row["risk_flag"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()


def run_engine() -> tuple[dict, list[dict]]:
    setup_schema()

    run_id = start_run(
        "aqsd_capital_protection",
        "Applying capital protection rules",
    )

    try:
        summary, allocations = build_protection()
        save_results(
            summary,
            allocations,
        )

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=len(allocations),
            errors_count=0,
            message=(
                f"Protection={summary['protection_status']}; "
                f"cash={summary['cash_reserve_percent']}%"
            ),
        )

        return summary, allocations

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


def latest_results() -> tuple[dict | None, pd.DataFrame]:
    setup_schema()

    with connect() as connection:
        summary_row = connection.execute(
            """
            SELECT *
            FROM capital_protection_summary
            ORDER BY protection_date DESC, summary_id DESC
            LIMIT 1
            """
        ).fetchone()

        if not summary_row:
            return None, pd.DataFrame()

        protection_date = summary_row[
            "protection_date"
        ]

        allocations = pd.read_sql_query(
            """
            SELECT *
            FROM capital_protection_allocations
            WHERE protection_date=?
            ORDER BY adjusted_allocation_percent DESC
            """,
            connection,
            params=(protection_date,),
        )

    return dict(summary_row), allocations


def write_reports(
    summary: dict | None = None,
    allocations: pd.DataFrame | None = None,
) -> None:
    if summary is None or allocations is None:
        summary, allocations = latest_results()

    if not summary:
        raise RuntimeError(
            "No Capital Protection result found. "
            "Run --run first."
        )

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    allocations.to_csv(
        CSV_REPORT,
        index=False,
        encoding="utf-8-sig",
    )

    if DASHBOARD.exists():
        workbook = load_workbook(DASHBOARD)
    else:
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    sheet_name = "Capital Protection"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(
        sheet_name,
        1,
    )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A11"

    ws.merge_cells("A1:L2")
    ws["A1"] = "AQSD PROFESSIONAL - CAPITAL PROTECTION ENGINE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Protection Status"
    ws["B4"] = summary[
        "protection_status"
    ]
    ws["D4"] = "Allowed Exposure %"
    ws["E4"] = summary[
        "maximum_portfolio_exposure"
    ]
    ws["G4"] = "Cash Reserve %"
    ws["H4"] = summary[
        "cash_reserve_percent"
    ]
    ws["J4"] = "Risk Score"
    ws["K4"] = summary[
        "portfolio_risk_score"
    ]

    for ref in ("A4", "D4", "G4", "J4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    status = str(
        summary["protection_status"]
    )

    ws["B4"].fill = PatternFill(
        "solid",
        fgColor=(
            RED
            if status in {
                "CRITICAL PROTECTION",
                "DEFENSIVE",
            }
            else YELLOW
            if status == "CAUTIOUS"
            else GREEN
        ),
    )
    ws["B4"].font = Font(bold=True)

    summary_rows = [
        (
            "Market Regime",
            summary["market_regime"],
        ),
        (
            "Maximum Position Size %",
            summary["maximum_position_size"],
        ),
        (
            "Maximum Sector Exposure %",
            summary["maximum_sector_exposure"],
        ),
        (
            "Maximum Concurrent Positions",
            summary[
                "maximum_concurrent_positions"
            ],
        ),
        (
            "Stop-Loss Policy",
            summary["stop_loss_policy"],
        ),
        (
            "Explanation",
            summary["explanation"],
        ),
    ]

    for row_no, (label, value) in enumerate(
        summary_rows,
        start=6,
    ):
        ws.cell(row_no, 1, label)
        ws.cell(row_no, 2, value)
        ws.cell(row_no, 1).font = Font(
            bold=True
        )
        ws.cell(row_no, 1).fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Symbol",
        "Sector",
        "Action",
        "Original Allocation %",
        "Adjusted Allocation %",
        "Master Score",
        "Confidence %",
        "Risk Level",
        "Sector Exposure %",
        "Risk Flag",
    ]

    header_row = 10

    for col, heading in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(
            header_row,
            col,
            heading,
        )
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        allocations.iterrows(),
        start=11,
    ):
        values = [
            row.get("nse_symbol"),
            row.get("sector"),
            row.get("action"),
            row.get(
                "original_allocation_percent"
            ),
            row.get(
                "adjusted_allocation_percent"
            ),
            row.get("master_score"),
            row.get("confidence_percent"),
            row.get("risk_level"),
            row.get(
                "sector_exposure_percent"
            ),
            row.get("risk_flag"),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(
                bottom=THIN
            )

        risk = str(
            row.get("risk_level") or ""
        )

        ws.cell(
            row_no,
            8,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                RED
                if risk == "HIGH"
                else YELLOW
                if risk == "MEDIUM"
                else GREEN
            ),
        )

    widths = {
        "A": 18,
        "B": 22,
        "C": 18,
        "D": 20,
        "E": 20,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 18,
        "J": 45,
        "K": 14,
        "L": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    workbook.save(DASHBOARD)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        summary = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                MIN(protection_date) AS first_date,
                MAX(protection_date) AS latest_date
            FROM capital_protection_summary
            """
        ).fetchone()

        allocations = connection.execute(
            """
            SELECT COUNT(*) AS total
            FROM capital_protection_allocations
            """
        ).fetchone()

    print("\nAQSD CAPITAL PROTECTION STATUS")
    print("=" * 72)
    print(
        f"Summary records:     "
        f"{summary['total'] or 0}"
    )
    print(
        f"Allocation records:  "
        f"{allocations['total'] or 0}"
    )
    print(
        f"First date:          "
        f"{summary['first_date'] or 'No data'}"
    )
    print(
        f"Latest date:         "
        f"{summary['latest_date'] or 'No data'}"
    )
    print("=" * 72)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Capital Protection Engine."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Apply capital-protection rules.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild CSV and Excel reports.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show Capital Protection status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.run:
        summary, allocation_rows = (
            run_engine()
        )

        allocation_frame = pd.DataFrame(
            allocation_rows
        )

        write_reports(
            summary,
            allocation_frame,
        )

        print("\nAQSD CAPITAL PROTECTION ENGINE")
        print("=" * 72)
        print(
            f"Protection Status: "
            f"{summary['protection_status']}"
        )
        print(
            f"Allowed Exposure:  "
            f"{summary['maximum_portfolio_exposure']}%"
        )
        print(
            f"Cash Reserve:      "
            f"{summary['cash_reserve_percent']}%"
        )
        print(
            f"Risk Score:        "
            f"{summary['portfolio_risk_score']}"
        )
        print(
            f"Positions:         "
            f"{len(allocation_rows)}"
        )
        print(f"CSV:               {CSV_REPORT}")
        print(f"Dashboard:         {DASHBOARD}")
        return

    if args.report:
        write_reports()
        print(f"CSV rebuilt:\n{CSV_REPORT}")
        print(f"Dashboard rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
