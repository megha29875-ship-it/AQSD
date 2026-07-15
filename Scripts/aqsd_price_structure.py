
"""
AQSD Market Intelligence
Module: Price Structure Intelligence Engine
Version: 1.0

Purpose
-------
Uses cached OHLCV data from aqsd_core.db to calculate:

- Swing Highs / Swing Lows
- Higher High / Higher Low / Lower High / Lower Low
- Break of Structure (BOS)
- Change of Character (CHOCH)
- ATR
- ADX
- Dynamic support / resistance
- Weekly pivots
- Monthly pivots
- CPR
- Trend strength
- Institutional Structure Score
- Excel report: Price Structure Intelligence

Commands
--------
python aqsd_price_structure.py --run
python aqsd_price_structure.py --run --limit 50
python aqsd_price_structure.py --symbol RELIANCE
python aqsd_price_structure.py --status
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"

DEFAULT_HISTORY_ROWS = 300
DEFAULT_SWING_WINDOW = 3

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


SCHEMA = """
CREATE TABLE IF NOT EXISTS price_structure_intelligence (
    structure_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    symbol_id INTEGER NOT NULL,
    nse_symbol TEXT NOT NULL,
    close_price REAL,
    atr_14 REAL,
    adx_14 REAL,
    latest_swing_high REAL,
    latest_swing_low REAL,
    market_structure TEXT,
    bos_signal TEXT,
    choch_signal TEXT,
    weekly_pivot REAL,
    monthly_pivot REAL,
    cpr_low REAL,
    cpr_high REAL,
    support_level REAL,
    resistance_level REAL,
    trend_strength TEXT,
    structure_score REAL,
    directional_bias TEXT,
    explanation TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(trade_date, symbol_id),
    FOREIGN KEY(symbol_id)
        REFERENCES symbols(symbol_id)
        ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_price_structure_symbol_date
ON price_structure_intelligence(symbol_id, trade_date);
"""


def setup_schema() -> None:
    setup_database()
    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


def get_symbols(limit: int = 0) -> list[dict]:
    query = """
        SELECT symbol_id, nse_symbol, yahoo_symbol
        FROM symbols
        WHERE active = 1
        ORDER BY nse_symbol
    """
    params: tuple = ()

    if limit > 0:
        query += " LIMIT ?"
        params = (limit,)

    with connect() as connection:
        rows = connection.execute(query, params).fetchall()

    return [dict(row) for row in rows]


def get_symbol(symbol: str) -> dict | None:
    text = str(symbol or "").strip().upper()
    nse_symbol = text[:-3] if text.endswith(".NS") else text
    yahoo_symbol = text if text.endswith(".NS") else f"{text}.NS"

    with connect() as connection:
        row = connection.execute(
            """
            SELECT symbol_id, nse_symbol, yahoo_symbol
            FROM symbols
            WHERE nse_symbol = ?
               OR yahoo_symbol = ?
            LIMIT 1
            """,
            (nse_symbol, yahoo_symbol),
        ).fetchone()

    return dict(row) if row else None


def get_history(symbol_id: int, rows: int) -> pd.DataFrame:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT *
            FROM (
                SELECT
                    trade_date,
                    open,
                    high,
                    low,
                    close,
                    adjusted_close,
                    volume
                FROM daily_prices
                WHERE symbol_id = ?
                ORDER BY trade_date DESC
                LIMIT ?
            )
            ORDER BY trade_date
            """,
            connection,
            params=(symbol_id, rows),
        )

    if frame.empty:
        return frame

    frame["trade_date"] = pd.to_datetime(frame["trade_date"])
    frame = frame.set_index("trade_date")
    return frame


def calculate_atr(frame: pd.DataFrame, period: int = 14) -> pd.Series:
    high_low = frame["high"] - frame["low"]
    high_close = (frame["high"] - frame["close"].shift(1)).abs()
    low_close = (frame["low"] - frame["close"].shift(1)).abs()

    true_range = pd.concat(
        [high_low, high_close, low_close],
        axis=1,
    ).max(axis=1)

    return true_range.rolling(period).mean()


def calculate_adx(frame: pd.DataFrame, period: int = 14) -> pd.Series:
    high_diff = frame["high"].diff()
    low_diff = -frame["low"].diff()

    plus_dm = np.where(
        (high_diff > low_diff) & (high_diff > 0),
        high_diff,
        0.0,
    )
    minus_dm = np.where(
        (low_diff > high_diff) & (low_diff > 0),
        low_diff,
        0.0,
    )

    atr = calculate_atr(frame, period)

    plus_di = 100 * (
        pd.Series(plus_dm, index=frame.index).rolling(period).sum()
        / atr
    )
    minus_di = 100 * (
        pd.Series(minus_dm, index=frame.index).rolling(period).sum()
        / atr
    )

    denominator = (plus_di + minus_di).replace(0, np.nan)
    dx = 100 * (plus_di - minus_di).abs() / denominator

    return dx.rolling(period).mean()


def swing_points(
    frame: pd.DataFrame,
    window: int,
) -> tuple[list[tuple[pd.Timestamp, float]], list[tuple[pd.Timestamp, float]]]:
    highs: list[tuple[pd.Timestamp, float]] = []
    lows: list[tuple[pd.Timestamp, float]] = []

    for index in range(window, len(frame) - window):
        high_value = float(frame["high"].iloc[index])
        low_value = float(frame["low"].iloc[index])

        high_slice = frame["high"].iloc[
            index - window:index + window + 1
        ]
        low_slice = frame["low"].iloc[
            index - window:index + window + 1
        ]

        if high_value == float(high_slice.max()):
            highs.append((frame.index[index], high_value))

        if low_value == float(low_slice.min()):
            lows.append((frame.index[index], low_value))

    return highs, lows


def classify_structure(
    swing_highs: list[tuple[pd.Timestamp, float]],
    swing_lows: list[tuple[pd.Timestamp, float]],
) -> str:
    if len(swing_highs) < 2 or len(swing_lows) < 2:
        return "INSUFFICIENT DATA"

    previous_high = swing_highs[-2][1]
    latest_high = swing_highs[-1][1]
    previous_low = swing_lows[-2][1]
    latest_low = swing_lows[-1][1]

    if latest_high > previous_high and latest_low > previous_low:
        return "HH-HL BULLISH"

    if latest_high < previous_high and latest_low < previous_low:
        return "LH-LL BEARISH"

    if latest_high > previous_high and latest_low < previous_low:
        return "EXPANDING RANGE"

    return "CONSOLIDATION"


def bos_choch(
    close_price: float,
    swing_highs: list[tuple[pd.Timestamp, float]],
    swing_lows: list[tuple[pd.Timestamp, float]],
    structure: str,
) -> tuple[str, str]:
    if not swing_highs or not swing_lows:
        return "NONE", "NONE"

    latest_high = swing_highs[-1][1]
    latest_low = swing_lows[-1][1]

    bos = "NONE"
    choch = "NONE"

    if close_price > latest_high:
        if "BULLISH" in structure:
            bos = "BULLISH BOS"
        elif "BEARISH" in structure:
            choch = "BULLISH CHOCH"

    elif close_price < latest_low:
        if "BEARISH" in structure:
            bos = "BEARISH BOS"
        elif "BULLISH" in structure:
            choch = "BEARISH CHOCH"

    return bos, choch


def previous_period_ohlc(
    frame: pd.DataFrame,
    frequency: str,
) -> tuple[float, float, float]:
    grouped = frame.resample(frequency).agg(
        {
            "high": "max",
            "low": "min",
            "close": "last",
        }
    ).dropna()

    if len(grouped) < 2:
        raise RuntimeError("Insufficient resampled history")

    row = grouped.iloc[-2]

    return (
        float(row["high"]),
        float(row["low"]),
        float(row["close"]),
    )


def classic_pivot(high: float, low: float, close: float) -> dict[str, float]:
    pivot = (high + low + close) / 3
    bc = (high + low) / 2
    tc = 2 * pivot - bc

    return {
        "pivot": pivot,
        "cpr_low": min(bc, tc),
        "cpr_high": max(bc, tc),
    }


def trend_strength_label(adx: float | None) -> str:
    if adx is None or np.isnan(adx):
        return "UNKNOWN"
    if adx >= 40:
        return "VERY STRONG"
    if adx >= 25:
        return "STRONG"
    if adx >= 20:
        return "DEVELOPING"
    return "WEAK"


def structure_score(
    structure: str,
    bos: str,
    choch: str,
    close_price: float,
    support: float | None,
    resistance: float | None,
    adx: float | None,
    weekly_pivot: float | None,
    monthly_pivot: float | None,
    cpr_low: float | None,
    cpr_high: float | None,
) -> tuple[float, str, str]:
    score = 50.0
    reasons: list[str] = []

    if "BULLISH" in structure:
        score += 15
        reasons.append("Bullish HH-HL structure")
    elif "BEARISH" in structure:
        score -= 15
        reasons.append("Bearish LH-LL structure")
    elif structure == "EXPANDING RANGE":
        reasons.append("Expanding range")
    else:
        reasons.append("Consolidation")

    if bos == "BULLISH BOS":
        score += 18
        reasons.append("Bullish break of structure")
    elif bos == "BEARISH BOS":
        score -= 18
        reasons.append("Bearish break of structure")

    if choch == "BULLISH CHOCH":
        score += 12
        reasons.append("Bullish change of character")
    elif choch == "BEARISH CHOCH":
        score -= 12
        reasons.append("Bearish change of character")

    if weekly_pivot is not None:
        if close_price > weekly_pivot:
            score += 5
            reasons.append("Above weekly pivot")
        else:
            score -= 5
            reasons.append("Below weekly pivot")

    if monthly_pivot is not None:
        if close_price > monthly_pivot:
            score += 7
            reasons.append("Above monthly pivot")
        else:
            score -= 7
            reasons.append("Below monthly pivot")

    if cpr_low is not None and cpr_high is not None:
        if close_price > cpr_high:
            score += 5
            reasons.append("Above CPR")
        elif close_price < cpr_low:
            score -= 5
            reasons.append("Below CPR")
        else:
            reasons.append("Inside CPR")

    if adx is not None and not np.isnan(adx):
        if adx >= 25:
            if score >= 50:
                score += 5
            else:
                score -= 5
            reasons.append("ADX confirms trend")
        elif adx < 20:
            reasons.append("Weak ADX")

    if support is not None and resistance is not None:
        support_distance = abs(close_price - support)
        resistance_distance = abs(resistance - close_price)

        if resistance_distance < support_distance:
            reasons.append("Closer to resistance")
        else:
            reasons.append("Closer to support")

    score = round(max(0, min(100, score)), 2)

    if score >= 75:
        bias = "STRONG BULLISH"
    elif score >= 60:
        bias = "BULLISH"
    elif score <= 25:
        bias = "STRONG BEARISH"
    elif score <= 40:
        bias = "BEARISH"
    else:
        bias = "NEUTRAL"

    return score, bias, " | ".join(reasons)


def analyse_symbol(
    symbol: dict,
    history_rows: int,
    swing_window: int,
) -> dict:
    frame = get_history(symbol["symbol_id"], history_rows)

    if len(frame) < 80:
        raise RuntimeError("Insufficient cached history")

    atr = calculate_atr(frame, 14)
    adx = calculate_adx(frame, 14)

    swing_highs, swing_lows = swing_points(
        frame,
        swing_window,
    )

    close_price = float(frame["close"].iloc[-1])
    trade_date = frame.index[-1].date().isoformat()

    structure = classify_structure(
        swing_highs,
        swing_lows,
    )

    bos, choch = bos_choch(
        close_price,
        swing_highs,
        swing_lows,
        structure,
    )

    latest_swing_high = (
        swing_highs[-1][1]
        if swing_highs
        else None
    )
    latest_swing_low = (
        swing_lows[-1][1]
        if swing_lows
        else None
    )

    weekly_high, weekly_low, weekly_close = previous_period_ohlc(
        frame,
        "W-FRI",
    )
    monthly_high, monthly_low, monthly_close = previous_period_ohlc(
        frame,
        "ME",
    )

    weekly = classic_pivot(
        weekly_high,
        weekly_low,
        weekly_close,
    )
    monthly = classic_pivot(
        monthly_high,
        monthly_low,
        monthly_close,
    )

    support = latest_swing_low
    resistance = latest_swing_high

    adx_value = (
        float(adx.iloc[-1])
        if not adx.empty and not np.isnan(adx.iloc[-1])
        else None
    )

    atr_value = (
        float(atr.iloc[-1])
        if not atr.empty and not np.isnan(atr.iloc[-1])
        else None
    )

    score, bias, explanation = structure_score(
        structure,
        bos,
        choch,
        close_price,
        support,
        resistance,
        adx_value,
        weekly["pivot"],
        monthly["pivot"],
        weekly["cpr_low"],
        weekly["cpr_high"],
    )

    return {
        "trade_date": trade_date,
        "symbol_id": symbol["symbol_id"],
        "nse_symbol": symbol["nse_symbol"],
        "close_price": round(close_price, 2),
        "atr_14": round(atr_value, 2) if atr_value is not None else None,
        "adx_14": round(adx_value, 2) if adx_value is not None else None,
        "latest_swing_high": (
            round(latest_swing_high, 2)
            if latest_swing_high is not None
            else None
        ),
        "latest_swing_low": (
            round(latest_swing_low, 2)
            if latest_swing_low is not None
            else None
        ),
        "market_structure": structure,
        "bos_signal": bos,
        "choch_signal": choch,
        "weekly_pivot": round(weekly["pivot"], 2),
        "monthly_pivot": round(monthly["pivot"], 2),
        "cpr_low": round(weekly["cpr_low"], 2),
        "cpr_high": round(weekly["cpr_high"], 2),
        "support_level": (
            round(support, 2)
            if support is not None
            else None
        ),
        "resistance_level": (
            round(resistance, 2)
            if resistance is not None
            else None
        ),
        "trend_strength": trend_strength_label(adx_value),
        "structure_score": score,
        "directional_bias": bias,
        "explanation": explanation,
    }


def save_result(connection, result: dict) -> None:
    connection.execute(
        """
        INSERT INTO price_structure_intelligence(
            trade_date,
            symbol_id,
            nse_symbol,
            close_price,
            atr_14,
            adx_14,
            latest_swing_high,
            latest_swing_low,
            market_structure,
            bos_signal,
            choch_signal,
            weekly_pivot,
            monthly_pivot,
            cpr_low,
            cpr_high,
            support_level,
            resistance_level,
            trend_strength,
            structure_score,
            directional_bias,
            explanation,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(trade_date, symbol_id)
        DO UPDATE SET
            close_price = excluded.close_price,
            atr_14 = excluded.atr_14,
            adx_14 = excluded.adx_14,
            latest_swing_high = excluded.latest_swing_high,
            latest_swing_low = excluded.latest_swing_low,
            market_structure = excluded.market_structure,
            bos_signal = excluded.bos_signal,
            choch_signal = excluded.choch_signal,
            weekly_pivot = excluded.weekly_pivot,
            monthly_pivot = excluded.monthly_pivot,
            cpr_low = excluded.cpr_low,
            cpr_high = excluded.cpr_high,
            support_level = excluded.support_level,
            resistance_level = excluded.resistance_level,
            trend_strength = excluded.trend_strength,
            structure_score = excluded.structure_score,
            directional_bias = excluded.directional_bias,
            explanation = excluded.explanation,
            created_at = excluded.created_at
        """,
        (
            result["trade_date"],
            result["symbol_id"],
            result["nse_symbol"],
            result["close_price"],
            result["atr_14"],
            result["adx_14"],
            result["latest_swing_high"],
            result["latest_swing_low"],
            result["market_structure"],
            result["bos_signal"],
            result["choch_signal"],
            result["weekly_pivot"],
            result["monthly_pivot"],
            result["cpr_low"],
            result["cpr_high"],
            result["support_level"],
            result["resistance_level"],
            result["trend_strength"],
            result["structure_score"],
            result["directional_bias"],
            result["explanation"],
            datetime.now().isoformat(timespec="seconds"),
        ),
    )


def run_engine(
    limit: int,
    history_rows: int,
    swing_window: int,
) -> list[dict]:
    setup_schema()
    symbols = get_symbols(limit)

    if not symbols:
        raise RuntimeError("No active symbols found.")

    run_id = start_run(
        "aqsd_price_structure",
        f"Analysing {len(symbols)} symbols",
    )

    completed = 0
    failed = 0
    results: list[dict] = []

    try:
        print("\nAQSD PRICE STRUCTURE INTELLIGENCE")
        print("=" * 78)

        with connect() as connection:
            for index, symbol in enumerate(symbols, start=1):
                print(
                    f"[{index}/{len(symbols)}] "
                    f"{symbol['nse_symbol']:<18}",
                    end="",
                )

                try:
                    result = analyse_symbol(
                        symbol,
                        history_rows,
                        swing_window,
                    )

                    save_result(connection, result)
                    results.append(result)
                    completed += 1

                    print(
                        f"OK  "
                        f"{result['structure_score']:>6.2f}  "
                        f"{result['directional_bias']}"
                    )

                except Exception as error:
                    failed += 1
                    print(f"FAILED: {error}")

            connection.commit()

        finish_run(
            run_id,
            status="SUCCESS" if failed == 0 else "PARTIAL",
            records_processed=completed,
            errors_count=failed,
            message=f"Completed={completed}; failed={failed}",
        )

        return sorted(
            results,
            key=lambda item: item["structure_score"],
            reverse=True,
        )

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=completed,
            errors_count=failed + 1,
            message=str(error),
        )
        raise


def latest_results() -> pd.DataFrame:
    setup_schema()

    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM price_structure_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
            ORDER BY structure_score DESC, nse_symbol
            """,
            connection,
        )


def write_report(results: list[dict] | None = None) -> None:
    if results is None:
        frame = latest_results()
        results = frame.to_dict("records")

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Price Structure Intelligence" in wb.sheetnames:
        del wb["Price Structure Intelligence"]

    ws = wb.create_sheet("Price Structure Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:W2")
    ws["A1"] = "AQSD PROFESSIONAL - PRICE STRUCTURE INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Symbols Analysed"
    ws["B4"] = len(results)
    ws["D4"] = "Updated"
    ws["E4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Rank",
        "Symbol",
        "Trade Date",
        "Close",
        "ATR 14",
        "ADX 14",
        "Swing High",
        "Swing Low",
        "Market Structure",
        "BOS",
        "CHOCH",
        "Weekly Pivot",
        "Monthly Pivot",
        "CPR Low",
        "CPR High",
        "Support",
        "Resistance",
        "Trend Strength",
        "Structure Score",
        "Directional Bias",
        "Explanation",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, result in enumerate(results, start=8):
        values = [
            row_no - 7,
            result["nse_symbol"],
            result["trade_date"],
            result["close_price"],
            result["atr_14"],
            result["adx_14"],
            result["latest_swing_high"],
            result["latest_swing_low"],
            result["market_structure"],
            result["bos_signal"],
            result["choch_signal"],
            result["weekly_pivot"],
            result["monthly_pivot"],
            result["cpr_low"],
            result["cpr_high"],
            result["support_level"],
            result["resistance_level"],
            result["trend_strength"],
            result["structure_score"],
            result["directional_bias"],
            result["explanation"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        score = float(result["structure_score"])
        ws.cell(row_no, 19).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )

        bias = str(result["directional_bias"])
        ws.cell(row_no, 20).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BULLISH" in bias
                else RED
                if "BEARISH" in bias
                else GREY
            ),
        )

    widths = {
        "A": 8,
        "B": 16,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 22,
        "J": 18,
        "K": 18,
        "L": 14,
        "M": 14,
        "N": 12,
        "O": 12,
        "P": 14,
        "Q": 14,
        "R": 16,
        "S": 16,
        "T": 18,
        "U": 70,
        "V": 14,
        "W": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT symbol_id) AS symbols,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM price_structure_intelligence
            """
        ).fetchone()

    print("\nAQSD PRICE STRUCTURE STATUS")
    print("=" * 72)
    print(f"Stored records:     {row['total'] or 0}")
    print(f"Symbols covered:    {row['symbols'] or 0}")
    print(f"First date:         {row['first_date'] or 'No data'}")
    print(f"Latest date:        {row['latest_date'] or 'No data'}")
    print("=" * 72)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Price Structure Intelligence Engine."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Run the full structure engine.",
    )

    parser.add_argument(
        "--symbol",
        metavar="SYMBOL",
        help="Analyse one symbol.",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Maximum active symbols. Use 0 for all.",
    )

    parser.add_argument(
        "--history-rows",
        type=int,
        default=DEFAULT_HISTORY_ROWS,
        help="Cached daily rows used per symbol.",
    )

    parser.add_argument(
        "--swing-window",
        type=int,
        default=DEFAULT_SWING_WINDOW,
        help="Swing-detection window.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild report from latest stored results.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show database status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.symbol:
        symbol = get_symbol(args.symbol)

        if not symbol:
            raise RuntimeError(
                f"Symbol not found: {args.symbol}"
            )

        result = analyse_symbol(
            symbol,
            args.history_rows,
            args.swing_window,
        )

        with connect() as connection:
            save_result(connection, result)
            connection.commit()

        print("\nAQSD SINGLE-SYMBOL STRUCTURE")
        print("=" * 72)

        for key, value in result.items():
            print(f"{key:<24}{value}")

        return

    if args.run:
        results = run_engine(
            args.limit,
            args.history_rows,
            args.swing_window,
        )

        write_report(results)

        print("=" * 78)
        print(f"Symbols completed: {len(results)}")
        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report()
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
