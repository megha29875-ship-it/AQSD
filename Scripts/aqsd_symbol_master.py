"""
AQSD Core - Symbol Master v1.0
"""
from __future__ import annotations
import argparse, csv, sqlite3
from datetime import datetime
from pathlib import Path
from typing import Iterable
import pandas as pd
from openpyxl import load_workbook
from aqsd_database import connect, setup_database, start_run, finish_run

SCRIPTS_DIR=Path(__file__).resolve().parent
BASE_DIR=SCRIPTS_DIR.parent
DATA_DIR=BASE_DIR/'Data'; OUTPUT_DIR=BASE_DIR/'Output'
FNO_XLSX=DATA_DIR/'FnO_Stocks.xlsx'; FNO_CSV=DATA_DIR/'FnO_Stocks.csv'
DASHBOARD=OUTPUT_DIR/'Dashboard.xlsx'; EXPORT_FILE=DATA_DIR/'AQSD_Symbol_Master.csv'

SECTOR_MAP={
'RELIANCE':('Energy','Oil, Gas & Consumable Fuels'),'ONGC':('Energy','Oil Exploration & Production'),
'OIL':('Energy','Oil Exploration & Production'),'IOC':('Energy','Refineries & Marketing'),
'BPCL':('Energy','Refineries & Marketing'),'HINDPETRO':('Energy','Refineries & Marketing'),
'GAIL':('Energy','Gas Transmission & Marketing'),'PETRONET':('Energy','Gas Infrastructure'),
'HDFCBANK':('Bank','Private Sector Bank'),'ICICIBANK':('Bank','Private Sector Bank'),
'AXISBANK':('Bank','Private Sector Bank'),'KOTAKBANK':('Bank','Private Sector Bank'),
'INDUSINDBK':('Bank','Private Sector Bank'),'FEDERALBNK':('Bank','Private Sector Bank'),
'IDFCFIRSTB':('Bank','Private Sector Bank'),'SBIN':('Bank','Public Sector Bank'),
'BANKBARODA':('Bank','Public Sector Bank'),'CANBK':('Bank','Public Sector Bank'),'PNB':('Bank','Public Sector Bank'),
'TCS':('IT','IT Services'),'INFY':('IT','IT Services'),'HCLTECH':('IT','IT Services'),
'WIPRO':('IT','IT Services'),'TECHM':('IT','IT Services'),'LTIM':('IT','IT Services'),
'MPHASIS':('IT','IT Services'),'PERSISTENT':('IT','IT Services'),'COFORGE':('IT','IT Services'),
'MARUTI':('Auto','Passenger Vehicles'),'TATAMOTORS':('Auto','Passenger & Commercial Vehicles'),
'M&M':('Auto','Automobiles'),'BAJAJ-AUTO':('Auto','Two Wheelers'),'EICHERMOT':('Auto','Two Wheelers'),
'HEROMOTOCO':('Auto','Two Wheelers'),'ASHOKLEY':('Auto','Commercial Vehicles'),'TVSMOTOR':('Auto','Two Wheelers'),
'SUNPHARMA':('Pharma','Pharmaceuticals'),'CIPLA':('Pharma','Pharmaceuticals'),
'DRREDDY':('Pharma','Pharmaceuticals'),'DIVISLAB':('Pharma','Pharmaceuticals'),'BIOCON':('Pharma','Biotechnology'),
'LUPIN':('Pharma','Pharmaceuticals'),'AUROPHARMA':('Pharma','Pharmaceuticals'),'TORNTPHARM':('Pharma','Pharmaceuticals'),
'HINDUNILVR':('FMCG','Personal Products'),'ITC':('FMCG','Diversified FMCG'),'NESTLEIND':('FMCG','Packaged Foods'),
'BRITANNIA':('FMCG','Packaged Foods'),'DABUR':('FMCG','Personal Products'),'MARICO':('FMCG','Personal Products'),
'GODREJCP':('FMCG','Personal Products'),'TATASTEEL':('Metal','Iron & Steel'),'JSWSTEEL':('Metal','Iron & Steel'),
'HINDALCO':('Metal','Aluminium'),'VEDL':('Metal','Diversified Metals'),'NMDC':('Metal','Mining'),
'SAIL':('Metal','Iron & Steel'),'NATIONALUM':('Metal','Aluminium'),'LT':('Infrastructure','Engineering & Construction'),
'SIEMENS':('Capital Goods','Electrical Equipment'),'ABB':('Capital Goods','Electrical Equipment'),
'BHEL':('Capital Goods','Heavy Electrical Equipment'),'CUMMINSIND':('Capital Goods','Industrial Machinery'),
'BEL':('Defence','Defence Electronics'),'HAL':('Defence','Aerospace & Defence'),'BDL':('Defence','Defence Equipment'),
'DLF':('Realty','Real Estate'),'GODREJPROP':('Realty','Real Estate'),'OBEROIRLTY':('Realty','Real Estate'),'PRESTIGE':('Realty','Real Estate'),
'ULTRACEMCO':('Cement','Cement'),'GRASIM':('Cement','Cement & Diversified'),'AMBUJACEM':('Cement','Cement'),
'ACC':('Cement','Cement'),'SHREECEM':('Cement','Cement'),'BHARTIARTL':('Telecom','Telecommunication Services'),
'IDEA':('Telecom','Telecommunication Services'),'POWERGRID':('Power','Power Transmission'),'NTPC':('Power','Power Generation'),
'TATAPOWER':('Power','Integrated Power'),'ADANIPOWER':('Power','Power Generation'),'APOLLOHOSP':('Healthcare','Hospitals'),
'MAXHEALTH':('Healthcare','Hospitals')}

FALLBACK_SYMBOLS=[('RELIANCE','RELIANCE.NS','Reliance Industries'),('HDFCBANK','HDFCBANK.NS','HDFC Bank'),
('ICICIBANK','ICICIBANK.NS','ICICI Bank'),('SBIN','SBIN.NS','State Bank of India'),('INFY','INFY.NS','Infosys'),
('TCS','TCS.NS','Tata Consultancy Services'),('LT','LT.NS','Larsen & Toubro'),('SUNPHARMA','SUNPHARMA.NS','Sun Pharmaceutical Industries'),
('BIOCON','BIOCON.NS','Biocon'),('TATAMOTORS','TATAMOTORS.NS','Tata Motors')]

def clean_text(v):
    if v is None:return ''
    t=str(v).strip()
    return '' if t.lower() in {'nan','none','null'} else t

def normalize_nse_symbol(v):
    s=clean_text(v).upper()
    if s.endswith('.NS'): s=s[:-3]
    return s.replace(' ','')

def normalize_yahoo_symbol(v):
    s=clean_text(v).upper()
    if not s:return ''
    return s if s.endswith('.NS') else normalize_nse_symbol(s)+'.NS'

def guess_sector_industry(s): return SECTOR_MAP.get(s,('',''))

def first_existing_column(columns:Iterable[str], candidates:Iterable[str]):
    m={str(c).strip().lower():str(c) for c in columns}
    for c in candidates:
        if c.strip().lower() in m:return m[c.strip().lower()]
    return None

def deduplicate_records(records):
    merged={}
    for r in records:
        s=r['nse_symbol']
        if s not in merged: merged[s]=r.copy(); continue
        for f in ('yahoo_symbol','company_name','sector','industry','source'):
            if not merged[s].get(f) and r.get(f): merged[s][f]=r[f]
    return sorted(merged.values(),key=lambda x:x['nse_symbol'])

def dataframe_to_records(df,source_name):
    nse=first_existing_column(df.columns,['NSE Symbol','NSE_SYMBOL','SYMBOL','Symbol','Ticker'])
    yahoo=first_existing_column(df.columns,['Yahoo Symbol','Yahoo_Symbol','YAHOO_SYMBOL'])
    company=first_existing_column(df.columns,['Company Name','Company','NAME OF COMPANY','Security Name','Name'])
    sector=first_existing_column(df.columns,['Sector','SECTOR'])
    industry=first_existing_column(df.columns,['Industry','INDUSTRY'])
    if not nse and not yahoo: raise RuntimeError('No symbol column found.')
    out=[]
    for _,row in df.iterrows():
        ns=normalize_nse_symbol(row[nse] if nse else row[yahoo]); ys=normalize_yahoo_symbol(row[yahoo] if yahoo else ns)
        if not ns or not ys: continue
        ms,mi=guess_sector_industry(ns)
        out.append({'nse_symbol':ns,'yahoo_symbol':ys,'company_name':clean_text(row[company]) if company else '',
                    'sector':clean_text(row[sector]) if sector else ms,'industry':clean_text(row[industry]) if industry else mi,
                    'fno_eligible':1,'active':1,'source':source_name})
    return deduplicate_records(out)

def read_fno_excel(): return dataframe_to_records(pd.read_excel(FNO_XLSX),FNO_XLSX.name) if FNO_XLSX.exists() else []
def read_fno_csv(): return dataframe_to_records(pd.read_csv(FNO_CSV),FNO_CSV.name) if FNO_CSV.exists() else []

def read_dashboard_candidates():
    if not DASHBOARD.exists(): return []
    wb=load_workbook(DASHBOARD,read_only=True,data_only=True); symbols=[]
    try:
        for sn in ('CALL Candidates','PUT Candidates'):
            if sn not in wb.sheetnames: continue
            ws=wb[sn]; col=None
            for cell in ws[1]:
                if str(cell.value or '').strip().lower()=='symbol': col=cell.column; break
            if col is None: continue
            for r in range(2,ws.max_row+1):
                s=normalize_nse_symbol(ws.cell(r,col).value)
                if s and s not in symbols:symbols.append(s)
    finally: wb.close()
    out=[]
    for s in symbols:
        sec,ind=guess_sector_industry(s)
        out.append({'nse_symbol':s,'yahoo_symbol':normalize_yahoo_symbol(s),'company_name':'','sector':sec,'industry':ind,
                    'fno_eligible':1,'active':1,'source':'Dashboard candidates'})
    return out

def read_fallback_symbols():
    out=[]
    for ns,ys,cn in FALLBACK_SYMBOLS:
        sec,ind=guess_sector_industry(ns)
        out.append({'nse_symbol':ns,'yahoo_symbol':ys,'company_name':cn,'sector':sec,'industry':ind,'fno_eligible':1,'active':1,'source':'Built-in fallback'})
    return out

def load_source_records():
    for fn,label in ((read_fno_excel,FNO_XLSX.name),(read_fno_csv,FNO_CSV.name),(read_dashboard_candidates,'Dashboard candidates')):
        records=fn()
        if records:return records,label
    return read_fallback_symbols(),'Built-in fallback'

def upsert_symbol(connection:sqlite3.Connection,record):
    ts=datetime.now().isoformat(timespec='seconds')
    connection.execute('''INSERT INTO symbols(nse_symbol,yahoo_symbol,company_name,sector,industry,fno_eligible,active,source,last_updated)
    VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(nse_symbol) DO UPDATE SET
    yahoo_symbol=excluded.yahoo_symbol,
    company_name=CASE WHEN excluded.company_name<>'' THEN excluded.company_name ELSE symbols.company_name END,
    sector=CASE WHEN excluded.sector<>'' THEN excluded.sector ELSE symbols.sector END,
    industry=CASE WHEN excluded.industry<>'' THEN excluded.industry ELSE symbols.industry END,
    fno_eligible=excluded.fno_eligible,active=excluded.active,source=excluded.source,last_updated=excluded.last_updated''',
    (record['nse_symbol'],record['yahoo_symbol'],record.get('company_name',''),record.get('sector',''),record.get('industry',''),
     int(record.get('fno_eligible',1)),int(record.get('active',1)),record.get('source',''),ts))

def import_symbol_master():
    setup_database(); records,source=load_source_records(); run_id=start_run('aqsd_symbol_master',f'Importing from {source}')
    try:
        with connect() as con:
            for r in records: upsert_symbol(con,r)
            con.commit()
        finish_run(run_id,'SUCCESS',len(records),0,f'Imported from {source}')
        return len(records),source
    except Exception as e:
        finish_run(run_id,'FAILED',0,1,str(e)); raise

def add_symbol(symbol,company_name='',sector='',industry='',fno_eligible=1):
    setup_database(); ns=normalize_nse_symbol(symbol)
    if not ns: raise ValueError('Symbol cannot be blank.')
    ms,mi=guess_sector_industry(ns)
    rec={'nse_symbol':ns,'yahoo_symbol':normalize_yahoo_symbol(ns),'company_name':clean_text(company_name),
         'sector':clean_text(sector) or ms,'industry':clean_text(industry) or mi,'fno_eligible':int(bool(fno_eligible)),
         'active':1,'source':'Manual'}
    with connect() as con: upsert_symbol(con,rec); con.commit()

def deactivate_symbol(symbol):
    setup_database()
    with connect() as con:
        cur=con.execute('UPDATE symbols SET active=0,last_updated=? WHERE nse_symbol=?',(datetime.now().isoformat(timespec='seconds'),normalize_nse_symbol(symbol)))
        con.commit(); return cur.rowcount>0

def get_all_symbols(active_only=True,fno_only=False):
    setup_database(); clauses=[]; params=[]
    if active_only: clauses.append('active=?'); params.append(1)
    if fno_only: clauses.append('fno_eligible=?'); params.append(1)
    q='SELECT * FROM symbols'+((' WHERE '+' AND '.join(clauses)) if clauses else '')+' ORDER BY nse_symbol'
    with connect() as con: rows=con.execute(q,params).fetchall()
    return [dict(r) for r in rows]

def search_symbols(text):
    setup_database(); p=f'%{clean_text(text)}%'
    with connect() as con:
        rows=con.execute('''SELECT * FROM symbols WHERE nse_symbol LIKE ? OR yahoo_symbol LIKE ? OR company_name LIKE ? OR sector LIKE ? OR industry LIKE ? ORDER BY active DESC,nse_symbol''',(p,p,p,p,p)).fetchall()
    return [dict(r) for r in rows]

def symbol_statistics():
    setup_database()
    with connect() as con:
        row=con.execute('''SELECT COUNT(*) total,SUM(CASE WHEN active=1 THEN 1 ELSE 0 END) active,
        SUM(CASE WHEN fno_eligible=1 THEN 1 ELSE 0 END) fno,
        SUM(CASE WHEN sector IS NULL OR TRIM(sector)='' THEN 1 ELSE 0 END) missing_sector,
        SUM(CASE WHEN company_name IS NULL OR TRIM(company_name)='' THEN 1 ELSE 0 END) missing_company FROM symbols''').fetchone()
        sectors=con.execute("SELECT COALESCE(NULLIF(TRIM(sector),''),'Unmapped') sector_name,COUNT(*) symbol_count FROM symbols WHERE active=1 GROUP BY sector_name ORDER BY symbol_count DESC,sector_name").fetchall()
    return {'total':int(row['total'] or 0),'active':int(row['active'] or 0),'fno':int(row['fno'] or 0),
            'missing_sector':int(row['missing_sector'] or 0),'missing_company':int(row['missing_company'] or 0),'sectors':[dict(x) for x in sectors]}

def validate_symbol_master():
    issues=[]; seen={}
    for item in get_all_symbols(active_only=False):
        ns=item['nse_symbol']; ys=item['yahoo_symbol']
        if not ys.endswith('.NS'):issues.append(f'{ns}: Yahoo symbol must end with .NS')
        if ys in seen and seen[ys]!=ns:issues.append(f'Duplicate Yahoo symbol {ys}: {seen[ys]}, {ns}')
        seen[ys]=ns
        if not item.get('sector'):issues.append(f'{ns}: sector not mapped')
    return issues

def export_symbol_master():
    rows=get_all_symbols(active_only=False); DATA_DIR.mkdir(parents=True,exist_ok=True)
    fields=['symbol_id','nse_symbol','yahoo_symbol','company_name','sector','industry','fno_eligible','active','source','last_updated']
    with EXPORT_FILE.open('w',newline='',encoding='utf-8-sig') as f:
        w=csv.DictWriter(f,fieldnames=fields); w.writeheader()
        for r in rows:w.writerow({k:r.get(k,'') for k in fields})
    return len(rows)

def print_symbol_rows(rows):
    if not rows: print('No matching symbols.'); return
    print(f"{'NSE Symbol':<16}{'Yahoo Symbol':<20}{'Sector':<20}{'Company':<35}{'Active':<8}"); print('-'*99)
    for x in rows: print(f"{x['nse_symbol']:<16}{x['yahoo_symbol']:<20}{(x.get('sector') or ''):<20}{(x.get('company_name') or '')[:33]:<35}{x['active']:<8}")

def show_status():
    s=symbol_statistics(); print('\nAQSD SYMBOL MASTER STATUS'); print('='*72)
    print(f"Total symbols:       {s['total']}\nActive symbols:      {s['active']}\nF&O eligible:        {s['fno']}\nMissing sector:      {s['missing_sector']}\nMissing company:     {s['missing_company']}")
    print('-'*72); print('Sector distribution')
    for x in s['sectors']: print(f"{x['sector_name']:<30}{x['symbol_count']:>8}")
    print('='*72)

def parse_arguments():
    p=argparse.ArgumentParser(description='Manage AQSD Core symbol master.')
    p.add_argument('--import',dest='import_symbols',action='store_true'); p.add_argument('--status',action='store_true'); p.add_argument('--list',action='store_true')
    p.add_argument('--search'); p.add_argument('--add'); p.add_argument('--company',default=''); p.add_argument('--sector',default=''); p.add_argument('--industry',default='')
    p.add_argument('--deactivate'); p.add_argument('--export',action='store_true'); p.add_argument('--validate',action='store_true')
    return p.parse_args()

def main():
    args=parse_arguments(); setup_database()
    if args.import_symbols:
        count,source=import_symbol_master(); print('\nAQSD SYMBOL MASTER'); print('='*72); print(f'Imported / refreshed: {count}\nSource: {source}'); print('='*72)
    elif args.status: show_status()
    elif args.list: print_symbol_rows(get_all_symbols())
    elif args.search: print_symbol_rows(search_symbols(args.search))
    elif args.add:
        add_symbol(args.add,args.company,args.sector,args.industry); print(f'Symbol added or updated: {normalize_nse_symbol(args.add)}')
    elif args.deactivate: print('Symbol deactivated.' if deactivate_symbol(args.deactivate) else 'Symbol not found.')
    elif args.export:
        count=export_symbol_master(); print(f'Exported {count} symbols.'); print(EXPORT_FILE)
    elif args.validate:
        issues=validate_symbol_master(); print('\nAQSD SYMBOL MASTER VALIDATION'); print('='*72)
        if issues:
            print(f'Issues found: {len(issues)}')
            for issue in issues: print('- '+issue)
        else: print('PASS: No symbol-master issues found.')
        print('='*72)
    else: show_status()

if __name__=='__main__': main()
