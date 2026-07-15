
"""
AQSD Market Intelligence
Module: News & Event Intelligence Foundation
Version: 1.0

Purpose
-------
Creates a structured news/event workflow for AQSD without depending on
a single news vendor.

Features
--------
- Imports news/events from CSV
- Adds events manually
- Prevents simple duplicates
- Scores sentiment, credibility, materiality, urgency and recency
- Calculates a News Impact Score from -100 to +100
- Stores events in aqsd_core.db
- Creates a "News Intelligence" sheet in Dashboard.xlsx
- Produces symbol-level news scores for future Master Intelligence use

Commands
--------
python aqsd_news_intelligence.py --setup
python aqsd_news_intelligence.py --import-csv Data/news_events.csv
python aqsd_news_intelligence.py --add-manual
python aqsd_news_intelligence.py --score
python aqsd_news_intelligence.py --report
python aqsd_news_intelligence.py --status

CSV columns supported
---------------------
event_time, source, headline, url, event_type, company_name, nse_symbol,
sector, country, sentiment_score, materiality_score, credibility_score,
urgency_score, expected_impact, time_horizon, status
"""

from __future__ import annotations

import argparse
import hashlib
import math
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE EXTENSION
# ============================================================

EXTRA_SCHEMA = """
ALTER TABLE news_events ADD COLUMN event_hash TEXT;
"""

INDEX_SCHEMA = """
CREATE UNIQUE INDEX IF NOT EXISTS idx_news_event_hash
ON news_events(event_hash)
WHERE event_hash IS NOT NULL AND event_hash <> '';
"""


def ensure_news_schema() -> None:
    setup_database()

    with connect() as connection:
        columns = {
            row["name"]
            for row in connection.execute(
                "PRAGMA table_info(news_events)"
            ).fetchall()
        }

        if "event_hash" not in columns:
            connection.execute(
                "ALTER TABLE news_events ADD COLUMN event_hash TEXT"
            )

        connection.executescript(INDEX_SCHEMA)
        connection.commit()


# ============================================================
# NORMALIZATION / SCORING
# ============================================================

def clean_text(value: Any) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"nan", "none", "null"} else text


def normalize_symbol(value: Any) -> str:
    symbol = clean_text(value).upper()

    if symbol.endswith(".NS"):
        symbol = symbol[:-3]

    return symbol.replace(" ", "")


def safe_float(
    value: Any,
    default: float | None = None,
) -> float | None:
    try:
        if value is None or value == "":
            return default

        return float(value)

    except (TypeError, ValueError):
        return default


def clamp(
    value: float,
    minimum: float,
    maximum: float,
) -> float:
    return max(minimum, min(maximum, value))


def parse_event_time(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")

    text = clean_text(value)

    if not text:
        return datetime.now().isoformat(timespec="seconds")

    parsed = pd.to_datetime(text, errors="coerce")

    if pd.isna(parsed):
        raise ValueError(f"Invalid event_time: {text}")

    return parsed.to_pydatetime().isoformat(timespec="seconds")


def event_hash(
    event_time: str,
    source: str,
    headline: str,
    symbol: str,
) -> str:
    raw = "|".join(
        [
            event_time[:10],
            source.strip().lower(),
            headline.strip().lower(),
            symbol.strip().upper(),
        ]
    )

    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def expected_impact_direction(
    sentiment_score: float,
    expected_impact: str,
) -> int:
    text = expected_impact.strip().upper()

    if "NEGATIVE" in text or "BEARISH" in text:
        return -1

    if "POSITIVE" in text or "BULLISH" in text:
        return 1

    if sentiment_score > 0:
        return 1

    if sentiment_score < 0:
        return -1

    return 0


def recency_score(
    event_time: str,
    half_life_hours: float = 48.0,
) -> float:
    parsed = datetime.fromisoformat(event_time)
    age_hours = max(
        0.0,
        (datetime.now() - parsed).total_seconds() / 3600,
    )

    return 100 * math.pow(
        0.5,
        age_hours / half_life_hours,
    )


def calculate_impact_score(record: dict) -> float:
    sentiment = clamp(
        float(record.get("sentiment_score") or 0),
        -100,
        100,
    )

    materiality = clamp(
        float(record.get("materiality_score") or 50),
        0,
        100,
    )

    credibility = clamp(
        float(record.get("credibility_score") or 50),
        0,
        100,
    )

    urgency = clamp(
        float(record.get("urgency_score") or 50),
        0,
        100,
    )

    recency = recency_score(record["event_time"])

    direction = expected_impact_direction(
        sentiment,
        record.get("expected_impact", ""),
    )

    magnitude = (
        materiality * 0.35
        + credibility * 0.30
        + urgency * 0.15
        + recency * 0.20
    )

    sentiment_strength = abs(sentiment) / 100

    if sentiment_strength == 0 and direction != 0:
        sentiment_strength = 0.5

    score = magnitude * sentiment_strength * direction

    return round(clamp(score, -100, 100), 2)


# ============================================================
# DATABASE OPERATIONS
# ============================================================

def insert_event(
    connection,
    record: dict,
) -> bool:
    hashed = event_hash(
        record["event_time"],
        record["source"],
        record["headline"],
        record["nse_symbol"],
    )

    try:
        connection.execute(
            """
            INSERT INTO news_events(
                event_time,
                source,
                headline,
                url,
                event_type,
                company_name,
                nse_symbol,
                sector,
                country,
                sentiment_score,
                materiality_score,
                credibility_score,
                urgency_score,
                expected_impact,
                time_horizon,
                status,
                raw_payload,
                created_at,
                event_hash
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                record["event_time"],
                record["source"],
                record["headline"],
                record.get("url", ""),
                record.get("event_type", ""),
                record.get("company_name", ""),
                record.get("nse_symbol", ""),
                record.get("sector", ""),
                record.get("country", "India"),
                record.get("sentiment_score"),
                record.get("materiality_score"),
                record.get("credibility_score"),
                record.get("urgency_score"),
                record.get("expected_impact", ""),
                record.get("time_horizon", ""),
                record.get("status", "NEW"),
                record.get("raw_payload", ""),
                datetime.now().isoformat(timespec="seconds"),
                hashed,
            ),
        )
        return True

    except Exception as error:
        if "UNIQUE constraint failed" in str(error):
            return False

        raise


def import_csv(path: Path) -> tuple[int, int]:
    ensure_news_schema()

    if not path.exists():
        raise FileNotFoundError(path)

    frame = pd.read_csv(path)

    required = {"source", "headline"}

    if not required.issubset(frame.columns):
        raise RuntimeError(
            "CSV must contain at least source and headline columns."
        )

    inserted = 0
    duplicates = 0

    run_id = start_run(
        "aqsd_news_intelligence",
        f"Importing {path.name}",
    )

    try:
        with connect() as connection:
            for _, row in frame.iterrows():
                record = {
                    "event_time": parse_event_time(
                        row.get("event_time")
                    ),
                    "source": clean_text(row.get("source")),
                    "headline": clean_text(row.get("headline")),
                    "url": clean_text(row.get("url")),
                    "event_type": clean_text(row.get("event_type")),
                    "company_name": clean_text(
                        row.get("company_name")
                    ),
                    "nse_symbol": normalize_symbol(
                        row.get("nse_symbol")
                    ),
                    "sector": clean_text(row.get("sector")),
                    "country": clean_text(
                        row.get("country")
                    ) or "India",
                    "sentiment_score": safe_float(
                        row.get("sentiment_score"),
                        0,
                    ),
                    "materiality_score": safe_float(
                        row.get("materiality_score"),
                        50,
                    ),
                    "credibility_score": safe_float(
                        row.get("credibility_score"),
                        50,
                    ),
                    "urgency_score": safe_float(
                        row.get("urgency_score"),
                        50,
                    ),
                    "expected_impact": clean_text(
                        row.get("expected_impact")
                    ),
                    "time_horizon": clean_text(
                        row.get("time_horizon")
                    ),
                    "status": clean_text(
                        row.get("status")
                    ) or "NEW",
                    "raw_payload": "",
                }

                if not record["headline"]:
                    continue

                if insert_event(connection, record):
                    inserted += 1
                else:
                    duplicates += 1

            connection.commit()

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=inserted,
            errors_count=duplicates,
            message=(
                f"Inserted={inserted}; "
                f"duplicates={duplicates}"
            ),
        )

        return inserted, duplicates

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=inserted,
            errors_count=duplicates + 1,
            message=str(error),
        )
        raise


def add_manual_event() -> None:
    print("\nAQSD MANUAL NEWS EVENT")
    print("=" * 72)

    record = {
        "event_time": parse_event_time(
            input("Event time [blank = now]: ").strip()
        ),
        "source": input("Source: ").strip(),
        "headline": input("Headline: ").strip(),
        "url": input("URL [optional]: ").strip(),
        "event_type": input("Event type: ").strip(),
        "company_name": input("Company name [optional]: ").strip(),
        "nse_symbol": normalize_symbol(
            input("NSE symbol [optional]: ").strip()
        ),
        "sector": input("Sector [optional]: ").strip(),
        "country": input("Country [India]: ").strip() or "India",
        "sentiment_score": safe_float(
            input("Sentiment -100 to +100 [0]: ").strip(),
            0,
        ),
        "materiality_score": safe_float(
            input("Materiality 0-100 [50]: ").strip(),
            50,
        ),
        "credibility_score": safe_float(
            input("Credibility 0-100 [50]: ").strip(),
            50,
        ),
        "urgency_score": safe_float(
            input("Urgency 0-100 [50]: ").strip(),
            50,
        ),
        "expected_impact": input(
            "Expected impact [Positive/Negative/Neutral]: "
        ).strip(),
        "time_horizon": input(
            "Time horizon [Intraday/Swing/Positional]: "
        ).strip(),
        "status": "NEW",
        "raw_payload": "",
    }

    if not record["source"] or not record["headline"]:
        raise ValueError("Source and headline are required.")

    with connect() as connection:
        inserted = insert_event(connection, record)
        connection.commit()

    print(
        "Event added."
        if inserted
        else "Duplicate event already exists."
    )


# ============================================================
# SCORING / AGGREGATION
# ============================================================

def load_recent_events(days: int = 30) -> pd.DataFrame:
    cutoff = (
        datetime.now() - timedelta(days=max(days, 1))
    ).isoformat(timespec="seconds")

    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT *
            FROM news_events
            WHERE event_time >= ?
            ORDER BY event_time DESC
            """,
            connection,
            params=(cutoff,),
        )

    return frame


def score_events(days: int = 30) -> pd.DataFrame:
    frame = load_recent_events(days)

    if frame.empty:
        return frame

    frame["news_impact_score"] = frame.apply(
        lambda row: calculate_impact_score(
            row.to_dict()
        ),
        axis=1,
    )

    return frame


def symbol_news_scores(days: int = 7) -> pd.DataFrame:
    frame = score_events(days)

    if frame.empty:
        return frame

    frame = frame[
        frame["nse_symbol"].fillna("").str.strip() != ""
    ].copy()

    if frame.empty:
        return frame

    grouped = (
        frame.groupby("nse_symbol")
        .agg(
            Event_Count=("event_id", "count"),
            News_Score=("news_impact_score", "sum"),
            Max_Impact=("news_impact_score", "max"),
            Min_Impact=("news_impact_score", "min"),
            Latest_Event=("event_time", "max"),
        )
        .reset_index()
    )

    grouped["News_Score"] = grouped["News_Score"].clip(
        -100,
        100,
    ).round(2)

    grouped["News_Bias"] = grouped["News_Score"].apply(
        lambda value: (
            "STRONG POSITIVE"
            if value >= 60
            else "POSITIVE"
            if value >= 20
            else "STRONG NEGATIVE"
            if value <= -60
            else "NEGATIVE"
            if value <= -20
            else "NEUTRAL"
        )
    )

    return grouped.sort_values(
        "News_Score",
        ascending=False,
    )


# ============================================================
# EXCEL REPORT
# ============================================================

def write_report(days: int = 30) -> None:
    events = score_events(days)
    symbol_scores = symbol_news_scores(min(days, 7))

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "News Intelligence" in wb.sheetnames:
        del wb["News Intelligence"]

    ws = wb.create_sheet("News Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:P2")
    ws["A1"] = "AQSD PROFESSIONAL - NEWS & EVENT INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Events Analysed"
    ws["B4"] = len(events)
    ws["D4"] = "Symbols Scored"
    ws["E4"] = len(symbol_scores)
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    summary_headers = [
        "Rank",
        "Symbol",
        "Event Count",
        "News Score",
        "News Bias",
        "Latest Event",
    ]

    for col, heading in enumerate(summary_headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)

    for row_no, (_, row) in enumerate(
        symbol_scores.iterrows(),
        start=8,
    ):
        values = [
            row_no - 7,
            row["nse_symbol"],
            row["Event_Count"],
            row["News_Score"],
            row["News_Bias"],
            row["Latest_Event"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        score = float(row["News_Score"])

        ws.cell(row_no, 4).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 20
                else RED
                if score <= -20
                else YELLOW
            ),
        )

    event_start = max(12, 10 + len(symbol_scores))

    ws.cell(
        event_start,
        1,
        "DETAILED EVENTS",
    ).font = Font(
        size=14,
        bold=True,
        color=WHITE,
    )
    ws.cell(
        event_start,
        1,
    ).fill = PatternFill("solid", fgColor=NAVY)

    event_headers = [
        "Event Time",
        "Source",
        "Headline",
        "Event Type",
        "Company",
        "Symbol",
        "Sector",
        "Country",
        "Sentiment",
        "Materiality",
        "Credibility",
        "Urgency",
        "Impact Score",
        "Expected Impact",
        "Time Horizon",
        "Status",
    ]

    for col, heading in enumerate(event_headers, start=1):
        cell = ws.cell(event_start + 2, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        events.iterrows(),
        start=event_start + 3,
    ):
        values = [
            row.get("event_time"),
            row.get("source"),
            row.get("headline"),
            row.get("event_type"),
            row.get("company_name"),
            row.get("nse_symbol"),
            row.get("sector"),
            row.get("country"),
            row.get("sentiment_score"),
            row.get("materiality_score"),
            row.get("credibility_score"),
            row.get("urgency_score"),
            row.get("news_impact_score"),
            row.get("expected_impact"),
            row.get("time_horizon"),
            row.get("status"),
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(row_no, col, value).border = Border(bottom=THIN)

        impact = float(row.get("news_impact_score") or 0)

        ws.cell(row_no, 13).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if impact >= 20
                else RED
                if impact <= -20
                else GREY
            ),
        )

    widths = {
        "A": 19,
        "B": 22,
        "C": 60,
        "D": 20,
        "E": 28,
        "F": 16,
        "G": 18,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 14,
        "N": 18,
        "O": 18,
        "P": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


# ============================================================
# STATUS
# ============================================================

def show_status() -> None:
    ensure_news_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT nse_symbol) AS symbols,
                MIN(event_time) AS first_event,
                MAX(event_time) AS latest_event
            FROM news_events
            """
        ).fetchone()

    print("\nAQSD NEWS INTELLIGENCE STATUS")
    print("=" * 72)
    print(f"Stored events:       {row['total'] or 0}")
    print(f"Symbols covered:     {row['symbols'] or 0}")
    print(f"First event:         {row['first_event'] or 'No data'}")
    print(f"Latest event:        {row['latest_event'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD News and Event Intelligence Foundation."
    )

    parser.add_argument(
        "--setup",
        action="store_true",
        help="Prepare the news-event schema.",
    )

    parser.add_argument(
        "--import-csv",
        type=Path,
        help="Import structured news events from CSV.",
    )

    parser.add_argument(
        "--add-manual",
        action="store_true",
        help="Add one news event interactively.",
    )

    parser.add_argument(
        "--score",
        action="store_true",
        help="Display recent symbol-level news scores.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Create the News Intelligence Excel sheet.",
    )

    parser.add_argument(
        "--days",
        type=int,
        default=30,
        help="Lookback days for scoring and reporting.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show news database status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    ensure_news_schema()

    if args.setup:
        print("AQSD News Intelligence schema is ready.")
        return

    if args.import_csv:
        inserted, duplicates = import_csv(args.import_csv)

        print("\nAQSD NEWS IMPORT")
        print("=" * 72)
        print(f"Inserted:   {inserted}")
        print(f"Duplicates: {duplicates}")
        return

    if args.add_manual:
        add_manual_event()
        return

    if args.score:
        frame = symbol_news_scores(args.days)

        if frame.empty:
            print("No scored symbol news found.")
        else:
            print(
                frame.to_string(index=False)
            )
        return

    if args.report:
        write_report(args.days)
        print(f"News Intelligence report created:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
