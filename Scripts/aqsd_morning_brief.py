
"""
AQSD Professional
Module: Morning Brief Generator
Version: 1.0

Purpose
-------
Creates a concise daily market brief from AQSD's latest intelligence.

Inputs
------
- Market Regime Intelligence
- Market Breadth Intelligence
- Sector Rotation Intelligence
- Unified Master Intelligence
- Trade Decision Engine
- Portfolio Allocation
- AQSD Alerts
- Global Markets Intelligence

Outputs
-------
- Console morning brief
- Text report
- CSV summary files
- Excel sheet: AQSD Morning Brief

Commands
--------
python aqsd_morning_brief.py --run
python aqsd_morning_brief.py --status
python aqsd_morning_brief.py --report
python aqsd_morning_brief.py --top 10
"""

from __future__ import annotations

import argparse
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data"
OUTPUT_DIR = BASE_DIR / "Output"

DATABASE = DATA_DIR / "aqsd_core.db"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"

TEXT_REPORT = OUTPUT_DIR / "AQSD_Morning_Brief.txt"
TOP_STOCKS_CSV = OUTPUT_DIR / "AQSD_Morning_Top_Stocks.csv"
TOP_SECTORS_CSV = OUTPUT_DIR / "AQSD_Morning_Top_Sectors.csv"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE HELPERS
# ============================================================

def connect() -> sqlite3.Connection:
    connection = sqlite3.connect(DATABASE)
    connection.row_factory = sqlite3.Row
    return connection


def table_exists(
    connection: sqlite3.Connection,
    table_name: str,
) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table'
          AND name=?
        """,
        (table_name,),
    ).fetchone()

    return row is not None


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


# ============================================================
# LOADERS
# ============================================================

def load_regime() -> dict:
    with connect() as connection:
        if not table_exists(
            connection,
            "market_regime_intelligence",
        ):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def load_market_breadth() -> dict:
    with connect() as connection:
        if not table_exists(
            connection,
            "market_breadth_intelligence",
        ):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM market_breadth_intelligence
            WHERE scope_type='MARKET'
              AND trade_date=(
                  SELECT MAX(trade_date)
                  FROM market_breadth_intelligence
              )
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def load_top_sectors(limit: int) -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "sector_rotation_intelligence",
        ):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                sector,
                sector_rotation_score,
                rotation_state,
                trend_state,
                bullish_breadth_percent,
                average_5d_return,
                average_20d_return,
                leader_symbol,
                leader_score
            FROM sector_rotation_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            ORDER BY sector_rotation_score DESC
            LIMIT ?
            """,
            connection,
            params=(limit,),
        )


def load_top_decisions(limit: int) -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "aqsd_trade_decisions",
        ):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                priority_rank,
                nse_symbol,
                sector,
                action,
                master_score,
                confidence_percent,
                completeness_percent,
                risk_level,
                entry_quality,
                last_price,
                entry_low,
                entry_high,
                stop_loss,
                target_1,
                target_2,
                reward_risk_1,
                priority_score
            FROM aqsd_trade_decisions
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
            ORDER BY priority_rank
            LIMIT ?
            """,
            connection,
            params=(limit,),
        )


def load_avoid_list(limit: int) -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "aqsd_trade_decisions",
        ):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                nse_symbol,
                sector,
                action,
                master_score,
                confidence_percent,
                risk_level,
                directional_bias
            FROM aqsd_trade_decisions
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
              AND action IN ('AVOID', 'EXIT / AVOID')
            ORDER BY master_score ASC
            LIMIT ?
            """,
            connection,
            params=(limit,),
        )


def load_portfolio_allocation(limit: int) -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "portfolio_allocation",
        ):
            return pd.DataFrame()

        columns = {
            row["name"]
            for row in connection.execute(
                "PRAGMA table_info(portfolio_allocation)"
            ).fetchall()
        }

        symbol_column = (
            "symbol"
            if "symbol" in columns
            else "nse_symbol"
        )

        query = f"""
            SELECT
                trade_date,
                {symbol_column} AS nse_symbol,
                action,
                priority_rank,
                allocation_percent,
                conviction
            FROM portfolio_allocation
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM portfolio_allocation
            )
            ORDER BY allocation_percent DESC
            LIMIT ?
        """

        return pd.read_sql_query(
            query,
            connection,
            params=(limit,),
        )


def load_alerts(limit: int) -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "aqsd_alerts",
        ):
            return pd.DataFrame()

        columns = {
            row["name"]
            for row in connection.execute(
                "PRAGMA table_info(aqsd_alerts)"
            ).fetchall()
        }

        priority_order = (
            "priority_score DESC"
            if "priority_score" in columns
            else """
                CASE severity
                    WHEN 'CRITICAL' THEN 1
                    WHEN 'HIGH' THEN 2
                    WHEN 'MEDIUM' THEN 3
                    WHEN 'LOW' THEN 4
                    ELSE 5
                END
            """
        )

        query = f"""
            SELECT
                severity,
                alert_type,
                nse_symbol,
                sector,
                title,
                message,
                status
            FROM aqsd_alerts
            WHERE alert_date=(
                SELECT MAX(alert_date)
                FROM aqsd_alerts
            )
            ORDER BY {priority_order}
            LIMIT ?
        """

        return pd.read_sql_query(
            query,
            connection,
            params=(limit,),
        )


def load_global_snapshot() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "global_markets",
        ):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT *
            FROM global_markets
            WHERE snapshot_date=(
                SELECT MAX(snapshot_date)
                FROM global_markets
            )
            ORDER BY symbol
            """,
            connection,
        )


# ============================================================
# BRIEF BUILDING
# ============================================================

def global_summary(frame: pd.DataFrame) -> list[str]:
    if frame.empty:
        return ["Global markets data unavailable."]

    lookup = {
        str(row.get("symbol")): row
        for _, row in frame.iterrows()
    }

    names = [
        ("^DJI", "Dow Jones"),
        ("^IXIC", "Nasdaq"),
        ("^GSPC", "S&P 500"),
        ("^VIX", "VIX"),
        ("DX-Y.NYB", "Dollar Index"),
        ("INR=X", "USD/INR"),
        ("^TNX", "US 10Y Yield"),
    ]

    lines = []

    for symbol, name in names:
        row = lookup.get(symbol)

        if row is None:
            continue

        daily = (
            safe_float(
                row.get("daily_change_percent")
            )
            if hasattr(row, "get")
            else None
        )

        five_day = (
            safe_float(
                row.get("five_day_change_percent")
            )
            if hasattr(row, "get")
            else None
        )

        parts = [name]

        if daily is not None:
            parts.append(f"1D {daily:+.2f}%")

        if five_day is not None:
            parts.append(f"5D {five_day:+.2f}%")

        lines.append(": ".join([parts[0], ", ".join(parts[1:])]))

    return lines or ["Global market symbols were not found."]


def build_text_brief(
    regime: dict,
    breadth: dict,
    sectors: pd.DataFrame,
    decisions: pd.DataFrame,
    avoid: pd.DataFrame,
    allocation: pd.DataFrame,
    alerts: pd.DataFrame,
    global_frame: pd.DataFrame,
) -> str:
    lines = [
        "AQSD PROFESSIONAL - MORNING MARKET BRIEF",
        "=" * 88,
        f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}",
        "",
        "MARKET REGIME",
        "-" * 88,
    ]

    if regime:
        lines.extend(
            [
                f"Regime: {regime.get('market_regime')}",
                f"Regime Score: {regime.get('regime_score')}",
                f"Bull Probability: {regime.get('bull_probability')}%",
                f"Bear Probability: {regime.get('bear_probability')}%",
                f"Range Probability: {regime.get('range_probability')}%",
                f"Risk State: {regime.get('risk_state')}",
                f"Volatility: {regime.get('volatility_state')}",
                f"Strategy: {regime.get('suggested_strategy')}",
                f"Suggested Capital Exposure: "
                f"{regime.get('capital_exposure_percent')}%",
            ]
        )
    else:
        lines.append("Market Regime data unavailable.")

    lines.extend(
        [
            "",
            "MARKET BREADTH",
            "-" * 88,
        ]
    )

    if breadth:
        lines.extend(
            [
                f"Breadth Score: {breadth.get('breadth_score')}",
                f"Breadth Regime: {breadth.get('breadth_regime')}",
                f"Advances / Declines: "
                f"{breadth.get('advances')} / {breadth.get('declines')}",
                f"Above 20DMA: {breadth.get('above_20dma_percent')}%",
                f"Above 50DMA: {breadth.get('above_50dma_percent')}%",
                f"Above 200DMA: {breadth.get('above_200dma_percent')}%",
                f"20D Highs / Lows: "
                f"{breadth.get('new_20d_highs')} / "
                f"{breadth.get('new_20d_lows')}",
                f"52W Highs / Lows: "
                f"{breadth.get('new_52w_highs')} / "
                f"{breadth.get('new_52w_lows')}",
            ]
        )
    else:
        lines.append("Market Breadth data unavailable.")

    lines.extend(
        [
            "",
            "GLOBAL MARKETS",
            "-" * 88,
            *global_summary(global_frame),
            "",
            "TOP SECTORS",
            "-" * 88,
        ]
    )

    if sectors.empty:
        lines.append("Sector Rotation data unavailable.")
    else:
        lines.append(
            sectors.to_string(index=False)
        )

    lines.extend(
        [
            "",
            "TOP TRADE DECISIONS",
            "-" * 88,
        ]
    )

    if decisions.empty:
        lines.append("No trade decisions available.")
    else:
        lines.append(
            decisions.to_string(index=False)
        )

    lines.extend(
        [
            "",
            "PORTFOLIO ALLOCATION",
            "-" * 88,
        ]
    )

    if allocation.empty:
        lines.append("No portfolio allocation available.")
    else:
        lines.append(
            allocation.to_string(index=False)
        )

    lines.extend(
        [
            "",
            "AVOID / EXIT LIST",
            "-" * 88,
        ]
    )

    if avoid.empty:
        lines.append("No Avoid or Exit signals.")
    else:
        lines.append(
            avoid.to_string(index=False)
        )

    lines.extend(
        [
            "",
            "IMPORTANT ALERTS",
            "-" * 88,
        ]
    )

    if alerts.empty:
        lines.append("No alerts available.")
    else:
        lines.append(
            alerts.to_string(index=False)
        )

    lines.extend(
        [
            "",
            "=" * 88,
            "AQSD is a decision-support system. Review risk, liquidity, "
            "price action and position size before trading.",
            "=" * 88,
        ]
    )

    return "\n".join(lines)


# ============================================================
# EXCEL REPORT
# ============================================================

def write_excel(
    regime: dict,
    breadth: dict,
    sectors: pd.DataFrame,
    decisions: pd.DataFrame,
    avoid: pd.DataFrame,
    allocation: pd.DataFrame,
    alerts: pd.DataFrame,
) -> None:
    if DASHBOARD.exists():
        workbook = load_workbook(DASHBOARD)
    else:
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    sheet_name = "AQSD Morning Brief"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(
        sheet_name,
        0,
    )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:P2")
    ws["A1"] = "AQSD PROFESSIONAL - MORNING MARKET BRIEF"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Market Regime"
    ws["B4"] = regime.get(
        "market_regime",
        "NO DATA",
    )
    ws["D4"] = "Breadth Score"
    ws["E4"] = breadth.get(
        "breadth_score",
    )
    ws["G4"] = "Capital Exposure"
    ws["H4"] = regime.get(
        "capital_exposure_percent",
    )
    ws["J4"] = "Updated"
    ws["K4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    for ref in ("A4", "D4", "G4", "J4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    regime_text = str(
        regime.get("market_regime", "")
    )

    ws["B4"].fill = PatternFill(
        "solid",
        fgColor=(
            GREEN
            if regime_text in {
                "BULL TREND",
                "ACCUMULATION",
                "RISK ON",
            }
            else RED
            if regime_text in {
                "BEAR TREND",
                "DISTRIBUTION",
                "RISK OFF",
                "HIGH VOLATILITY",
            }
            else YELLOW
        ),
    )
    ws["B4"].font = Font(bold=True)

    row = 7

    def section_title(
        title: str,
        start_row: int,
    ) -> int:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=16,
        )
        cell = ws.cell(
            start_row,
            1,
            title,
        )
        cell.font = Font(
            size=14,
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        return start_row + 1

    def write_dataframe(
        frame: pd.DataFrame,
        start_row: int,
    ) -> int:
        if frame.empty:
            ws.cell(
                start_row,
                1,
                "No data available.",
            )
            return start_row + 2

        for col, heading in enumerate(
            frame.columns,
            start=1,
        ):
            cell = ws.cell(
                start_row,
                col,
                str(heading),
            )
            cell.font = Font(
                bold=True,
                color=WHITE,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=NAVY,
            )
            cell.alignment = Alignment(
                horizontal="center",
                wrap_text=True,
            )

        for row_no, values in enumerate(
            frame.itertuples(
                index=False,
                name=None,
            ),
            start=start_row + 1,
        ):
            for col, value in enumerate(
                values,
                start=1,
            ):
                ws.cell(
                    row_no,
                    col,
                    value,
                ).border = Border(
                    bottom=THIN
                )

        return start_row + len(frame) + 3

    row = section_title(
        "MARKET OVERVIEW",
        row,
    )

    overview = pd.DataFrame(
        [
            {
                "Regime": regime.get("market_regime"),
                "Regime Score": regime.get("regime_score"),
                "Bull %": regime.get("bull_probability"),
                "Bear %": regime.get("bear_probability"),
                "Range %": regime.get("range_probability"),
                "Risk State": regime.get("risk_state"),
                "Volatility": regime.get("volatility_state"),
                "Strategy": regime.get("suggested_strategy"),
                "Exposure %": regime.get("capital_exposure_percent"),
                "Breadth Score": breadth.get("breadth_score"),
                "Advances": breadth.get("advances"),
                "Declines": breadth.get("declines"),
            }
        ]
    )

    row = write_dataframe(
        overview,
        row,
    )

    row = section_title(
        "TOP SECTORS",
        row,
    )
    row = write_dataframe(
        sectors,
        row,
    )

    row = section_title(
        "TOP TRADE DECISIONS",
        row,
    )
    row = write_dataframe(
        decisions,
        row,
    )

    row = section_title(
        "PORTFOLIO ALLOCATION",
        row,
    )
    row = write_dataframe(
        allocation,
        row,
    )

    row = section_title(
        "AVOID / EXIT LIST",
        row,
    )
    row = write_dataframe(
        avoid,
        row,
    )

    row = section_title(
        "IMPORTANT ALERTS",
        row,
    )
    write_dataframe(
        alerts,
        row,
    )

    for column in "ABCDEFGHIJKLMNOP":
        ws.column_dimensions[column].width = 18

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 25
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 20
    ws.column_dimensions["O"].width = 20
    ws.column_dimensions["P"].width = 20

    workbook.save(DASHBOARD)


# ============================================================
# MAIN WORKFLOW
# ============================================================

def generate_brief(
    top: int,
) -> str:
    regime = load_regime()
    breadth = load_market_breadth()
    sectors = load_top_sectors(top)
    decisions = load_top_decisions(top)
    avoid = load_avoid_list(top)
    allocation = load_portfolio_allocation(top)
    alerts = load_alerts(top)
    global_frame = load_global_snapshot()

    brief = build_text_brief(
        regime,
        breadth,
        sectors,
        decisions,
        avoid,
        allocation,
        alerts,
        global_frame,
    )

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    TEXT_REPORT.write_text(
        brief,
        encoding="utf-8",
    )

    sectors.to_csv(
        TOP_SECTORS_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    decisions.to_csv(
        TOP_STOCKS_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    write_excel(
        regime,
        breadth,
        sectors,
        decisions,
        avoid,
        allocation,
        alerts,
    )

    return brief


def show_status() -> None:
    print("\nAQSD MORNING BRIEF STATUS")
    print("=" * 72)
    print(f"Database:    {DATABASE}")
    print(f"Dashboard:   {DASHBOARD}")
    print(f"Text report: {TEXT_REPORT}")
    print("-" * 72)

    with connect() as connection:
        tables = [
            "market_regime_intelligence",
            "market_breadth_intelligence",
            "sector_rotation_intelligence",
            "aqsd_trade_decisions",
            "portfolio_allocation",
            "aqsd_alerts",
            "global_markets",
        ]

        for table in tables:
            exists = table_exists(
                connection,
                table,
            )
            count = 0

            if exists:
                count = connection.execute(
                    f"SELECT COUNT(*) FROM {table}"
                ).fetchone()[0]

            print(
                f"{table:<38}"
                f"{'READY' if exists else 'MISSING':<10}"
                f"Rows: {count}"
            )

    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Morning Brief Generator."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Generate the latest morning brief.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild all morning-brief reports.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show morning-brief dependency status.",
    )

    parser.add_argument(
        "--top",
        type=int,
        default=10,
        help="Number of stocks, sectors and alerts to include.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if args.status:
        show_status()
        return

    top = max(
        1,
        min(50, args.top),
    )

    brief = generate_brief(top)

    print(brief)
    print()
    print(f"Text report: {TEXT_REPORT}")
    print(f"Top stocks:  {TOP_STOCKS_CSV}")
    print(f"Top sectors: {TOP_SECTORS_CSV}")
    print(f"Dashboard:   {DASHBOARD}")


if __name__ == "__main__":
    main()
