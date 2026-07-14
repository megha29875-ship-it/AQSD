
"""
AQSD Professional
Module: Trade Journal & Analytics
Version: 1.0
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def create_journal_sheet(wb):
    if "Trade Journal" in wb.sheetnames:
        return wb["Trade Journal"]

    ws = wb.create_sheet("Trade Journal")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:N2")
    ws["A1"] = "AQSD PROFESSIONAL - TRADE JOURNAL"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    summary = [
        ("Total Trades", 0),
        ("Winning Trades", 0),
        ("Losing Trades", 0),
        ("Win Rate", 0),
        ("Total P/L", 0),
        ("Average P/L", 0),
    ]

    for row_no, (label, value) in enumerate(summary, start=4):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)

    headers = [
        "Trade ID",
        "Symbol",
        "Side",
        "Entry Date",
        "Exit Date",
        "Qty",
        "Entry",
        "Exit",
        "Stop Loss",
        "Target",
        "P/L",
        "P/L %",
        "Result",
        "Notes",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    widths = {
        "A": 10, "B": 18, "C": 12, "D": 14, "E": 14,
        "F": 10, "G": 12, "H": 12, "I": 12, "J": 12,
        "K": 14, "L": 12, "M": 12, "N": 28,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    return ws


def existing_keys(ws) -> set[tuple]:
    keys = set()

    for row in range(11, ws.max_row + 1):
        trade_id = ws.cell(row, 1).value
        symbol = ws.cell(row, 2).value
        entry_date = ws.cell(row, 4).value

        if trade_id and symbol:
            keys.add((trade_id, str(symbol), str(entry_date)))

    return keys


def copy_closed_trades(wb, journal_ws) -> int:
    if "Portfolio" not in wb.sheetnames:
        raise RuntimeError("Portfolio sheet not found.")

    portfolio_ws = wb["Portfolio"]
    p_headers = header_map(portfolio_ws, 12)

    required = [
        "Trade ID", "Symbol", "Side", "Entry Date", "Qty",
        "Entry", "CMP", "Stop Loss", "P/L", "P/L %",
        "Target", "Status",
    ]

    missing = [name for name in required if name not in p_headers]

    if missing:
        raise RuntimeError(
            "Missing Portfolio columns: " + ", ".join(missing)
        )

    keys = existing_keys(journal_ws)
    added = 0

    for row in range(13, portfolio_ws.max_row + 1):
        status = str(
            portfolio_ws.cell(row, p_headers["Status"]).value or ""
        ).upper().strip()

        if status not in {"TARGET", "STOPPED"}:
            continue

        trade_id = portfolio_ws.cell(row, p_headers["Trade ID"]).value
        symbol = portfolio_ws.cell(row, p_headers["Symbol"]).value
        entry_date = portfolio_ws.cell(row, p_headers["Entry Date"]).value

        key = (trade_id, str(symbol), str(entry_date))

        if key in keys:
            continue

        pl = float(portfolio_ws.cell(row, p_headers["P/L"]).value or 0)
        result = "WIN" if pl > 0 else "LOSS"

        output_row = max(journal_ws.max_row + 1, 11)

        values = [
            trade_id,
            symbol,
            portfolio_ws.cell(row, p_headers["Side"]).value,
            entry_date,
            datetime.now(),
            portfolio_ws.cell(row, p_headers["Qty"]).value,
            portfolio_ws.cell(row, p_headers["Entry"]).value,
            portfolio_ws.cell(row, p_headers["CMP"]).value,
            portfolio_ws.cell(row, p_headers["Stop Loss"]).value,
            portfolio_ws.cell(row, p_headers["Target"]).value,
            pl,
            portfolio_ws.cell(row, p_headers["P/L %"]).value,
            result,
            "",
        ]

        for col, value in enumerate(values, start=1):
            cell = journal_ws.cell(output_row, col, value)
            cell.border = Border(bottom=THIN)

        journal_ws.cell(output_row, 13).fill = PatternFill(
            "solid",
            fgColor=GREEN if result == "WIN" else RED,
        )
        journal_ws.cell(output_row, 13).font = Font(bold=True)

        for col in [7, 8, 9, 10, 11]:
            journal_ws.cell(output_row, col).number_format = '₹#,##0.00'

        journal_ws.cell(output_row, 12).number_format = '0.00"%"'
        journal_ws.cell(output_row, 4).number_format = "dd-mm-yyyy"
        journal_ws.cell(output_row, 5).number_format = "dd-mm-yyyy"

        keys.add(key)
        added += 1

    return added


def update_analytics(ws) -> None:
    total = 0
    wins = 0
    losses = 0
    total_pl = 0.0

    for row in range(11, ws.max_row + 1):
        symbol = ws.cell(row, 2).value

        if not symbol:
            continue

        total += 1
        pl = float(ws.cell(row, 11).value or 0)
        total_pl += pl

        if pl > 0:
            wins += 1
        elif pl < 0:
            losses += 1

    win_rate = (wins / total * 100) if total else 0.0
    average_pl = (total_pl / total) if total else 0.0

    ws["B4"] = total
    ws["B5"] = wins
    ws["B6"] = losses
    ws["B7"] = round(win_rate, 2)
    ws["B8"] = round(total_pl, 2)
    ws["B9"] = round(average_pl, 2)

    ws["B7"].number_format = '0.00"%"'
    ws["B8"].number_format = '₹#,##0.00'
    ws["B9"].number_format = '₹#,##0.00'


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    journal_ws = create_journal_sheet(wb)
    added = copy_closed_trades(wb, journal_ws)
    update_analytics(journal_ws)

    wb.save(DASHBOARD)

    print("Trade Journal updated successfully.")
    print(f"New completed trades added: {added}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
