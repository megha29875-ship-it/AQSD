"""
AQSD Professional
Module: Sector Rotation Engine (Starter)
Version: 1.0

Ranks major NSE sectors using representative ETFs/indices.

Creates a "Sector Rotation" sheet in Dashboard.xlsx.
"""

from pathlib import Path
import yfinance as yf
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

SECTORS = {
    "Bank": "^NSEBANK",
    "IT": "^CNXIT",
    "Auto": "^CNXAUTO",
    "FMCG": "^CNXFMCG",
    "Pharma": "^CNXPHARMA",
    "Metal": "^CNXMETAL",
    "Energy": "^CNXENERGY",
    "Realty": "^CNXREALTY",
}

def ret(series, days):
    if len(series) <= days:
        return None
    return round((series.iloc[-1]/series.iloc[-1-days]-1)*100,2)

rows=[]
for sector,ticker in SECTORS.items():
    try:
        df=yf.download(ticker,period="1y",progress=False,auto_adjust=True)
        if df.empty:
            continue
        if isinstance(df.columns,pd.MultiIndex):
            df.columns=df.columns.get_level_values(0)
        c=df["Close"].dropna()
        r20=ret(c,20)
        r60=ret(c,60)
        score=(r20*0.6+r60*0.4) if r20 is not None and r60 is not None else 0
        rows.append([sector,ticker,r20,r60,round(score,2)])
    except Exception:
        pass

rows.sort(key=lambda x:x[4],reverse=True)

if DASHBOARD.exists():
    wb=load_workbook(DASHBOARD)
else:
    wb=Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

if "Sector Rotation" in wb.sheetnames:
    del wb["Sector Rotation"]

ws=wb.create_sheet("Sector Rotation",1)
ws.merge_cells("A1:F2")
c=ws["A1"]
c.value="AQSD - SECTOR ROTATION ENGINE"
c.font=Font(size=18,bold=True,color="FFFFFF")
c.fill=PatternFill("solid",fgColor="17365D")
c.alignment=Alignment(horizontal="center")

headers=["Rank","Sector","Ticker","20D %","60D %","Rotation Score"]
for i,h in enumerate(headers,1):
    cell=ws.cell(4,i,h)
    cell.font=Font(bold=True,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor="17365D")

for r,row in enumerate(rows,5):
    ws.cell(r,1,r-4)
    for cidx,val in enumerate(row,2):
        ws.cell(r,cidx,val)

for col in "ABCDEF":
    ws.column_dimensions[col].width=18

wb.save(DASHBOARD)
print("Sector Rotation sheet created.")
