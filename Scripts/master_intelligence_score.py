
"""
AQSD Professional
Module: Master Intelligence Score
Version: 2.0

Combines:
- Market Structure
- Trend Intelligence
- Relative Strength
- Sector Rotation
- Pivot & CPR Intelligence

Creates:
- Master Intelligence sheet
- AQSD Master Score from 0 to 100
- Directional bias
- Explainable recommendation
- Confidence grade
- Engine agreement
- Data completeness

Run:
    python master_intelligence_score_v2.py
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

WEIGHTS = {
    "Structure": 0.25,
    "Trend": 0.25,
    "Relative Strength": 0.20,
    "Sector": 0.10,
    "Pivot": 0.20,
}

SECTOR_MAP = {
    "RELIANCE.NS": "Energy",
    "ONGC.NS": "Energy",
    "IOC.NS": "Energy",
    "BPCL.NS": "Energy",

    "HDFCBANK.NS": "Bank",
    "ICICIBANK.NS": "Bank",
    "SBIN.NS": "Bank",
    "AXISBANK.NS": "Bank",
    "KOTAKBANK.NS": "Bank",
    "INDUSINDBK.NS": "Bank",
    "BANKBARODA.NS": "Bank",
    "FEDERALBNK.NS": "Bank",

    "TCS.NS": "IT",
    "INFY.NS": "IT",
    "HCLTECH.NS": "IT",
    "WIPRO.NS": "IT",
    "TECHM.NS": "IT",
    "LTIM.NS": "IT",

    "TATAMOTORS.NS": "Auto",
    "M&M.NS": "Auto",
    "MARUTI.NS": "Auto",
    "BAJAJ-AUTO.NS": "Auto",
    "EICHERMOT.NS": "Auto",
    "HEROMOTOCO.NS": "Auto",

    "HINDUNILVR.NS": "FMCG",
    "ITC.NS": "FMCG",
    "NESTLEIND.NS": "FMCG",
    "BRITANNIA.NS": "FMCG",
    "DABUR.NS": "FMCG",
    "MARICO.NS": "FMCG",

    "SUNPHARMA.NS": "Pharma",
    "CIPLA.NS": "Pharma",
    "DRREDDY.NS": "Pharma",
    "DIVISLAB.NS": "Pharma",
    "BIOCON.NS": "Pharma",
    "LUPIN.NS": "Pharma",

    "TATASTEEL.NS": "Metal",
    "JSWSTEEL.NS": "Metal",
    "HINDALCO.NS": "Metal",
    "VEDL.NS": "Metal",
    "NMDC.NS": "Metal",

    "DLF.NS": "Realty",
    "GODREJPROP.NS": "Realty",
    "OBEROIRLTY.NS": "Realty",
    "PRESTIGE.NS": "Realty",

    "LT.NS": "Infrastructure",
    "SIEMENS.NS": "Capital Goods",
    "ABB.NS": "Capital Goods",
    "BHEL.NS": "Capital Goods",

    "BEL.NS": "Defence",
    "HAL.NS": "Defence",
}

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_symbol(value) -> str:
    symbol = str(value or "").strip().upper()

    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"

    return symbol


def read_score_sheet(
    wb,
    sheet_name: str,
    header_row: int,
    score_header: str,
    extra_headers: list[str],
) -> dict[str, dict]:
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    headers = header_map(ws, header_row)

    if "Symbol" not in headers or score_header not in headers:
        return {}

    data: dict[str, dict] = {}

    for row in range(header_row + 1, ws.max_row + 1):
        symbol = normalize_symbol(
            ws.cell(row, headers["Symbol"]).value
        )

        score = safe_float(
            ws.cell(row, headers[score_header]).value
        )

        if not symbol or score is None:
            continue

        record = {
            "Score": max(0.0, min(100.0, score)),
        }

        for header in extra_headers:
            record[header] = (
                ws.cell(row, headers[header]).value
                if header in headers
                else ""
            )

        data[symbol] = record

    return data


def read_sector_rotation(wb) -> dict[str, dict]:
    if "Sector Rotation" not in wb.sheetnames:
        return {}

    ws = wb["Sector Rotation"]
    headers = header_map(ws, 4)

    if "Sector" not in headers or "Rotation Score" not in headers:
        return {}

    raw = []

    for row in range(5, ws.max_row + 1):
        sector = str(
            ws.cell(row, headers["Sector"]).value or ""
        ).strip()

        score = safe_float(
            ws.cell(row, headers["Rotation Score"]).value
        )

        if sector and score is not None:
            raw.append((sector, score))

    if not raw:
        return {}

    values = [score for _, score in raw]
    minimum = min(values)
    maximum = max(values)

    output = {}

    for rank, (sector, raw_score) in enumerate(
        sorted(raw, key=lambda item: item[1], reverse=True),
        start=1,
    ):
        normalized = (
            50.0
            if maximum == minimum
            else (raw_score - minimum) / (maximum - minimum) * 100
        )

        output[sector.upper()] = {
            "Score": round(normalized, 2),
            "Rank": rank,
            "Raw Score": raw_score,
        }

    return output


def directional_bias(score: float) -> str:
    if score >= 80:
        return "VERY BULLISH"
    if score >= 65:
        return "BULLISH"
    if score >= 55:
        return "MILD BULLISH"
    if score <= 20:
        return "VERY BEARISH"
    if score <= 35:
        return "BEARISH"
    if score <= 45:
        return "MILD BEARISH"
    return "NEUTRAL"


def recommendation(score: float, completeness: float) -> str:
    if completeness < 50:
        return "INSUFFICIENT DATA"

    if score >= 85:
        return "STRONG BUY CANDIDATE"
    if score >= 72:
        return "BUY CANDIDATE"
    if score >= 60:
        return "WATCH FOR LONG"
    if score <= 15:
        return "STRONG SELL CANDIDATE"
    if score <= 28:
        return "SELL CANDIDATE"
    if score <= 40:
        return "WATCH FOR SHORT"
    return "NEUTRAL / WAIT"


def confidence_grade(
    score: float,
    completeness: float,
    agreement: float,
) -> str:
    directional_strength = abs(score - 50) * 2

    confidence = (
        directional_strength * 0.5
        + completeness * 0.25
        + agreement * 0.25
    )

    if confidence >= 85:
        return "A+"
    if confidence >= 70:
        return "A"
    if confidence >= 55:
        return "B"
    if confidence >= 40:
        return "C"
    return "D"


def combine_scores(
    structure: dict[str, dict],
    trend: dict[str, dict],
    relative_strength: dict[str, dict],
    pivots: dict[str, dict],
    sectors: dict[str, dict],
) -> list[dict]:
    symbols = sorted(
        set(structure)
        | set(trend)
        | set(relative_strength)
        | set(pivots)
    )

    results = []

    for symbol in symbols:
        sector = SECTOR_MAP.get(symbol, "Others")
        sector_data = sectors.get(sector.upper())

        component_values = {
            "Structure": structure.get(symbol, {}).get("Score"),
            "Trend": trend.get(symbol, {}).get("Score"),
            "Relative Strength": relative_strength.get(symbol, {}).get("Score"),
            "Sector": (
                sector_data.get("Score")
                if sector_data
                else None
            ),
            "Pivot": pivots.get(symbol, {}).get("Score"),
        }

        weighted_sum = 0.0
        available_weight = 0.0
        available_values = []

        for engine, value in component_values.items():
            if value is None:
                continue

            weight = WEIGHTS[engine]
            weighted_sum += float(value) * weight
            available_weight += weight
            available_values.append(float(value))

        if not available_values or available_weight == 0:
            continue

        master_score = weighted_sum / available_weight
        completeness = available_weight / sum(WEIGHTS.values()) * 100

        bullish_votes = sum(value >= 60 for value in available_values)
        bearish_votes = sum(value <= 40 for value in available_values)

        agreement = (
            max(bullish_votes, bearish_votes)
            / len(available_values)
            * 100
        )

        spread = max(available_values) - min(available_values)

        reasons = []

        for label, source, key in (
            ("Structure", structure, "Reason"),
            ("Trend", trend, "Reason"),
            ("Relative Strength", relative_strength, "Reason"),
            ("Pivot", pivots, "Reason"),
        ):
            text = source.get(symbol, {}).get(key)

            if text:
                reasons.append(f"{label}: {text}")

        if sector_data:
            reasons.append(
                f"Sector: {sector}, rank {sector_data['Rank']}"
            )
        else:
            reasons.append(
                f"Sector: {sector}, score unavailable"
            )

        bias = directional_bias(master_score)
        rec = recommendation(master_score, completeness)
        grade = confidence_grade(
            master_score,
            completeness,
            agreement,
        )

        results.append(
            {
                "Symbol": symbol,
                "Sector": sector,
                "Structure Score": component_values["Structure"],
                "Trend Score": component_values["Trend"],
                "Relative Strength Score": component_values["Relative Strength"],
                "Sector Score": component_values["Sector"],
                "Pivot Score": component_values["Pivot"],
                "AQSD Master Score": round(master_score, 2),
                "Directional Bias": bias,
                "Recommendation": rec,
                "Confidence Grade": grade,
                "Engine Agreement %": round(agreement, 2),
                "Engine Spread": round(spread, 2),
                "Data Completeness %": round(completeness, 2),
                "Structure Event": structure.get(
                    symbol,
                    {},
                ).get("Structure Event", ""),
                "Trend Regime": trend.get(
                    symbol,
                    {},
                ).get("Trend Regime", ""),
                "RS Classification": relative_strength.get(
                    symbol,
                    {},
                ).get("Classification", ""),
                "Pivot Bias": pivots.get(
                    symbol,
                    {},
                ).get("Pivot Bias", ""),
                "CPR Type": pivots.get(
                    symbol,
                    {},
                ).get("CPR Type", ""),
                "CPR Position": pivots.get(
                    symbol,
                    {},
                ).get("CPR Position", ""),
                "Explanation": " || ".join(reasons),
            }
        )

    return sorted(
        results,
        key=lambda item: (
            item["AQSD Master Score"],
            item["Data Completeness %"],
            item["Engine Agreement %"],
        ),
        reverse=True,
    )


def write_results(wb, results: list[dict]) -> None:
    if "Master Intelligence" in wb.sheetnames:
        del wb["Master Intelligence"]

    ws = wb.create_sheet("Master Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:U2")
    ws["A1"] = "AQSD PROFESSIONAL - MASTER INTELLIGENCE SCORE v2"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Last Updated"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["D4"] = "Stocks Ranked"
    ws["E4"] = len(results)

    for ref in ("A4", "D4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    ws["A5"] = "Weights"
    ws["B5"] = (
        "Structure 25% | Trend 25% | Relative Strength 20% | "
        "Sector 10% | Pivot/CPR 20%"
    )
    ws["A5"].font = Font(bold=True)
    ws["A5"].fill = PatternFill(
        "solid",
        fgColor=BLUE,
    )

    headers = [
        "Rank",
        "Symbol",
        "Sector",
        "Structure Score",
        "Trend Score",
        "Relative Strength Score",
        "Sector Score",
        "Pivot Score",
        "AQSD Master Score",
        "Directional Bias",
        "Recommendation",
        "Confidence Grade",
        "Engine Agreement %",
        "Engine Spread",
        "Data Completeness %",
        "Structure Event",
        "Trend Regime",
        "RS Classification",
        "Pivot Bias",
        "CPR Type",
        "CPR Position",
        "Explanation",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, result in enumerate(results, start=8):
        values = [
            row_no - 7,
            result["Symbol"],
            result["Sector"],
            result["Structure Score"],
            result["Trend Score"],
            result["Relative Strength Score"],
            result["Sector Score"],
            result["Pivot Score"],
            result["AQSD Master Score"],
            result["Directional Bias"],
            result["Recommendation"],
            result["Confidence Grade"],
            result["Engine Agreement %"],
            result["Engine Spread"],
            result["Data Completeness %"],
            result["Structure Event"],
            result["Trend Regime"],
            result["RS Classification"],
            result["Pivot Bias"],
            result["CPR Type"],
            result["CPR Position"],
            result["Explanation"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            cell.border = Border(bottom=THIN)

        for col in range(4, 10):
            ws.cell(
                row_no,
                col,
            ).number_format = "0.00"

        for col in (13, 15):
            ws.cell(
                row_no,
                col,
            ).number_format = '0.00"%"'

        score = result["AQSD Master Score"]

        ws.cell(
            row_no,
            9,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 65
                else RED
                if score <= 35
                else YELLOW
            ),
        )
        ws.cell(row_no, 9).font = Font(bold=True)

        bias = result["Directional Bias"]

        ws.cell(
            row_no,
            10,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BULLISH" in bias
                else RED
                if "BEARISH" in bias
                else GREY
            ),
        )
        ws.cell(row_no, 10).font = Font(bold=True)

        rec = result["Recommendation"]

        ws.cell(
            row_no,
            11,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BUY" in rec or "LONG" in rec
                else RED
                if "SELL" in rec or "SHORT" in rec
                else YELLOW
            ),
        )
        ws.cell(row_no, 11).font = Font(bold=True)

        completeness = result["Data Completeness %"]

        ws.cell(
            row_no,
            15,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if completeness >= 90
                else YELLOW
                if completeness >= 60
                else RED
            ),
        )

    widths = {
        "A": 8,
        "B": 18,
        "C": 18,
        "D": 15,
        "E": 13,
        "F": 20,
        "G": 13,
        "H": 13,
        "I": 17,
        "J": 17,
        "K": 24,
        "L": 14,
        "M": 17,
        "N": 14,
        "O": 18,
        "P": 18,
        "Q": 16,
        "R": 18,
        "S": 18,
        "T": 14,
        "U": 14,
        "V": 80,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    structure = read_score_sheet(
        wb,
        "Market Structure",
        6,
        "Structure Score",
        [
            "Structure Event",
            "Trend Phase",
            "Reason",
        ],
    )

    trend = read_score_sheet(
        wb,
        "Trend Intelligence",
        6,
        "Trend Score",
        [
            "Trend Classification",
            "Trend Strength",
            "Trend Regime",
            "Reason",
        ],
    )

    relative_strength = read_score_sheet(
        wb,
        "Relative Strength",
        6,
        "Relative Strength Score",
        [
            "Classification",
            "Reason",
        ],
    )

    pivots = read_score_sheet(
        wb,
        "Pivot Intelligence",
        6,
        "Pivot Score",
        [
            "Pivot Bias",
            "CPR Type",
            "CPR Position",
            "Reason",
        ],
    )

    sectors = read_sector_rotation(wb)

    results = combine_scores(
        structure,
        trend,
        relative_strength,
        pivots,
        sectors,
    )

    write_results(wb, results)

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("\nAQSD MASTER INTELLIGENCE SCORE v2")
    print("=" * 72)
    print(f"Structure records: {len(structure)}")
    print(f"Trend records: {len(trend)}")
    print(f"Relative-strength records: {len(relative_strength)}")
    print(f"Pivot records: {len(pivots)}")
    print(f"Sector records: {len(sectors)}")
    print(f"Stocks ranked: {len(results)}")

    if results:
        strongest = results[0]
        weakest = results[-1]

        print(
            f"Highest score: {strongest['Symbol']} "
            f"({strongest['AQSD Master Score']:.2f})"
        )

        print(
            f"Lowest score: {weakest['Symbol']} "
            f"({weakest['AQSD Master Score']:.2f})"
        )

    print(DASHBOARD)


if __name__ == "__main__":
    main()
