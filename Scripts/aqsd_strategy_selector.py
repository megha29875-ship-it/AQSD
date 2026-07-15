
"""
AQSD Professional
Module: Dynamic Strategy Selector
Version: 1.0

Purpose
-------
Selects the most suitable trading style for the current AQSD market regime.

Inputs
------
- Market Regime Intelligence
- Market Breadth Intelligence
- Sector Rotation Intelligence
- Unified Master Intelligence
- Trade Decision Engine
- Price Structure Intelligence

Outputs
-------
- Recommended primary strategy
- Secondary strategy
- Strategies to avoid
- Position-size multiplier
- Maximum concurrent positions
- Preferred holding period
- Minimum Master Score
- Minimum Confidence
- Minimum Reward/Risk
- Stop-loss style
- Entry style
- SQLite history
- Excel sheet: Strategy Selector

Commands
--------
python aqsd_strategy_selector.py --run
python aqsd_strategy_selector.py --status
python aqsd_strategy_selector.py --report
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


SCHEMA = """
CREATE TABLE IF NOT EXISTS aqsd_strategy_selector (
    selector_id INTEGER PRIMARY KEY AUTOINCREMENT,
    selection_date TEXT NOT NULL UNIQUE,
    market_regime TEXT NOT NULL,
    primary_strategy TEXT NOT NULL,
    secondary_strategy TEXT,
    avoid_strategies TEXT,
    position_size_multiplier REAL,
    max_concurrent_positions INTEGER,
    preferred_holding_period TEXT,
    minimum_master_score REAL,
    minimum_confidence_percent REAL,
    minimum_reward_risk REAL,
    stop_loss_style TEXT,
    entry_style TEXT,
    exit_style TEXT,
    sector_preference TEXT,
    stock_profile TEXT,
    capital_exposure_percent REAL,
    strategy_confidence REAL,
    explanation TEXT,
    created_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_strategy_selector_date
ON aqsd_strategy_selector(selection_date);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table'
          AND name=?
        """,
        (table_name,),
    ).fetchone()

    return row is not None


def safe_float(value: Any, default: float | None = None) -> float | None:
    try:
        if value is None or value == "":
            return default

        return float(value)

    except (TypeError, ValueError):
        return default


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def latest_regime() -> dict:
    with connect() as connection:
        if not table_exists(
            connection,
            "market_regime_intelligence",
        ):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM market_regime_intelligence
            ORDER BY regime_date DESC, regime_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def latest_breadth() -> dict:
    with connect() as connection:
        if not table_exists(
            connection,
            "market_breadth_intelligence",
        ):
            return {}

        row = connection.execute(
            """
            SELECT *
            FROM market_breadth_intelligence
            WHERE scope_type='MARKET'
              AND trade_date=(
                  SELECT MAX(trade_date)
                  FROM market_breadth_intelligence
              )
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else {}


def latest_sector_summary() -> dict:
    with connect() as connection:
        if not table_exists(
            connection,
            "sector_rotation_intelligence",
        ):
            return {}

        row = connection.execute(
            """
            SELECT
                AVG(sector_rotation_score) AS average_score,
                SUM(
                    CASE
                        WHEN rotation_state IN (
                            'STRONG INFLOW',
                            'INFLOW',
                            'EARLY ROTATION'
                        )
                        THEN 1 ELSE 0
                    END
                ) AS positive_sectors,
                SUM(
                    CASE
                        WHEN rotation_state IN (
                            'STRONG OUTFLOW',
                            'OUTFLOW',
                            'WEAKENING LEADERSHIP'
                        )
                        THEN 1 ELSE 0
                    END
                ) AS negative_sectors,
                COUNT(*) AS sector_count
            FROM sector_rotation_intelligence
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            """
        ).fetchone()

    return dict(row) if row else {}


def latest_decision_summary() -> dict:
    with connect() as connection:
        if not table_exists(
            connection,
            "aqsd_trade_decisions",
        ):
            return {}

        row = connection.execute(
            """
            SELECT
                AVG(master_score) AS average_master_score,
                AVG(confidence_percent) AS average_confidence,
                AVG(completeness_percent) AS average_completeness,
                SUM(
                    CASE
                        WHEN action IN (
                            'STRONG BUY',
                            'BUY',
                            'BUY ON DIP'
                        )
                        THEN 1 ELSE 0
                    END
                ) AS actionable_longs,
                SUM(
                    CASE
                        WHEN action IN (
                            'AVOID',
                            'EXIT / AVOID'
                        )
                        THEN 1 ELSE 0
                    END
                ) AS avoid_count,
                COUNT(*) AS stock_count
            FROM aqsd_trade_decisions
            WHERE trade_date=(
                SELECT MAX(trade_date)
                FROM aqsd_trade_decisions
            )
            """
        ).fetchone()

    return dict(row) if row else {}


def strategy_template(regime: str) -> dict:
    templates = {
        "BULL TREND": {
            "primary_strategy": "TREND FOLLOWING",
            "secondary_strategy": "BUY ON DIPS",
            "avoid_strategies": "AGGRESSIVE MEAN REVERSION|EARLY SHORTING",
            "position_size_multiplier": 1.00,
            "max_concurrent_positions": 8,
            "preferred_holding_period": "5-20 TRADING DAYS",
            "minimum_master_score": 68,
            "minimum_confidence_percent": 60,
            "minimum_reward_risk": 1.8,
            "stop_loss_style": "ATR OR SWING-LOW TRAILING STOP",
            "entry_style": "BREAKOUT OR CONTROLLED PULLBACK",
            "exit_style": "TRAIL PROFITS; PARTIAL EXIT AT TARGET 1",
            "sector_preference": "STRONG INFLOW AND LEADING SECTORS",
            "stock_profile": "HIGH RS, HH-HL, BULLISH BOS",
        },
        "BEAR TREND": {
            "primary_strategy": "CAPITAL PROTECTION",
            "secondary_strategy": "SHORT BIAS / CASH",
            "avoid_strategies": "BOTTOM FISHING|AVERAGING DOWN|WEAK BREAKOUTS",
            "position_size_multiplier": 0.30,
            "max_concurrent_positions": 2,
            "preferred_holding_period": "1-5 TRADING DAYS",
            "minimum_master_score": 82,
            "minimum_confidence_percent": 75,
            "minimum_reward_risk": 2.2,
            "stop_loss_style": "TIGHT ATR STOP",
            "entry_style": "ONLY EXCEPTIONAL RELATIVE-STRENGTH SETUPS",
            "exit_style": "QUICK PROFIT BOOKING",
            "sector_preference": "DEFENSIVE OR RELATIVE-STRENGTH LEADERS",
            "stock_profile": "VERY STRONG RS WITH LOW RISK",
        },
        "RANGE BOUND": {
            "primary_strategy": "MEAN REVERSION",
            "secondary_strategy": "SUPPORT-RESISTANCE SWING",
            "avoid_strategies": "LATE BREAKOUT CHASING|WIDE STOPS",
            "position_size_multiplier": 0.60,
            "max_concurrent_positions": 5,
            "preferred_holding_period": "2-8 TRADING DAYS",
            "minimum_master_score": 62,
            "minimum_confidence_percent": 55,
            "minimum_reward_risk": 1.5,
            "stop_loss_style": "STRUCTURE-BASED TIGHT STOP",
            "entry_style": "BUY NEAR SUPPORT; SELL NEAR RESISTANCE",
            "exit_style": "BOOK NEAR OPPOSITE RANGE BOUNDARY",
            "sector_preference": "SELECTIVE SECTORS WITH STABLE BREADTH",
            "stock_profile": "LIQUID STOCKS WITH CLEAR SUPPORT/RESISTANCE",
        },
        "HIGH VOLATILITY": {
            "primary_strategy": "CAPITAL PRESERVATION",
            "secondary_strategy": "SELECTIVE EVENT-DRIVEN SWING",
            "avoid_strategies": "LEVERAGE|OVERSIZED POSITIONS|TIGHT CHASING",
            "position_size_multiplier": 0.35,
            "max_concurrent_positions": 3,
            "preferred_holding_period": "1-5 TRADING DAYS",
            "minimum_master_score": 78,
            "minimum_confidence_percent": 70,
            "minimum_reward_risk": 2.0,
            "stop_loss_style": "WIDER ATR STOP WITH SMALLER SIZE",
            "entry_style": "WAIT FOR CONFIRMATION",
            "exit_style": "PARTIAL PROFITS AND ACTIVE TRAILING",
            "sector_preference": "ONLY STRONGEST SECTORS",
            "stock_profile": "HIGH LIQUIDITY, HIGH CONFIDENCE, LOW GAP RISK",
        },
        "ACCUMULATION": {
            "primary_strategy": "EARLY POSITION BUILDING",
            "secondary_strategy": "BUY ON CONFIRMED BREAKOUT",
            "avoid_strategies": "FULL-SIZE EARLY ENTRY|WEAK SECTORS",
            "position_size_multiplier": 0.70,
            "max_concurrent_positions": 6,
            "preferred_holding_period": "5-15 TRADING DAYS",
            "minimum_master_score": 64,
            "minimum_confidence_percent": 58,
            "minimum_reward_risk": 1.8,
            "stop_loss_style": "SWING-LOW STOP",
            "entry_style": "STAGGERED ENTRY",
            "exit_style": "HOLD LEADERS; EXIT FAILED BREAKOUTS",
            "sector_preference": "EARLY ROTATION SECTORS",
            "stock_profile": "IMPROVING RS, BULLISH CHOCH, RISING BREADTH",
        },
        "DISTRIBUTION": {
            "primary_strategy": "REDUCE EXPOSURE",
            "secondary_strategy": "SELL INTO STRENGTH",
            "avoid_strategies": "NEW AGGRESSIVE LONGS|AVERAGING DOWN",
            "position_size_multiplier": 0.40,
            "max_concurrent_positions": 3,
            "preferred_holding_period": "1-5 TRADING DAYS",
            "minimum_master_score": 76,
            "minimum_confidence_percent": 68,
            "minimum_reward_risk": 2.0,
            "stop_loss_style": "TIGHT TRAILING STOP",
            "entry_style": "ONLY HIGH-CONVICTION LEADERS",
            "exit_style": "REDUCE ON WEAKNESS",
            "sector_preference": "DEFENSIVE OR FRESH INFLOW SECTORS",
            "stock_profile": "LOW RISK, HIGH CONFIDENCE, STRONG RS",
        },
        "RISK ON": {
            "primary_strategy": "SECTOR LEADERSHIP",
            "secondary_strategy": "MOMENTUM SWING",
            "avoid_strategies": "LAGGARDS|LOW-LIQUIDITY STOCKS",
            "position_size_multiplier": 0.90,
            "max_concurrent_positions": 7,
            "preferred_holding_period": "4-15 TRADING DAYS",
            "minimum_master_score": 66,
            "minimum_confidence_percent": 58,
            "minimum_reward_risk": 1.7,
            "stop_loss_style": "ATR TRAILING STOP",
            "entry_style": "BREAKOUT OR PULLBACK IN LEADING SECTOR",
            "exit_style": "TRAIL WINNERS",
            "sector_preference": "STRONG INFLOW SECTORS",
            "stock_profile": "SECTOR LEADERS WITH OUTPERFORMANCE",
        },
        "RISK OFF": {
            "primary_strategy": "DEFENSIVE / CASH HEAVY",
            "secondary_strategy": "SELECTIVE LOW-BETA SWING",
            "avoid_strategies": "LEVERAGE|CYCLICAL LAGGARDS|WEAK BREAKOUTS",
            "position_size_multiplier": 0.25,
            "max_concurrent_positions": 2,
            "preferred_holding_period": "1-5 TRADING DAYS",
            "minimum_master_score": 84,
            "minimum_confidence_percent": 78,
            "minimum_reward_risk": 2.2,
            "stop_loss_style": "TIGHT RISK-CONTROL STOP",
            "entry_style": "ONLY EXCEPTIONAL SETUPS",
            "exit_style": "FAST PROFIT PROTECTION",
            "sector_preference": "DEFENSIVES AND RELATIVE-STRENGTH LEADERS",
            "stock_profile": "LOW VOLATILITY, HIGH CONFIDENCE",
        },
        "TRANSITION": {
            "primary_strategy": "WAIT FOR CONFIRMATION",
            "secondary_strategy": "SELECTIVE SWING",
            "avoid_strategies": "AGGRESSIVE POSITIONING|HIGH CONCENTRATION",
            "position_size_multiplier": 0.45,
            "max_concurrent_positions": 4,
            "preferred_holding_period": "2-7 TRADING DAYS",
            "minimum_master_score": 70,
            "minimum_confidence_percent": 65,
            "minimum_reward_risk": 1.9,
            "stop_loss_style": "STRUCTURE-BASED STOP",
            "entry_style": "CONFIRMED BOS OR CHOCH ONLY",
            "exit_style": "QUICKLY EXIT FAILED CONFIRMATION",
            "sector_preference": "ONLY CLEAR LEADERS",
            "stock_profile": "HIGH DATA COMPLETENESS AND AGREEMENT",
        },
    }

    return templates.get(
        regime,
        templates["TRANSITION"],
    ).copy()


def calculate_selection() -> dict:
    regime = latest_regime()
    breadth = latest_breadth()
    sectors = latest_sector_summary()
    decisions = latest_decision_summary()

    if not regime:
        raise RuntimeError(
            "No Market Regime data found. "
            "Run aqsd_market_regime.py --run first."
        )

    regime_name = str(
        regime.get("market_regime") or "TRANSITION"
    )

    result = strategy_template(regime_name)

    breadth_score = safe_float(
        breadth.get("breadth_score"),
        50,
    ) or 50

    average_confidence = safe_float(
        decisions.get("average_confidence"),
        50,
    ) or 50

    average_completeness = safe_float(
        decisions.get("average_completeness"),
        50,
    ) or 50

    actionable_longs = int(
        decisions.get("actionable_longs") or 0
    )
    stock_count = int(
        decisions.get("stock_count") or 0
    )

    positive_sectors = int(
        sectors.get("positive_sectors") or 0
    )
    sector_count = int(
        sectors.get("sector_count") or 0
    )

    actionable_percent = (
        actionable_longs / stock_count * 100
        if stock_count
        else 0
    )

    positive_sector_percent = (
        positive_sectors / sector_count * 100
        if sector_count
        else 0
    )

    multiplier = float(
        result["position_size_multiplier"]
    )

    if breadth_score < 40:
        multiplier *= 0.80

    if average_confidence < 55:
        multiplier *= 0.85

    if average_completeness < 55:
        multiplier *= 0.85

    if actionable_percent < 10:
        multiplier *= 0.80

    if positive_sector_percent >= 60:
        multiplier *= 1.10

    multiplier = round(
        clamp(multiplier, 0.15, 1.00),
        2,
    )

    base_exposure = safe_float(
        regime.get("capital_exposure_percent"),
        50,
    ) or 50

    capital_exposure = round(
        clamp(
            base_exposure * multiplier,
            10,
            90,
        ),
        2,
    )

    confidence = (
        safe_float(
            regime.get("regime_score"),
            50,
        )
        * 0.40
        + breadth_score * 0.20
        + average_confidence * 0.20
        + average_completeness * 0.10
        + positive_sector_percent * 0.10
    )

    confidence = round(
        clamp(confidence, 0, 100),
        2,
    )

    explanation = " | ".join(
        [
            f"Regime {regime_name}",
            f"Breadth {breadth_score:.1f}",
            f"Average confidence {average_confidence:.1f}%",
            f"Average completeness {average_completeness:.1f}%",
            f"Actionable longs {actionable_percent:.1f}%",
            f"Positive sectors {positive_sector_percent:.1f}%",
            f"Position multiplier {multiplier:.2f}",
        ]
    )

    return {
        "selection_date": str(
            regime.get("regime_date")
            or datetime.now().date().isoformat()
        ),
        "market_regime": regime_name,
        **result,
        "position_size_multiplier": multiplier,
        "capital_exposure_percent": capital_exposure,
        "strategy_confidence": confidence,
        "explanation": explanation,
    }


def save_result(result: dict) -> None:
    with connect() as connection:
        connection.execute(
            """
            INSERT INTO aqsd_strategy_selector(
                selection_date,
                market_regime,
                primary_strategy,
                secondary_strategy,
                avoid_strategies,
                position_size_multiplier,
                max_concurrent_positions,
                preferred_holding_period,
                minimum_master_score,
                minimum_confidence_percent,
                minimum_reward_risk,
                stop_loss_style,
                entry_style,
                exit_style,
                sector_preference,
                stock_profile,
                capital_exposure_percent,
                strategy_confidence,
                explanation,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(selection_date)
            DO UPDATE SET
                market_regime=excluded.market_regime,
                primary_strategy=excluded.primary_strategy,
                secondary_strategy=excluded.secondary_strategy,
                avoid_strategies=excluded.avoid_strategies,
                position_size_multiplier=excluded.position_size_multiplier,
                max_concurrent_positions=excluded.max_concurrent_positions,
                preferred_holding_period=excluded.preferred_holding_period,
                minimum_master_score=excluded.minimum_master_score,
                minimum_confidence_percent=excluded.minimum_confidence_percent,
                minimum_reward_risk=excluded.minimum_reward_risk,
                stop_loss_style=excluded.stop_loss_style,
                entry_style=excluded.entry_style,
                exit_style=excluded.exit_style,
                sector_preference=excluded.sector_preference,
                stock_profile=excluded.stock_profile,
                capital_exposure_percent=excluded.capital_exposure_percent,
                strategy_confidence=excluded.strategy_confidence,
                explanation=excluded.explanation,
                created_at=excluded.created_at
            """,
            (
                result["selection_date"],
                result["market_regime"],
                result["primary_strategy"],
                result["secondary_strategy"],
                result["avoid_strategies"],
                result["position_size_multiplier"],
                result["max_concurrent_positions"],
                result["preferred_holding_period"],
                result["minimum_master_score"],
                result["minimum_confidence_percent"],
                result["minimum_reward_risk"],
                result["stop_loss_style"],
                result["entry_style"],
                result["exit_style"],
                result["sector_preference"],
                result["stock_profile"],
                result["capital_exposure_percent"],
                result["strategy_confidence"],
                result["explanation"],
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )

        connection.commit()


def run_engine() -> dict:
    setup_schema()

    run_id = start_run(
        "aqsd_strategy_selector",
        "Selecting dynamic AQSD strategy",
    )

    try:
        result = calculate_selection()
        save_result(result)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=1,
            errors_count=0,
            message=(
                f"Primary={result['primary_strategy']}; "
                f"confidence={result['strategy_confidence']}"
            ),
        )

        return result

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


def latest_result() -> dict | None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT *
            FROM aqsd_strategy_selector
            ORDER BY selection_date DESC, selector_id DESC
            LIMIT 1
            """
        ).fetchone()

    return dict(row) if row else None


def write_report(
    result: dict | None = None,
) -> None:
    if result is None:
        result = latest_result()

    if not result:
        raise RuntimeError(
            "No Strategy Selector result found. Run --run first."
        )

    if DASHBOARD.exists():
        workbook = load_workbook(DASHBOARD)
    else:
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    sheet_name = "Strategy Selector"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(
        sheet_name,
        1,
    )
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H2")
    ws["A1"] = "AQSD PROFESSIONAL - DYNAMIC STRATEGY SELECTOR"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Market Regime"
    ws["B4"] = result["market_regime"]
    ws["D4"] = "Primary Strategy"
    ws["E4"] = result["primary_strategy"]
    ws["G4"] = "Confidence"
    ws["H4"] = result["strategy_confidence"]

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    rows = [
        ("Secondary Strategy", result["secondary_strategy"]),
        ("Avoid Strategies", result["avoid_strategies"]),
        (
            "Position Size Multiplier",
            result["position_size_multiplier"],
        ),
        (
            "Maximum Concurrent Positions",
            result["max_concurrent_positions"],
        ),
        (
            "Preferred Holding Period",
            result["preferred_holding_period"],
        ),
        (
            "Minimum Master Score",
            result["minimum_master_score"],
        ),
        (
            "Minimum Confidence %",
            result["minimum_confidence_percent"],
        ),
        (
            "Minimum Reward/Risk",
            result["minimum_reward_risk"],
        ),
        ("Stop-Loss Style", result["stop_loss_style"]),
        ("Entry Style", result["entry_style"]),
        ("Exit Style", result["exit_style"]),
        ("Sector Preference", result["sector_preference"]),
        ("Preferred Stock Profile", result["stock_profile"]),
        (
            "Suggested Capital Exposure %",
            result["capital_exposure_percent"],
        ),
        ("Explanation", result["explanation"]),
    ]

    for row_no, (label, value) in enumerate(
        rows,
        start=7,
    ):
        ws.cell(row_no, 1, label)
        ws.cell(row_no, 2, value)

        ws.cell(row_no, 1).font = Font(
            bold=True
        )
        ws.cell(row_no, 1).fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )
        ws.cell(row_no, 1).border = Border(
            bottom=THIN
        )
        ws.cell(row_no, 2).border = Border(
            bottom=THIN
        )

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    workbook.save(DASHBOARD)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                MIN(selection_date) AS first_date,
                MAX(selection_date) AS latest_date
            FROM aqsd_strategy_selector
            """
        ).fetchone()

    print("\nAQSD STRATEGY SELECTOR STATUS")
    print("=" * 72)
    print(f"Stored records: {row['total'] or 0}")
    print(f"First date:     {row['first_date'] or 'No data'}")
    print(f"Latest date:    {row['latest_date'] or 'No data'}")
    print("=" * 72)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Dynamic Strategy Selector."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Select the latest AQSD strategy.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild the Excel report.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show Strategy Selector status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.run:
        result = run_engine()
        write_report(result)

        print("\nAQSD DYNAMIC STRATEGY SELECTOR")
        print("=" * 72)
        print(f"Market Regime:       {result['market_regime']}")
        print(f"Primary Strategy:    {result['primary_strategy']}")
        print(f"Secondary Strategy:  {result['secondary_strategy']}")
        print(f"Position Multiplier: {result['position_size_multiplier']}")
        print(
            f"Maximum Positions:   "
            f"{result['max_concurrent_positions']}"
        )
        print(
            f"Capital Exposure:    "
            f"{result['capital_exposure_percent']}%"
        )
        print(
            f"Strategy Confidence: "
            f"{result['strategy_confidence']}%"
        )
        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report()
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
