
"""
AQSD Professional
Module: Strategy Scorecard
Version: 1.0

Compares strategy signals from CALL and PUT candidates.
"""

from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference

BASE = Path(__file__).resolve().parent.parent
FILE = BASE / "Output" / "Dashboard.xlsx"

wb = load_workbook(FILE)

if "Strategy Scorecard" in wb.sheetnames:
    del wb["Strategy Scorecard"]

ws = wb.create_sheet("Strategy Scorecard")

ws["A1"] = "AQSD PROFESSIONAL - STRATEGY SCORECARD"
ws["A1"].font = Font(size=18,bold=True,color="FFFFFF")
ws["A1"].fill = PatternFill("solid",fgColor="1F4E78")

counter = Counter()

for sheet in ("CALL Candidates","PUT Candidates"):
    if sheet not in wb.sheetnames:
        continue
    s = wb[sheet]
    headers={}
    for c in s[1]:
        if c.value:
            headers[str(c.value)] = c.column

    if "Recommendation" not in headers:
        continue

    for r in range(2, s.max_row+1):
        rec = str(s.cell(r, headers["Recommendation"]).value or "").strip().upper()
        if rec:
            counter[rec] += 1

ws["A3"]="Recommendation"
ws["B3"]="Count"

row=4
for name,count in counter.most_common():
    ws.cell(row,1).value=name
    ws.cell(row,2).value=count
    row+=1

if row>4:
    chart=BarChart()
    data=Reference(ws,min_col=2,min_row=3,max_row=row-1)
    cats=Reference(ws,min_col=1,min_row=4,max_row=row-1)
    chart.add_data(data,titles_from_data=True)
    chart.set_categories(cats)
    chart.title="Recommendation Distribution"
    chart.height=8
    chart.width=14
    ws.add_chart(chart,"D3")

wb.save(FILE)
print("Strategy Scorecard created.")
print(FILE)
