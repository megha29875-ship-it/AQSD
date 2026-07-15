
"""
AQSD Professional
Module: Pivot & CPR Intelligence Engine
Version: 1.0

Builds pivot-based intelligence for NSE F&O symbols.

Features
--------
- Daily Classic Pivots
- Weekly Classic Pivots
- Monthly Classic Pivots
- Central Pivot Range (CPR)
- Narrow / Normal / Wide CPR classification
- Price location versus CPR
- Nearest support and resistance
- Pivot confluence detection
- Pivot Intelligence Score from 0 to 100
- Explainable bullish / bearish bias
- Excel output: Pivot Intelligence

Run
---
python pivot_cpr_engine.py
python pivot_cpr_engine.py --period 1y --limit 100

Notes
-----
--limit 0 means analyse the full available F&O universe.
Yahoo Finance prices may be delayed.
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"
FNO_FILE = BASE / "Data" / "FnO_Stocks.xlsx"


# ============================================================
# SETTINGS
# ============================================================

DEFAULT_PERIOD = "1y"
DEFAULT_LIMIT = 0
CONFLUENCE_THRESHOLD_PERCENT = 0.35


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def normalize_symbol(value) -> str:
    symbol = str(value or "").strip().upper()

    if symbol and not symbol.endswith(".NS"):
        symbol += ".NS"

    return symbol


def load_symbols(limit: int) -> list[str]:
    symbols: list[str] = []

    if FNO_FILE.exists():
        try:
            df = pd.read_excel(FNO_FILE)

            symbol_col = next(
                (
                    column
                    for column in (
                        "Yahoo Symbol",
                        "Symbol",
                        "SYMBOL",
                        "Ticker",
                    )
                    if column in df.columns
                ),
                None,
            )

            if symbol_col:
                for value in df[symbol_col].dropna():
                    symbol = normalize_symbol(value)

                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

                    if limit > 0 and len(symbols) >= limit:
                        return symbols

        except Exception:
            pass

    if DASHBOARD.exists() and (limit == 0 or len(symbols) < limit):
        try:
            wb = load_workbook(
                DASHBOARD,
                read_only=True,
                data_only=True,
            )

            for sheet_name in ("CALL Candidates", "PUT Candidates"):
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                headers = {
                    str(cell.value).strip(): cell.column
                    for cell in ws[1]
                    if cell.value is not None
                }

                symbol_col = headers.get("Symbol")

                if not symbol_col:
                    continue

                for row in range(2, ws.max_row + 1):
                    symbol = normalize_symbol(
                        ws.cell(row, symbol_col).value
                    )

                    if symbol and symbol not in symbols:
                        symbols.append(symbol)

                    if limit > 0 and len(symbols) >= limit:
                        wb.close()
                        return symbols

            wb.close()

        except Exception:
            pass

    fallback = [
        "RELIANCE.NS",
        "HDFCBANK.NS",
        "ICICIBANK.NS",
        "SBIN.NS",
        "INFY.NS",
        "TCS.NS",
        "LT.NS",
        "SUNPHARMA.NS",
        "BIOCON.NS",
        "TATAMOTORS.NS",
    ]

    for symbol in fallback:
        if symbol not in symbols:
            symbols.append(symbol)

        if limit > 0 and len(symbols) >= limit:
            break

    return symbols


def download_ohlc(symbol: str, period: str) -> pd.DataFrame:
    df = yf.download(
        symbol,
        period=period,
        interval="1d",
        auto_adjust=True,
        progress=False,
    )

    if df.empty:
        raise RuntimeError("No OHLC data")

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)

    required = ["Open", "High", "Low", "Close"]

    if any(column not in df.columns for column in required):
        raise RuntimeError("Required OHLC columns missing")

    df = df.dropna(subset=required).copy()

    if len(df) < 40:
        raise RuntimeError("Insufficient history")

    return df


def classic_pivots(
    high: float,
    low: float,
    close: float,
) -> dict[str, float]:
    pivot = (high + low + close) / 3
    r1 = 2 * pivot - low
    s1 = 2 * pivot - high
    r2 = pivot + (high - low)
    s2 = pivot - (high - low)
    r3 = high + 2 * (pivot - low)
    s3 = low - 2 * (high - pivot)

    bc = (high + low) / 2
    tc = 2 * pivot - bc

    cpr_low = min(bc, tc)
    cpr_high = max(bc, tc)

    return {
        "Pivot": pivot,
        "R1": r1,
        "R2": r2,
        "R3": r3,
        "S1": s1,
        "S2": s2,
        "S3": s3,
        "BC": bc,
        "TC": tc,
        "CPR Low": cpr_low,
        "CPR High": cpr_high,
        "CPR Width": cpr_high - cpr_low,
    }


def previous_period_ohlc(
    df: pd.DataFrame,
    period: str,
) -> tuple[float, float, float]:
    working = df.copy()

    if period == "week":
        grouped = working.resample("W-FRI").agg(
            {
                "High": "max",
                "Low": "min",
                "Close": "last",
            }
        )
    elif period == "month":
        grouped = working.resample("ME").agg(
            {
                "High": "max",
                "Low": "min",
                "Close": "last",
            }
        )
    else:
        raise ValueError("Unsupported period")

    grouped = grouped.dropna()

    if len(grouped) < 2:
        raise RuntimeError(f"Insufficient {period} data")

    previous = grouped.iloc[-2]

    return (
        float(previous["High"]),
        float(previous["Low"]),
        float(previous["Close"]),
    )


def pct_distance(
    current: float,
    reference: float | None,
) -> float | None:
    if reference in (None, 0):
        return None

    return round(
        (current - reference) / reference * 100,
        2,
    )


def nearest_level(
    current: float,
    levels: dict[str, float],
    direction: str,
) -> tuple[str, float] | tuple[str, None]:
    if direction == "support":
        candidates = {
            name: value
            for name, value in levels.items()
            if value <= current
        }

        if not candidates:
            return "", None

        name = max(candidates, key=candidates.get)
        return name, candidates[name]

    candidates = {
        name: value
        for name, value in levels.items()
        if value >= current
    }

    if not candidates:
        return "", None

    name = min(candidates, key=candidates.get)
    return name, candidates[name]


def detect_confluence(
    current: float,
    levels: dict[str, float],
) -> str:
    if current == 0:
        return ""

    keys = list(levels)

    for i, first in enumerate(keys):
        for second in keys[i + 1:]:
            first_value = levels[first]
            second_value = levels[second]

            distance = abs(first_value - second_value) / current * 100

            if distance <= CONFLUENCE_THRESHOLD_PERCENT:
                return (
                    f"{first} + {second} "
                    f"({distance:.2f}% apart)"
                )

    return "NONE"


def classify_cpr_width(
    width_percent: float,
) -> str:
    if width_percent <= 0.25:
        return "NARROW CPR"

    if width_percent >= 0.75:
        return "WIDE CPR"

    return "NORMAL CPR"


def analyse_symbol(symbol: str, df: pd.DataFrame) -> dict:
    close = float(df["Close"].iloc[-1])

    previous_day = df.iloc[-2]
    daily = classic_pivots(
        float(previous_day["High"]),
        float(previous_day["Low"]),
        float(previous_day["Close"]),
    )

    weekly_high, weekly_low, weekly_close = previous_period_ohlc(
        df,
        "week",
    )

    monthly_high, monthly_low, monthly_close = previous_period_ohlc(
        df,
        "month",
    )

    weekly = classic_pivots(
        weekly_high,
        weekly_low,
        weekly_close,
    )

    monthly = classic_pivots(
        monthly_high,
        monthly_low,
        monthly_close,
    )

    daily_cpr_width_percent = (
        daily["CPR Width"] / close * 100
        if close
        else 0.0
    )

    cpr_type = classify_cpr_width(
        daily_cpr_width_percent
    )

    if close > daily["CPR High"]:
        cpr_position = "ABOVE CPR"
    elif close < daily["CPR Low"]:
        cpr_position = "BELOW CPR"
    else:
        cpr_position = "INSIDE CPR"

    combined_levels = {
        "Daily Pivot": daily["Pivot"],
        "Daily R1": daily["R1"],
        "Daily R2": daily["R2"],
        "Daily S1": daily["S1"],
        "Daily S2": daily["S2"],
        "Weekly Pivot": weekly["Pivot"],
        "Weekly R1": weekly["R1"],
        "Weekly S1": weekly["S1"],
        "Monthly Pivot": monthly["Pivot"],
        "Monthly R1": monthly["R1"],
        "Monthly S1": monthly["S1"],
    }

    support_name, support_value = nearest_level(
        close,
        combined_levels,
        "support",
    )

    resistance_name, resistance_value = nearest_level(
        close,
        combined_levels,
        "resistance",
    )

    confluence = detect_confluence(
        close,
        combined_levels,
    )

    score = 50
    reasons: list[str] = []

    if close > daily["Pivot"]:
        score += 10
        reasons.append("Above daily pivot")
    else:
        score -= 10
        reasons.append("Below daily pivot")

    if close > weekly["Pivot"]:
        score += 12
        reasons.append("Above weekly pivot")
    else:
        score -= 12
        reasons.append("Below weekly pivot")

    if close > monthly["Pivot"]:
        score += 15
        reasons.append("Above monthly pivot")
    else:
        score -= 15
        reasons.append("Below monthly pivot")

    if cpr_position == "ABOVE CPR":
        score += 8
        reasons.append("Above daily CPR")
    elif cpr_position == "BELOW CPR":
        score -= 8
        reasons.append("Below daily CPR")
    else:
        reasons.append("Inside daily CPR")

    if cpr_type == "NARROW CPR":
        reasons.append("Narrow CPR: breakout potential")
    elif cpr_type == "WIDE CPR":
        reasons.append("Wide CPR: range risk")

    if confluence != "NONE":
        reasons.append("Pivot confluence present")

    score = max(0, min(100, score))

    if score >= 75:
        bias = "STRONG BULLISH"
    elif score >= 60:
        bias = "BULLISH"
    elif score <= 25:
        bias = "STRONG BEARISH"
    elif score <= 40:
        bias = "BEARISH"
    else:
        bias = "NEUTRAL"

    return {
        "Symbol": symbol,
        "Close": round(close, 2),
        "Daily Pivot": round(daily["Pivot"], 2),
        "Daily R1": round(daily["R1"], 2),
        "Daily S1": round(daily["S1"], 2),
        "Weekly Pivot": round(weekly["Pivot"], 2),
        "Weekly R1": round(weekly["R1"], 2),
        "Weekly S1": round(weekly["S1"], 2),
        "Monthly Pivot": round(monthly["Pivot"], 2),
        "Monthly R1": round(monthly["R1"], 2),
        "Monthly S1": round(monthly["S1"], 2),
        "CPR Low": round(daily["CPR Low"], 2),
        "CPR High": round(daily["CPR High"], 2),
        "CPR Width %": round(daily_cpr_width_percent, 2),
        "CPR Type": cpr_type,
        "CPR Position": cpr_position,
        "Nearest Support": support_name,
        "Support Value": (
            round(support_value, 2)
            if support_value is not None
            else None
        ),
        "Support Distance %": pct_distance(
            close,
            support_value,
        ),
        "Nearest Resistance": resistance_name,
        "Resistance Value": (
            round(resistance_value, 2)
            if resistance_value is not None
            else None
        ),
        "Resistance Distance %": pct_distance(
            resistance_value,
            close,
        ) if resistance_value is not None else None,
        "Confluence": confluence,
        "Pivot Bias": bias,
        "Pivot Score": score,
        "Reason": " | ".join(reasons),
    }


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_results(results: list[dict]) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Pivot Intelligence" in wb.sheetnames:
        del wb["Pivot Intelligence"]

    ws = wb.create_sheet("Pivot Intelligence", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:Z2")
    ws["A1"] = "AQSD PROFESSIONAL - PIVOT & CPR INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Last Updated"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["D4"] = "Stocks Analysed"
    ws["E4"] = len(results)

    for ref in ("A4", "D4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Rank",
        "Symbol",
        "Close",
        "Daily Pivot",
        "Daily R1",
        "Daily S1",
        "Weekly Pivot",
        "Weekly R1",
        "Weekly S1",
        "Monthly Pivot",
        "Monthly R1",
        "Monthly S1",
        "CPR Low",
        "CPR High",
        "CPR Width %",
        "CPR Type",
        "CPR Position",
        "Nearest Support",
        "Support Value",
        "Support Distance %",
        "Nearest Resistance",
        "Resistance Value",
        "Resistance Distance %",
        "Confluence",
        "Pivot Bias",
        "Pivot Score",
        "Reason",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(6, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    ranked = sorted(
        results,
        key=lambda item: item["Pivot Score"],
        reverse=True,
    )

    for row_no, result in enumerate(ranked, start=7):
        values = [
            row_no - 6,
            result["Symbol"],
            result["Close"],
            result["Daily Pivot"],
            result["Daily R1"],
            result["Daily S1"],
            result["Weekly Pivot"],
            result["Weekly R1"],
            result["Weekly S1"],
            result["Monthly Pivot"],
            result["Monthly R1"],
            result["Monthly S1"],
            result["CPR Low"],
            result["CPR High"],
            result["CPR Width %"],
            result["CPR Type"],
            result["CPR Position"],
            result["Nearest Support"],
            result["Support Value"],
            result["Support Distance %"],
            result["Nearest Resistance"],
            result["Resistance Value"],
            result["Resistance Distance %"],
            result["Confluence"],
            result["Pivot Bias"],
            result["Pivot Score"],
            result["Reason"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            cell.border = Border(bottom=THIN)

        for col in (
            3, 4, 5, 6, 7, 8, 9,
            10, 11, 12, 13, 14, 19, 22,
        ):
            ws.cell(
                row_no,
                col,
            ).number_format = '₹#,##0.00'

        for col in (15, 20, 23):
            ws.cell(
                row_no,
                col,
            ).number_format = '0.00"%"'

        cpr_type = result["CPR Type"]

        ws.cell(
            row_no,
            16,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if cpr_type == "NARROW CPR"
                else RED
                if cpr_type == "WIDE CPR"
                else YELLOW
            ),
        )

        cpr_position = result["CPR Position"]

        ws.cell(
            row_no,
            17,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if cpr_position == "ABOVE CPR"
                else RED
                if cpr_position == "BELOW CPR"
                else GREY
            ),
        )

        bias = result["Pivot Bias"]

        ws.cell(
            row_no,
            25,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BULLISH" in bias
                else RED
                if "BEARISH" in bias
                else YELLOW
            ),
        )
        ws.cell(row_no, 25).font = Font(bold=True)

        score = result["Pivot Score"]

        ws.cell(
            row_no,
            26,
        ).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 70
                else RED
                if score <= 30
                else YELLOW
            ),
        )
        ws.cell(row_no, 26).font = Font(bold=True)

    widths = {
        "A": 8,
        "B": 18,
        "C": 12,
        "D": 13,
        "E": 12,
        "F": 12,
        "G": 14,
        "H": 12,
        "I": 12,
        "J": 14,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 14,
        "Q": 14,
        "R": 20,
        "S": 13,
        "T": 14,
        "U": 20,
        "V": 14,
        "W": 16,
        "X": 28,
        "Y": 18,
        "Z": 12,
        "AA": 70,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.auto_filter.ref = ws.dimensions

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build AQSD pivot and CPR intelligence."
    )

    parser.add_argument(
        "--period",
        default=DEFAULT_PERIOD,
        help="Yahoo Finance period, normally 1y or 2y.",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=DEFAULT_LIMIT,
        help=(
            "Maximum number of symbols. "
            "Use 0 for the full available F&O universe."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    symbols = load_symbols(args.limit)

    print("\nAQSD PIVOT & CPR INTELLIGENCE")
    print("=" * 72)
    print(f"Symbols requested: {len(symbols)}")
    print(f"Period: {args.period}")

    results: list[dict] = []

    for index, symbol in enumerate(symbols, start=1):
        print(f"[{index}/{len(symbols)}] {symbol}")

        try:
            df = download_ohlc(
                symbol,
                args.period,
            )

            results.append(
                analyse_symbol(
                    symbol,
                    df,
                )
            )

        except Exception as error:
            print(f"  Skipped: {error}")

    write_results(results)

    print("=" * 72)
    print(f"Stocks completed: {len(results)}")

    if results:
        strongest = max(
            results,
            key=lambda item: item["Pivot Score"],
        )

        weakest = min(
            results,
            key=lambda item: item["Pivot Score"],
        )

        print(
            f"Strongest pivot bias: {strongest['Symbol']} "
            f"({strongest['Pivot Score']})"
        )

        print(
            f"Weakest pivot bias: {weakest['Symbol']} "
            f"({weakest['Pivot Score']})"
        )

    print(DASHBOARD)


if __name__ == "__main__":
    main()
