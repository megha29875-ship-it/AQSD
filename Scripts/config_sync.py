
"""
AQSD Professional
Module: Configuration Sync
Version: 1.0

Synchronises AQSD_Config.json with an editable Settings sheet
inside Dashboard.xlsx.

Commands
--------
python config_sync.py
    Push JSON settings into the Excel Settings sheet.

python config_sync.py --pull
    Read edited values from the Settings sheet and save them
    back into AQSD_Config.json.

python config_sync.py --show
    Display current settings in Command Prompt.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
CONFIG_FILE = BASE / "Config" / "AQSD_Config.json"
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# STYLING
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
GREY = "E7E6E6"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# HELPERS
# ============================================================

def flatten_dict(
    data: dict[str, Any],
    prefix: str = "",
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []

    for key, value in data.items():
        full_key = f"{prefix}.{key}" if prefix else key

        if isinstance(value, dict):
            rows.extend(flatten_dict(value, full_key))
        else:
            rows.append((full_key, value))

    return rows


def set_nested_value(
    data: dict[str, Any],
    dotted_key: str,
    value: Any,
) -> None:
    parts = dotted_key.split(".")
    current = data

    for part in parts[:-1]:
        if part not in current or not isinstance(current[part], dict):
            current[part] = {}

        current = current[part]

    current[parts[-1]] = value


def convert_value(value: Any) -> Any:
    if isinstance(value, (bool, int, float)) or value is None:
        return value

    text = str(value).strip()

    if text.lower() in {"true", "yes", "on"}:
        return True

    if text.lower() in {"false", "no", "off"}:
        return False

    try:
        return int(text)
    except ValueError:
        pass

    try:
        return float(text)
    except ValueError:
        return text


def load_config() -> dict[str, Any]:
    if not CONFIG_FILE.exists():
        raise FileNotFoundError(
            f"Configuration file not found:\n{CONFIG_FILE}\n"
            "Run: python config_manager.py"
        )

    try:
        return json.loads(
            CONFIG_FILE.read_text(encoding="utf-8")
        )

    except json.JSONDecodeError as error:
        raise RuntimeError(
            f"Invalid configuration JSON:\n{error}"
        ) from error


def save_config(config: dict[str, Any]) -> None:
    CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)

    CONFIG_FILE.write_text(
        json.dumps(
            config,
            indent=4,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def load_or_create_workbook():
    if DASHBOARD.exists():
        return load_workbook(DASHBOARD)

    DASHBOARD.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    return wb


# ============================================================
# JSON -> EXCEL
# ============================================================

def push_to_excel() -> None:
    config = load_config()
    wb = load_or_create_workbook()

    if "Settings" in wb.sheetnames:
        del wb["Settings"]

    ws = wb.create_sheet("Settings", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws.merge_cells("A1:D2")
    ws["A1"] = "AQSD PROFESSIONAL - CONFIGURATION"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    headers = [
        "Setting",
        "Value",
        "Type",
        "Instructions",
    ]

    for column, heading in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=column, value=heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    descriptions = {
        "trading.capital": "Total trading capital.",
        "trading.risk_percent": "Maximum risk per trade in percent.",
        "trading.maximum_open_positions": "Maximum simultaneous open positions.",
        "trading.maximum_position_percent": "Maximum capital in one position.",
        "trading.maximum_portfolio_risk_percent": "Maximum total portfolio risk.",
        "scanner.minimum_call_score": "Minimum score for CALL candidates.",
        "scanner.minimum_put_score": "Minimum score for PUT candidates.",
        "scanner.minimum_confidence": "Minimum trade confidence.",
        "scanner.top_call_candidates": "CALL candidates shown in watchlist.",
        "scanner.top_put_candidates": "PUT candidates shown in watchlist.",
        "risk_management.trailing_stop_enabled": "TRUE or FALSE.",
        "risk_management.trailing_trigger_percent": "Profit required before trailing starts.",
        "risk_management.trailing_distance_percent": "Trailing stop distance.",
        "backup.keep_latest": "Number of recent backups retained.",
        "workflow.open_dashboard_after_run": "TRUE or FALSE.",
        "workflow.update_fno_daily": "TRUE or FALSE.",
    }

    for row_no, (key, value) in enumerate(
        flatten_dict(config),
        start=5,
    ):
        ws.cell(row_no, 1, key)
        ws.cell(row_no, 2, value)
        ws.cell(row_no, 3, type(value).__name__)
        ws.cell(
            row_no,
            4,
            descriptions.get(key, ""),
        )

        ws.cell(row_no, 1).fill = PatternFill(
            fill_type="solid",
            fgColor=BLUE,
        )
        ws.cell(row_no, 1).font = Font(bold=True)

        ws.cell(row_no, 2).fill = PatternFill(
            fill_type="solid",
            fgColor=YELLOW,
        )

        for column in range(1, 5):
            ws.cell(row_no, column).border = Border(bottom=THIN)

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55

    ws.auto_filter.ref = f"A4:D{ws.max_row}"

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Configuration pushed to Settings sheet.")
    print(DASHBOARD)


# ============================================================
# EXCEL -> JSON
# ============================================================

def pull_from_excel() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(
        DASHBOARD,
        data_only=False,
    )

    if "Settings" not in wb.sheetnames:
        raise RuntimeError(
            "Settings sheet not found.\n"
            "Run: python config_sync.py"
        )

    ws = wb["Settings"]
    config: dict[str, Any] = {}

    changed = 0

    for row in range(5, ws.max_row + 1):
        key = str(ws.cell(row, 1).value or "").strip()

        if not key:
            continue

        raw_value = ws.cell(row, 2).value
        value = convert_value(raw_value)

        set_nested_value(
            config,
            key,
            value,
        )

        ws.cell(row, 2).fill = PatternFill(
            fill_type="solid",
            fgColor=GREEN,
        )

        changed += 1

    save_config(config)

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Settings pulled from Excel into JSON.")
    print(f"Settings saved: {changed}")
    print(CONFIG_FILE)


# ============================================================
# CLI
# ============================================================

def show_config() -> None:
    config = load_config()

    print("\nAQSD CONFIGURATION")
    print("=" * 70)
    print(
        json.dumps(
            config,
            indent=4,
            ensure_ascii=False,
        )
    )
    print("=" * 70)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Synchronise AQSD JSON and Excel settings."
    )

    parser.add_argument(
        "--pull",
        action="store_true",
        help="Save edited Excel Settings values into JSON.",
    )

    parser.add_argument(
        "--show",
        action="store_true",
        help="Display current JSON configuration.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if args.show:
        show_config()
        return

    if args.pull:
        pull_from_excel()
        return

    push_to_excel()


if __name__ == "__main__":
    main()
