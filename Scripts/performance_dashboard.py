
"""
AQSD Professional
Performance Dashboard v1.0
Creates a Performance Dashboard sheet.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

def num(v):
    try:
        return float(v)
    except:
        return 0.0

wb = load_workbook(DASHBOARD)

if "Portfolio" not in wb.sheetnames:
    raise RuntimeError("Portfolio sheet not found")

pws = wb["Portfolio"]

if "Performance Dashboard" in wb.sheetnames:
    del wb["Performance Dashboard"]

ws = wb.create_sheet("Performance Dashboard")

ws.merge_cells("A1:H2")
c = ws["A1"]
c.value = "AQSD PROFESSIONAL - PERFORMANCE DASHBOARD"
c.font = Font(size=18,bold=True,color="FFFFFF")
c.fill = PatternFill("solid", fgColor="17365D")
c.alignment = Alignment(horizontal="center")

headers={}
for cell in pws[12]:
    if cell.value:
        headers[str(cell.value)] = cell.column

open_pos=0
mtm=0
capital=0
risk=0
winner=0
loser=0

for r in range(13,pws.max_row+1):
    if not pws.cell(r,headers["Symbol"]).value:
        continue
    if str(pws.cell(r,headers["Status"]).value).upper()!="OPEN":
        continue
    open_pos+=1
    pl=num(pws.cell(r,headers["P/L"]).value)
    mtm+=pl
    capital+=num(pws.cell(r,headers["Capital Used"]).value)
    risk+=num(pws.cell(r,headers["Risk Amount"]).value)
    if pl>=0:
        winner+=1
    else:
        loser+=1

items=[
("Open Positions",open_pos),
("Capital Used",capital),
("Total Risk",risk),
("Open MTM",mtm),
("Winning Positions",winner),
("Losing Positions",loser),
]

row=4
for k,v in items:
    ws[f"A{row}"]=k
    ws[f"B{row}"]=v
    ws[f"A{row}"].font=Font(bold=True)
    row+=1

for col in "AB":
    ws.column_dimensions[col].width=24

wb.save(DASHBOARD)

print("Performance Dashboard created.")
print(DASHBOARD)
