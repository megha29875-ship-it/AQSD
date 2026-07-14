
"""
AQSD Professional
Module: Daily Trading Report
Version: 1.0

Creates a one-page Daily Report sheet using:
- Market Pulse
- CALL Candidates
- PUT Candidates
- Portfolio
- Alerts
- Performance Analytics
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


def header_map(ws, row_number: int = 1) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def safe_value(value, default=""):
    return default if value is None else value


def get_market_pulse(wb) -> dict:
    result = {
        "Market Bias": "UNKNOWN",
        "Confidence": "",
        "Strategy": "",
    }

    if "Market Pulse" not in wb.sheetnames:
        return result

    ws = wb["Market Pulse"]

    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()

        if label in result:
            result[label] = safe_value(ws.cell(row, 2).value)

    return result


def get_top_candidates(
    wb,
    sheet_name: str,
    limit: int = 5,
) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    headers = header_map(ws, 1)

    required = [
        "Symbol",
        "Trade Score",
        "Trade Confidence",
        "Trade Grade",
        "Recommendation",
    ]

    if any(name not in headers for name in required):
        return []

    rows = []

    for row in range(2, min(ws.max_row, limit + 1) + 1):
        rows.append(
            {
                "Symbol": safe_value(
                    ws.cell(row, headers["Symbol"]).value
                ),
                "Score": safe_value(
                    ws.cell(row, headers["Trade Score"]).value
                ),
                "Confidence": safe_value(
                    ws.cell(row, headers["Trade Confidence"]).value
                ),
                "Grade": safe_value(
                    ws.cell(row, headers["Trade Grade"]).value
                ),
                "Recommendation": safe_value(
                    ws.cell(row, headers["Recommendation"]).value
                ),
            }
        )

    return rows


def get_portfolio_summary(wb) -> dict:
    result = {
        "Trading Capital": 0,
        "Capital Invested": 0,
        "Available Cash": 0,
        "Open Positions": 0,
        "Total P/L": 0,
        "Portfolio Return %": 0,
    }

    if "Portfolio" not in wb.sheetnames:
        return result

    ws = wb["Portfolio"]

    mapping = {
        "Trading Capital": "B4",
        "Capital Invested": "B5",
        "Available Cash": "B6",
        "Open Positions": "B7",
        "Total P/L": "B8",
        "Portfolio Return %": "B9",
    }

    for label, cell_ref in mapping.items():
        result[label] = safe_value(ws[cell_ref].value, 0)

    return result


def get_alert_summary(wb) -> dict:
    result = {
        "High": 0,
        "Medium": 0,
        "Low": 0,
        "Total": 0,
    }

    if "Alerts" not in wb.sheetnames:
        return result

    ws = wb["Alerts"]

    for row in range(7, ws.max_row + 1):
        priority = str(ws.cell(row, 1).value or "").strip().upper()

        if priority == "HIGH":
            result["High"] += 1
        elif priority == "MEDIUM":
            result["Medium"] += 1
        elif priority == "LOW":
            result["Low"] += 1

    result["Total"] = (
        result["High"]
        + result["Medium"]
        + result["Low"]
    )

    return result


def create_daily_report(wb) -> None:
    if "Daily Report" in wb.sheetnames:
        del wb["Daily Report"]

    ws = wb.create_sheet("Daily Report", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L2")
    ws["A1"] = "AQSD PROFESSIONAL - DAILY TRADING REPORT"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Report Time"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["A4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE)

    pulse = get_market_pulse(wb)
    portfolio = get_portfolio_summary(wb)
    alerts = get_alert_summary(wb)
    calls = get_top_candidates(wb, "CALL Candidates", 5)
    puts = get_top_candidates(wb, "PUT Candidates", 5)

    # Market section
    ws.merge_cells("A6:C6")
    ws["A6"] = "MARKET OUTLOOK"
    ws["A6"].font = Font(bold=True, color=WHITE)
    ws["A6"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A6"].alignment = Alignment(horizontal="center")

    market_rows = [
        ("Market Bias", pulse["Market Bias"]),
        ("Confidence", pulse["Confidence"]),
        ("Strategy", pulse["Strategy"]),
    ]

    for row_no, (label, value) in enumerate(market_rows, start=7):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)

    bias = str(pulse["Market Bias"]).upper()

    ws["B7"].fill = PatternFill(
        "solid",
        fgColor=(
            GREEN
            if bias == "BULLISH"
            else RED
            if bias == "BEARISH"
            else YELLOW
        ),
    )
    ws["B7"].font = Font(bold=True)

    # Portfolio section
    ws.merge_cells("A12:C12")
    ws["A12"] = "PORTFOLIO SNAPSHOT"
    ws["A12"].font = Font(bold=True, color=WHITE)
    ws["A12"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A12"].alignment = Alignment(horizontal="center")

    portfolio_rows = [
        ("Trading Capital", portfolio["Trading Capital"]),
        ("Capital Invested", portfolio["Capital Invested"]),
        ("Available Cash", portfolio["Available Cash"]),
        ("Open Positions", portfolio["Open Positions"]),
        ("Total P/L", portfolio["Total P/L"]),
        ("Portfolio Return %", portfolio["Portfolio Return %"]),
    ]

    for row_no, (label, value) in enumerate(
        portfolio_rows,
        start=13,
    ):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)

    for row_no in [13, 14, 15, 17]:
        ws[f"B{row_no}"].number_format = '₹#,##0.00'

    ws["B18"].number_format = '0.00"%"'

    # Alerts section
    ws.merge_cells("A21:C21")
    ws["A21"] = "ALERT SUMMARY"
    ws["A21"].font = Font(bold=True, color=WHITE)
    ws["A21"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A21"].alignment = Alignment(horizontal="center")

    alert_rows = [
        ("High Priority", alerts["High"]),
        ("Medium Priority", alerts["Medium"]),
        ("Low Priority", alerts["Low"]),
        ("Total Alerts", alerts["Total"]),
    ]

    for row_no, (label, value) in enumerate(alert_rows, start=22):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)

    ws["B22"].fill = PatternFill("solid", fgColor=RED)
    ws["B23"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["B24"].fill = PatternFill("solid", fgColor=GREY)

    # Candidate tables
    def add_table(
        title: str,
        start_col: int,
        start_row: int,
        candidates: list[dict],
        bullish: bool,
    ) -> None:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + 4,
        )

        title_cell = ws.cell(start_row, start_col, title)
        title_cell.font = Font(bold=True, color=WHITE)
        title_cell.fill = PatternFill("solid", fgColor=NAVY)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Rank",
            "Symbol",
            "Score",
            "Grade",
            "Recommendation",
        ]

        for offset, heading in enumerate(headers):
            cell = ws.cell(
                start_row + 1,
                start_col + offset,
                heading,
            )
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=GREY)
            cell.border = Border(bottom=THIN)

        for index, candidate in enumerate(candidates, start=1):
            row_no = start_row + 1 + index

            values = [
                index,
                candidate["Symbol"],
                candidate["Score"],
                candidate["Grade"],
                candidate["Recommendation"],
            ]

            for offset, value in enumerate(values):
                cell = ws.cell(
                    row_no,
                    start_col + offset,
                    value,
                )
                cell.border = Border(bottom=THIN)

            ws.cell(
                row_no,
                start_col + 4,
            ).fill = PatternFill(
                "solid",
                fgColor=GREEN if bullish else RED,
            )

    add_table(
        "TOP 5 CALL CANDIDATES",
        5,
        6,
        calls,
        True,
    )

    add_table(
        "TOP 5 PUT CANDIDATES",
        5,
        16,
        puts,
        False,
    )

    widths = {
        "A": 22,
        "B": 18,
        "C": 4,
        "D": 4,
        "E": 8,
        "F": 18,
        "G": 10,
        "H": 10,
        "I": 18,
        "J": 4,
        "K": 4,
        "L": 4,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)
    create_daily_report(wb)

    try:
        wb.save(DASHBOARD)
    except PermissionError as error:
        raise PermissionError(
            "Close Dashboard.xlsx in Excel and run again."
        ) from error

    print("Daily Trading Report created successfully.")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
