
"""
AQSD Professional
Module: Dashboard Formatter
Version: 2.0
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


NAVY = "17365D"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"
GREEN = "C6EFCE"
DARK_GREEN = "006100"
YELLOW = "FFF2CC"
RED = "FFC7CE"
DARK_RED = "9C0006"
GOLD = "FFD966"
WHITE = "FFFFFF"
GREY = "E7E6E6"

THIN = Side(style="thin", color="D9D9D9")


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def style_table(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="center")

    for col_index, cells in enumerate(ws.columns, start=1):
        maximum = 0

        for cell in cells:
            value = "" if cell.value is None else str(cell.value)
            maximum = max(maximum, len(value))

        width = min(max(maximum + 2, 10), 38)
        ws.column_dimensions[get_column_letter(col_index)].width = width


def format_option_sheet(ws) -> None:
    style_table(ws)
    headers = header_map(ws)

    action_col = (
        headers.get("Recommendation")
        or headers.get("Action")
    )
    rank_col = headers.get("Rank")
    score_col = (
        headers.get("Trade Score")
        or headers.get("Option Score")
    )
    confidence_col = (
        headers.get("Trade Confidence")
        or headers.get("Confidence")
    )
    reasons_col = headers.get("Reasons")

    for row in range(2, ws.max_row + 1):
        action = (
            str(ws.cell(row, action_col).value or "").upper()
            if action_col
            else ""
        )

        if action in {"STRONG BUY", "BUY"}:
            fill = PatternFill("solid", fgColor=GREEN)
            action_font = Font(color=DARK_GREEN, bold=True)
        elif action in {"WATCH", "WAIT"}:
            fill = PatternFill("solid", fgColor=YELLOW)
            action_font = Font(bold=True)
        elif action == "AVOID":
            fill = PatternFill("solid", fgColor=RED)
            action_font = Font(color=DARK_RED, bold=True)
        else:
            fill = None
            action_font = Font()

        if fill:
            for cell in ws[row]:
                cell.fill = fill

        if action_col:
            ws.cell(row, action_col).font = action_font

        if rank_col:
            rank = ws.cell(row, rank_col).value
            if isinstance(rank, int) and rank <= 10:
                ws.cell(row, rank_col).fill = PatternFill(
                    "solid",
                    fgColor=GOLD,
                )
                ws.cell(row, rank_col).font = Font(bold=True)

    if score_col and ws.max_row >= 2:
        score_letter = get_column_letter(score_col)
        ws.conditional_formatting.add(
            f"{score_letter}2:{score_letter}{ws.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=60,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="63BE7B",
            ),
        )

    if confidence_col:
        for row in range(2, ws.max_row + 1):
            ws.cell(row, confidence_col).number_format = '0"%"'

    if reasons_col:
        ws.column_dimensions[
            get_column_letter(reasons_col)
        ].width = 55

        for row in range(2, ws.max_row + 1):
            ws.cell(row, reasons_col).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )


def format_long_term_sheet(ws) -> None:
    style_table(ws)
    headers = header_map(ws)

    action_col = headers.get("Investment Action")
    rank_col = headers.get("Rank")
    score_col = headers.get("Investment Score")
    reasons_col = headers.get("Investment Reasons")

    for row in range(2, ws.max_row + 1):
        action = (
            str(ws.cell(row, action_col).value or "").upper()
            if action_col
            else ""
        )

        if action in {"STRONG", "ACCUMULATE"}:
            fill = PatternFill("solid", fgColor=GREEN)
        elif action == "WATCH":
            fill = PatternFill("solid", fgColor=YELLOW)
        else:
            fill = PatternFill("solid", fgColor=RED)

        for cell in ws[row]:
            cell.fill = fill

        if rank_col:
            rank = ws.cell(row, rank_col).value
            if isinstance(rank, int) and rank <= 10:
                ws.cell(row, rank_col).fill = PatternFill(
                    "solid",
                    fgColor=GOLD,
                )
                ws.cell(row, rank_col).font = Font(bold=True)

    if score_col and ws.max_row >= 2:
        score_letter = get_column_letter(score_col)
        ws.conditional_formatting.add(
            f"{score_letter}2:{score_letter}{ws.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=60,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="63BE7B",
            ),
        )

    if reasons_col:
        ws.column_dimensions[
            get_column_letter(reasons_col)
        ].width = 55


def get_market_summary(wb) -> tuple[str, str, str]:
    if "Market Pulse" not in wb.sheetnames:
        return "UNKNOWN", "", ""

    ws = wb["Market Pulse"]
    bias = "UNKNOWN"
    confidence = ""
    strategy = ""

    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()

        if label == "Market Bias":
            bias = str(ws.cell(row, 2).value or "")
        elif label == "Confidence":
            confidence = str(ws.cell(row, 2).value or "")
        elif label == "Strategy":
            strategy = str(ws.cell(row, 2).value or "")

    return bias, confidence, strategy


def create_home_sheet(wb) -> None:
    if "HOME" in wb.sheetnames:
        del wb["HOME"]

    ws = wb.create_sheet("HOME", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H2")
    ws["A1"] = "AQSD PROFESSIONAL DASHBOARD"
    ws["A1"].font = Font(size=22, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Last Updated"
    ws["B4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    bias, confidence, strategy = get_market_summary(wb)

    ws["A6"] = "MARKET PULSE"
    ws["A6"].font = Font(bold=True, color=WHITE)
    ws["A6"].fill = PatternFill("solid", fgColor=NAVY)

    market_rows = [
        ("Market Bias", bias),
        ("Confidence", confidence),
        ("Strategy", strategy),
    ]

    for row_number, (label, value) in enumerate(
        market_rows,
        start=7,
    ):
        ws[f"A{row_number}"] = label
        ws[f"B{row_number}"] = value
        ws[f"A{row_number}"].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )
        ws[f"A{row_number}"].font = Font(bold=True)

    if str(bias).upper() == "BULLISH":
        ws["B7"].fill = PatternFill("solid", fgColor=GREEN)
    elif str(bias).upper() == "BEARISH":
        ws["B7"].fill = PatternFill("solid", fgColor=RED)
    else:
        ws["B7"].fill = PatternFill("solid", fgColor=YELLOW)

    option_ws = wb["Option Buying"]
    headers = header_map(option_ws)

    action_col = (
        headers.get("Recommendation")
        or headers.get("Action")
    )
    score_col = (
        headers.get("Trade Score")
        or headers.get("Option Score")
    )
    symbol_col = headers.get("Symbol")
    grade_col = (
        headers.get("Trade Grade")
        or headers.get("Grade")
    )
    stars_col = headers.get("Stars")

    counts = {
        "Stocks Scanned": option_ws.max_row - 1,
        "Strong Buy": 0,
        "Buy": 0,
        "Watch": 0,
        "Wait": 0,
        "Avoid": 0,
    }

    if action_col:
        for row in range(2, option_ws.max_row + 1):
            action = str(
                option_ws.cell(row, action_col).value or ""
            ).title()

            if action in counts:
                counts[action] += 1

    ws["A12"] = "MARKET STATISTICS"
    ws["A12"].font = Font(bold=True, color=WHITE)
    ws["A12"].fill = PatternFill("solid", fgColor=NAVY)

    for row_number, (label, value) in enumerate(
        counts.items(),
        start=13,
    ):
        ws[f"A{row_number}"] = label
        ws[f"B{row_number}"] = value
        ws[f"A{row_number}"].fill = PatternFill(
            "solid",
            fgColor=LIGHT_BLUE,
        )
        ws[f"A{row_number}"].font = Font(bold=True)

    ws["D4"] = "TOP 10 TRADE SETUPS"
    ws["D4"].font = Font(bold=True, color=WHITE)
    ws["D4"].fill = PatternFill("solid", fgColor=NAVY)

    top_headers = ["Rank", "Symbol", "Score", "Grade", "Stars"]
    for column, value in enumerate(top_headers, start=4):
        cell = ws.cell(row=5, column=column, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=GREY)

    for home_row, source_row in enumerate(
        range(2, min(option_ws.max_row, 11) + 1),
        start=6,
    ):
        ws.cell(home_row, 4, home_row - 5)
        ws.cell(
            home_row,
            5,
            option_ws.cell(source_row, symbol_col).value
            if symbol_col
            else "",
        )
        ws.cell(
            home_row,
            6,
            option_ws.cell(source_row, score_col).value
            if score_col
            else "",
        )
        ws.cell(
            home_row,
            7,
            option_ws.cell(source_row, grade_col).value
            if grade_col
            else "",
        )
        ws.cell(
            home_row,
            8,
            option_ws.cell(source_row, stars_col).value
            if stars_col
            else "",
        )

        if home_row <= 15:
            ws.cell(home_row, 4).fill = PatternFill(
                "solid",
                fgColor=GOLD,
            )

    widths = {
        "A": 22,
        "B": 20,
        "C": 4,
        "D": 10,
        "E": 20,
        "F": 12,
        "G": 10,
        "H": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    if "Option Buying" in wb.sheetnames:
        format_option_sheet(wb["Option Buying"])

    if "Long Term" in wb.sheetnames:
        format_long_term_sheet(wb["Long Term"])

    if "Summary" in wb.sheetnames:
        style_table(wb["Summary"])

    if "Option Buying" in wb.sheetnames:
        create_home_sheet(wb)

    wb.save(DASHBOARD)

    print("AQSD Dashboard v2 formatted successfully.")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
