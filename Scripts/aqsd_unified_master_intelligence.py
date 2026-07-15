
"""
AQSD Professional
Module: Unified Master Intelligence Engine
Version: 1.0

Purpose
-------
Combines the latest available intelligence from:

- Price Structure
- Sector Rotation
- Relative Strength
- Market Breadth
- News
- Macro
- Futures
- Options

Outputs
-------
- Unified Master Score from 0 to 100
- Recommendation
- Confidence %
- Engine agreement
- Data completeness
- Risk level
- Entry quality
- Explainable reasons
- SQLite history
- Excel sheet: Unified Master Intelligence

Commands
--------
python aqsd_unified_master_intelligence.py --run
python aqsd_unified_master_intelligence.py --status
python aqsd_unified_master_intelligence.py --report
python aqsd_unified_master_intelligence.py --symbol RELIANCE
"""

from __future__ import annotations

import argparse
import math
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# CONFIGURATION
# ============================================================

WEIGHTS = {
    "price_structure_score": 0.22,
    "sector_rotation_score": 0.14,
    "relative_strength_score": 0.14,
    "market_breadth_score": 0.10,
    "news_score": 0.10,
    "macro_score": 0.08,
    "futures_score": 0.12,
    "options_score": 0.10,
}

MINIMUM_COMPLETENESS_FOR_SIGNAL = 45.0


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS unified_master_intelligence (
    master_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    symbol_id INTEGER NOT NULL,
    nse_symbol TEXT NOT NULL,
    sector TEXT,
    price_structure_score REAL,
    sector_rotation_score REAL,
    relative_strength_score REAL,
    market_breadth_score REAL,
    news_score REAL,
    macro_score REAL,
    futures_score REAL,
    options_score REAL,
    master_score REAL,
    recommendation TEXT,
    confidence_percent REAL,
    engine_agreement_percent REAL,
    data_completeness_percent REAL,
    risk_level TEXT,
    entry_quality TEXT,
    directional_bias TEXT,
    reasons TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(trade_date, symbol_id),
    FOREIGN KEY(symbol_id)
        REFERENCES symbols(symbol_id)
        ON DELETE CASCADE
);

CREATE INDEX IF NOT EXISTS idx_unified_master_date
ON unified_master_intelligence(trade_date);

CREATE INDEX IF NOT EXISTS idx_unified_master_symbol
ON unified_master_intelligence(symbol_id, trade_date);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# HELPERS
# ============================================================

def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table'
          AND name = ?
        """,
        (table_name,),
    ).fetchone()

    return row is not None


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None

        return float(value)

    except (TypeError, ValueError):
        return None


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def normalize_signed_score(value: float | None) -> float | None:
    """
    Convert -100..+100 scores into 0..100 scores.
    """

    if value is None:
        return None

    return round(clamp(50 + value / 2, 0, 100), 2)


def normalize_rs_score(value: float | None) -> float | None:
    if value is None:
        return None

    return round(clamp(value, 0, 100), 2)


def recommendation(score: float, completeness: float) -> str:
    if completeness < MINIMUM_COMPLETENESS_FOR_SIGNAL:
        return "INSUFFICIENT DATA"

    if score >= 85:
        return "STRONG BUY"
    if score >= 72:
        return "BUY"
    if score >= 60:
        return "BUY ON DIPS"
    if score >= 48:
        return "WATCH"
    if score >= 35:
        return "AVOID"
    return "EXIT / STRONG AVOID"


def directional_bias(score: float) -> str:
    if score >= 78:
        return "STRONG BULLISH"
    if score >= 60:
        return "BULLISH"
    if score <= 22:
        return "STRONG BEARISH"
    if score <= 40:
        return "BEARISH"
    return "NEUTRAL"


def risk_level(
    score: float,
    agreement: float,
    completeness: float,
) -> str:
    if completeness < 50:
        return "HIGH"

    if agreement >= 75 and abs(score - 50) >= 20:
        return "LOW"

    if agreement >= 55:
        return "MEDIUM"

    return "HIGH"


def entry_quality(
    score: float,
    agreement: float,
    structure_score: float | None,
) -> str:
    if (
        score >= 80
        and agreement >= 75
        and (structure_score or 0) >= 70
    ):
        return "EXCELLENT"

    if (
        score >= 68
        and agreement >= 60
    ):
        return "GOOD"

    if score >= 50:
        return "AVERAGE"

    return "POOR"


def confidence_percent(
    score: float,
    agreement: float,
    completeness: float,
) -> float:
    directional_strength = abs(score - 50) * 2

    confidence = (
        directional_strength * 0.45
        + agreement * 0.35
        + completeness * 0.20
    )

    return round(clamp(confidence, 0, 100), 2)


# ============================================================
# DATA LOADERS
# ============================================================

def load_symbols() -> pd.DataFrame:
    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT
                symbol_id,
                nse_symbol,
                COALESCE(NULLIF(TRIM(sector), ''), 'Unmapped') AS sector
            FROM symbols
            WHERE active = 1
            ORDER BY nse_symbol
            """,
            connection,
        )


def latest_price_structure() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "price_structure_intelligence",
        ):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                symbol_id,
                nse_symbol,
                trade_date,
                structure_score AS price_structure_score,
                directional_bias AS structure_bias,
                market_structure,
                bos_signal,
                choch_signal,
                trend_strength
            FROM price_structure_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
            """,
            connection,
        )


def latest_sector_rotation() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(
            connection,
            "sector_rotation_intelligence",
        ):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                sector,
                trade_date,
                sector_rotation_score,
                rotation_state,
                trend_state
            FROM sector_rotation_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            """,
            connection,
        )


def latest_relative_strength() -> pd.DataFrame:
    with connect() as connection:
        if table_exists(connection, "relative_strength"):
            frame = pd.read_sql_query(
                """
                SELECT
                    symbol AS nse_symbol,
                    sector,
                    trade_date,
                    rs_score AS relative_strength_score,
                    rating AS rs_rating
                FROM relative_strength
                """,
                connection,
            )

            return frame

        if table_exists(connection, "intelligence_scores"):
            return pd.read_sql_query(
                """
                SELECT
                    s.nse_symbol,
                    s.sector,
                    i.score_date AS trade_date,
                    i.relative_strength_score,
                    '' AS rs_rating
                FROM intelligence_scores i
                JOIN symbols s
                    ON s.symbol_id = i.symbol_id
                WHERE i.score_date = (
                    SELECT MAX(score_date)
                    FROM intelligence_scores
                )
                  AND i.relative_strength_score IS NOT NULL
                """,
                connection,
            )

    return pd.DataFrame()


def latest_market_breadth() -> tuple[float | None, dict[str, float]]:
    with connect() as connection:
        if not table_exists(
            connection,
            "market_breadth_intelligence",
        ):
            return None, {}

        frame = pd.read_sql_query(
            """
            SELECT
                scope_type,
                scope_name,
                breadth_score
            FROM market_breadth_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM market_breadth_intelligence
            )
            """,
            connection,
        )

    if frame.empty:
        return None, {}

    market_rows = frame[
        frame["scope_type"] == "MARKET"
    ]

    market_score = (
        safe_float(
            market_rows["breadth_score"].iloc[0]
        )
        if not market_rows.empty
        else None
    )

    sector_scores = {
        str(row["scope_name"]).upper(): float(
            row["breadth_score"]
        )
        for _, row in frame[
            frame["scope_type"] == "SECTOR"
        ].iterrows()
    }

    return market_score, sector_scores


def latest_news_scores() -> tuple[dict[str, float], dict[str, float]]:
    """
    Returns:
        symbol_scores, sector_scores
    """

    with connect() as connection:
        if not table_exists(connection, "news_events"):
            return {}, {}

        frame = pd.read_sql_query(
            """
            SELECT
                nse_symbol,
                sector,
                sentiment_score,
                materiality_score,
                credibility_score,
                event_time
            FROM news_events
            WHERE event_time >= DATETIME('now', '-14 day')
            """,
            connection,
        )

    if frame.empty:
        return {}, {}

    frame["impact"] = (
        frame["sentiment_score"].fillna(0)
        * frame["materiality_score"].fillna(50)
        / 100
        * frame["credibility_score"].fillna(50)
        / 100
    )

    symbol_scores: dict[str, float] = {}
    sector_scores: dict[str, float] = {}

    symbol_frame = frame[
        frame["nse_symbol"].fillna("").str.strip() != ""
    ]

    for symbol, group in symbol_frame.groupby("nse_symbol"):
        symbol_scores[str(symbol).upper()] = round(
            float(group["impact"].mean()),
            2,
        )

    sector_frame = frame[
        frame["sector"].fillna("").str.strip() != ""
    ]

    for sector, group in sector_frame.groupby("sector"):
        sector_scores[str(sector).upper()] = round(
            float(group["impact"].mean()),
            2,
        )

    return symbol_scores, sector_scores


def latest_macro_scores() -> tuple[float | None, dict[str, float]]:
    """
    Returns:
        broad_macro_score, sector_scores
    """

    with connect() as connection:
        if table_exists(
            connection,
            "macro_policy_events",
        ):
            frame = pd.read_sql_query(
                """
                SELECT
                    affected_sectors,
                    macro_impact_score,
                    materiality_score,
                    credibility_score
                FROM macro_policy_events
                WHERE event_date >= DATE('now', '-90 day')
                """,
                connection,
            )

            if frame.empty:
                return None, {}

            weights = (
                frame["materiality_score"].fillna(50)
                * frame["credibility_score"].fillna(70)
            )

            broad = (
                float(
                    (
                        frame["macro_impact_score"].fillna(0)
                        * weights
                    ).sum()
                    / weights.sum()
                )
                if weights.sum() != 0
                else 0.0
            )

            sector_values: dict[str, list[float]] = {}

            for _, row in frame.iterrows():
                impact = safe_float(
                    row["macro_impact_score"]
                )

                if impact is None:
                    continue

                for sector in str(
                    row["affected_sectors"] or ""
                ).split("|"):
                    sector = sector.strip().upper()

                    if sector:
                        sector_values.setdefault(
                            sector,
                            [],
                        ).append(impact)

            sector_scores = {
                key: round(
                    sum(values) / len(values),
                    2,
                )
                for key, values in sector_values.items()
                if values
            }

            return round(broad, 2), sector_scores

        if table_exists(connection, "macro_events"):
            frame = pd.read_sql_query(
                """
                SELECT surprise_score
                FROM macro_events
                WHERE event_date >= DATE('now', '-90 day')
                """,
                connection,
            )

            if frame.empty:
                return None, {}

            return round(
                float(
                    frame["surprise_score"]
                    .fillna(0)
                    .mean()
                ),
                2,
            ), {}

    return None, {}


def latest_futures_scores() -> dict[str, float]:
    with connect() as connection:
        if not table_exists(connection, "futures_oi"):
            return {}

        frame = pd.read_sql_query(
            """
            SELECT
                nse_symbol,
                smart_money_score
            FROM futures_oi
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM futures_oi
            )
            """,
            connection,
        )

    if frame.empty:
        return {}

    return {
        str(symbol).upper(): round(
            float(group["smart_money_score"].mean()),
            2,
        )
        for symbol, group in frame.groupby("nse_symbol")
    }


def latest_options_scores() -> dict[str, float]:
    with connect() as connection:
        if not table_exists(
            connection,
            "options_intelligence",
        ):
            return {}

        frame = pd.read_sql_query(
            """
            SELECT
                nse_symbol,
                options_score
            FROM options_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM options_intelligence
            )
            """,
            connection,
        )

    if frame.empty:
        return {}

    return {
        str(symbol).upper(): round(
            float(group["options_score"].mean()),
            2,
        )
        for symbol, group in frame.groupby("nse_symbol")
    }


# ============================================================
# ENGINE
# ============================================================

def build_master_results() -> list[dict]:
    symbols = load_symbols()

    if symbols.empty:
        raise RuntimeError(
            "No active symbols found."
        )

    structure = latest_price_structure()
    sector_rotation = latest_sector_rotation()
    relative_strength = latest_relative_strength()

    market_breadth_score, sector_breadth_scores = (
        latest_market_breadth()
    )

    symbol_news, sector_news = latest_news_scores()
    broad_macro, sector_macro = latest_macro_scores()
    futures_scores = latest_futures_scores()
    options_scores = latest_options_scores()

    if not structure.empty:
        symbols = symbols.merge(
            structure,
            on=[
                "symbol_id",
                "nse_symbol",
            ],
            how="left",
        )

    if not relative_strength.empty:
        symbols = symbols.merge(
            relative_strength[
                [
                    "nse_symbol",
                    "relative_strength_score",
                    "rs_rating",
                ]
            ],
            on="nse_symbol",
            how="left",
        )

    if not sector_rotation.empty:
        sector_rotation = sector_rotation.copy()
        sector_rotation["sector_key"] = (
            sector_rotation["sector"]
            .astype(str)
            .str.upper()
        )

    trade_dates = []

    for frame, column in (
        (structure, "trade_date"),
        (relative_strength, "trade_date"),
        (sector_rotation, "trade_date"),
    ):
        if not frame.empty and column in frame.columns:
            trade_dates.extend(
                frame[column]
                .dropna()
                .astype(str)
                .tolist()
            )

    trade_date = (
        max(trade_dates)
        if trade_dates
        else datetime.now().date().isoformat()
    )

    results = []

    for _, row in symbols.iterrows():
        symbol = str(row["nse_symbol"]).upper()
        sector = str(row["sector"] or "Unmapped")
        sector_key = sector.upper()

        structure_score = safe_float(
            row.get("price_structure_score")
        )

        rs_score = normalize_rs_score(
            safe_float(
                row.get("relative_strength_score")
            )
        )

        sector_row = (
            sector_rotation[
                sector_rotation["sector_key"]
                == sector_key
            ]
            if not sector_rotation.empty
            else pd.DataFrame()
        )

        sector_rotation_score = (
            safe_float(
                sector_row[
                    "sector_rotation_score"
                ].iloc[0]
            )
            if not sector_row.empty
            else None
        )

        breadth_score = (
            sector_breadth_scores.get(
                sector_key,
                market_breadth_score,
            )
        )

        raw_news = symbol_news.get(
            symbol,
            sector_news.get(sector_key),
        )
        news_score = normalize_signed_score(
            raw_news
        )

        raw_macro = sector_macro.get(
            sector_key,
            broad_macro,
        )
        macro_score = normalize_signed_score(
            raw_macro
        )

        futures_score = futures_scores.get(symbol)
        options_score = options_scores.get(symbol)

        components = {
            "price_structure_score": structure_score,
            "sector_rotation_score": sector_rotation_score,
            "relative_strength_score": rs_score,
            "market_breadth_score": breadth_score,
            "news_score": news_score,
            "macro_score": macro_score,
            "futures_score": futures_score,
            "options_score": options_score,
        }

        weighted_sum = 0.0
        available_weight = 0.0
        available_values = []

        for key, value in components.items():
            if value is None:
                continue

            weight = WEIGHTS[key]
            weighted_sum += float(value) * weight
            available_weight += weight
            available_values.append(float(value))

        if available_weight == 0:
            continue

        master_score = round(
            weighted_sum / available_weight,
            2,
        )

        completeness = round(
            available_weight
            / sum(WEIGHTS.values())
            * 100,
            2,
        )

        bullish_votes = sum(
            value >= 60
            for value in available_values
        )
        bearish_votes = sum(
            value <= 40
            for value in available_values
        )

        agreement = round(
            max(
                bullish_votes,
                bearish_votes,
                len(available_values)
                - bullish_votes
                - bearish_votes,
            )
            / len(available_values)
            * 100,
            2,
        )

        rec = recommendation(
            master_score,
            completeness,
        )
        bias = directional_bias(master_score)
        confidence = confidence_percent(
            master_score,
            agreement,
            completeness,
        )
        risk = risk_level(
            master_score,
            agreement,
            completeness,
        )
        quality = entry_quality(
            master_score,
            agreement,
            structure_score,
        )

        reasons = []

        structure_bias = str(
            row.get("structure_bias") or ""
        )
        market_structure = str(
            row.get("market_structure") or ""
        )
        bos_signal = str(
            row.get("bos_signal") or ""
        )
        choch_signal = str(
            row.get("choch_signal") or ""
        )
        trend_strength = str(
            row.get("trend_strength") or ""
        )
        rs_rating = str(
            row.get("rs_rating") or ""
        )

        if structure_score is not None:
            reasons.append(
                f"Structure {structure_score:.1f}"
            )

        if market_structure:
            reasons.append(market_structure)

        if bos_signal and bos_signal != "NONE":
            reasons.append(bos_signal)

        if choch_signal and choch_signal != "NONE":
            reasons.append(choch_signal)

        if trend_strength:
            reasons.append(
                f"Trend {trend_strength}"
            )

        if sector_rotation_score is not None:
            reasons.append(
                f"Sector {sector_rotation_score:.1f}"
            )

        if rs_score is not None:
            reasons.append(
                f"RS {rs_score:.1f} {rs_rating}"
            )

        if breadth_score is not None:
            reasons.append(
                f"Breadth {breadth_score:.1f}"
            )

        if raw_news is not None:
            reasons.append(
                f"News impact {raw_news:.1f}"
            )

        if raw_macro is not None:
            reasons.append(
                f"Macro impact {raw_macro:.1f}"
            )

        if futures_score is not None:
            reasons.append(
                f"Futures {futures_score:.1f}"
            )

        if options_score is not None:
            reasons.append(
                f"Options {options_score:.1f}"
            )

        results.append(
            {
                "trade_date": trade_date,
                "symbol_id": int(row["symbol_id"]),
                "nse_symbol": symbol,
                "sector": sector,
                **components,
                "master_score": master_score,
                "recommendation": rec,
                "confidence_percent": confidence,
                "engine_agreement_percent": agreement,
                "data_completeness_percent": completeness,
                "risk_level": risk,
                "entry_quality": quality,
                "directional_bias": bias,
                "reasons": " | ".join(reasons),
            }
        )

    return sorted(
        results,
        key=lambda item: (
            item["master_score"],
            item["confidence_percent"],
            item["data_completeness_percent"],
        ),
        reverse=True,
    )


def save_results(results: list[dict]) -> None:
    with connect() as connection:
        for result in results:
            connection.execute(
                """
                INSERT INTO unified_master_intelligence(
                    trade_date,
                    symbol_id,
                    nse_symbol,
                    sector,
                    price_structure_score,
                    sector_rotation_score,
                    relative_strength_score,
                    market_breadth_score,
                    news_score,
                    macro_score,
                    futures_score,
                    options_score,
                    master_score,
                    recommendation,
                    confidence_percent,
                    engine_agreement_percent,
                    data_completeness_percent,
                    risk_level,
                    entry_quality,
                    directional_bias,
                    reasons,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(trade_date, symbol_id)
                DO UPDATE SET
                    sector = excluded.sector,
                    price_structure_score =
                        excluded.price_structure_score,
                    sector_rotation_score =
                        excluded.sector_rotation_score,
                    relative_strength_score =
                        excluded.relative_strength_score,
                    market_breadth_score =
                        excluded.market_breadth_score,
                    news_score = excluded.news_score,
                    macro_score = excluded.macro_score,
                    futures_score = excluded.futures_score,
                    options_score = excluded.options_score,
                    master_score = excluded.master_score,
                    recommendation = excluded.recommendation,
                    confidence_percent =
                        excluded.confidence_percent,
                    engine_agreement_percent =
                        excluded.engine_agreement_percent,
                    data_completeness_percent =
                        excluded.data_completeness_percent,
                    risk_level = excluded.risk_level,
                    entry_quality = excluded.entry_quality,
                    directional_bias =
                        excluded.directional_bias,
                    reasons = excluded.reasons,
                    created_at = excluded.created_at
                """,
                (
                    result["trade_date"],
                    result["symbol_id"],
                    result["nse_symbol"],
                    result["sector"],
                    result["price_structure_score"],
                    result["sector_rotation_score"],
                    result["relative_strength_score"],
                    result["market_breadth_score"],
                    result["news_score"],
                    result["macro_score"],
                    result["futures_score"],
                    result["options_score"],
                    result["master_score"],
                    result["recommendation"],
                    result["confidence_percent"],
                    result["engine_agreement_percent"],
                    result["data_completeness_percent"],
                    result["risk_level"],
                    result["entry_quality"],
                    result["directional_bias"],
                    result["reasons"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()


def run_engine() -> list[dict]:
    setup_schema()

    run_id = start_run(
        "aqsd_unified_master_intelligence",
        "Running unified master intelligence",
    )

    try:
        results = build_master_results()
        save_results(results)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=len(results),
            errors_count=0,
            message=f"Symbols ranked={len(results)}",
        )

        return results

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


# ============================================================
# REPORTING
# ============================================================

def latest_results() -> pd.DataFrame:
    setup_schema()

    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM unified_master_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM unified_master_intelligence
            )
            ORDER BY
                master_score DESC,
                confidence_percent DESC,
                nse_symbol
            """,
            connection,
        )


def write_report(
    results: list[dict] | None = None,
) -> None:
    if results is None:
        results = latest_results().to_dict(
            "records"
        )

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Unified Master Intelligence" in wb.sheetnames:
        del wb["Unified Master Intelligence"]

    ws = wb.create_sheet(
        "Unified Master Intelligence",
        1,
    )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:W2")
    ws["A1"] = "AQSD PROFESSIONAL - UNIFIED MASTER INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Stocks Ranked"
    ws["B4"] = len(results)
    ws["D4"] = "Updated"
    ws["E4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    for ref in ("A4", "D4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Rank",
        "Symbol",
        "Sector",
        "Price Structure",
        "Sector Rotation",
        "Relative Strength",
        "Market Breadth",
        "News",
        "Macro",
        "Futures",
        "Options",
        "Master Score",
        "Recommendation",
        "Confidence %",
        "Engine Agreement %",
        "Data Completeness %",
        "Risk",
        "Entry Quality",
        "Directional Bias",
        "Reasons",
    ]

    for col, heading in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, result in enumerate(
        results,
        start=8,
    ):
        values = [
            row_no - 7,
            result["nse_symbol"],
            result["sector"],
            result["price_structure_score"],
            result["sector_rotation_score"],
            result["relative_strength_score"],
            result["market_breadth_score"],
            result["news_score"],
            result["macro_score"],
            result["futures_score"],
            result["options_score"],
            result["master_score"],
            result["recommendation"],
            result["confidence_percent"],
            result["engine_agreement_percent"],
            result["data_completeness_percent"],
            result["risk_level"],
            result["entry_quality"],
            result["directional_bias"],
            result["reasons"],
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(bottom=THIN)

        score = float(result["master_score"])

        ws.cell(row_no, 12).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )
        ws.cell(row_no, 12).font = Font(
            bold=True
        )

        recommendation_text = str(
            result["recommendation"]
        )

        ws.cell(row_no, 13).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BUY" in recommendation_text
                else RED
                if (
                    "AVOID" in recommendation_text
                    or "EXIT" in recommendation_text
                )
                else YELLOW
            ),
        )
        ws.cell(row_no, 13).font = Font(
            bold=True
        )

        risk = str(result["risk_level"])

        ws.cell(row_no, 17).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if risk == "LOW"
                else RED
                if risk == "HIGH"
                else YELLOW
            ),
        )

        bias = str(
            result["directional_bias"]
        )

        ws.cell(row_no, 19).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if "BULLISH" in bias
                else RED
                if "BEARISH" in bias
                else GREY
            ),
        )

    widths = {
        "A": 8,
        "B": 16,
        "C": 20,
        "D": 16,
        "E": 16,
        "F": 17,
        "G": 16,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 15,
        "M": 22,
        "N": 14,
        "O": 18,
        "P": 18,
        "Q": 12,
        "R": 16,
        "S": 18,
        "T": 90,
        "U": 14,
        "V": 14,
        "W": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


def show_symbol(symbol: str) -> None:
    text = symbol.strip().upper()

    if text.endswith(".NS"):
        text = text[:-3]

    frame = latest_results()

    if frame.empty:
        print("No unified master results found.")
        return

    selected = frame[
        frame["nse_symbol"].str.upper()
        == text
    ]

    if selected.empty:
        print(f"Symbol not found: {text}")
        return

    row = selected.iloc[0]

    print("\nAQSD UNIFIED MASTER SNAPSHOT")
    print("=" * 72)

    fields = [
        "nse_symbol",
        "sector",
        "price_structure_score",
        "sector_rotation_score",
        "relative_strength_score",
        "market_breadth_score",
        "news_score",
        "macro_score",
        "futures_score",
        "options_score",
        "master_score",
        "recommendation",
        "confidence_percent",
        "engine_agreement_percent",
        "data_completeness_percent",
        "risk_level",
        "entry_quality",
        "directional_bias",
        "reasons",
    ]

    for field in fields:
        print(
            f"{field:<28}"
            f"{row[field]}"
        )

    print("=" * 72)


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT symbol_id) AS symbols,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM unified_master_intelligence
            """
        ).fetchone()

    print("\nAQSD UNIFIED MASTER STATUS")
    print("=" * 72)
    print(f"Stored records:   {row['total'] or 0}")
    print(f"Symbols covered:  {row['symbols'] or 0}")
    print(f"First date:       {row['first_date'] or 'No data'}")
    print(f"Latest date:      {row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "AQSD Unified Master "
            "Intelligence Engine."
        )
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Run unified master intelligence.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild Excel report from stored data.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show database status.",
    )

    parser.add_argument(
        "--symbol",
        metavar="SYMBOL",
        help="Show one symbol's unified intelligence.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.symbol:
        show_symbol(args.symbol)
        return

    if args.run:
        results = run_engine()
        write_report(results)

        print("\nAQSD UNIFIED MASTER INTELLIGENCE")
        print("=" * 72)
        print(f"Stocks ranked: {len(results)}")

        if results:
            best = results[0]
            worst = results[-1]

            print(
                f"Highest score: "
                f"{best['nse_symbol']} "
                f"({best['master_score']})"
            )

            print(
                f"Lowest score:  "
                f"{worst['nse_symbol']} "
                f"({worst['master_score']})"
            )

        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report()
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
