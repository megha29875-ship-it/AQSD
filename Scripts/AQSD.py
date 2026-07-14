
"""
AQSD Professional
Master Controller
Version: 4.0

Adds:
- Smart Alerts
- Daily Trading Report
- Strategy Scorecard
- Full one-click daily + portfolio workflow
"""

from __future__ import annotations

import argparse
import os
import subprocess
import sys
import time
from pathlib import Path


SCRIPTS_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPTS_DIR.parent
OUTPUT_FILE = BASE_DIR / "Output" / "Dashboard.xlsx"

SCRIPTS = {
    "update": SCRIPTS_DIR / "update_fno_nse_v2.py",
    "scanner": SCRIPTS_DIR / "scanner.py",
    "market_pulse": SCRIPTS_DIR / "market_pulse.py",
    "format_dashboard": SCRIPTS_DIR / "format_dashboard.py",
    "risk_dashboard": SCRIPTS_DIR / "risk_dashboard.py",
    "live_watchlist": SCRIPTS_DIR / "live_watchlist.py",
    "smart_alerts": SCRIPTS_DIR / "smart_alerts.py",
    "daily_report": SCRIPTS_DIR / "daily_trading_report.py",
    "strategy_scorecard": SCRIPTS_DIR / "strategy_scorecard.py",
    "portfolio_manager": SCRIPTS_DIR / "portfolio_manager.py",
    "portfolio_live": SCRIPTS_DIR / "portfolio_live_assistant.py",
    "trade_journal": SCRIPTS_DIR / "trade_journal.py",
    "analytics": SCRIPTS_DIR / "performance_analytics.py",
    "performance_dashboard": SCRIPTS_DIR / "performance_dashboard.py",
    "portfolio_heatmap": SCRIPTS_DIR / "portfolio_heatmap.py",
    "equity_curve": SCRIPTS_DIR / "equity_curve.py",
    "drawdown": SCRIPTS_DIR / "drawdown_analyzer.py",
    "sector_exposure": SCRIPTS_DIR / "sector_exposure.py",
    "portfolio_allocation": SCRIPTS_DIR / "portfolio_allocation.py",
    "backup": SCRIPTS_DIR / "auto_backup.py",
}


def run_script(
    label: str,
    script_path: Path,
    *,
    required: bool = True,
) -> bool:
    if not script_path.exists():
        if required:
            raise FileNotFoundError(
                f"{label} script not found:\n{script_path}"
            )

        print(f"\nSKIPPED: {label}")
        print(f"Missing optional file: {script_path.name}")
        return False

    print("\n" + "=" * 70)
    print(label)
    print("=" * 70)

    completed = subprocess.run(
        [sys.executable, str(script_path)],
        cwd=str(SCRIPTS_DIR),
        check=False,
    )

    if completed.returncode != 0:
        raise RuntimeError(
            f"{label} failed with exit code {completed.returncode}."
        )

    return True


def dashboard_has_sheet(sheet_name: str) -> bool:
    if not OUTPUT_FILE.exists():
        return False

    try:
        from openpyxl import load_workbook

        wb = load_workbook(
            OUTPUT_FILE,
            read_only=True,
            data_only=False,
        )
        exists = sheet_name in wb.sheetnames
        wb.close()
        return exists

    except Exception:
        return False


def open_dashboard() -> None:
    if not OUTPUT_FILE.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{OUTPUT_FILE}"
        )

    if os.name == "nt":
        os.startfile(OUTPUT_FILE)  # type: ignore[attr-defined]
    else:
        print(f"\nOpen manually:\n{OUTPUT_FILE}")


def run_backup() -> None:
    run_script(
        "CREATE VERIFIED DASHBOARD BACKUP",
        SCRIPTS["backup"],
        required=False,
    )


def run_daily(skip_update: bool) -> None:
    if skip_update:
        print("\nNSE F&O update skipped.")
    else:
        run_script(
            "UPDATE NSE F&O UNIVERSE",
            SCRIPTS["update"],
        )

    run_script(
        "SCAN NSE F&O STOCKS",
        SCRIPTS["scanner"],
    )

    run_script(
        "CREATE MARKET PULSE",
        SCRIPTS["market_pulse"],
    )

    run_script(
        "FORMAT PROFESSIONAL DASHBOARD",
        SCRIPTS["format_dashboard"],
    )

    run_script(
        "ADD RISK & POSITION PLAN",
        SCRIPTS["risk_dashboard"],
        required=False,
    )

    run_script(
        "CREATE LIVE WATCHLIST",
        SCRIPTS["live_watchlist"],
        required=False,
    )

    run_script(
        "CREATE STRATEGY SCORECARD",
        SCRIPTS["strategy_scorecard"],
        required=False,
    )

    run_script(
        "CREATE SMART ALERTS",
        SCRIPTS["smart_alerts"],
        required=False,
    )

    run_script(
        "CREATE DAILY TRADING REPORT",
        SCRIPTS["daily_report"],
        required=False,
    )

    run_backup()


def run_portfolio() -> None:
    if not dashboard_has_sheet("Portfolio"):
        raise RuntimeError(
            "Portfolio sheet is missing.\n"
            "Run: python AQSD.py --mode setup-portfolio"
        )

    if not dashboard_has_sheet("Trade Journal"):
        run_script(
            "CREATE TRADE JOURNAL",
            SCRIPTS["trade_journal"],
        )

    run_script(
        "UPDATE PORTFOLIO, MTM & TRAILING STOP",
        SCRIPTS["portfolio_live"],
    )

    run_script(
        "REFRESH TRADE JOURNAL",
        SCRIPTS["trade_journal"],
    )

    run_script(
        "CREATE PORTFOLIO HEAT MAP",
        SCRIPTS["portfolio_heatmap"],
        required=False,
    )

    run_script(
        "CREATE PERFORMANCE ANALYTICS",
        SCRIPTS["analytics"],
        required=False,
    )

    run_script(
        "CREATE PERFORMANCE DASHBOARD",
        SCRIPTS["performance_dashboard"],
        required=False,
    )

    run_script(
        "CREATE EQUITY CURVE",
        SCRIPTS["equity_curve"],
        required=False,
    )

    run_script(
        "CREATE DRAWDOWN ANALYSIS",
        SCRIPTS["drawdown"],
        required=False,
    )

    run_script(
        "CREATE SECTOR EXPOSURE",
        SCRIPTS["sector_exposure"],
        required=False,
    )

    run_script(
        "CREATE PORTFOLIO ALLOCATION",
        SCRIPTS["portfolio_allocation"],
        required=False,
    )

    run_script(
        "REFRESH SMART ALERTS",
        SCRIPTS["smart_alerts"],
        required=False,
    )

    run_script(
        "REFRESH DAILY TRADING REPORT",
        SCRIPTS["daily_report"],
        required=False,
    )

    run_backup()


def setup_portfolio() -> None:
    print("\nWARNING: Existing Portfolio sheet will be replaced.")

    run_script(
        "CREATE / RESET PORTFOLIO",
        SCRIPTS["portfolio_manager"],
    )

    if not dashboard_has_sheet("Trade Journal"):
        run_script(
            "CREATE TRADE JOURNAL",
            SCRIPTS["trade_journal"],
        )

    run_script(
        "CREATE PERFORMANCE ANALYTICS",
        SCRIPTS["analytics"],
        required=False,
    )

    run_script(
        "CREATE SMART ALERTS",
        SCRIPTS["smart_alerts"],
        required=False,
    )

    run_script(
        "CREATE DAILY TRADING REPORT",
        SCRIPTS["daily_report"],
        required=False,
    )

    run_backup()


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run AQSD Professional workflows."
    )

    parser.add_argument(
        "--mode",
        choices=[
            "daily",
            "portfolio",
            "all",
            "setup-portfolio",
        ],
        default="daily",
    )

    parser.add_argument(
        "--skip-update",
        action="store_true",
        help="Use the existing NSE F&O list.",
    )

    parser.add_argument(
        "--no-open",
        action="store_true",
        help="Do not open Dashboard.xlsx after completion.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    started = time.perf_counter()

    print("\n" + "=" * 70)
    print("AQSD PROFESSIONAL v4.0")
    print("NSE F&O TRADING, ALERTS, PORTFOLIO & ANALYTICS WORKSTATION")
    print("=" * 70)
    print(f"Mode: {args.mode}")

    try:
        if args.mode == "daily":
            run_daily(args.skip_update)

        elif args.mode == "portfolio":
            run_portfolio()

        elif args.mode == "all":
            run_daily(args.skip_update)

            if dashboard_has_sheet("Portfolio"):
                run_portfolio()
            else:
                print("\nPortfolio workflow skipped.")
                print("Create it with:")
                print("python AQSD.py --mode setup-portfolio")

        elif args.mode == "setup-portfolio":
            setup_portfolio()

    except (FileNotFoundError, RuntimeError, PermissionError) as error:
        print("\n" + "=" * 70)
        print("AQSD STOPPED")
        print("=" * 70)
        print(error)

        if isinstance(error, PermissionError):
            print("\nClose Dashboard.xlsx in Excel and run again.")

        raise SystemExit(1)

    elapsed = time.perf_counter() - started

    print("\n" + "=" * 70)
    print("AQSD COMPLETED SUCCESSFULLY")
    print(f"Mode: {args.mode}")
    print(f"Time taken: {elapsed:.1f} seconds")
    print(f"Dashboard: {OUTPUT_FILE}")
    print("=" * 70)

    if not args.no_open and OUTPUT_FILE.exists():
        try:
            open_dashboard()
        except OSError as error:
            print(f"\nDashboard created but could not be opened: {error}")


if __name__ == "__main__":
    main()
