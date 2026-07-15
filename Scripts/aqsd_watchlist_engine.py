
"""
AQSD Professional
Module: Watchlist Intelligence Engine
Version: 1.0

Purpose
-------
Creates a dynamic institutional watchlist from AQSD intelligence.

Commands
--------
python aqsd_watchlist_engine.py --run
python aqsd_watchlist_engine.py --status
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from aqsd_database import connect, setup_database

BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT = BASE_DIR / "Output"
DASHBOARD = OUTPUT / "Dashboard.xlsx"


def load_watchlist() -> pd.DataFrame:
    with connect() as con:
        try:
            return pd.read_sql_query(
                """
                SELECT
                    priority_rank,
                    nse_symbol,
                    sector,
                    action,
                    master_score,
                    confidence_percent,
                    target_1,
                    target_2
                FROM aqsd_trade_decisions
                WHERE trade_date=(
                    SELECT MAX(trade_date)
                    FROM aqsd_trade_decisions
                )
                AND action IN ('STRONG BUY','BUY','BUY ON DIP')
                ORDER BY priority_rank
                LIMIT 50
                """,
                con,
            )
        except Exception:
            return pd.DataFrame()


def write_excel(df: pd.DataFrame) -> None:
    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

    if "AQSD Watchlist" in wb.sheetnames:
        del wb["AQSD Watchlist"]

    ws = wb.create_sheet("AQSD Watchlist")

    ws["A1"] = "AQSD PROFESSIONAL WATCHLIST"
    ws["A2"] = f"Generated : {datetime.now():%d-%m-%Y %H:%M}"

    if df.empty:
        ws["A4"] = "No qualifying stocks."
    else:
        for c, h in enumerate(df.columns, 1):
            ws.cell(4, c).value = h
        for r, row in enumerate(df.itertuples(index=False), 5):
            for c, value in enumerate(row, 1):
                ws.cell(r, c).value = value

    wb.save(DASHBOARD)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run", action="store_true")
    parser.add_argument("--status", action="store_true")
    args = parser.parse_args()

    setup_database()

    if args.status:
        print("AQSD Watchlist Engine : READY")
        return

    df = load_watchlist()
    OUTPUT.mkdir(exist_ok=True)
    df.to_csv(OUTPUT / "AQSD_Watchlist.csv", index=False)
    write_excel(df)

    print("AQSD WATCHLIST ENGINE")
    print("=" * 50)
    print(f"Stocks Selected : {len(df)}")
    print(f"CSV            : {OUTPUT/'AQSD_Watchlist.csv'}")
    print(f"Dashboard      : {DASHBOARD}")


if __name__ == "__main__":
    main()
