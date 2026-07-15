
"""
AQSD Market Intelligence
Module: Sector Rotation Intelligence Engine
Version: 1.0

Purpose
-------
Ranks sectors using data already stored in aqsd_core.db.

Inputs
------
- Price Structure Intelligence
- Latest cached prices
- Symbol Master sector mapping
- Macro & Policy Intelligence, when available
- News Intelligence, when available
- Futures Intelligence, when available
- Options Intelligence, when available

Outputs
-------
- Sector Rotation Score from 0 to 100
- Sector trend and rotation state
- Breadth inside each sector
- Sector leaders and laggards
- Early rotation detection
- Excel report: Sector Rotation Intelligence
- SQLite history table

Commands
--------
python aqsd_sector_rotation.py --run
python aqsd_sector_rotation.py --run --minimum-stocks 2
python aqsd_sector_rotation.py --status
python aqsd_sector_rotation.py --report
python aqsd_sector_rotation.py --sector BANK
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# SETTINGS
# ============================================================

DEFAULT_MINIMUM_STOCKS = 2
DEFAULT_TOP_STOCKS = 5


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS sector_rotation_intelligence (
    sector_rotation_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    sector TEXT NOT NULL,
    stock_count INTEGER,
    average_structure_score REAL,
    average_5d_return REAL,
    average_20d_return REAL,
    bullish_breadth_percent REAL,
    bearish_breadth_percent REAL,
    breakout_percent REAL,
    macro_score REAL,
    news_score REAL,
    futures_score REAL,
    options_score REAL,
    sector_rotation_score REAL,
    rotation_state TEXT,
    trend_state TEXT,
    leader_symbol TEXT,
    leader_score REAL,
    laggard_symbol TEXT,
    laggard_score REAL,
    explanation TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(trade_date, sector)
);

CREATE INDEX IF NOT EXISTS idx_sector_rotation_date
ON sector_rotation_intelligence(trade_date);

CREATE INDEX IF NOT EXISTS idx_sector_rotation_sector
ON sector_rotation_intelligence(sector);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# BASIC HELPERS
# ============================================================

def safe_float(value) -> float | None:
    try:
        if value is None or value == "":
            return None

        return float(value)

    except (TypeError, ValueError):
        return None


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def normalize_sector(value: object) -> str:
    text = str(value or "").strip()

    return text if text else "Unmapped"


# ============================================================
# SOURCE DATA
# ============================================================

def latest_structure_frame() -> pd.DataFrame:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                p.trade_date,
                p.symbol_id,
                p.nse_symbol,
                s.sector,
                p.close_price,
                p.structure_score,
                p.directional_bias,
                p.market_structure,
                p.bos_signal,
                p.choch_signal,
                p.adx_14
            FROM price_structure_intelligence p
            JOIN symbols s
                ON s.symbol_id = p.symbol_id
            WHERE p.trade_date = (
                SELECT MAX(trade_date)
                FROM price_structure_intelligence
            )
              AND s.active = 1
            ORDER BY s.sector, p.structure_score DESC
            """,
            connection,
        )

    if not frame.empty:
        frame["sector"] = frame["sector"].apply(normalize_sector)

    return frame


def price_returns_frame() -> pd.DataFrame:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                s.symbol_id,
                s.nse_symbol,
                s.sector,
                p.trade_date,
                p.close
            FROM symbols s
            JOIN daily_prices p
                ON p.symbol_id = s.symbol_id
            WHERE s.active = 1
              AND p.trade_date >= (
                  SELECT DATE(MAX(trade_date), '-45 day')
                  FROM daily_prices
              )
            ORDER BY s.symbol_id, p.trade_date
            """,
            connection,
        )

    if frame.empty:
        return frame

    frame["trade_date"] = pd.to_datetime(frame["trade_date"])
    frame["sector"] = frame["sector"].apply(normalize_sector)

    rows = []

    for symbol_id, group in frame.groupby("symbol_id"):
        group = group.sort_values("trade_date")
        close = group["close"].dropna()

        if close.empty:
            continue

        latest = float(close.iloc[-1])

        def return_for(period: int) -> float | None:
            if len(close) <= period:
                return None

            previous = float(close.iloc[-1 - period])

            if previous == 0:
                return None

            return round((latest / previous - 1) * 100, 2)

        rows.append(
            {
                "symbol_id": int(symbol_id),
                "nse_symbol": group["nse_symbol"].iloc[-1],
                "sector": group["sector"].iloc[-1],
                "return_5d": return_for(5),
                "return_20d": return_for(20),
            }
        )

    return pd.DataFrame(rows)


def latest_macro_sector_scores() -> dict[str, float]:
    with connect() as connection:
        exists = connection.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
              AND name = 'macro_policy_events'
            """
        ).fetchone()

        if not exists:
            return {}

        frame = pd.read_sql_query(
            """
            SELECT
                affected_sectors,
                macro_impact_score
            FROM macro_policy_events
            WHERE event_date >= DATE('now', '-90 day')
            """,
            connection,
        )

    scores: dict[str, list[float]] = {}

    for _, row in frame.iterrows():
        sectors = str(row.get("affected_sectors") or "").split("|")
        impact = safe_float(row.get("macro_impact_score"))

        if impact is None:
            continue

        for sector in sectors:
            sector = sector.strip()

            if sector:
                scores.setdefault(sector.upper(), []).append(impact)

    return {
        sector: round(float(np.mean(values)), 2)
        for sector, values in scores.items()
        if values
    }


def latest_news_sector_scores() -> dict[str, float]:
    with connect() as connection:
        exists = connection.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
              AND name = 'news_events'
            """
        ).fetchone()

        if not exists:
            return {}

        frame = pd.read_sql_query(
            """
            SELECT
                sector,
                sentiment_score,
                materiality_score,
                credibility_score,
                event_time
            FROM news_events
            WHERE event_time >= DATETIME('now', '-14 day')
              AND sector IS NOT NULL
              AND TRIM(sector) <> ''
            """,
            connection,
        )

    if frame.empty:
        return {}

    rows = []

    for _, row in frame.iterrows():
        sentiment = safe_float(row.get("sentiment_score")) or 0
        materiality = safe_float(row.get("materiality_score")) or 50
        credibility = safe_float(row.get("credibility_score")) or 50

        impact = (
            sentiment
            * (materiality / 100)
            * (credibility / 100)
        )

        rows.append(
            {
                "sector": normalize_sector(row.get("sector")).upper(),
                "impact": impact,
            }
        )

    if not rows:
        return {}

    temp = pd.DataFrame(rows)

    return {
        sector: round(float(group["impact"].mean()), 2)
        for sector, group in temp.groupby("sector")
    }


def latest_futures_sector_scores() -> dict[str, float]:
    with connect() as connection:
        exists = connection.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
              AND name = 'futures_oi'
            """
        ).fetchone()

        if not exists:
            return {}

        frame = pd.read_sql_query(
            """
            SELECT
                s.sector,
                f.smart_money_score
            FROM futures_oi f
            JOIN symbols s
                ON s.nse_symbol = f.nse_symbol
            WHERE f.trade_date = (
                SELECT MAX(trade_date)
                FROM futures_oi
            )
              AND s.active = 1
            """,
            connection,
        )

    if frame.empty:
        return {}

    frame["sector"] = frame["sector"].apply(normalize_sector)

    return {
        sector.upper(): round(float(group["smart_money_score"].mean()), 2)
        for sector, group in frame.groupby("sector")
    }


def latest_options_sector_scores() -> dict[str, float]:
    with connect() as connection:
        exists = connection.execute(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
              AND name = 'options_intelligence'
            """
        ).fetchone()

        if not exists:
            return {}

        frame = pd.read_sql_query(
            """
            SELECT
                s.sector,
                o.options_score
            FROM options_intelligence o
            JOIN symbols s
                ON s.nse_symbol = o.nse_symbol
            WHERE o.trade_date = (
                SELECT MAX(trade_date)
                FROM options_intelligence
            )
              AND s.active = 1
            """,
            connection,
        )

    if frame.empty:
        return {}

    frame["sector"] = frame["sector"].apply(normalize_sector)

    return {
        sector.upper(): round(float(group["options_score"].mean()), 2)
        for sector, group in frame.groupby("sector")
    }


# ============================================================
# SCORING LOGIC
# ============================================================

def score_from_return(value: float | None, scale: float) -> float:
    if value is None or np.isnan(value):
        return 50.0

    return clamp(50 + value * scale, 0, 100)


def normalize_signed_score(value: float | None) -> float:
    if value is None:
        return 50.0

    return clamp(50 + value / 2, 0, 100)


def rotation_state(
    score: float,
    average_5d_return: float,
    average_20d_return: float,
    bullish_breadth: float,
) -> str:
    if (
        score >= 75
        and average_5d_return > 0
        and average_20d_return > 0
        and bullish_breadth >= 60
    ):
        return "STRONG INFLOW"

    if (
        score >= 60
        and average_5d_return > 0
        and bullish_breadth >= 50
    ):
        return "INFLOW"

    if (
        score <= 25
        and average_5d_return < 0
        and average_20d_return < 0
    ):
        return "STRONG OUTFLOW"

    if score <= 40 and average_5d_return < 0:
        return "OUTFLOW"

    if average_5d_return > 0 and average_20d_return <= 0:
        return "EARLY ROTATION"

    if average_5d_return < 0 and average_20d_return > 0:
        return "WEAKENING LEADERSHIP"

    return "NEUTRAL"


def trend_state(
    average_structure_score: float,
    average_20d_return: float,
) -> str:
    if average_structure_score >= 65 and average_20d_return > 0:
        return "BULLISH"

    if average_structure_score <= 35 and average_20d_return < 0:
        return "BEARISH"

    return "SIDEWAYS"


def calculate_sector_score(
    average_structure_score: float,
    average_5d_return: float,
    average_20d_return: float,
    bullish_breadth: float,
    breakout_percent: float,
    macro_score: float | None,
    news_score: float | None,
    futures_score: float | None,
    options_score: float | None,
) -> float:
    components = {
        "structure": (average_structure_score, 0.30),
        "return_5d": (
            score_from_return(average_5d_return, 5),
            0.15,
        ),
        "return_20d": (
            score_from_return(average_20d_return, 3),
            0.15,
        ),
        "breadth": (bullish_breadth, 0.15),
        "breakout": (breakout_percent, 0.10),
        "macro": (normalize_signed_score(macro_score), 0.05),
        "news": (normalize_signed_score(news_score), 0.04),
        "futures": (
            futures_score if futures_score is not None else 50,
            0.04,
        ),
        "options": (
            options_score if options_score is not None else 50,
            0.02,
        ),
    }

    weighted_sum = 0.0
    available_weight = 0.0

    for value, weight in components.values():
        if value is None:
            continue

        weighted_sum += float(value) * weight
        available_weight += weight

    if available_weight == 0:
        return 50.0

    return round(
        clamp(weighted_sum / available_weight, 0, 100),
        2,
    )


# ============================================================
# ENGINE
# ============================================================

def build_sector_results(
    minimum_stocks: int,
) -> tuple[list[dict], pd.DataFrame]:
    structure = latest_structure_frame()
    returns = price_returns_frame()

    if structure.empty:
        raise RuntimeError(
            "No Price Structure results found. "
            "Run aqsd_price_structure.py --run first."
        )

    merged = structure.merge(
        returns,
        on=[
            "symbol_id",
            "nse_symbol",
            "sector",
        ],
        how="left",
    )

    macro_scores = latest_macro_sector_scores()
    news_scores = latest_news_sector_scores()
    futures_scores = latest_futures_sector_scores()
    options_scores = latest_options_sector_scores()

    results = []

    for sector, group in merged.groupby("sector"):
        if len(group) < minimum_stocks:
            continue

        average_structure = float(
            group["structure_score"].fillna(50).mean()
        )

        average_5d = float(
            group["return_5d"].dropna().mean()
        ) if group["return_5d"].notna().any() else 0.0

        average_20d = float(
            group["return_20d"].dropna().mean()
        ) if group["return_20d"].notna().any() else 0.0

        bullish_breadth = float(
            (group["structure_score"] >= 60).mean() * 100
        )

        bearish_breadth = float(
            (group["structure_score"] <= 40).mean() * 100
        )

        breakout_percent = float(
            (
                (
                    group["bos_signal"] == "BULLISH BOS"
                )
                | (
                    group["choch_signal"] == "BULLISH CHOCH"
                )
            ).mean()
            * 100
        )

        sector_key = sector.upper()

        macro_score = macro_scores.get(sector_key)
        news_score = news_scores.get(sector_key)
        futures_score = futures_scores.get(sector_key)
        options_score = options_scores.get(sector_key)

        rotation_score = calculate_sector_score(
            average_structure,
            average_5d,
            average_20d,
            bullish_breadth,
            breakout_percent,
            macro_score,
            news_score,
            futures_score,
            options_score,
        )

        ranked = group.sort_values(
            "structure_score",
            ascending=False,
        )

        leader = ranked.iloc[0]
        laggard = ranked.iloc[-1]

        state = rotation_state(
            rotation_score,
            average_5d,
            average_20d,
            bullish_breadth,
        )

        trend = trend_state(
            average_structure,
            average_20d,
        )

        explanation_parts = [
            f"Structure {average_structure:.1f}",
            f"5D return {average_5d:.2f}%",
            f"20D return {average_20d:.2f}%",
            f"Bullish breadth {bullish_breadth:.1f}%",
            f"Breakout participation {breakout_percent:.1f}%",
        ]

        if macro_score is not None:
            explanation_parts.append(
                f"Macro {macro_score:.1f}"
            )

        if news_score is not None:
            explanation_parts.append(
                f"News {news_score:.1f}"
            )

        if futures_score is not None:
            explanation_parts.append(
                f"Futures {futures_score:.1f}"
            )

        if options_score is not None:
            explanation_parts.append(
                f"Options {options_score:.1f}"
            )

        results.append(
            {
                "trade_date": str(
                    group["trade_date"].max()
                ),
                "sector": sector,
                "stock_count": int(len(group)),
                "average_structure_score": round(
                    average_structure,
                    2,
                ),
                "average_5d_return": round(
                    average_5d,
                    2,
                ),
                "average_20d_return": round(
                    average_20d,
                    2,
                ),
                "bullish_breadth_percent": round(
                    bullish_breadth,
                    2,
                ),
                "bearish_breadth_percent": round(
                    bearish_breadth,
                    2,
                ),
                "breakout_percent": round(
                    breakout_percent,
                    2,
                ),
                "macro_score": macro_score,
                "news_score": news_score,
                "futures_score": futures_score,
                "options_score": options_score,
                "sector_rotation_score": rotation_score,
                "rotation_state": state,
                "trend_state": trend,
                "leader_symbol": str(
                    leader["nse_symbol"]
                ),
                "leader_score": round(
                    float(leader["structure_score"]),
                    2,
                ),
                "laggard_symbol": str(
                    laggard["nse_symbol"]
                ),
                "laggard_score": round(
                    float(laggard["structure_score"]),
                    2,
                ),
                "explanation": " | ".join(
                    explanation_parts
                ),
            }
        )

    results = sorted(
        results,
        key=lambda item: item["sector_rotation_score"],
        reverse=True,
    )

    return results, merged


def save_results(results: list[dict]) -> None:
    with connect() as connection:
        for result in results:
            connection.execute(
                """
                INSERT INTO sector_rotation_intelligence(
                    trade_date,
                    sector,
                    stock_count,
                    average_structure_score,
                    average_5d_return,
                    average_20d_return,
                    bullish_breadth_percent,
                    bearish_breadth_percent,
                    breakout_percent,
                    macro_score,
                    news_score,
                    futures_score,
                    options_score,
                    sector_rotation_score,
                    rotation_state,
                    trend_state,
                    leader_symbol,
                    leader_score,
                    laggard_symbol,
                    laggard_score,
                    explanation,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(trade_date, sector)
                DO UPDATE SET
                    stock_count = excluded.stock_count,
                    average_structure_score =
                        excluded.average_structure_score,
                    average_5d_return =
                        excluded.average_5d_return,
                    average_20d_return =
                        excluded.average_20d_return,
                    bullish_breadth_percent =
                        excluded.bullish_breadth_percent,
                    bearish_breadth_percent =
                        excluded.bearish_breadth_percent,
                    breakout_percent =
                        excluded.breakout_percent,
                    macro_score = excluded.macro_score,
                    news_score = excluded.news_score,
                    futures_score = excluded.futures_score,
                    options_score = excluded.options_score,
                    sector_rotation_score =
                        excluded.sector_rotation_score,
                    rotation_state = excluded.rotation_state,
                    trend_state = excluded.trend_state,
                    leader_symbol = excluded.leader_symbol,
                    leader_score = excluded.leader_score,
                    laggard_symbol = excluded.laggard_symbol,
                    laggard_score = excluded.laggard_score,
                    explanation = excluded.explanation,
                    created_at = excluded.created_at
                """,
                (
                    result["trade_date"],
                    result["sector"],
                    result["stock_count"],
                    result["average_structure_score"],
                    result["average_5d_return"],
                    result["average_20d_return"],
                    result["bullish_breadth_percent"],
                    result["bearish_breadth_percent"],
                    result["breakout_percent"],
                    result["macro_score"],
                    result["news_score"],
                    result["futures_score"],
                    result["options_score"],
                    result["sector_rotation_score"],
                    result["rotation_state"],
                    result["trend_state"],
                    result["leader_symbol"],
                    result["leader_score"],
                    result["laggard_symbol"],
                    result["laggard_score"],
                    result["explanation"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()


def run_engine(
    minimum_stocks: int,
) -> tuple[list[dict], pd.DataFrame]:
    setup_schema()

    run_id = start_run(
        "aqsd_sector_rotation",
        (
            "Running sector rotation with "
            f"minimum {minimum_stocks} stocks"
        ),
    )

    try:
        results, stock_frame = build_sector_results(
            minimum_stocks
        )

        save_results(results)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=len(results),
            errors_count=0,
            message=f"Sectors ranked={len(results)}",
        )

        return results, stock_frame

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


# ============================================================
# REPORTING
# ============================================================

def latest_results() -> pd.DataFrame:
    setup_schema()

    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM sector_rotation_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM sector_rotation_intelligence
            )
            ORDER BY sector_rotation_score DESC, sector
            """,
            connection,
        )


def write_report(
    results: list[dict] | None = None,
    stock_frame: pd.DataFrame | None = None,
    top_stocks: int = DEFAULT_TOP_STOCKS,
) -> None:
    if results is None:
        frame = latest_results()
        results = frame.to_dict("records")

    if stock_frame is None:
        stock_frame = latest_structure_frame()

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Sector Rotation Intelligence" in wb.sheetnames:
        del wb["Sector Rotation Intelligence"]

    ws = wb.create_sheet(
        "Sector Rotation Intelligence",
        1,
    )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:V2")
    ws["A1"] = "AQSD PROFESSIONAL - SECTOR ROTATION INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A4"] = "Sectors Ranked"
    ws["B4"] = len(results)
    ws["D4"] = "Updated"
    ws["E4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    for ref in ("A4", "D4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    headers = [
        "Rank",
        "Sector",
        "Stocks",
        "Structure Score",
        "5D Return %",
        "20D Return %",
        "Bullish Breadth %",
        "Bearish Breadth %",
        "Breakout %",
        "Macro Score",
        "News Score",
        "Futures Score",
        "Options Score",
        "Rotation Score",
        "Rotation State",
        "Trend State",
        "Leader",
        "Leader Score",
        "Laggard",
        "Laggard Score",
        "Explanation",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, result in enumerate(
        results,
        start=8,
    ):
        values = [
            row_no - 7,
            result["sector"],
            result["stock_count"],
            result["average_structure_score"],
            result["average_5d_return"],
            result["average_20d_return"],
            result["bullish_breadth_percent"],
            result["bearish_breadth_percent"],
            result["breakout_percent"],
            result["macro_score"],
            result["news_score"],
            result["futures_score"],
            result["options_score"],
            result["sector_rotation_score"],
            result["rotation_state"],
            result["trend_state"],
            result["leader_symbol"],
            result["leader_score"],
            result["laggard_symbol"],
            result["laggard_score"],
            result["explanation"],
        ]

        for col, value in enumerate(values, start=1):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(bottom=THIN)

        score = float(
            result["sector_rotation_score"]
        )

        ws.cell(row_no, 14).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )

        state = str(result["rotation_state"])

        ws.cell(row_no, 15).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if state in {
                    "STRONG INFLOW",
                    "INFLOW",
                    "EARLY ROTATION",
                }
                else RED
                if state in {
                    "STRONG OUTFLOW",
                    "OUTFLOW",
                    "WEAKENING LEADERSHIP",
                }
                else GREY
            ),
        )

    detail_start = max(
        12,
        10 + len(results),
    )

    ws.cell(
        detail_start,
        1,
        "TOP STOCKS BY SECTOR",
    ).font = Font(
        size=14,
        bold=True,
        color=WHITE,
    )
    ws.cell(
        detail_start,
        1,
    ).fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    detail_headers = [
        "Sector",
        "Rank",
        "Symbol",
        "Structure Score",
        "Directional Bias",
        "Market Structure",
        "BOS",
        "CHOCH",
        "ADX",
    ]

    for col, heading in enumerate(
        detail_headers,
        start=1,
    ):
        cell = ws.cell(
            detail_start + 2,
            col,
            heading,
        )
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )

    current_row = detail_start + 3

    if not stock_frame.empty:
        ranked_sectors = [
            result["sector"]
            for result in results
        ]

        for sector in ranked_sectors:
            group = stock_frame[
                stock_frame["sector"] == sector
            ].sort_values(
                "structure_score",
                ascending=False,
            ).head(top_stocks)

            for rank, (_, stock) in enumerate(
                group.iterrows(),
                start=1,
            ):
                values = [
                    sector,
                    rank,
                    stock["nse_symbol"],
                    stock["structure_score"],
                    stock["directional_bias"],
                    stock["market_structure"],
                    stock["bos_signal"],
                    stock["choch_signal"],
                    stock["adx_14"],
                ]

                for col, value in enumerate(
                    values,
                    start=1,
                ):
                    ws.cell(
                        current_row,
                        col,
                        value,
                    ).border = Border(
                        bottom=THIN
                    )

                current_row += 1

    widths = {
        "A": 8,
        "B": 22,
        "C": 10,
        "D": 16,
        "E": 13,
        "F": 13,
        "G": 17,
        "H": 17,
        "I": 13,
        "J": 13,
        "K": 13,
        "L": 14,
        "M": 14,
        "N": 16,
        "O": 22,
        "P": 14,
        "Q": 16,
        "R": 14,
        "S": 16,
        "T": 14,
        "U": 70,
        "V": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


def show_sector(
    sector: str,
) -> None:
    structure = latest_structure_frame()

    if structure.empty:
        print("No price-structure records found.")
        return

    selected = structure[
        structure["sector"].str.upper()
        == sector.strip().upper()
    ].sort_values(
        "structure_score",
        ascending=False,
    )

    if selected.empty:
        print(f"No records found for sector: {sector}")
        return

    columns = [
        "nse_symbol",
        "structure_score",
        "directional_bias",
        "market_structure",
        "bos_signal",
        "choch_signal",
        "adx_14",
    ]

    print(
        selected[columns].to_string(
            index=False
        )
    )


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT sector) AS sectors,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM sector_rotation_intelligence
            """
        ).fetchone()

    print("\nAQSD SECTOR ROTATION STATUS")
    print("=" * 72)
    print(
        f"Stored records:   "
        f"{row['total'] or 0}"
    )
    print(
        f"Sectors covered:  "
        f"{row['sectors'] or 0}"
    )
    print(
        f"First date:       "
        f"{row['first_date'] or 'No data'}"
    )
    print(
        f"Latest date:      "
        f"{row['latest_date'] or 'No data'}"
    )
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "AQSD Sector Rotation "
            "Intelligence Engine."
        )
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Run sector rotation analysis.",
    )

    parser.add_argument(
        "--minimum-stocks",
        type=int,
        default=DEFAULT_MINIMUM_STOCKS,
        help=(
            "Minimum stocks required "
            "to rank a sector."
        ),
    )

    parser.add_argument(
        "--top-stocks",
        type=int,
        default=DEFAULT_TOP_STOCKS,
        help=(
            "Top stocks per sector "
            "in Excel report."
        ),
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help=(
            "Rebuild report from "
            "stored results."
        ),
    )

    parser.add_argument(
        "--sector",
        metavar="SECTOR",
        help=(
            "Show latest stock rankings "
            "inside one sector."
        ),
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show database status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.sector:
        show_sector(args.sector)
        return

    if args.run:
        results, stock_frame = run_engine(
            args.minimum_stocks
        )

        write_report(
            results,
            stock_frame,
            args.top_stocks,
        )

        print("\nAQSD SECTOR ROTATION")
        print("=" * 78)
        print(
            f"Sectors ranked: "
            f"{len(results)}"
        )

        if results:
            print(
                f"Strongest sector: "
                f"{results[0]['sector']} "
                f"({results[0]['sector_rotation_score']})"
            )

            print(
                f"Weakest sector: "
                f"{results[-1]['sector']} "
                f"({results[-1]['sector_rotation_score']})"
            )

        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report(
            top_stocks=args.top_stocks
        )
        print(
            f"Report rebuilt:\n"
            f"{DASHBOARD}"
        )
        return

    show_status()


if __name__ == "__main__":
    main()
