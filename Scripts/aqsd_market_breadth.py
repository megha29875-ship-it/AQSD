
"""
AQSD Market Intelligence
Module: Market Breadth Intelligence Engine
Version: 1.0

Purpose
-------
Measures market participation using cached daily prices in aqsd_core.db.

Breadth metrics
---------------
- Advances / Declines / Unchanged
- Advance-Decline Ratio
- Advance-Decline Net
- Percentage above 20 DMA
- Percentage above 50 DMA
- Percentage above 200 DMA
- New 20-day highs
- New 20-day lows
- New 52-week highs
- New 52-week lows
- Sector breadth
- Market Breadth Score from 0 to 100
- Breadth Regime
- Excel report: Market Breadth Intelligence

Commands
--------
python aqsd_market_breadth.py --run
python aqsd_market_breadth.py --status
python aqsd_market_breadth.py --report
python aqsd_market_breadth.py --sector BANK
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


# ============================================================
# DATABASE SCHEMA
# ============================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS market_breadth_intelligence (
    breadth_id INTEGER PRIMARY KEY AUTOINCREMENT,
    trade_date TEXT NOT NULL,
    scope_type TEXT NOT NULL,
    scope_name TEXT NOT NULL,
    stock_count INTEGER,
    advances INTEGER,
    declines INTEGER,
    unchanged INTEGER,
    advance_decline_ratio REAL,
    advance_decline_net INTEGER,
    above_20dma_percent REAL,
    above_50dma_percent REAL,
    above_200dma_percent REAL,
    new_20d_highs INTEGER,
    new_20d_lows INTEGER,
    new_52w_highs INTEGER,
    new_52w_lows INTEGER,
    breadth_score REAL,
    breadth_regime TEXT,
    explanation TEXT,
    created_at TEXT NOT NULL,
    UNIQUE(trade_date, scope_type, scope_name)
);

CREATE INDEX IF NOT EXISTS idx_market_breadth_date
ON market_breadth_intelligence(trade_date);

CREATE INDEX IF NOT EXISTS idx_market_breadth_scope
ON market_breadth_intelligence(scope_type, scope_name);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


# ============================================================
# DATA LOADING
# ============================================================

def load_recent_prices() -> pd.DataFrame:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                p.symbol_id,
                s.nse_symbol,
                COALESCE(NULLIF(TRIM(s.sector), ''), 'Unmapped') AS sector,
                p.trade_date,
                p.close
            FROM daily_prices p
            JOIN symbols s
                ON s.symbol_id = p.symbol_id
            WHERE s.active = 1
              AND p.trade_date >= (
                  SELECT DATE(MAX(trade_date), '-430 day')
                  FROM daily_prices
              )
            ORDER BY p.symbol_id, p.trade_date
            """,
            connection,
        )

    if frame.empty:
        return frame

    frame["trade_date"] = pd.to_datetime(frame["trade_date"])
    return frame


def build_symbol_snapshot(frame: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for symbol_id, group in frame.groupby("symbol_id"):
        group = group.sort_values("trade_date")
        close = group["close"].dropna()

        if len(close) < 2:
            continue

        latest = float(close.iloc[-1])
        previous = float(close.iloc[-2])

        dma20 = float(close.tail(20).mean()) if len(close) >= 20 else None
        dma50 = float(close.tail(50).mean()) if len(close) >= 50 else None
        dma200 = float(close.tail(200).mean()) if len(close) >= 200 else None

        high20 = float(close.tail(20).max()) if len(close) >= 20 else None
        low20 = float(close.tail(20).min()) if len(close) >= 20 else None
        high252 = float(close.tail(252).max()) if len(close) >= 252 else None
        low252 = float(close.tail(252).min()) if len(close) >= 252 else None

        rows.append(
            {
                "symbol_id": int(symbol_id),
                "nse_symbol": group["nse_symbol"].iloc[-1],
                "sector": group["sector"].iloc[-1],
                "trade_date": group["trade_date"].iloc[-1].date().isoformat(),
                "latest_close": latest,
                "previous_close": previous,
                "change": latest - previous,
                "above_20dma": (
                    latest > dma20 if dma20 is not None else None
                ),
                "above_50dma": (
                    latest > dma50 if dma50 is not None else None
                ),
                "above_200dma": (
                    latest > dma200 if dma200 is not None else None
                ),
                "new_20d_high": (
                    latest >= high20 if high20 is not None else False
                ),
                "new_20d_low": (
                    latest <= low20 if low20 is not None else False
                ),
                "new_52w_high": (
                    latest >= high252 if high252 is not None else False
                ),
                "new_52w_low": (
                    latest <= low252 if low252 is not None else False
                ),
            }
        )

    return pd.DataFrame(rows)


# ============================================================
# SCORING
# ============================================================

def safe_percent(series: pd.Series) -> float:
    valid = series.dropna()

    if valid.empty:
        return 0.0

    return round(float(valid.mean() * 100), 2)


def breadth_regime(score: float) -> str:
    if score >= 80:
        return "VERY STRONG"
    if score >= 65:
        return "STRONG"
    if score >= 55:
        return "POSITIVE"
    if score <= 20:
        return "VERY WEAK"
    if score <= 35:
        return "WEAK"
    if score <= 45:
        return "NEGATIVE"
    return "NEUTRAL"


def calculate_breadth_score(
    ad_ratio: float,
    above20: float,
    above50: float,
    above200: float,
    new20_highs: int,
    new20_lows: int,
    new52_highs: int,
    new52_lows: int,
    stock_count: int,
) -> float:
    if stock_count <= 0:
        return 50.0

    ad_component = max(
        0.0,
        min(100.0, ad_ratio / 2 * 100),
    )

    new20_net = (
        (new20_highs - new20_lows)
        / stock_count
        * 100
    )

    new52_net = (
        (new52_highs - new52_lows)
        / stock_count
        * 100
    )

    score = (
        ad_component * 0.25
        + above20 * 0.20
        + above50 * 0.20
        + above200 * 0.20
        + max(0, min(100, 50 + new20_net * 2)) * 0.10
        + max(0, min(100, 50 + new52_net * 3)) * 0.05
    )

    return round(max(0, min(100, score)), 2)


def summarise_scope(
    group: pd.DataFrame,
    scope_type: str,
    scope_name: str,
) -> dict:
    stock_count = len(group)
    advances = int((group["change"] > 0).sum())
    declines = int((group["change"] < 0).sum())
    unchanged = int((group["change"] == 0).sum())

    ad_ratio = (
        advances / declines
        if declines > 0
        else float(advances)
    )

    above20 = safe_percent(group["above_20dma"])
    above50 = safe_percent(group["above_50dma"])
    above200 = safe_percent(group["above_200dma"])

    new20_highs = int(group["new_20d_high"].sum())
    new20_lows = int(group["new_20d_low"].sum())
    new52_highs = int(group["new_52w_high"].sum())
    new52_lows = int(group["new_52w_low"].sum())

    score = calculate_breadth_score(
        ad_ratio,
        above20,
        above50,
        above200,
        new20_highs,
        new20_lows,
        new52_highs,
        new52_lows,
        stock_count,
    )

    explanation = " | ".join(
        [
            f"Advances {advances}",
            f"Declines {declines}",
            f"A/D ratio {ad_ratio:.2f}",
            f"Above 20DMA {above20:.1f}%",
            f"Above 50DMA {above50:.1f}%",
            f"Above 200DMA {above200:.1f}%",
            f"20D highs/lows {new20_highs}/{new20_lows}",
            f"52W highs/lows {new52_highs}/{new52_lows}",
        ]
    )

    return {
        "trade_date": str(group["trade_date"].max()),
        "scope_type": scope_type,
        "scope_name": scope_name,
        "stock_count": stock_count,
        "advances": advances,
        "declines": declines,
        "unchanged": unchanged,
        "advance_decline_ratio": round(ad_ratio, 2),
        "advance_decline_net": advances - declines,
        "above_20dma_percent": above20,
        "above_50dma_percent": above50,
        "above_200dma_percent": above200,
        "new_20d_highs": new20_highs,
        "new_20d_lows": new20_lows,
        "new_52w_highs": new52_highs,
        "new_52w_lows": new52_lows,
        "breadth_score": score,
        "breadth_regime": breadth_regime(score),
        "explanation": explanation,
    }


# ============================================================
# ENGINE
# ============================================================

def build_results() -> tuple[list[dict], pd.DataFrame]:
    prices = load_recent_prices()

    if prices.empty:
        raise RuntimeError(
            "No cached prices found. "
            "Run aqsd_price_cache.py first."
        )

    snapshot = build_symbol_snapshot(prices)

    if snapshot.empty:
        raise RuntimeError(
            "Could not build market breadth snapshot."
        )

    results = [
        summarise_scope(
            snapshot,
            "MARKET",
            "ALL ACTIVE STOCKS",
        )
    ]

    for sector, group in snapshot.groupby("sector"):
        results.append(
            summarise_scope(
                group,
                "SECTOR",
                str(sector),
            )
        )

    results = sorted(
        results,
        key=lambda item: (
            0 if item["scope_type"] == "MARKET" else 1,
            -item["breadth_score"],
        ),
    )

    return results, snapshot


def save_results(results: list[dict]) -> None:
    with connect() as connection:
        for result in results:
            connection.execute(
                """
                INSERT INTO market_breadth_intelligence(
                    trade_date,
                    scope_type,
                    scope_name,
                    stock_count,
                    advances,
                    declines,
                    unchanged,
                    advance_decline_ratio,
                    advance_decline_net,
                    above_20dma_percent,
                    above_50dma_percent,
                    above_200dma_percent,
                    new_20d_highs,
                    new_20d_lows,
                    new_52w_highs,
                    new_52w_lows,
                    breadth_score,
                    breadth_regime,
                    explanation,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(trade_date, scope_type, scope_name)
                DO UPDATE SET
                    stock_count = excluded.stock_count,
                    advances = excluded.advances,
                    declines = excluded.declines,
                    unchanged = excluded.unchanged,
                    advance_decline_ratio =
                        excluded.advance_decline_ratio,
                    advance_decline_net =
                        excluded.advance_decline_net,
                    above_20dma_percent =
                        excluded.above_20dma_percent,
                    above_50dma_percent =
                        excluded.above_50dma_percent,
                    above_200dma_percent =
                        excluded.above_200dma_percent,
                    new_20d_highs = excluded.new_20d_highs,
                    new_20d_lows = excluded.new_20d_lows,
                    new_52w_highs = excluded.new_52w_highs,
                    new_52w_lows = excluded.new_52w_lows,
                    breadth_score = excluded.breadth_score,
                    breadth_regime = excluded.breadth_regime,
                    explanation = excluded.explanation,
                    created_at = excluded.created_at
                """,
                (
                    result["trade_date"],
                    result["scope_type"],
                    result["scope_name"],
                    result["stock_count"],
                    result["advances"],
                    result["declines"],
                    result["unchanged"],
                    result["advance_decline_ratio"],
                    result["advance_decline_net"],
                    result["above_20dma_percent"],
                    result["above_50dma_percent"],
                    result["above_200dma_percent"],
                    result["new_20d_highs"],
                    result["new_20d_lows"],
                    result["new_52w_highs"],
                    result["new_52w_lows"],
                    result["breadth_score"],
                    result["breadth_regime"],
                    result["explanation"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()


def run_engine() -> tuple[list[dict], pd.DataFrame]:
    setup_schema()

    run_id = start_run(
        "aqsd_market_breadth",
        "Running market breadth intelligence",
    )

    try:
        results, snapshot = build_results()
        save_results(results)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=len(results),
            errors_count=0,
            message=f"Scopes stored={len(results)}",
        )

        return results, snapshot

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


# ============================================================
# REPORTING
# ============================================================

def latest_results() -> pd.DataFrame:
    setup_schema()

    with connect() as connection:
        return pd.read_sql_query(
            """
            SELECT *
            FROM market_breadth_intelligence
            WHERE trade_date = (
                SELECT MAX(trade_date)
                FROM market_breadth_intelligence
            )
            ORDER BY
                CASE scope_type
                    WHEN 'MARKET' THEN 0
                    ELSE 1
                END,
                breadth_score DESC,
                scope_name
            """,
            connection,
        )


def write_report(
    results: list[dict] | None = None,
    snapshot: pd.DataFrame | None = None,
) -> None:
    if results is None:
        results = latest_results().to_dict("records")

    if snapshot is None:
        snapshot = build_symbol_snapshot(
            load_recent_prices()
        )

    if DASHBOARD.exists():
        wb = load_workbook(DASHBOARD)
    else:
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "Market Breadth Intelligence" in wb.sheetnames:
        del wb["Market Breadth Intelligence"]

    ws = wb.create_sheet(
        "Market Breadth Intelligence",
        1,
    )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:T2")
    ws["A1"] = "AQSD PROFESSIONAL - MARKET BREADTH INTELLIGENCE"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    market_row = next(
        (
            item
            for item in results
            if item["scope_type"] == "MARKET"
        ),
        None,
    )

    ws["A4"] = "Market Breadth Score"
    ws["B4"] = (
        market_row["breadth_score"]
        if market_row
        else None
    )
    ws["D4"] = "Market Regime"
    ws["E4"] = (
        market_row["breadth_regime"]
        if market_row
        else "NO DATA"
    )
    ws["G4"] = "Updated"
    ws["H4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    for ref in ("A4", "D4", "G4"):
        ws[ref].font = Font(bold=True)
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    if market_row:
        score = float(market_row["breadth_score"])

        ws["B4"].fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )
        ws["B4"].font = Font(bold=True)

    headers = [
        "Scope",
        "Name",
        "Stocks",
        "Advances",
        "Declines",
        "Unchanged",
        "A/D Ratio",
        "A-D Net",
        "Above 20DMA %",
        "Above 50DMA %",
        "Above 200DMA %",
        "20D Highs",
        "20D Lows",
        "52W Highs",
        "52W Lows",
        "Breadth Score",
        "Breadth Regime",
        "Explanation",
    ]

    for col, heading in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN)

    for row_no, result in enumerate(
        results,
        start=8,
    ):
        values = [
            result["scope_type"],
            result["scope_name"],
            result["stock_count"],
            result["advances"],
            result["declines"],
            result["unchanged"],
            result["advance_decline_ratio"],
            result["advance_decline_net"],
            result["above_20dma_percent"],
            result["above_50dma_percent"],
            result["above_200dma_percent"],
            result["new_20d_highs"],
            result["new_20d_lows"],
            result["new_52w_highs"],
            result["new_52w_lows"],
            result["breadth_score"],
            result["breadth_regime"],
            result["explanation"],
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(bottom=THIN)

        score = float(result["breadth_score"])

        ws.cell(row_no, 16).fill = PatternFill(
            "solid",
            fgColor=(
                GREEN
                if score >= 60
                else RED
                if score <= 40
                else YELLOW
            ),
        )

    detail_start = max(
        12,
        10 + len(results),
    )

    ws.cell(
        detail_start,
        1,
        "STOCK-LEVEL BREADTH SNAPSHOT",
    ).font = Font(
        size=14,
        bold=True,
        color=WHITE,
    )
    ws.cell(
        detail_start,
        1,
    ).fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    detail_headers = [
        "Symbol",
        "Sector",
        "Latest Close",
        "Change",
        "Above 20DMA",
        "Above 50DMA",
        "Above 200DMA",
        "New 20D High",
        "New 20D Low",
        "New 52W High",
        "New 52W Low",
    ]

    for col, heading in enumerate(
        detail_headers,
        start=1,
    ):
        cell = ws.cell(
            detail_start + 2,
            col,
            heading,
        )
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )

    if snapshot is not None and not snapshot.empty:
        for row_no, (_, row) in enumerate(
            snapshot.sort_values(
                ["sector", "nse_symbol"]
            ).iterrows(),
            start=detail_start + 3,
        ):
            values = [
                row["nse_symbol"],
                row["sector"],
                row["latest_close"],
                row["change"],
                row["above_20dma"],
                row["above_50dma"],
                row["above_200dma"],
                row["new_20d_high"],
                row["new_20d_low"],
                row["new_52w_high"],
                row["new_52w_low"],
            ]

            for col, value in enumerate(
                values,
                start=1,
            ):
                ws.cell(
                    row_no,
                    col,
                    value,
                ).border = Border(
                    bottom=THIN
                )

    widths = {
        "A": 14,
        "B": 24,
        "C": 10,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 16,
        "J": 16,
        "K": 17,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 15,
        "Q": 18,
        "R": 70,
        "S": 14,
        "T": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wb.save(DASHBOARD)


def show_sector(
    sector: str,
) -> None:
    snapshot = build_symbol_snapshot(
        load_recent_prices()
    )

    if snapshot.empty:
        print("No breadth snapshot available.")
        return

    selected = snapshot[
        snapshot["sector"].str.upper()
        == sector.strip().upper()
    ].copy()

    if selected.empty:
        print(f"No records found for sector: {sector}")
        return

    columns = [
        "nse_symbol",
        "latest_close",
        "change",
        "above_20dma",
        "above_50dma",
        "above_200dma",
        "new_20d_high",
        "new_20d_low",
        "new_52w_high",
        "new_52w_low",
    ]

    print(
        selected[columns].to_string(
            index=False
        )
    )


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT scope_name) AS scopes,
                MIN(trade_date) AS first_date,
                MAX(trade_date) AS latest_date
            FROM market_breadth_intelligence
            """
        ).fetchone()

    print("\nAQSD MARKET BREADTH STATUS")
    print("=" * 72)
    print(f"Stored records:   {row['total'] or 0}")
    print(f"Scopes covered:   {row['scopes'] or 0}")
    print(f"First date:       {row['first_date'] or 'No data'}")
    print(f"Latest date:      {row['latest_date'] or 'No data'}")
    print("=" * 72)


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="AQSD Market Breadth Intelligence Engine."
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Run market breadth analysis.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild Excel report from stored data.",
    )

    parser.add_argument(
        "--sector",
        metavar="SECTOR",
        help="Show stock breadth for one sector.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show market breadth database status.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.sector:
        show_sector(args.sector)
        return

    if args.run:
        results, snapshot = run_engine()
        write_report(results, snapshot)

        market = next(
            item
            for item in results
            if item["scope_type"] == "MARKET"
        )

        print("\nAQSD MARKET BREADTH")
        print("=" * 72)
        print(
            f"Market Breadth Score: "
            f"{market['breadth_score']}"
        )
        print(
            f"Market Regime:       "
            f"{market['breadth_regime']}"
        )
        print(
            f"Advances / Declines: "
            f"{market['advances']} / "
            f"{market['declines']}"
        )
        print(f"Report: {DASHBOARD}")
        return

    if args.report:
        write_report()
        print(f"Report rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
