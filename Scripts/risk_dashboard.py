
"""
AQSD Professional
Module: Risk Dashboard Integration
Version: 1.0
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from risk_manager import calculate_position_size


# ============================================================
# USER SETTINGS
# ============================================================

TRADING_CAPITAL = 200000
RISK_PERCENT = 1.0
MAX_CAPITAL_PERCENT = 100.0


# ============================================================
# PATHS
# ============================================================

BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"


# ============================================================
# COLORS
# ============================================================

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"


def header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def get_best_trade(wb) -> dict:
    if "Option Buying" not in wb.sheetnames:
        raise RuntimeError("Option Buying sheet not found.")

    ws = wb["Option Buying"]
    headers = header_map(ws)

    required = [
        "Symbol",
        "Entry",
        "Stop Loss",
        "Target 1",
        "Target 2",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            "Missing columns in Option Buying: "
            + ", ".join(missing)
        )

    if ws.max_row < 2:
        raise RuntimeError("Option Buying sheet has no trades.")

    row = 2

    return {
        "Symbol": ws.cell(row, headers["Symbol"]).value,
        "Entry": float(ws.cell(row, headers["Entry"]).value),
        "Stop Loss": float(ws.cell(row, headers["Stop Loss"]).value),
        "Target 1": float(ws.cell(row, headers["Target 1"]).value),
        "Target 2": float(ws.cell(row, headers["Target 2"]).value),
    }


def write_risk_plan(wb, trade: dict) -> None:
    if "HOME" not in wb.sheetnames:
        raise RuntimeError(
            "HOME sheet not found. Run format_dashboard.py first."
        )

    ws = wb["HOME"]

    plan = calculate_position_size(
        capital=TRADING_CAPITAL,
        risk_percent=RISK_PERCENT,
        entry=trade["Entry"],
        stop_loss=trade["Stop Loss"],
        target_1=trade["Target 1"],
        target_2=trade["Target 2"],
        max_capital_percent=MAX_CAPITAL_PERCENT,
    )

    # Risk Plan title
    ws.merge_cells("A21:C21")
    ws["A21"] = "RISK & POSITION PLAN"
    ws["A21"].font = Font(
        bold=True,
        color=WHITE,
    )
    ws["A21"].fill = PatternFill(
        fill_type="solid",
        fgColor=NAVY,
    )
    ws["A21"].alignment = Alignment(
        horizontal="center",
    )

    rows = [
        ("Trading Capital", plan.capital),
        ("Risk per Trade", f"{plan.risk_percent:.2f}%"),
        ("Maximum Risk", plan.risk_amount),
        ("Quantity", plan.quantity),
        ("Capital Required", plan.capital_required),
        ("Maximum Loss", plan.max_loss),
        ("RR Target 1", f"1:{plan.rr_1:.2f}"),
        ("RR Target 2", f"1:{plan.rr_2:.2f}"),
        ("Expected Reward T1", plan.reward_1),
        ("Expected Reward T2", plan.reward_2),
    ]

    for row_no, (label, value) in enumerate(
        rows,
        start=22,
    ):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value

        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill(
            fill_type="solid",
            fgColor=BLUE,
        )

    currency_rows = [22, 24, 26, 27, 30, 31]

    for row_no in currency_rows:
        ws[f"B{row_no}"].number_format = '₹#,##0.00'

    ws["B25"].number_format = "0"

    if plan.quantity > 0:
        ws["B25"].fill = PatternFill(
            fill_type="solid",
            fgColor=GREEN,
        )
    else:
        ws["B25"].fill = PatternFill(
            fill_type="solid",
            fgColor=YELLOW,
        )

    # Add summary to Today's Best Trade card
    ws.merge_cells("H13:N13")
    ws["H13"] = (
        f"Quantity: {plan.quantity}   |   "
        f"Max Loss: ₹{plan.max_loss:,.2f}   |   "
        f"Capital Used: ₹{plan.capital_required:,.2f}"
    )
    ws["H13"].font = Font(bold=True)
    ws["H13"].fill = PatternFill(
        fill_type="solid",
        fgColor=GREEN,
    )
    ws["H13"].alignment = Alignment(
        horizontal="center",
    )

    ws.merge_cells("H14:N14")
    ws["H14"] = (
        f"RR T1: 1:{plan.rr_1:.2f}   |   "
        f"RR T2: 1:{plan.rr_2:.2f}"
    )
    ws["H14"].font = Font(bold=True)
    ws["H14"].fill = PatternFill(
        fill_type="solid",
        fgColor=YELLOW,
    )
    ws["H14"].alignment = Alignment(
        horizontal="center",
    )


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    trade = get_best_trade(wb)
    write_risk_plan(wb, trade)

    wb.save(DASHBOARD)

    print("Risk plan added to Dashboard.xlsx")
    print(f"Best Trade: {trade['Symbol']}")
    print(f"Capital: {TRADING_CAPITAL}")
    print(f"Risk per Trade: {RISK_PERCENT}%")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
