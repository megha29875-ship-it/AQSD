
"""
AQSD Professional
Module: Equity Curve
Creates an Equity Curve sheet from the Trade Journal.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference

BASE = Path(__file__).resolve().parent.parent
FILE = BASE/"Output"/"Dashboard.xlsx"

wb = load_workbook(FILE)

if "Trade Journal" not in wb.sheetnames:
    raise RuntimeError("Trade Journal sheet not found.")

tws = wb["Trade Journal"]

if "Equity Curve" in wb.sheetnames:
    del wb["Equity Curve"]

ws = wb.create_sheet("Equity Curve")

ws["A1"]="AQSD PROFESSIONAL - EQUITY CURVE"
ws["A1"].font=Font(bold=True,size=18,color="FFFFFF")
ws["A1"].fill=PatternFill("solid",fgColor="1F4E78")

headers={}
for c in tws[10]:
    if c.value:
        headers[str(c.value)] = c.column

ws["A3"]="Trade"
ws["B3"]="Equity"

equity=0
row=4
trade=1

if "P/L" in headers:
    for r in range(11,tws.max_row+1):
        v=tws.cell(r,headers["P/L"]).value
        if isinstance(v,(int,float)):
            equity += float(v)
            ws.cell(row,1).value=trade
            ws.cell(row,2).value=equity
            trade+=1
            row+=1

if row>4:
    chart=LineChart()
    data=Reference(ws,min_col=2,min_row=3,max_row=row-1)
    cats=Reference(ws,min_col=1,min_row=4,max_row=row-1)
    chart.add_data(data,titles_from_data=True)
    chart.set_categories(cats)
    chart.title="Equity Curve"
    chart.height=10
    chart.width=20
    ws.add_chart(chart,"D3")

wb.save(FILE)

print("Equity Curve created.")
print(FILE)
