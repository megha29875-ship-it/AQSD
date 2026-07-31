"""
AQSD
NSE Market Breadth Data Builder

Module : MBI-002
Version: 1.0.0
Author : AQSD

Description
-----------
Downloads official NSE Cash Market UDiFF bhavcopy files and builds
the market breadth snapshot required by:

    Scripts.aqsd_intelligence.market_breadth_engine

Official data source
--------------------
NSE CM-UDiFF Common Bhavcopy Final ZIP files.

Output
------
Data/Market_Breadth/market_breadth_snapshot.xlsx

Output columns
--------------
Symbol
Close
Previous_Close
EMA20
EMA50
EMA200
High_52W
Low_52W
Sector
Market_Cap_Category
Volume
Average_Volume_20
Is_FnO

Important
---------
- Only NSE EQ-series equity securities are retained.
- No Yahoo Finance or Fyers data is used.
- Missing weekends and exchange holidays are skipped automatically.
- Existing downloaded bhavcopy files are reused.
- The latest available trading date on or before the requested date
  becomes the snapshot date.
"""

from __future__ import annotations

import argparse
import io
import re
import time
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Final

import pandas as pd
import requests


# ==========================================================
# MODULE SETTINGS
# ==========================================================

MODULE_ID: Final[str] = "MBI-002"
MODULE_VERSION: Final[str] = "1.0.0"

BASE_DIR: Final[Path] = Path(__file__).resolve().parents[2]

RAW_DIRECTORY: Final[Path] = (
    BASE_DIR
    / "Data"
    / "Raw"
    / "NSE"
    / "Cash_Market"
    / "UDiFF"
)

OUTPUT_DIRECTORY: Final[Path] = (
    BASE_DIR
    / "Data"
    / "Market_Breadth"
)

OUTPUT_FILE: Final[Path] = (
    OUTPUT_DIRECTORY
    / "market_breadth_snapshot.xlsx"
)

HISTORY_FILE: Final[Path] = (
    OUTPUT_DIRECTORY
    / "market_breadth_price_history.csv"
)

MANIFEST_FILE: Final[Path] = (
    OUTPUT_DIRECTORY
    / "market_breadth_builder_manifest.csv"
)

DEFAULT_LOOKBACK_CALENDAR_DAYS: Final[int] = 430
MINIMUM_REQUIRED_SESSIONS: Final[int] = 200

REQUEST_TIMEOUT_SECONDS: Final[int] = 30
REQUEST_RETRIES: Final[int] = 3
REQUEST_DELAY_SECONDS: Final[float] = 0.20

NSE_ARCHIVE_BASE_URL: Final[str] = (
    "https://nsearchives.nseindia.com/content/cm"
)

NSE_HOME_URL: Final[str] = "https://www.nseindia.com"

USER_AGENT: Final[str] = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)

EQ_SERIES: Final[str] = "EQ"


# ==========================================================
# RESULT MODEL
# ==========================================================

@dataclass(frozen=True)
class BreadthBuilderResult:
    """
    Final NSE breadth-builder result.
    """

    requested_date: date
    snapshot_date: date | None

    calendar_dates_checked: int
    trading_files_available: int
    downloaded_files: int
    cached_files: int
    unavailable_dates: int
    failed_dates: int

    historical_rows: int
    historical_symbols: int
    eligible_symbols: int
    output_rows: int

    history_file: Path
    output_file: Path
    manifest_file: Path

    status: str
    message: str


# ==========================================================
# DATE HELPERS
# ==========================================================

def parse_date(value: str) -> date:
    """
    Parse a YYYY-MM-DD date.
    """

    try:
        return datetime.strptime(
            value,
            "%Y-%m-%d",
        ).date()

    except ValueError as exc:
        raise ValueError(
            "Invalid date format. Use YYYY-MM-DD."
        ) from exc


def is_weekend(value: date) -> bool:
    """
    Return True for Saturday or Sunday.
    """

    return value.weekday() >= 5


def build_candidate_dates(
    *,
    requested_date: date,
    lookback_calendar_days: int,
) -> list[date]:
    """
    Build weekday dates in ascending order.
    """

    start_date = (
        requested_date
        - timedelta(days=lookback_calendar_days)
    )

    dates: list[date] = []

    current_date = start_date

    while current_date <= requested_date:
        if not is_weekend(current_date):
            dates.append(current_date)

        current_date += timedelta(days=1)

    return dates


# ==========================================================
# HTTP SESSION
# ==========================================================

def create_nse_session() -> requests.Session:
    """
    Create an HTTP session for NSE archive downloads.
    """

    session = requests.Session()

    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": (
                "text/html,application/xhtml+xml,application/xml,"
                "application/zip,text/csv,*/*"
            ),
            "Accept-Language": "en-IN,en;q=0.9",
            "Connection": "keep-alive",
            "Referer": NSE_HOME_URL,
        }
    )

    try:
        session.get(
            NSE_HOME_URL,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
    except requests.RequestException:
        # Archive downloads can still work without the homepage request.
        pass

    return session


# ==========================================================
# URL AND FILE HELPERS
# ==========================================================

def build_udiff_filename(
    trade_date: date,
) -> str:
    """
    Build official CM UDiFF ZIP filename.
    """

    return (
        "BhavCopy_NSE_CM_0_0_0_"
        f"{trade_date.strftime('%Y%m%d')}"
        "_F_0000.csv.zip"
    )


def build_udiff_url(
    trade_date: date,
) -> str:
    """
    Build NSE archive URL for one trade date.
    """

    return (
        f"{NSE_ARCHIVE_BASE_URL}/"
        f"{build_udiff_filename(trade_date)}"
    )


def local_zip_path(
    trade_date: date,
) -> Path:
    """
    Return local cache path for one UDiFF ZIP.
    """

    year_directory = (
        RAW_DIRECTORY
        / f"{trade_date.year:04d}"
        / f"{trade_date.month:02d}"
    )

    return (
        year_directory
        / build_udiff_filename(trade_date)
    )


def content_is_zip(
    content: bytes,
) -> bool:
    """
    Validate ZIP file signature.
    """

    return (
        len(content) >= 4
        and content[:4] == b"PK\x03\x04"
    )


# ==========================================================
# DOWNLOAD ENGINE
# ==========================================================

def download_udiff_bhavcopy(
    *,
    session: requests.Session,
    trade_date: date,
    force_download: bool,
) -> tuple[str, Path | None, str]:
    """
    Download or reuse one NSE CM UDiFF bhavcopy.

    Returns
    -------
    status:
        DOWNLOADED, CACHED, UNAVAILABLE or FAILED
    file_path:
        Local ZIP path when available
    message:
        Diagnostic message
    """

    destination = local_zip_path(
        trade_date
    )

    if (
        destination.exists()
        and destination.stat().st_size > 100
        and not force_download
    ):
        return (
            "CACHED",
            destination,
            "Existing local file reused.",
        )

    url = build_udiff_url(
        trade_date
    )

    last_error = ""

    for attempt in range(
        1,
        REQUEST_RETRIES + 1,
    ):
        try:
            response = session.get(
                url,
                timeout=REQUEST_TIMEOUT_SECONDS,
            )

        except requests.RequestException as exc:
            last_error = str(exc)

            if attempt < REQUEST_RETRIES:
                time.sleep(
                    attempt
                )

            continue

        if response.status_code == 404:
            return (
                "UNAVAILABLE",
                None,
                "No NSE bhavcopy was available for this date.",
            )

        if response.status_code != 200:
            last_error = (
                f"HTTP {response.status_code}"
            )

            if attempt < REQUEST_RETRIES:
                time.sleep(
                    attempt
                )

            continue

        if not content_is_zip(
            response.content
        ):
            last_error = (
                "Response was not a valid ZIP file."
            )

            if attempt < REQUEST_RETRIES:
                time.sleep(
                    attempt
                )

            continue

        destination.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        destination.write_bytes(
            response.content
        )

        return (
            "DOWNLOADED",
            destination,
            "Downloaded successfully.",
        )

    return (
        "FAILED",
        None,
        last_error or "Unknown download failure.",
    )


# ==========================================================
# COLUMN NORMALIZATION
# ==========================================================

def normalize_header(value: object) -> str:
    """
    Normalize an NSE column heading.
    """

    text = str(value).strip().upper()

    text = re.sub(
        r"[^A-Z0-9]+",
        "_",
        text,
    )

    text = re.sub(
        r"_+",
        "_",
        text,
    )

    return text.strip("_")


def identify_column(
    columns: list[str],
    aliases: tuple[str, ...],
) -> str | None:
    """
    Find the first matching column alias.
    """

    normalized_aliases = {
        normalize_header(alias)
        for alias in aliases
    }

    for column in columns:
        if normalize_header(column) in normalized_aliases:
            return column

    return None


# ==========================================================
# ZIP READING
# ==========================================================

def extract_csv_bytes(
    zip_path: Path,
) -> bytes:
    """
    Extract the first CSV file from an NSE ZIP.
    """

    with zipfile.ZipFile(
        zip_path,
        mode="r",
    ) as archive:
        csv_names = [
            name
            for name in archive.namelist()
            if name.lower().endswith(".csv")
        ]

        if not csv_names:
            raise RuntimeError(
                f"No CSV file found inside {zip_path.name}"
            )

        return archive.read(
            csv_names[0]
        )


def read_csv_flexibly(
    csv_bytes: bytes,
) -> pd.DataFrame:
    """
    Read NSE CSV data with common encoding fallbacks.
    """

    last_error: Exception | None = None

    for encoding in (
        "utf-8-sig",
        "utf-8",
        "latin-1",
    ):
        try:
            return pd.read_csv(
                io.BytesIO(csv_bytes),
                encoding=encoding,
                low_memory=False,
            )

        except Exception as exc:
            last_error = exc

    raise RuntimeError(
        f"Could not read NSE CSV: {last_error}"
    )


# ==========================================================
# BHAVCOPY PARSER
# ==========================================================

def parse_udiff_bhavcopy(
    *,
    zip_path: Path,
    expected_trade_date: date,
) -> pd.DataFrame:
    """
    Parse one NSE CM UDiFF bhavcopy into a standard format.
    """

    csv_bytes = extract_csv_bytes(
        zip_path
    )

    raw = read_csv_flexibly(
        csv_bytes
    )

    raw.columns = [
        str(column).strip()
        for column in raw.columns
    ]

    columns = list(
        raw.columns
    )

    symbol_column = identify_column(
        columns,
        (
            "TckrSymb",
            "Symbol",
            "SYMBOL",
            "Security Symbol",
        ),
    )

    series_column = identify_column(
        columns,
        (
            "SctySrs",
            "Series",
            "SERIES",
            "Security Series",
        ),
    )

    close_column = identify_column(
        columns,
        (
            "ClsPric",
            "Close",
            "CLOSE",
            "Closing Price",
        ),
    )

    previous_close_column = identify_column(
        columns,
        (
            "PrvsClsgPric",
            "Previous Close",
            "PREV_CLOSE",
            "PREVCLOSE",
            "Previous_Close",
        ),
    )

    volume_column = identify_column(
        columns,
        (
            "TtlTradgVol",
            "TOTTRDQTY",
            "Total Traded Quantity",
            "Volume",
        ),
    )

    trade_date_column = identify_column(
        columns,
        (
            "TradDt",
            "Trade Date",
            "TIMESTAMP",
            "Date",
        ),
    )

    security_type_column = identify_column(
        columns,
        (
            "FinInstrmTp",
            "Financial Instrument Type",
            "Instrument Type",
        ),
    )

    required_mapping = {
        "Symbol": symbol_column,
        "Series": series_column,
        "Close": close_column,
        "Volume": volume_column,
    }

    missing = [
        field
        for field, column
        in required_mapping.items()
        if column is None
    ]

    if missing:
        raise KeyError(
            "Could not identify required NSE columns: "
            + ", ".join(missing)
            + ". Available columns: "
            + ", ".join(columns)
        )

    parsed = pd.DataFrame(
        {
            "Symbol": raw[symbol_column],
            "Series": raw[series_column],
            "Close": raw[close_column],
            "Volume": raw[volume_column],
        }
    )

    if previous_close_column is not None:
        parsed["Bhavcopy_Previous_Close"] = (
            raw[previous_close_column]
        )
    else:
        parsed["Bhavcopy_Previous_Close"] = pd.NA

    if trade_date_column is not None:
        parsed["Trade_Date"] = pd.to_datetime(
            raw[trade_date_column],
            errors="coerce",
            dayfirst=True,
        ).dt.date

        parsed["Trade_Date"] = (
            parsed["Trade_Date"]
            .fillna(expected_trade_date)
        )
    else:
        parsed["Trade_Date"] = expected_trade_date

    if security_type_column is not None:
        parsed["Instrument_Type"] = (
            raw[security_type_column]
            .astype(str)
            .str.strip()
            .str.upper()
        )
    else:
        parsed["Instrument_Type"] = ""

    parsed["Symbol"] = (
        parsed["Symbol"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    parsed["Series"] = (
        parsed["Series"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    parsed["Close"] = pd.to_numeric(
        parsed["Close"],
        errors="coerce",
    )

    parsed["Volume"] = pd.to_numeric(
        parsed["Volume"],
        errors="coerce",
    )

    parsed["Bhavcopy_Previous_Close"] = pd.to_numeric(
        parsed["Bhavcopy_Previous_Close"],
        errors="coerce",
    )

    parsed = parsed.loc[
        parsed["Series"].eq(EQ_SERIES)
        & parsed["Symbol"].ne("")
        & parsed["Symbol"].ne("NAN")
        & parsed["Close"].notna()
        & parsed["Close"].gt(0)
    ].copy()

    parsed["Volume"] = (
        parsed["Volume"]
        .fillna(0)
        .clip(lower=0)
    )

    parsed = parsed.drop_duplicates(
        subset=[
            "Trade_Date",
            "Symbol",
        ],
        keep="last",
    )

    return parsed.reset_index(
        drop=True
    )


# ==========================================================
# HISTORICAL DATA ASSEMBLY
# ==========================================================

def read_cached_history() -> pd.DataFrame:
    """
    Read any previously assembled price history.
    """

    if not HISTORY_FILE.exists():
        return pd.DataFrame(
            columns=[
                "Trade_Date",
                "Symbol",
                "Close",
                "Volume",
                "Bhavcopy_Previous_Close",
            ]
        )

    history = pd.read_csv(
        HISTORY_FILE,
        low_memory=False,
    )

    if history.empty:
        return history

    history["Trade_Date"] = pd.to_datetime(
        history["Trade_Date"],
        errors="coerce",
    ).dt.date

    history["Close"] = pd.to_numeric(
        history["Close"],
        errors="coerce",
    )

    history["Volume"] = pd.to_numeric(
        history["Volume"],
        errors="coerce",
    )

    if "Bhavcopy_Previous_Close" in history.columns:
        history["Bhavcopy_Previous_Close"] = pd.to_numeric(
            history["Bhavcopy_Previous_Close"],
            errors="coerce",
        )
    else:
        history["Bhavcopy_Previous_Close"] = pd.NA

    history = history.dropna(
        subset=[
            "Trade_Date",
            "Symbol",
            "Close",
        ]
    )

    return history


def save_history(
    history: pd.DataFrame,
) -> None:
    """
    Save normalized historical price data.
    """

    OUTPUT_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    export_data = history.copy()

    export_data = export_data.sort_values(
        [
            "Symbol",
            "Trade_Date",
        ]
    )

    export_data.to_csv(
        HISTORY_FILE,
        index=False,
    )


def combine_history(
    *,
    existing_history: pd.DataFrame,
    new_frames: list[pd.DataFrame],
) -> pd.DataFrame:
    """
    Combine old and newly parsed history safely.
    """

    frames: list[pd.DataFrame] = []

    if not existing_history.empty:
        frames.append(
            existing_history
        )

    frames.extend(
        frame
        for frame in new_frames
        if not frame.empty
    )

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(
        frames,
        ignore_index=True,
    )

    combined["Trade_Date"] = pd.to_datetime(
        combined["Trade_Date"],
        errors="coerce",
    ).dt.date

    combined["Symbol"] = (
        combined["Symbol"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    combined["Close"] = pd.to_numeric(
        combined["Close"],
        errors="coerce",
    )

    combined["Volume"] = pd.to_numeric(
        combined["Volume"],
        errors="coerce",
    ).fillna(0)

    combined = combined.dropna(
        subset=[
            "Trade_Date",
            "Symbol",
            "Close",
        ]
    )

    combined = combined.loc[
        combined["Close"].gt(0)
    ]

    combined = combined.drop_duplicates(
        subset=[
            "Trade_Date",
            "Symbol",
        ],
        keep="last",
    )

    return combined.sort_values(
        [
            "Trade_Date",
            "Symbol",
        ]
    ).reset_index(
        drop=True
    )


# ==========================================================
# F&O FLAG
# ==========================================================

def locate_local_fno_file() -> Path | None:
    """
    Locate an existing AQSD F&O stock list.

    The list should have a Symbol column or symbols in its first column.
    """

    candidates = (
        BASE_DIR / "Data" / "FnO_Stocks.xlsx",
        BASE_DIR / "Data" / "FNO_Stocks.xlsx",
        BASE_DIR / "Data" / "FnO_Stocks.csv",
        BASE_DIR / "Config" / "FnO_Stocks.xlsx",
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def normalize_equity_symbol(
    value: object,
) -> str:
    """
    Normalize NSE or Yahoo-style equity symbols.
    """

    text = str(value).strip().upper()

    text = text.replace(
        "NSE:",
        "",
    )

    text = text.replace(
        "-EQ",
        "",
    )

    text = text.replace(
        ".NS",
        "",
    )

    return text


def load_fno_symbols() -> set[str]:
    """
    Read the existing AQSD F&O list when available.
    """

    source_file = locate_local_fno_file()

    if source_file is None:
        return set()

    if source_file.suffix.lower() == ".csv":
        dataframe = pd.read_csv(
            source_file
        )
    else:
        dataframe = pd.read_excel(
            source_file,
            engine="openpyxl",
        )

    if dataframe.empty:
        return set()

    symbol_column: str | None = None

    for column in dataframe.columns:
        if normalize_header(column) in {
            "SYMBOL",
            "TICKER",
            "STOCK",
        }:
            symbol_column = column
            break

    if symbol_column is None:
        symbol_column = dataframe.columns[0]

    return {
        normalize_equity_symbol(value)
        for value in dataframe[symbol_column]
        if normalize_equity_symbol(value)
    }


# ==========================================================
# INDICATORS
# ==========================================================

def calculate_symbol_snapshot(
    *,
    symbol: str,
    symbol_history: pd.DataFrame,
    snapshot_date: date,
    fno_symbols: set[str],
) -> dict[str, object] | None:
    """
    Calculate one stock's breadth snapshot.
    """

    history = symbol_history.loc[
        symbol_history["Trade_Date"]
        <= snapshot_date
    ].sort_values(
        "Trade_Date"
    ).copy()

    history = history.drop_duplicates(
        subset=["Trade_Date"],
        keep="last",
    )

    if len(history) < MINIMUM_REQUIRED_SESSIONS:
        return None

    history["EMA20"] = history["Close"].ewm(
        span=20,
        adjust=False,
        min_periods=20,
    ).mean()

    history["EMA50"] = history["Close"].ewm(
        span=50,
        adjust=False,
        min_periods=50,
    ).mean()

    history["EMA200"] = history["Close"].ewm(
        span=200,
        adjust=False,
        min_periods=200,
    ).mean()

    history["High_52W"] = history["Close"].rolling(
        window=252,
        min_periods=200,
    ).max()

    history["Low_52W"] = history["Close"].rolling(
        window=252,
        min_periods=200,
    ).min()

    history["Average_Volume_20"] = history["Volume"].rolling(
        window=20,
        min_periods=10,
    ).mean()

    latest = history.iloc[-1]

    previous = history.iloc[-2]

    required_values = (
        latest["Close"],
        latest["EMA20"],
        latest["EMA50"],
        latest["EMA200"],
        latest["High_52W"],
        latest["Low_52W"],
    )

    if any(
        pd.isna(value)
        for value in required_values
    ):
        return None

    bhavcopy_previous_close = (
        latest.get(
            "Bhavcopy_Previous_Close",
            pd.NA,
        )
    )

    if pd.notna(
        bhavcopy_previous_close
    ) and float(
        bhavcopy_previous_close
    ) > 0:
        previous_close = float(
            bhavcopy_previous_close
        )
    else:
        previous_close = float(
            previous["Close"]
        )

    return {
        "Symbol": symbol,
        "Close": round(
            float(latest["Close"]),
            4,
        ),
        "Previous_Close": round(
            previous_close,
            4,
        ),
        "EMA20": round(
            float(latest["EMA20"]),
            4,
        ),
        "EMA50": round(
            float(latest["EMA50"]),
            4,
        ),
        "EMA200": round(
            float(latest["EMA200"]),
            4,
        ),
        "High_52W": round(
            float(latest["High_52W"]),
            4,
        ),
        "Low_52W": round(
            float(latest["Low_52W"]),
            4,
        ),
        "Sector": "UNKNOWN",
        "Market_Cap_Category": "UNKNOWN",
        "Volume": int(
            round(float(latest["Volume"]))
        ),
        "Average_Volume_20": round(
            float(latest["Average_Volume_20"]),
            2,
        ),
        "Is_FnO": symbol in fno_symbols,
        "Snapshot_Date": snapshot_date,
        "History_Sessions": len(history),
    }


def build_snapshot(
    *,
    history: pd.DataFrame,
    requested_date: date,
) -> tuple[pd.DataFrame, date]:
    """
    Build the latest valid market breadth snapshot.
    """

    eligible_history = history.loc[
        history["Trade_Date"]
        <= requested_date
    ].copy()

    if eligible_history.empty:
        raise RuntimeError(
            "No NSE history is available on or before "
            f"{requested_date}."
        )

    snapshot_date = max(
        eligible_history["Trade_Date"]
    )

    fno_symbols = load_fno_symbols()

    snapshot_rows: list[dict[str, object]] = []

    grouped = eligible_history.groupby(
        "Symbol",
        sort=True,
    )

    for symbol, symbol_history in grouped:
        row = calculate_symbol_snapshot(
            symbol=str(symbol),
            symbol_history=symbol_history,
            snapshot_date=snapshot_date,
            fno_symbols=fno_symbols,
        )

        if row is not None:
            snapshot_rows.append(
                row
            )

    if not snapshot_rows:
        raise RuntimeError(
            "No stock had sufficient history for EMA200 and "
            "52-week calculations."
        )

    snapshot = pd.DataFrame(
        snapshot_rows
    )

    snapshot = snapshot.sort_values(
        "Symbol"
    ).reset_index(
        drop=True
    )

    required_output_columns = [
        "Symbol",
        "Close",
        "Previous_Close",
        "EMA20",
        "EMA50",
        "EMA200",
        "High_52W",
        "Low_52W",
        "Sector",
        "Market_Cap_Category",
        "Volume",
        "Average_Volume_20",
        "Is_FnO",
    ]

    snapshot = snapshot[
        required_output_columns
    ]

    return (
        snapshot,
        snapshot_date,
    )


# ==========================================================
# EXCEL EXPORT
# ==========================================================

def format_snapshot_workbook(
    output_file: Path,
) -> None:
    """
    Apply simple professional formatting to the workbook.
    """

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(
        output_file
    )

    worksheet = workbook["Breadth Snapshot"]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_cells in worksheet.columns:
        maximum_length = 0

        column_number = column_cells[0].column

        for cell in column_cells:
            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            maximum_length = max(
                maximum_length,
                len(value),
            )

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = min(
            max(
                maximum_length + 2,
                12,
            ),
            28,
        )

    workbook.save(
        output_file
    )


def export_snapshot(
    *,
    snapshot: pd.DataFrame,
    result_summary: dict[str, object],
) -> None:
    """
    Export snapshot and builder summary to Excel.
    """

    OUTPUT_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    summary = pd.DataFrame(
        [
            {
                "Field": key,
                "Value": value,
            }
            for key, value
            in result_summary.items()
        ]
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        snapshot.to_excel(
            writer,
            sheet_name="Breadth Snapshot",
            index=False,
        )

        summary.to_excel(
            writer,
            sheet_name="Builder Summary",
            index=False,
        )

    format_snapshot_workbook(
        OUTPUT_FILE
    )


# ==========================================================
# MANIFEST
# ==========================================================

def save_manifest(
    rows: list[dict[str, object]],
) -> None:
    """
    Save download and parse history.
    """

    OUTPUT_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    pd.DataFrame(
        rows
    ).to_csv(
        MANIFEST_FILE,
        index=False,
    )


# ==========================================================
# MAIN BUILDER
# ==========================================================

def run_nse_market_breadth_builder(
    *,
    requested_date: date,
    lookback_calendar_days: int,
    force_download: bool,
) -> BreadthBuilderResult:
    """
    Download official NSE data and create the breadth workbook.
    """

    print()
    print("=" * 94)
    print("AQSD NSE MARKET BREADTH DATA BUILDER")
    print("=" * 94)
    print(f"Module                    : {MODULE_ID}")
    print(f"Version                   : {MODULE_VERSION}")
    print(f"Requested Date            : {requested_date}")
    print(f"Calendar Lookback         : {lookback_calendar_days} days")
    print(f"Force Download            : {force_download}")
    print(f"Official Source           : NSE CM-UDiFF Bhavcopy")
    print("-" * 94)

    candidate_dates = build_candidate_dates(
        requested_date=requested_date,
        lookback_calendar_days=lookback_calendar_days,
    )

    existing_history = read_cached_history()

    existing_dates: set[date] = set()

    if not existing_history.empty:
        existing_dates = set(
            existing_history["Trade_Date"]
        )

    session = create_nse_session()

    downloaded_files = 0
    cached_files = 0
    unavailable_dates = 0
    failed_dates = 0

    available_files = 0

    parsed_frames: list[pd.DataFrame] = []
    manifest_rows: list[dict[str, object]] = []

    for sequence, trade_date in enumerate(
        candidate_dates,
        start=1,
    ):
        if (
            trade_date in existing_dates
            and not force_download
        ):
            manifest_rows.append(
                {
                    "Trade_Date": trade_date,
                    "Download_Status": "HISTORY CACHE",
                    "Parse_Status": "SKIPPED",
                    "Rows": 0,
                    "Message": (
                        "Trade date already exists in assembled history."
                    ),
                }
            )

            continue

        print(
            f"[{sequence:03d}/{len(candidate_dates):03d}] "
            f"{trade_date}",
            end="",
        )

        status, zip_path, message = (
            download_udiff_bhavcopy(
                session=session,
                trade_date=trade_date,
                force_download=force_download,
            )
        )

        if status == "DOWNLOADED":
            downloaded_files += 1

        elif status == "CACHED":
            cached_files += 1

        elif status == "UNAVAILABLE":
            unavailable_dates += 1

            print(" : unavailable")

            manifest_rows.append(
                {
                    "Trade_Date": trade_date,
                    "Download_Status": status,
                    "Parse_Status": "NOT RUN",
                    "Rows": 0,
                    "Message": message,
                }
            )

            time.sleep(
                REQUEST_DELAY_SECONDS
            )

            continue

        else:
            failed_dates += 1

            print(
                f" : failed — {message}"
            )

            manifest_rows.append(
                {
                    "Trade_Date": trade_date,
                    "Download_Status": status,
                    "Parse_Status": "NOT RUN",
                    "Rows": 0,
                    "Message": message,
                }
            )

            time.sleep(
                REQUEST_DELAY_SECONDS
            )

            continue

        available_files += 1

        assert zip_path is not None

        try:
            parsed = parse_udiff_bhavcopy(
                zip_path=zip_path,
                expected_trade_date=trade_date,
            )

        except Exception as exc:
            failed_dates += 1

            print(
                f" : parse failed — {exc}"
            )

            manifest_rows.append(
                {
                    "Trade_Date": trade_date,
                    "Download_Status": status,
                    "Parse_Status": "FAILED",
                    "Rows": 0,
                    "Message": str(exc),
                }
            )

            time.sleep(
                REQUEST_DELAY_SECONDS
            )

            continue

        parsed_frames.append(
            parsed
        )

        print(
            f" : {status.lower()}, "
            f"{len(parsed):,} EQ rows"
        )

        manifest_rows.append(
            {
                "Trade_Date": trade_date,
                "Download_Status": status,
                "Parse_Status": "SUCCESS",
                "Rows": len(parsed),
                "Message": message,
            }
        )

        time.sleep(
            REQUEST_DELAY_SECONDS
        )

    history = combine_history(
        existing_history=existing_history,
        new_frames=parsed_frames,
    )

    if history.empty:
        save_manifest(
            manifest_rows
        )

        return BreadthBuilderResult(
            requested_date=requested_date,
            snapshot_date=None,
            calendar_dates_checked=len(candidate_dates),
            trading_files_available=available_files,
            downloaded_files=downloaded_files,
            cached_files=cached_files,
            unavailable_dates=unavailable_dates,
            failed_dates=failed_dates,
            historical_rows=0,
            historical_symbols=0,
            eligible_symbols=0,
            output_rows=0,
            history_file=HISTORY_FILE,
            output_file=OUTPUT_FILE,
            manifest_file=MANIFEST_FILE,
            status="FAILED",
            message="No valid NSE history could be assembled.",
        )

    save_history(
        history
    )

    snapshot, snapshot_date = build_snapshot(
        history=history,
        requested_date=requested_date,
    )

    historical_symbols = int(
        history["Symbol"].nunique()
    )

    sessions_per_symbol = (
        history.groupby("Symbol")["Trade_Date"]
        .nunique()
    )

    eligible_symbols = int(
        (
            sessions_per_symbol
            >= MINIMUM_REQUIRED_SESSIONS
        ).sum()
    )

    summary = {
        "Module": MODULE_ID,
        "Version": MODULE_VERSION,
        "Requested Date": requested_date,
        "Snapshot Date": snapshot_date,
        "Official Source": "NSE CM-UDiFF Common Bhavcopy Final",
        "Calendar Dates Checked": len(candidate_dates),
        "Trading Files Available": available_files,
        "Files Downloaded": downloaded_files,
        "Files Reused From Cache": cached_files,
        "Unavailable Dates": unavailable_dates,
        "Failed Dates": failed_dates,
        "Historical Rows": len(history),
        "Historical Symbols": historical_symbols,
        "Symbols With 200+ Sessions": eligible_symbols,
        "Snapshot Rows": len(snapshot),
        "History File": HISTORY_FILE,
        "Output File": OUTPUT_FILE,
        "Status": "SUCCESS",
    }

    export_snapshot(
        snapshot=snapshot,
        result_summary=summary,
    )

    save_manifest(
        manifest_rows
    )

    return BreadthBuilderResult(
        requested_date=requested_date,
        snapshot_date=snapshot_date,
        calendar_dates_checked=len(candidate_dates),
        trading_files_available=available_files,
        downloaded_files=downloaded_files,
        cached_files=cached_files,
        unavailable_dates=unavailable_dates,
        failed_dates=failed_dates,
        historical_rows=len(history),
        historical_symbols=historical_symbols,
        eligible_symbols=eligible_symbols,
        output_rows=len(snapshot),
        history_file=HISTORY_FILE,
        output_file=OUTPUT_FILE,
        manifest_file=MANIFEST_FILE,
        status="SUCCESS",
        message=(
            "Official NSE market breadth snapshot created successfully."
        ),
    )


# ==========================================================
# DISPLAY
# ==========================================================

def display_result(
    result: BreadthBuilderResult,
) -> None:
    """
    Display the final builder result.
    """

    print()
    print("=" * 94)
    print("AQSD NSE MARKET BREADTH DATA BUILDER — RESULT")
    print("=" * 94)
    print(f"Requested Date             : {result.requested_date}")
    print(f"Snapshot Date              : {result.snapshot_date}")
    print(f"Calendar Dates Checked     : {result.calendar_dates_checked}")
    print(
        f"Trading Files Available    : "
        f"{result.trading_files_available}"
    )
    print(f"Files Downloaded           : {result.downloaded_files}")
    print(f"Files Reused               : {result.cached_files}")
    print(f"Unavailable Dates          : {result.unavailable_dates}")
    print(f"Failed Dates               : {result.failed_dates}")
    print("-" * 94)
    print(f"Historical Rows            : {result.historical_rows:,}")
    print(f"Historical Symbols         : {result.historical_symbols:,}")
    print(f"Eligible Symbols           : {result.eligible_symbols:,}")
    print(f"Snapshot Rows              : {result.output_rows:,}")
    print("-" * 94)
    print(f"History File               : {result.history_file}")
    print(f"Snapshot File              : {result.output_file}")
    print(f"Manifest File              : {result.manifest_file}")
    print("-" * 94)
    print(f"Status                     : {result.status}")
    print(f"Message                    : {result.message}")
    print("=" * 94)


# ==========================================================
# COMMAND LINE
# ==========================================================

def parse_arguments() -> argparse.Namespace:
    """
    Read command-line arguments.
    """

    parser = argparse.ArgumentParser(
        description=(
            "Download official NSE CM UDiFF bhavcopies and create "
            "the AQSD market breadth snapshot."
        )
    )

    parser.add_argument(
        "--date",
        required=True,
        help="Requested snapshot date in YYYY-MM-DD format.",
    )

    parser.add_argument(
        "--lookback-days",
        type=int,
        default=DEFAULT_LOOKBACK_CALENDAR_DAYS,
        help=(
            "Calendar-day history to inspect. "
            f"Default: {DEFAULT_LOOKBACK_CALENDAR_DAYS}"
        ),
    )

    parser.add_argument(
        "--force-download",
        action="store_true",
        help="Download files again even when local files exist.",
    )

    return parser.parse_args()


def main() -> None:
    """
    Command-line entry point.
    """

    arguments = parse_arguments()

    if arguments.lookback_days < 300:
        print()
        print(
            "WARNING: Less than 300 calendar days may be "
            "insufficient for EMA200."
        )

    try:
        result = run_nse_market_breadth_builder(
            requested_date=parse_date(
                arguments.date
            ),
            lookback_calendar_days=(
                arguments.lookback_days
            ),
            force_download=(
                arguments.force_download
            ),
        )

    except Exception as exc:
        print()
        print("=" * 94)
        print("AQSD NSE MARKET BREADTH DATA BUILDER")
        print("=" * 94)
        print("Status : FAILED")
        print(f"Reason : {exc}")
        print("=" * 94)

        raise SystemExit(1) from exc

    display_result(
        result
    )

    if result.status != "SUCCESS":
        raise SystemExit(1)


if __name__ == "__main__":
    main()