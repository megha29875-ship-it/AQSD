"""
AQSD
NSE Security Classification Builder

Module : MBI-003
Version: 1.0.0
Author : AQSD

Description
-----------
Enriches the AQSD market-breadth snapshot with official NSE Indices
classification information.

The module adds or updates:

- Company Name
- Industry
- Sector
- Market_Cap_Category
- Is_Nifty_100
- Is_Nifty_200
- Is_Nifty_500
- Is_FnO

Official sources
----------------
- Nifty 100 constituent list
- Nifty 200 constituent list
- Nifty 500 constituent list

Market-cap classification
-------------------------
LARGE CAP:
    Member of Nifty 100.

MID CAP:
    Member of Nifty 200 but not Nifty 100.

SMALL CAP:
    Member of Nifty 500 but not Nifty 200.

UNCLASSIFIED:
    Equity is outside the Nifty 500 constituent universe.

Important
---------
This classification is designed for AQSD breadth segmentation.

Stocks outside Nifty 500 remain in the breadth universe but their
sector and market-cap category may remain UNKNOWN or UNCLASSIFIED.

This module does not use Yahoo Finance or Fyers.
"""

from __future__ import annotations

import argparse
import io
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Final

import pandas as pd
import requests


# ==========================================================
# MODULE SETTINGS
# ==========================================================

MODULE_ID: Final[str] = "MBI-003"
MODULE_VERSION: Final[str] = "1.0.0"

BASE_DIR: Final[Path] = Path(__file__).resolve().parents[2]

BREADTH_DIRECTORY: Final[Path] = (
    BASE_DIR
    / "Data"
    / "Market_Breadth"
)

BREADTH_FILE: Final[Path] = (
    BREADTH_DIRECTORY
    / "market_breadth_snapshot.xlsx"
)

CLASSIFICATION_DIRECTORY: Final[Path] = (
    BREADTH_DIRECTORY
    / "Classification"
)

CLASSIFICATION_MASTER_FILE: Final[Path] = (
    CLASSIFICATION_DIRECTORY
    / "nse_security_classification_master.xlsx"
)

CLASSIFICATION_MANIFEST_FILE: Final[Path] = (
    CLASSIFICATION_DIRECTORY
    / "classification_manifest.csv"
)

REQUEST_TIMEOUT_SECONDS: Final[int] = 40

USER_AGENT: Final[str] = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)

NSE_HOME_URL: Final[str] = "https://www.nseindia.com"

INDEX_URLS: Final[dict[str, str]] = {
    "NIFTY_100": (
        "https://nsearchives.nseindia.com/"
        "content/indices/ind_nifty100list.csv"
    ),
    "NIFTY_200": (
        "https://nsearchives.nseindia.com/"
        "content/indices/ind_nifty200list.csv"
    ),
    "NIFTY_500": (
        "https://nsearchives.nseindia.com/"
        "content/indices/ind_nifty500list.csv"
    ),
}


# ==========================================================
# RESULT MODEL
# ==========================================================

@dataclass(frozen=True)
class ClassificationBuilderResult:
    """
    Final result from the classification builder.
    """

    breadth_file: Path
    classification_master_file: Path
    manifest_file: Path

    breadth_rows: int
    nifty_100_symbols: int
    nifty_200_symbols: int
    nifty_500_symbols: int
    fno_symbols: int

    sector_classified_rows: int
    market_cap_classified_rows: int
    fno_classified_rows: int
    unclassified_rows: int

    status: str
    message: str


# ==========================================================
# GENERAL HELPERS
# ==========================================================

def normalize_header(
    value: object,
) -> str:
    """
    Convert a heading into a normalized uppercase form.
    """

    text = str(value).strip().upper()

    text = re.sub(
        r"[^A-Z0-9]+",
        "_",
        text,
    )

    text = re.sub(
        r"_+",
        "_",
        text,
    )

    return text.strip("_")


def normalize_symbol(
    value: object,
) -> str:
    """
    Normalize NSE, Yahoo or Fyers-style equity symbols.
    """

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass

    text = str(value).strip().upper()

    text = text.replace(
        "NSE:",
        "",
    )

    text = text.replace(
        "-EQ",
        "",
    )

    text = text.replace(
        ".NS",
        "",
    )

    return text.strip()


def normalize_text(
    value: object,
    default: str = "UNKNOWN",
) -> str:
    """
    Normalize a text value.
    """

    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except TypeError:
        pass

    text = str(value).strip()

    if not text:
        return default

    return text.upper()


def create_session() -> requests.Session:
    """
    Create a requests session suitable for NSE downloads.
    """

    session = requests.Session()

    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": (
                "text/csv,text/plain,application/octet-stream,"
                "application/vnd.ms-excel,*/*"
            ),
            "Accept-Language": "en-IN,en;q=0.9",
            "Referer": NSE_HOME_URL,
            "Connection": "keep-alive",
        }
    )

    try:
        session.get(
            NSE_HOME_URL,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
    except requests.RequestException:
        pass

    return session


# ==========================================================
# CSV READING
# ==========================================================

def read_csv_bytes(
    content: bytes,
) -> pd.DataFrame:
    """
    Read CSV bytes using common encoding fallbacks.
    """

    errors: list[str] = []

    for encoding in (
        "utf-8-sig",
        "utf-8",
        "latin-1",
    ):
        try:
            dataframe = pd.read_csv(
                io.BytesIO(content),
                encoding=encoding,
                low_memory=False,
            )

            if not dataframe.empty:
                return dataframe

        except Exception as exc:
            errors.append(
                f"{encoding}: {exc}"
            )

    raise RuntimeError(
        "Could not read downloaded index CSV. "
        + " | ".join(errors)
    )


def download_index_file(
    *,
    session: requests.Session,
    index_name: str,
    url: str,
) -> pd.DataFrame:
    """
    Download one official NSE Indices constituent file.
    """

    response = session.get(
        url,
        timeout=REQUEST_TIMEOUT_SECONDS,
    )

    if response.status_code != 200:
        raise RuntimeError(
            f"{index_name} download failed with "
            f"HTTP {response.status_code}."
        )

    if not response.content:
        raise RuntimeError(
            f"{index_name} download returned an empty file."
        )

    dataframe = read_csv_bytes(
        response.content
    )

    dataframe.columns = [
        str(column).strip()
        for column in dataframe.columns
    ]

    return dataframe


# ==========================================================
# INDEX FILE NORMALIZATION
# ==========================================================

def identify_column(
    dataframe: pd.DataFrame,
    aliases: set[str],
) -> str | None:
    """
    Find a dataframe column using normalized aliases.
    """

    normalized_aliases = {
        normalize_header(alias)
        for alias in aliases
    }

    for column in dataframe.columns:
        if normalize_header(column) in normalized_aliases:
            return str(column)

    return None


def normalize_index_constituents(
    *,
    dataframe: pd.DataFrame,
    index_name: str,
) -> pd.DataFrame:
    """
    Normalize an NSE Indices constituent file.
    """

    symbol_column = identify_column(
        dataframe,
        {
            "Symbol",
            "Ticker",
            "Security Symbol",
        },
    )

    company_column = identify_column(
        dataframe,
        {
            "Company Name",
            "Company",
            "Security Name",
            "Name",
        },
    )

    industry_column = identify_column(
        dataframe,
        {
            "Industry",
            "Sector",
            "Basic Industry",
        },
    )

    series_column = identify_column(
        dataframe,
        {
            "Series",
        },
    )

    isin_column = identify_column(
        dataframe,
        {
            "ISIN Code",
            "ISIN",
        },
    )

    if symbol_column is None:
        raise KeyError(
            f"{index_name} file does not contain a Symbol column. "
            f"Available columns: {list(dataframe.columns)}"
        )

    normalized = pd.DataFrame()

    normalized["Symbol"] = (
        dataframe[symbol_column]
        .apply(normalize_symbol)
    )

    if company_column is not None:
        normalized["Company_Name"] = (
            dataframe[company_column]
            .apply(normalize_text)
        )
    else:
        normalized["Company_Name"] = "UNKNOWN"

    if industry_column is not None:
        normalized["Industry"] = (
            dataframe[industry_column]
            .apply(normalize_text)
        )
    else:
        normalized["Industry"] = "UNKNOWN"

    if series_column is not None:
        normalized["Series"] = (
            dataframe[series_column]
            .apply(normalize_text)
        )
    else:
        normalized["Series"] = "EQ"

    if isin_column is not None:
        normalized["ISIN"] = (
            dataframe[isin_column]
            .apply(normalize_text)
        )
    else:
        normalized["ISIN"] = "UNKNOWN"

    normalized["Index_Name"] = index_name

    normalized = normalized.loc[
        normalized["Symbol"].ne("")
    ].copy()

    normalized = normalized.drop_duplicates(
        subset=["Symbol"],
        keep="last",
    )

    return normalized.reset_index(
        drop=True
    )


# ==========================================================
# SECTOR DERIVATION
# ==========================================================

def derive_sector_from_industry(
    industry: object,
) -> str:
    """
    Convert detailed NSE industry labels into broad AQSD sectors.

    The detailed Industry value is retained separately.
    """

    text = normalize_text(
        industry
    )

    sector_rules: tuple[
        tuple[tuple[str, ...], str],
        ...
    ] = (
        (
            (
                "BANK",
                "FINANC",
                "INSURANCE",
                "ASSET MANAGEMENT",
                "CAPITAL MARKET",
                "HOUSING FINANCE",
                "MICROFINANCE",
                "NBFC",
            ),
            "FINANCIAL SERVICES",
        ),
        (
            (
                "SOFTWARE",
                "INFORMATION TECHNOLOGY",
                "IT SERVICES",
                "COMPUTER",
                "DIGITAL",
                "TECHNOLOGY",
            ),
            "INFORMATION TECHNOLOGY",
        ),
        (
            (
                "PHARMA",
                "HEALTHCARE",
                "HOSPITAL",
                "DIAGNOSTIC",
                "BIOTECH",
                "MEDICAL",
            ),
            "HEALTHCARE",
        ),
        (
            (
                "AUTOMOBILE",
                "AUTO COMPONENT",
                "TYRE",
                "TRACTOR",
                "TWO WHEELER",
                "PASSENGER CAR",
                "COMMERCIAL VEHICLE",
            ),
            "AUTOMOBILE",
        ),
        (
            (
                "OIL",
                "GAS",
                "PETROLEUM",
                "REFINERY",
                "COAL",
                "ENERGY",
                "POWER",
                "ELECTRICITY",
            ),
            "ENERGY",
        ),
        (
            (
                "FMCG",
                "FOOD",
                "BEVERAGE",
                "TOBACCO",
                "PERSONAL CARE",
                "HOUSEHOLD",
                "CONSUMER FOOD",
            ),
            "FMCG",
        ),
        (
            (
                "CONSUMER DURABLE",
                "ELECTRONICS",
                "APPLIANCE",
                "JEWELLERY",
                "RETAIL",
                "TEXTILE",
                "APPAREL",
                "FOOTWEAR",
            ),
            "CONSUMER DISCRETIONARY",
        ),
        (
            (
                "METAL",
                "STEEL",
                "ALUMINIUM",
                "COPPER",
                "MINING",
                "MINERAL",
                "ZINC",
                "FERRO",
            ),
            "METALS & MINING",
        ),
        (
            (
                "CEMENT",
                "CONSTRUCTION",
                "INFRASTRUCTURE",
                "REAL ESTATE",
                "BUILDING",
                "ENGINEERING",
            ),
            "CONSTRUCTION & INFRASTRUCTURE",
        ),
        (
            (
                "CHEMICAL",
                "FERTILIZER",
                "PESTICIDE",
                "AGROCHEMICAL",
                "SPECIALITY CHEMICAL",
            ),
            "CHEMICALS",
        ),
        (
            (
                "TELECOM",
                "COMMUNICATION",
            ),
            "TELECOMMUNICATION",
        ),
        (
            (
                "MEDIA",
                "ENTERTAINMENT",
                "BROADCAST",
                "PRINTING",
            ),
            "MEDIA & ENTERTAINMENT",
        ),
        (
            (
                "LOGISTIC",
                "TRANSPORT",
                "AIRLINE",
                "PORT",
                "SHIPPING",
                "RAILWAY",
            ),
            "TRANSPORTATION & LOGISTICS",
        ),
        (
            (
                "DEFENCE",
                "AEROSPACE",
            ),
            "DEFENCE",
        ),
        (
            (
                "AGRICULTURE",
                "PLANTATION",
                "SUGAR",
                "TEA",
                "COFFEE",
            ),
            "AGRICULTURE",
        ),
    )

    for keywords, sector in sector_rules:
        if any(
            keyword in text
            for keyword in keywords
        ):
            return sector

    if text == "UNKNOWN":
        return "UNKNOWN"

    return "OTHER"


# ==========================================================
# CLASSIFICATION MASTER
# ==========================================================

def build_classification_master(
    *,
    nifty_100: pd.DataFrame,
    nifty_200: pd.DataFrame,
    nifty_500: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build one consolidated classification master.
    """

    nifty_100_symbols = set(
        nifty_100["Symbol"]
    )

    nifty_200_symbols = set(
        nifty_200["Symbol"]
    )

    nifty_500_symbols = set(
        nifty_500["Symbol"]
    )

    master = nifty_500[
        [
            "Symbol",
            "Company_Name",
            "Industry",
            "Series",
            "ISIN",
        ]
    ].copy()

    master["Sector"] = master["Industry"].apply(
        derive_sector_from_industry
    )

    master["Is_Nifty_100"] = master["Symbol"].isin(
        nifty_100_symbols
    )

    master["Is_Nifty_200"] = master["Symbol"].isin(
        nifty_200_symbols
    )

    master["Is_Nifty_500"] = master["Symbol"].isin(
        nifty_500_symbols
    )

    def classify_market_cap(
        symbol: str,
    ) -> str:
        if symbol in nifty_100_symbols:
            return "LARGE CAP"

        if symbol in nifty_200_symbols:
            return "MID CAP"

        if symbol in nifty_500_symbols:
            return "SMALL CAP"

        return "UNCLASSIFIED"

    master["Market_Cap_Category"] = (
        master["Symbol"]
        .apply(classify_market_cap)
    )

    master = master.sort_values(
        [
            "Market_Cap_Category",
            "Sector",
            "Symbol",
        ]
    ).reset_index(
        drop=True
    )

    return master


# ==========================================================
# F&O LIST
# ==========================================================

def locate_fno_file() -> Path | None:
    """
    Locate the existing AQSD F&O securities file.
    """

    candidates = (
        BASE_DIR / "Data" / "FnO_Stocks.xlsx",
        BASE_DIR / "Data" / "FNO_Stocks.xlsx",
        BASE_DIR / "Data" / "FnO_Stocks.csv",
        BASE_DIR / "Data" / "FNO_Stocks.csv",
        BASE_DIR / "Config" / "FnO_Stocks.xlsx",
        BASE_DIR / "Config" / "FNO_Stocks.xlsx",
    )

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def load_fno_symbols() -> set[str]:
    """
    Load the existing AQSD F&O symbol list.
    """

    source_file = locate_fno_file()

    if source_file is None:
        return set()

    if source_file.suffix.lower() == ".csv":
        dataframe = pd.read_csv(
            source_file,
            low_memory=False,
        )
    else:
        dataframe = pd.read_excel(
            source_file,
            engine="openpyxl",
        )

    if dataframe.empty:
        return set()

    symbol_column = identify_column(
        dataframe,
        {
            "Symbol",
            "Ticker",
            "Stock",
            "F&O Symbol",
        },
    )

    if symbol_column is None:
        symbol_column = str(
            dataframe.columns[0]
        )

    return {
        symbol
        for symbol in (
            normalize_symbol(value)
            for value in dataframe[symbol_column]
        )
        if symbol
    }


# ==========================================================
# BREADTH FILE
# ==========================================================

def read_breadth_snapshot() -> pd.DataFrame:
    """
    Read the existing AQSD breadth snapshot.
    """

    if not BREADTH_FILE.exists():
        raise FileNotFoundError(
            f"Breadth snapshot not found: {BREADTH_FILE}"
        )

    try:
        dataframe = pd.read_excel(
            BREADTH_FILE,
            sheet_name="Breadth Snapshot",
            engine="openpyxl",
        )

    except ValueError:
        dataframe = pd.read_excel(
            BREADTH_FILE,
            sheet_name=0,
            engine="openpyxl",
        )

    if dataframe.empty:
        raise RuntimeError(
            "The breadth snapshot workbook is empty."
        )

    symbol_column = identify_column(
        dataframe,
        {
            "Symbol",
        },
    )

    if symbol_column is None:
        raise KeyError(
            "Breadth snapshot does not contain a Symbol column."
        )

    if symbol_column != "Symbol":
        dataframe = dataframe.rename(
            columns={
                symbol_column: "Symbol",
            }
        )

    dataframe["Symbol"] = dataframe["Symbol"].apply(
        normalize_symbol
    )

    dataframe = dataframe.loc[
        dataframe["Symbol"].ne("")
    ].copy()

    dataframe = dataframe.drop_duplicates(
        subset=["Symbol"],
        keep="last",
    )

    return dataframe.reset_index(
        drop=True
    )


def enrich_breadth_snapshot(
    *,
    breadth: pd.DataFrame,
    classification_master: pd.DataFrame,
    fno_symbols: set[str],
) -> pd.DataFrame:
    """
    Merge official index classification into the breadth snapshot.
    """

    classification_columns = [
        "Symbol",
        "Company_Name",
        "Industry",
        "Sector",
        "Market_Cap_Category",
        "Is_Nifty_100",
        "Is_Nifty_200",
        "Is_Nifty_500",
    ]

    enriched = breadth.drop(
        columns=[
            "Company_Name",
            "Industry",
            "Sector",
            "Market_Cap_Category",
            "Is_Nifty_100",
            "Is_Nifty_200",
            "Is_Nifty_500",
            "Is_FnO",
        ],
        errors="ignore",
    )

    enriched = enriched.merge(
        classification_master[
            classification_columns
        ],
        on="Symbol",
        how="left",
        validate="one_to_one",
    )

    text_defaults = {
        "Company_Name": "UNKNOWN",
        "Industry": "UNKNOWN",
        "Sector": "UNKNOWN",
        "Market_Cap_Category": "UNCLASSIFIED",
    }

    for column, default in text_defaults.items():
        enriched[column] = (
            enriched[column]
            .fillna(default)
            .astype(str)
            .str.strip()
            .str.upper()
        )

    boolean_columns = (
        "Is_Nifty_100",
        "Is_Nifty_200",
        "Is_Nifty_500",
    )

    for column in boolean_columns:
        enriched[column] = (
            enriched[column]
            .fillna(False)
            .astype(bool)
        )

    enriched["Is_FnO"] = enriched["Symbol"].isin(
        fno_symbols
    )

    preferred_order = [
        "Symbol",
        "Company_Name",
        "Close",
        "Previous_Close",
        "EMA20",
        "EMA50",
        "EMA200",
        "High_52W",
        "Low_52W",
        "Sector",
        "Industry",
        "Market_Cap_Category",
        "Volume",
        "Average_Volume_20",
        "Is_FnO",
        "Is_Nifty_100",
        "Is_Nifty_200",
        "Is_Nifty_500",
    ]

    ordered_columns = [
        column
        for column in preferred_order
        if column in enriched.columns
    ]

    remaining_columns = [
        column
        for column in enriched.columns
        if column not in ordered_columns
    ]

    enriched = enriched[
        ordered_columns
        + remaining_columns
    ]

    return enriched.sort_values(
        "Symbol"
    ).reset_index(
        drop=True
    )


# ==========================================================
# EXPORT
# ==========================================================

def format_workbook(
    workbook_path: Path,
) -> None:
    """
    Apply simple workbook formatting.
    """

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(
        workbook_path
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for column_cells in worksheet.columns:
            maximum_length = 0

            column_number = (
                column_cells[0].column
            )

            for cell in column_cells:
                value = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                maximum_length = max(
                    maximum_length,
                    len(value),
                )

            worksheet.column_dimensions[
                get_column_letter(column_number)
            ].width = min(
                max(
                    maximum_length + 2,
                    12,
                ),
                36,
            )

    workbook.save(
        workbook_path
    )


def export_classification_master(
    *,
    master: pd.DataFrame,
    nifty_100: pd.DataFrame,
    nifty_200: pd.DataFrame,
    nifty_500: pd.DataFrame,
) -> None:
    """
    Export the classification master workbook.
    """

    CLASSIFICATION_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        CLASSIFICATION_MASTER_FILE,
        engine="openpyxl",
    ) as writer:
        master.to_excel(
            writer,
            sheet_name="Classification Master",
            index=False,
        )

        nifty_100.to_excel(
            writer,
            sheet_name="Nifty 100",
            index=False,
        )

        nifty_200.to_excel(
            writer,
            sheet_name="Nifty 200",
            index=False,
        )

        nifty_500.to_excel(
            writer,
            sheet_name="Nifty 500",
            index=False,
        )

    format_workbook(
        CLASSIFICATION_MASTER_FILE
    )


def export_enriched_breadth(
    *,
    enriched_breadth: pd.DataFrame,
    classification_summary: pd.DataFrame,
) -> None:
    """
    Overwrite the breadth workbook with enriched information.
    """

    with pd.ExcelWriter(
        BREADTH_FILE,
        engine="openpyxl",
    ) as writer:
        enriched_breadth.to_excel(
            writer,
            sheet_name="Breadth Snapshot",
            index=False,
        )

        classification_summary.to_excel(
            writer,
            sheet_name="Classification Summary",
            index=False,
        )

    format_workbook(
        BREADTH_FILE
    )


# ==========================================================
# MAIN BUILDER
# ==========================================================

def run_classification_builder() -> ClassificationBuilderResult:
    """
    Download official NSE classifications and enrich breadth data.
    """

    print()
    print("=" * 96)
    print("AQSD NSE SECURITY CLASSIFICATION BUILDER")
    print("=" * 96)
    print(f"Module                    : {MODULE_ID}")
    print(f"Version                   : {MODULE_VERSION}")
    print(f"Breadth File              : {BREADTH_FILE}")
    print("-" * 96)

    breadth = read_breadth_snapshot()

    session = create_session()

    downloaded_indexes: dict[str, pd.DataFrame] = {}

    manifest_rows: list[dict[str, object]] = []

    for index_name, url in INDEX_URLS.items():
        print(
            f"Downloading {index_name}...",
            end=" ",
        )

        raw_dataframe = download_index_file(
            session=session,
            index_name=index_name,
            url=url,
        )

        normalized = normalize_index_constituents(
            dataframe=raw_dataframe,
            index_name=index_name,
        )

        downloaded_indexes[
            index_name
        ] = normalized

        print(
            f"{len(normalized):,} symbols"
        )

        manifest_rows.append(
            {
                "Downloaded_At": (
                    datetime.now()
                    .astimezone()
                    .isoformat()
                ),
                "Index_Name": index_name,
                "URL": url,
                "Rows": len(normalized),
                "Status": "SUCCESS",
            }
        )

    nifty_100 = downloaded_indexes[
        "NIFTY_100"
    ]

    nifty_200 = downloaded_indexes[
        "NIFTY_200"
    ]

    nifty_500 = downloaded_indexes[
        "NIFTY_500"
    ]

    classification_master = (
        build_classification_master(
            nifty_100=nifty_100,
            nifty_200=nifty_200,
            nifty_500=nifty_500,
        )
    )

    fno_symbols = load_fno_symbols()

    enriched_breadth = enrich_breadth_snapshot(
        breadth=breadth,
        classification_master=classification_master,
        fno_symbols=fno_symbols,
    )

    sector_classified_rows = int(
        enriched_breadth["Sector"]
        .ne("UNKNOWN")
        .sum()
    )

    market_cap_classified_rows = int(
        enriched_breadth[
            "Market_Cap_Category"
        ]
        .ne("UNCLASSIFIED")
        .sum()
    )

    fno_classified_rows = int(
        enriched_breadth["Is_FnO"].sum()
    )

    unclassified_rows = int(
        enriched_breadth[
            "Market_Cap_Category"
        ]
        .eq("UNCLASSIFIED")
        .sum()
    )

    summary_values = {
        "Module": MODULE_ID,
        "Version": MODULE_VERSION,
        "Generated At": (
            datetime.now()
            .astimezone()
            .isoformat()
        ),
        "Breadth Rows": len(enriched_breadth),
        "Nifty 100 Symbols": len(nifty_100),
        "Nifty 200 Symbols": len(nifty_200),
        "Nifty 500 Symbols": len(nifty_500),
        "Local F&O Symbols": len(fno_symbols),
        "Rows With Sector": sector_classified_rows,
        "Rows With Market Cap": market_cap_classified_rows,
        "Rows Marked F&O": fno_classified_rows,
        "Unclassified Rows": unclassified_rows,
        "Classification Method": (
            "Nifty 100 / Nifty 200 / Nifty 500 membership"
        ),
        "Status": "SUCCESS",
    }

    classification_summary = pd.DataFrame(
        [
            {
                "Field": key,
                "Value": value,
            }
            for key, value
            in summary_values.items()
        ]
    )

    export_classification_master(
        master=classification_master,
        nifty_100=nifty_100,
        nifty_200=nifty_200,
        nifty_500=nifty_500,
    )

    export_enriched_breadth(
        enriched_breadth=enriched_breadth,
        classification_summary=(
            classification_summary
        ),
    )

    CLASSIFICATION_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    pd.DataFrame(
        manifest_rows
    ).to_csv(
        CLASSIFICATION_MANIFEST_FILE,
        index=False,
    )

    return ClassificationBuilderResult(
        breadth_file=BREADTH_FILE,
        classification_master_file=(
            CLASSIFICATION_MASTER_FILE
        ),
        manifest_file=(
            CLASSIFICATION_MANIFEST_FILE
        ),
        breadth_rows=len(enriched_breadth),
        nifty_100_symbols=len(nifty_100),
        nifty_200_symbols=len(nifty_200),
        nifty_500_symbols=len(nifty_500),
        fno_symbols=len(fno_symbols),
        sector_classified_rows=(
            sector_classified_rows
        ),
        market_cap_classified_rows=(
            market_cap_classified_rows
        ),
        fno_classified_rows=(
            fno_classified_rows
        ),
        unclassified_rows=(
            unclassified_rows
        ),
        status="SUCCESS",
        message=(
            "NSE classification data was merged into the "
            "market breadth snapshot successfully."
        ),
    )


# ==========================================================
# DISPLAY
# ==========================================================

def display_result(
    result: ClassificationBuilderResult,
) -> None:
    """
    Display the builder result.
    """

    print()
    print("=" * 96)
    print("AQSD NSE SECURITY CLASSIFICATION BUILDER — RESULT")
    print("=" * 96)
    print(
        f"Breadth Rows                  : "
        f"{result.breadth_rows:,}"
    )
    print(
        f"Nifty 100 Symbols             : "
        f"{result.nifty_100_symbols:,}"
    )
    print(
        f"Nifty 200 Symbols             : "
        f"{result.nifty_200_symbols:,}"
    )
    print(
        f"Nifty 500 Symbols             : "
        f"{result.nifty_500_symbols:,}"
    )
    print(
        f"F&O Symbols Loaded            : "
        f"{result.fno_symbols:,}"
    )
    print("-" * 96)
    print(
        f"Rows With Sector              : "
        f"{result.sector_classified_rows:,}"
    )
    print(
        f"Rows With Market-Cap Category : "
        f"{result.market_cap_classified_rows:,}"
    )
    print(
        f"Rows Marked F&O               : "
        f"{result.fno_classified_rows:,}"
    )
    print(
        f"Unclassified Rows             : "
        f"{result.unclassified_rows:,}"
    )
    print("-" * 96)
    print(
        f"Breadth File                  : "
        f"{result.breadth_file}"
    )
    print(
        f"Classification Master         : "
        f"{result.classification_master_file}"
    )
    print(
        f"Manifest                      : "
        f"{result.manifest_file}"
    )
    print("-" * 96)
    print(
        f"Status                        : "
        f"{result.status}"
    )
    print(
        f"Message                       : "
        f"{result.message}"
    )
    print("=" * 96)


# ==========================================================
# COMMAND LINE
# ==========================================================

def parse_arguments() -> argparse.Namespace:
    """
    Read command-line arguments.
    """

    parser = argparse.ArgumentParser(
        description=(
            "Add NSE sector, industry, market-cap and F&O "
            "classification to the AQSD breadth snapshot."
        )
    )

    return parser.parse_args()


def main() -> None:
    """
    Command-line entry point.
    """

    parse_arguments()

    try:
        result = run_classification_builder()

    except Exception as exc:
        print()
        print("=" * 96)
        print("AQSD NSE SECURITY CLASSIFICATION BUILDER")
        print("=" * 96)
        print("Status : FAILED")
        print(f"Reason : {exc}")
        print("=" * 96)

        raise SystemExit(1) from exc

    display_result(
        result
    )


if __name__ == "__main__":
    main()