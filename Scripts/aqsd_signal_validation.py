
"""
AQSD Professional
Module: Signal Validation & Learning Engine
Version: 1.0

Purpose
-------
Validates AQSD historical signals against subsequent market performance.

Inputs
------
- aqsd_trade_decisions
- unified_master_intelligence
- daily_prices
- symbols

Outputs
-------
- Forward returns after 5, 10 and 20 trading days
- Maximum favourable excursion (MFE)
- Maximum adverse excursion (MAE)
- Target-1 and stop-loss hit analysis
- Win rate by AQSD action
- Win rate by Master Score band
- Confidence calibration
- SQLite validation history
- CSV reports
- Excel sheet: Signal Validation

Commands
--------
python aqsd_signal_validation.py --run
python aqsd_signal_validation.py --status
python aqsd_signal_validation.py --report
python aqsd_signal_validation.py --minimum-age 20
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aqsd_database import connect, setup_database, start_run, finish_run


BASE_DIR = Path(__file__).resolve().parent.parent
OUTPUT_DIR = BASE_DIR / "Output"
DASHBOARD = OUTPUT_DIR / "Dashboard.xlsx"

DETAIL_CSV = OUTPUT_DIR / "AQSD_Signal_Validation_Detail.csv"
SUMMARY_CSV = OUTPUT_DIR / "AQSD_Signal_Validation_Summary.csv"

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFF2CC"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9D9D9")


SCHEMA = """
CREATE TABLE IF NOT EXISTS aqsd_signal_validation (
    validation_id INTEGER PRIMARY KEY AUTOINCREMENT,
    signal_date TEXT NOT NULL,
    symbol_id INTEGER NOT NULL,
    nse_symbol TEXT NOT NULL,
    action TEXT,
    master_score REAL,
    confidence_percent REAL,
    entry_price REAL,
    stop_loss REAL,
    target_1 REAL,
    target_2 REAL,
    forward_return_5d REAL,
    forward_return_10d REAL,
    forward_return_20d REAL,
    mfe_20d_percent REAL,
    mae_20d_percent REAL,
    target_1_hit INTEGER,
    target_2_hit INTEGER,
    stop_loss_hit INTEGER,
    outcome_20d TEXT,
    evaluated_at TEXT NOT NULL,
    UNIQUE(signal_date, symbol_id)
);

CREATE TABLE IF NOT EXISTS aqsd_signal_validation_summary (
    summary_id INTEGER PRIMARY KEY AUTOINCREMENT,
    summary_date TEXT NOT NULL,
    grouping_type TEXT NOT NULL,
    grouping_value TEXT NOT NULL,
    signal_count INTEGER,
    win_rate_5d REAL,
    win_rate_10d REAL,
    win_rate_20d REAL,
    average_return_5d REAL,
    average_return_10d REAL,
    average_return_20d REAL,
    average_mfe_20d REAL,
    average_mae_20d REAL,
    target_1_hit_rate REAL,
    stop_loss_hit_rate REAL,
    created_at TEXT NOT NULL,
    UNIQUE(summary_date, grouping_type, grouping_value)
);

CREATE INDEX IF NOT EXISTS idx_signal_validation_date
ON aqsd_signal_validation(signal_date);

CREATE INDEX IF NOT EXISTS idx_signal_validation_symbol
ON aqsd_signal_validation(symbol_id, signal_date);

CREATE INDEX IF NOT EXISTS idx_signal_validation_summary
ON aqsd_signal_validation_summary(summary_date, grouping_type);
"""


def setup_schema() -> None:
    setup_database()

    with connect() as connection:
        connection.executescript(SCHEMA)
        connection.commit()


def table_exists(connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND name=?
        """,
        (table_name,),
    ).fetchone()

    return row is not None


def safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def score_band(score: float | None) -> str:
    if score is None:
        return "NO SCORE"
    if score >= 85:
        return "85-100"
    if score >= 72:
        return "72-84.99"
    if score >= 60:
        return "60-71.99"
    if score >= 48:
        return "48-59.99"
    if score >= 35:
        return "35-47.99"
    return "0-34.99"


def confidence_band(confidence: float | None) -> str:
    if confidence is None:
        return "NO CONFIDENCE"
    if confidence >= 80:
        return "80-100"
    if confidence >= 65:
        return "65-79.99"
    if confidence >= 50:
        return "50-64.99"
    return "0-49.99"


def load_decisions() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "aqsd_trade_decisions"):
            return pd.DataFrame()

        return pd.read_sql_query(
            """
            SELECT
                trade_date AS signal_date,
                symbol_id,
                nse_symbol,
                action,
                master_score,
                confidence_percent,
                last_price,
                entry_low,
                entry_high,
                stop_loss,
                target_1,
                target_2
            FROM aqsd_trade_decisions
            ORDER BY trade_date, symbol_id
            """,
            connection,
        )


def load_prices() -> pd.DataFrame:
    with connect() as connection:
        if not table_exists(connection, "daily_prices"):
            return pd.DataFrame()

        frame = pd.read_sql_query(
            """
            SELECT
                symbol_id,
                trade_date,
                open,
                high,
                low,
                close
            FROM daily_prices
            ORDER BY symbol_id, trade_date
            """,
            connection,
        )

    if not frame.empty:
        frame["trade_date"] = pd.to_datetime(
            frame["trade_date"]
        )

    return frame


def build_validation(
    minimum_age: int,
) -> pd.DataFrame:
    decisions = load_decisions()
    prices = load_prices()

    if decisions.empty:
        raise RuntimeError(
            "No historical Trade Decision records found."
        )

    if prices.empty:
        raise RuntimeError(
            "No daily price history found."
        )

    decisions["signal_date"] = pd.to_datetime(
        decisions["signal_date"]
    )

    latest_price_date = prices["trade_date"].max()

    eligible = decisions[
        decisions["signal_date"]
        <= latest_price_date - pd.Timedelta(days=minimum_age)
    ].copy()

    if eligible.empty:
        raise RuntimeError(
            "No signals are old enough to validate. "
            "Reduce --minimum-age or accumulate more daily history."
        )

    rows: list[dict] = []

    price_groups = {
        int(symbol_id): group.reset_index(drop=True)
        for symbol_id, group in prices.groupby("symbol_id")
    }

    for _, decision in eligible.iterrows():
        symbol_id = int(decision["symbol_id"])
        history = price_groups.get(symbol_id)

        if history is None or history.empty:
            continue

        after = history[
            history["trade_date"] > decision["signal_date"]
        ].reset_index(drop=True)

        if after.empty:
            continue

        entry_price = safe_float(
            decision.get("last_price")
        )

        if entry_price is None:
            entry_low = safe_float(
                decision.get("entry_low")
            )
            entry_high = safe_float(
                decision.get("entry_high")
            )

            if (
                entry_low is not None
                and entry_high is not None
            ):
                entry_price = (
                    entry_low + entry_high
                ) / 2

        if entry_price is None or entry_price <= 0:
            continue

        def forward_return(days: int) -> float | None:
            if len(after) < days:
                return None

            close = safe_float(
                after.iloc[days - 1]["close"]
            )

            if close is None:
                return None

            return round(
                (close / entry_price - 1) * 100,
                2,
            )

        window = after.head(20)

        max_high = safe_float(
            window["high"].max()
        )
        min_low = safe_float(
            window["low"].min()
        )

        mfe = (
            round(
                (max_high / entry_price - 1) * 100,
                2,
            )
            if max_high is not None
            else None
        )

        mae = (
            round(
                (min_low / entry_price - 1) * 100,
                2,
            )
            if min_low is not None
            else None
        )

        target_1 = safe_float(
            decision.get("target_1")
        )
        target_2 = safe_float(
            decision.get("target_2")
        )
        stop_loss = safe_float(
            decision.get("stop_loss")
        )

        target_1_hit = int(
            target_1 is not None
            and max_high is not None
            and max_high >= target_1
        )

        target_2_hit = int(
            target_2 is not None
            and max_high is not None
            and max_high >= target_2
        )

        stop_loss_hit = int(
            stop_loss is not None
            and min_low is not None
            and min_low <= stop_loss
        )

        return_20d = forward_return(20)

        if target_1_hit and not stop_loss_hit:
            outcome = "TARGET HIT"
        elif stop_loss_hit and not target_1_hit:
            outcome = "STOP HIT"
        elif target_1_hit and stop_loss_hit:
            outcome = "BOTH HIT"
        elif return_20d is None:
            outcome = "INCOMPLETE"
        elif return_20d > 0:
            outcome = "POSITIVE"
        elif return_20d < 0:
            outcome = "NEGATIVE"
        else:
            outcome = "FLAT"

        rows.append(
            {
                "signal_date": decision[
                    "signal_date"
                ].date().isoformat(),
                "symbol_id": symbol_id,
                "nse_symbol": str(
                    decision["nse_symbol"]
                ),
                "action": str(
                    decision.get("action") or ""
                ),
                "master_score": safe_float(
                    decision.get("master_score")
                ),
                "confidence_percent": safe_float(
                    decision.get(
                        "confidence_percent"
                    )
                ),
                "entry_price": round(
                    entry_price,
                    2,
                ),
                "stop_loss": stop_loss,
                "target_1": target_1,
                "target_2": target_2,
                "forward_return_5d": forward_return(5),
                "forward_return_10d": forward_return(10),
                "forward_return_20d": return_20d,
                "mfe_20d_percent": mfe,
                "mae_20d_percent": mae,
                "target_1_hit": target_1_hit,
                "target_2_hit": target_2_hit,
                "stop_loss_hit": stop_loss_hit,
                "outcome_20d": outcome,
            }
        )

    return pd.DataFrame(rows)


def save_validation(frame: pd.DataFrame) -> None:
    if frame.empty:
        return

    with connect() as connection:
        for _, row in frame.iterrows():
            connection.execute(
                """
                INSERT INTO aqsd_signal_validation(
                    signal_date,
                    symbol_id,
                    nse_symbol,
                    action,
                    master_score,
                    confidence_percent,
                    entry_price,
                    stop_loss,
                    target_1,
                    target_2,
                    forward_return_5d,
                    forward_return_10d,
                    forward_return_20d,
                    mfe_20d_percent,
                    mae_20d_percent,
                    target_1_hit,
                    target_2_hit,
                    stop_loss_hit,
                    outcome_20d,
                    evaluated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(signal_date, symbol_id)
                DO UPDATE SET
                    action=excluded.action,
                    master_score=excluded.master_score,
                    confidence_percent=excluded.confidence_percent,
                    entry_price=excluded.entry_price,
                    stop_loss=excluded.stop_loss,
                    target_1=excluded.target_1,
                    target_2=excluded.target_2,
                    forward_return_5d=excluded.forward_return_5d,
                    forward_return_10d=excluded.forward_return_10d,
                    forward_return_20d=excluded.forward_return_20d,
                    mfe_20d_percent=excluded.mfe_20d_percent,
                    mae_20d_percent=excluded.mae_20d_percent,
                    target_1_hit=excluded.target_1_hit,
                    target_2_hit=excluded.target_2_hit,
                    stop_loss_hit=excluded.stop_loss_hit,
                    outcome_20d=excluded.outcome_20d,
                    evaluated_at=excluded.evaluated_at
                """,
                (
                    row["signal_date"],
                    int(row["symbol_id"]),
                    row["nse_symbol"],
                    row["action"],
                    row["master_score"],
                    row["confidence_percent"],
                    row["entry_price"],
                    row["stop_loss"],
                    row["target_1"],
                    row["target_2"],
                    row["forward_return_5d"],
                    row["forward_return_10d"],
                    row["forward_return_20d"],
                    row["mfe_20d_percent"],
                    row["mae_20d_percent"],
                    int(row["target_1_hit"]),
                    int(row["target_2_hit"]),
                    int(row["stop_loss_hit"]),
                    row["outcome_20d"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()


def aggregate_summary(
    frame: pd.DataFrame,
) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()

    working = frame.copy()

    working["score_band"] = working[
        "master_score"
    ].apply(score_band)

    working["confidence_band"] = working[
        "confidence_percent"
    ].apply(confidence_band)

    summaries: list[dict] = []
    summary_date = datetime.now().date().isoformat()

    def summarise(
        grouping_type: str,
        grouping_value: str,
        group: pd.DataFrame,
    ) -> None:
        def win_rate(column: str) -> float | None:
            valid = group[column].dropna()

            if valid.empty:
                return None

            return round(
                float((valid > 0).mean() * 100),
                2,
            )

        summaries.append(
            {
                "summary_date": summary_date,
                "grouping_type": grouping_type,
                "grouping_value": grouping_value,
                "signal_count": len(group),
                "win_rate_5d": win_rate(
                    "forward_return_5d"
                ),
                "win_rate_10d": win_rate(
                    "forward_return_10d"
                ),
                "win_rate_20d": win_rate(
                    "forward_return_20d"
                ),
                "average_return_5d": round(
                    float(
                        group[
                            "forward_return_5d"
                        ].mean()
                    ),
                    2,
                )
                if group[
                    "forward_return_5d"
                ].notna().any()
                else None,
                "average_return_10d": round(
                    float(
                        group[
                            "forward_return_10d"
                        ].mean()
                    ),
                    2,
                )
                if group[
                    "forward_return_10d"
                ].notna().any()
                else None,
                "average_return_20d": round(
                    float(
                        group[
                            "forward_return_20d"
                        ].mean()
                    ),
                    2,
                )
                if group[
                    "forward_return_20d"
                ].notna().any()
                else None,
                "average_mfe_20d": round(
                    float(
                        group[
                            "mfe_20d_percent"
                        ].mean()
                    ),
                    2,
                )
                if group[
                    "mfe_20d_percent"
                ].notna().any()
                else None,
                "average_mae_20d": round(
                    float(
                        group[
                            "mae_20d_percent"
                        ].mean()
                    ),
                    2,
                )
                if group[
                    "mae_20d_percent"
                ].notna().any()
                else None,
                "target_1_hit_rate": round(
                    float(
                        group[
                            "target_1_hit"
                        ].mean()
                        * 100
                    ),
                    2,
                ),
                "stop_loss_hit_rate": round(
                    float(
                        group[
                            "stop_loss_hit"
                        ].mean()
                        * 100
                    ),
                    2,
                ),
            }
        )

    summarise(
        "OVERALL",
        "ALL SIGNALS",
        working,
    )

    for action, group in working.groupby(
        "action"
    ):
        summarise(
            "ACTION",
            str(action),
            group,
        )

    for band, group in working.groupby(
        "score_band"
    ):
        summarise(
            "MASTER SCORE BAND",
            str(band),
            group,
        )

    for band, group in working.groupby(
        "confidence_band"
    ):
        summarise(
            "CONFIDENCE BAND",
            str(band),
            group,
        )

    return pd.DataFrame(summaries)


def save_summary(frame: pd.DataFrame) -> None:
    if frame.empty:
        return

    with connect() as connection:
        for _, row in frame.iterrows():
            connection.execute(
                """
                INSERT INTO aqsd_signal_validation_summary(
                    summary_date,
                    grouping_type,
                    grouping_value,
                    signal_count,
                    win_rate_5d,
                    win_rate_10d,
                    win_rate_20d,
                    average_return_5d,
                    average_return_10d,
                    average_return_20d,
                    average_mfe_20d,
                    average_mae_20d,
                    target_1_hit_rate,
                    stop_loss_hit_rate,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(
                    summary_date,
                    grouping_type,
                    grouping_value
                )
                DO UPDATE SET
                    signal_count=excluded.signal_count,
                    win_rate_5d=excluded.win_rate_5d,
                    win_rate_10d=excluded.win_rate_10d,
                    win_rate_20d=excluded.win_rate_20d,
                    average_return_5d=
                        excluded.average_return_5d,
                    average_return_10d=
                        excluded.average_return_10d,
                    average_return_20d=
                        excluded.average_return_20d,
                    average_mfe_20d=
                        excluded.average_mfe_20d,
                    average_mae_20d=
                        excluded.average_mae_20d,
                    target_1_hit_rate=
                        excluded.target_1_hit_rate,
                    stop_loss_hit_rate=
                        excluded.stop_loss_hit_rate,
                    created_at=excluded.created_at
                """,
                (
                    row["summary_date"],
                    row["grouping_type"],
                    row["grouping_value"],
                    int(row["signal_count"]),
                    row["win_rate_5d"],
                    row["win_rate_10d"],
                    row["win_rate_20d"],
                    row["average_return_5d"],
                    row["average_return_10d"],
                    row["average_return_20d"],
                    row["average_mfe_20d"],
                    row["average_mae_20d"],
                    row["target_1_hit_rate"],
                    row["stop_loss_hit_rate"],
                    datetime.now().isoformat(
                        timespec="seconds"
                    ),
                ),
            )

        connection.commit()


def run_engine(
    minimum_age: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    setup_schema()

    run_id = start_run(
        "aqsd_signal_validation",
        (
            "Validating historical AQSD signals "
            f"with minimum age {minimum_age} days"
        ),
    )

    try:
        detail = build_validation(
            minimum_age
        )
        save_validation(detail)

        summary = aggregate_summary(
            detail
        )
        save_summary(summary)

        finish_run(
            run_id,
            status="SUCCESS",
            records_processed=len(detail),
            errors_count=0,
            message=(
                f"Validated={len(detail)}; "
                f"summary rows={len(summary)}"
            ),
        )

        return detail, summary

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=0,
            errors_count=1,
            message=str(error),
        )
        raise


def latest_results() -> tuple[pd.DataFrame, pd.DataFrame]:
    setup_schema()

    with connect() as connection:
        detail = pd.read_sql_query(
            """
            SELECT *
            FROM aqsd_signal_validation
            ORDER BY signal_date DESC, nse_symbol
            """,
            connection,
        )

        summary = pd.read_sql_query(
            """
            SELECT *
            FROM aqsd_signal_validation_summary
            WHERE summary_date=(
                SELECT MAX(summary_date)
                FROM aqsd_signal_validation_summary
            )
            ORDER BY grouping_type, grouping_value
            """,
            connection,
        )

    return detail, summary


def write_reports(
    detail: pd.DataFrame | None = None,
    summary: pd.DataFrame | None = None,
) -> None:
    if detail is None or summary is None:
        detail, summary = latest_results()

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    detail.to_csv(
        DETAIL_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    summary.to_csv(
        SUMMARY_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    if DASHBOARD.exists():
        workbook = load_workbook(
            DASHBOARD
        )
    else:
        workbook = Workbook()

        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

    sheet_name = "Signal Validation"

    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    ws = workbook.create_sheet(
        sheet_name,
        1,
    )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A8"

    ws.merge_cells("A1:P2")
    ws["A1"] = "AQSD PROFESSIONAL - SIGNAL VALIDATION & LEARNING"
    ws["A1"].font = Font(
        size=20,
        bold=True,
        color=WHITE,
    )
    ws["A1"].fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    overall = summary[
        (
            summary["grouping_type"]
            == "OVERALL"
        )
        & (
            summary["grouping_value"]
            == "ALL SIGNALS"
        )
    ]

    ws["A4"] = "Signals Validated"
    ws["B4"] = (
        int(overall.iloc[0]["signal_count"])
        if not overall.empty
        else len(detail)
    )
    ws["D4"] = "20D Win Rate %"
    ws["E4"] = (
        overall.iloc[0]["win_rate_20d"]
        if not overall.empty
        else None
    )
    ws["G4"] = "Average 20D Return %"
    ws["H4"] = (
        overall.iloc[0][
            "average_return_20d"
        ]
        if not overall.empty
        else None
    )
    ws["J4"] = "Target-1 Hit Rate %"
    ws["K4"] = (
        overall.iloc[0][
            "target_1_hit_rate"
        ]
        if not overall.empty
        else None
    )

    for ref in (
        "A4",
        "D4",
        "G4",
        "J4",
    ):
        ws[ref].font = Font(
            bold=True
        )
        ws[ref].fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

    summary_headers = [
        "Grouping Type",
        "Grouping Value",
        "Signals",
        "Win Rate 5D %",
        "Win Rate 10D %",
        "Win Rate 20D %",
        "Avg Return 5D %",
        "Avg Return 10D %",
        "Avg Return 20D %",
        "Avg MFE 20D %",
        "Avg MAE 20D %",
        "Target-1 Hit %",
        "Stop Hit %",
    ]

    for col, heading in enumerate(
        summary_headers,
        start=1,
    ):
        cell = ws.cell(7, col, heading)
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        summary.iterrows(),
        start=8,
    ):
        values = [
            row.get("grouping_type"),
            row.get("grouping_value"),
            row.get("signal_count"),
            row.get("win_rate_5d"),
            row.get("win_rate_10d"),
            row.get("win_rate_20d"),
            row.get("average_return_5d"),
            row.get("average_return_10d"),
            row.get("average_return_20d"),
            row.get("average_mfe_20d"),
            row.get("average_mae_20d"),
            row.get("target_1_hit_rate"),
            row.get("stop_loss_hit_rate"),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(
                bottom=THIN
            )

        win_rate = safe_float(
            row.get("win_rate_20d")
        )

        if win_rate is not None:
            ws.cell(
                row_no,
                6,
            ).fill = PatternFill(
                "solid",
                fgColor=(
                    GREEN
                    if win_rate >= 55
                    else RED
                    if win_rate < 45
                    else YELLOW
                ),
            )

    detail_start = max(
        14,
        len(summary) + 11,
    )

    ws.cell(
        detail_start,
        1,
        "VALIDATION DETAIL",
    ).font = Font(
        size=14,
        bold=True,
        color=WHITE,
    )
    ws.cell(
        detail_start,
        1,
    ).fill = PatternFill(
        "solid",
        fgColor=NAVY,
    )

    detail_headers = [
        "Signal Date",
        "Symbol",
        "Action",
        "Master Score",
        "Confidence %",
        "Entry Price",
        "5D Return %",
        "10D Return %",
        "20D Return %",
        "MFE 20D %",
        "MAE 20D %",
        "Target-1 Hit",
        "Target-2 Hit",
        "Stop Hit",
        "Outcome",
    ]

    for col, heading in enumerate(
        detail_headers,
        start=1,
    ):
        cell = ws.cell(
            detail_start + 2,
            col,
            heading,
        )
        cell.font = Font(
            bold=True,
            color=WHITE,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )
        cell.alignment = Alignment(
            horizontal="center",
            wrap_text=True,
        )

    for row_no, (_, row) in enumerate(
        detail.head(1000).iterrows(),
        start=detail_start + 3,
    ):
        values = [
            row.get("signal_date"),
            row.get("nse_symbol"),
            row.get("action"),
            row.get("master_score"),
            row.get("confidence_percent"),
            row.get("entry_price"),
            row.get("forward_return_5d"),
            row.get("forward_return_10d"),
            row.get("forward_return_20d"),
            row.get("mfe_20d_percent"),
            row.get("mae_20d_percent"),
            row.get("target_1_hit"),
            row.get("target_2_hit"),
            row.get("stop_loss_hit"),
            row.get("outcome_20d"),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            ws.cell(
                row_no,
                col,
                value,
            ).border = Border(
                bottom=THIN
            )

    widths = {
        "A": 18,
        "B": 20,
        "C": 18,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 17,
        "I": 17,
        "J": 16,
        "K": 16,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 18,
        "P": 18,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    workbook.save(
        DASHBOARD
    )


def show_status() -> None:
    setup_schema()

    with connect() as connection:
        detail = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                MIN(signal_date) AS first_date,
                MAX(signal_date) AS latest_date
            FROM aqsd_signal_validation
            """
        ).fetchone()

        summary = connection.execute(
            """
            SELECT COUNT(*) AS total
            FROM aqsd_signal_validation_summary
            """
        ).fetchone()

    print("\nAQSD SIGNAL VALIDATION STATUS")
    print("=" * 72)
    print(
        f"Validated signals: "
        f"{detail['total'] or 0}"
    )
    print(
        f"Summary records:   "
        f"{summary['total'] or 0}"
    )
    print(
        f"First signal date: "
        f"{detail['first_date'] or 'No data'}"
    )
    print(
        f"Latest signal date:"
        f" {detail['latest_date'] or 'No data'}"
    )
    print("=" * 72)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "AQSD Signal Validation "
            "and Learning Engine."
        )
    )

    parser.add_argument(
        "--run",
        action="store_true",
        help="Validate historical AQSD signals.",
    )

    parser.add_argument(
        "--report",
        action="store_true",
        help="Rebuild CSV and Excel reports.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show validation-engine status.",
    )

    parser.add_argument(
        "--minimum-age",
        type=int,
        default=20,
        help=(
            "Minimum calendar age of signals "
            "to evaluate. Default: 20."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_schema()

    if args.run:
        detail, summary = run_engine(
            max(1, args.minimum_age)
        )

        write_reports(
            detail,
            summary,
        )

        print("\nAQSD SIGNAL VALIDATION ENGINE")
        print("=" * 72)
        print(
            f"Signals validated: "
            f"{len(detail)}"
        )

        overall = summary[
            (
                summary["grouping_type"]
                == "OVERALL"
            )
            & (
                summary["grouping_value"]
                == "ALL SIGNALS"
            )
        ]

        if not overall.empty:
            row = overall.iloc[0]

            print(
                f"20D Win Rate:     "
                f"{row['win_rate_20d']}%"
            )
            print(
                f"Average 20D Return:"
                f" {row['average_return_20d']}%"
            )
            print(
                f"Target-1 Hit Rate:"
                f" {row['target_1_hit_rate']}%"
            )
            print(
                f"Stop-Loss Hit Rate:"
                f" {row['stop_loss_hit_rate']}%"
            )

        print(f"Detail CSV:       {DETAIL_CSV}")
        print(f"Summary CSV:      {SUMMARY_CSV}")
        print(f"Dashboard:        {DASHBOARD}")
        return

    if args.report:
        write_reports()
        print(f"Detail CSV rebuilt:\n{DETAIL_CSV}")
        print(f"Summary CSV rebuilt:\n{SUMMARY_CSV}")
        print(f"Dashboard rebuilt:\n{DASHBOARD}")
        return

    show_status()


if __name__ == "__main__":
    main()
