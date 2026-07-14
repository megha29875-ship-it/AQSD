
"""
AQSD Professional
Module: Portfolio Manager
Version: 2.0

Fixes the summary/table overlap, auto-adds the top trade,
downloads latest prices, and calculates portfolio P/L.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent.parent
DASHBOARD = BASE / "Output" / "Dashboard.xlsx"

TRADING_CAPITAL = 200000
DEFAULT_QUANTITY = 100

NAVY = "17365D"
BLUE = "D9EAF7"
GREEN = "C6EFCE"
YELLOW = "FFF2CC"
RED = "FFC7CE"
WHITE = "FFFFFF"
GREY = "E7E6E6"
INPUT = "FFFDF2"
THIN = Side(style="thin", color="D9D9D9")

HEADER_ROW = 12
FIRST_DATA_ROW = 13
LAST_DATA_ROW = 62


def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def latest_price(symbol: str) -> float | None:
    data = yf.download(
        symbol,
        period="5d",
        interval="1d",
        progress=False,
        auto_adjust=True,
    )

    if data.empty:
        return None

    if hasattr(data.columns, "nlevels") and data.columns.nlevels > 1:
        data.columns = data.columns.get_level_values(0)

    close = data["Close"].dropna()

    if close.empty:
        return None

    return float(close.iloc[-1])


def get_best_trade(wb) -> dict:
    if "Option Buying" not in wb.sheetnames:
        raise RuntimeError("Option Buying sheet not found.")

    ws = wb["Option Buying"]
    headers = header_map(ws, 1)

    required = [
        "Symbol",
        "Entry",
        "Stop Loss",
        "Target 1",
    ]

    missing = [name for name in required if name not in headers]

    if missing:
        raise RuntimeError(
            "Missing Option Buying columns: " + ", ".join(missing)
        )

    if ws.max_row < 2:
        raise RuntimeError("Option Buying sheet has no trades.")

    recommendation_col = (
        headers.get("Recommendation")
        or headers.get("Action")
    )

    recommendation = (
        str(ws.cell(2, recommendation_col).value or "")
        if recommendation_col
        else ""
    )

    side = "PUT" if "PUT" in recommendation.upper() else "CALL"

    return {
        "Symbol": str(ws.cell(2, headers["Symbol"]).value),
        "Side": side,
        "Entry": float(ws.cell(2, headers["Entry"]).value),
        "Stop Loss": float(ws.cell(2, headers["Stop Loss"]).value),
        "Target": float(ws.cell(2, headers["Target 1"]).value),
    }


def create_clean_portfolio(wb, best_trade: dict) -> None:
    if "Portfolio" in wb.sheetnames:
        del wb["Portfolio"]

    ws = wb.create_sheet("Portfolio")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{FIRST_DATA_ROW}"

    ws.merge_cells("A1:N2")
    ws["A1"] = "AQSD PROFESSIONAL - PORTFOLIO MANAGER"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    summary = [
        ("Trading Capital", TRADING_CAPITAL),
        ("Capital Invested", 0),
        ("Available Cash", TRADING_CAPITAL),
        ("Open Positions", 0),
        ("Total P/L", 0),
        ("Portfolio Return %", 0),
    ]

    for row_no, (label, value) in enumerate(summary, start=4):
        ws[f"A{row_no}"] = label
        ws[f"B{row_no}"] = value
        ws[f"A{row_no}"].font = Font(bold=True)
        ws[f"A{row_no}"].fill = PatternFill("solid", fgColor=BLUE)

    for cell_ref in ("B4", "B5", "B6", "B8"):
        ws[cell_ref].number_format = '₹#,##0.00'

    ws["B9"].number_format = '0.00"%"'

    ws["D4"] = "Last Price Update"
    ws["E4"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["D4"].font = Font(bold=True)
    ws["D4"].fill = PatternFill("solid", fgColor=BLUE)

    headers = [
        "Trade ID",
        "Symbol",
        "Side",
        "Entry Date",
        "Qty",
        "Entry",
        "CMP",
        "Stop Loss",
        "Capital Used",
        "Risk Amount",
        "P/L",
        "P/L %",
        "Target",
        "Status",
    ]

    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(HEADER_ROW, col, heading)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=THIN)

    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        ws.cell(row, 1, row - FIRST_DATA_ROW + 1)

        for col in range(1, 15):
            ws.cell(row, col).border = Border(bottom=THIN)

        for col in [2, 3, 4, 5, 6, 7, 8, 13]:
            ws.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)

        for col in [9, 10, 11, 12, 14]:
            ws.cell(row, col).fill = PatternFill("solid", fgColor=GREY)

    # Auto-add the top AQSD trade into the first row
    row = FIRST_DATA_ROW

    ws.cell(row, 2, best_trade["Symbol"])
    ws.cell(row, 3, best_trade["Side"])
    ws.cell(row, 4, datetime.now())
    ws.cell(row, 5, DEFAULT_QUANTITY)
    ws.cell(row, 6, best_trade["Entry"])
    ws.cell(row, 8, best_trade["Stop Loss"])
    ws.cell(row, 13, best_trade["Target"])

    widths = {
        "A": 10,
        "B": 18,
        "C": 12,
        "D": 14,
        "E": 10,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 15,
        "J": 15,
        "K": 14,
        "L": 12,
        "M": 12,
        "N": 12,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.auto_filter.ref = f"A{HEADER_ROW}:N{LAST_DATA_ROW}"


def update_portfolio(wb) -> None:
    ws = wb["Portfolio"]
    headers = header_map(ws, HEADER_ROW)

    total_invested = 0.0
    total_pl = 0.0
    open_positions = 0
    updated = 0

    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        symbol = str(ws.cell(row, headers["Symbol"]).value or "").strip()

        if not symbol:
            continue

        if not symbol.endswith(".NS"):
            symbol += ".NS"
            ws.cell(row, headers["Symbol"]).value = symbol

        qty_value = ws.cell(row, headers["Qty"]).value
        entry_value = ws.cell(row, headers["Entry"]).value

        if not qty_value or not entry_value:
            continue

        qty = float(qty_value)
        entry = float(entry_value)
        side = str(
            ws.cell(row, headers["Side"]).value or "CALL"
        ).upper().strip()

        stop_value = ws.cell(row, headers["Stop Loss"]).value
        target_value = ws.cell(row, headers["Target"]).value

        stop_loss = (
            float(stop_value)
            if stop_value not in (None, "")
            else None
        )

        target = (
            float(target_value)
            if target_value not in (None, "")
            else None
        )

        cmp_price = latest_price(symbol)

        if cmp_price is None:
            ws.cell(row, headers["Status"]).value = "PRICE ERROR"
            ws.cell(row, headers["Status"]).fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )
            continue

        capital_used = qty * entry

        risk_amount = (
            qty * abs(entry - stop_loss)
            if stop_loss is not None
            else 0.0
        )

        if side in {"BUY", "CALL"}:
            profit_loss = (cmp_price - entry) * qty
        else:
            profit_loss = (entry - cmp_price) * qty

        pl_percent = (
            profit_loss / capital_used * 100
            if capital_used
            else 0.0
        )

        status = "OPEN"

        if stop_loss is not None and target is not None:
            if side in {"BUY", "CALL"}:
                if cmp_price <= stop_loss:
                    status = "STOPPED"
                elif cmp_price >= target:
                    status = "TARGET"
            else:
                if cmp_price >= stop_loss:
                    status = "STOPPED"
                elif cmp_price <= target:
                    status = "TARGET"

        ws.cell(row, headers["CMP"]).value = round(cmp_price, 2)
        ws.cell(row, headers["Capital Used"]).value = round(capital_used, 2)
        ws.cell(row, headers["Risk Amount"]).value = round(risk_amount, 2)
        ws.cell(row, headers["P/L"]).value = round(profit_loss, 2)
        ws.cell(row, headers["P/L %"]).value = round(pl_percent, 2)
        ws.cell(row, headers["Status"]).value = status

        for col_name in [
            "Entry",
            "CMP",
            "Stop Loss",
            "Capital Used",
            "Risk Amount",
            "P/L",
            "Target",
        ]:
            ws.cell(row, headers[col_name]).number_format = '₹#,##0.00'

        ws.cell(row, headers["P/L %"]).number_format = '0.00"%"'
        ws.cell(row, headers["Entry Date"]).number_format = "dd-mm-yyyy"

        pl_fill = GREEN if profit_loss >= 0 else RED

        ws.cell(row, headers["P/L"]).fill = PatternFill(
            "solid",
            fgColor=pl_fill,
        )
        ws.cell(row, headers["P/L %"]).fill = PatternFill(
            "solid",
            fgColor=pl_fill,
        )

        if status == "TARGET":
            status_fill = GREEN
        elif status == "STOPPED":
            status_fill = RED
        else:
            status_fill = YELLOW

        ws.cell(row, headers["Status"]).fill = PatternFill(
            "solid",
            fgColor=status_fill,
        )
        ws.cell(row, headers["Status"]).font = Font(bold=True)

        total_invested += capital_used
        total_pl += profit_loss

        if status == "OPEN":
            open_positions += 1

        updated += 1

    trading_capital = float(ws["B4"].value or TRADING_CAPITAL)
    available_cash = trading_capital - total_invested

    portfolio_return = (
        total_pl / total_invested * 100
        if total_invested
        else 0.0
    )

    ws["B5"] = round(total_invested, 2)
    ws["B6"] = round(available_cash, 2)
    ws["B7"] = open_positions
    ws["B8"] = round(total_pl, 2)
    ws["B9"] = round(portfolio_return, 2)

    ws["E4"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    print(f"Positions updated: {updated}")
    print(f"Capital invested: ₹{total_invested:,.2f}")
    print(f"Total P/L: ₹{total_pl:,.2f}")


def main() -> None:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    wb = load_workbook(DASHBOARD)

    best_trade = get_best_trade(wb)
    create_clean_portfolio(wb, best_trade)
    update_portfolio(wb)

    wb.save(DASHBOARD)

    print("Portfolio Manager v2 created successfully.")
    print(f"Auto-added: {best_trade['Symbol']}")
    print(DASHBOARD)


if __name__ == "__main__":
    main()
