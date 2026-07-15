
"""
AQSD Core
Module: Intelligence History Recorder
Version: 1.0

Reads the latest intelligence sheets from Dashboard.xlsx and stores one
daily intelligence snapshot per symbol in aqsd_core.db.

Supported sheets
----------------
- Market Structure
- Trend Intelligence
- Relative Strength
- Sector Rotation
- Pivot Intelligence
- Master Intelligence

Creates historical memory for:
- Structure Score
- Trend Score
- Relative Strength Score
- Sector Score
- Pivot Score
- Master Score
- Directional Bias
- Recommendation
- Confidence Grade
- Explanation

Commands
--------
python aqsd_intelligence_recorder.py --record
python aqsd_intelligence_recorder.py --record --date 2026-07-15
python aqsd_intelligence_recorder.py --status
python aqsd_intelligence_recorder.py --history RELIANCE --days 30
python aqsd_intelligence_recorder.py --improving --days 5
python aqsd_intelligence_recorder.py --changes
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from aqsd_database import connect, setup_database, start_run, finish_run


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent
DASHBOARD = BASE_DIR / "Output" / "Dashboard.xlsx"


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_nse_symbol(value: object) -> str:
    symbol = str(value or "").strip().upper()

    if symbol.endswith(".NS"):
        symbol = symbol[:-3]

    return symbol.replace(" ", "")


def safe_float(value: object) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def header_map(ws, row_number: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[row_number]
        if cell.value is not None
    }


def resolve_score_date(value: str | None) -> str:
    if not value:
        return date.today().isoformat()

    parsed = datetime.strptime(value, "%Y-%m-%d")
    return parsed.date().isoformat()


# ============================================================
# SYMBOL LOOKUP
# ============================================================

def get_symbol_ids() -> dict[str, int]:
    setup_database()

    with connect() as connection:
        rows = connection.execute(
            """
            SELECT symbol_id, nse_symbol
            FROM symbols
            """
        ).fetchall()

    return {
        str(row["nse_symbol"]).upper(): int(row["symbol_id"])
        for row in rows
    }


# ============================================================
# SHEET READERS
# ============================================================

def read_generic_score_sheet(
    wb,
    sheet_name: str,
    header_row: int,
    score_header: str,
    extra_headers: list[str],
) -> dict[str, dict]:
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    headers = header_map(ws, header_row)

    if "Symbol" not in headers or score_header not in headers:
        return {}

    output: dict[str, dict] = {}

    for row_number in range(header_row + 1, ws.max_row + 1):
        symbol = normalize_nse_symbol(
            ws.cell(row_number, headers["Symbol"]).value
        )

        if not symbol:
            continue

        score = safe_float(
            ws.cell(row_number, headers[score_header]).value
        )

        if score is None:
            continue

        record = {"Score": score}

        for header in extra_headers:
            record[header] = (
                ws.cell(row_number, headers[header]).value
                if header in headers
                else ""
            )

        output[symbol] = record

    return output


def read_sector_rotation(wb) -> dict[str, dict]:
    if "Sector Rotation" not in wb.sheetnames:
        return {}

    ws = wb["Sector Rotation"]
    headers = header_map(ws, 4)

    if "Sector" not in headers or "Rotation Score" not in headers:
        return {}

    raw = []

    for row_number in range(5, ws.max_row + 1):
        sector = str(
            ws.cell(row_number, headers["Sector"]).value or ""
        ).strip()

        score = safe_float(
            ws.cell(row_number, headers["Rotation Score"]).value
        )

        if sector and score is not None:
            raw.append((sector.upper(), score))

    if not raw:
        return {}

    values = [item[1] for item in raw]
    minimum = min(values)
    maximum = max(values)

    output = {}

    for rank, (sector, raw_score) in enumerate(
        sorted(raw, key=lambda item: item[1], reverse=True),
        start=1,
    ):
        normalized = (
            50.0
            if maximum == minimum
            else (raw_score - minimum) / (maximum - minimum) * 100
        )

        output[sector] = {
            "Score": round(normalized, 2),
            "Rank": rank,
            "Raw Score": raw_score,
        }

    return output


# ============================================================
# MASTER RECORDING
# ============================================================

def build_records(wb) -> list[dict]:
    structure = read_generic_score_sheet(
        wb,
        "Market Structure",
        6,
        "Structure Score",
        ["Structure Event", "Reason"],
    )

    trend = read_generic_score_sheet(
        wb,
        "Trend Intelligence",
        6,
        "Trend Score",
        ["Trend Regime", "Reason"],
    )

    relative_strength = read_generic_score_sheet(
        wb,
        "Relative Strength",
        6,
        "Relative Strength Score",
        ["Classification", "Reason"],
    )

    pivots = read_generic_score_sheet(
        wb,
        "Pivot Intelligence",
        6,
        "Pivot Score",
        ["Pivot Bias", "CPR Type", "CPR Position", "Reason"],
    )

    master = read_generic_score_sheet(
        wb,
        "Master Intelligence",
        7,
        "AQSD Master Score",
        [
            "Sector",
            "Sector Score",
            "Directional Bias",
            "Recommendation",
            "Confidence Grade",
            "Explanation",
        ],
    )

    sectors = read_sector_rotation(wb)

    symbols = sorted(
        set(structure)
        | set(trend)
        | set(relative_strength)
        | set(pivots)
        | set(master)
    )

    records = []

    for symbol in symbols:
        master_row = master.get(symbol, {})
        sector_name = str(master_row.get("Sector") or "").strip()
        sector_row = sectors.get(sector_name.upper(), {})

        explanation_parts = []

        for label, source in (
            ("Structure", structure),
            ("Trend", trend),
            ("Relative Strength", relative_strength),
            ("Pivot", pivots),
        ):
            reason = source.get(symbol, {}).get("Reason")

            if reason:
                explanation_parts.append(f"{label}: {reason}")

        master_explanation = master_row.get("Explanation")

        if master_explanation:
            explanation_parts.append(
                f"Master: {master_explanation}"
            )

        records.append(
            {
                "symbol": symbol,
                "structure_score": structure.get(symbol, {}).get("Score"),
                "trend_score": trend.get(symbol, {}).get("Score"),
                "relative_strength_score": relative_strength.get(
                    symbol,
                    {},
                ).get("Score"),
                "sector_score": (
                    safe_float(master_row.get("Sector Score"))
                    if master_row
                    else sector_row.get("Score")
                ),
                "pivot_score": pivots.get(symbol, {}).get("Score"),
                "master_score": master_row.get("Score"),
                "directional_bias": str(
                    master_row.get("Directional Bias") or ""
                ),
                "recommendation": str(
                    master_row.get("Recommendation") or ""
                ),
                "confidence_grade": str(
                    master_row.get("Confidence Grade") or ""
                ),
                "explanation": " || ".join(explanation_parts),
            }
        )

    return records


def upsert_intelligence_record(
    connection,
    symbol_id: int,
    score_date: str,
    record: dict,
) -> None:
    connection.execute(
        """
        INSERT INTO intelligence_scores(
            score_date,
            symbol_id,
            structure_score,
            trend_score,
            relative_strength_score,
            sector_score,
            pivot_score,
            master_score,
            directional_bias,
            recommendation,
            confidence_grade,
            explanation,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(score_date, symbol_id)
        DO UPDATE SET
            structure_score = excluded.structure_score,
            trend_score = excluded.trend_score,
            relative_strength_score =
                excluded.relative_strength_score,
            sector_score = excluded.sector_score,
            pivot_score = excluded.pivot_score,
            master_score = excluded.master_score,
            directional_bias = excluded.directional_bias,
            recommendation = excluded.recommendation,
            confidence_grade = excluded.confidence_grade,
            explanation = excluded.explanation,
            created_at = excluded.created_at
        """,
        (
            score_date,
            symbol_id,
            record["structure_score"],
            record["trend_score"],
            record["relative_strength_score"],
            record["sector_score"],
            record["pivot_score"],
            record["master_score"],
            record["directional_bias"],
            record["recommendation"],
            record["confidence_grade"],
            record["explanation"],
            datetime.now().isoformat(timespec="seconds"),
        ),
    )


def record_dashboard(score_date: str) -> tuple[int, int]:
    if not DASHBOARD.exists():
        raise FileNotFoundError(
            f"Dashboard not found:\n{DASHBOARD}"
        )

    run_id = start_run(
        "aqsd_intelligence_recorder",
        f"Recording dashboard intelligence for {score_date}",
    )

    symbol_ids = get_symbol_ids()
    wb = load_workbook(
        DASHBOARD,
        read_only=True,
        data_only=True,
    )

    stored = 0
    skipped = 0

    try:
        records = build_records(wb)

        with connect() as connection:
            for record in records:
                symbol_id = symbol_ids.get(record["symbol"])

                if symbol_id is None:
                    skipped += 1
                    print(
                        f"Skipped unknown symbol: "
                        f"{record['symbol']}"
                    )
                    continue

                upsert_intelligence_record(
                    connection,
                    symbol_id,
                    score_date,
                    record,
                )
                stored += 1

            connection.commit()

        finish_run(
            run_id,
            status="SUCCESS" if skipped == 0 else "PARTIAL",
            records_processed=stored,
            errors_count=skipped,
            message=(
                f"Stored={stored}; skipped={skipped}; "
                f"date={score_date}"
            ),
        )

        return stored, skipped

    except Exception as error:
        finish_run(
            run_id,
            status="FAILED",
            records_processed=stored,
            errors_count=skipped + 1,
            message=str(error),
        )
        raise

    finally:
        wb.close()


# ============================================================
# REPORTS
# ============================================================

def show_status() -> None:
    setup_database()

    with connect() as connection:
        summary = connection.execute(
            """
            SELECT
                COUNT(*) AS rows_count,
                COUNT(DISTINCT symbol_id) AS symbol_count,
                MIN(score_date) AS first_date,
                MAX(score_date) AS latest_date
            FROM intelligence_scores
            """
        ).fetchone()

        latest_count = connection.execute(
            """
            SELECT COUNT(*)
            FROM intelligence_scores
            WHERE score_date = (
                SELECT MAX(score_date)
                FROM intelligence_scores
            )
            """
        ).fetchone()[0]

    print("\nAQSD INTELLIGENCE HISTORY STATUS")
    print("=" * 72)
    print(f"Stored records:       {summary['rows_count'] or 0}")
    print(f"Symbols with history: {summary['symbol_count'] or 0}")
    print(f"First score date:     {summary['first_date'] or 'No data'}")
    print(f"Latest score date:    {summary['latest_date'] or 'No data'}")
    print(f"Latest-day records:   {latest_count or 0}")
    print("=" * 72)


def get_history(
    symbol: str,
    days: int,
) -> pd.DataFrame:
    nse_symbol = normalize_nse_symbol(symbol)
    cutoff = (
        date.today() - timedelta(days=max(days, 1))
    ).isoformat()

    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                i.score_date,
                s.nse_symbol,
                i.structure_score,
                i.trend_score,
                i.relative_strength_score,
                i.sector_score,
                i.pivot_score,
                i.master_score,
                i.directional_bias,
                i.recommendation,
                i.confidence_grade
            FROM intelligence_scores i
            JOIN symbols s
                ON s.symbol_id = i.symbol_id
            WHERE s.nse_symbol = ?
              AND i.score_date >= ?
            ORDER BY i.score_date
            """,
            connection,
            params=(nse_symbol, cutoff),
        )

    return frame


def get_improving(
    days: int,
    minimum_change: float,
) -> pd.DataFrame:
    cutoff = (
        date.today() - timedelta(days=max(days, 2) + 10)
    ).isoformat()

    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                s.nse_symbol,
                s.sector,
                i.score_date,
                i.master_score
            FROM intelligence_scores i
            JOIN symbols s
                ON s.symbol_id = i.symbol_id
            WHERE i.score_date >= ?
              AND i.master_score IS NOT NULL
              AND s.active = 1
            ORDER BY s.nse_symbol, i.score_date
            """,
            connection,
            params=(cutoff,),
        )

    if frame.empty:
        return frame

    results = []

    for symbol, group in frame.groupby("nse_symbol"):
        group = group.sort_values("score_date").tail(days)

        if len(group) < 2:
            continue

        first_score = float(group["master_score"].iloc[0])
        latest_score = float(group["master_score"].iloc[-1])
        change = latest_score - first_score

        if change >= minimum_change:
            results.append(
                {
                    "Symbol": symbol,
                    "Sector": group["sector"].iloc[-1],
                    "First Score": round(first_score, 2),
                    "Latest Score": round(latest_score, 2),
                    "Change": round(change, 2),
                    "Observations": len(group),
                }
            )

    if not results:
        return pd.DataFrame()

    return pd.DataFrame(results).sort_values(
        "Change",
        ascending=False,
    )


def get_recommendation_changes() -> pd.DataFrame:
    with connect() as connection:
        frame = pd.read_sql_query(
            """
            SELECT
                s.nse_symbol,
                s.sector,
                i.score_date,
                i.master_score,
                i.recommendation
            FROM intelligence_scores i
            JOIN symbols s
                ON s.symbol_id = i.symbol_id
            WHERE s.active = 1
            ORDER BY s.nse_symbol, i.score_date
            """,
            connection,
        )

    if frame.empty:
        return frame

    results = []

    for symbol, group in frame.groupby("nse_symbol"):
        group = group.sort_values("score_date").tail(2)

        if len(group) < 2:
            continue

        previous = str(group["recommendation"].iloc[0] or "")
        current = str(group["recommendation"].iloc[1] or "")

        if previous != current:
            results.append(
                {
                    "Symbol": symbol,
                    "Sector": group["sector"].iloc[-1],
                    "Previous Date": group["score_date"].iloc[0],
                    "Previous Recommendation": previous,
                    "Current Date": group["score_date"].iloc[1],
                    "Current Recommendation": current,
                    "Current Score": group["master_score"].iloc[1],
                }
            )

    return pd.DataFrame(results)


def print_frame(frame: pd.DataFrame) -> None:
    if frame.empty:
        print("No records found.")
        return

    with pd.option_context(
        "display.max_rows",
        100,
        "display.max_columns",
        30,
        "display.width",
        180,
    ):
        print(frame.to_string(index=False))


# ============================================================
# CLI
# ============================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Record and analyse AQSD intelligence history."
    )

    parser.add_argument(
        "--record",
        action="store_true",
        help="Record the latest Dashboard intelligence.",
    )

    parser.add_argument(
        "--date",
        help="Score date in YYYY-MM-DD format.",
    )

    parser.add_argument(
        "--status",
        action="store_true",
        help="Show intelligence-history status.",
    )

    parser.add_argument(
        "--history",
        metavar="SYMBOL",
        help="Show intelligence history for one symbol.",
    )

    parser.add_argument(
        "--days",
        type=int,
        default=30,
        help="Lookback days for history or improving report.",
    )

    parser.add_argument(
        "--improving",
        action="store_true",
        help="Show stocks whose Master Score is improving.",
    )

    parser.add_argument(
        "--minimum-change",
        type=float,
        default=5.0,
        help="Minimum score improvement.",
    )

    parser.add_argument(
        "--changes",
        action="store_true",
        help="Show latest recommendation changes.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    setup_database()

    if args.record:
        score_date = resolve_score_date(args.date)
        stored, skipped = record_dashboard(score_date)

        print("\nAQSD INTELLIGENCE RECORDER")
        print("=" * 72)
        print(f"Score date:     {score_date}")
        print(f"Records stored: {stored}")
        print(f"Records skipped:{skipped}")
        print("=" * 72)
        return

    if args.history:
        print_frame(
            get_history(
                args.history,
                args.days,
            )
        )
        return

    if args.improving:
        print_frame(
            get_improving(
                args.days,
                args.minimum_change,
            )
        )
        return

    if args.changes:
        print_frame(
            get_recommendation_changes()
        )
        return

    show_status()


if __name__ == "__main__":
    main()
