"""
AQSD
Sector Strength Engine

Module : MBI-006
Version: 1.0.0
Author : AQSD

Description
-----------
Analyses NSE sector participation using the enriched AQSD market
breadth snapshot.

The engine calculates:

- Sector advance and decline participation
- EMA20 breadth
- EMA50 breadth
- EMA200 breadth
- Volume participation
- 52-week positioning
- Sector breadth score
- Sector strength
- Sector trend
- Sector momentum
- Sector leadership status
- Improving or deteriorating participation
- Sector rank
- Confidence
- Expected behaviour
- Explanation

Inputs
------
Data/Market_Breadth/market_breadth_snapshot.xlsx

Optional historical comparison
------------------------------
Output/Market_Breadth/market_breadth_report_YYYYMMDD.xlsx

When previous dated breadth reports are available, the engine compares
the current sector score with the previous saved sector score.

Important
---------
This module performs analytical decision support only.

It does not generate BUY, SELL or SHORT instructions.

Relative strength against NIFTY or BANKNIFTY is not calculated in this
version because the current breadth snapshot does not contain benchmark
price history.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Final

import pandas as pd


# ==========================================================
# MODULE SETTINGS
# ==========================================================

MODULE_ID: Final[str] = "MBI-006"
MODULE_VERSION: Final[str] = "1.0.0"

BASE_DIR: Final[Path] = Path(__file__).resolve().parents[2]

DEFAULT_INPUT_FILE: Final[Path] = (
    BASE_DIR
    / "Data"
    / "Market_Breadth"
    / "market_breadth_snapshot.xlsx"
)

BREADTH_REPORT_DIRECTORY: Final[Path] = (
    BASE_DIR
    / "Output"
    / "Market_Breadth"
)

OUTPUT_DIRECTORY: Final[Path] = (
    BASE_DIR
    / "Output"
    / "Sector_Strength"
)

REQUIRED_COLUMNS: Final[tuple[str, ...]] = (
    "Symbol",
    "Close",
    "Previous_Close",
    "EMA20",
    "EMA50",
    "EMA200",
    "High_52W",
    "Low_52W",
    "Sector",
)

MINIMUM_SECTOR_STOCKS: Final[int] = 3

NEAR_HIGH_MULTIPLIER: Final[float] = 0.95
NEAR_LOW_MULTIPLIER: Final[float] = 1.05


# ==========================================================
# RESULT MODELS
# ==========================================================

@dataclass(frozen=True)
class SectorStrengthResult:
    """
    Analytical result for one sector.
    """

    rank: int
    sector: str

    total_stocks: int
    advances: int
    declines: int
    unchanged: int

    advance_percentage: float
    decline_percentage: float
    advance_decline_ratio: float | None

    above_ema20_count: int
    above_ema50_count: int
    above_ema200_count: int

    above_ema20_percentage: float
    above_ema50_percentage: float
    above_ema200_percentage: float

    near_52w_high_count: int
    near_52w_low_count: int

    near_52w_high_percentage: float
    near_52w_low_percentage: float

    high_volume_advances: int
    high_volume_declines: int
    volume_breadth_ratio: float | None
    volume_participation: str

    fno_stocks: int
    fno_percentage: float

    breadth_score: int
    previous_score: int | None
    score_change: int | None

    strength: str
    trend: str
    momentum: str
    rotation: str
    leadership: str

    risk_level: str
    confidence: int

    expected_behaviour: str
    concise_summary: str
    explanation: str

    warnings: tuple[str, ...]
    status: str


@dataclass(frozen=True)
class SectorStrengthEngineResult:
    """
    Complete MBI-006 result.
    """

    requested_date: date
    analysis_date: date
    source_file: Path
    previous_reference_date: date | None

    total_input_rows: int
    valid_rows: int
    invalid_rows: int

    sectors_analysed: int
    bullish_sectors: int
    bearish_sectors: int
    neutral_sectors: int

    strongest_sector: str | None
    weakest_sector: str | None

    broad_rotation: str
    market_sector_health: str

    confidence: int
    expected_behaviour: str
    concise_summary: str
    explanation: str

    sectors: tuple[SectorStrengthResult, ...]
    warnings: tuple[str, ...]

    csv_file: Path | None
    excel_file: Path | None

    status: str


# ==========================================================
# GENERAL HELPERS
# ==========================================================

def clamp_score(
    value: float,
) -> int:
    """
    Restrict a score to 0-100.
    """

    return max(
        0,
        min(
            round(value),
            100,
        ),
    )


def safe_percentage(
    numerator: float,
    denominator: float,
) -> float:
    """
    Calculate a percentage safely.
    """

    if denominator == 0:
        return 0.0

    return round(
        numerator / denominator * 100,
        2,
    )


def safe_ratio(
    numerator: float,
    denominator: float,
) -> float | None:
    """
    Calculate a ratio safely.
    """

    if denominator == 0:
        return None

    return round(
        numerator / denominator,
        2,
    )


def format_optional_number(
    value: int | float | None,
    decimals: int = 2,
    suffix: str = "",
) -> str:
    """
    Format an optional value.
    """

    if value is None:
        return "NOT AVAILABLE"

    if isinstance(value, int):
        return f"{value:+d}{suffix}"

    return f"{value:+.{decimals}f}{suffix}"


def normalize_header(
    value: object,
) -> str:
    """
    Normalize a column heading.
    """

    text = str(value).strip()

    text = re.sub(
        r"[^A-Za-z0-9]+",
        "_",
        text,
    )

    text = re.sub(
        r"_+",
        "_",
        text,
    )

    return text.strip("_")


def normalize_boolean(
    value: object,
) -> bool:
    """
    Convert common truth markers into Boolean values.
    """

    if isinstance(value, bool):
        return value

    if value is None:
        return False

    try:
        if pd.isna(value):
            return False
    except TypeError:
        pass

    return str(value).strip().upper() in {
        "TRUE",
        "YES",
        "Y",
        "1",
        "F&O",
        "FNO",
    }


# ==========================================================
# INPUT READING
# ==========================================================

def consolidate_duplicate_columns(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Consolidate duplicate columns safely.
    """

    if dataframe.columns.is_unique:
        return dataframe.copy()

    result = pd.DataFrame(
        index=dataframe.index
    )

    handled: set[str] = set()

    for column in dataframe.columns:
        column_name = str(column)

        if column_name in handled:
            continue

        handled.add(
            column_name
        )

        positions = [
            position
            for position, existing
            in enumerate(dataframe.columns)
            if str(existing) == column_name
        ]

        if len(positions) == 1:
            result[column_name] = dataframe.iloc[
                :,
                positions[0],
            ]

            continue

        duplicate_data = dataframe.iloc[
            :,
            positions,
        ]

        combined = duplicate_data.iloc[
            :,
            0,
        ].copy()

        for position in range(
            1,
            duplicate_data.shape[1],
        ):
            candidate = duplicate_data.iloc[
                :,
                position,
            ]

            current_empty = (
                combined.isna()
                | combined.astype(str)
                .str.strip()
                .str.upper()
                .isin(
                    {
                        "",
                        "NAN",
                        "NONE",
                        "UNKNOWN",
                        "UNCLASSIFIED",
                    }
                )
            )

            candidate_valid = (
                candidate.notna()
                & ~candidate.astype(str)
                .str.strip()
                .str.upper()
                .isin(
                    {
                        "",
                        "NAN",
                        "NONE",
                    }
                )
            )

            replacement_mask = (
                current_empty
                & candidate_valid
            )

            combined.loc[
                replacement_mask
            ] = candidate.loc[
                replacement_mask
            ]

        result[column_name] = combined

    return result


def normalize_columns(
    dataframe: pd.DataFrame,
) -> pd.DataFrame:
    """
    Normalize headings and known aliases.
    """

    result = dataframe.copy()

    result.columns = [
        normalize_header(column)
        for column in result.columns
    ]

    aliases = {
        "SYMBOL": "Symbol",
        "TICKER": "Symbol",
        "STOCK": "Symbol",

        "CLOSE": "Close",
        "LTP": "Close",

        "PREVIOUS_CLOSE": "Previous_Close",
        "PREV_CLOSE": "Previous_Close",

        "EMA20": "EMA20",
        "EMA_20": "EMA20",

        "EMA50": "EMA50",
        "EMA_50": "EMA50",

        "EMA200": "EMA200",
        "EMA_200": "EMA200",

        "HIGH_52W": "High_52W",
        "52W_HIGH": "High_52W",
        "52_WEEK_HIGH": "High_52W",

        "LOW_52W": "Low_52W",
        "52W_LOW": "Low_52W",
        "52_WEEK_LOW": "Low_52W",

        "SECTOR": "Sector",
        "INDUSTRY": "Industry",

        "VOLUME": "Volume",

        "AVERAGE_VOLUME_20": "Average_Volume_20",
        "AVG_VOLUME_20": "Average_Volume_20",

        "IS_FNO": "Is_FnO",
        "IS_F_O": "Is_FnO",
        "FNO": "Is_FnO",

        "MARKET_CAP_CATEGORY": "Market_Cap_Category",
    }

    rename_map: dict[str, str] = {}

    for column in result.columns:
        key = str(column).upper()

        if key in aliases:
            rename_map[str(column)] = aliases[
                key
            ]

    result = result.rename(
        columns=rename_map
    )

    return consolidate_duplicate_columns(
        result
    )


def read_sector_snapshot(
    source_file: Path,
) -> pd.DataFrame:
    """
    Read the enriched breadth snapshot.
    """

    if not source_file.exists():
        raise FileNotFoundError(
            f"Sector-strength input file not found: {source_file}"
        )

    suffix = source_file.suffix.lower()

    if suffix == ".csv":
        dataframe = pd.read_csv(
            source_file,
            low_memory=False,
        )

    elif suffix in {
        ".xlsx",
        ".xlsm",
    }:
        try:
            dataframe = pd.read_excel(
                source_file,
                sheet_name="Breadth Snapshot",
                engine="openpyxl",
            )

        except ValueError:
            dataframe = pd.read_excel(
                source_file,
                sheet_name=0,
                engine="openpyxl",
            )

    else:
        raise ValueError(
            "Supported input formats are CSV, XLSX and XLSM."
        )

    dataframe = dataframe.dropna(
        how="all"
    ).reset_index(
        drop=True
    )

    dataframe = normalize_columns(
        dataframe
    )

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in dataframe.columns
    ]

    if missing_columns:
        raise KeyError(
            "Sector-strength input is missing required columns: "
            + ", ".join(missing_columns)
        )

    return dataframe


# ==========================================================
# DATA PREPARATION
# ==========================================================

def prepare_snapshot(
    dataframe: pd.DataFrame,
) -> tuple[pd.DataFrame, int]:
    """
    Prepare stock-level data for sector analysis.
    """

    prepared = consolidate_duplicate_columns(
        dataframe
    )

    numeric_columns = (
        "Close",
        "Previous_Close",
        "EMA20",
        "EMA50",
        "EMA200",
        "High_52W",
        "Low_52W",
        "Volume",
        "Average_Volume_20",
    )

    for column in numeric_columns:
        if column in prepared.columns:
            prepared[column] = pd.to_numeric(
                prepared[column],
                errors="coerce",
            )

    prepared["Symbol"] = (
        prepared["Symbol"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    prepared["Sector"] = (
        prepared["Sector"]
        .fillna("UNKNOWN")
        .astype(str)
        .str.strip()
        .str.upper()
        .replace(
            {
                "": "UNKNOWN",
                "NAN": "UNKNOWN",
                "NONE": "UNKNOWN",
            }
        )
    )

    valid_mask = (
        prepared["Symbol"].ne("")
        & prepared["Symbol"].ne("NAN")
        & prepared["Sector"].ne("UNKNOWN")
    )

    for column in (
        "Close",
        "Previous_Close",
        "EMA20",
        "EMA50",
        "EMA200",
        "High_52W",
        "Low_52W",
    ):
        valid_mask &= (
            prepared[column].notna()
            & prepared[column].gt(0)
        )

    invalid_rows = int(
        (~valid_mask).sum()
    )

    prepared = prepared.loc[
        valid_mask
    ].copy()

    prepared = prepared.drop_duplicates(
        subset=["Symbol"],
        keep="last",
    )

    if "Volume" not in prepared.columns:
        prepared["Volume"] = 0.0

    if "Average_Volume_20" not in prepared.columns:
        prepared["Average_Volume_20"] = 0.0

    if "Is_FnO" not in prepared.columns:
        prepared["Is_FnO"] = False

    prepared["Is_FnO"] = prepared[
        "Is_FnO"
    ].apply(
        normalize_boolean
    )

    prepared["Advance"] = (
        prepared["Close"]
        > prepared["Previous_Close"]
    )

    prepared["Decline"] = (
        prepared["Close"]
        < prepared["Previous_Close"]
    )

    prepared["Unchanged"] = (
        prepared["Close"]
        == prepared["Previous_Close"]
    )

    prepared["Above_EMA20"] = (
        prepared["Close"]
        > prepared["EMA20"]
    )

    prepared["Above_EMA50"] = (
        prepared["Close"]
        > prepared["EMA50"]
    )

    prepared["Above_EMA200"] = (
        prepared["Close"]
        > prepared["EMA200"]
    )

    prepared["Near_52W_High"] = (
        prepared["Close"]
        >= prepared["High_52W"]
        * NEAR_HIGH_MULTIPLIER
    )

    prepared["Near_52W_Low"] = (
        prepared["Close"]
        <= prepared["Low_52W"]
        * NEAR_LOW_MULTIPLIER
    )

    prepared["High_Volume"] = (
        prepared["Average_Volume_20"].notna()
        & prepared["Average_Volume_20"].gt(0)
        & prepared["Volume"].notna()
        & (
            prepared["Volume"]
            >= prepared["Average_Volume_20"]
        )
    )

    return (
        prepared.reset_index(drop=True),
        invalid_rows,
    )


# ==========================================================
# PREVIOUS SECTOR REPORT
# ==========================================================

def extract_report_date(
    file_path: Path,
) -> date | None:
    """
    Extract YYYYMMDD from a report filename.
    """

    match = re.search(
        r"(\d{8})",
        file_path.stem,
    )

    if match is None:
        return None

    try:
        return datetime.strptime(
            match.group(1),
            "%Y%m%d",
        ).date()

    except ValueError:
        return None


def discover_previous_sector_report(
    requested_date: date,
) -> tuple[date | None, Path | None]:
    """
    Locate the latest breadth report before the requested date.
    """

    if not BREADTH_REPORT_DIRECTORY.exists():
        return (
            None,
            None,
        )

    candidates: list[
        tuple[date, Path]
    ] = []

    for file_path in BREADTH_REPORT_DIRECTORY.glob(
        "market_breadth_report_*.xlsx"
    ):
        report_date = extract_report_date(
            file_path
        )

        if (
            report_date is not None
            and report_date < requested_date
        ):
            candidates.append(
                (
                    report_date,
                    file_path,
                )
            )

    if not candidates:
        return (
            None,
            None,
        )

    return max(
        candidates,
        key=lambda item: item[0],
    )


def read_previous_sector_scores(
    report_file: Path | None,
) -> dict[str, int]:
    """
    Read sector scores from the previous breadth report.
    """

    if report_file is None:
        return {}

    try:
        dataframe = pd.read_excel(
            report_file,
            sheet_name="Sectors",
            engine="openpyxl",
        )

    except (
        FileNotFoundError,
        ValueError,
    ):
        return {}

    if dataframe.empty:
        return {}

    dataframe.columns = [
        normalize_header(column)
        for column in dataframe.columns
    ]

    required = {
        "Sector",
        "Score",
    }

    if not required.issubset(
        dataframe.columns
    ):
        return {}

    dataframe["Sector"] = (
        dataframe["Sector"]
        .fillna("UNKNOWN")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    dataframe["Score"] = pd.to_numeric(
        dataframe["Score"],
        errors="coerce",
    )

    dataframe = dataframe.dropna(
        subset=["Score"]
    )

    return {
        str(row["Sector"]): int(
            round(float(row["Score"]))
        )
        for _, row
        in dataframe.iterrows()
        if str(row["Sector"]) != "UNKNOWN"
    }


# ==========================================================
# CLASSIFICATION HELPERS
# ==========================================================

def calculate_sector_score(
    *,
    advance_percentage: float,
    ema20_percentage: float,
    ema50_percentage: float,
    ema200_percentage: float,
    near_high_percentage: float,
    near_low_percentage: float,
    volume_ratio: float | None,
) -> int:
    """
    Calculate sector strength score.
    """

    score = (
        advance_percentage * 0.30
        + ema20_percentage * 0.22
        + ema50_percentage * 0.18
        + ema200_percentage * 0.17
        + near_high_percentage * 0.08
    )

    score -= min(
        near_low_percentage * 0.10,
        10,
    )

    if volume_ratio is not None:
        if volume_ratio >= 2.0:
            score += 5

        elif volume_ratio >= 1.25:
            score += 3

        elif volume_ratio <= 0.50:
            score -= 5

        elif volume_ratio <= 0.80:
            score -= 3

    return clamp_score(
        score
    )


def determine_strength(
    score: int,
) -> str:
    """
    Determine sector strength.
    """

    if score >= 75:
        return "VERY STRONG"

    if score >= 65:
        return "STRONG"

    if score >= 55:
        return "MODERATE TO STRONG"

    if score >= 45:
        return "MODERATE"

    if score >= 35:
        return "WEAK"

    return "VERY WEAK"


def determine_trend(
    *,
    ema20_percentage: float,
    ema50_percentage: float,
    ema200_percentage: float,
) -> str:
    """
    Determine sector trend structure.
    """

    if (
        ema20_percentage >= 65
        and ema50_percentage >= 60
        and ema200_percentage >= 55
    ):
        return "BULLISH"

    if (
        ema20_percentage <= 35
        and ema50_percentage <= 40
        and ema200_percentage <= 45
    ):
        return "BEARISH"

    if (
        ema20_percentage
        > ema50_percentage
        > ema200_percentage
    ):
        return "IMPROVING BULLISH"

    if (
        ema20_percentage
        < ema50_percentage
        < ema200_percentage
    ):
        return "DETERIORATING BEARISH"

    return "NEUTRAL TO MIXED"


def determine_momentum(
    *,
    advance_percentage: float,
    ema20_percentage: float,
    score_change: int | None,
) -> str:
    """
    Determine sector momentum.
    """

    if score_change is not None:
        if (
            score_change >= 8
            and advance_percentage >= 60
        ):
            return "STRONGLY IMPROVING"

        if score_change >= 3:
            return "IMPROVING"

        if (
            score_change <= -8
            and advance_percentage <= 40
        ):
            return "STRONGLY DETERIORATING"

        if score_change <= -3:
            return "DETERIORATING"

    if (
        advance_percentage >= 60
        and ema20_percentage >= 60
    ):
        return "POSITIVE"

    if (
        advance_percentage <= 40
        and ema20_percentage <= 40
    ):
        return "NEGATIVE"

    return "MIXED"


def determine_rotation(
    score_change: int | None,
) -> str:
    """
    Determine sector rotation.
    """

    if score_change is None:
        return "INSUFFICIENT HISTORY"

    if score_change >= 8:
        return "STRONG POSITIVE ROTATION"

    if score_change >= 3:
        return "POSITIVE ROTATION"

    if score_change <= -8:
        return "STRONG NEGATIVE ROTATION"

    if score_change <= -3:
        return "NEGATIVE ROTATION"

    return "STABLE"


def determine_leadership(
    *,
    rank: int,
    total_sectors: int,
    score: int,
    momentum: str,
) -> str:
    """
    Determine sector leadership classification.
    """

    if (
        rank <= 3
        and score >= 60
        and "DETERIORATING" not in momentum
    ):
        return "MARKET LEADER"

    if (
        rank <= max(
            5,
            round(total_sectors * 0.35),
        )
        and score >= 50
    ):
        return "LEADING"

    if (
        rank >= max(
            1,
            total_sectors - 2,
        )
        and score <= 40
    ):
        return "MARKET LAGGARD"

    if score <= 45:
        return "LAGGING"

    return "NEUTRAL"


def determine_volume_participation(
    volume_ratio: float | None,
) -> str:
    """
    Determine sector volume participation.
    """

    if volume_ratio is None:
        return "INSUFFICIENT VOLUME DATA"

    if volume_ratio >= 1.50:
        return "BULLISH VOLUME PARTICIPATION"

    if volume_ratio <= 0.67:
        return "BEARISH VOLUME PARTICIPATION"

    return "BALANCED VOLUME PARTICIPATION"


def determine_risk_level(
    *,
    score: int,
    trend: str,
    momentum: str,
    near_low_percentage: float,
) -> str:
    """
    Determine analytical sector risk.
    """

    risk_score = 20

    if score <= 40:
        risk_score += 25

    if "BEARISH" in trend:
        risk_score += 20

    if "DETERIORATING" in momentum:
        risk_score += 20

    if near_low_percentage >= 20:
        risk_score += 15

    if score >= 65:
        risk_score -= 15

    risk_score = clamp_score(
        risk_score
    )

    if risk_score >= 75:
        return "VERY HIGH"

    if risk_score >= 60:
        return "HIGH"

    if risk_score >= 45:
        return "MODERATE"

    if risk_score >= 30:
        return "LOW TO MODERATE"

    return "LOW"


def calculate_confidence(
    *,
    total_stocks: int,
    previous_score_available: bool,
    volume_available: bool,
    fno_stocks: int,
) -> int:
    """
    Calculate sector-analysis confidence.
    """

    score = 45.0

    if total_stocks >= 40:
        score += 25

    elif total_stocks >= 20:
        score += 20

    elif total_stocks >= 10:
        score += 14

    elif total_stocks >= 5:
        score += 7

    else:
        score -= 10

    if previous_score_available:
        score += 12

    if volume_available:
        score += 8

    if fno_stocks >= 5:
        score += 7

    elif fno_stocks > 0:
        score += 3

    return clamp_score(
        score
    )


# ==========================================================
# INTERPRETATION
# ==========================================================

def determine_expected_behaviour(
    *,
    strength: str,
    trend: str,
    momentum: str,
    rotation: str,
    leadership: str,
) -> str:
    """
    Determine expected sector behaviour.
    """

    if (
        leadership == "MARKET LEADER"
        and "BULLISH" in trend
        and "IMPROVING" in momentum
    ):
        return (
            "THE SECTOR MAY CONTINUE TO ATTRACT BROAD PARTICIPATION "
            "AND MAINTAIN MARKET LEADERSHIP"
        )

    if (
        "POSITIVE ROTATION" in rotation
        and "BEARISH" not in trend
    ):
        return (
            "SECTOR PARTICIPATION IS IMPROVING AND MAY DEVELOP INTO "
            "STRONGER LEADERSHIP IF THE TREND CONFIRMS"
        )

    if (
        "BEARISH" in trend
        and "IMPROVING" in momentum
    ):
        return (
            "A SECTOR RECOVERY MAY CONTINUE, BUT THE LONGER-TERM "
            "STRUCTURE REMAINS WEAK"
        )

    if (
        "BULLISH" in trend
        and "DETERIORATING" in momentum
    ):
        return (
            "THE SECTOR TREND REMAINS POSITIVE, BUT WEAKENING "
            "PARTICIPATION INCREASES PULLBACK RISK"
        )

    if (
        strength in {
            "VERY WEAK",
            "WEAK",
        }
        and "BEARISH" in trend
    ):
        return (
            "THE SECTOR MAY REMAIN UNDER PRESSURE UNTIL ADVANCE "
            "PARTICIPATION AND MOVING-AVERAGE BREADTH IMPROVE"
        )

    return (
        "SECTOR BEHAVIOUR MAY REMAIN MIXED OR STOCK-SPECIFIC "
        "UNTIL PARTICIPATION BECOMES MORE DIRECTIONAL"
    )


def build_sector_warnings(
    *,
    total_stocks: int,
    previous_score: int | None,
    volume_available: bool,
    trend: str,
    momentum: str,
) -> tuple[str, ...]:
    """
    Build warnings for one sector.
    """

    warnings: list[str] = []

    if total_stocks < 5:
        warnings.append(
            "Sector contains fewer than five analysed stocks."
        )

    if previous_score is None:
        warnings.append(
            "Historical sector score comparison is unavailable."
        )

    if not volume_available:
        warnings.append(
            "Volume participation is unavailable."
        )

    if (
        "BULLISH" in trend
        and "DETERIORATING" in momentum
    ):
        warnings.append(
            "Bullish sector structure conflicts with deteriorating momentum."
        )

    if (
        "BEARISH" in trend
        and "IMPROVING" in momentum
    ):
        warnings.append(
            "Bearish sector structure conflicts with improving momentum."
        )

    if not warnings:
        warnings.append(
            "No major sector-strength warning is active."
        )

    return tuple(
        dict.fromkeys(warnings)
    )


def build_sector_summary(
    *,
    sector: str,
    rank: int,
    score: int,
    strength: str,
    trend: str,
    momentum: str,
    rotation: str,
    confidence: int,
) -> str:
    """
    Build concise sector summary.
    """

    return (
        f"RANK {rank} | "
        f"{sector} | "
        f"SCORE {score}% | "
        f"{strength} | "
        f"{trend} | "
        f"{momentum} | "
        f"{rotation} | "
        f"{confidence}% CONFIDENCE"
    )


def build_sector_explanation(
    *,
    sector: str,
    total_stocks: int,
    advances: int,
    declines: int,
    advance_percentage: float,
    ema20_percentage: float,
    ema50_percentage: float,
    ema200_percentage: float,
    score: int,
    previous_score: int | None,
    score_change: int | None,
    strength: str,
    trend: str,
    momentum: str,
    rotation: str,
    leadership: str,
    risk_level: str,
) -> str:
    """
    Build explainable sector conclusion.
    """

    previous_text = (
        "No previous sector score is available."
        if previous_score is None
        else (
            f"The previous sector score was {previous_score}%, "
            f"producing a change of "
            f"{format_optional_number(score_change)} points."
        )
    )

    return (
        f"{sector} contains {total_stocks} valid stocks. "
        f"{advances} advanced and {declines} declined, producing an "
        f"advance rate of {advance_percentage:.2f}%. "
        f"{ema20_percentage:.2f}% are above EMA20, "
        f"{ema50_percentage:.2f}% are above EMA50 and "
        f"{ema200_percentage:.2f}% are above EMA200. "
        f"The sector breadth score is {score}%. {previous_text} "
        f"Sector strength is {strength.lower()}, trend is "
        f"{trend.lower()}, momentum is {momentum.lower()} and rotation "
        f"is {rotation.lower()}. Leadership is classified as "
        f"{leadership.lower()} and analytical risk is "
        f"{risk_level.lower()}."
    )


# ==========================================================
# SECTOR CALCULATION
# ==========================================================

def calculate_raw_sector_results(
    *,
    dataframe: pd.DataFrame,
    previous_scores: dict[str, int],
) -> list[dict[str, object]]:
    """
    Calculate raw sector results before ranking.
    """

    results: list[dict[str, object]] = []

    for sector, sector_data in dataframe.groupby(
        "Sector",
        sort=True,
    ):
        total_stocks = len(
            sector_data
        )

        if total_stocks < MINIMUM_SECTOR_STOCKS:
            continue

        advances = int(
            sector_data["Advance"].sum()
        )

        declines = int(
            sector_data["Decline"].sum()
        )

        unchanged = int(
            sector_data["Unchanged"].sum()
        )

        advance_percentage = safe_percentage(
            advances,
            total_stocks,
        )

        decline_percentage = safe_percentage(
            declines,
            total_stocks,
        )

        advance_decline_ratio = safe_ratio(
            advances,
            declines,
        )

        above_ema20_count = int(
            sector_data["Above_EMA20"].sum()
        )

        above_ema50_count = int(
            sector_data["Above_EMA50"].sum()
        )

        above_ema200_count = int(
            sector_data["Above_EMA200"].sum()
        )

        ema20_percentage = safe_percentage(
            above_ema20_count,
            total_stocks,
        )

        ema50_percentage = safe_percentage(
            above_ema50_count,
            total_stocks,
        )

        ema200_percentage = safe_percentage(
            above_ema200_count,
            total_stocks,
        )

        near_52w_high_count = int(
            sector_data["Near_52W_High"].sum()
        )

        near_52w_low_count = int(
            sector_data["Near_52W_Low"].sum()
        )

        near_high_percentage = safe_percentage(
            near_52w_high_count,
            total_stocks,
        )

        near_low_percentage = safe_percentage(
            near_52w_low_count,
            total_stocks,
        )

        high_volume_advances = int(
            (
                sector_data["High_Volume"]
                & sector_data["Advance"]
            ).sum()
        )

        high_volume_declines = int(
            (
                sector_data["High_Volume"]
                & sector_data["Decline"]
            ).sum()
        )

        volume_ratio = safe_ratio(
            high_volume_advances,
            high_volume_declines,
        )

        volume_participation = (
            determine_volume_participation(
                volume_ratio
            )
        )

        fno_stocks = int(
            sector_data["Is_FnO"].sum()
        )

        fno_percentage = safe_percentage(
            fno_stocks,
            total_stocks,
        )

        score = calculate_sector_score(
            advance_percentage=advance_percentage,
            ema20_percentage=ema20_percentage,
            ema50_percentage=ema50_percentage,
            ema200_percentage=ema200_percentage,
            near_high_percentage=near_high_percentage,
            near_low_percentage=near_low_percentage,
            volume_ratio=volume_ratio,
        )

        previous_score = previous_scores.get(
            str(sector)
        )

        score_change = (
            score - previous_score
            if previous_score is not None
            else None
        )

        strength = determine_strength(
            score
        )

        trend = determine_trend(
            ema20_percentage=ema20_percentage,
            ema50_percentage=ema50_percentage,
            ema200_percentage=ema200_percentage,
        )

        momentum = determine_momentum(
            advance_percentage=advance_percentage,
            ema20_percentage=ema20_percentage,
            score_change=score_change,
        )

        rotation = determine_rotation(
            score_change
        )

        risk_level = determine_risk_level(
            score=score,
            trend=trend,
            momentum=momentum,
            near_low_percentage=near_low_percentage,
        )

        volume_available = bool(
            sector_data["Average_Volume_20"]
            .fillna(0)
            .gt(0)
            .any()
        )

        confidence = calculate_confidence(
            total_stocks=total_stocks,
            previous_score_available=(
                previous_score is not None
            ),
            volume_available=volume_available,
            fno_stocks=fno_stocks,
        )

        results.append(
            {
                "sector": str(sector),
                "total_stocks": total_stocks,
                "advances": advances,
                "declines": declines,
                "unchanged": unchanged,
                "advance_percentage": advance_percentage,
                "decline_percentage": decline_percentage,
                "advance_decline_ratio": advance_decline_ratio,
                "above_ema20_count": above_ema20_count,
                "above_ema50_count": above_ema50_count,
                "above_ema200_count": above_ema200_count,
                "above_ema20_percentage": ema20_percentage,
                "above_ema50_percentage": ema50_percentage,
                "above_ema200_percentage": ema200_percentage,
                "near_52w_high_count": near_52w_high_count,
                "near_52w_low_count": near_52w_low_count,
                "near_52w_high_percentage": near_high_percentage,
                "near_52w_low_percentage": near_low_percentage,
                "high_volume_advances": high_volume_advances,
                "high_volume_declines": high_volume_declines,
                "volume_breadth_ratio": volume_ratio,
                "volume_participation": volume_participation,
                "fno_stocks": fno_stocks,
                "fno_percentage": fno_percentage,
                "breadth_score": score,
                "previous_score": previous_score,
                "score_change": score_change,
                "strength": strength,
                "trend": trend,
                "momentum": momentum,
                "rotation": rotation,
                "risk_level": risk_level,
                "confidence": confidence,
                "volume_available": volume_available,
            }
        )

    return sorted(
        results,
        key=lambda item: (
            int(item["breadth_score"]),
            float(item["advance_percentage"]),
            float(item["above_ema200_percentage"]),
        ),
        reverse=True,
    )


def finalize_sector_results(
    raw_results: list[dict[str, object]],
) -> tuple[SectorStrengthResult, ...]:
    """
    Add ranks, leadership and final interpretations.
    """

    final_results: list[
        SectorStrengthResult
    ] = []

    total_sectors = len(
        raw_results
    )

    for rank, raw in enumerate(
        raw_results,
        start=1,
    ):
        leadership = determine_leadership(
            rank=rank,
            total_sectors=total_sectors,
            score=int(raw["breadth_score"]),
            momentum=str(raw["momentum"]),
        )

        expected_behaviour = (
            determine_expected_behaviour(
                strength=str(raw["strength"]),
                trend=str(raw["trend"]),
                momentum=str(raw["momentum"]),
                rotation=str(raw["rotation"]),
                leadership=leadership,
            )
        )

        warnings = build_sector_warnings(
            total_stocks=int(raw["total_stocks"]),
            previous_score=(
                int(raw["previous_score"])
                if raw["previous_score"] is not None
                else None
            ),
            volume_available=bool(
                raw["volume_available"]
            ),
            trend=str(raw["trend"]),
            momentum=str(raw["momentum"]),
        )

        concise_summary = build_sector_summary(
            sector=str(raw["sector"]),
            rank=rank,
            score=int(raw["breadth_score"]),
            strength=str(raw["strength"]),
            trend=str(raw["trend"]),
            momentum=str(raw["momentum"]),
            rotation=str(raw["rotation"]),
            confidence=int(raw["confidence"]),
        )

        explanation = build_sector_explanation(
            sector=str(raw["sector"]),
            total_stocks=int(raw["total_stocks"]),
            advances=int(raw["advances"]),
            declines=int(raw["declines"]),
            advance_percentage=float(
                raw["advance_percentage"]
            ),
            ema20_percentage=float(
                raw["above_ema20_percentage"]
            ),
            ema50_percentage=float(
                raw["above_ema50_percentage"]
            ),
            ema200_percentage=float(
                raw["above_ema200_percentage"]
            ),
            score=int(raw["breadth_score"]),
            previous_score=(
                int(raw["previous_score"])
                if raw["previous_score"] is not None
                else None
            ),
            score_change=(
                int(raw["score_change"])
                if raw["score_change"] is not None
                else None
            ),
            strength=str(raw["strength"]),
            trend=str(raw["trend"]),
            momentum=str(raw["momentum"]),
            rotation=str(raw["rotation"]),
            leadership=leadership,
            risk_level=str(raw["risk_level"]),
        )

        final_results.append(
            SectorStrengthResult(
                rank=rank,
                sector=str(raw["sector"]),

                total_stocks=int(raw["total_stocks"]),
                advances=int(raw["advances"]),
                declines=int(raw["declines"]),
                unchanged=int(raw["unchanged"]),

                advance_percentage=float(
                    raw["advance_percentage"]
                ),
                decline_percentage=float(
                    raw["decline_percentage"]
                ),
                advance_decline_ratio=(
                    float(raw["advance_decline_ratio"])
                    if raw["advance_decline_ratio"] is not None
                    else None
                ),

                above_ema20_count=int(
                    raw["above_ema20_count"]
                ),
                above_ema50_count=int(
                    raw["above_ema50_count"]
                ),
                above_ema200_count=int(
                    raw["above_ema200_count"]
                ),

                above_ema20_percentage=float(
                    raw["above_ema20_percentage"]
                ),
                above_ema50_percentage=float(
                    raw["above_ema50_percentage"]
                ),
                above_ema200_percentage=float(
                    raw["above_ema200_percentage"]
                ),

                near_52w_high_count=int(
                    raw["near_52w_high_count"]
                ),
                near_52w_low_count=int(
                    raw["near_52w_low_count"]
                ),

                near_52w_high_percentage=float(
                    raw["near_52w_high_percentage"]
                ),
                near_52w_low_percentage=float(
                    raw["near_52w_low_percentage"]
                ),

                high_volume_advances=int(
                    raw["high_volume_advances"]
                ),
                high_volume_declines=int(
                    raw["high_volume_declines"]
                ),
                volume_breadth_ratio=(
                    float(raw["volume_breadth_ratio"])
                    if raw["volume_breadth_ratio"] is not None
                    else None
                ),
                volume_participation=str(
                    raw["volume_participation"]
                ),

                fno_stocks=int(raw["fno_stocks"]),
                fno_percentage=float(
                    raw["fno_percentage"]
                ),

                breadth_score=int(
                    raw["breadth_score"]
                ),
                previous_score=(
                    int(raw["previous_score"])
                    if raw["previous_score"] is not None
                    else None
                ),
                score_change=(
                    int(raw["score_change"])
                    if raw["score_change"] is not None
                    else None
                ),

                strength=str(raw["strength"]),
                trend=str(raw["trend"]),
                momentum=str(raw["momentum"]),
                rotation=str(raw["rotation"]),
                leadership=leadership,

                risk_level=str(raw["risk_level"]),
                confidence=int(raw["confidence"]),

                expected_behaviour=expected_behaviour,
                concise_summary=concise_summary,
                explanation=explanation,

                warnings=warnings,
                status="SUCCESS",
            )
        )

    return tuple(
        final_results
    )


# ==========================================================
# OVERALL SECTOR MARKET ASSESSMENT
# ==========================================================

def determine_broad_rotation(
    sectors: tuple[SectorStrengthResult, ...],
) -> str:
    """
    Determine broad sector rotation.
    """

    positive = sum(
        "POSITIVE ROTATION" in sector.rotation
        for sector in sectors
    )

    negative = sum(
        "NEGATIVE ROTATION" in sector.rotation
        for sector in sectors
    )

    available = sum(
        sector.rotation
        != "INSUFFICIENT HISTORY"
        for sector in sectors
    )

    if available == 0:
        return "INSUFFICIENT HISTORY"

    if (
        positive >= max(
            5,
            round(available * 0.60),
        )
    ):
        return "BROAD POSITIVE SECTOR ROTATION"

    if (
        negative >= max(
            5,
            round(available * 0.60),
        )
    ):
        return "BROAD NEGATIVE SECTOR ROTATION"

    if positive > negative:
        return "SELECTIVE POSITIVE SECTOR ROTATION"

    if negative > positive:
        return "SELECTIVE NEGATIVE SECTOR ROTATION"

    return "MIXED SECTOR ROTATION"


def determine_market_sector_health(
    *,
    bullish_sectors: int,
    bearish_sectors: int,
    neutral_sectors: int,
) -> str:
    """
    Determine overall sector health.
    """

    total = (
        bullish_sectors
        + bearish_sectors
        + neutral_sectors
    )

    if total == 0:
        return "UNKNOWN"

    bullish_percentage = safe_percentage(
        bullish_sectors,
        total,
    )

    bearish_percentage = safe_percentage(
        bearish_sectors,
        total,
    )

    if bullish_percentage >= 65:
        return "VERY HEALTHY"

    if bullish_percentage >= 50:
        return "HEALTHY"

    if bearish_percentage >= 65:
        return "VERY WEAK"

    if bearish_percentage >= 50:
        return "WEAK"

    return "MIXED"


def calculate_engine_confidence(
    sectors: tuple[SectorStrengthResult, ...],
) -> int:
    """
    Calculate overall engine confidence.
    """

    if not sectors:
        return 0

    average_confidence = sum(
        sector.confidence
        for sector in sectors
    ) / len(sectors)

    return clamp_score(
        average_confidence
    )


def build_engine_warnings(
    *,
    sectors: tuple[SectorStrengthResult, ...],
    previous_reference_date: date | None,
) -> tuple[str, ...]:
    """
    Build engine warnings.
    """

    warnings: list[str] = []

    if previous_reference_date is None:
        warnings.append(
            "Previous sector-score history is unavailable."
        )

    small_sectors = [
        sector.sector
        for sector in sectors
        if sector.total_stocks < 5
    ]

    if small_sectors:
        warnings.append(
            "Some sectors contain fewer than five analysed stocks: "
            + ", ".join(small_sectors)
        )

    if not sectors:
        warnings.append(
            "No sectors were eligible for analysis."
        )

    if not warnings:
        warnings.append(
            "No major sector-strength engine warning is active."
        )

    return tuple(
        dict.fromkeys(warnings)
    )


# ==========================================================
# EXPORT
# ==========================================================

def sector_results_dataframe(
    sectors: tuple[SectorStrengthResult, ...],
) -> pd.DataFrame:
    """
    Convert sector results to a dataframe.
    """

    return pd.DataFrame(
        [
            {
                "Rank": sector.rank,
                "Sector": sector.sector,
                "Total_Stocks": sector.total_stocks,
                "Advances": sector.advances,
                "Declines": sector.declines,
                "Unchanged": sector.unchanged,
                "Advance_Percentage": sector.advance_percentage,
                "Decline_Percentage": sector.decline_percentage,
                "Advance_Decline_Ratio": (
                    sector.advance_decline_ratio
                ),
                "Above_EMA20_Percentage": (
                    sector.above_ema20_percentage
                ),
                "Above_EMA50_Percentage": (
                    sector.above_ema50_percentage
                ),
                "Above_EMA200_Percentage": (
                    sector.above_ema200_percentage
                ),
                "Near_52W_High_Percentage": (
                    sector.near_52w_high_percentage
                ),
                "Near_52W_Low_Percentage": (
                    sector.near_52w_low_percentage
                ),
                "High_Volume_Advances": (
                    sector.high_volume_advances
                ),
                "High_Volume_Declines": (
                    sector.high_volume_declines
                ),
                "Volume_Breadth_Ratio": (
                    sector.volume_breadth_ratio
                ),
                "Volume_Participation": (
                    sector.volume_participation
                ),
                "FNO_Stocks": sector.fno_stocks,
                "FNO_Percentage": sector.fno_percentage,
                "Breadth_Score": sector.breadth_score,
                "Previous_Score": sector.previous_score,
                "Score_Change": sector.score_change,
                "Strength": sector.strength,
                "Trend": sector.trend,
                "Momentum": sector.momentum,
                "Rotation": sector.rotation,
                "Leadership": sector.leadership,
                "Risk_Level": sector.risk_level,
                "Confidence": sector.confidence,
                "Expected_Behaviour": (
                    sector.expected_behaviour
                ),
                "Concise_Summary": (
                    sector.concise_summary
                ),
                "Explanation": sector.explanation,
                "Status": sector.status,
            }
            for sector in sectors
        ]
    )


def format_workbook(
    workbook_path: Path,
) -> None:
    """
    Apply simple workbook formatting.
    """

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(
        workbook_path
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for column_cells in worksheet.columns:
            maximum_length = 0

            column_number = column_cells[
                0
            ].column

            for cell in column_cells:
                value = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                maximum_length = max(
                    maximum_length,
                    len(value),
                )

            worksheet.column_dimensions[
                get_column_letter(
                    column_number
                )
            ].width = min(
                max(
                    maximum_length + 2,
                    12,
                ),
                45,
            )

    workbook.save(
        workbook_path
    )


def export_result(
    result: SectorStrengthEngineResult,
) -> tuple[Path, Path]:
    """
    Export sector-strength results.
    """

    OUTPUT_DIRECTORY.mkdir(
        parents=True,
        exist_ok=True,
    )

    date_text = result.analysis_date.strftime(
        "%Y%m%d"
    )

    csv_file = (
        OUTPUT_DIRECTORY
        / f"sector_strength_summary_{date_text}.csv"
    )

    excel_file = (
        OUTPUT_DIRECTORY
        / f"sector_strength_report_{date_text}.xlsx"
    )

    sector_dataframe = sector_results_dataframe(
        result.sectors
    )

    summary_dataframe = pd.DataFrame(
        [
            {
                "Requested_Date": result.requested_date,
                "Analysis_Date": result.analysis_date,
                "Previous_Reference_Date": (
                    result.previous_reference_date
                ),
                "Total_Input_Rows": (
                    result.total_input_rows
                ),
                "Valid_Rows": result.valid_rows,
                "Invalid_Rows": result.invalid_rows,
                "Sectors_Analysed": (
                    result.sectors_analysed
                ),
                "Bullish_Sectors": (
                    result.bullish_sectors
                ),
                "Bearish_Sectors": (
                    result.bearish_sectors
                ),
                "Neutral_Sectors": (
                    result.neutral_sectors
                ),
                "Strongest_Sector": (
                    result.strongest_sector
                ),
                "Weakest_Sector": (
                    result.weakest_sector
                ),
                "Broad_Rotation": (
                    result.broad_rotation
                ),
                "Market_Sector_Health": (
                    result.market_sector_health
                ),
                "Confidence": result.confidence,
                "Expected_Behaviour": (
                    result.expected_behaviour
                ),
                "Concise_Summary": (
                    result.concise_summary
                ),
                "Explanation": result.explanation,
                "Status": result.status,
            }
        ]
    )

    warning_dataframe = pd.DataFrame(
        {
            "Warning": list(
                result.warnings
            )
        }
    )

    sector_dataframe.to_csv(
        csv_file,
        index=False,
    )

    with pd.ExcelWriter(
        excel_file,
        engine="openpyxl",
    ) as writer:
        summary_dataframe.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )

        sector_dataframe.to_excel(
            writer,
            sheet_name="Sector Rankings",
            index=False,
        )

        sector_dataframe.head(5).to_excel(
            writer,
            sheet_name="Top Sectors",
            index=False,
        )

        sector_dataframe.tail(5).to_excel(
            writer,
            sheet_name="Weakest Sectors",
            index=False,
        )

        warning_dataframe.to_excel(
            writer,
            sheet_name="Warnings",
            index=False,
        )

    format_workbook(
        excel_file
    )

    return (
        csv_file,
        excel_file,
    )


# ==========================================================
# MAIN ENGINE
# ==========================================================

def run_sector_strength_engine(
    *,
    requested_date: date,
    source_file: Path = DEFAULT_INPUT_FILE,
    export: bool = True,
) -> SectorStrengthEngineResult:
    """
    Run the complete Sector Strength Engine.
    """

    raw_dataframe = read_sector_snapshot(
        source_file
    )

    total_input_rows = len(
        raw_dataframe
    )

    prepared_dataframe, invalid_rows = (
        prepare_snapshot(
            raw_dataframe
        )
    )

    valid_rows = len(
        prepared_dataframe
    )

    if valid_rows == 0:
        raise RuntimeError(
            "No valid classified stock records are available."
        )

    (
        previous_reference_date,
        previous_report_file,
    ) = discover_previous_sector_report(
        requested_date
    )

    previous_scores = (
        read_previous_sector_scores(
            previous_report_file
        )
    )

    raw_sector_results = (
        calculate_raw_sector_results(
            dataframe=prepared_dataframe,
            previous_scores=previous_scores,
        )
    )

    sectors = finalize_sector_results(
        raw_sector_results
    )

    if not sectors:
        raise RuntimeError(
            "No sector had enough valid stocks for analysis."
        )

    bullish_sectors = sum(
        (
            "BULLISH" in sector.trend
            or sector.breadth_score >= 58
        )
        for sector in sectors
    )

    bearish_sectors = sum(
        (
            "BEARISH" in sector.trend
            or sector.breadth_score <= 42
        )
        for sector in sectors
    )

    neutral_sectors = (
        len(sectors)
        - bullish_sectors
        - bearish_sectors
    )

    if neutral_sectors < 0:
        neutral_sectors = 0

    strongest_sector = (
        sectors[0].sector
        if sectors
        else None
    )

    weakest_sector = (
        sectors[-1].sector
        if sectors
        else None
    )

    broad_rotation = determine_broad_rotation(
        sectors
    )

    market_sector_health = (
        determine_market_sector_health(
            bullish_sectors=bullish_sectors,
            bearish_sectors=bearish_sectors,
            neutral_sectors=neutral_sectors,
        )
    )

    confidence = calculate_engine_confidence(
        sectors
    )

    if broad_rotation == (
        "BROAD POSITIVE SECTOR ROTATION"
    ):
        expected_behaviour = (
            "BROADER SECTOR PARTICIPATION MAY SUPPORT A MORE "
            "CONSTRUCTIVE MARKET ENVIRONMENT"
        )

    elif broad_rotation == (
        "BROAD NEGATIVE SECTOR ROTATION"
    ):
        expected_behaviour = (
            "BROAD SECTOR DETERIORATION MAY SUPPORT A MORE "
            "DEFENSIVE OR RISK-OFF MARKET ENVIRONMENT"
        )

    elif "POSITIVE" in broad_rotation:
        expected_behaviour = (
            "LEADERSHIP IS IMPROVING SELECTIVELY, SO MARKET "
            "BEHAVIOUR MAY REMAIN ROTATIONAL"
        )

    elif "NEGATIVE" in broad_rotation:
        expected_behaviour = (
            "SECTOR WEAKNESS IS EXPANDING SELECTIVELY AND MAY "
            "CREATE UNEVEN MARKET PRESSURE"
        )

    else:
        expected_behaviour = (
            "SECTOR PARTICIPATION REMAINS MIXED AND STOCK-SPECIFIC "
            "BEHAVIOUR MAY CONTINUE"
        )

    concise_summary = (
        f"{len(sectors)} SECTORS | "
        f"STRONGEST {strongest_sector} | "
        f"WEAKEST {weakest_sector} | "
        f"{broad_rotation} | "
        f"{market_sector_health} HEALTH | "
        f"{confidence}% CONFIDENCE"
    )

    explanation = (
        f"The Sector Strength Engine analysed {len(sectors)} sectors "
        f"using {valid_rows:,} valid classified stocks. "
        f"{bullish_sectors} sectors are bullish, "
        f"{bearish_sectors} are bearish and "
        f"{neutral_sectors} are neutral. "
        f"The strongest sector is {strongest_sector}, while the "
        f"weakest sector is {weakest_sector}. "
        f"Broad rotation is classified as "
        f"{broad_rotation.lower()}, producing "
        f"{market_sector_health.lower()} sector health. "
        f"Overall confidence is {confidence}%."
    )

    warnings = build_engine_warnings(
        sectors=sectors,
        previous_reference_date=(
            previous_reference_date
        ),
    )

    result = SectorStrengthEngineResult(
        requested_date=requested_date,
        analysis_date=requested_date,
        source_file=source_file,
        previous_reference_date=(
            previous_reference_date
        ),

        total_input_rows=total_input_rows,
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,

        sectors_analysed=len(sectors),
        bullish_sectors=bullish_sectors,
        bearish_sectors=bearish_sectors,
        neutral_sectors=neutral_sectors,

        strongest_sector=strongest_sector,
        weakest_sector=weakest_sector,

        broad_rotation=broad_rotation,
        market_sector_health=(
            market_sector_health
        ),

        confidence=confidence,
        expected_behaviour=expected_behaviour,
        concise_summary=concise_summary,
        explanation=explanation,

        sectors=sectors,
        warnings=warnings,

        csv_file=None,
        excel_file=None,

        status="SUCCESS",
    )

    if export:
        csv_file, excel_file = export_result(
            result
        )

        result = SectorStrengthEngineResult(
            requested_date=result.requested_date,
            analysis_date=result.analysis_date,
            source_file=result.source_file,
            previous_reference_date=(
                result.previous_reference_date
            ),

            total_input_rows=(
                result.total_input_rows
            ),
            valid_rows=result.valid_rows,
            invalid_rows=result.invalid_rows,

            sectors_analysed=(
                result.sectors_analysed
            ),
            bullish_sectors=(
                result.bullish_sectors
            ),
            bearish_sectors=(
                result.bearish_sectors
            ),
            neutral_sectors=(
                result.neutral_sectors
            ),

            strongest_sector=(
                result.strongest_sector
            ),
            weakest_sector=(
                result.weakest_sector
            ),

            broad_rotation=(
                result.broad_rotation
            ),
            market_sector_health=(
                result.market_sector_health
            ),

            confidence=result.confidence,
            expected_behaviour=(
                result.expected_behaviour
            ),
            concise_summary=(
                result.concise_summary
            ),
            explanation=result.explanation,

            sectors=result.sectors,
            warnings=result.warnings,

            csv_file=csv_file,
            excel_file=excel_file,

            status=result.status,
        )

    return result


# ==========================================================
# DISPLAY
# ==========================================================

def display_result(
    result: SectorStrengthEngineResult,
) -> None:
    """
    Display the complete MBI-006 result.
    """

    print()
    print("=" * 112)
    print("AQSD SECTOR STRENGTH ENGINE")
    print("=" * 112)
    print(f"Module                              : {MODULE_ID}")
    print(f"Version                             : {MODULE_VERSION}")
    print(f"Requested Date                      : {result.requested_date}")
    print(f"Analysis Date                       : {result.analysis_date}")
    print(f"Source File                         : {result.source_file}")
    print(
        f"Previous Reference Date             : "
        f"{result.previous_reference_date}"
    )
    print("-" * 112)

    print("SECTOR MARKET SUMMARY")
    print("-" * 112)
    print(
        f"Valid Classified Stocks             : "
        f"{result.valid_rows:,}"
    )
    print(
        f"Invalid / Excluded Rows             : "
        f"{result.invalid_rows:,}"
    )
    print(
        f"Sectors Analysed                    : "
        f"{result.sectors_analysed}"
    )
    print(
        f"Bullish Sectors                     : "
        f"{result.bullish_sectors}"
    )
    print(
        f"Bearish Sectors                     : "
        f"{result.bearish_sectors}"
    )
    print(
        f"Neutral Sectors                     : "
        f"{result.neutral_sectors}"
    )
    print(
        f"Strongest Sector                    : "
        f"{result.strongest_sector}"
    )
    print(
        f"Weakest Sector                      : "
        f"{result.weakest_sector}"
    )
    print(
        f"Broad Rotation                      : "
        f"{result.broad_rotation}"
    )
    print(
        f"Market Sector Health                : "
        f"{result.market_sector_health}"
    )
    print(
        f"Confidence                          : "
        f"{result.confidence}%"
    )
    print("-" * 112)

    print("SECTOR RANKINGS")
    print("-" * 112)

    for sector in result.sectors:
        score_change = format_optional_number(
            sector.score_change
        )

        print(
            f"{sector.rank:>2}. "
            f"{sector.sector:<34} | "
            f"Score {sector.breadth_score:>3}% | "
            f"Change {score_change:<13} | "
            f"{sector.strength:<18} | "
            f"{sector.trend:<22} | "
            f"{sector.momentum}"
        )

    print("-" * 112)
    print("TOP FIVE SECTORS")
    print("-" * 112)

    for sector in result.sectors[:5]:
        print(
            f"{sector.rank}. {sector.sector}"
        )
        print(
            f"   Breadth Score       : "
            f"{sector.breadth_score}%"
        )
        print(
            f"   Advance Rate        : "
            f"{sector.advance_percentage:.2f}%"
        )
        print(
            f"   Above EMA20         : "
            f"{sector.above_ema20_percentage:.2f}%"
        )
        print(
            f"   Above EMA50         : "
            f"{sector.above_ema50_percentage:.2f}%"
        )
        print(
            f"   Above EMA200        : "
            f"{sector.above_ema200_percentage:.2f}%"
        )
        print(
            f"   Strength            : "
            f"{sector.strength}"
        )
        print(
            f"   Trend               : "
            f"{sector.trend}"
        )
        print(
            f"   Momentum            : "
            f"{sector.momentum}"
        )
        print(
            f"   Rotation            : "
            f"{sector.rotation}"
        )
        print(
            f"   Leadership          : "
            f"{sector.leadership}"
        )
        print(
            f"   Risk                : "
            f"{sector.risk_level}"
        )
        print(
            f"   Confidence          : "
            f"{sector.confidence}%"
        )
        print("-" * 112)

    print("WEAKEST FIVE SECTORS")
    print("-" * 112)

    for sector in result.sectors[-5:]:
        print(
            f"{sector.rank}. "
            f"{sector.sector:<35} | "
            f"Score {sector.breadth_score}% | "
            f"{sector.strength} | "
            f"{sector.trend} | "
            f"{sector.risk_level} RISK"
        )

    print("-" * 112)
    print("EXPECTED BEHAVIOUR")
    print("-" * 112)
    print(result.expected_behaviour)

    print("-" * 112)
    print("CONCISE SUMMARY")
    print("-" * 112)
    print(result.concise_summary)

    print("-" * 112)
    print("WARNINGS")
    print("-" * 112)

    for number, warning in enumerate(
        result.warnings,
        start=1,
    ):
        print(
            f"{number}. {warning}"
        )

    print("-" * 112)
    print("EXPLANATION")
    print("-" * 112)
    print(result.explanation)

    print("-" * 112)
    print(
        f"CSV Output                          : "
        f"{result.csv_file}"
    )
    print(
        f"Excel Output                        : "
        f"{result.excel_file}"
    )
    print(
        "Method                              : "
        "RULE-BASED SECTOR STRENGTH ANALYSIS"
    )
    print(
        f"Status                              : "
        f"{result.status}"
    )
    print("=" * 112)


# ==========================================================
# COMMAND LINE
# ==========================================================

def parse_date(
    value: str,
) -> date:
    """
    Parse YYYY-MM-DD.
    """

    try:
        return datetime.strptime(
            value,
            "%Y-%m-%d",
        ).date()

    except ValueError as exc:
        raise ValueError(
            "Invalid date format. Use YYYY-MM-DD."
        ) from exc


def parse_arguments() -> argparse.Namespace:
    """
    Read command-line arguments.
    """

    parser = argparse.ArgumentParser(
        description=(
            "Run the AQSD Sector Strength Engine."
        )
    )

    parser.add_argument(
        "--date",
        required=True,
        help="Requested analysis date in YYYY-MM-DD format.",
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_FILE,
        help=(
            "Path to the enriched market breadth snapshot. "
            f"Default: {DEFAULT_INPUT_FILE}"
        ),
    )

    parser.add_argument(
        "--no-export",
        action="store_true",
        help="Run without creating CSV and Excel outputs.",
    )

    return parser.parse_args()


def main() -> None:
    """
    Command-line entry point.
    """

    arguments = parse_arguments()

    try:
        result = run_sector_strength_engine(
            requested_date=parse_date(
                arguments.date
            ),
            source_file=(
                arguments.input
                .expanduser()
                .resolve()
            ),
            export=not arguments.no_export,
        )

    except Exception as exc:
        print()
        print("=" * 112)
        print("AQSD SECTOR STRENGTH ENGINE")
        print("=" * 112)
        print("Status : FAILED")
        print(f"Reason : {exc}")
        print("=" * 112)

        raise SystemExit(1) from exc

    display_result(
        result
    )


if __name__ == "__main__":
    main()